from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from hashlib import sha256
from io import BytesIO
from pathlib import Path
import re
import shutil
import unicodedata
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pypdf import PdfReader


MAX_ACADEMIC_DOCUMENT_PAGES = 40
_COMMON_TESSERACT_PATHS = (
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
)
_INVALID_SHEET_TITLE = re.compile(r"[\\/*?:\[\]]")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_MALLA_ADM_TEMPLATE_PATH = Path(__file__).resolve().parents[1] / "templates" / "malla_adm_template.xlsx"
_MALLA_ADM_TEMPLATE_SHA256 = "4f5f82ff48714839407c3c6f454fac3db63725d45e28ec561acadfbc7dd76c96"
_MALLA_ADM_TEMPLATE_SHEET = "Malla ADM"
_MALLA_ADM_SUBJECT_ROWS = (
    3, 4, 5, 6, 7, 8,
    10, 11, 12, 13, 14, 15,
    17, 18, 19, 20, 21, 22,
    24, 25, 26, 27, 28, 29,
)


@dataclass(frozen=True)
class SheetLayout:
    sheet_name: str
    header_row: int
    subject_rows: tuple[int, ...]
    subject_col: int
    period_col: int
    curricular_unit_col: int
    hours_col: int
    field_col: int | None = None
    outcomes_col: int | None = None
    contents_col: int | None = None

    @property
    def enriched(self) -> bool:
        return all((self.field_col, self.outcomes_col, self.contents_col))


@dataclass(frozen=True)
class PdfTextPage:
    text: str
    width: float
    height: float
    items: tuple[tuple[float, float, float, str], ...]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _normalized(value: Any) -> str:
    text = unicodedata.normalize("NFD", _clean(value).casefold())
    text = "".join(character for character in text if unicodedata.category(character) != "Mn")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _compact_multiline(value: Any) -> str:
    lines = [_clean(line) for line in str(value or "").replace("\r", "\n").split("\n")]
    return "\n".join(line for line in lines if line)


def _safe_excel_text(value: Any) -> str:
    text = _compact_multiline(value)
    if text.startswith(_FORMULA_PREFIXES):
        return f"'{text}"
    return text


def _header_column_map(sheet: Worksheet, header_row: int) -> dict[str, int]:
    columns: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = _normalized(sheet.cell(header_row, column).value)
        if value:
            columns[value] = column
    return columns


def _find_column(columns: dict[str, int], *needles: str) -> int | None:
    normalized_needles = tuple(_normalized(needle) for needle in needles)
    for label, column in columns.items():
        if any(label == needle or label.startswith(needle) for needle in normalized_needles):
            return column
    return None


def _subject_rows(
    sheet: Worksheet,
    *,
    header_row: int,
    subject_col: int,
    period_col: int,
    curricular_unit_col: int,
    hours_col: int,
) -> tuple[int, ...]:
    rows: list[int] = []
    for row in range(header_row + 2, sheet.max_row + 1):
        subject = _clean(sheet.cell(row, subject_col).value)
        period = _clean(sheet.cell(row, period_col).value)
        curricular_unit = _clean(sheet.cell(row, curricular_unit_col).value)
        hours = sheet.cell(row, hours_col).value
        if not subject or not period or not curricular_unit or hours in (None, ""):
            continue
        if _normalized(subject).startswith(("total", "subtotal")):
            continue
        rows.append(row)
    return tuple(rows)


def inspect_sheet(sheet: Worksheet) -> SheetLayout | None:
    for header_row in range(1, min(sheet.max_row, 12) + 1):
        columns = _header_column_map(sheet, header_row)
        subject_col = _find_column(columns, "Asignaturas")
        period_col = _find_column(columns, "Período Académico")
        curricular_unit_col = _find_column(columns, "Unidades de organización curricular")
        hours_col = _find_column(columns, "Horas / Créditos")
        if not all((subject_col, period_col, curricular_unit_col, hours_col)):
            continue

        field_col = _find_column(columns, "Campos de formación", "Campo de formación")
        outcomes_col = _find_column(columns, "Resultados de aprendizaje")
        contents_col = _find_column(columns, "Contenidos mínimos", "Contenido mínimo")
        rows = _subject_rows(
            sheet,
            header_row=header_row,
            subject_col=subject_col,
            period_col=period_col,
            curricular_unit_col=curricular_unit_col,
            hours_col=hours_col,
        )
        if not rows:
            continue
        return SheetLayout(
            sheet_name=sheet.title,
            header_row=header_row,
            subject_rows=rows,
            subject_col=subject_col,
            period_col=period_col,
            curricular_unit_col=curricular_unit_col,
            hours_col=hours_col,
            field_col=field_col,
            outcomes_col=outcomes_col,
            contents_col=contents_col,
        )
    return None


def _is_empty_sheet(sheet: Worksheet) -> bool:
    return not any(
        cell.value not in (None, "")
        for row in sheet.iter_rows()
        for cell in row
    )


def _career_from_title(sheet: Worksheet, header_row: int) -> str:
    for row in range(1, header_row):
        values = [_clean(sheet.cell(row, column).value) for column in range(1, min(sheet.max_column, 20) + 1)]
        text = next((value for value in values if value), "")
        if not text:
            continue
        match = re.search(r"(?i)\bcarrera(?:\s+en)?\s*:?\s*(.+)", text)
        if match:
            value = re.sub(r"\s+-\s+[A-Z0-9][A-Z0-9-]{3,}\s*$", "", match.group(1)).strip()
            if value:
                return value
    return ""


def infer_career_name(
    workbook: Any,
    filename: str,
    base_layout: SheetLayout,
    requested_name: str = "",
) -> str:
    requested = _clean(requested_name)
    if requested:
        return requested[:180]

    base_sheet = workbook[base_layout.sheet_name]
    title_career = _career_from_title(base_sheet, base_layout.header_row)
    if title_career:
        return title_career[:180]

    generic_sheet_names = {"malla", "malla academica", "malla curricular"}
    base_name = _clean(base_sheet.title)
    if _normalized(base_name) not in generic_sheet_names:
        if _normalized(base_name).startswith("malla "):
            base_name = re.sub(r"(?i)^malla\s+", "", base_name).strip()
        if base_name:
            return base_name[:180]

    for sheet in workbook.worksheets:
        if _is_empty_sheet(sheet) and _normalized(sheet.title).startswith("malla "):
            value = re.sub(r"(?i)^malla\s+", "", sheet.title).strip()
            if value:
                return value[:180]

    stem = Path(filename).stem
    stem = re.sub(r"(?i)^malla(?:\s+carreras?)?\s*", "", stem)
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem).strip(" -_")
    return (stem or "Carrera")[:180]


def _target_sheet_title(career_name: str) -> str:
    cleaned = _INVALID_SHEET_TITLE.sub(" ", _clean(career_name))
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return f"Malla {cleaned}"[:31].rstrip() or "Malla actualizada"


def _layout_subject_names(workbook: Any, layout: SheetLayout) -> set[str]:
    sheet = workbook[layout.sheet_name]
    return {
        name
        for row in layout.subject_rows
        if (name := _normalized(sheet.cell(row, layout.subject_col).value))
    }


def _subject_overlap(workbook: Any, left: SheetLayout, right: SheetLayout) -> float:
    left_names = _layout_subject_names(workbook, left)
    right_names = _layout_subject_names(workbook, right)
    if not left_names or not right_names:
        return 0.0
    return len(left_names & right_names) / max(len(left_names), len(right_names))


def _empty_target_candidate(
    workbook: Any,
    base_layout: SheetLayout,
    career_name: str,
) -> Worksheet | None:
    candidates = [
        sheet
        for sheet in workbook.worksheets
        if sheet.title != base_layout.sheet_name
        and _is_empty_sheet(sheet)
        and _normalized(sheet.title).startswith("malla")
    ]
    if not candidates:
        return None

    expected_title = _normalized(_target_sheet_title(career_name))
    for sheet in candidates:
        if _normalized(sheet.title) == expected_title:
            return sheet
    if len(candidates) == 1:
        return candidates[0]

    career = _normalized(career_name)
    ranked = sorted(
        candidates,
        key=lambda sheet: SequenceMatcher(None, career, re.sub(r"^malla\s+", "", _normalized(sheet.title))).ratio(),
        reverse=True,
    )
    best = ranked[0]
    score = SequenceMatcher(None, career, re.sub(r"^malla\s+", "", _normalized(best.title))).ratio()
    return best if score >= 0.55 else None


def inspect_workbook(content: bytes, filename: str, requested_career: str = "") -> tuple[Any, dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=False, keep_links=False)
    except Exception as exc:
        raise ValueError("No se pudo abrir el archivo Excel. Verifique que sea un .xlsx válido.") from exc

    layouts = [layout for sheet in workbook.worksheets if (layout := inspect_sheet(sheet))]
    enriched = sorted(
        (layout for layout in layouts if layout.enriched),
        key=lambda item: len(item.subject_rows),
        reverse=True,
    )
    bases = sorted(
        (layout for layout in layouts if not layout.enriched),
        key=lambda item: len(item.subject_rows),
        reverse=True,
    )
    base_layout = bases[0] if bases else (enriched[0] if enriched else None)
    if not base_layout:
        raise ValueError(
            "No se encontró una hoja de malla con los encabezados Asignaturas, Período Académico, "
            "Unidades de organización curricular y Horas / Créditos."
        )

    career_name = infer_career_name(workbook, filename, base_layout, requested_career)
    target_layout: SheetLayout | None = None
    if enriched:
        if bases:
            ranked_targets = sorted(
                enriched,
                key=lambda layout: (_subject_overlap(workbook, base_layout, layout), len(layout.subject_rows)),
                reverse=True,
            )
            if _subject_overlap(workbook, base_layout, ranked_targets[0]) >= 0.65:
                target_layout = ranked_targets[0]
        else:
            target_layout = enriched[0]
    blank_target = _empty_target_candidate(workbook, base_layout, career_name) if not target_layout else None
    target_name = target_layout.sheet_name if target_layout else (
        blank_target.title if blank_target else _target_sheet_title(career_name)
    )

    warnings: list[str] = []
    if target_layout:
        warnings.append(
            f"Se conservarán los datos existentes de {target_layout.sheet_name} y la hoja se regenerará "
            "desde la plantilla institucional Malla ADM."
        )
    else:
        warnings.append(
            f"No existe una estructura enriquecida; se creará la hoja {target_name} "
            "desde la plantilla institucional Malla ADM."
        )

    normalized_names: dict[str, int] = {}
    reference_layout = target_layout or base_layout
    reference_sheet = workbook[reference_layout.sheet_name]
    for row in reference_layout.subject_rows:
        name = _normalized(reference_sheet.cell(row, reference_layout.subject_col).value)
        normalized_names[name] = normalized_names.get(name, 0) + 1
    duplicates = [name for name, count in normalized_names.items() if name and count > 1]
    if duplicates:
        warnings.append("La malla contiene asignaturas repetidas; deberán revisarse antes de aplicar los documentos.")

    if target_layout and bases:
        base_by_name = {
            _normalized(workbook[base_layout.sheet_name].cell(row, base_layout.subject_col).value): row
            for row in base_layout.subject_rows
        }
        differences: list[str] = []
        target_sheet = workbook[target_layout.sheet_name]
        base_sheet = workbook[base_layout.sheet_name]
        for row in target_layout.subject_rows:
            subject = _clean(target_sheet.cell(row, target_layout.subject_col).value)
            base_row = base_by_name.get(_normalized(subject))
            if not base_row:
                continue
            target_unit = _clean(target_sheet.cell(row, target_layout.curricular_unit_col).value)
            base_unit = _clean(base_sheet.cell(base_row, base_layout.curricular_unit_col).value)
            if _normalized(target_unit) != _normalized(base_unit):
                differences.append(f"{subject}: {target_unit} / {base_unit}")
        if differences:
            warnings.append(
                "Clasificaciones distintas entre la malla enriquecida y la base: "
                + "; ".join(differences[:8])
                + ("." if len(differences) <= 8 else "; entre otras.")
            )

    info = {
        "filename": filename,
        "career_name": career_name,
        "source_sheet": base_layout.sheet_name,
        "target_sheet": target_name,
        "target_exists": target_layout is not None,
        "target_will_be_created": target_layout is None,
        "header_row": reference_layout.header_row,
        "subject_count": len(reference_layout.subject_rows),
        "period_count": len(
            {
                _clean(reference_sheet.cell(row, reference_layout.period_col).value)
                for row in reference_layout.subject_rows
            }
        ),
        "warnings": warnings,
        "base_layout": base_layout,
        "target_layout": target_layout,
        "blank_target": blank_target.title if blank_target else "",
    }
    return workbook, info


def _tesseract_executable() -> str:
    executable = shutil.which("tesseract")
    if executable:
        return executable
    for candidate in _COMMON_TESSERACT_PATHS:
        if Path(candidate).exists():
            return candidate
    return ""


def ocr_available() -> bool:
    if not _tesseract_executable():
        return False
    try:
        import pypdfium2  # type: ignore[import-not-found]  # noqa: F401
        import pytesseract  # type: ignore[import-not-found]  # noqa: F401
    except Exception:
        return False
    return True


def _extract_selectable_pdf_text(data: bytes) -> tuple[str, int, list[str], list[PdfTextPage]]:
    warnings: list[str] = []
    try:
        reader = PdfReader(BytesIO(data))
    except Exception as exc:
        raise ValueError("El PDF no se pudo abrir.") from exc
    if reader.is_encrypted:
        try:
            if not reader.decrypt(""):
                raise ValueError("El PDF está protegido con contraseña.")
        except Exception as exc:
            raise ValueError("El PDF está protegido con contraseña.") from exc

    page_count = len(reader.pages)
    if page_count > MAX_ACADEMIC_DOCUMENT_PAGES:
        warnings.append(
            f"El documento tiene {page_count} páginas; se analizaron las primeras {MAX_ACADEMIC_DOCUMENT_PAGES}."
        )
    parts: list[str] = []
    positioned_pages: list[PdfTextPage] = []
    for index, page in enumerate(reader.pages[:MAX_ACADEMIC_DOCUMENT_PAGES]):
        items: list[tuple[float, float, float, str]] = []

        def visitor_text(
            text: str,
            _current_transformation_matrix: Any,
            text_matrix: Any,
            _font_dictionary: Any,
            font_size: float,
        ) -> None:
            value = _clean(text)
            if not value or len(text_matrix) < 6:
                return
            items.append((float(text_matrix[4]), float(text_matrix[5]), float(font_size), value))

        try:
            page_text = page.extract_text(visitor_text=visitor_text) or ""
            parts.append(page_text)
        except Exception:
            page_text = ""
            warnings.append(f"No se pudo extraer el texto seleccionable de la página {index + 1}.")
        positioned_pages.append(
            PdfTextPage(
                text=page_text,
                width=float(page.mediabox.width),
                height=float(page.mediabox.height),
                items=tuple(items),
            )
        )
    return "\n".join(parts), page_count, warnings, positioned_pages


def _extract_layout_pdf_pages(data: bytes) -> list[str]:
    try:
        reader = PdfReader(BytesIO(data))
    except Exception:
        return []
    pages: list[str] = []
    for page in reader.pages[:MAX_ACADEMIC_DOCUMENT_PAGES]:
        try:
            pages.append(page.extract_text(extraction_mode="layout") or "")
        except Exception:
            pages.append("")
    return pages


def _extract_ocr_pdf_text(data: bytes) -> str:
    tesseract = _tesseract_executable()
    if not tesseract:
        return ""
    try:
        import pypdfium2 as pdfium  # type: ignore[import-not-found]
        import pytesseract  # type: ignore[import-not-found]
    except Exception:
        return ""

    pytesseract.pytesseract.tesseract_cmd = tesseract
    try:
        document = pdfium.PdfDocument(data)
    except Exception:
        return ""

    parts: list[str] = []
    try:
        for page_index in range(min(len(document), MAX_ACADEMIC_DOCUMENT_PAGES)):
            page = document[page_index]
            try:
                bitmap = page.render(scale=2.4, rotation=0)
                image = bitmap.to_pil()
                try:
                    text = pytesseract.image_to_string(image, lang="spa+eng", config="--psm 6")
                except Exception:
                    text = pytesseract.image_to_string(image, config="--psm 6")
                parts.append(text or "")
            finally:
                try:
                    page.close()
                except Exception:
                    pass
    finally:
        try:
            document.close()
        except Exception:
            pass
    return "\n".join(parts)


def _document_lines(text: str) -> list[str]:
    normalized = text.replace("\r", "\n").replace("\xa0", " ").replace("\u00ad", "")
    return [_clean(line) for line in normalized.split("\n") if _clean(line)]


def _is_document_noise(line: str) -> bool:
    value = _normalized(line)
    if not value:
        return True
    if value.startswith("instituto superior tecnologico intec"):
        return True
    if value.startswith("vicerrectorado general academico"):
        return True
    if re.fullmatch(r"pagina?\s*\d+\s+de\s*\d+", value):
        return True
    if re.fullmatch(r"\d+", value):
        return True
    if " pagina " in f" {value} " and re.search(r"pagina\s+\d+\s+de", value):
        return True
    return False


def _value_after_label(
    lines: list[str],
    label_pattern: str,
    stop_patterns: Iterable[str],
    *,
    max_following_lines: int = 2,
) -> str:
    label = re.compile(label_pattern, flags=re.IGNORECASE)
    stops = [re.compile(pattern, flags=re.IGNORECASE) for pattern in stop_patterns]
    for index, line in enumerate(lines):
        match = label.search(line)
        if not match:
            continue
        rest = line[match.end() :].strip(" :-")
        for stop in stops:
            stop_match = stop.search(rest)
            if stop_match:
                rest = rest[: stop_match.start()].strip()
        collected = [rest] if rest else []
        if collected:
            return _clean(" ".join(collected))
        for following in lines[index + 1 : index + 1 + max_following_lines]:
            if any(stop.search(following) for stop in stops) or _is_document_noise(following):
                break
            collected.append(following)
        value = _clean(" ".join(collected))
        if value:
            return value
    return ""


def _filename_subject(filename: str) -> str:
    stem = Path(filename).stem
    stem = re.sub(r"(?i)^\s*(?:pea|silabo|sílabo)(?:\s*-\s*|\s+)", "", stem)
    stem = re.sub(r"(?i)^hom\s+", "", stem)
    stem = re.sub(r"(?i)\s*-\s*actualizado\s*$", "", stem)
    stem = re.sub(r"\s*\(\d+\)\s*$", "", stem)
    return _clean(stem)


def _extract_subject(lines: list[str], filename: str) -> tuple[str, bool]:
    subject = _value_after_label(
        lines,
        r"nombre\s+de\s+la\s+asignatura",
        (
            r"nivel\s+de\s+la\s+asignatura",
            r"unidad\s+de\s+organizaci[oó]n",
            r"campo\s+de\s+formaci[oó]n",
        ),
        max_following_lines=3,
    )
    if subject:
        return subject, False

    for index, line in enumerate(lines):
        if re.search(r"(?i)PEA\s+DE\s+LA\s+ASIGNATURA", line) and index + 1 < len(lines):
            candidate = lines[index + 1]
            if candidate and not _is_document_noise(candidate):
                return _clean(candidate), False
    return _filename_subject(filename), True


def _extract_section(lines: list[str], start_pattern: str, stop_patterns: Iterable[str]) -> str:
    start = re.compile(start_pattern, flags=re.IGNORECASE)
    stops = [re.compile(pattern, flags=re.IGNORECASE) for pattern in stop_patterns]
    collected: list[str] = []
    active = False
    for line in lines:
        if not active:
            match = start.search(line)
            if not match:
                continue
            active = True
            rest = line[match.end() :].strip(" :-")
            if rest:
                collected.append(rest)
            continue
        if any(stop.search(line) for stop in stops):
            break
        if not _is_document_noise(line):
            collected.append(line)

    if not collected:
        return ""

    items: list[str] = []
    prose: list[str] = []
    bullet_pattern = re.compile(r"^[\u2022\uf0b7\uf0fc\u2713\u25aa\-]+\s*")
    for line in collected:
        if bullet_pattern.match(line):
            if prose:
                items.append(_clean(" ".join(prose)))
                prose = []
            items.append(bullet_pattern.sub("", line).strip())
        else:
            prose.append(line)
    if prose:
        items.append(_clean(" ".join(prose)))
    items = [item for item in items if item]
    if len(items) > 1:
        return "\n".join(f"{index}. {item}" for index, item in enumerate(items, start=1))
    return items[0] if items else ""


def _extract_units(lines: list[str]) -> list[dict[str, Any]]:
    unit_pattern = re.compile(r"^\s*unidad\s+(\d{1,2})\s*[:.\-–]\s*(.*)$", flags=re.IGNORECASE)
    boundary_pattern = re.compile(
        r"^(?:resultado\s+de\s+aprendizaje|contenidos?|horas?\s+de\s+la\s+unidad|"
        r"observaciones?|tema|semana|docencia|pr[aá]ctica|aut[oó]nomo|componente|actividad\s+calificada)",
        flags=re.IGNORECASE,
    )
    bullet_pattern = re.compile(r"^[\u2022\uf0b7\uf0fc\u2713\u25aa\-]")
    result_pattern = re.compile(
        r"resultados?\s+de\s+aprendizaje(?:\s+de\s+la\s+unidad)?\s*:?\s*(.*)",
        flags=re.IGNORECASE,
    )
    units: list[dict[str, Any]] = []
    units_by_number: dict[int, dict[str, Any]] = {}

    for index, line in enumerate(lines):
        match = unit_pattern.match(line)
        if not match:
            continue
        number = int(match.group(1))
        name_parts = [_clean(match.group(2))] if _clean(match.group(2)) else []
        cursor = index + 1
        while cursor < len(lines) and len(name_parts) < 4:
            candidate = lines[cursor]
            if unit_pattern.match(candidate) or boundary_pattern.match(candidate) or bullet_pattern.match(candidate):
                break
            if _is_document_noise(candidate) or re.fullmatch(r"\d+(?:\s+\d+)*", candidate):
                break
            name_parts.append(candidate)
            cursor += 1

        name = _clean(" ".join(name_parts))
        if not name:
            continue

        result = ""
        result_start = None
        for result_index in range(index + 1, min(len(lines), index + 12)):
            result_match = result_pattern.search(lines[result_index])
            if result_match:
                result_start = result_index
                result_parts = [_clean(result_match.group(1))] if _clean(result_match.group(1)) else []
                for result_line in lines[result_index + 1 :]:
                    if unit_pattern.match(result_line) or re.match(r"(?i)^contenidos?\b", result_line):
                        break
                    if not _is_document_noise(result_line):
                        result_parts.append(result_line)
                result = _clean(" ".join(result_parts))
                break
            if unit_pattern.match(lines[result_index]):
                break
        if result_start is None:
            result = ""

        existing = units_by_number.get(number)
        if existing:
            if not existing["name"] and name:
                existing["name"] = name[:600]
            if not existing["learning_outcome"] and result:
                existing["learning_outcome"] = result[:4000]
            continue

        unit = {"number": number, "name": name[:600], "learning_outcome": result[:4000]}
        units.append(unit)
        units_by_number[number] = unit

    return units


def _learning_outcomes_from_units(units: list[dict[str, Any]]) -> str:
    ordered_units = sorted(units, key=lambda unit: int(unit.get("number") or 0))
    return "\n".join(
        f"{unit['number']}. {_clean(unit.get('learning_outcome'))}"
        for unit in ordered_units
        if _clean(unit.get("learning_outcome"))
    )


def _detect_academic_document_type(filename: str, lines: list[str]) -> str:
    filename_value = _normalized(Path(filename).stem)
    if any(token in filename_value.split() for token in ("silabo", "syllabus")):
        return "SILABO"
    first_page = " ".join(_normalized(line) for line in lines[:40])
    if "silabo de la asignatura" in first_page or "syllabus de la asignatura" in first_page:
        return "SILABO"
    return "PEA"


def _positioned_rows(
    items: Iterable[tuple[float, float, float, str]],
) -> list[list[tuple[float, float, float, str]]]:
    rows: list[list[tuple[float, float, float, str]]] = []
    for item in sorted(items, key=lambda value: (-value[1], value[0])):
        if rows and abs(rows[-1][0][1] - item[1]) <= 1.6:
            rows[-1].append(item)
        else:
            rows.append([item])
    return rows


def _positioned_row_text(row: Iterable[tuple[float, float, float, str]]) -> str:
    return _clean(" ".join(item[3] for item in sorted(row, key=lambda value: value[0])))


def _extract_syllabus_outcomes(layout_pages: list[str]) -> str:
    for page_text in layout_pages:
        lines = page_text.splitlines()
        for index, line in enumerate(lines):
            normalized_line = _normalized(line)
            if "perfil de egreso" not in normalized_line or "resultado" not in normalized_line:
                continue
            result_match = re.search(r"(?i)resultados?", line)
            evidence_match = re.search(r"(?i)evidencia", line)
            if not result_match or not evidence_match or evidence_match.start() <= result_match.start():
                continue

            result_start = result_match.start()
            evidence_start = evidence_match.start()
            collected: list[str] = []
            for following in lines[index + 1 :]:
                following_normalized = _normalized(following)
                if following_normalized.startswith("se considera copia") or (
                    "contenidos de la asignatura" in following_normalized
                ):
                    break
                value = _clean(following[result_start:evidence_start])
                if _normalized(value) in {"", "de la asignatura", "resultado de aprendizaje"}:
                    continue
                collected.append(value)
            result = _clean(" ".join(collected))
            if len(result) >= 30:
                return result
    return ""


def _extract_syllabus_units(positioned_pages: list[PdfTextPage]) -> list[dict[str, Any]]:
    start_page: int | None = None
    header_y = 0.0
    week_x = 0.0
    for page_index, page in enumerate(positioned_pages):
        if "contenidos de la asignatura" not in _normalized(page.text):
            continue
        unit_headers = [item for item in page.items if _normalized(item[3]) == "unidad"]
        topic_headers = [item for item in page.items if _normalized(item[3]) == "tema"]
        week_headers = [item for item in page.items if _normalized(item[3]) == "semana"]
        if not unit_headers or not topic_headers:
            continue
        unit_header = unit_headers[0]
        topic_header = min(topic_headers, key=lambda item: abs(item[1] - unit_header[1]))
        week_header = (
            min(week_headers, key=lambda item: abs(item[1] - unit_header[1]))
            if week_headers
            else None
        )
        start_page = page_index
        header_y = min(unit_header[1], topic_header[1])
        week_x = week_header[0] if week_header else page.width * 0.48
        break
    if start_page is None:
        return []

    stop_headings = (
        "estrategias metodologicas",
        "formacion ciudadana",
        "recursos didacticos",
        "bibliografia",
    )
    ignored_fragments = {
        "unidad",
        "tema",
        "semana",
        "no horas",
        "docencia",
        "practica",
        "autonomo",
    }
    candidates: list[tuple[int, float, float, float, str]] = []
    final_page = len(positioned_pages) - 1
    for page_index in range(start_page, len(positioned_pages)):
        page = positioned_pages[page_index]
        stop_y: float | None = None
        for row in _positioned_rows(page.items):
            row_value = _normalized(_positioned_row_text(row))
            if any(row_value.startswith(heading) for heading in stop_headings):
                stop_y = row[0][1]
                break

        minimum_y = max(50.0, (stop_y + 2.0) if stop_y is not None else 50.0)
        maximum_y = header_y - 5.0 if page_index == start_page else page.height - 50.0
        for x, y, font_size, text in page.items:
            normalized_text = _normalized(text)
            if not (minimum_y <= y <= maximum_y) or x >= week_x:
                continue
            if not any(character.isalpha() for character in text):
                continue
            if normalized_text in ignored_fragments:
                continue
            candidates.append((page_index, x, y, font_size, text))
        if stop_y is not None:
            final_page = page_index
            break

    anchor_counts: dict[float, int] = {}
    for _page_index, x, _y, _font_size, _text in candidates:
        anchor = round(x, 1)
        anchor_counts[anchor] = anchor_counts.get(anchor, 0) + 1
    frequent_anchors = sorted(
        (anchor for anchor, count in anchor_counts.items() if count >= 2),
    )
    if not frequent_anchors:
        return []
    unit_left = frequent_anchors[0]
    topic_options = [anchor for anchor in frequent_anchors if anchor - unit_left >= 45.0]
    if not topic_options:
        return []
    topic_left = max(topic_options, key=lambda anchor: anchor_counts[anchor])

    selected = [
        item
        for item in candidates
        if unit_left - 8.0 <= item[1] < topic_left - 2.0
    ]
    names: list[str] = []
    normalized_names: set[str] = set()
    current: list[str] = []

    def flush_current() -> None:
        value = _clean(" ".join(current)).strip(" :-")
        current.clear()
        normalized_value = _normalized(value)
        if not value or normalized_value.startswith(("evaluacion", "recuperacion")):
            return
        if normalized_value in normalized_names:
            return
        normalized_names.add(normalized_value)
        names.append(value)

    for page_index in range(start_page, final_page + 1):
        page_items = [
            (x, y, font_size, text)
            for candidate_page, x, y, font_size, text in selected
            if candidate_page == page_index
        ]
        previous_y: float | None = None
        for row in _positioned_rows(page_items):
            row_y = sum(item[1] for item in row) / len(row)
            row_font_size = max(item[2] for item in row)
            row_value = _positioned_row_text(row)
            if previous_y is not None and previous_y - row_y <= max(20.0, row_font_size * 1.9):
                current.append(row_value)
            else:
                flush_current()
                current.append(row_value)
            previous_y = row_y
        flush_current()

    return [
        {"number": index, "name": name[:600], "learning_outcome": ""}
        for index, name in enumerate(names, start=1)
    ]


def parse_pea_pdf(data: bytes, filename: str, index: int = 0) -> dict[str, Any]:
    warnings: list[str] = []
    selectable_text, page_count, selectable_warnings, positioned_pages = _extract_selectable_pdf_text(data)
    warnings.extend(selectable_warnings)
    text = selectable_text
    method = "TEXTO"
    normalized_selectable = _normalized(selectable_text)
    text_is_sufficient = (
        len(normalized_selectable) >= 500
        and "asignatura" in normalized_selectable
        and any(
            marker in normalized_selectable
            for marker in ("unidad 1", "resultados de aprendizaje", "contenidos de la asignatura")
        )
    )
    if not text_is_sufficient:
        ocr_text = _extract_ocr_pdf_text(data)
        if ocr_text:
            text = f"{selectable_text}\n{ocr_text}" if selectable_text else ocr_text
            method = "TEXTO+OCR" if selectable_text.strip() else "OCR"
        elif not ocr_available():
            warnings.append("El PDF requiere OCR, pero Tesseract no está disponible en el servidor.")
        else:
            warnings.append("El OCR no devolvió texto suficiente para este documento.")

    lines = _document_lines(text)
    document_type = _detect_academic_document_type(filename, lines)
    document_label = "El sílabo" if document_type == "SILABO" else "El PEA"
    subject_name, subject_from_filename = _extract_subject(lines, filename)
    if subject_from_filename:
        warnings.append("El nombre de la asignatura se tomó del nombre del archivo.")

    code = _value_after_label(
        lines,
        r"c[oó]digo\s+de\s+la\s+asignatura",
        (r"nombre\s+de\s+la\s+asignatura", r"nivel\s+de\s+la\s+asignatura"),
    )
    career = _value_after_label(
        lines,
        r"(?:^|\s)carrera\s*:",
        (r"tipos?\s+de\s+asignatura", r"datos\s+generales", r"escuela\s*:"),
    )
    field = _value_after_label(
        lines,
        r"campo\s+de\s+formaci[oó]n",
        (r"distribuci[oó]n\s+de\s+horas", r"docencia\s*:"),
        max_following_lines=1,
    )
    general_outcomes = _extract_section(
        lines,
        r"resultados?\s+de\s+aprendizaje\s+de\s+la\s+asignatura(?:\s+y\s+como\s+aporta\s+al\s+perfil\s+profesional)?",
        (r"^alineamiento\s+curricular", r"^misi[oó]n\s+intec", r"^contenidos\s+de\s+la\s+asignatura"),
    )
    units = _extract_units(lines)
    if document_type == "SILABO" and (not general_outcomes or not units):
        layout_pages = _extract_layout_pdf_pages(data)
        if not general_outcomes:
            general_outcomes = _extract_syllabus_outcomes(layout_pages)
        if not units:
            units = _extract_syllabus_units(positioned_pages)
    unit_outcomes = _learning_outcomes_from_units(units)
    outcomes = unit_outcomes
    if unit_outcomes:
        missing_outcomes = [
            str(unit["number"])
            for unit in units
            if not _clean(unit.get("learning_outcome"))
        ]
        if missing_outcomes:
            warnings.append(
                "No se identificó un resultado de aprendizaje para las unidades: "
                + ", ".join(missing_outcomes)
                + "."
            )
    elif general_outcomes:
        warnings.append(
            "El resultado general se omitió porque no se identificaron resultados de aprendizaje por unidad."
        )
    minimum_contents = "\n".join(
        f"UNIDAD {unit['number']}: {unit['name']}"
        for unit in units
    )

    if not subject_name:
        warnings.append("No se identificó la asignatura.")
    if not field:
        warnings.append(f"{document_label} no contiene un campo de formación identificable.")
    if not outcomes:
        warnings.append("No se identificaron resultados de aprendizaje.")
    if not units:
        warnings.append("No se identificaron unidades en el documento.")

    confidence = 0
    confidence += 30 if subject_name and not subject_from_filename else 14 if subject_name else 0
    confidence += 10 if code else 0
    confidence += 10 if career else 0
    confidence += 15 if field else 0
    confidence += 20 if outcomes else 0
    confidence += 15 if units else 0
    return {
        "index": index,
        "filename": filename,
        "document_type": document_type,
        "method": method,
        "page_count": page_count,
        "course_code": code[:180],
        "subject_name": subject_name[:400],
        "career_name": career[:300],
        "field": field[:1000],
        "learning_outcomes": outcomes[:12000],
        "minimum_contents": minimum_contents[:12000],
        "units": units,
        "confidence": min(confidence, 100),
        "warnings": warnings,
    }


def _subject_similarity(left: str, right: str) -> float:
    normalized_left = _normalized(left)
    normalized_right = _normalized(right)
    if not normalized_left or not normalized_right:
        return 0.0
    if normalized_left == normalized_right:
        return 1.0
    sequence = SequenceMatcher(None, normalized_left, normalized_right).ratio()
    left_tokens = set(normalized_left.split())
    right_tokens = set(normalized_right.split())
    union = left_tokens | right_tokens
    token_score = len(left_tokens & right_tokens) / len(union) if union else 0.0
    containment = 0.93 if normalized_left in normalized_right or normalized_right in normalized_left else 0.0
    return max(sequence, (sequence * 0.6) + (token_score * 0.4), containment)


def _assign_documents(
    subjects: list[dict[str, Any]],
    documents: list[dict[str, Any]],
) -> dict[int, list[tuple[int, float]]]:
    assignments: dict[int, list[tuple[int, float]]] = {}
    for document_index, document in enumerate(documents):
        candidates: list[tuple[float, int]] = []
        for subject_index, subject in enumerate(subjects):
            score = _subject_similarity(subject["subject_name"], document["subject_name"])
            if score < 0.84:
                continue
            candidates.append((score, subject_index))
        if not candidates:
            continue
        score, subject_index = max(candidates)
        assignments.setdefault(subject_index, []).append((document_index, score))
    return assignments


def _preferred_document(
    documents: list[dict[str, Any]],
    field: str,
    preferred_type: str,
) -> dict[str, Any] | None:
    candidates = [document for document in documents if _clean(document.get(field))]
    if not candidates:
        return None
    return max(
        candidates,
        key=lambda document: (
            document.get("document_type") == preferred_type,
            int(document.get("confidence") or 0),
            len(_clean(document.get(field))),
        ),
    )


def _unique_warnings(documents: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []
    seen: set[str] = set()
    for document in documents:
        for warning in document.get("warnings") or []:
            value = f"{document['filename']}: {warning}"
            normalized_value = _normalized(value)
            if normalized_value in seen:
                continue
            seen.add(normalized_value)
            warnings.append(value)
    return warnings


def analyze_curriculum(
    workbook_content: bytes,
    workbook_filename: str,
    academic_documents: list[tuple[str, bytes]],
    requested_career: str = "",
) -> dict[str, Any]:
    workbook, info = inspect_workbook(workbook_content, workbook_filename, requested_career)
    reference_layout: SheetLayout = info["target_layout"] or info["base_layout"]
    reference_sheet = workbook[reference_layout.sheet_name]
    documents: list[dict[str, Any]] = []
    for index, (filename, content) in enumerate(academic_documents):
        try:
            documents.append(parse_pea_pdf(content, filename, index))
        except ValueError as exc:
            documents.append(
                {
                    "index": index,
                    "filename": filename,
                    "document_type": _detect_academic_document_type(filename, []),
                    "method": "ERROR",
                    "page_count": 0,
                    "course_code": "",
                    "subject_name": _filename_subject(filename),
                    "career_name": "",
                    "field": "",
                    "learning_outcomes": "",
                    "minimum_contents": "",
                    "units": [],
                    "confidence": 0,
                    "warnings": [str(exc)],
                }
            )

    subjects: list[dict[str, Any]] = []
    for row in reference_layout.subject_rows:
        subjects.append(
            {
                "row_number": row,
                "subject_name": _clean(reference_sheet.cell(row, reference_layout.subject_col).value),
                "period": _clean(reference_sheet.cell(row, reference_layout.period_col).value),
                "curricular_unit": _clean(reference_sheet.cell(row, reference_layout.curricular_unit_col).value),
                "current": {
                    "field": _compact_multiline(reference_sheet.cell(row, reference_layout.field_col).value)
                    if reference_layout.field_col else "",
                    "learning_outcomes": _compact_multiline(reference_sheet.cell(row, reference_layout.outcomes_col).value)
                    if reference_layout.outcomes_col else "",
                    "minimum_contents": _compact_multiline(reference_sheet.cell(row, reference_layout.contents_col).value)
                    if reference_layout.contents_col else "",
                },
            }
        )

    assignments = _assign_documents(subjects, documents)
    used_documents: set[int] = set()
    rows: list[dict[str, Any]] = []
    for subject_index, subject in enumerate(subjects):
        assignment = assignments.get(subject_index)
        if not assignment:
            current_has_data = any(subject["current"].values())
            rows.append(
                {
                    **subject,
                    "document_index": None,
                    "document_indices": [],
                    "source_file": "",
                    "source_files": [],
                    "match_score": 0,
                    "match_type": "SIN_PEA",
                    "status": "DATOS_EXISTENTES" if current_has_data else "SIN_PEA",
                    "apply_recommended": False,
                    "proposal": dict(subject["current"]),
                    "warnings": [],
                }
            )
            continue

        assigned_documents = [documents[document_index] for document_index, _score in assignment]
        for document_index, _score in assignment:
            used_documents.add(document_index)
        match_score = max(score for _document_index, score in assignment)
        primary_document = max(
            assigned_documents,
            key=lambda document: (
                document.get("document_type") == "PEA",
                int(document.get("confidence") or 0),
            ),
        )
        if match_score >= 0.995:
            status_value = "LISTO"
            match_type = "NOMBRE_EXACTO"
        else:
            status_value = "LISTO"
            match_type = "NOMBRE_SIMILAR"

        field_document = _preferred_document(assigned_documents, "field", "PEA")
        outcomes_document = _preferred_document(assigned_documents, "learning_outcomes", "PEA")
        contents_document = _preferred_document(assigned_documents, "minimum_contents", "SILABO")
        proposal = {
            "field": (field_document or {}).get("field") or subject["current"]["field"],
            "learning_outcomes": (outcomes_document or {}).get("learning_outcomes")
            or subject["current"]["learning_outcomes"],
            "minimum_contents": (contents_document or {}).get("minimum_contents")
            or subject["current"]["minimum_contents"],
        }
        row_warnings = _unique_warnings(assigned_documents)
        source_documents = assigned_documents
        source_files = [document["filename"] for document in source_documents]
        if len({document.get("document_type") for document in source_documents}) > 1:
            row_warnings.insert(
                0,
                "Se combinaron PEA y sílabo: resultados priorizados del PEA y contenidos mínimos del sílabo.",
            )
        rows.append(
            {
                **subject,
                "document_index": primary_document["index"],
                "document_indices": [document["index"] for document in source_documents],
                "source_file": " | ".join(source_files),
                "source_files": source_files,
                "match_score": round(match_score * 100, 1),
                "match_type": match_type,
                "status": status_value,
                "apply_recommended": status_value == "LISTO" and bool(
                    proposal["learning_outcomes"] or proposal["minimum_contents"]
                ),
                "proposal": proposal,
                "warnings": row_warnings,
            }
        )

    unmatched_documents = [document for index, document in enumerate(documents) if index not in used_documents]
    summary = {
        "subjects": len(rows),
        "documents": len(documents),
        "ready": sum(row["status"] == "LISTO" for row in rows),
        "requires_review": sum(row["status"] == "REQUIERE_REVISION" for row in rows),
        "without_pea": sum(row["status"] == "SIN_PEA" for row in rows),
        "existing_data": sum(row["status"] == "DATOS_EXISTENTES" for row in rows),
        "unmatched_documents": len(unmatched_documents),
    }
    public_workbook = {
        key: value
        for key, value in info.items()
        if key not in {"base_layout", "target_layout", "blank_target"}
    }
    return {
        "workbook": public_workbook,
        "ocr_available": ocr_available(),
        "documents": documents,
        "rows": rows,
        "unmatched_documents": unmatched_documents,
        "summary": summary,
    }


def _copy_template_cell(source: Any, target: Any) -> None:
    target.value = source.value
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def _load_malla_adm_template() -> Any:
    try:
        content = _MALLA_ADM_TEMPLATE_PATH.read_bytes()
    except OSError as exc:
        raise ValueError("No se encontró la plantilla institucional Malla ADM.") from exc
    if sha256(content).hexdigest() != _MALLA_ADM_TEMPLATE_SHA256:
        raise ValueError("La plantilla institucional Malla ADM fue modificada y no puede utilizarse.")
    try:
        return load_workbook(BytesIO(content), data_only=False, keep_links=False)
    except Exception as exc:
        raise ValueError("No se pudo abrir la plantilla institucional Malla ADM.") from exc


def _clone_malla_adm_sheet(workbook: Any, title: str, position: int) -> Worksheet:
    template_workbook = _load_malla_adm_template()
    template = template_workbook[_MALLA_ADM_TEMPLATE_SHEET]
    workbook.loaded_theme = template_workbook.loaded_theme
    target = workbook.create_sheet(title=title, index=position)

    for row in template.iter_rows(min_row=1, max_row=template.max_row, min_col=1, max_col=template.max_column):
        for source_cell in row:
            _copy_template_cell(source_cell, target.cell(source_cell.row, source_cell.column))
    for cell_range in template.merged_cells.ranges:
        target.merge_cells(str(cell_range))

    target.sheet_format = copy(template.sheet_format)
    target.sheet_properties = copy(template.sheet_properties)
    target.page_margins = copy(template.page_margins)
    target.page_setup = copy(template.page_setup)
    target.print_options = copy(template.print_options)
    target.views = copy(template.views)
    target.freeze_panes = template.freeze_panes
    target.sheet_state = template.sheet_state
    target.auto_filter = copy(template.auto_filter)

    for index, source_dimension in template.row_dimensions.items():
        dimension = target.row_dimensions[index]
        dimension.height = source_dimension.height
        dimension.hidden = source_dimension.hidden
        dimension.outlineLevel = source_dimension.outlineLevel
        dimension.collapsed = source_dimension.collapsed
    for key, source_dimension in template.column_dimensions.items():
        dimension = target.column_dimensions[key]
        dimension.width = source_dimension.width
        dimension.hidden = source_dimension.hidden
        dimension.bestFit = source_dimension.bestFit
        dimension.outlineLevel = source_dimension.outlineLevel
        dimension.collapsed = source_dimension.collapsed

    return target


def _translate_formula(value: Any, origin: str, destination: str) -> Any:
    if not isinstance(value, str) or not value.startswith("="):
        return value
    try:
        return Translator(value, origin=origin).translate_formula(destination)
    except Exception:
        return value


def _existing_enrichment(
    workbook: Any,
    layout: SheetLayout | None,
) -> dict[tuple[str, str], list[tuple[Any, Any, Any]]]:
    values: dict[tuple[str, str], list[tuple[Any, Any, Any]]] = {}
    if not layout or not layout.enriched:
        return values
    sheet = workbook[layout.sheet_name]
    for row in layout.subject_rows:
        key = (
            _normalized(sheet.cell(row, layout.subject_col).value),
            _normalized(sheet.cell(row, layout.period_col).value),
        )
        values.setdefault(key, []).append(
            (
                sheet.cell(row, layout.field_col).value,
                sheet.cell(row, layout.outcomes_col).value,
                sheet.cell(row, layout.contents_col).value,
            )
        )
    return values


def _base_malla_payload(sheet: Worksheet, layout: SheetLayout) -> list[tuple[int, int, Any]]:
    payload: list[tuple[int, int, Any]] = []
    source_core_start = 7 if layout.enriched else 4
    source_columns = ((1, 1), (2, 2), (3, 3)) + tuple(
        (source_core_start + offset, 7 + offset) for offset in range(10)
    )
    for source_row in range(layout.header_row + 2, sheet.max_row + 1):
        target_row = source_row - layout.header_row + 1
        populated = any(
            sheet.cell(source_row, source_column).value not in (None, "")
            for source_column, _target_column in source_columns
        )
        if target_row > 37:
            if populated:
                raise ValueError(
                    "La hoja Malla supera las 37 filas admitidas por la plantilla institucional Malla ADM."
                )
            continue
        for source_column, target_column in source_columns:
            value = sheet.cell(source_row, source_column).value
            if value in (None, ""):
                continue
            origin = f"{get_column_letter(source_column)}{source_row}"
            destination = f"{get_column_letter(target_column)}{target_row}"
            payload.append((target_row, target_column, _translate_formula(value, origin, destination)))
    return payload


def _unique_sheet_title(workbook: Any, desired: str) -> str:
    existing = {_normalized(sheet.title) for sheet in workbook.worksheets}
    if _normalized(desired) not in existing:
        return desired[:31]
    for index in range(2, 100):
        suffix = f" {index}"
        candidate = f"{desired[:31 - len(suffix)]}{suffix}"
        if _normalized(candidate) not in existing:
            return candidate
    return f"Malla actualizada {datetime.now().strftime('%H%M%S')}"[:31]


def _create_enriched_sheet(workbook: Any, info: dict[str, Any]) -> tuple[Worksheet, SheetLayout]:
    base_layout: SheetLayout = info["base_layout"]
    base_sheet = workbook[base_layout.sheet_name]
    payload = _base_malla_payload(base_sheet, base_layout)
    enrichment = _existing_enrichment(workbook, info.get("target_layout"))

    existing_layout: SheetLayout | None = info.get("target_layout")
    existing_name = existing_layout.sheet_name if existing_layout else ""
    blank_name = info.get("blank_target") or ""
    removable_name = existing_name or blank_name
    if removable_name and removable_name in workbook.sheetnames:
        removable = workbook[removable_name]
        target_position = workbook.index(removable)
        desired_title = removable.title
        workbook.remove(removable)
    else:
        target_position = workbook.index(base_sheet) if base_sheet in workbook.worksheets else 0
        desired_title = info.get("target_sheet") or _target_sheet_title(info["career_name"])

    target = _clone_malla_adm_sheet(
        workbook,
        _unique_sheet_title(workbook, desired_title),
        target_position,
    )
    for row, column, value in payload:
        target.cell(row, column).value = value

    for row in _MALLA_ADM_SUBJECT_ROWS:
        subject = _normalized(target.cell(row, 1).value)
        period = _normalized(target.cell(row, 2).value)
        if not subject:
            for column in range(16, 26):
                target.cell(row, column).value = None
            continue
        existing_values = enrichment.get((subject, period), [])
        if existing_values:
            field, outcomes, contents = existing_values.pop(0)
            target.cell(row, 4).value = field
            target.cell(row, 5).value = outcomes
            target.cell(row, 6).value = contents
        target.cell(row, 17).value = target.cell(row, 7).value

    layout = inspect_sheet(target)
    if not layout or not layout.enriched:
        raise ValueError("No se pudo completar la plantilla institucional Malla ADM con la malla cargada.")
    info["target_sheet"] = target.title
    info["target_will_be_created"] = info.get("target_layout") is None
    return target, layout


def generate_curriculum_workbook(
    workbook_content: bytes,
    workbook_filename: str,
    requested_career: str,
    updates: list[dict[str, Any]],
    generated_by: str,
) -> tuple[bytes, str, dict[str, Any]]:
    _ = generated_by
    workbook, info = inspect_workbook(workbook_content, workbook_filename, requested_career)
    if "Resultado OCR" in workbook.sheetnames:
        workbook.remove(workbook["Resultado OCR"])
    target, target_layout = _create_enriched_sheet(workbook, info)

    rows_by_name: dict[str, list[int]] = {}
    for row in target_layout.subject_rows:
        name = _normalized(target.cell(row, target_layout.subject_col).value)
        rows_by_name.setdefault(name, []).append(row)
    valid_rows = set(target_layout.subject_rows)
    applied = 0
    for raw_update in updates:
        update = dict(raw_update)
        subject_name = _clean(update.get("subject_name"))
        expected_row = int(update.get("row_number") or 0)
        normalized_subject = _normalized(subject_name)
        row: int | None = None
        if expected_row in valid_rows:
            expected_subject = _normalized(target.cell(expected_row, target_layout.subject_col).value)
            if expected_subject == normalized_subject:
                row = expected_row
        matching_rows = rows_by_name.get(normalized_subject, [])
        if row is None and len(matching_rows) == 1:
            row = matching_rows[0]
        if row is None and len(matching_rows) > 1:
            continue
        if not row:
            continue

        proposal = update.get("proposal") or {}
        if update.get("apply"):
            field = _safe_excel_text(proposal.get("field"))
            outcomes = _safe_excel_text(proposal.get("learning_outcomes"))
            contents = _safe_excel_text(proposal.get("minimum_contents"))
            target.cell(row, target_layout.field_col).value = field
            target.cell(row, target_layout.outcomes_col).value = outcomes
            target.cell(row, target_layout.contents_col).value = contents
            applied += 1
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = BytesIO()
    workbook.save(output)
    stem = re.sub(r"[^A-Za-z0-9_-]+", "-", Path(workbook_filename).stem).strip("-") or "malla"
    output_filename = f"{stem}-actualizada.xlsx"
    return output.getvalue(), output_filename, {
        "target_sheet": target.title,
        "target_created": info["target_layout"] is None,
        "format": "Malla ADM",
        "applied": applied,
        "total": len(updates),
    }
