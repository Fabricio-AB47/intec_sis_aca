from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from html import escape
from io import BytesIO
import json
from pathlib import Path
import re
import unicodedata
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage
from pydantic import BaseModel, Field
from reportlab.graphics import renderPDF
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Flowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from svglib.svglib import svg2rlg

from app.core.security import SessionUser, require_roles
from app.services.db import get_connection

router = APIRouter(prefix="/api/certificados", tags=["certificados"])

_CERTIFICATES_ACCESS = require_roles("ADMINISTRADOR", "ACADEMICO", "RECTOR")
_BACKEND_ROOT = Path(__file__).resolve().parents[2]
_PROJECT_ROOT = _BACKEND_ROOT.parent
_LOGO_PNG_PATH = _PROJECT_ROOT / "SisAcademicoV1" / "Intec" / "public" / "images" / "intec.png"
_LOGO_SVG_PATH = _PROJECT_ROOT / "frontend" / "public" / "Intec-Logowithslogangray.svg"
_PROMO_EXCLUDED_CODES = {"VGA-ID-2023-114", "VGA-ID-2023-115"}
_PROMO_EXCLUDED_NAMES = {"A1 - BEGINNER", "A1+ - ELEMENTARY"}
_GASTRONOMIA_MATRICULA_BASE = 100.0
_GASTRONOMIA_ARANCEL_BASE = 1000.0


class CertificateGeneratePayload(BaseModel):
    tipo_beca: str | None = ""
    tipo_certificado: str = "ambos"
    periodo: str | None = ""
    proximo_periodo: str | None = ""
    semestre: int | None = Field(default=None, ge=1, le=4)
    estudiantes: list[str] = Field(min_length=1, max_length=500)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _upper(value: Any) -> str:
    return _clean(value).upper()


def _text_key(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _upper(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def _is_gastronomia_career(value: Any) -> bool:
    return "GASTRONOMIA" in _text_key(value)


def _is_futuro_femenino_scholarship(value: Any) -> bool:
    key = _text_key(value)
    return "FUTURO" in key and "FEMENINO" in key


def _is_mintel_scholarship(value: Any) -> bool:
    return "MINTEL" in _text_key(value)


def _date_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _clean(value)
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text[:19]).date().isoformat()
    except ValueError:
        return text


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return float(value)
    text = _clean(value).replace(",", ".")
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _int_value(value: Any) -> int | None:
    number = _number(value)
    return int(number) if number is not None else None


def _percentage(value: Any) -> float:
    number = _number(value) or 0.0
    return max(0.0, min(number, 100.0))


def _safe_filename(value: Any) -> str:
    text = _clean(value).lower()
    text = re.sub(r"[^a-z0-9._-]+", "_", text)
    return text.strip("._") or "certificado"


def _certificate_ref(codigo_estud: Any, cod_anio_basica: Any = "", codigo_periodo: Any = "") -> str:
    return "|".join([_clean(codigo_estud), _clean(cod_anio_basica), _clean(codigo_periodo)])


def _parse_certificate_ref(value: Any) -> dict[str, str]:
    parts = [_clean(part) for part in _clean(value).split("|")]
    return {
        "codigo_estud": parts[0] if parts else "",
        "cod_anio_basica": parts[1] if len(parts) > 1 else "",
        "codigo_periodo": parts[2] if len(parts) > 2 else "",
    }


def _legacy_certificate_scholarship(cursor: Any, codigo_estud: str) -> dict[str, Any]:
    cursor.execute(
        """
        SELECT TOP (1)
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca))) AS tipo_beca,
            porcentaje_beca
        FROM dbo.Becas
        WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), codestud))) = ?
        ORDER BY
            CASE WHEN NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca))), '') IS NULL THEN 1 ELSE 0 END,
            TRY_CONVERT(decimal(10, 2), porcentaje_beca) DESC
        """,
        codigo_estud,
    )
    row = _one_from_cursor(cursor) or {}
    tipo_beca = _clean(row.get("tipo_beca")) or "Sin Beca"
    porcentaje_beca = 50.0 if _is_futuro_femenino_scholarship(tipo_beca) else _percentage(row.get("porcentaje_beca"))
    es_suzuki = "suzuki" in tipo_beca.lower()

    matricula_base = 75.0
    arancel_base = 725.0 if _is_mintel_scholarship(tipo_beca) else 750.0
    matricula_financiada = round(matricula_base * (porcentaje_beca / 100), 2)
    arancel_financiado = round(arancel_base * (porcentaje_beca / 100), 2)

    return {
        "tipo_beca": tipo_beca,
        "porcentaje_beca": porcentaje_beca,
        "es_suzuki": es_suzuki,
        "matricula_base": matricula_base,
        "arancel_base": arancel_base,
        "matricula_financiada": matricula_financiada,
        "arancel_financiado": arancel_financiado,
        "total_financiado": round(matricula_financiada + arancel_financiado, 2),
    }


def _certificate_costs_for_career(scholarship: dict[str, Any], carrera: Any) -> dict[str, float]:
    porcentaje_beca = _percentage(scholarship.get("porcentaje_beca"))
    tipo_beca = scholarship.get("tipo_beca")
    if _is_gastronomia_career(carrera):
        matricula_base = _GASTRONOMIA_MATRICULA_BASE
        arancel_base = _GASTRONOMIA_ARANCEL_BASE
    else:
        matricula_base = _number(scholarship.get("matricula_base")) or 0.0
        arancel_base = _number(scholarship.get("arancel_base")) or 0.0
    if _is_mintel_scholarship(tipo_beca):
        porcentaje_beca = 100.0
    matricula_financiada = round(matricula_base * (porcentaje_beca / 100), 2)
    arancel_financiado = round(arancel_base * (porcentaje_beca / 100), 2)
    return {
        "matricula_base": matricula_base,
        "arancel_base": arancel_base,
        "matricula_financiada": matricula_financiada,
        "arancel_financiado": arancel_financiado,
        "total_financiado": round(matricula_financiada + arancel_financiado, 2),
    }


def _matricula_certificate_block_reason(context: dict[str, Any]) -> str:
    if not context.get("cabecera_matricula"):
        return "No existe cabecera de matricula para generar el certificado de matricula"
    return ""


def _document_digits(value: Any) -> str:
    return re.sub(r"\D+", "", _clean(value))


def _split_documents(value: Any) -> list[str]:
    seen: set[str] = set()
    items: list[str] = []
    for raw in re.split(r"[\s,;|]+", _clean(value)):
        digits = _document_digits(raw)
        if digits and digits not in seen:
            seen.add(digits)
            items.append(digits)
    return items


def _active_sql(alias: str) -> str:
    return (
        f"("
        f"UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), {alias}.Estado)))) IN ('A', 'ACTIVO', 'ACTIVA')"
        f")"
    )


def _is_active_state(value: Any) -> bool:
    return _upper(value) in {"A", "ACTIVO", "ACTIVA"}


def _db_error_detail(action: str, exc: Exception) -> str:
    text = str(exc)
    lowered = text.lower()
    if "no such host" in lowered or "login timeout" in lowered or "sql server" in lowered and "server is not found" in lowered:
        return f"{action}. No se pudo conectar a SQL Server: revisa DB_HOST/DB_PORT en backend/.env y que el servidor este disponible."
    return action


def _semester_text(value: int | None) -> str:
    if value == 1:
        return "PRIMER SEMESTRE"
    if value == 2:
        return "SEGUNDO SEMESTRE"
    if value == 3:
        return "TERCER SEMESTRE"
    if value == 4:
        return "CUARTO SEMESTRE"
    return f"{value or ''} SEMESTRE".strip() or "SEMESTRE"


def _rows_from_cursor(cursor: Any) -> list[dict[str, Any]]:
    columns = [column[0] for column in cursor.description or []]
    return [{column: value for column, value in zip(columns, row)} for row in cursor.fetchall()]


def _one_from_cursor(cursor: Any) -> dict[str, Any] | None:
    columns = [column[0] for column in cursor.description or []]
    row = cursor.fetchone()
    return {column: value for column, value in zip(columns, row)} if row else None


def _period_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "cod_periodo": _clean(row.get("cod_periodo")),
        "detalle_periodo": _clean(row.get("detalle_periodo")),
        "fecha_inicio": _date_text(row.get("fecha_inicio")),
        "fecha_fin": _date_text(row.get("fecha_fin")),
        "orden": _int_value(row.get("orden")),
    }


def _period_meta(cursor: Any, codigo_periodo: str) -> dict[str, Any] | None:
    codigo = _clean(codigo_periodo)
    if not codigo:
        return None
    cursor.execute(
        """
        SELECT TOP (1)
            TRY_CONVERT(varchar(50), cod_periodo) AS cod_periodo,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Detalle_Periodo))) AS detalle_periodo,
            fechain AS fecha_inicio,
            fechafin AS fecha_fin,
            TRY_CONVERT(int, Orden) AS orden
        FROM dbo.PERIODO
        WHERE TRY_CONVERT(varchar(50), cod_periodo) = ?
        """,
        codigo,
    )
    row = _one_from_cursor(cursor)
    return _period_row(row) if row else None


def _active_student(cursor: Any, codigo_estud: str) -> bool:
    cursor.execute(
        """
        SELECT TOP (1)
            CE.Estado AS estado_correos,
            D.Estado AS estado_datos
        FROM dbo.DATOS_ESTUD D
        LEFT JOIN dbo.CorreosEstudIntec CE
            ON LTRIM(RTRIM(TRY_CONVERT(varchar(50), CE.codestud)))
             = LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.codigo_estud)))
        WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.codigo_estud))) = ?
        """,
        codigo_estud,
    )
    row = cursor.fetchone()
    return bool(row and (_is_active_state(row[0]) or _is_active_state(row[1])))


def _fetch_reprobadas(
    cursor: Any,
    codigo_periodo: str,
    student_codes: list[str],
    semestre: int | None = None,
) -> dict[str, list[dict[str, Any]]]:
    if not codigo_periodo or not student_codes:
        return {}

    details: dict[str, list[dict[str, Any]]] = {code: [] for code in student_codes}
    for start in range(0, len(student_codes), 700):
        chunk = student_codes[start : start + 700]
        placeholders = ", ".join("?" for _ in chunk)
        join_semestre = ""
        where_semestre = ""
        params: list[Any] = [codigo_periodo]
        if semestre is not None:
            join_semestre = """
                INNER JOIN dbo.PENSUM PS
                    ON LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), PS.codigo_materia)))
                       = LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), CX.codigo_materia)))
            """
            where_semestre = "AND TRY_CONVERT(int, PS.Semestre) = ?"
            params.append(semestre)
        params.extend(chunk)

        cursor.execute(
            f"""
            WITH REPROB AS (
                SELECT
                    LTRIM(RTRIM(TRY_CONVERT(varchar(50), CX.codigo_estud))) AS codigo_estud,
                    LTRIM(RTRIM(TRY_CONVERT(varchar(100), CX.codigo_materia))) AS codigo_materia,
                    TRY_CONVERT(float, REPLACE(TRY_CONVERT(varchar(60), CX.PromedioFinal), ',', '.')) AS promedio_final,
                    CX.caprueba,
                    CX.ControlAprueba,
                    P.Nomb_Materia AS nombre,
                    P.cod_materia,
                    ROW_NUMBER() OVER (
                        PARTITION BY
                            LTRIM(RTRIM(TRY_CONVERT(varchar(50), CX.codigo_estud))),
                            CX.codigo_periodo,
                            LTRIM(RTRIM(TRY_CONVERT(varchar(100), CX.codigo_materia)))
                        ORDER BY ISNULL(CX.num, 0) DESC, ISNULL(CX.Num_Matricula, 0) DESC
                    ) AS rn
                FROM dbo.CARRERAXESTUD CX
                LEFT JOIN dbo.PENSUM P
                    ON LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), P.codigo_materia)))
                       = LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), CX.codigo_materia)))
                   AND TRY_CONVERT(varchar(50), P.Cod_AnioBasica)
                       = TRY_CONVERT(varchar(50), CX.cod_anio_Basica)
                {join_semestre}
                WHERE TRY_CONVERT(varchar(50), CX.codigo_periodo) = ?
                    {where_semestre}
                    AND LTRIM(RTRIM(TRY_CONVERT(varchar(50), CX.codigo_estud))) IN ({placeholders})
                    AND (
                        TRY_CONVERT(float, REPLACE(TRY_CONVERT(varchar(60), CX.PromedioFinal), ',', '.')) < 7
                        OR TRY_CONVERT(int, CX.caprueba) = 0
                        OR TRY_CONVERT(int, CX.ControlAprueba) = 0
                        OR UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), CX.caprueba)))) IN ('REPROBADO', 'REPROBADA', 'NO', 'N')
                    )
            )
            SELECT codigo_estud, codigo_materia, promedio_final, caprueba, ControlAprueba, nombre, cod_materia
            FROM REPROB
            WHERE rn = 1
            ORDER BY codigo_estud, codigo_materia
            """,
            params,
        )
        for row in _rows_from_cursor(cursor):
            code = _clean(row.get("codigo_estud"))
            details.setdefault(code, []).append(
                {
                    "codigo_materia": _clean(row.get("codigo_materia")),
                    "cod_materia": _clean(row.get("cod_materia")),
                    "nombre": _clean(row.get("nombre")),
                    "promedioFinal": _number(row.get("promedio_final")),
                    "caprueba": _clean(row.get("caprueba")),
                    "controlAprueba": _clean(row.get("ControlAprueba")),
                }
            )
    return details


def _certificate_reprobadas(
    cursor: Any,
    codigo_periodo: str,
    codigo_estud: str,
    semestre: int | None = None,
) -> list[dict[str, Any]]:
    if not codigo_periodo or not codigo_estud:
        return []
    return _fetch_reprobadas(cursor, codigo_periodo, [codigo_estud], semestre).get(codigo_estud, [])


def _student_list_row(row: dict[str, Any], reprobadas: list[dict[str, Any]] | None) -> dict[str, Any]:
    count = len(reprobadas or [])
    code = _clean(row.get("codestud"))
    cod_carrera = _clean(row.get("cod_anio_basica"))
    codigo_periodo = _clean(row.get("codigo_periodo_matricula"))
    carrera = _clean(row.get("carrera"))
    has_matricula_header = bool(cod_carrera and codigo_periodo)
    matricula_allowed = bool(code and has_matricula_header)
    promo_allowed = bool(code and count == 0)
    matricula_block = ""
    if code and not matricula_allowed:
        matricula_block = "No existe cabecera de matricula para generar el certificado de matricula"
    return {
        "codestud": code,
        "certificado_ref": _certificate_ref(code, cod_carrera, codigo_periodo) if cod_carrera or codigo_periodo else code,
        "nombres": _clean(row.get("nombres")),
        "correo_personal": _clean(row.get("correo_personal")),
        "correo_intec": _clean(row.get("correo_intec")),
        "estado": _clean(row.get("estado")),
        "cod_anio_basica": cod_carrera,
        "carrera": carrera,
        "codigo_periodo_matricula": codigo_periodo,
        "periodo_matricula": _clean(row.get("periodo_matricula")),
        "num_matricula": _clean(row.get("num_matricula")),
        "reprobadas_count": count,
        "reprobadas_detalle": reprobadas or [],
        "puede_generar": bool(code and has_matricula_header),
        "puede_generar_matricula": matricula_allowed,
        "puede_generar_promocion": bool(promo_allowed and codigo_periodo),
        "motivo_bloqueo_matricula": matricula_block,
        "motivo_bloqueo": "" if promo_allowed else "Estudiante reprobado; no se permite generar certificado de promoción",
    }


def _logo_reader() -> tuple[ImageReader | None, BytesIO | None, float, float]:
    if _LOGO_PNG_PATH.exists():
        try:
            image = PILImage.open(_LOGO_PNG_PATH).convert("RGBA")
            bbox = image.getbbox()
            if bbox:
                image = image.crop(bbox)
            buffer = BytesIO()
            image.save(buffer, format="PNG")
            buffer.seek(0)
            reader = ImageReader(buffer)
            width, height = reader.getSize()
            return reader, buffer, float(width), float(height)
        except Exception:
            pass
    return None, None, 1.0, 1.0


class _LogoFlowable(Flowable):
    def __init__(self, width: float) -> None:
        super().__init__()
        self.hAlign = "CENTER"
        self.drawing = None
        self.reader = None
        self._buffer = None
        if _LOGO_SVG_PATH.exists():
            self.drawing = svg2rlg(str(_LOGO_SVG_PATH))
            self.scale = width / float(self.drawing.width or width)
            self.width = width
            self.height = float(self.drawing.height or 0) * self.scale
        else:
            self.reader, self._buffer, source_width, source_height = _logo_reader()
            if self.reader:
                self.width = width
                self.height = width * (source_height / max(source_width, 1.0))
                self.scale = 1.0
            else:
                self.width = width
                self.height = 0.55 * inch
                self.scale = 1.0

    def draw(self) -> None:
        if self.drawing:
            self.canv.saveState()
            self.canv.scale(self.scale, self.scale)
            clip = self.canv.beginPath()
            clip.rect(0, 0, float(self.drawing.width or self.width), float(self.drawing.height or self.height))
            self.canv.clipPath(clip, stroke=0)
            renderPDF.draw(self.drawing, self.canv, 0, 0)
            self.canv.restoreState()
            return
        if self.reader:
            self.canv.drawImage(self.reader, 0, 0, width=self.width, height=self.height, mask="auto")
            return
        self.canv.setFont("Helvetica-Bold", 22)
        self.canv.setFillColor(colors.HexColor("#808285"))
        self.canv.drawString(0, 0.2 * inch, "intec")


def _draw_logo_top(canv: Any, top_y: float, width: float) -> float:
    x = (letter[0] - width) / 2
    if _LOGO_SVG_PATH.exists():
        drawing = svg2rlg(str(_LOGO_SVG_PATH))
        scale = width / float(drawing.width or width)
        height = float(drawing.height or 0) * scale
        canv.saveState()
        canv.translate(x, top_y - height)
        canv.scale(scale, scale)
        clip = canv.beginPath()
        clip.rect(0, 0, float(drawing.width or width), float(drawing.height or height))
        canv.clipPath(clip, stroke=0)
        renderPDF.draw(drawing, canv, 0, 0)
        canv.restoreState()
        return height
    reader, _buffer, source_width, source_height = _logo_reader()
    if reader:
        height = width * (source_height / max(source_width, 1.0))
        canv.drawImage(reader, x, top_y - height, width=width, height=height, mask="auto")
        return height
    canv.setFont("Helvetica-Bold", 24)
    canv.setFillColor(colors.HexColor("#808285"))
    canv.drawCentredString(letter[0] / 2, top_y - 24, "intec")
    return 34


def _wrap_text(canv: Any, text: str, font_name: str, font_size: float, max_width: float) -> list[str]:
    words = [word for word in _clean(text).split(" ") if word]
    lines: list[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if current and canv.stringWidth(candidate, font_name, font_size) > max_width:
            lines.append(current)
            current = word
        else:
            current = candidate
    if current:
        lines.append(current)
    return lines or [""]


def _draw_wrapped(
    canv: Any,
    text: str,
    x: float,
    y: float,
    max_width: float,
    font_name: str,
    font_size: float,
    leading: float,
) -> float:
    canv.setFont(font_name, font_size)
    for line in _wrap_text(canv, text, font_name, font_size, max_width):
        canv.drawString(x, y, line)
        y -= leading
    return y


def _draw_rich_wrapped(
    canv: Any,
    parts: list[tuple[str, str]],
    x: float,
    y: float,
    max_width: float,
    font_size: float,
    leading: float,
) -> float:
    line: list[tuple[str, str]] = []
    line_width = 0.0

    def flush(current_y: float) -> float:
        nonlocal line, line_width
        draw_x = x
        for text, font_name in line:
            canv.setFont(font_name, font_size)
            canv.drawString(draw_x, current_y, text)
            draw_x += canv.stringWidth(text, font_name, font_size)
        line = []
        line_width = 0.0
        return current_y - leading

    for raw_text, font_name in parts:
        tokens = re.findall(r"\S+\s*", _clean(raw_text) + (" " if raw_text.endswith(" ") else ""))
        for token in tokens:
            token_width = canv.stringWidth(token, font_name, font_size)
            if line and line_width + token_width > max_width:
                y = flush(y)
                token = token.lstrip()
                token_width = canv.stringWidth(token, font_name, font_size)
            if token:
                line.append((token, font_name))
                line_width += token_width
    if line:
        y = flush(y)
    return y


def _pdf_styles() -> dict[str, ParagraphStyle]:
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="CertTitle",
            parent=styles["Title"],
            alignment=TA_CENTER,
            fontName="Times-Bold",
            fontSize=15,
            leading=18,
            textColor=colors.black,
            spaceAfter=8,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CertHeader",
            parent=styles["BodyText"],
            alignment=TA_CENTER,
            fontName="Times-Bold",
            fontSize=9,
            leading=10.5,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CertBody",
            parent=styles["BodyText"],
            fontName="Times-Roman",
            fontSize=8.5,
            leading=10.5,
        )
    )
    styles.add(
        ParagraphStyle(
            name="CertBodyBold",
            parent=styles["CertBody"],
            fontName="Times-Bold",
        )
    )
    styles.add(
        ParagraphStyle(
            name="CertCell",
            parent=styles["BodyText"],
            fontName="Times-Roman",
            fontSize=7.5,
            leading=9,
        )
    )
    return styles


def _paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(_clean(value)) or "-", style)


def _grade_state(value: Any) -> str:
    number = _number(value)
    return "APROBADO" if number is not None and number >= 7 else "REPROBADO"


def _filtered_promo_subjects(subjects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    for item in subjects:
        code = re.sub(r"\s+", "", _upper(item.get("codigo_materia") or item.get("cod_materia")))
        code_text = _upper(item.get("cod_materia"))
        name = _upper(item.get("nombre_materia"))
        if code in _PROMO_EXCLUDED_CODES or code_text in _PROMO_EXCLUDED_CODES:
            continue
        if name in _PROMO_EXCLUDED_NAMES:
            continue
        filtered.append(item)
    return filtered


def _promocion_story(context: dict[str, Any], styles: Any) -> list[Any]:
    story: list[Any] = [
        _LogoFlowable(1.45 * inch),
        Spacer(1, 0.08 * inch),
        Paragraph(
            "Instituto Superior Tecnológico de Técnicas Empresariales y del Conocimiento INTEC",
            styles["CertHeader"],
        ),
        Paragraph("Direc.: Calle Isla Seymour N44-91 y Av. Rio Coca", styles["CertHeader"]),
        Paragraph("Teléfono: +593 98 376 0020 / (02) 501-7702", styles["CertHeader"]),
        Spacer(1, 0.28 * inch),
        Paragraph("CERTIFICADO DE PROMOCIÓN", styles["CertTitle"]),
    ]

    period = context["periodo"]
    if period.get("fecha_inicio") or period.get("fecha_fin"):
        story.append(
            Paragraph(
                f"Del <b>{escape(period.get('fecha_inicio') or '-')}</b> al <b>{escape(period.get('fecha_fin') or '-')}</b>",
                styles["CertHeader"],
            )
        )

    meta_rows = [
        [
            Paragraph(f"<b>Nombre:</b> {escape(context['nombre'])}", styles["CertBody"]),
            Paragraph(f"<b>Cédula:</b> {escape(context['cedula'])}", styles["CertBody"]),
        ],
        [
            Paragraph(f"<b>Carrera:</b> {escape(context['carrera'])}", styles["CertBody"]),
            Paragraph(f"<b>Periodo:</b> {escape(period.get('detalle_periodo') or '-')}", styles["CertBody"]),
        ],
    ]
    scholarship_label = "Incentivos tributarios patrocinado por" if context.get("es_suzuki") else "Beca"
    scholarship_text = f"<b>{scholarship_label}:</b> {escape(context['tipo_beca'])} ({context['porcentaje_beca']:.0f}%)"
    if context.get("periodo_matricula", {}).get("detalle_periodo"):
        meta_rows.append(
            [
                Paragraph(
                    f"<b>Próximo periodo:</b> {escape(context['periodo_matricula']['detalle_periodo'])}",
                    styles["CertBody"],
                ),
                Paragraph(scholarship_text, styles["CertBody"]),
            ]
        )
    else:
        meta_rows.append(
            [
                Paragraph(scholarship_text, styles["CertBody"]),
                "",
            ]
        )
    story.extend([Spacer(1, 0.18 * inch), Table(meta_rows, colWidths=[3.7 * inch, 3.0 * inch]), Spacer(1, 0.2 * inch)])

    table_rows: list[list[Any]] = [
        [
            Paragraph("<b>Semestre</b>", styles["CertCell"]),
            Paragraph("<b>Código</b>", styles["CertCell"]),
            Paragraph("<b>Materia</b>", styles["CertCell"]),
            Paragraph("<b>Créditos</b>", styles["CertCell"]),
            Paragraph("<b>Nota</b>", styles["CertCell"]),
            Paragraph("<b>Estado</b>", styles["CertCell"]),
        ]
    ]
    for item in _filtered_promo_subjects(context["subjects"]):
        semester = context.get("semestre") or _int_value(item.get("semestre"))
        table_rows.append(
            [
                _paragraph(_semester_text(semester), styles["CertCell"]),
                _paragraph(item.get("cod_materia") or item.get("codigo_materia"), styles["CertCell"]),
                _paragraph(item.get("nombre_materia"), styles["CertCell"]),
                _paragraph(item.get("creditos"), styles["CertCell"]),
                _paragraph(item.get("promedio_final"), styles["CertCell"]),
                _paragraph(_grade_state(item.get("promedio_final")), styles["CertCell"]),
            ]
        )

    if len(table_rows) == 1:
        table_rows.append([Paragraph("Sin materias para mostrar.", styles["CertCell"]), "", "", "", "", ""])

    table = Table(table_rows, colWidths=[1.05 * inch, 0.88 * inch, 2.45 * inch, 0.65 * inch, 0.62 * inch, 0.95 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#600000")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Times-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    story.extend(
        [
            Spacer(1, 0.18 * inch),
            Paragraph(
                "El estudiante para aprobar una asignatura debe obtener un mínimo de 7/10 y cumplir con el reglamento respectivo.",
                styles["CertBody"],
            ),
            Spacer(1, 0.4 * inch),
            Paragraph("<b>Ing. Verónica Cevallos Calderón Mgtr.</b>", styles["CertHeader"]),
            Paragraph("<b>VICERRECTORA GENERAL ACADÉMICA</b>", styles["CertHeader"]),
        ]
    )
    return story


def _build_promocion_pdf(context: dict[str, Any]) -> bytes:
    styles = _pdf_styles()
    story = _promocion_story(context, styles)
    output = BytesIO()
    SimpleDocTemplate(
        output,
        pagesize=letter,
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.45 * inch,
        title="Certificado de Promoción",
    ).build(story)
    output.seek(0)
    return output.getvalue()


def _build_promocion_pdf_bundle(contexts: list[dict[str, Any]]) -> bytes:
    styles = _pdf_styles()
    story: list[Any] = []
    for index, context in enumerate(contexts):
        if index:
            story.append(PageBreak())
        story.extend(_promocion_story(context, styles))

    output = BytesIO()
    SimpleDocTemplate(
        output,
        pagesize=letter,
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.45 * inch,
        title="Certificados de Promoción",
    ).build(story)
    output.seek(0)
    return output.getvalue()


def _draw_signature(canv: Any, center_x: float, top_y: float, width: float) -> None:
    line_width = min(width, 150)
    canv.setStrokeColor(colors.HexColor("#7a1414"))
    canv.setLineWidth(1)
    canv.line(center_x - line_width / 2, top_y, center_x + line_width / 2, top_y)
    canv.setFillColor(colors.black)
    canv.setFont("Times-Bold", 10.8)
    canv.drawCentredString(center_x, top_y - 20, "Ing. Verónica Cevallos Calderón Mgtr.")
    canv.setFont("Times-Roman", 11.2)
    canv.drawCentredString(center_x, top_y - 42, "VICERRECTORA GENERAL ACADÉMICA")


def _draw_matricula_certificate_page(canv: Any, context: dict[str, Any]) -> None:
    page_width, page_height = letter
    left = 0.9 * inch
    max_width = page_width - (1.8 * inch)

    logo_top = page_height - 1.08 * inch
    logo_height = _draw_logo_top(canv, logo_top, 82)
    header_y = logo_top - logo_height - 42

    canv.setFillColor(colors.black)
    canv.setFont("Times-Bold", 8.8)
    canv.drawCentredString(
        page_width / 2,
        header_y,
        "Instituto Superior Tecnológico de Técnicas Empresariales y del Conocimiento INTEC",
    )
    canv.setFont("Times-Bold", 8.5)
    canv.drawCentredString(page_width / 2, header_y - 12, "Direc.: Calle Isla Seymour N44-91 y Av. Río Coca")
    canv.drawCentredString(page_width / 2, header_y - 24, "Teléfono: +593 98 376 0020 / (02) 501-7702")

    title_y = header_y - 54
    canv.setFont("Times-Bold", 15.2)
    canv.drawCentredString(page_width / 2, title_y, "CERTIFICADO DE MATRÍCULA")

    body_y = title_y - 32
    body_y = _draw_rich_wrapped(
        canv,
        [
            (
                "Previa revisión de nuestros registros y de la información suministrada, "
                "el Instituto Superior Tecnológico de Técnicas Empresariales y del Conocimiento INTEC "
                "certifica que el/la estudiante: ",
                "Times-Roman",
            ),
            (context["nombre"], "Times-Bold"),
            (" , titular de la Cédula No. ", "Times-Roman"),
            (context["cedula"], "Times-Bold"),
            (", se encuentra matrículado(a).", "Times-Roman"),
        ],
        left,
        body_y,
        max_width,
        8.6,
        9.8,
    )

    period = context["periodo_matricula"]
    detail_y = body_y - 18
    detail_lines = [
        f"Carrera: {context['carrera']}",
        f"Semestre: {_semester_text(context['semestre_matricula'])}",
        f"Periodo Académico: {period.get('detalle_periodo') or context['periodo'].get('detalle_periodo') or '-'}",
        f"Fecha de inicio: {period.get('fecha_inicio') or '-'}",
        f"Fecha fin: {period.get('fecha_fin') or '-'}",
    ]
    for line in detail_lines:
        detail_y = _draw_wrapped(canv, line, left, detail_y, max_width, "Times-Bold", 8.8, 10.2)

    cost_y = detail_y - 18
    cost_y = _draw_rich_wrapped(
        canv,
        [
            ("Costo Matrícula (por periodo académico):", "Times-Bold"),
            (f"${context['matricula_base']:.2f}", "Times-Roman"),
        ],
        left,
        cost_y,
        max_width,
        8.8,
        10.2,
    )
    cost_y = _draw_rich_wrapped(
        canv,
        [
            ("Costo Arancel (por periodo académico):", "Times-Bold"),
            (f"${context['arancel_base']:.2f}", "Times-Roman"),
        ],
        left,
        cost_y,
        max_width,
        8.8,
        10.2,
    )

    if _percentage(context.get("porcentaje_beca")) > 0 or (_number(context.get("total_financiado")) or 0.0) > 0:
        cost_y = _draw_rich_wrapped(
            canv,
            [
                ("Costo Matrícula Financiada (por periodo académico):", "Times-Bold"),
                (f"${context['matricula_financiada']:.2f}", "Times-Roman"),
            ],
            left,
            cost_y,
            max_width,
            8.8,
            10.2,
        )
        cost_y = _draw_rich_wrapped(
            canv,
            [
                ("Costo Arancel Financiado (por periodo académico):", "Times-Bold"),
                (f"${context['arancel_financiado']:.2f}", "Times-Roman"),
            ],
            left,
            cost_y,
            max_width,
            8.8,
            10.2,
        )
        cost_y = _draw_rich_wrapped(
            canv,
            [
                ("Total Financiado (por periodo académico):", "Times-Bold"),
                (f"${context['total_financiado']:.2f}", "Times-Roman"),
            ],
            left,
            cost_y,
            max_width,
            8.8,
            10.2,
        )

    closing_y = cost_y - 18
    closing_y = _draw_wrapped(
        canv,
        "A petición del interesado/a extiendo el presente certificado para que lo utilice para los fines legales pertinentes.",
        left,
        closing_y,
        max_width,
        "Times-Roman",
        8.8,
        10.2,
    )
    _draw_signature(canv, page_width / 2, max(closing_y - 42, 210), 260)
    canv.setFont("Times-Roman", 10)
    canv.setFillColor(colors.black)
    canv.drawRightString(page_width - 0.45 * inch, 0.45 * inch, f"Fecha: {date.today().strftime('%d/%m/%Y')}")


def _build_matricula_pdf(context: dict[str, Any]) -> bytes:
    output = BytesIO()
    canv = pdf_canvas.Canvas(output, pagesize=letter)
    canv.setTitle("Certificado de Matrícula")
    _draw_matricula_certificate_page(canv, context)
    canv.showPage()
    canv.save()
    output.seek(0)
    return output.getvalue()


def _build_matricula_pdf_bundle(contexts: list[dict[str, Any]]) -> bytes:
    output = BytesIO()
    canv = pdf_canvas.Canvas(output, pagesize=letter)
    canv.setTitle("Certificados de Matrícula")
    for context in contexts:
        _draw_matricula_certificate_page(canv, context)
        canv.showPage()
    canv.save()
    output.seek(0)
    return output.getvalue()


_MATRICULA_EXCEL_COLUMNS = [
    "nombres_apellidos",
    "numero_cedula",
    "carrera",
    "semestre",
]


def _normalize_excel_header(value: Any) -> str:
    text = unicodedata.normalize("NFKD", _clean(value).lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _excel_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if _clean(value):
            return value
    return ""


def _matricula_excel_costs(carrera: str, row: dict[str, Any]) -> tuple[float, float]:
    del row
    matricula = _GASTRONOMIA_MATRICULA_BASE if _is_gastronomia_career(carrera) else 75.0
    arancel = _GASTRONOMIA_ARANCEL_BASE if _is_gastronomia_career(carrera) else 750.0
    return matricula, arancel


def _is_excluded_matricula_excel_career(value: Any) -> bool:
    key = _text_key(value)
    return "EDUCACION CONTINUA" in key or key in {"INGLES", "IDIOMAS INGLES"} or " INGLES " in f" {key} "


def _build_matricula_excel_template(periodos: list[dict[str, Any]], carreras: list[dict[str, Any]] | None = None) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Matriculas"
    sheet.append(_MATRICULA_EXCEL_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="600000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.append(["NOMBRE APELLIDO", "0102030405", "TECNOLOGIA SUPERIOR EN ADMINISTRACION", 1])
    sheet.append(["ESTUDIANTE EJEMPLO", "0999999999", "GASTRONOMIA", 2])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    career_sheet = workbook.create_sheet("Carreras")
    career_sheet.append(["carrera"])
    career_names: list[str] = []
    for item in carreras or []:
        name = _clean(item.get("carrera") or item.get("nombre") or item.get("Nombre_Basica"))
        if name and not _is_excluded_matricula_excel_career(name) and name not in career_names:
            career_names.append(name)
    if not career_names:
        career_names = [
            "TECNOLOGIA SUPERIOR EN ADMINISTRACION",
            "TECNOLOGIA SUPERIOR EN DESARROLLO DE SOFTWARE",
            "GASTRONOMIA",
        ]
    for name in career_names:
        career_sheet.append([name])
    for cell in career_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="600000")
    _autosize_worksheet(career_sheet)

    career_validation = DataValidation(
        type="list",
        formula1=f"'Carreras'!$A$2:$A${len(career_names) + 1}",
        allow_blank=False,
    )
    career_validation.error = "Selecciona una carrera de la hoja Carreras."
    career_validation.errorTitle = "Carrera no valida"
    sheet.add_data_validation(career_validation)
    career_validation.add("C2:C501")

    semester_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="4", allow_blank=False)
    semester_validation.error = "El semestre debe ser un numero entre 1 y 4."
    semester_validation.errorTitle = "Semestre no valido"
    sheet.add_data_validation(semester_validation)
    semester_validation.add("D2:D501")

    _autosize_worksheet(sheet)

    help_sheet = workbook.create_sheet("Instrucciones")
    help_rows = [
        ["Campo", "Detalle"],
        ["nombres_apellidos", "Obligatorio. Nombre completo tal como debe aparecer en el certificado."],
        ["numero_cedula", "Obligatorio. Numero de cedula o documento."],
        ["carrera", "Obligatorio. Selecciona una carrera desde la lista. No incluye Educacion Continua ni Ingles."],
        ["semestre", "Obligatorio. Valor numerico entre 1 y 4."],
        ["costos", "No van en el Excel. Se calculan automaticamente por carrera; Gastronomia usa su valor diferenciado."],
        ["periodo", "No va en el Excel. Seleccionalo en la pantalla antes de subir el documento."],
    ]
    for row in help_rows:
        help_sheet.append(row)
    for cell in help_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="600000")
    _autosize_worksheet(help_sheet)

    period_sheet = workbook.create_sheet("Periodos")
    period_sheet.append(["cod_periodo", "detalle_periodo", "fecha_inicio", "fecha_fin"])
    for period in periodos:
        period_sheet.append([
            period.get("cod_periodo"),
            period.get("detalle_periodo"),
            period.get("fecha_inicio"),
            period.get("fecha_fin"),
        ])
    for cell in period_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="600000")
    _autosize_worksheet(period_sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _parse_matricula_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo Excel no es valido") from exc
    sheet = workbook["Matriculas"] if "Matriculas" in workbook.sheetnames else workbook.active
    headers = [_normalize_excel_header(cell.value) for cell in sheet[1]]
    rows: list[dict[str, Any]] = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_clean(value) for value in values):
            continue
        row = {headers[column_index]: value for column_index, value in enumerate(values) if column_index < len(headers)}
        nombre = _upper(_excel_value(row, "nombres_apellidos", "nombre", "nombres", "estudiante"))
        cedula = _clean(_excel_value(row, "numero_cedula", "cedula", "numero_de_cedula", "documento"))
        carrera = _clean(_excel_value(row, "carrera", "nombre_carrera"))
        errors: list[str] = []
        if not nombre:
            errors.append("nombres_apellidos")
        if not cedula:
            errors.append("numero_cedula")
        if not carrera:
            errors.append("carrera")
        if carrera and _is_excluded_matricula_excel_career(carrera):
            errors.append("carrera no permitida para esta plantilla")
        semestre = _int_value(_excel_value(row, "semestre"))
        if semestre is None or semestre < 1 or semestre > 4:
            errors.append("semestre debe estar entre 1 y 4")
        matricula_base, arancel_base = _matricula_excel_costs(carrera, row)
        rows.append(
            {
                "row": index,
                "nombre": nombre,
                "cedula": cedula,
                "carrera": carrera,
                "semestre": semestre or 1,
                "matricula_base": matricula_base,
                "arancel_base": arancel_base,
                "errors": errors,
            }
        )
    return rows


def _matricula_excel_context(row: dict[str, Any], period: dict[str, Any]) -> dict[str, Any]:
    return {
        "codigo_estud": "",
        "cedula": row["cedula"],
        "nombre": row["nombre"],
        "correo_personal": "",
        "correo_intec": "",
        "estado": "",
        "cod_anio_basica": "",
        "carrera": row["carrera"],
        "periodo": period,
        "periodo_matricula": period,
        "cabecera_matricula": {},
        "num_matricula": "",
        "tipo_beca": "Sin Beca",
        "porcentaje_beca": 0,
        "es_suzuki": False,
        "is_presencial": False,
        "subjects": [],
        "semestre": row["semestre"],
        "semestre_matricula": row["semestre"],
        "matricula_base": row["matricula_base"],
        "arancel_base": row["arancel_base"],
        "matricula_financiada": 0,
        "arancel_financiado": 0,
        "total_financiado": 0,
    }


def _fetch_subjects(cursor: Any, codigo_estud: str, codigo_periodo: str) -> list[dict[str, Any]]:
    cursor.execute(
        """
        SELECT
            LTRIM(RTRIM(TRY_CONVERT(varchar(100), CX.codigo_materia))) AS codigo_materia,
            CX.PromedioFinal AS promedio_final,
            CX.caprueba,
            CX.ControlAprueba,
            P.cod_materia,
            P.Nomb_Materia AS nombre_materia,
            P.Creditos AS creditos,
            P.Semestre AS semestre,
            P.Horas AS horas,
            P.ValorHora AS valor_hora,
            P.ValorHoraVirtual AS valor_hora_virtual
        FROM dbo.CARRERAXESTUD CX
        LEFT JOIN dbo.PENSUM P
            ON LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), P.codigo_materia)))
               = LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), CX.codigo_materia)))
           AND TRY_CONVERT(varchar(50), P.Cod_AnioBasica)
               = TRY_CONVERT(varchar(50), CX.cod_anio_Basica)
        WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), CX.codigo_estud))) = ?
          AND TRY_CONVERT(varchar(50), CX.codigo_periodo) = ?
        ORDER BY TRY_CONVERT(int, P.Semestre), CX.codigo_materia
        """,
        codigo_estud,
        codigo_periodo,
    )
    return _rows_from_cursor(cursor)


def _fetch_matricula_header(
    cursor: Any,
    codigo_estud: str,
    codigo_periodo: str,
    cod_anio_basica: str = "",
) -> dict[str, Any] | None:
    wheres = ["LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.codigo_estud))) = ?"]
    params: list[Any] = [codigo_estud]
    if codigo_periodo:
        wheres.append("TRY_CONVERT(varchar(50), CM.codigo_periodo) = ?")
        params.append(codigo_periodo)
    if cod_anio_basica:
        wheres.append("TRY_CONVERT(varchar(50), CM.cod_anio_Basica) = ?")
        params.append(cod_anio_basica)

    cursor.execute(
        f"""
        SELECT TOP (1)
            LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.codigo_estud))) AS codigo_estud,
            TRY_CONVERT(varchar(50), CM.cod_anio_Basica) AS cod_anio_basica,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), C.Nombre_Basica))) AS carrera,
            TRY_CONVERT(varchar(50), CM.codigo_periodo) AS codigo_periodo,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), P.Detalle_Periodo))) AS periodo_matricula,
            P.fechain AS fecha_inicio,
            P.fechafin AS fecha_fin,
            CM.fecha_pago,
            CM.Num_Matricula AS num_matricula,
            CM.InscripValor AS inscrip_valor,
            CM.MatriValor AS matri_valor,
            CM.valor AS valor,
            CM.Jornada AS jornada,
            CM.ControlMatricula AS control_matricula
        FROM dbo.CABECERA_MATRICULA CM
        LEFT JOIN dbo.CARRERAS C
            ON TRY_CONVERT(varchar(50), C.Cod_AnioBasica) = TRY_CONVERT(varchar(50), CM.cod_anio_Basica)
        LEFT JOIN dbo.PERIODO P
            ON TRY_CONVERT(varchar(50), P.cod_periodo) = TRY_CONVERT(varchar(50), CM.codigo_periodo)
        WHERE {" AND ".join(wheres)}
        ORDER BY
            TRY_CONVERT(int, CM.Num_Matricula) DESC,
            TRY_CONVERT(datetime2, CM.fecha_pago) DESC,
            TRY_CONVERT(varchar(50), CM.cod_anio_Basica)
        """,
        params,
    )
    return _one_from_cursor(cursor)


def _fetch_certificate_context(
    cursor: Any,
    codigo_estud: str,
    codigo_periodo: str,
    proximo_periodo: str,
    semestre: int | None,
    cod_anio_basica: str = "",
    matricula_periodo: str = "",
) -> dict[str, Any] | None:
    cursor.execute(
        """
        SELECT TOP (1)
            LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))) AS cedula,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), D.Apellidos_nombre))) AS nombre,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoPersonal, D.correo)))) AS correo_personal,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoIntec, D.correointec)))) AS correo_intec,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), COALESCE(CE.Estado, D.Estado)))) AS estado
        FROM dbo.DATOS_ESTUD D
        LEFT JOIN dbo.CorreosEstudIntec CE
            ON TRY_CONVERT(varchar(50), CE.codestud) = TRY_CONVERT(varchar(50), D.codigo_estud)
        WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.codigo_estud))) = ?
        """,
        codigo_estud,
    )
    student = _one_from_cursor(cursor)
    if not student:
        return None

    matricula_period_code = _clean(matricula_periodo) or _clean(proximo_periodo) or codigo_periodo
    header = _fetch_matricula_header(cursor, codigo_estud, matricula_period_code, _clean(cod_anio_basica))
    if not header and matricula_period_code != codigo_periodo:
        header = _fetch_matricula_header(cursor, codigo_estud, codigo_periodo, _clean(cod_anio_basica))

    cursor.execute(
        """
        SELECT TOP (1)
            TRY_CONVERT(varchar(50), CX.cod_anio_Basica) AS cod_carrera,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), C.Nombre_Basica))) AS carrera
        FROM dbo.CARRERAXESTUD CX
        LEFT JOIN dbo.CARRERAS C
            ON TRY_CONVERT(varchar(50), C.Cod_AnioBasica) = TRY_CONVERT(varchar(50), CX.cod_anio_Basica)
        WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), CX.codigo_estud))) = ?
          AND TRY_CONVERT(varchar(50), CX.codigo_periodo) = ?
        ORDER BY CX.cod_anio_Basica
        """,
        codigo_estud,
        codigo_periodo,
    )
    career = _one_from_cursor(cursor) or {}

    period = _period_meta(cursor, codigo_periodo) or {
        "cod_periodo": codigo_periodo,
        "detalle_periodo": codigo_periodo,
        "fecha_inicio": "",
        "fecha_fin": "",
    }
    next_period = _period_meta(cursor, proximo_periodo) if proximo_periodo else None
    header_period_code = _clean((header or {}).get("codigo_periodo"))
    header_period = (
        {
            "cod_periodo": header_period_code,
            "detalle_periodo": _clean((header or {}).get("periodo_matricula")) or header_period_code,
            "fecha_inicio": _date_text((header or {}).get("fecha_inicio")),
            "fecha_fin": _date_text((header or {}).get("fecha_fin")),
            "orden": None,
        }
        if header
        else None
    )
    period_matricula = header_period or next_period or period

    context_carrera = _clean((header or {}).get("carrera")) or _clean(career.get("carrera")) or "-"
    scholarship = _legacy_certificate_scholarship(cursor, codigo_estud)
    certificate_costs = _certificate_costs_for_career(scholarship, context_carrera)

    cursor.execute(
        """
        SELECT TOP (1) codmodalida
        FROM dbo.PREINSCRIPCION
        WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), Codestu))) = ?
        """,
        codigo_estud,
    )
    modality = cursor.fetchone()
    is_presencial = _int_value(modality[0] if modality else None) == 3

    subjects = _fetch_subjects(cursor, codigo_estud, codigo_periodo)
    matricula_subjects = (
        _fetch_subjects(cursor, codigo_estud, header_period_code)
        if header_period_code and header_period_code != codigo_periodo
        else subjects
    )
    max_semester = max((_int_value(item.get("semestre")) or 0 for item in matricula_subjects), default=0)
    if semestre is not None:
        semestre_matricula = max(1, min(semestre, 4))
    else:
        semestre_matricula = max(1, min(max_semester + 1, 4))

    return {
        "codigo_estud": codigo_estud,
        "cedula": _clean(student.get("cedula")),
        "nombre": _upper(student.get("nombre") or codigo_estud),
        "correo_personal": _clean(student.get("correo_personal")),
        "correo_intec": _clean(student.get("correo_intec")),
        "estado": _clean(student.get("estado")),
        "cod_anio_basica": _clean((header or {}).get("cod_anio_basica")) or _clean(career.get("cod_carrera")),
        "carrera": context_carrera,
        "periodo": period,
        "periodo_matricula": period_matricula,
        "cabecera_matricula": header or {},
        "num_matricula": _clean((header or {}).get("num_matricula")),
        "tipo_beca": scholarship["tipo_beca"],
        "porcentaje_beca": scholarship["porcentaje_beca"],
        "es_suzuki": scholarship["es_suzuki"],
        "is_presencial": is_presencial,
        "subjects": subjects,
        "semestre": semestre,
        "semestre_matricula": semestre_matricula,
        "matricula_base": certificate_costs["matricula_base"],
        "arancel_base": certificate_costs["arancel_base"],
        "matricula_financiada": certificate_costs["matricula_financiada"],
        "arancel_financiado": certificate_costs["arancel_financiado"],
        "total_financiado": certificate_costs["total_financiado"],
    }


@router.get("/catalog")
def catalog(_: SessionUser = Depends(_CERTIFICATES_ACCESS)) -> dict[str, Any]:
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT DISTINCT LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca))) AS tipo_beca
                FROM dbo.Becas
                WHERE tipo_beca IS NOT NULL
                  AND LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca))) <> ''
                ORDER BY LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca))) ASC
                """
            )
            becas = [_clean(row.tipo_beca) for row in cursor.fetchall() if _clean(row.tipo_beca)]

            cursor.execute(
                """
                SELECT
                    TRY_CONVERT(varchar(50), cod_periodo) AS cod_periodo,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Detalle_Periodo))) AS detalle_periodo,
                    fechain AS fecha_inicio,
                    fechafin AS fecha_fin,
                    TRY_CONVERT(int, Orden) AS orden
                FROM dbo.PERIODO
                ORDER BY Orden ASC
                """
            )
            periodos = [_period_row(row) for row in _rows_from_cursor(cursor)]
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo cargar el catalogo de certificados",
        ) from exc

    return {
        "becas": becas,
        "periodos": periodos,
        "semestres": [
            {"value": "1", "label": "Primero"},
            {"value": "2", "label": "Segundo"},
            {"value": "3", "label": "Tercero"},
            {"value": "4", "label": "Cuarto"},
        ],
    }


@router.get("/matricula-excel/plantilla")
def download_matricula_excel_template(_: SessionUser = Depends(_CERTIFICATES_ACCESS)) -> StreamingResponse:
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    TRY_CONVERT(varchar(50), cod_periodo) AS cod_periodo,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Detalle_Periodo))) AS detalle_periodo,
                    fechain AS fecha_inicio,
                    fechafin AS fecha_fin,
                    TRY_CONVERT(int, Orden) AS orden
                FROM dbo.PERIODO
                ORDER BY Orden ASC
                """
            )
            periodos = [_period_row(row) for row in _rows_from_cursor(cursor)]
            cursor.execute(
                """
                SELECT DISTINCT
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Nombre_Basica))) AS carrera
                FROM dbo.CARRERAS
                WHERE Nombre_Basica IS NOT NULL
                  AND LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Nombre_Basica))) <> ''
                ORDER BY LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Nombre_Basica)))
                """
            )
            carreras = [
                row
                for row in _rows_from_cursor(cursor)
                if not _is_excluded_matricula_excel_career(row.get("carrera"))
            ]
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo generar la plantilla de matricula",
        ) from exc

    return StreamingResponse(
        _build_matricula_excel_template(periodos, carreras),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="plantilla_matricula_certificados_v3.xlsx"',
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )


@router.post("/matricula-excel/generar")
async def generate_matricula_from_excel(
    periodo: str = Form(...),
    file: UploadFile = File(...),
    _: SessionUser = Depends(_CERTIFICATES_ACCESS),
) -> StreamingResponse:
    period_code = _clean(periodo)
    if not period_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selecciona un periodo")
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Sube un archivo Excel .xlsx")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo esta vacio")

    rows = _parse_matricula_excel(file_bytes)
    if not rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel no contiene filas para procesar")

    invalid_rows = [row for row in rows if row["errors"]]
    if invalid_rows:
        details = [
            f"Fila {row['row']}: {', '.join(row['errors'])}"
            for row in invalid_rows[:12]
        ]
        suffix = f" y {len(invalid_rows) - 12} mas" if len(invalid_rows) > 12 else ""
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Corrige el Excel antes de generar: {'; '.join(details)}{suffix}",
        )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            period = _period_meta(cursor, period_code)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo validar el periodo seleccionado",
        ) from exc

    if not period:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El periodo seleccionado no existe")

    contexts = [_matricula_excel_context(row, period) for row in rows]
    pdf_bytes = _build_matricula_pdf_bundle(contexts)
    filename = f"certificados_matricula_excel_{period_code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/estudiantes")
def list_students(
    tipo_beca: str = Query(default=""),
    periodo: str = Query(default=""),
    busqueda: str = Query(default=""),
    cedulas: str = Query(default=""),
    matricula_scope: str = Query(default="todas"),
    semestre: int | None = Query(default=None, ge=1, le=4),
    limit: int = Query(default=500, ge=1, le=1000),
    _: SessionUser = Depends(_CERTIFICATES_ACCESS),
) -> dict[str, Any]:
    tipo = _clean(tipo_beca)
    if tipo.lower() in {"todos", "todas", "todo", "all"}:
        tipo = ""
    codigo_periodo = _clean(periodo)
    search = _clean(busqueda)
    document_numbers = _split_documents(cedulas)
    only_latest_matricula = _clean(matricula_scope).lower() in {"ultima", "última", "last", "reciente"}
    params: list[Any] = []

    if codigo_periodo:
        wheres = ["TRY_CONVERT(varchar(50), CM.codigo_periodo) = ?"]
        params.append(codigo_periodo)
        if tipo:
            if tipo == "Sin beca":
                wheres.append("NOT EXISTS (SELECT 1 FROM dbo.Becas B WHERE B.codestud = CM.codigo_estud)")
            else:
                wheres.append(
                    """
                    EXISTS (
                        SELECT 1
                        FROM dbo.Becas B
                        WHERE B.codestud = CM.codigo_estud
                          AND LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), B.tipo_beca))) = ?
                    )
                    """
                )
                params.append(tipo)
        if search:
            wheres.append(
                """
                (
                    TRY_CONVERT(nvarchar(max), COALESCE(CE.Nombres, D.Apellidos_nombre)) LIKE ?
                    OR TRY_CONVERT(varchar(50), CM.codigo_estud) LIKE ?
                    OR REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))), '-', ''), ' ', '') LIKE ?
                    OR TRY_CONVERT(nvarchar(max), C.Nombre_Basica) LIKE ?
                )
                """
            )
            params.extend([f"%{search}%", f"%{search}%", f"%{_document_digits(search) or search}%", f"%{search}%"])
        if document_numbers:
            placeholders = ", ".join("?" for _ in document_numbers)
            wheres.append(f"REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))), '-', ''), ' ', '') IN ({placeholders})")
            params.extend(document_numbers)
        sql = f"""
            SELECT TOP ({limit})
                LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.codigo_estud))) AS codestud,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.Nombres, D.Apellidos_nombre)))) AS nombres,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoPersonal, D.correo)))) AS correo_personal,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoIntec, D.correointec)))) AS correo_intec,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), COALESCE(CE.Estado, D.Estado)))) AS estado,
                TRY_CONVERT(varchar(50), CM.cod_anio_Basica) AS cod_anio_basica,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), C.Nombre_Basica))) AS carrera,
                TRY_CONVERT(varchar(50), CM.codigo_periodo) AS codigo_periodo_matricula,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), P.Detalle_Periodo))) AS periodo_matricula,
                LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.Num_Matricula))) AS num_matricula
            FROM dbo.CABECERA_MATRICULA CM
            INNER JOIN dbo.DATOS_ESTUD D
                ON TRY_CONVERT(varchar(50), D.codigo_estud) = TRY_CONVERT(varchar(50), CM.codigo_estud)
            LEFT JOIN dbo.CorreosEstudIntec CE
                ON TRY_CONVERT(varchar(50), CE.codestud) = TRY_CONVERT(varchar(50), CM.codigo_estud)
            LEFT JOIN dbo.CARRERAS C
                ON TRY_CONVERT(varchar(50), C.Cod_AnioBasica) = TRY_CONVERT(varchar(50), CM.cod_anio_Basica)
            LEFT JOIN dbo.PERIODO P
                ON TRY_CONVERT(varchar(50), P.cod_periodo) = TRY_CONVERT(varchar(50), CM.codigo_periodo)
            WHERE {" AND ".join(wheres)}
            ORDER BY C.Nombre_Basica, COALESCE(CE.Nombres, D.Apellidos_nombre), CM.Num_Matricula
        """
    else:
        if document_numbers:
            placeholders = ", ".join("?" for _ in document_numbers)
            wheres = [
                _active_sql("D"),
                f"REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))), '-', ''), ' ', '') IN ({placeholders})",
            ]
            params.extend(document_numbers)
            if tipo:
                if tipo == "Sin beca":
                    wheres.append("NOT EXISTS (SELECT 1 FROM dbo.Becas B WHERE B.codestud = CM.codigo_estud)")
                else:
                    wheres.append(
                        """
                        EXISTS (
                            SELECT 1
                            FROM dbo.Becas B
                            WHERE B.codestud = CM.codigo_estud
                              AND LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), B.tipo_beca))) = ?
                        )
                        """
                    )
                    params.append(tipo)
            if search:
                wheres.append(
                    """
                    (
                        TRY_CONVERT(nvarchar(max), COALESCE(CE.Nombres, D.Apellidos_nombre)) LIKE ?
                        OR TRY_CONVERT(varchar(50), CM.codigo_estud) LIKE ?
                        OR REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))), '-', ''), ' ', '') LIKE ?
                        OR TRY_CONVERT(nvarchar(max), C.Nombre_Basica) LIKE ?
                    )
                    """
                )
                params.extend([f"%{search}%", f"%{search}%", f"%{_document_digits(search) or search}%", f"%{search}%"])
            if only_latest_matricula:
                sql = f"""
                    WITH MatriculasFiltradas AS (
                        SELECT
                            LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.codigo_estud))) AS codestud,
                            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.Nombres, D.Apellidos_nombre)))) AS nombres,
                            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoPersonal, D.correo)))) AS correo_personal,
                            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoIntec, D.correointec)))) AS correo_intec,
                            LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), COALESCE(CE.Estado, D.Estado)))) AS estado,
                            TRY_CONVERT(varchar(50), CM.cod_anio_Basica) AS cod_anio_basica,
                            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), C.Nombre_Basica))) AS carrera,
                            TRY_CONVERT(varchar(50), CM.codigo_periodo) AS codigo_periodo_matricula,
                            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), P.Detalle_Periodo))) AS periodo_matricula,
                            LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.Num_Matricula))) AS num_matricula,
                            REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))), '-', ''), ' ', '') AS cedula_orden,
                            ISNULL(TRY_CONVERT(int, P.Orden), 0) AS periodo_orden,
                            TRY_CONVERT(datetime2, CM.fecha_pago) AS fecha_pago_orden,
                            TRY_CONVERT(int, CM.Num_Matricula) AS num_matricula_orden,
                            ROW_NUMBER() OVER (
                                PARTITION BY LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.codigo_estud)))
                                ORDER BY
                                    ISNULL(TRY_CONVERT(int, P.Orden), 0) DESC,
                                    TRY_CONVERT(datetime2, CM.fecha_pago) DESC,
                                    TRY_CONVERT(int, CM.Num_Matricula) DESC,
                                    TRY_CONVERT(varchar(50), CM.cod_anio_Basica) DESC
                            ) AS rn
                        FROM dbo.CABECERA_MATRICULA CM
                        INNER JOIN dbo.DATOS_ESTUD D
                            ON TRY_CONVERT(varchar(50), D.codigo_estud) = TRY_CONVERT(varchar(50), CM.codigo_estud)
                        LEFT JOIN dbo.CorreosEstudIntec CE
                            ON TRY_CONVERT(varchar(50), CE.codestud) = TRY_CONVERT(varchar(50), CM.codigo_estud)
                        LEFT JOIN dbo.CARRERAS C
                            ON TRY_CONVERT(varchar(50), C.Cod_AnioBasica) = TRY_CONVERT(varchar(50), CM.cod_anio_Basica)
                        LEFT JOIN dbo.PERIODO P
                            ON TRY_CONVERT(varchar(50), P.cod_periodo) = TRY_CONVERT(varchar(50), CM.codigo_periodo)
                        WHERE {" AND ".join(wheres)}
                    )
                    SELECT TOP ({limit})
                        codestud,
                        nombres,
                        correo_personal,
                        correo_intec,
                        estado,
                        cod_anio_basica,
                        carrera,
                        codigo_periodo_matricula,
                        periodo_matricula,
                        num_matricula
                    FROM MatriculasFiltradas
                    WHERE rn = 1
                    ORDER BY cedula_orden, periodo_orden DESC, fecha_pago_orden DESC, num_matricula_orden DESC
                """
            else:
                sql = f"""
                    SELECT TOP ({limit})
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.codigo_estud))) AS codestud,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.Nombres, D.Apellidos_nombre)))) AS nombres,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoPersonal, D.correo)))) AS correo_personal,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoIntec, D.correointec)))) AS correo_intec,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), COALESCE(CE.Estado, D.Estado)))) AS estado,
                        TRY_CONVERT(varchar(50), CM.cod_anio_Basica) AS cod_anio_basica,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), C.Nombre_Basica))) AS carrera,
                        TRY_CONVERT(varchar(50), CM.codigo_periodo) AS codigo_periodo_matricula,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), P.Detalle_Periodo))) AS periodo_matricula,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.Num_Matricula))) AS num_matricula
                    FROM dbo.CABECERA_MATRICULA CM
                    INNER JOIN dbo.DATOS_ESTUD D
                        ON TRY_CONVERT(varchar(50), D.codigo_estud) = TRY_CONVERT(varchar(50), CM.codigo_estud)
                    LEFT JOIN dbo.CorreosEstudIntec CE
                        ON TRY_CONVERT(varchar(50), CE.codestud) = TRY_CONVERT(varchar(50), CM.codigo_estud)
                    LEFT JOIN dbo.CARRERAS C
                        ON TRY_CONVERT(varchar(50), C.Cod_AnioBasica) = TRY_CONVERT(varchar(50), CM.cod_anio_Basica)
                    LEFT JOIN dbo.PERIODO P
                        ON TRY_CONVERT(varchar(50), P.cod_periodo) = TRY_CONVERT(varchar(50), CM.codigo_periodo)
                    WHERE {" AND ".join(wheres)}
                    ORDER BY
                        REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))), '-', ''), ' ', ''),
                        ISNULL(TRY_CONVERT(int, P.Orden), 0) ASC,
                        TRY_CONVERT(varchar(50), CM.codigo_periodo),
                        TRY_CONVERT(varchar(50), CM.cod_anio_Basica),
                        TRY_CONVERT(int, CM.Num_Matricula)
                """
        else:
            wheres = [f"({_active_sql('CE')} OR {_active_sql('D')})"]
            if tipo:
                if tipo == "Sin beca":
                    wheres.append("NOT EXISTS (SELECT 1 FROM dbo.Becas B WHERE B.codestud = CE.codestud)")
                else:
                    wheres.append(
                        """
                        EXISTS (
                            SELECT 1
                            FROM dbo.Becas B
                            WHERE B.codestud = CE.codestud
                              AND LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), B.tipo_beca))) = ?
                        )
                        """
                    )
                    params.append(tipo)
            if search:
                wheres.append(
                    """
                    (
                        TRY_CONVERT(nvarchar(max), COALESCE(CE.Nombres, D.Apellidos_nombre)) LIKE ?
                        OR TRY_CONVERT(varchar(50), CE.codestud) LIKE ?
                        OR REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), D.Cedula_Est))), '-', ''), ' ', '') LIKE ?
                        OR TRY_CONVERT(nvarchar(max), CMX.carrera) LIKE ?
                    )
                    """
                )
                params.extend([f"%{search}%", f"%{search}%", f"%{_document_digits(search) or search}%", f"%{search}%"])
            sql = f"""
                SELECT TOP ({limit})
                    LTRIM(RTRIM(TRY_CONVERT(varchar(50), CE.codestud))) AS codestud,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.Nombres, D.Apellidos_nombre)))) AS nombres,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoPersonal, D.correo)))) AS correo_personal,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), COALESCE(CE.CorreoIntec, D.correointec)))) AS correo_intec,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), COALESCE(CE.Estado, D.Estado)))) AS estado,
                    CMX.cod_anio_basica,
                    CMX.carrera,
                    CMX.codigo_periodo_matricula,
                    CMX.periodo_matricula,
                    CMX.num_matricula
                FROM dbo.CorreosEstudIntec CE
                LEFT JOIN dbo.DATOS_ESTUD D
                    ON TRY_CONVERT(varchar(50), D.codigo_estud) = TRY_CONVERT(varchar(50), CE.codestud)
                OUTER APPLY (
                    SELECT TOP (1)
                        TRY_CONVERT(varchar(50), CM.cod_anio_Basica) AS cod_anio_basica,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), C.Nombre_Basica))) AS carrera,
                        TRY_CONVERT(varchar(50), CM.codigo_periodo) AS codigo_periodo_matricula,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), P.Detalle_Periodo))) AS periodo_matricula,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), CM.Num_Matricula))) AS num_matricula,
                        TRY_CONVERT(int, P.Orden) AS periodo_orden
                    FROM dbo.CABECERA_MATRICULA CM
                    LEFT JOIN dbo.CARRERAS C
                        ON TRY_CONVERT(varchar(50), C.Cod_AnioBasica) = TRY_CONVERT(varchar(50), CM.cod_anio_Basica)
                    LEFT JOIN dbo.PERIODO P
                        ON TRY_CONVERT(varchar(50), P.cod_periodo) = TRY_CONVERT(varchar(50), CM.codigo_periodo)
                    WHERE TRY_CONVERT(varchar(50), CM.codigo_estud) = TRY_CONVERT(varchar(50), CE.codestud)
                    ORDER BY
                        ISNULL(TRY_CONVERT(int, P.Orden), 0) DESC,
                        TRY_CONVERT(datetime2, CM.fecha_pago) DESC,
                        TRY_CONVERT(int, CM.Num_Matricula) DESC
                ) CMX
                WHERE {" AND ".join(wheres)}
                ORDER BY ISNULL(CMX.periodo_orden, 0) DESC, COALESCE(CE.Nombres, D.Apellidos_nombre)
            """

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(sql, params)
            rows = _rows_from_cursor(cursor)
            reprobadas_by_ref: dict[str, list[dict[str, Any]]] = {}
            if codigo_periodo:
                codes = [_clean(row.get("codestud")) for row in rows if _clean(row.get("codestud"))]
                reprobadas_by_code = _fetch_reprobadas(cursor, codigo_periodo, codes, semestre)
                for row in rows:
                    code = _clean(row.get("codestud"))
                    row_ref = _certificate_ref(code, _clean(row.get("cod_anio_basica")), _clean(row.get("codigo_periodo_matricula")))
                    reprobadas_by_ref[row_ref] = reprobadas_by_code.get(code, [])
            else:
                period_groups: dict[str, set[str]] = {}
                for row in rows:
                    code = _clean(row.get("codestud"))
                    row_period = _clean(row.get("codigo_periodo_matricula"))
                    if code and row_period:
                        period_groups.setdefault(row_period, set()).add(code)
                reprobadas_by_period: dict[str, dict[str, list[dict[str, Any]]]] = {}
                for row_period, codes_set in period_groups.items():
                    reprobadas_by_period[row_period] = _fetch_reprobadas(cursor, row_period, sorted(codes_set), semestre)
                for row in rows:
                    code = _clean(row.get("codestud"))
                    row_period = _clean(row.get("codigo_periodo_matricula"))
                    row_ref = _certificate_ref(code, _clean(row.get("cod_anio_basica")), row_period)
                    reprobadas_by_ref[row_ref] = reprobadas_by_period.get(row_period, {}).get(code, [])
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=_db_error_detail("No se pudo consultar estudiantes para certificados", exc),
        ) from exc

    items = [
        _student_list_row(
            row,
            reprobadas_by_ref.get(
                _certificate_ref(
                    _clean(row.get("codestud")),
                    _clean(row.get("cod_anio_basica")),
                    _clean(row.get("codigo_periodo_matricula")),
                )
            ),
        )
        for row in rows
    ]
    return {
        "items": items,
        "total": len(items),
        "criteria": {
            "tipo_beca": tipo,
            "periodo": codigo_periodo,
            "busqueda": search,
            "cedulas": cedulas,
            "matricula_scope": "ultima" if only_latest_matricula else "todas",
            "semestre": semestre,
            "limit": limit,
        },
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@router.post("/generar")
def generate_certificates(
    payload: CertificateGeneratePayload,
    _: SessionUser = Depends(_CERTIFICATES_ACCESS),
) -> StreamingResponse:
    periodo = _clean(payload.periodo)
    proximo_periodo = _clean(payload.proximo_periodo)
    certificate_type = _clean(payload.tipo_certificado).lower() or "ambos"
    if certificate_type not in {"ambos", "matricula", "promocion"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de certificado no valido")
    student_refs: list[dict[str, str]] = []
    seen: set[str] = set()
    for raw_code in payload.estudiantes:
        ref = _parse_certificate_ref(raw_code)
        code = ref["codigo_estud"]
        ref_key = _certificate_ref(code, ref["cod_anio_basica"], ref["codigo_periodo"])
        if code and ref_key not in seen:
            seen.add(ref_key)
            student_refs.append(ref)

    if not student_refs:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selecciona estudiantes")

    batch_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    zip_buffer = BytesIO()
    manifest: dict[str, Any] = {
        "batchId": batch_id,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "tipo_beca": _clean(payload.tipo_beca),
        "tipo_certificado": certificate_type,
        "periodo": periodo,
        "proximo_periodo": proximo_periodo,
        "semestre": payload.semestre,
        "items": [],
        "omitidos": [],
    }
    generated_count = 0

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            with ZipFile(zip_buffer, "w", ZIP_DEFLATED) as archive:
                for ref in student_refs:
                    code = ref["codigo_estud"]
                    generation_period = periodo or ref["codigo_periodo"]
                    if not generation_period:
                        manifest["omitidos"].append(
                            {
                                "codestud": code,
                                "motivo": "No se encontro periodo base ni cabecera de matricula para generar",
                            }
                        )
                        continue
                    is_active = _active_student(cursor, code)
                    if not is_active:
                        manifest["omitidos"].append({"codestud": code, "motivo": "Estudiante no activo"})
                        continue
                    reprobadas = (
                        _certificate_reprobadas(cursor, generation_period, code, payload.semestre)
                        if certificate_type in {"ambos", "promocion"}
                        else []
                    )
                    if certificate_type == "promocion" and reprobadas:
                        manifest["omitidos"].append(
                            {
                                "codestud": code,
                                "motivo": "Estudiante reprobado; no se permite generar certificado de promocion",
                                "reprobadas": reprobadas,
                            }
                        )
                        continue

                    context = _fetch_certificate_context(
                        cursor,
                        code,
                        generation_period,
                        proximo_periodo,
                        payload.semestre,
                        ref["cod_anio_basica"],
                        ref["codigo_periodo"],
                    )
                    if not context:
                        manifest["omitidos"].append({"codestud": code, "motivo": "No se encontraron datos generales"})
                        continue
                    matricula_block_reason = _matricula_certificate_block_reason(context)
                    if certificate_type == "matricula" and matricula_block_reason:
                        manifest["omitidos"].append(
                            {
                                "codestud": code,
                                "motivo": matricula_block_reason,
                            }
                        )
                        continue

                    safe_name = _safe_filename(context["nombre"])
                    career_suffix = _safe_filename(context.get("cod_anio_basica") or ref["cod_anio_basica"] or "carrera")
                    matricula_period = _safe_filename(
                        context.get("periodo_matricula", {}).get("cod_periodo") or ref["codigo_periodo"] or generation_period
                    )
                    files: list[str] = []
                    if certificate_type in {"ambos", "promocion"}:
                        if not is_active:
                            manifest["omitidos"].append(
                                {
                                    "codestud": code,
                                    "tipo_certificado": "promocion",
                                    "motivo": "Estudiante no activo",
                                }
                            )
                        elif reprobadas:
                            manifest["omitidos"].append(
                                {
                                    "codestud": code,
                                    "tipo_certificado": "promocion",
                                    "motivo": "Estudiante reprobado; no se permite generar certificado de promocion",
                                    "reprobadas": reprobadas,
                                }
                            )
                        else:
                            promo_pdf = _build_promocion_pdf(context)
                            promo_name = f"C_P_{code}_{generation_period}_{batch_id}_{safe_name}_{career_suffix}.pdf"
                            archive.writestr(promo_name, promo_pdf)
                            files.append(promo_name)
                    if certificate_type in {"ambos", "matricula"}:
                        if matricula_block_reason:
                            manifest["omitidos"].append(
                                {
                                    "codestud": code,
                                    "tipo_certificado": "matricula",
                                    "motivo": matricula_block_reason,
                                }
                            )
                        else:
                            matricula_pdf = _build_matricula_pdf(context)
                            matricula_name = f"C_M_{code}_{matricula_period}_{batch_id}_{safe_name}_{career_suffix}.pdf"
                            archive.writestr(matricula_name, matricula_pdf)
                            files.append(matricula_name)
                    manifest["items"].append(
                        {
                            "codestud": code,
                            "cedula": context["cedula"],
                            "nombre": context["nombre"],
                            "carrera": context.get("carrera"),
                            "periodo_matricula": context.get("periodo_matricula", {}).get("detalle_periodo"),
                            "archivos": files,
                        }
                    )
                    generated_count += len(files)
                archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudieron generar los certificados",
        ) from exc

    if generated_count == 0:
        omitted = manifest.get("omitidos") or []
        if any("reprobado" in _clean(item.get("motivo")).lower() for item in omitted):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se genero ningun certificado: los estudiantes seleccionados tienen materias reprobadas.",
            )
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se genero ningun certificado")

    zip_buffer.seek(0)
    filename = f"certificados_{certificate_type}_{batch_id}.zip"
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/generar-pdf")
def generate_certificates_pdf(
    payload: CertificateGeneratePayload,
    _: SessionUser = Depends(_CERTIFICATES_ACCESS),
) -> StreamingResponse:
    periodo = _clean(payload.periodo)
    proximo_periodo = _clean(payload.proximo_periodo)
    certificate_type = _clean(payload.tipo_certificado).lower() or "matricula"
    if certificate_type not in {"matricula", "promocion"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de certificado no valido para PDF masivo")

    student_refs: list[dict[str, str]] = []
    seen: set[str] = set()
    for raw_code in payload.estudiantes:
        ref = _parse_certificate_ref(raw_code)
        code = ref["codigo_estud"]
        ref_key = _certificate_ref(code, ref["cod_anio_basica"], ref["codigo_periodo"])
        if code and ref_key not in seen:
            seen.add(ref_key)
            student_refs.append(ref)

    if not student_refs:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selecciona estudiantes")

    contexts: list[dict[str, Any]] = []
    omitted: list[dict[str, Any]] = []
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            for ref in student_refs:
                code = ref["codigo_estud"]
                generation_period = periodo or ref["codigo_periodo"]
                if not generation_period:
                    omitted.append({"codestud": code, "motivo": "No se encontro periodo base ni cabecera de matricula"})
                    continue
                if not _active_student(cursor, code):
                    omitted.append({"codestud": code, "motivo": "Estudiante no activo"})
                    continue
                reprobadas = _certificate_reprobadas(cursor, generation_period, code, payload.semestre) if certificate_type == "promocion" else []
                if certificate_type == "promocion" and reprobadas:
                    omitted.append(
                        {
                            "codestud": code,
                            "motivo": "Estudiante reprobado; no se permite generar certificado de promocion",
                            "reprobadas": reprobadas,
                        }
                    )
                    continue

                context = _fetch_certificate_context(
                    cursor,
                    code,
                    generation_period,
                    proximo_periodo,
                    payload.semestre,
                    ref["cod_anio_basica"],
                    ref["codigo_periodo"],
                )
                if not context:
                    omitted.append({"codestud": code, "motivo": "No se encontraron datos generales"})
                    continue
                matricula_block_reason = _matricula_certificate_block_reason(context)
                if certificate_type == "matricula" and matricula_block_reason:
                    omitted.append({"codestud": code, "motivo": matricula_block_reason})
                    continue
                contexts.append(context)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="No se pudo generar el PDF") from exc

    if not contexts:
        if any("reprobado" in _clean(item.get("motivo")).lower() for item in omitted):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se genero ningun certificado: los estudiantes seleccionados tienen materias reprobadas.",
            )
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se genero ningun certificado")

    pdf_bytes = _build_matricula_pdf_bundle(contexts) if certificate_type == "matricula" else _build_promocion_pdf_bundle(contexts)
    filename = f"certificados_{certificate_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _certificate_types(certificate_type: str) -> list[str]:
    if certificate_type == "ambos":
        return ["matricula", "promocion"]
    return [certificate_type]


def _reprobadas_text(reprobadas: list[dict[str, Any]]) -> str:
    details: list[str] = []
    for item in reprobadas:
        subject = _clean(item.get("cod_materia") or item.get("codigo_materia")) or "-"
        name = _clean(item.get("nombre"))
        grade = item.get("promedioFinal")
        grade_text = f" Nota {grade}" if grade is not None else ""
        details.append(f"{subject} {name}{grade_text}".strip())
    return "; ".join(details)


def _certificate_excel_filename(code: str, certificate_type: str, period: str, context: dict[str, Any] | None) -> str:
    safe_name = _safe_filename((context or {}).get("nombre") or code)
    career_suffix = _safe_filename((context or {}).get("cod_anio_basica") or "carrera")
    prefix = "C_M" if certificate_type == "matricula" else "C_P"
    return f"{prefix}_{code}_{period}_{safe_name}_{career_suffix}.pdf"


def _certificate_excel_row(
    certificate_type: str,
    status_text: str,
    reason: str,
    code: str,
    generation_period: str,
    context: dict[str, Any] | None = None,
    reprobadas: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    period = (context or {}).get("periodo") or {}
    matricula_period = (context or {}).get("periodo_matricula") or {}
    return {
        "Tipo certificado": "Matrícula" if certificate_type == "matricula" else "Promoción",
        "Estado": status_text,
        "Motivo": reason,
        "Código estudiante": code,
        "Cédula": _clean((context or {}).get("cedula")),
        "Estudiante": _clean((context or {}).get("nombre")) or code,
        "Estado estudiante": _clean((context or {}).get("estado")),
        "Correo INTEC": _clean((context or {}).get("correo_intec")),
        "Correo personal": _clean((context or {}).get("correo_personal")),
        "Código carrera": _clean((context or {}).get("cod_anio_basica")),
        "Carrera": _clean((context or {}).get("carrera")),
        "Periodo base código": _clean(period.get("cod_periodo")) or generation_period,
        "Periodo base": _clean(period.get("detalle_periodo")),
        "Inicio periodo base": _clean(period.get("fecha_inicio")),
        "Fin periodo base": _clean(period.get("fecha_fin")),
        "Periodo matrícula código": _clean(matricula_period.get("cod_periodo")),
        "Periodo matrícula": _clean(matricula_period.get("detalle_periodo")),
        "Inicio matrícula": _clean(matricula_period.get("fecha_inicio")),
        "Fin matrícula": _clean(matricula_period.get("fecha_fin")),
        "Número matrícula": _clean((context or {}).get("num_matricula")),
        "Semestre": _semester_text(_int_value((context or {}).get("semestre_matricula"))),
        "Tipo beca": _clean((context or {}).get("tipo_beca")),
        "Porcentaje beca": (context or {}).get("porcentaje_beca"),
        "Costo matrícula": (context or {}).get("matricula_base"),
        "Costo arancel": (context or {}).get("arancel_base"),
        "Costo matrícula financiada": (context or {}).get("matricula_financiada"),
        "Costo arancel financiado": (context or {}).get("arancel_financiado"),
        "Total financiado": (context or {}).get("total_financiado"),
        "Reprobadas": len(reprobadas or []),
        "Detalle reprobadas": _reprobadas_text(reprobadas or []),
        "Archivo sugerido": _certificate_excel_filename(code, certificate_type, generation_period, context),
    }


def _autosize_worksheet(sheet: Any) -> None:
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 58)


def _build_certificates_workbook(rows: list[dict[str, Any]], payload: CertificateGeneratePayload, certificate_type: str) -> BytesIO:
    columns = [
        "Tipo certificado",
        "Estado",
        "Motivo",
        "Código estudiante",
        "Cédula",
        "Estudiante",
        "Estado estudiante",
        "Correo INTEC",
        "Correo personal",
        "Código carrera",
        "Carrera",
        "Periodo base código",
        "Periodo base",
        "Inicio periodo base",
        "Fin periodo base",
        "Periodo matrícula código",
        "Periodo matrícula",
        "Inicio matrícula",
        "Fin matrícula",
        "Número matrícula",
        "Semestre",
        "Tipo beca",
        "Porcentaje beca",
        "Costo matrícula",
        "Costo arancel",
        "Costo matrícula financiada",
        "Costo arancel financiado",
        "Total financiado",
        "Reprobadas",
        "Detalle reprobadas",
        "Archivo sugerido",
    ]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certificados"
    sheet.append(columns)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="600000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append([row.get(column) for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _autosize_worksheet(sheet)

    summary = workbook.create_sheet("Resumen")
    summary.append(["Campo", "Valor"])
    summary_rows = [
        ("Generado", datetime.now(timezone.utc).isoformat()),
        ("Tipo certificado", certificate_type),
        ("Total filas", len(rows)),
        ("Habilitados", sum(1 for row in rows if row.get("Estado") == "Habilitado")),
        ("Omitidos", sum(1 for row in rows if row.get("Estado") != "Habilitado")),
        ("Periodo base", _clean(payload.periodo)),
        ("Periodo matrícula", _clean(payload.proximo_periodo)),
        ("Semestre", payload.semestre or ""),
        ("Tipo beca", _clean(payload.tipo_beca)),
    ]
    for item in summary_rows:
        summary.append(list(item))
    for cell in summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="600000")
    _autosize_worksheet(summary)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@router.post("/exportar-excel")
def export_certificates_excel(
    payload: CertificateGeneratePayload,
    _: SessionUser = Depends(_CERTIFICATES_ACCESS),
) -> StreamingResponse:
    periodo = _clean(payload.periodo)
    proximo_periodo = _clean(payload.proximo_periodo)
    certificate_type = _clean(payload.tipo_certificado).lower() or "ambos"
    if certificate_type not in {"ambos", "matricula", "promocion"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de certificado no valido")

    student_refs: list[dict[str, str]] = []
    seen: set[str] = set()
    for raw_code in payload.estudiantes:
        ref = _parse_certificate_ref(raw_code)
        code = ref["codigo_estud"]
        ref_key = _certificate_ref(code, ref["cod_anio_basica"], ref["codigo_periodo"])
        if code and ref_key not in seen:
            seen.add(ref_key)
            student_refs.append(ref)
    if not student_refs:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selecciona estudiantes")

    rows: list[dict[str, Any]] = []
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            for ref in student_refs:
                code = ref["codigo_estud"]
                generation_period = periodo or ref["codigo_periodo"]
                for item_type in _certificate_types(certificate_type):
                    if not generation_period:
                        rows.append(
                            _certificate_excel_row(
                                item_type,
                                "Omitido",
                                "No se encontro periodo base ni cabecera de matricula",
                                code,
                                generation_period,
                            )
                        )
                        continue
                    if not _active_student(cursor, code):
                        rows.append(_certificate_excel_row(item_type, "Omitido", "Estudiante no activo", code, generation_period))
                        continue
                    reprobadas = _certificate_reprobadas(cursor, generation_period, code, payload.semestre)
                    context = _fetch_certificate_context(
                        cursor,
                        code,
                        generation_period,
                        proximo_periodo,
                        payload.semestre,
                        ref["cod_anio_basica"],
                        ref["codigo_periodo"],
                    )
                    if not context:
                        rows.append(
                            _certificate_excel_row(
                                item_type,
                                "Omitido",
                                "No se encontraron datos generales",
                                code,
                                generation_period,
                                reprobadas=reprobadas,
                            )
                        )
                        continue
                    if item_type == "promocion" and reprobadas:
                        rows.append(
                            _certificate_excel_row(
                                item_type,
                                "Omitido",
                                "Estudiante reprobado; no se permite generar certificado de promocion",
                                code,
                                generation_period,
                                context,
                                reprobadas,
                            )
                        )
                        continue
                    if item_type == "matricula" and not context.get("cabecera_matricula"):
                        rows.append(
                            _certificate_excel_row(
                                item_type,
                                "Omitido",
                                "No existe cabecera de matricula para generar el certificado de matricula",
                                code,
                                generation_period,
                                context,
                                reprobadas,
                            )
                        )
                        continue
                    if item_type == "matricula":
                        matricula_block_reason = _matricula_certificate_block_reason(context)
                        if matricula_block_reason:
                            rows.append(
                                _certificate_excel_row(
                                    item_type,
                                    "Omitido",
                                    matricula_block_reason,
                                    code,
                                    generation_period,
                                    context,
                                    reprobadas,
                                )
                            )
                            continue
                    rows.append(_certificate_excel_row(item_type, "Habilitado", "", code, generation_period, context, reprobadas))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="No se pudo generar el Excel") from exc

    filename = f"certificados_{certificate_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        _build_certificates_workbook(rows, payload, certificate_type),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{codigo_estud}/preview")
def preview_certificate(
    codigo_estud: str,
    periodo: str = Query(default=""),
    proximo_periodo: str = Query(default=""),
    cod_anio_basica: str = Query(default=""),
    periodo_matricula: str = Query(default=""),
    semestre: int | None = Query(default=None, ge=1, le=4),
    tipo: str = Query(default="matricula"),
    _: SessionUser = Depends(_CERTIFICATES_ACCESS),
) -> StreamingResponse:
    code = _clean(codigo_estud)
    period_code = _clean(periodo) or _clean(periodo_matricula)
    certificate_type = _clean(tipo).lower() or "matricula"
    if certificate_type not in {"matricula", "promocion"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de certificado no valido")
    if not code or not period_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Indica estudiante y periodo")

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            if not _active_student(cursor, code):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Estudiante no activo")
            reprobadas = _certificate_reprobadas(cursor, period_code, code, semestre) if certificate_type == "promocion" else []
            if certificate_type == "promocion" and reprobadas:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="No se puede generar el certificado de promocion: el estudiante tiene materias reprobadas.",
                )
            context = _fetch_certificate_context(
                cursor,
                code,
                period_code,
                _clean(proximo_periodo),
                semestre,
                _clean(cod_anio_basica),
                _clean(periodo_matricula),
            )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo preparar la vista previa del certificado",
        ) from exc

    if not context:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No se encontraron datos del estudiante")

    matricula_block_reason = _matricula_certificate_block_reason(context)
    if certificate_type == "matricula" and matricula_block_reason:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=matricula_block_reason,
        )
    pdf_bytes = _build_promocion_pdf(context) if certificate_type == "promocion" else _build_matricula_pdf(context)
    filename_prefix = "certificado-promocion" if certificate_type == "promocion" else "certificado-matricula"
    filename = f"{filename_prefix}-{_safe_filename(code)}.pdf"
    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )
