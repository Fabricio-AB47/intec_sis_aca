from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
import re
import unicodedata
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook

from app.core.security import SessionUser, require_roles
from app.services.db import get_connection

router = APIRouter(prefix="/api/students/validar-excel", tags=["validar-excel"])

AllowedUser = Depends(require_roles("ADMINISTRADOR", "ACADEMICO", "RECTOR"))

_MAX_FILE_SIZE = 12 * 1024 * 1024
_MAX_ROWS = 5000
_COLUMN_ALIASES = {
    "codigo": {
        "codigo",
        "codigo_estud",
        "cod_estud",
        "codestud",
        "codigo_estudiante",
        "estudiante_codigo",
        "id_estudiante",
        "cod_alumno",
    },
    "cedula": {
        "cedula",
        "cedula_est",
        "cedula_estudiante",
        "identificacion",
        "numero_identificacion",
        "documento",
        "num_documento",
        "dni",
        "cc",
    },
    "correo": {
        "correo",
        "email",
        "mail",
        "correo_personal",
        "correo_electronico",
        "email_personal",
    },
    "correo_intec": {
        "correo_intec",
        "correointec",
        "correo_institucional",
        "email_institucional",
        "mail_institucional",
    },
    "nombre": {
        "nombre",
        "nombres",
        "estudiante",
        "apellidos_nombre",
        "apellidos_nombres",
        "alumno",
        "participante",
        "nombre_completo",
    },
}


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _serializable(value: Any) -> Any:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bytes):
        return None
    return value


def _slug(value: Any) -> str:
    text = unicodedata.normalize("NFD", _clean(value).lower())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _normalize_name(value: Any) -> str:
    text = unicodedata.normalize("NFD", _clean(value).upper())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def _normalize_email(value: Any) -> str:
    return _clean(value).lower()


def _normalize_document(value: Any) -> str:
    digits = re.sub(r"\D+", "", _clean(value))
    if not digits:
        return ""
    if len(digits) == 13 and digits.endswith("001"):
        digits = digits[:10]
    elif len(digits) > 10:
        digits = digits[-10:]
    return digits.zfill(10) if len(digits) <= 10 else digits


def _excel_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _rows_from_cursor(cursor: Any) -> list[dict[str, Any]]:
    columns = [column[0] for column in cursor.description or []]
    return [{column: _serializable(value) for column, value in zip(columns, row)} for row in cursor.fetchall()]


def _chunks(values: set[str], size: int = 700) -> list[list[str]]:
    items = [value for value in values if value]
    return [items[index : index + size] for index in range(0, len(items), size)]


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows[:10]):
        non_empty = [_clean(cell) for cell in row if _clean(cell)]
        if len(non_empty) >= 2:
            return index
    return 0


def _unique_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        raw = _clean(value) or f"columna_{index + 1}"
        base = _slug(raw) or f"columna_{index + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _detect_columns(headers: list[str]) -> dict[str, str | None]:
    detected: dict[str, str | None] = {key: None for key in _COLUMN_ALIASES}
    for header in headers:
        normalized = _slug(header)
        for field_name, aliases in _COLUMN_ALIASES.items():
            if detected[field_name]:
                continue
            if normalized in aliases:
                detected[field_name] = header
    return detected


def _read_excel(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]], list[str]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix != "xlsx":
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Sube un archivo .xlsx")
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo supera el maximo permitido de 12 MB")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se pudo leer el Excel") from exc

    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel no contiene filas")

    header_index = _find_header_row(raw_rows)
    headers = _unique_headers(list(raw_rows[header_index]))
    data_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row_number, row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        values = list(row)
        if not any(_clean(value) for value in values):
            continue
        item = {
            headers[index]: _excel_value(values[index]) if index < len(values) else None
            for index in range(len(headers))
        }
        item["_excel_row"] = row_number
        data_rows.append(item)
        if len(data_rows) >= _MAX_ROWS:
            warnings.append(f"Se procesaron las primeras {_MAX_ROWS} filas con informacion.")
            break

    if not data_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel no contiene datos despues de la cabecera")
    return headers, data_rows, warnings


def _student_base_sql() -> str:
    return """
        SELECT
            TRY_CONVERT(varchar(50), de.codigo_estud) AS codigo_estud,
            LTRIM(RTRIM(TRY_CONVERT(varchar(50), de.Cedula_Est))) AS cedula,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.Apellidos_nombre))) AS estudiante,
            LTRIM(RTRIM(TRY_CONVERT(varchar(20), de.Estado))) AS estado_codigo,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(120), es.ESTADO))) AS estado,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correo))) AS correo,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correointec))) AS correo_intec,
            de.Fecha_Nac AS fecha_nacimiento,
            beca.tipo_beca,
            beca.porcentaje_beca,
            mat.periodo_codigo,
            mat.periodo,
            mat.carrera_codigo,
            mat.carrera,
            mat.materias_matriculadas
        FROM dbo.DATOS_ESTUD de
        LEFT JOIN dbo.ESTADO es ON de.Estado = es.IDESTADO
        OUTER APPLY (
            SELECT TOP (1)
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), B.tipo_beca))) AS tipo_beca,
                B.porcentaje_beca
            FROM dbo.Becas B
            WHERE TRY_CONVERT(varchar(50), B.codestud) = TRY_CONVERT(varchar(50), de.codigo_estud)
            ORDER BY TRY_CONVERT(decimal(10, 2), B.porcentaje_beca) DESC
        ) beca
        OUTER APPLY (
            SELECT TOP (1)
                TRY_CONVERT(varchar(50), ce.codigo_periodo) AS periodo_codigo,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), p.Detalle_Periodo))) AS periodo,
                TRY_CONVERT(varchar(50), ce.cod_anio_Basica) AS carrera_codigo,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), c.Nombre_Basica))) AS carrera,
                COUNT(*) OVER (PARTITION BY ce.codigo_estud, ce.codigo_periodo) AS materias_matriculadas,
                p.anio,
                p.Orden
            FROM dbo.CARRERAXESTUD ce
            LEFT JOIN dbo.PERIODO p ON ce.codigo_periodo = p.cod_periodo
            LEFT JOIN dbo.CARRERAS c ON ce.cod_anio_Basica = c.Cod_AnioBasica
            WHERE TRY_CONVERT(varchar(50), ce.codigo_estud) = TRY_CONVERT(varchar(50), de.codigo_estud)
            ORDER BY
                COALESCE(TRY_CONVERT(int, p.anio), 0) DESC,
                COALESCE(TRY_CONVERT(int, p.Orden), 0) DESC,
                TRY_CONVERT(int, ce.codigo_periodo) DESC
        ) mat
    """


def _lookup_students(cursor: Any, identifiers: dict[str, set[str]]) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    by_code: dict[str, dict[str, Any]] = {}
    by_cedula: dict[str, dict[str, Any]] = {}
    by_email: dict[str, dict[str, Any]] = {}
    by_name: dict[str, dict[str, Any]] = {}

    def add(row: dict[str, Any]) -> None:
        code = _clean(row.get("codigo_estud"))
        cedula = _normalize_document(row.get("cedula"))
        correo = _normalize_email(row.get("correo"))
        correo_intec = _normalize_email(row.get("correo_intec"))
        name = _normalize_name(row.get("estudiante"))
        if code and code not in by_code:
            by_code[code] = row
        if cedula and cedula not in by_cedula:
            by_cedula[cedula] = row
        for email in [correo, correo_intec]:
            if email and email not in by_email:
                by_email[email] = row
        if name and name not in by_name:
            by_name[name] = row

    lookup_plan = [
        ("codigo", "TRY_CONVERT(varchar(50), de.codigo_estud)", identifiers["codigo"]),
        ("cedula", "LTRIM(RTRIM(TRY_CONVERT(varchar(50), de.Cedula_Est)))", identifiers["cedula_raw"] | identifiers["cedula"]),
        ("correo", "LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correo))))", identifiers["correo"]),
        ("correo_intec", "LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correointec))))", identifiers["correo_intec"]),
        ("nombre", "UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.Apellidos_nombre))))", identifiers["nombre"]),
    ]
    for _key, column_sql, values in lookup_plan:
        for chunk in _chunks(values):
            placeholders = ", ".join("?" for _ in chunk)
            cursor.execute(f"{_student_base_sql()} WHERE {column_sql} IN ({placeholders})", chunk)
            for row in _rows_from_cursor(cursor):
                add(row)
    return by_code, by_cedula, by_email, by_name


def _lookup_simple_table(
    cursor: Any,
    table_sql: str,
    identifiers: dict[str, set[str]],
    field_map: dict[str, str],
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    by_code: dict[str, dict[str, Any]] = {}
    by_cedula: dict[str, dict[str, Any]] = {}
    by_email: dict[str, dict[str, Any]] = {}
    by_name: dict[str, dict[str, Any]] = {}

    def add(row: dict[str, Any]) -> None:
        code = _clean(row.get("codigo_estud"))
        cedula = _normalize_document(row.get("cedula"))
        correo = _normalize_email(row.get("correo"))
        correo_intec = _normalize_email(row.get("correo_intec"))
        name = _normalize_name(row.get("estudiante"))
        if code and code not in by_code:
            by_code[code] = row
        if cedula and cedula not in by_cedula:
            by_cedula[cedula] = row
        for email in [correo, correo_intec]:
            if email and email not in by_email:
                by_email[email] = row
        if name and name not in by_name:
            by_name[name] = row

    lookup_plan = [
        ("codigo", field_map.get("codigo"), identifiers["codigo"]),
        ("cedula", field_map.get("cedula"), identifiers["cedula_raw"] | identifiers["cedula"]),
        ("correo", field_map.get("correo"), identifiers["correo"]),
        ("correo_intec", field_map.get("correo_intec"), identifiers["correo_intec"]),
        ("nombre", field_map.get("nombre"), identifiers["nombre"]),
    ]
    for _key, column_sql, values in lookup_plan:
        if not column_sql:
            continue
        for chunk in _chunks(values):
            placeholders = ", ".join("?" for _ in chunk)
            cursor.execute(f"{table_sql} WHERE {column_sql} IN ({placeholders})", chunk)
            for row in _rows_from_cursor(cursor):
                add(row)
    return by_code, by_cedula, by_email, by_name


def _match_from_maps(
    identifiers: dict[str, str],
    maps: tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]]],
) -> tuple[dict[str, Any] | None, str]:
    by_code, by_cedula, by_email, by_name = maps
    if identifiers["codigo"] and identifiers["codigo"] in by_code:
        return by_code[identifiers["codigo"]], "codigo"
    if identifiers["cedula"] and identifiers["cedula"] in by_cedula:
        return by_cedula[identifiers["cedula"]], "cedula"
    for key in ("correo_intec", "correo"):
        if identifiers[key] and identifiers[key] in by_email:
            return by_email[identifiers[key]], key
    if identifiers["nombre"] and identifiers["nombre"] in by_name:
        return by_name[identifiers["nombre"]], "nombre"
    return None, ""


def _extract_identifiers(row: dict[str, Any], detected: dict[str, str | None]) -> dict[str, str]:
    def field_value(field: str) -> Any:
        column = detected.get(field)
        return row.get(column) if column else ""

    cedula_raw = _clean(field_value("cedula"))
    return {
        "codigo": _clean(field_value("codigo")),
        "cedula": _normalize_document(cedula_raw),
        "cedula_raw": cedula_raw,
        "correo": _normalize_email(field_value("correo")),
        "correo_intec": _normalize_email(field_value("correo_intec")),
        "nombre": _normalize_name(field_value("nombre")),
        "nombre_raw": _clean(field_value("nombre")),
    }


def _status_label(found_datos: bool, found_correos: bool, found_preinscripcion: bool, found_matricula: bool, has_identifier: bool) -> str:
    if not has_identifier:
        return "SIN_IDENTIFICADOR"
    if found_datos:
        return "ENCONTRADO"
    if found_correos or found_preinscripcion or found_matricula:
        return "PARCIAL"
    return "NO_ENCONTRADO"


@router.post("")
async def validate_excel(
    file: UploadFile = File(...),
    _: SessionUser = AllowedUser,
) -> dict[str, Any]:
    content = await file.read()
    headers, excel_rows, warnings = _read_excel(content, file.filename or "archivo.xlsx")
    detected = _detect_columns(headers)

    if not any(detected.values()):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se detectaron columnas de codigo, cedula, correo o nombre para cruzar.",
        )

    identifier_sets = {
        "codigo": set(),
        "cedula": set(),
        "cedula_raw": set(),
        "correo": set(),
        "correo_intec": set(),
        "nombre": set(),
    }
    row_identifiers: list[dict[str, str]] = []
    seen_keys: set[str] = set()
    duplicate_count = 0
    for row in excel_rows:
        identifiers = _extract_identifiers(row, detected)
        row_identifiers.append(identifiers)
        for field_name in identifier_sets:
            if identifiers.get(field_name):
                identifier_sets[field_name].add(identifiers[field_name])
        row_key = next((identifiers[key] for key in ["codigo", "cedula", "correo_intec", "correo", "nombre"] if identifiers.get(key)), "")
        if row_key:
            if row_key in seen_keys:
                duplicate_count += 1
            seen_keys.add(row_key)

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            datos_maps = _lookup_students(cursor, identifier_sets)
            correos_maps = _lookup_simple_table(
                cursor,
                """
                SELECT
                    TRY_CONVERT(varchar(50), codestud) AS codigo_estud,
                    '' AS cedula,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Nombres))) AS estudiante,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), CorreoPersonal))) AS correo,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), CorreoIntec))) AS correo_intec,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(60), Estado))) AS estado,
                    TRY_CONVERT(varchar(50), Periodo) AS periodo_codigo
                FROM dbo.CorreosEstudIntec
                """,
                identifier_sets,
                {
                    "codigo": "TRY_CONVERT(varchar(50), codestud)",
                    "correo": "LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), CorreoPersonal))))",
                    "correo_intec": "LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), CorreoIntec))))",
                    "nombre": "UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Nombres))))",
                },
            )
            preins_maps = _lookup_simple_table(
                cursor,
                """
                SELECT
                    TRY_CONVERT(varchar(50), Codestu) AS codigo_estud,
                    LTRIM(RTRIM(TRY_CONVERT(varchar(50), Cedula))) AS cedula,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Apellidos_nombre))) AS estudiante,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), correo))) AS correo,
                    '' AS correo_intec,
                    TRY_CONVERT(varchar(50), codperiodo) AS periodo_codigo,
                    TRY_CONVERT(varchar(50), codcarrera) AS carrera_codigo,
                    Prematricula AS prematricula,
                    ProcesoFinalilzado AS proceso_finalizado
                FROM dbo.PREINSCRIPCION
                """,
                identifier_sets,
                {
                    "codigo": "TRY_CONVERT(varchar(50), Codestu)",
                    "cedula": "LTRIM(RTRIM(TRY_CONVERT(varchar(50), Cedula)))",
                    "correo": "LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), correo))))",
                    "nombre": "UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), Apellidos_nombre))))",
                },
            )
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="No se pudo cruzar el Excel con SQL Server") from exc

    results: list[dict[str, Any]] = []
    totals = {
        "total": len(excel_rows),
        "encontrados": 0,
        "parciales": 0,
        "no_encontrados": 0,
        "sin_identificador": 0,
        "duplicados_excel": duplicate_count,
        "en_datos_estud": 0,
        "en_correos_intec": 0,
        "en_preinscripcion": 0,
        "con_matricula": 0,
    }

    for index, row in enumerate(excel_rows):
        identifiers = row_identifiers[index]
        has_identifier = any(identifiers.get(key) for key in ["codigo", "cedula", "correo", "correo_intec", "nombre"])
        datos_match, match_field = _match_from_maps(identifiers, datos_maps)
        correos_match, correos_field = _match_from_maps(identifiers, correos_maps)
        preins_match, preins_field = _match_from_maps(identifiers, preins_maps)
        found_matricula = bool(datos_match and datos_match.get("periodo_codigo"))
        status_label = _status_label(bool(datos_match), bool(correos_match), bool(preins_match), found_matricula, has_identifier)

        if status_label == "ENCONTRADO":
            totals["encontrados"] += 1
        elif status_label == "PARCIAL":
            totals["parciales"] += 1
        elif status_label == "NO_ENCONTRADO":
            totals["no_encontrados"] += 1
        else:
            totals["sin_identificador"] += 1
        if datos_match:
            totals["en_datos_estud"] += 1
        if correos_match:
            totals["en_correos_intec"] += 1
        if preins_match:
            totals["en_preinscripcion"] += 1
        if found_matricula:
            totals["con_matricula"] += 1

        db = datos_match or correos_match or preins_match or {}
        results.append(
            {
                "row_number": row.get("_excel_row"),
                "status": status_label,
                "match_field": match_field or correos_field or preins_field,
                "excel": {
                    "codigo": identifiers["codigo"],
                    "cedula": identifiers["cedula_raw"] or identifiers["cedula"],
                    "correo": identifiers["correo"],
                    "correo_intec": identifiers["correo_intec"],
                    "nombre": identifiers["nombre_raw"],
                },
                "exists": {
                    "datos_estud": bool(datos_match),
                    "correos_intec": bool(correos_match),
                    "preinscripcion": bool(preins_match),
                    "matricula": found_matricula,
                },
                "db": {
                    "codigo_estud": _clean(db.get("codigo_estud")),
                    "cedula": _clean(db.get("cedula")),
                    "estudiante": _clean(db.get("estudiante")),
                    "estado": _clean(db.get("estado")),
                    "correo": _clean(db.get("correo")),
                    "correo_intec": _clean(db.get("correo_intec")),
                    "tipo_beca": _clean(db.get("tipo_beca")),
                    "porcentaje_beca": _serializable(db.get("porcentaje_beca")),
                    "periodo": _clean(db.get("periodo")),
                    "periodo_codigo": _clean(db.get("periodo_codigo")),
                    "carrera": _clean(db.get("carrera")),
                    "carrera_codigo": _clean(db.get("carrera_codigo")),
                    "materias_matriculadas": _serializable(db.get("materias_matriculadas")),
                },
                "raw": {key: value for key, value in row.items() if key != "_excel_row"},
            }
        )

    if detected.get("nombre") and not any(detected.get(key) for key in ["codigo", "cedula", "correo", "correo_intec"]):
        warnings.append("El cruce se realizo solo por nombre; puede producir coincidencias parciales si existen homonimos.")

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "filename": file.filename,
        "sheet": "Activa",
        "columns": headers,
        "detected_columns": detected,
        "summary": totals,
        "rows": results,
        "warnings": warnings,
    }
