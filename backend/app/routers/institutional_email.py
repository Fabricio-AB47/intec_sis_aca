from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
import re
import unicodedata
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field, field_validator

from app.core.security import SessionUser, require_screen_access
from app.services.db import get_connection


router = APIRouter(prefix="/api/institutional-email", tags=["institutional-email"])
_SCREEN_ACCESS = require_screen_access("actualizar-correo-intec")

_MAX_FILE_SIZE = 12 * 1024 * 1024
_MAX_ROWS = 5000
_CEDULA_PATTERN = re.compile(r"^\d{10}$")
_INSTITUTIONAL_EMAIL_PATTERN = re.compile(
    r"^[a-z0-9.!#$%&'*+/=?^_`{|}~-]+@intec\.edu\.ec$",
    re.IGNORECASE,
)
_HEADER_ALIASES = {
    "cedula": {
        "cedula",
        "cedula_est",
        "cedula_estudiante",
        "identificacion",
        "numero_identificacion",
        "numero_de_identificacion",
        "documento",
    },
    "correo_intec": {
        "correo_intec",
        "correointec",
        "correo_intec_nuevo",
        "correo_institucional",
        "email_institucional",
    },
    "password": {
        "password",
        "password_nueva",
        "contrasena",
        "clave",
        "clave_nueva",
    },
}
_TEMPLATE_HEADERS = ["cedula", "correo_intec_nuevo", "password_nueva"]


class InstitutionalEmailUpdatePayload(BaseModel):
    correo_intec: str = Field(min_length=1, max_length=100)
    password: str = Field(min_length=6, max_length=20)

    @field_validator("correo_intec")
    @classmethod
    def validate_email(cls, value: str) -> str:
        return _validate_institutional_email(value)

    @field_validator("password")
    @classmethod
    def validate_password(cls, value: str) -> str:
        return _validate_password(value)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal) and value == value.to_integral_value():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _slug(value: Any) -> str:
    text = unicodedata.normalize("NFD", _clean(value).lower())
    text = "".join(character for character in text if unicodedata.category(character) != "Mn")
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def _normalize_excel_cedula(value: Any) -> str:
    raw = _clean(value)
    if not raw:
        return ""
    if raw.endswith(".0") and raw[:-2].isdigit():
        raw = raw[:-2]
    if not raw.isdigit() or len(raw) > 10:
        return raw
    return raw.zfill(10)


def _validate_cedula(value: str) -> str:
    cedula = str(value or "").strip()
    if not _CEDULA_PATTERN.fullmatch(cedula):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="La cedula debe contener exactamente 10 digitos.",
        )
    return cedula


def _validate_institutional_email(value: str) -> str:
    email = str(value or "").strip().lower()
    if not _INSTITUTIONAL_EMAIL_PATTERN.fullmatch(email):
        raise ValueError("El correo debe pertenecer al dominio @intec.edu.ec.")
    return email


def _validate_password(value: str) -> str:
    password = str(value or "")
    if password != password.strip():
        raise ValueError("La contrasena no puede iniciar ni terminar con espacios.")
    if len(password) < 6 or len(password) > 20:
        raise ValueError("La contrasena debe tener entre 6 y 20 caracteres.")
    if any(character in "\r\n\t" for character in password):
        raise ValueError("La contrasena contiene caracteres no permitidos.")
    return password


def _serialize(value: Any) -> Any:
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _row_dict(cursor: Any, row: Any) -> dict[str, Any]:
    columns = [column[0] for column in cursor.description or []]
    return {column: _serialize(value) for column, value in zip(columns, row)}


def _rows_dict(cursor: Any) -> list[dict[str, Any]]:
    columns = [column[0] for column in cursor.description or []]
    return [
        {column: _serialize(value) for column, value in zip(columns, row)}
        for row in cursor.fetchall()
    ]


def _student_select_sql() -> str:
    return """
        SELECT
            TRY_CONVERT(int, de.codigo_estud) AS codigo_estud,
            LTRIM(RTRIM(TRY_CONVERT(varchar(50), de.Cedula_Est))) AS cedula,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.Apellidos_nombre))) AS estudiante,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correo))) AS correo_personal,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), de.Estado))) AS estado,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correointec))) AS correo_intec_datos,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), ce.CorreoIntec))) AS correo_intec_registro,
            CASE
                WHEN NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), ce.CorreoIntec))), N'') IS NOT NULL
                    THEN LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), ce.CorreoIntec)))
                ELSE LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correointec)))
            END AS correo_intec,
            CASE WHEN ce.codestud IS NULL THEN CAST(0 AS bit) ELSE CAST(1 AS bit) END AS tiene_registro,
            CASE
                WHEN NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), ce.[Password]))), N'') IS NULL
                    THEN CAST(0 AS bit)
                ELSE CAST(1 AS bit)
            END AS password_configurada,
            CASE
                WHEN COALESCE(LOWER(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correointec))), N'')), N'') =
                     COALESCE(LOWER(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), ce.CorreoIntec))), N'')), N'')
                    THEN CAST(1 AS bit)
                ELSE CAST(0 AS bit)
            END AS sincronizado,
            carrera.carrera,
            carrera.periodo_codigo
        FROM dbo.DATOS_ESTUD de
        LEFT JOIN dbo.CorreosEstudIntec ce
          ON TRY_CONVERT(int, ce.codestud) = TRY_CONVERT(int, de.codigo_estud)
        OUTER APPLY (
            SELECT TOP (1)
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), c.Nombre_Basica))) AS carrera,
                TRY_CONVERT(int, cxe.codigo_periodo) AS periodo_codigo
            FROM dbo.CARRERAXESTUD cxe
            LEFT JOIN dbo.CARRERAS c
              ON c.Cod_AnioBasica = cxe.cod_anio_Basica
            LEFT JOIN dbo.PERIODO p
              ON p.cod_periodo = cxe.codigo_periodo
            WHERE TRY_CONVERT(int, cxe.codigo_estud) = TRY_CONVERT(int, de.codigo_estud)
            ORDER BY
                COALESCE(TRY_CONVERT(int, p.anio), 0) DESC,
                COALESCE(TRY_CONVERT(int, p.Orden), 0) DESC,
                TRY_CONVERT(int, cxe.codigo_periodo) DESC
        ) carrera
    """


def _ensure_tables(cursor: Any) -> None:
    cursor.execute(
        """
        SELECT
            CASE WHEN OBJECT_ID(N'dbo.DATOS_ESTUD', N'U') IS NOT NULL THEN 1 ELSE 0 END,
            CASE WHEN OBJECT_ID(N'dbo.CorreosEstudIntec', N'U') IS NOT NULL THEN 1 ELSE 0 END
        """
    )
    row = cursor.fetchone()
    if not row or not bool(row[0]) or not bool(row[1]):
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="No estan disponibles DATOS_ESTUD y CorreosEstudIntec en INTECBDD.",
        )


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows[:10]):
        normalized = {_slug(cell) for cell in row if _clean(cell)}
        recognized = sum(
            1
            for aliases in _HEADER_ALIASES.values()
            if any(value in aliases for value in normalized)
        )
        if recognized >= 2:
            return index
    return 0


def _detect_columns(headers: list[str]) -> dict[str, int]:
    detected: dict[str, int] = {}
    for index, header in enumerate(headers):
        for target, aliases in _HEADER_ALIASES.items():
            if target not in detected and header in aliases:
                detected[target] = index
    missing = [name for name in ("cedula", "correo_intec", "password") if name not in detected]
    if missing:
        labels = {
            "cedula": "cedula",
            "correo_intec": "correo_intec_nuevo",
            "password": "password_nueva",
        }
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan columnas obligatorias: {', '.join(labels[name] for name in missing)}.",
        )
    return detected


def _read_workbook(content: bytes, filename: str) -> list[dict[str, Any]]:
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sube un archivo Excel con extension .xlsx.")
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="El archivo supera el maximo permitido de 12 MB.")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel.") from exc

    worksheet = workbook["Actualizacion"] if "Actualizacion" in workbook.sheetnames else workbook.active
    raw_rows = list(worksheet.iter_rows(values_only=True, max_row=_MAX_ROWS + 12))
    workbook.close()
    if not raw_rows:
        raise HTTPException(status_code=400, detail="El archivo Excel esta vacio.")

    header_index = _find_header_row(raw_rows)
    headers = [_slug(value) for value in raw_rows[header_index]]
    columns = _detect_columns(headers)
    parsed: list[dict[str, Any]] = []
    for excel_row, values in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        cedula = _normalize_excel_cedula(values[columns["cedula"]] if columns["cedula"] < len(values) else None)
        email = _clean(values[columns["correo_intec"]] if columns["correo_intec"] < len(values) else None)
        password_value = values[columns["password"]] if columns["password"] < len(values) else None
        if isinstance(password_value, float) and password_value.is_integer():
            password = str(int(password_value))
        elif isinstance(password_value, Decimal) and password_value == password_value.to_integral_value():
            password = str(int(password_value))
        else:
            password = "" if password_value is None else str(password_value)

        # La plantilla incluye a todos los estudiantes como referencia. Solo se
        # procesa una fila cuando el operador informa al menos un dato nuevo.
        if not email and not password:
            continue
        parsed.append(
            {
                "row": excel_row,
                "cedula": cedula,
                "correo_intec": email,
                "password": password,
            }
        )
        if len(parsed) > _MAX_ROWS:
            raise HTTPException(status_code=400, detail=f"El archivo supera {_MAX_ROWS} filas de datos.")
    if not parsed:
        raise HTTPException(status_code=400, detail="No existen filas para actualizar en el archivo.")
    return parsed


def _build_template_bytes(cedulas: list[str]) -> bytes:
    """Build a plain XLSX range that Excel can open without repairing it."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Actualizacion"
    worksheet.append(_TEMPLATE_HEADERS)
    for cedula in cedulas:
        # Avoid materializing empty inline strings in thousands of table rows.
        worksheet.append([cedula])

    header_fill = PatternFill("solid", fgColor="A91F1B")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:C{max(len(cedulas) + 1, 1)}"
    worksheet.column_dimensions["A"].width = 18
    worksheet.column_dimensions["B"].width = 38
    worksheet.column_dimensions["C"].width = 24
    for cell in worksheet["A"][1:]:
        cell.number_format = "@"

    instructions = workbook.create_sheet("Instrucciones")
    instructions.append(["Actualizacion masiva de correo institucional"])
    instructions.append(["La cedula es la unica clave de identificacion incluida en la plantilla."])
    instructions.append(["El sistema valida su existencia y obtiene automaticamente el codigo_estud unico desde DATOS_ESTUD."])
    instructions.append(["No agregue codigo_estud, nombres, carrera ni otros campos de referencia."])
    instructions.append(["Complete correo_intec_nuevo y password_nueva solo para las cedulas que desea actualizar."])
    instructions.append(["La cedula debe tener 10 digitos y el correo debe terminar en @intec.edu.ec."])
    instructions.append(["La contrasena debe tener entre 6 y 20 caracteres."])
    instructions.append(["Las contrasenas existentes nunca se exportan."])
    instructions["A1"].font = Font(bold=True, size=14, color="A91F1B")
    instructions.column_dimensions["A"].width = 110

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _chunks(values: list[str], size: int = 500) -> list[list[str]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def _students_by_cedula(cursor: Any, cedulas: list[str]) -> dict[str, dict[str, Any]]:
    candidates: dict[str, list[dict[str, Any]]] = {}
    for chunk in _chunks(sorted(set(cedulas))):
        placeholders = ",".join("?" for _ in chunk)
        cursor.execute(
            f"""
            {_student_select_sql()}
            WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), de.Cedula_Est))) IN ({placeholders})
            """,
            *chunk,
        )
        for row in _rows_dict(cursor):
            cedula = str(row.get("cedula") or "")
            candidates.setdefault(cedula, []).append(row)

    found: dict[str, dict[str, Any]] = {}
    for cedula, rows in candidates.items():
        codes = sorted(
            {
                int(row.get("codigo_estud") or 0)
                for row in rows
                if int(row.get("codigo_estud") or 0) > 0
            }
        )
        selected = next(
            (
                row
                for row in rows
                if codes and int(row.get("codigo_estud") or 0) == codes[0]
            ),
            rows[0],
        )
        found[cedula] = {**selected, "_codigos_estud": codes}
    return found


def _unique_student_code(student: dict[str, Any] | None) -> int | None:
    if not student:
        return None
    codes = student.get("_codigos_estud")
    if isinstance(codes, list):
        return int(codes[0]) if len(codes) == 1 else None
    code = int(student.get("codigo_estud") or 0)
    return code if code > 0 else None


def _email_owners(cursor: Any, emails: list[str]) -> dict[str, set[int]]:
    owners: dict[str, set[int]] = {}
    for chunk in _chunks(sorted(set(emails))):
        placeholders = ",".join("?" for _ in chunk)
        cursor.execute(
            f"""
            SELECT correo, codigo_estud
            FROM (
                SELECT
                    LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), ce.CorreoIntec)))) AS correo,
                    TRY_CONVERT(int, ce.codestud) AS codigo_estud
                FROM dbo.CorreosEstudIntec ce
                UNION ALL
                SELECT
                    LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.correointec)))) AS correo,
                    TRY_CONVERT(int, de.codigo_estud) AS codigo_estud
                FROM dbo.DATOS_ESTUD de
            ) source
            WHERE source.correo IN ({placeholders})
            """,
            *chunk,
        )
        for row in cursor.fetchall():
            email = _clean(row[0]).lower()
            code = int(row[1]) if row[1] is not None else 0
            owners.setdefault(email, set()).add(code)
    return owners


def _analyze_rows(cursor: Any, parsed: list[dict[str, Any]]) -> dict[str, Any]:
    students = _students_by_cedula(cursor, [str(item["cedula"]) for item in parsed])
    normalized_emails: list[str] = []
    for item in parsed:
        try:
            normalized_emails.append(_validate_institutional_email(str(item["correo_intec"])))
        except ValueError:
            continue
    owners = _email_owners(cursor, normalized_emails) if normalized_emails else {}

    cedula_counts: dict[str, int] = {}
    email_counts: dict[str, int] = {}
    for item in parsed:
        cedula = str(item["cedula"])
        cedula_counts[cedula] = cedula_counts.get(cedula, 0) + 1
        email = str(item["correo_intec"]).strip().lower()
        if email:
            email_counts[email] = email_counts.get(email, 0) + 1

    preview: list[dict[str, Any]] = []
    valid_rows: list[dict[str, Any]] = []
    for item in parsed:
        row_number = int(item["row"])
        cedula = str(item["cedula"])
        student = students.get(cedula)
        student_code = _unique_student_code(student)
        messages: list[str] = []
        new_email = ""
        password = str(item["password"])

        if not _CEDULA_PATTERN.fullmatch(cedula):
            messages.append("La cedula debe contener exactamente 10 digitos")
        if cedula_counts.get(cedula, 0) > 1:
            messages.append("La cedula esta repetida en el archivo")
        if student is None and _CEDULA_PATTERN.fullmatch(cedula):
            messages.append("La cedula no existe en DATOS_ESTUD")
        elif student is not None and student_code is None:
            codes = student.get("_codigos_estud")
            if isinstance(codes, list) and len(codes) > 1:
                messages.append("La cedula esta asociada a mas de un codigo de estudiante")
            else:
                messages.append("La cedula no tiene un codigo unico valido")
        try:
            new_email = _validate_institutional_email(str(item["correo_intec"]))
        except ValueError as exc:
            messages.append(str(exc))
        try:
            password = _validate_password(password)
        except ValueError as exc:
            messages.append(str(exc))
        if new_email and email_counts.get(new_email, 0) > 1:
            messages.append("El correo esta repetido en el archivo")
        if student_code is not None and new_email:
            conflicting_owners = {owner for owner in owners.get(new_email, set()) if owner != student_code}
            if conflicting_owners:
                messages.append("El correo ya pertenece a otro estudiante")

        status_value = "ERROR" if messages else "VALIDO"
        preview_row = {
            "row": row_number,
            "cedula": cedula,
            "codigo_estud": student_code,
            "estudiante": student.get("estudiante") if student else "",
            "correo_actual": student.get("correo_intec") if student else "",
            "correo_nuevo": new_email or str(item["correo_intec"]).strip().lower(),
            "password_informada": bool(password),
            "estado": status_value,
            "detalle": "; ".join(messages) if messages else "Lista para actualizar",
        }
        preview.append(preview_row)
        if not messages and student and student_code is not None:
            valid_rows.append(
                {
                    **student,
                    "codigo_estud": student_code,
                    "correo_nuevo": new_email,
                    "password_nueva": password,
                }
            )

    errors = sum(1 for row in preview if row["estado"] == "ERROR")
    return {
        "rows": preview,
        "valid_rows": valid_rows,
        "summary": {
            "total": len(preview),
            "validos": len(preview) - errors,
            "errores": errors,
        },
    }


def _upsert_student_email(cursor: Any, row: dict[str, Any]) -> None:
    code = int(row.get("codigo_estud") or 0)
    if code <= 0:
        raise ValueError("El estudiante no tiene codigo valido.")
    email = str(row["correo_nuevo"])
    password = str(row["password_nueva"])
    student_name = _clean(row.get("estudiante"))[:100]
    personal_email = _clean(row.get("correo_personal"))[:100]
    period_code = int(row.get("periodo_codigo") or 0)

    cursor.execute(
        """
        IF EXISTS (
            SELECT 1
            FROM dbo.CorreosEstudIntec WITH (UPDLOCK, HOLDLOCK)
            WHERE TRY_CONVERT(int, codestud) = ?
        )
        BEGIN
            UPDATE dbo.CorreosEstudIntec
               SET Nombres = ?,
                   CorreoPersonal = ?,
                   CorreoIntec = ?,
                   [Password] = ?,
                   fecha = CONVERT(date, GETDATE()),
                   Periodo = CASE WHEN ? > 0 THEN ? ELSE Periodo END,
                   Estado = COALESCE(NULLIF(LTRIM(RTRIM(Estado)), ''), 'PENDIENTE')
             WHERE TRY_CONVERT(int, codestud) = ?;
        END
        ELSE
        BEGIN
            INSERT INTO dbo.CorreosEstudIntec
                (codestud, Nombres, CorreoPersonal, CorreoIntec, [Password], fecha,
                 Periodo, CorreoEnviado, Estado, TipoCursoMigra)
            VALUES
                (?, ?, ?, ?, ?, CONVERT(date, GETDATE()), ?, 0, 'PENDIENTE', 'N');
        END;

        UPDATE dbo.DATOS_ESTUD
           SET correointec = ?,
               clave = ?
         WHERE TRY_CONVERT(int, codigo_estud) = ?;
        """,
        code,
        student_name,
        personal_email,
        email,
        password,
        period_code,
        period_code,
        code,
        code,
        student_name,
        personal_email,
        email,
        password,
        period_code,
        email,
        password,
        code,
    )


@router.get("/students")
def list_students(
    response: Response,
    cedula: str = Query(default=""),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=10, le=100),
    _: SessionUser = Depends(_SCREEN_ACCESS),
) -> dict[str, Any]:
    normalized_cedula = _validate_cedula(cedula) if cedula.strip() else ""
    offset = (page - 1) * page_size
    response.headers["Cache-Control"] = "no-store"
    with get_connection() as connection:
        cursor = connection.cursor()
        _ensure_tables(cursor)
        where_sql = ""
        params: list[Any] = []
        if normalized_cedula:
            where_sql = "WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), de.Cedula_Est))) = ?"
            params.append(normalized_cedula)
        cursor.execute(f"SELECT COUNT_BIG(1) FROM dbo.DATOS_ESTUD de {where_sql}", *params)
        total = int(cursor.fetchone()[0])
        cursor.execute(
            f"""
            {_student_select_sql()}
            {where_sql}
            ORDER BY
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.Apellidos_nombre))),
                TRY_CONVERT(int, de.codigo_estud)
            OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
            """,
            *params,
            offset,
            page_size,
        )
        rows = _rows_dict(cursor)
    return {"rows": rows, "total": total, "page": page, "page_size": page_size, "cedula": normalized_cedula}


@router.get("/template")
def download_template(
    _: SessionUser = Depends(_SCREEN_ACCESS),
) -> StreamingResponse:
    with get_connection() as connection:
        cursor = connection.cursor()
        _ensure_tables(cursor)
        cursor.execute(
            """
            SELECT DISTINCT
                LTRIM(RTRIM(TRY_CONVERT(varchar(50), Cedula_Est))) AS cedula
            FROM dbo.DATOS_ESTUD
            WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), Cedula_Est))) LIKE
                  REPLICATE('[0-9]', 10)
            ORDER BY cedula
            """
        )
        cedulas = [str(row[0]) for row in cursor.fetchall() if row[0]]

    output = BytesIO(_build_template_bytes(cedulas))
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="actualizacion_correo_intec.xlsx"',
            "Cache-Control": "no-store",
        },
    )


@router.post("/analyze")
async def analyze_workbook(
    response: Response,
    file: UploadFile = File(...),
    _: SessionUser = Depends(_SCREEN_ACCESS),
) -> dict[str, Any]:
    content = await file.read()
    parsed = _read_workbook(content, file.filename or "")
    with get_connection() as connection:
        cursor = connection.cursor()
        _ensure_tables(cursor)
        analysis = _analyze_rows(cursor, parsed)
    response.headers["Cache-Control"] = "no-store"
    analysis.pop("valid_rows", None)
    return analysis


@router.post("/apply")
async def apply_workbook(
    response: Response,
    file: UploadFile = File(...),
    _: SessionUser = Depends(_SCREEN_ACCESS),
) -> dict[str, Any]:
    content = await file.read()
    parsed = _read_workbook(content, file.filename or "")
    connection = get_connection()
    try:
        cursor = connection.cursor()
        _ensure_tables(cursor)
        analysis = _analyze_rows(cursor, parsed)
        errors = int(analysis["summary"]["errores"])
        if errors:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"El archivo contiene {errors} fila(s) con errores. Analiza y corrige antes de aplicar.",
            )
        for row in analysis["valid_rows"]:
            _upsert_student_email(cursor, row)
        connection.commit()
    except HTTPException:
        connection.rollback()
        raise
    except Exception as exc:
        connection.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo aplicar la actualizacion de correos institucionales.",
        ) from exc
    finally:
        connection.close()
    response.headers["Cache-Control"] = "no-store"
    return {
        "ok": True,
        "actualizados": len(analysis["valid_rows"]),
        "message": f"Se actualizaron {len(analysis['valid_rows'])} estudiante(s).",
    }


@router.put("/students/{cedula}")
def update_student_email(
    cedula: str,
    payload: InstitutionalEmailUpdatePayload,
    response: Response,
    _: SessionUser = Depends(_SCREEN_ACCESS),
) -> dict[str, Any]:
    normalized_cedula = _validate_cedula(cedula)
    connection = get_connection()
    try:
        cursor = connection.cursor()
        _ensure_tables(cursor)
        students = _students_by_cedula(cursor, [normalized_cedula])
        student = students.get(normalized_cedula)
        if not student:
            raise HTTPException(status_code=404, detail="No existe el estudiante en DATOS_ESTUD.")
        student_code = _unique_student_code(student)
        if student_code is None:
            raise HTTPException(
                status_code=409,
                detail="La cedula no esta asociada a un codigo unico de estudiante.",
            )
        owners = _email_owners(cursor, [payload.correo_intec])
        if any(owner != student_code for owner in owners.get(payload.correo_intec, set())):
            raise HTTPException(status_code=409, detail="El correo ya pertenece a otro estudiante.")
        _upsert_student_email(
            cursor,
            {
                **student,
                "codigo_estud": student_code,
                "correo_nuevo": payload.correo_intec,
                "password_nueva": payload.password,
            },
        )
        connection.commit()
    except HTTPException:
        connection.rollback()
        raise
    except Exception as exc:
        connection.rollback()
        raise HTTPException(status_code=500, detail="No se pudo actualizar el correo institucional.") from exc
    finally:
        connection.close()
    response.headers["Cache-Control"] = "no-store"
    return {
        "ok": True,
        "message": "Correo institucional y contrasena actualizados.",
        "cedula": normalized_cedula,
        "correo_intec": payload.correo_intec,
    }
