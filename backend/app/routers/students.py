import csv
from datetime import date, datetime, timezone
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
import math
import re
import subprocess
from typing import Annotated, Any
import unicodedata
from urllib.parse import unquote

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from pydantic import BaseModel, Field
import pyodbc

from app.core.security import SessionUser, require_roles
from app.services.db import get_connection

router = APIRouter(prefix="/api/students", tags=["students"])

_ALLOWED_TIPO_MATRICULA = {"R", "H"}
_STUDENT_ACCESS = require_roles("ADMINISTRADOR", "ACADEMICO", "BIENESTAR", "RECTOR", "VICERRECTOR")
_DASHBOARD_ACCESS = require_roles("ADMINISTRADOR", "ACADEMICO", "BIENESTAR", "RECTOR", "VICERRECTOR", "ADMISIONES")
_GRADUATION_ACCESS = require_roles("ADMINISTRADOR", "ACADEMICO", "BIENESTAR", "RECTOR", "VICERRECTOR", "SECRETARIA")
_MAIN_ESTADOS = (
    ("A", "Activo"),
    ("G", "Graduado"),
    ("P", "Inactivo"),
    ("R", "Retirado"),
)
_PROJECT_ROOT = Path(__file__).resolve().parents[3]
_REGISTRO_PATH = _PROJECT_ROOT / "registro.xlsx"
_DATA_MOODLE_PATH = _PROJECT_ROOT / "data_moodle.xlsx"
_MATRIZ_DESAGREGADA_PATH = _PROJECT_ROOT / "matriz_desagregada.xls"
_MATRIZ_DESAGREGADA_CACHE_PATH = _PROJECT_ROOT / "matriz_desagregada_nombres.csv"
_MATRIZ_DESAGREGADA_TEMP_CACHE_PATH = _PROJECT_ROOT / "tmp_matriz_names.csv"
_COMPACT_KEY_MIN_LENGTH = 8
_DEFAULT_CROSS_DB_LIMIT = 0
_EXPECTED_ACTIVE_SQL_TOTAL = 866
_DASHBOARD_IGNORED_CEDULA = "1708531189"
_DASHBOARD_IGNORED_CEDULA_SQL = (
    "COALESCE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), bc.Cedula_Est))), '') <> '1708531189'"
)
_SPANISH_MONTHS = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


class GraduationDateItem(BaseModel):
    codigo_estud: str = Field(min_length=1, max_length=50)
    fecha_grado: str | None = None
    fecha_emision_senescyt: str | None = None
    cod_refrendacion: str | None = Field(default=None, max_length=50)


class GraduationDateSavePayload(BaseModel):
    items: list[GraduationDateItem] = Field(min_length=1, max_length=1000)


class DataUpdatePayload(BaseModel):
    fields: dict[str, str | int | float | bool | None] = Field(default_factory=dict)


class _HtmlTableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() == "tr":
            self._current_row = []
        elif tag.lower() in {"td", "th"} and self._current_row is not None:
            self._current_cell = []
            self._in_cell = True

    def handle_data(self, data: str) -> None:
        if self._in_cell and self._current_cell is not None:
            self._current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        normalized_tag = tag.lower()
        if normalized_tag in {"td", "th"} and self._current_row is not None:
            value = _clean_cell(" ".join(self._current_cell or []))
            self._current_row.append(value)
            self._current_cell = None
            self._in_cell = False
        elif normalized_tag == "tr" and self._current_row is not None:
            if any(cell for cell in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _normalize_name(value: Any) -> str:
    text = _clean_cell(value)
    if " - " in text:
        text = text.split(" - ", 1)[0]
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _name_match_keys(*values: Any) -> list[str]:
    keys: list[str] = []
    for value in values:
        normalized = _normalize_name(value)
        if normalized and normalized not in keys:
            keys.append(normalized)
        compact = normalized.replace(" ", "")
        if len(compact) >= _COMPACT_KEY_MIN_LENGTH and compact not in keys:
            keys.append(compact)
    return keys


def _looks_like_email(value: str) -> bool:
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value.strip().lower()))


def _normalize_email(value: Any) -> str:
    return _clean_cell(value).lower()


def _comparison_flag(left: Any, right: Any, *, kind: str = "text") -> str:
    if kind == "email":
        left_value = _normalize_email(left)
        right_value = _normalize_email(right)
    elif kind == "name":
        left_value = _normalize_name(left)
        right_value = _normalize_name(right)
    else:
        left_value = _clean_cell(left).upper()
        right_value = _clean_cell(right).upper()
    if not left_value or not right_value:
        return "SIN_DATO"
    return "SI" if left_value == right_value else "NO"


def _looks_like_hash(value: str) -> bool:
    return bool(re.fullmatch(r"[a-f0-9]{24,}", value.strip().lower()))


def _normalize_identifier(value: Any) -> str:
    digits = re.sub(r"\D+", "", _clean_cell(value))
    if not digits:
        return ""
    if len(digits) == 13 and digits.endswith("001"):
        digits = digits[:10]
    elif len(digits) > 10:
        digits = digits[-10:]
    return digits.zfill(10)


def _identifier_match_keys(*values: Any) -> list[str]:
    keys: list[str] = []
    for value in values:
        normalized = _normalize_identifier(value)
        if normalized and normalized not in keys:
            keys.append(f"CEDULA:{normalized}")
    return keys


def _read_html_table_rows(path: Path) -> list[list[str]]:
    parser = _HtmlTableParser()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    return parser.rows


def _excel_html_sheet_path(path: Path) -> Path | None:
    text = path.read_text(encoding="utf-8", errors="replace")
    matches = re.findall(r'(?:HRef|href|src)=["\']([^"\']*sheet\d+\.htm)["\']', text, flags=re.IGNORECASE)
    if not matches:
        return None
    relative_sheet = Path(unquote(matches[0]))
    candidates = [
        path.parent / relative_sheet,
        path.parent / relative_sheet.name,
        path.parent / f"{path.stem}_files" / relative_sheet.name,
        path.parent / f"{path.name}_files" / relative_sheet.name,
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]


def _column_index(headers: list[str], patterns: list[str]) -> int | None:
    normalized_headers = [_normalize_name(header) for header in headers]
    for pattern in patterns:
        normalized_pattern = _normalize_name(pattern)
        for index, header in enumerate(normalized_headers):
            if header == normalized_pattern:
                return index
    for pattern in patterns:
        normalized_pattern = _normalize_name(pattern)
        for index, header in enumerate(normalized_headers):
            if normalized_pattern and normalized_pattern in header:
                return index
    return None


def _cell_from_row(row: list[str], index: int | None) -> str:
    if index is None or index >= len(row):
        return ""
    return _clean_cell(row[index])


def _read_xlsx_rows(path: Path) -> list[list[str]]:
    dataframe = pd.read_excel(path, sheet_name=0, header=None, dtype=object, engine="openpyxl")
    dataframe = dataframe.where(pd.notna(dataframe), "")
    rows: list[list[str]] = []
    for row in dataframe.itertuples(index=False, name=None):
        values = [_clean_cell(cell) for cell in row]
        if any(values):
            rows.append(values)
    return rows


def _quote_powershell_literal(value: Path) -> str:
    return "'" + str(value).replace("'", "''") + "'"


def _refresh_matriz_desagregada_cache() -> None:
    if not _MATRIZ_DESAGREGADA_PATH.exists():
        raise FileNotFoundError(f"No se encontro {_MATRIZ_DESAGREGADA_PATH.name}")

    script = f"""
$ErrorActionPreference = 'Stop'
$path = {_quote_powershell_literal(_MATRIZ_DESAGREGADA_PATH)}
$out = {_quote_powershell_literal(_MATRIZ_DESAGREGADA_CACHE_PATH)}
$excel = $null
$workbook = $null
$worksheet = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($path)
    $worksheet = $workbook.Worksheets.Item(1)
    $used = $worksheet.UsedRange
    $rows = New-Object System.Collections.Generic.List[object]
    $currentCareer = ''
    for ($r = 1; $r -le $used.Rows.Count; $r++) {{
        $c1 = [string]$worksheet.Cells.Item($r, 1).Text
        $c2 = [string]$worksheet.Cells.Item($r, 2).Text
        $c3 = [string]$worksheet.Cells.Item($r, 3).Text
        if ($c1.Trim() -like 'Carrera:*') {{
            $currentCareer = $c2
        }}
        $num = 0
        if ([int]::TryParse($c1.Trim(), [ref]$num) -and -not [string]::IsNullOrWhiteSpace($c2)) {{
            $rows.Add([pscustomobject]@{{
                numero = $num
                nombre = $c2.Trim()
                documento = $c3.Trim()
                carrera = $currentCareer.Trim()
            }})
        }}
    }}
    $rows | Export-Csv -Path $out -NoTypeInformation -Encoding UTF8
}} finally {{
    if ($workbook -ne $null) {{
        $workbook.Close($false) | Out-Null
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null
    }}
    if ($worksheet -ne $null) {{
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($worksheet) | Out-Null
    }}
    if ($excel -ne $null) {{
        $excel.Quit() | Out-Null
        [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
    }}
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
}}
"""
    result = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
        capture_output=True,
        text=True,
        timeout=90,
    )
    if result.returncode != 0:
        detail = _clean_cell(result.stderr or result.stdout)
        raise RuntimeError(detail or "No se pudo leer matriz_desagregada.xls con Excel")


def _matriz_desagregada_cache_path(warnings: list[str]) -> Path | None:
    if not _MATRIZ_DESAGREGADA_PATH.exists():
        warnings.append(f"No se encontro {_MATRIZ_DESAGREGADA_PATH.name}; se usa solo la base de datos.")
        return None

    cache_is_current = (
        _MATRIZ_DESAGREGADA_CACHE_PATH.exists()
        and _MATRIZ_DESAGREGADA_CACHE_PATH.stat().st_mtime >= _MATRIZ_DESAGREGADA_PATH.stat().st_mtime
    )
    if cache_is_current:
        return _MATRIZ_DESAGREGADA_CACHE_PATH

    try:
        _refresh_matriz_desagregada_cache()
        if _MATRIZ_DESAGREGADA_CACHE_PATH.exists():
            return _MATRIZ_DESAGREGADA_CACHE_PATH
    except Exception as exc:
        warnings.append(f"No se pudo actualizar la lectura de {_MATRIZ_DESAGREGADA_PATH.name}: {exc}")

    for fallback in (_MATRIZ_DESAGREGADA_CACHE_PATH, _MATRIZ_DESAGREGADA_TEMP_CACHE_PATH):
        if fallback.exists():
            warnings.append(f"Se usa cache existente {fallback.name} para la referencia de matriz.")
            return fallback
    return None


def _parse_matriz_cutoff(value: Any) -> date | None:
    text = unicodedata.normalize("NFD", _clean_cell(value))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    normalized = re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()
    match = re.search(r"(\d{1,2}) ([A-Z]+) (\d{4})", normalized)
    if not match:
        return None
    day = int(match.group(1))
    month = _SPANISH_MONTHS.get(match.group(2))
    year = int(match.group(3))
    if not month:
        return None
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _coerce_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_cell(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%Y %H:%M:%S"):
        try:
            return datetime.strptime(text[:19] if "%H" in fmt else text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _read_matriz_desagregada_reference() -> dict[str, Any]:
    warnings: list[str] = []
    cache_path = _matriz_desagregada_cache_path(warnings)
    if cache_path is None:
        return {
            "names": set(),
            "rows": 0,
            "unique_names": 0,
            "duplicate_names": 0,
            "cutoff_date": None,
            "warnings": warnings,
        }

    rows_count = 0
    names: set[str] = set()
    cutoff_date: date | None = None
    with cache_path.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        for row in reader:
            name_key = _normalize_name(row.get("nombre"))
            if not name_key:
                continue
            rows_count += 1
            names.add(name_key)
            cutoff_date = cutoff_date or _parse_matriz_cutoff(row.get("carrera"))

    if cutoff_date is None and _MATRIZ_DESAGREGADA_PATH.exists():
        cutoff_date = datetime.fromtimestamp(_MATRIZ_DESAGREGADA_PATH.stat().st_mtime).date()

    return {
        "names": names,
        "rows": rows_count,
        "unique_names": len(names),
        "duplicate_names": max(rows_count - len(names), 0),
        "cutoff_date": cutoff_date,
        "warnings": warnings,
    }


def _row_matches_matriz_reference(row: Any, reference: dict[str, Any]) -> bool:
    names: set[str] = reference.get("names") or set()
    if not names:
        return True

    if _normalize_name(getattr(row, "Apellidos_nombre", "")) in names:
        return True

    cutoff_date = reference.get("cutoff_date")
    if cutoff_date is None:
        return False
    for field_name in ("Fecha_Ingreso", "fechaMatricula"):
        row_date = _coerce_date(getattr(row, field_name, None))
        if row_date and row_date > cutoff_date:
            return True
    return False


def _matricula_reference_meta(reference: dict[str, Any], total_base: int, total_filtrado: int) -> dict[str, Any]:
    cutoff_date = reference.get("cutoff_date")
    return {
        "archivo": _MATRIZ_DESAGREGADA_PATH.name,
        "filas_matriz": int(reference.get("rows") or 0),
        "nombres_unicos_matriz": int(reference.get("unique_names") or 0),
        "duplicados_por_nombre": int(reference.get("duplicate_names") or 0),
        "fecha_corte": cutoff_date.isoformat() if isinstance(cutoff_date, date) else None,
        "total_base_sql": total_base,
        "total_filtrado": total_filtrado,
        "incluye_ingresos_posteriores": True,
        "warnings": reference.get("warnings") or [],
    }


def _records_dataframe(records: list[dict[str, Any]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    return pd.DataFrame.from_records(records).fillna("")


def _dataframe_records(dataframe: pd.DataFrame) -> list[dict[str, Any]]:
    if dataframe.empty:
        return []
    return dataframe.to_dict(orient="records")


def _read_registro(path: Path = _REGISTRO_PATH) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if not path.exists():
        return [], [f"No se encontro {path.name}"]

    if path.suffix.lower() == ".xlsx":
        rows = _read_xlsx_rows(path)
    else:
        rows = _read_html_table_rows(path)
    if not rows and path.suffix.lower() != ".xlsx":
        sheet_path = _excel_html_sheet_path(path)
        if sheet_path is None:
            return [], [f"No se encontro una hoja de datos en {path.name}"]
        if not sheet_path.exists():
            return [], [
                f"{path.name} es un contenedor HTML de Excel y referencia {sheet_path.parent.name}\\{sheet_path.name}, pero esa hoja no existe en la carpeta del proyecto"
            ]
        rows = _read_html_table_rows(sheet_path)
    if not rows:
        return [], [f"{path.name} no contiene filas de datos"]

    header_row: int | None = None
    cedula_index: int | None = None
    nombre_index: int | None = None
    nombres_index: int | None = None
    apellidos_index: int | None = None
    nombre_comercial_index: int | None = None
    correo_index: int | None = None
    balance_index: int | None = None
    credito_index: int | None = None
    canal_index: int | None = None
    ciudad_index: int | None = None
    direccion_index: int | None = None

    for row_index, row in enumerate(rows):
        cedula_candidate = _column_index(row, ["CEDULA", "IDENTIFICACION", "IDENTIFICACION", "DOCUMENTO", "CI", "RUC"])
        nombre_candidate = _column_index(
            row,
            [
                "APELLIDOS NOMBRES",
                "APELLIDOS Y NOMBRES",
                "NOMBRES APELLIDOS",
                "NOMBRE COMPLETO",
                "ESTUDIANTE",
                "CLIENTE",
                "RAZON SOCIAL",
                "NOMBRE",
            ],
        )
        nombres_candidate = _column_index(row, ["NOMBRES", "NOMBRE"])
        apellidos_candidate = _column_index(row, ["APELLIDOS", "APELLIDO"])
        nombre_comercial_candidate = _column_index(row, ["N. COMERCIAL", "N COMERCIAL", "NOMBRE COMERCIAL"])
        if cedula_candidate is not None and (
            nombre_candidate is not None or nombres_candidate is not None or apellidos_candidate is not None
        ):
            header_row = row_index
            cedula_index = cedula_candidate
            nombre_index = nombre_candidate
            nombres_index = nombres_candidate
            apellidos_index = apellidos_candidate
            nombre_comercial_index = nombre_comercial_candidate
            correo_index = _column_index(row, ["CORREO", "EMAIL", "MAIL"])
            balance_index = _column_index(row, ["BALANCE", "SALDO"])
            credito_index = _column_index(row, ["CREDITO", "CREDITO"])
            canal_index = _column_index(row, ["CANAL"])
            ciudad_index = _column_index(row, ["CIUDAD"])
            direccion_index = _column_index(row, ["DIRECCION", "DIRECCION"])
            break

    if header_row is None:
        return [], [f"No se detecto la cabecera de cedula/nombres en {path.name}"]

    records: list[dict[str, Any]] = []
    seen_records: set[str] = set()
    for row in rows[header_row + 1 :]:
        cedula_raw = _cell_from_row(row, cedula_index)
        cedula_normalizada = _normalize_identifier(cedula_raw)
        nombre_completo = _cell_from_row(row, nombre_index)
        nombre_comercial = _cell_from_row(row, nombre_comercial_index)
        if not nombre_completo:
            nombre_completo = " ".join(
                item
                for item in [_cell_from_row(row, apellidos_index), _cell_from_row(row, nombres_index)]
                if item
            ).strip()
        if not nombre_completo:
            nombre_completo = nombre_comercial
        if not nombre_completo and not cedula_normalizada:
            continue

        keys = _identifier_match_keys(cedula_raw) + _name_match_keys(nombre_completo, nombre_comercial)
        unique_key = cedula_normalizada or (keys[0] if keys else _normalize_name(nombre_completo))
        if unique_key in seen_records:
            continue
        seen_records.add(unique_key)

        records.append(
            {
                "source": "registro",
                "razon_social": nombre_completo,
                "nombre_comercial": nombre_comercial,
                "identificacion": cedula_normalizada,
                "identificacion_raw": cedula_raw,
                "correo": _cell_from_row(row, correo_index),
                "balance": _cell_from_row(row, balance_index),
                "credito": _cell_from_row(row, credito_index),
                "canal": _cell_from_row(row, canal_index),
                "ciudad": _cell_from_row(row, ciudad_index),
                "direccion": _cell_from_row(row, direccion_index),
                "nombre_validacion": nombre_completo,
                "keys": keys,
            }
        )

    if not records:
        warnings.append(f"{path.name} no tiene registros para cruzar")
    return records, warnings


def _read_moodle_users(path: Path = _DATA_MOODLE_PATH) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    if not path.exists():
        return [], [f"No se encontro {path.name}"]

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    raw_headers = next(worksheet.iter_rows(max_row=1, values_only=True), ())
    headers = [_clean_cell(header).lower() for header in raw_headers]
    header_index = {header: index for index, header in enumerate(headers) if header}
    required_headers = {"firstname", "lastname", "email", "username"}

    if not required_headers.issubset(header_index):
        return [], [f"{path.name} no tiene las columnas Moodle esperadas"]

    def cell(row: tuple[Any, ...], header: str) -> Any:
        index = header_index.get(header)
        if index is None or index >= len(row):
            return None
        return row[index]

    candidates: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        firstname = _clean_cell(cell(row, "firstname"))
        lastname = _clean_cell(cell(row, "lastname"))
        email = _clean_cell(cell(row, "email"))
        username = _clean_cell(cell(row, "username"))
        idnumber = _clean_cell(cell(row, "idnumber"))
        deleted = _clean_cell(cell(row, "deleted"))
        suspended = _clean_cell(cell(row, "suspended"))
        if not firstname and not lastname and not email and not username:
            continue

        nombre_apellidos_primero = f"{lastname} {firstname}".strip()
        nombre_nombres_primero = f"{firstname} {lastname}".strip()
        keys = _identifier_match_keys(idnumber) + _name_match_keys(nombre_apellidos_primero, nombre_nombres_primero)
        primary_name = keys[0] if keys else ""
        normalized_email = email.lower()
        normalized_username = username.lower()
        if (
            not primary_name
            or primary_name == "INVITADO"
            or normalized_username == "guest"
            or normalized_email == "root@localhost"
        ):
            continue

        candidates.append(
            {
                "source": "moodle",
                "moodle_id": _clean_cell(cell(row, "id")),
                "firstname": firstname,
                "lastname": lastname,
                "nombre_validacion": nombre_apellidos_primero or nombre_nombres_primero,
                "email": email,
                "username": username,
                "idnumber": idnumber,
                "deleted": deleted,
                "suspended": suspended,
                "lastaccess": _clean_cell(cell(row, "lastaccess")),
                "lastlogin": _clean_cell(cell(row, "lastlogin")),
                "timemodified": _clean_cell(cell(row, "timemodified")),
                "keys": keys,
            }
        )

    records, duplicate_count = _deduplicate_moodle_records(candidates)

    if not records:
        warnings.append(f"{path.name} no tiene usuarios Moodle para cruzar")
    if duplicate_count > 0:
        warnings.append(f"Moodle: {duplicate_count} duplicados eliminados por nombre normalizado")
    return records, warnings


def _moodle_record_score(record: dict[str, Any]) -> tuple[int, int, int, int, int]:
    email = str(record.get("email") or "").strip().lower()
    username = str(record.get("username") or "").strip().lower()
    is_active = str(record.get("deleted") or "") != "1" and str(record.get("suspended") or "") != "1"
    has_valid_email = _looks_like_email(email) and email != "root@localhost"
    has_intec_email = has_valid_email and email.endswith("@intec.edu.ec")
    username_has_email = _looks_like_email(username.split(".1", 1)[0]) or _looks_like_email(username)
    is_hash_email = _looks_like_hash(email)
    activity = max(
        int(value)
        for value in [
            str(record.get("lastlogin") or "0"),
            str(record.get("lastaccess") or "0"),
            str(record.get("timemodified") or "0"),
            "0",
        ]
        if value.isdigit()
    )
    return (
        1 if is_active else 0,
        1 if has_intec_email else 0,
        1 if has_valid_email and not is_hash_email else 0,
        1 if username_has_email else 0,
        activity,
    )


def _merge_moodle_candidate(
    entities: list[dict[str, Any]],
    key_index: dict[str, int],
    record: dict[str, Any],
) -> None:
    keys = [key for key in record.get("keys", []) if key]
    matched_indexes = sorted({key_index[key] for key in keys if key in key_index})

    if not matched_indexes:
        entities.append({"keys": set(keys), "records": [record]})
        target_index = len(entities) - 1
    else:
        target_index = matched_indexes[0]
        for merge_index in reversed(matched_indexes[1:]):
            target_entity = entities[target_index]
            merged_entity = entities[merge_index]
            target_entity["keys"].update(merged_entity["keys"])
            target_entity["records"].extend(merged_entity["records"])
            for key in merged_entity["keys"]:
                key_index[key] = target_index
            entities.pop(merge_index)
            for key, index in list(key_index.items()):
                if index == merge_index:
                    key_index[key] = target_index
                elif index > merge_index:
                    key_index[key] = index - 1
        entities[target_index]["keys"].update(keys)
        entities[target_index]["records"].append(record)

    for key in entities[target_index]["keys"]:
        key_index[key] = target_index


def _deduplicate_moodle_records(candidates: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    entities: list[dict[str, Any]] = []
    key_index: dict[str, int] = {}
    for candidate in candidates:
        _merge_moodle_candidate(entities, key_index, candidate)

    deduplicated: list[dict[str, Any]] = []
    duplicate_count = 0
    for entity in entities:
        records = entity["records"]
        best_record = max(records, key=_moodle_record_score)
        best_record = dict(best_record)
        best_record["registros_fuente"] = len(records)
        best_record["duplicados_eliminados"] = max(0, len(records) - 1)
        best_record["keys"] = sorted(entity["keys"])
        duplicate_count += max(0, len(records) - 1)
        deduplicated.append(best_record)

    deduplicated.sort(key=lambda record: str(record.get("nombre_validacion") or ""))
    return deduplicated, duplicate_count


def _read_sql_student_sample(db_limit: int) -> tuple[list[dict[str, Any]], list[str]]:
    top_clause = "TOP (?)" if db_limit > 0 else ""
    query = f"""
        WITH latest_student_data AS (
            SELECT
                d.*,
                ROW_NUMBER() OVER (
                    PARTITION BY TRY_CONVERT(varchar(50), d.codigo_estud)
                    ORDER BY
                        COALESCE(
                            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 121),
                            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 120),
                            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 103),
                            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 105),
                            TRY_CONVERT(datetime2, d.Fecha_Ingreso),
                            CAST('1900-01-01' AS datetime2)
                        ) DESC,
                        COALESCE(TRY_CONVERT(int, d.NumMigracion), -1) DESC
                ) AS rn_data
            FROM [dbo].[DATOS_ESTUD] d
            WHERE d.Apellidos_nombre IS NOT NULL
        )
        SELECT {top_clause}
            d.codigo_estud,
            d.Cedula_Est,
            d.Apellidos_nombre,
            d.correo,
            d.correointec,
            d.Estado,
            cx.cod_anio_Basica,
            cx.codigo_materia,
            cx.codigo_periodo,
            cx.Num_Matricula,
            cx.paralelo,
            cx.NumGrupo,
            cx.Promedio,
            cx.Asistencia,
            cx.Recuperacion,
            cx.PromedioFinal,
            cx.caprueba,
            cx.Num_Creditos,
            cx.Fecha_Matricula,
            cx.Num_Folio,
            cx.TipoMatricula,
            cx.PromedioAux,
            cx.ControlAprueba,
            cx.ControlMatricula,
            cx.estadoMoodle,
            cx.NumMigracion AS cxe_num_migracion,
            cx.TipoCursoMigra AS cxe_tipo_curso_migra,
            cx.num AS cxe_num,
            pensum.Unidad_Organiza AS pensum_unidad_organiza,
            pensum.Nomb_Materia AS pensum_nomb_materia,
            pensum.Semestre AS pensum_semestre,
            pensum.Creditos AS pensum_creditos,
            pensum.Orden AS pensum_orden,
            pensum.NumMalla AS pensum_num_malla,
            pensum.cod_materia AS pensum_cod_materia,
            pensum.Horas AS pensum_horas,
            pensum.ValorHora AS pensum_valor_hora,
            pensum.ValorHoraVirtual AS pensum_valor_hora_virtual,
            pensum.CombinarMateria AS pensum_combinar_materia,
            pensum.verreporte AS pensum_ver_reporte,
            pensum.SecuenciaMateria AS pensum_secuencia_materia,
            pensum.tipomateria AS pensum_tipo_materia
        FROM latest_student_data d
        OUTER APPLY (
            SELECT TOP (1)
                cxe.cod_anio_Basica,
                cxe.codigo_materia,
                cxe.codigo_periodo,
                cxe.Num_Matricula,
                cxe.paralelo,
                cxe.NumGrupo,
                cxe.Promedio,
                cxe.Asistencia,
                cxe.Recuperacion,
                cxe.PromedioFinal,
                cxe.caprueba,
                cxe.Num_Creditos,
                cxe.Fecha_Matricula,
                cxe.Num_Folio,
                cxe.TipoMatricula,
                cxe.PromedioAux,
                cxe.ControlAprueba,
                cxe.ControlMatricula,
                cxe.estadoMoodle,
                cxe.NumMigracion,
                cxe.TipoCursoMigra,
                cxe.num
            FROM [dbo].[CARRERAXESTUD] cxe
            WHERE TRY_CONVERT(varchar(50), cxe.codigo_estud) = TRY_CONVERT(varchar(50), d.codigo_estud)
            ORDER BY
                COALESCE(
                    TRY_CONVERT(datetime2, cxe.Fecha_Matricula, 121),
                    TRY_CONVERT(datetime2, cxe.Fecha_Matricula, 120),
                    TRY_CONVERT(datetime2, cxe.Fecha_Matricula, 103),
                    TRY_CONVERT(datetime2, cxe.Fecha_Matricula, 105),
                    TRY_CONVERT(datetime2, cxe.Fecha_Matricula),
                    CAST('1900-01-01' AS datetime2)
                ) DESC,
                TRY_CONVERT(int, cxe.codigo_periodo) DESC,
                TRY_CONVERT(int, cxe.Num_Matricula) DESC,
                TRY_CONVERT(int, cxe.Num_Reg_Mat) DESC,
                TRY_CONVERT(int, cxe.num) DESC
        ) cx
        OUTER APPLY (
            SELECT TOP (1)
                p.Unidad_Organiza,
                p.Nomb_Materia,
                p.Semestre,
                p.Creditos,
                p.Orden,
                p.NumMalla,
                p.cod_materia,
                p.Horas,
                p.ValorHora,
                p.ValorHoraVirtual,
                p.CombinarMateria,
                p.verreporte,
                p.SecuenciaMateria,
                p.tipomateria
            FROM [dbo].[PENSUM] p
            WHERE TRY_CONVERT(varchar(50), p.Cod_AnioBasica) = TRY_CONVERT(varchar(50), cx.cod_anio_Basica)
              AND (
                  TRY_CONVERT(varchar(50), p.codigo_materia) = TRY_CONVERT(varchar(50), cx.codigo_materia)
                  OR TRY_CONVERT(varchar(50), p.cod_materia) = TRY_CONVERT(varchar(50), cx.codigo_materia)
              )
            ORDER BY
                TRY_CONVERT(int, p.NumMalla) DESC,
                TRY_CONVERT(int, p.Orden) DESC,
                TRY_CONVERT(int, p.Semestre) DESC
        ) pensum
        WHERE d.rn_data = 1
          AND UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Estado)))) IN ('A', 'G', 'P', 'R')
          AND NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), cx.codigo_materia))), '') IS NOT NULL
        ORDER BY d.Apellidos_nombre
    """

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            if db_limit > 0:
                cursor.execute(query, (db_limit,))
            else:
                cursor.execute(query)
            rows = cursor.fetchall()
    except Exception as exc:
        return [], [f"No se pudo consultar DATOS_ESTUD: {exc}"]

    records: list[dict[str, Any]] = []
    seen_students: set[str] = set()
    for row in rows:
        nombre = _clean_cell(row.Apellidos_nombre)
        cedula = _clean_cell(row.Cedula_Est)
        correo_datos = _clean_cell(row.correo)
        correointec_datos = _clean_cell(row.correointec)
        codigo_estud = _clean_cell(row.codigo_estud)
        estado_codigo = _clean_cell(row.Estado)
        estado_nombre = {
            "A": "Activo",
            "G": "Graduado",
            "P": "Inactivo",
            "R": "Retirado",
        }.get(estado_codigo.upper(), estado_codigo)
        keys = _identifier_match_keys(cedula) + _name_match_keys(nombre)
        unique_key = codigo_estud or _normalize_identifier(cedula) or (keys[0] if keys else nombre)
        if unique_key in seen_students:
            continue
        seen_students.add(unique_key)

        records.append(
            {
                "source": "tablas",
                "codigo_estud": codigo_estud,
                "cedula": _normalize_identifier(cedula),
                "cedula_raw": cedula,
                "nombre_validacion": nombre,
                "correo": correo_datos,
                "correointec": correointec_datos,
                "nombre_final": nombre,
                "correo_final": correo_datos,
                "correointec_final": correointec_datos,
                "estado": estado_codigo,
                "estado_nombre": estado_nombre,
                "estado_final": estado_nombre,
                "cod_anio_basica": _clean_cell(row.cod_anio_Basica),
                "nombre_basica": "",
                "carrera_estado": "",
                "carrera_abrevia": "",
                "tp_escuela": "",
                "codigo_materia": _clean_cell(row.codigo_materia),
                "codigo_periodo": _clean_cell(row.codigo_periodo),
                "periodo_final": _clean_cell(row.codigo_periodo),
                "fecha_pago": "",
                "valor_matricula": "",
                "inscrip_valor": "",
                "matri_valor": "",
                "beca": "",
                "descuento": "",
                "jornada_matricula": "",
                "control_matricula": "",
                "periodo_nombre": "",
                "detalle_periodo": "",
                "tipo_matricula_periodo": "",
                "anio_periodo": "",
                "estado_periodo": "",
                "periodo_fecha_inicio": "",
                "periodo_fecha_fin": "",
                "origen_tablas_sql": "DATOS_ESTUD+CARRERAXESTUD+PENSUM"
                if _clean_cell(row.codigo_materia)
                else "DATOS_ESTUD",
                "en_carreraxestud": bool(_clean_cell(row.codigo_materia)),
                "periodos_vinculados": "",
                "num_matricula": _clean_cell(row.Num_Matricula),
                "paralelo": _clean_cell(row.paralelo),
                "num_grupo": _clean_cell(row.NumGrupo),
                "promedio": _clean_cell(row.Promedio),
                "asistencia": _clean_cell(row.Asistencia),
                "recuperacion": _clean_cell(row.Recuperacion),
                "promedio_final": _clean_cell(row.PromedioFinal),
                "caprueba": _clean_cell(row.caprueba),
                "num_creditos": _clean_cell(row.Num_Creditos),
                "fecha_matricula": _clean_cell(row.Fecha_Matricula),
                "num_folio": _clean_cell(row.Num_Folio),
                "tipo_matricula_cxe": _clean_cell(row.TipoMatricula),
                "promedio_aux": _clean_cell(row.PromedioAux),
                "control_aprueba": _clean_cell(row.ControlAprueba),
                "control_matricula_cxe": _clean_cell(row.ControlMatricula),
                "estado_moodle": _clean_cell(row.estadoMoodle),
                "cxe_num_migracion": _clean_cell(row.cxe_num_migracion),
                "cxe_tipo_curso_migra": _clean_cell(row.cxe_tipo_curso_migra),
                "cxe_num": _clean_cell(row.cxe_num),
                "nivel_semestre": _clean_cell(row.pensum_semestre),
                "pensum_unidad_organiza": _clean_cell(row.pensum_unidad_organiza),
                "pensum_nomb_materia": _clean_cell(row.pensum_nomb_materia),
                "pensum_semestre": _clean_cell(row.pensum_semestre),
                "pensum_creditos": _clean_cell(row.pensum_creditos),
                "pensum_orden": _clean_cell(row.pensum_orden),
                "pensum_num_malla": _clean_cell(row.pensum_num_malla),
                "pensum_cod_materia": _clean_cell(row.pensum_cod_materia),
                "pensum_horas": _clean_cell(row.pensum_horas),
                "pensum_valor_hora": _clean_cell(row.pensum_valor_hora),
                "pensum_valor_hora_virtual": _clean_cell(row.pensum_valor_hora_virtual),
                "pensum_combinar_materia": _clean_cell(row.pensum_combinar_materia),
                "pensum_ver_reporte": _clean_cell(row.pensum_ver_reporte),
                "pensum_secuencia_materia": _clean_cell(row.pensum_secuencia_materia),
                "pensum_tipo_materia": _clean_cell(row.pensum_tipo_materia),
                "correos_intec_encontrado": False,
                "correos_intec_usado": False,
                "correos_intec_ignorado": False,
                "datos_estud_prevalece": True,
                "correos_codestud": "",
                "correos_nombres": "",
                "correos_correo_personal": "",
                "correos_correo_intec": "",
                "correos_fecha": "",
                "correos_periodo": "",
                "correos_correo_enviado": "",
                "correos_estado": "",
                "correos_descripcion": "",
                "correos_ult_acceso_moodle": "",
                "correos_num_migracion": "",
                "correos_tipo_curso_migra": "",
                "correos_nombre_coincide": "",
                "correos_personal_coincide": "",
                "correos_intec_coincide": "",
                "correos_periodo_coincide": "",
                "correos_estado_coincide": "",
                "keys": keys,
            }
        )

    return records, []


def _read_sql_active_validation_counts() -> tuple[dict[str, int], list[str]]:
    records, warnings = _read_sql_student_sample(0)
    total = len(records)
    return (
        {
            "datos_estud_activos": total,
            "activos_con_carreraxestud": total,
            "activos_con_carrera_no_ingles": total,
            "activos_sin_carreraxestud": 0,
            "activos_excluidos_ingles_o_sin_carrera": 0,
        },
        warnings,
    )


def _merge_entity_records(
    entities: list[dict[str, Any]],
    key_index: dict[str, int],
    source: str,
    record: dict[str, Any],
) -> None:
    keys = [key for key in record.get("keys", []) if key]
    matched_indexes = sorted({key_index[key] for key in keys if key in key_index})

    if not matched_indexes:
        entities.append({"keys": set(keys), "balance": [], "moodle": [], "tablas": []})
        target_index = len(entities) - 1
    else:
        target_index = matched_indexes[0]
        for merge_index in reversed(matched_indexes[1:]):
            target_entity = entities[target_index]
            merged_entity = entities[merge_index]
            target_entity["keys"].update(merged_entity["keys"])
            target_entity["balance"].extend(merged_entity["balance"])
            target_entity["moodle"].extend(merged_entity["moodle"])
            target_entity["tablas"].extend(merged_entity["tablas"])
            for key in merged_entity["keys"]:
                key_index[key] = target_index
            entities.pop(merge_index)
            for key, index in list(key_index.items()):
                if index == merge_index:
                    key_index[key] = target_index
                elif index > merge_index:
                    key_index[key] = index - 1
        entities[target_index]["keys"].update(keys)

    entities[target_index][source].append(record)
    for key in entities[target_index]["keys"]:
        key_index[key] = target_index


def _cruce_estado(has_balance: bool, has_moodle: bool, has_tablas: bool) -> str:
    if has_balance and has_moodle and has_tablas:
        return "EN_TODOS"
    if has_balance and has_moodle:
        return "BALANCE_MOODLE"
    if has_balance and has_tablas:
        return "BALANCE_TABLAS"
    if has_moodle and has_tablas:
        return "MOODLE_TABLAS"
    if has_balance:
        return "SOLO_BALANCE"
    if has_moodle:
        return "SOLO_MOODLE"
    return "SOLO_TABLAS"


def _cross_result(has_balance: bool, has_moodle: bool, has_tablas: bool) -> str:
    sources_count = int(has_balance) + int(has_moodle) + int(has_tablas)
    return "CRUZADA" if sources_count >= 2 else "NO_CRUZADA"


def _single_source_label(has_balance: bool, has_moodle: bool, has_tablas: bool) -> str:
    if has_balance and not has_moodle and not has_tablas:
        return "Registro"
    if has_moodle and not has_balance and not has_tablas:
        return "Moodle"
    if has_tablas and not has_balance and not has_moodle:
        return "DATOS_ESTUD"
    return ""


def _first_record(records: list[dict[str, Any]]) -> dict[str, Any]:
    return records[0] if records else {}


def _principal_source(has_balance: bool, has_moodle: bool, has_tablas: bool) -> str:
    if has_tablas:
        return "DATOS_ESTUD"
    if has_moodle:
        return "Moodle"
    if has_balance:
        return "Registro"
    return ""


def _count_duplicate_values(values: list[str]) -> int:
    seen: set[str] = set()
    duplicated: set[str] = set()
    for value in values:
        if not value:
            continue
        if value in seen:
            duplicated.add(value)
        else:
            seen.add(value)
    return len(duplicated)


def _entity_to_cross_row(entity: dict[str, Any]) -> dict[str, Any]:
    balance = _first_record(entity["balance"])
    moodle = _first_record(entity["moodle"])
    tablas = _first_record(entity["tablas"])
    has_balance = bool(entity["balance"])
    has_moodle = bool(entity["moodle"])
    has_tablas = bool(entity["tablas"])
    estado = _cruce_estado(has_balance, has_moodle, has_tablas)
    resultado = _cross_result(has_balance, has_moodle, has_tablas)
    nombre_validado = (
        tablas.get("nombre_validacion")
        or moodle.get("nombre_validacion")
        or balance.get("razon_social")
        or ""
    )
    correos_moodle_email_coincide = _comparison_flag(
        moodle.get("email", ""),
        tablas.get("correos_correo_intec", ""),
        kind="email",
    )

    return {
        "fuente_principal": _principal_source(has_balance, has_moodle, has_tablas),
        "nombre_validado": nombre_validado,
        "clave_normalizada": _normalize_name(nombre_validado),
        "estado_cruce": estado,
        "resultado_cruce": resultado,
        "origen_no_cruzado": _single_source_label(has_balance, has_moodle, has_tablas)
        if resultado == "NO_CRUZADA"
        else "",
        "en_balance": has_balance,
        "en_moodle": has_moodle,
        "en_tablas": has_tablas,
        "balance": {
            "razon_social": balance.get("razon_social", ""),
            "nombre_comercial": balance.get("nombre_comercial", ""),
            "identificacion": balance.get("identificacion", ""),
            "identificacion_raw": balance.get("identificacion_raw", ""),
            "correo": balance.get("correo", ""),
            "balance": balance.get("balance", ""),
            "credito": balance.get("credito", ""),
            "canal": balance.get("canal", ""),
            "ciudad": balance.get("ciudad", ""),
            "direccion": balance.get("direccion", ""),
            "registros": len(entity["balance"]),
        },
        "moodle": {
            "nombre": moodle.get("nombre_validacion", ""),
            "email": moodle.get("email", ""),
            "username": moodle.get("username", ""),
            "idnumber": moodle.get("idnumber", ""),
            "deleted": moodle.get("deleted", ""),
            "suspended": moodle.get("suspended", ""),
            "registros": moodle.get("registros_fuente", len(entity["moodle"])),
            "duplicados_eliminados": moodle.get("duplicados_eliminados", 0),
        },
        "tablas": {
            "codigo_estud": tablas.get("codigo_estud", ""),
            "cedula": tablas.get("cedula", ""),
            "cedula_raw": tablas.get("cedula_raw", ""),
            "nombre": tablas.get("nombre_validacion", ""),
            "correo": tablas.get("correo", ""),
            "correointec": tablas.get("correointec", ""),
            "nombre_final": tablas.get("nombre_final", tablas.get("nombre_validacion", "")),
            "correo_final": tablas.get("correo_final", tablas.get("correo", "")),
            "correointec_final": tablas.get("correointec_final", tablas.get("correointec", "")),
            "estado_final": tablas.get("estado_final", tablas.get("estado_nombre", "")),
            "periodo_final": tablas.get("periodo_final", tablas.get("codigo_periodo", "")),
            "estado": tablas.get("estado", ""),
            "estado_nombre": tablas.get("estado_nombre", ""),
            "cod_anio_basica": tablas.get("cod_anio_basica", ""),
            "nombre_basica": tablas.get("nombre_basica", ""),
            "carrera_estado": tablas.get("carrera_estado", ""),
            "carrera_abrevia": tablas.get("carrera_abrevia", ""),
            "tp_escuela": tablas.get("tp_escuela", ""),
            "codigo_periodo": tablas.get("codigo_periodo", ""),
            "fecha_pago": tablas.get("fecha_pago", ""),
            "valor_matricula": tablas.get("valor_matricula", ""),
            "inscrip_valor": tablas.get("inscrip_valor", ""),
            "matri_valor": tablas.get("matri_valor", ""),
            "beca": tablas.get("beca", ""),
            "descuento": tablas.get("descuento", ""),
            "jornada_matricula": tablas.get("jornada_matricula", ""),
            "control_matricula": tablas.get("control_matricula", ""),
            "periodo_nombre": tablas.get("periodo_nombre", ""),
            "detalle_periodo": tablas.get("detalle_periodo", ""),
            "tipo_matricula_periodo": tablas.get("tipo_matricula_periodo", ""),
            "anio_periodo": tablas.get("anio_periodo", ""),
            "estado_periodo": tablas.get("estado_periodo", ""),
            "periodo_fecha_inicio": tablas.get("periodo_fecha_inicio", ""),
            "periodo_fecha_fin": tablas.get("periodo_fecha_fin", ""),
            "origen_tablas_sql": tablas.get("origen_tablas_sql", ""),
            "en_carreraxestud": bool(tablas.get("en_carreraxestud")),
            "periodos_vinculados": tablas.get("periodos_vinculados", ""),
            "codigo_materia": tablas.get("codigo_materia", ""),
            "num_matricula": tablas.get("num_matricula", ""),
            "paralelo": tablas.get("paralelo", ""),
            "num_grupo": tablas.get("num_grupo", ""),
            "promedio": tablas.get("promedio", ""),
            "asistencia": tablas.get("asistencia", ""),
            "recuperacion": tablas.get("recuperacion", ""),
            "promedio_final": tablas.get("promedio_final", ""),
            "caprueba": tablas.get("caprueba", ""),
            "num_creditos": tablas.get("num_creditos", ""),
            "fecha_matricula": tablas.get("fecha_matricula", ""),
            "num_folio": tablas.get("num_folio", ""),
            "tipo_matricula_cxe": tablas.get("tipo_matricula_cxe", ""),
            "promedio_aux": tablas.get("promedio_aux", ""),
            "control_aprueba": tablas.get("control_aprueba", ""),
            "control_matricula_cxe": tablas.get("control_matricula_cxe", ""),
            "estado_moodle": tablas.get("estado_moodle", ""),
            "cxe_num_migracion": tablas.get("cxe_num_migracion", ""),
            "cxe_tipo_curso_migra": tablas.get("cxe_tipo_curso_migra", ""),
            "cxe_num": tablas.get("cxe_num", ""),
            "nivel_semestre": tablas.get("nivel_semestre", ""),
            "pensum_unidad_organiza": tablas.get("pensum_unidad_organiza", ""),
            "pensum_nomb_materia": tablas.get("pensum_nomb_materia", ""),
            "pensum_semestre": tablas.get("pensum_semestre", ""),
            "pensum_creditos": tablas.get("pensum_creditos", ""),
            "pensum_orden": tablas.get("pensum_orden", ""),
            "pensum_num_malla": tablas.get("pensum_num_malla", ""),
            "pensum_cod_materia": tablas.get("pensum_cod_materia", ""),
            "pensum_horas": tablas.get("pensum_horas", ""),
            "pensum_valor_hora": tablas.get("pensum_valor_hora", ""),
            "pensum_valor_hora_virtual": tablas.get("pensum_valor_hora_virtual", ""),
            "pensum_combinar_materia": tablas.get("pensum_combinar_materia", ""),
            "pensum_ver_reporte": tablas.get("pensum_ver_reporte", ""),
            "pensum_secuencia_materia": tablas.get("pensum_secuencia_materia", ""),
            "pensum_tipo_materia": tablas.get("pensum_tipo_materia", ""),
            "correos_intec_encontrado": bool(tablas.get("correos_intec_encontrado")),
            "correos_intec_usado": bool(tablas.get("correos_intec_usado")),
            "correos_intec_ignorado": bool(tablas.get("correos_intec_ignorado")),
            "datos_estud_prevalece": bool(tablas.get("datos_estud_prevalece")),
            "correos_codestud": tablas.get("correos_codestud", ""),
            "correos_nombres": tablas.get("correos_nombres", ""),
            "correos_correo_personal": tablas.get("correos_correo_personal", ""),
            "correos_correo_intec": tablas.get("correos_correo_intec", ""),
            "correos_fecha": tablas.get("correos_fecha", ""),
            "correos_periodo": tablas.get("correos_periodo", ""),
            "correos_correo_enviado": tablas.get("correos_correo_enviado", ""),
            "correos_estado": tablas.get("correos_estado", ""),
            "correos_descripcion": tablas.get("correos_descripcion", ""),
            "correos_ult_acceso_moodle": tablas.get("correos_ult_acceso_moodle", ""),
            "correos_num_migracion": tablas.get("correos_num_migracion", ""),
            "correos_tipo_curso_migra": tablas.get("correos_tipo_curso_migra", ""),
            "correos_nombre_coincide": tablas.get("correos_nombre_coincide", ""),
            "correos_personal_coincide": tablas.get("correos_personal_coincide", ""),
            "correos_intec_coincide": tablas.get("correos_intec_coincide", ""),
            "correos_periodo_coincide": tablas.get("correos_periodo_coincide", ""),
            "correos_estado_coincide": tablas.get("correos_estado_coincide", ""),
            "correos_moodle_email_coincide": correos_moodle_email_coincide,
            "registros": len(entity["tablas"]),
        },
    }


def _build_excel_sql_cross(limit: int, db_limit: int) -> dict[str, Any]:
    registro_records, registro_warnings = _read_registro()
    moodle_records, moodle_warnings = _read_moodle_users()
    sql_records, sql_warnings = _read_sql_student_sample(db_limit)
    active_validation, active_validation_warnings = _read_sql_active_validation_counts()
    warnings = registro_warnings + moodle_warnings + sql_warnings + active_validation_warnings
    registro_df = _records_dataframe(registro_records)
    moodle_df = _records_dataframe(moodle_records)
    sql_df = _records_dataframe(sql_records)

    entities: list[dict[str, Any]] = []
    key_index: dict[str, int] = {}
    for record in _dataframe_records(sql_df):
        _merge_entity_records(entities, key_index, "tablas", record)
    for record in _dataframe_records(moodle_df):
        _merge_entity_records(entities, key_index, "moodle", record)
    for record in _dataframe_records(registro_df):
        _merge_entity_records(entities, key_index, "balance", record)

    all_rows_df = _records_dataframe([_entity_to_cross_row(entity) for entity in entities])
    if all_rows_df.empty:
        rows_df = all_rows_df
    else:
        rows_df = all_rows_df[all_rows_df["en_tablas"].astype(bool)].copy()
        rows_df["_sort_cruzada"] = rows_df["resultado_cruce"].eq("CRUZADA").map({True: 0, False: 1})
        rows_df["_sort_todos"] = rows_df["estado_cruce"].eq("EN_TODOS").map({True: 0, False: 1})
        rows_df = rows_df.sort_values(["_sort_cruzada", "_sort_todos", "nombre_validado"], kind="stable")
        rows_df = rows_df.drop(columns=["_sort_cruzada", "_sort_todos"])
    rows = _dataframe_records(rows_df)
    table_codes = [
        str((row.get("tablas") or {}).get("codigo_estud") or "").strip()
        for row in rows
    ]
    table_ids = [
        str((row.get("tablas") or {}).get("cedula") or "").strip()
        for row in rows
    ]
    duplicate_code_count = _count_duplicate_values(table_codes)
    duplicate_id_count = _count_duplicate_values(table_ids)
    active_difference = 0
    cxe_rows = [row for row in rows if (row.get("tablas") or {}).get("en_carreraxestud")]
    pensum_rows = [row for row in rows if (row.get("tablas") or {}).get("pensum_semestre")]
    correos_intec_rows = [row for row in rows if (row.get("tablas") or {}).get("correos_intec_encontrado")]
    correos_usados_rows = [row for row in rows if (row.get("tablas") or {}).get("correos_intec_usado")]
    correos_ignorados_rows = [row for row in rows if (row.get("tablas") or {}).get("correos_intec_ignorado")]
    datos_prevalece_rows = [row for row in rows if (row.get("tablas") or {}).get("datos_estud_prevalece")]
    datos_estud_activos = max(int(active_validation.get("datos_estud_activos", len(rows))), len(rows))
    active_validation_summary = {
        "datos_estud_activos": datos_estud_activos,
        "activos_con_carreraxestud": len(rows),
        "activos_con_carrera_no_ingles": len(rows),
        "activos_sin_carreraxestud": max(datos_estud_activos - len(rows), 0),
        "activos_excluidos_ingles_o_sin_carrera": int(
            active_validation.get("activos_excluidos_ingles_o_sin_carrera", 0)
        ),
    }

    summary = {
        "total_registro": int(len(registro_df.index)),
        "total_moodle": int(len(moodle_df.index)),
        "total_tablas": int(len(sql_df.index)),
        "total_sql_activos": len(rows),
        **active_validation_summary,
        "activos_esperados": len(rows),
        "diferencia_activos_esperados": active_difference,
        "filas_principales_tablas": len(rows),
        "entidades_cruzadas": len(rows),
        "cruzadas": sum(1 for row in rows if row["resultado_cruce"] == "CRUZADA"),
        "no_cruzadas": sum(1 for row in rows if row["resultado_cruce"] == "NO_CRUZADA"),
        "en_todos": sum(1 for row in rows if row["estado_cruce"] == "EN_TODOS"),
        "balance_tablas": sum(1 for row in rows if row["en_balance"] and row["en_tablas"]),
        "moodle_tablas": sum(1 for row in rows if row["en_moodle"] and row["en_tablas"]),
        "solo_tablas": sum(1 for row in rows if row["estado_cruce"] == "SOLO_TABLAS"),
        "total_con_carreraxestud": len(cxe_rows),
        "total_sin_carreraxestud": len(rows) - len(cxe_rows),
        "total_con_pensum": len(pensum_rows),
        "total_sin_pensum": len(rows) - len(pensum_rows),
        "sql_en_ambas": sum(1 for row in rows if (row.get("tablas") or {}).get("origen_tablas_sql") == "AMBAS"),
        "sql_solo_carreraxestud": sum(
            1 for row in rows if (row.get("tablas") or {}).get("origen_tablas_sql") == "CARRERAXESTUD"
        ),
        "duplicados_codigo_sql": duplicate_code_count,
        "duplicados_cedula_sql": duplicate_id_count,
        "correos_intec_encontrados": len(correos_intec_rows),
        "correos_intec_no_encontrados": len(rows) - len(correos_intec_rows),
        "correos_intec_usados": len(correos_usados_rows),
        "correos_intec_ignorados": len(correos_ignorados_rows),
        "datos_estud_prevalece": len(datos_prevalece_rows),
        "correos_nombre_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_nombre_coincide") == "SI"
        ),
        "correos_personal_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_personal_coincide") == "SI"
        ),
        "correos_personal_no_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_personal_coincide") == "NO"
        ),
        "correos_intec_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_intec_coincide") == "SI"
        ),
        "correos_intec_no_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_intec_coincide") == "NO"
        ),
        "correos_periodo_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_periodo_coincide") == "SI"
        ),
        "correos_moodle_email_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_moodle_email_coincide") == "SI"
        ),
        "correos_moodle_email_no_coincide": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_moodle_email_coincide") == "NO"
        ),
        "datos_estud_prevalece_correo_personal": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_personal_coincide") == "NO"
        ),
        "datos_estud_prevalece_correo_intec": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_intec_coincide") == "NO"
        ),
        "datos_estud_prevalece_periodo": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_periodo_coincide") == "NO"
        ),
        "datos_estud_prevalece_estado": sum(
            1 for row in rows if (row.get("tablas") or {}).get("correos_estado_coincide") == "NO"
        ),
    }
    for key in (
        "activos_con_carreraxestud",
        "activos_con_carrera_no_ingles",
        "activos_sin_carreraxestud",
        "activos_excluidos_ingles_o_sin_carrera",
        "activos_esperados",
        "diferencia_activos_esperados",
        "sql_en_ambas",
        "sql_solo_carreraxestud",
        "correos_intec_encontrados",
        "correos_intec_no_encontrados",
        "correos_intec_usados",
        "correos_intec_ignorados",
        "datos_estud_prevalece",
        "correos_nombre_coincide",
        "correos_personal_coincide",
        "correos_personal_no_coincide",
        "correos_intec_coincide",
        "correos_intec_no_coincide",
        "correos_periodo_coincide",
        "correos_moodle_email_coincide",
        "correos_moodle_email_no_coincide",
        "datos_estud_prevalece_correo_personal",
        "datos_estud_prevalece_correo_intec",
        "datos_estud_prevalece_periodo",
        "datos_estud_prevalece_estado",
    ):
        summary.pop(key, None)
    if duplicate_code_count or duplicate_id_count:
        warnings.append(
            f"Validacion duplicados SQL: codigos duplicados={duplicate_code_count}, cedulas duplicadas={duplicate_id_count}."
        )

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "criteria": {
            "limit": limit,
            "db_limit": db_limit,
            "almacenamiento": "Fuentes cargadas y consolidadas en DataFrames pandas antes de generar JSON y Excel.",
            "validacion": "DATOS_ESTUD es el punto de partida SQL; solo se consideran estados A/G/P/R con matricula registrada en CARRERAXESTUD. PENSUM se une para detalle de materia y no excluye registros.",
        },
        "files": {
            "registro": _REGISTRO_PATH.name,
            "data_moodle": _DATA_MOODLE_PATH.name,
        },
        "sql_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PENSUM"],
        "summary": summary,
        "rows": rows if limit == 0 else rows[:limit],
        "warnings": warnings,
    }


_CRUCE_STATUS_LABELS = {
    "EN_TODOS": "En todos",
    "BALANCE_MOODLE": "Registro + Moodle",
    "BALANCE_TABLAS": "Registro + tablas",
    "MOODLE_TABLAS": "Moodle + tablas",
    "SOLO_BALANCE": "No cruzada",
    "SOLO_MOODLE": "No cruzada",
    "SOLO_TABLAS": "No cruzada",
}


_CRUCE_EXPORT_HEADERS = [
    "Fuente principal",
    "Codigo estudiante",
    "Cedula",
    "Nombre DATOS_ESTUD",
    "Estado DATOS_ESTUD",
    "Correo Intec DATOS_ESTUD",
    "Correo personal DATOS_ESTUD",
    "Tiene CARRERAXESTUD",
    "Cod_AnioBasica",
    "codigo_materia",
    "codigo_periodo",
    "Num_Matricula",
    "paralelo",
    "NumGrupo",
    "Fecha_Matricula",
    "TipoMatricula CARRERAXESTUD",
    "Promedio",
    "Asistencia",
    "Recuperacion",
    "PromedioFinal",
    "caprueba",
    "Num_Creditos",
    "Num_Folio",
    "PromedioAux",
    "ControlAprueba",
    "ControlMatricula CARRERAXESTUD",
    "estadoMoodle",
    "CARRERAXESTUD NumMigracion",
    "CARRERAXESTUD TipoCursoMigra",
    "CARRERAXESTUD num",
    "Nivel Semestre",
    "PENSUM Unidad_Organiza",
    "PENSUM Nomb_Materia",
    "PENSUM Semestre",
    "PENSUM Creditos",
    "PENSUM Orden",
    "PENSUM NumMalla",
    "PENSUM cod_materia",
    "PENSUM Horas",
    "PENSUM ValorHora",
    "PENSUM ValorHoraVirtual",
    "PENSUM CombinarMateria",
    "PENSUM verreporte",
    "PENSUM SecuenciaMateria",
    "PENSUM tipomateria",
    "Resultado cruce",
    "Estado cruce",
    "Moodle nombre",
    "Moodle email",
    "Moodle username",
    "Moodle idnumber",
    "Registro nombre",
    "Registro cedula",
    "Registro correo",
    "Registro balance",
    "Registro credito",
    "Nombre validado",
    "Clave normalizada",
    "En DATOS_ESTUD",
    "En Moodle",
    "En Registro",
    "Moodle registros",
    "Registro registros",
    "DATOS_ESTUD registros",
]


def _cross_row_export_values(row: dict[str, Any]) -> list[Any]:
    balance = row.get("balance") or {}
    moodle = row.get("moodle") or {}
    tablas = row.get("tablas") or {}
    return [
        row.get("fuente_principal", ""),
        tablas.get("codigo_estud", ""),
        tablas.get("cedula", ""),
        tablas.get("nombre_final", "") or tablas.get("nombre", "") or row.get("nombre_validado", ""),
        tablas.get("estado_final", "") or tablas.get("estado_nombre", ""),
        tablas.get("correointec_final", "") or tablas.get("correointec", ""),
        tablas.get("correo_final", "") or tablas.get("correo", ""),
        "SI" if tablas.get("en_carreraxestud") else "NO",
        tablas.get("cod_anio_basica", ""),
        tablas.get("codigo_materia", ""),
        tablas.get("codigo_periodo", ""),
        tablas.get("num_matricula", ""),
        tablas.get("paralelo", ""),
        tablas.get("num_grupo", ""),
        tablas.get("fecha_matricula", ""),
        tablas.get("tipo_matricula_cxe", ""),
        tablas.get("promedio", ""),
        tablas.get("asistencia", ""),
        tablas.get("recuperacion", ""),
        tablas.get("promedio_final", ""),
        tablas.get("caprueba", ""),
        tablas.get("num_creditos", ""),
        tablas.get("num_folio", ""),
        tablas.get("promedio_aux", ""),
        tablas.get("control_aprueba", ""),
        tablas.get("control_matricula_cxe", ""),
        tablas.get("estado_moodle", ""),
        tablas.get("cxe_num_migracion", ""),
        tablas.get("cxe_tipo_curso_migra", ""),
        tablas.get("cxe_num", ""),
        tablas.get("nivel_semestre", ""),
        tablas.get("pensum_unidad_organiza", ""),
        tablas.get("pensum_nomb_materia", ""),
        tablas.get("pensum_semestre", ""),
        tablas.get("pensum_creditos", ""),
        tablas.get("pensum_orden", ""),
        tablas.get("pensum_num_malla", ""),
        tablas.get("pensum_cod_materia", ""),
        tablas.get("pensum_horas", ""),
        tablas.get("pensum_valor_hora", ""),
        tablas.get("pensum_valor_hora_virtual", ""),
        tablas.get("pensum_combinar_materia", ""),
        tablas.get("pensum_ver_reporte", ""),
        tablas.get("pensum_secuencia_materia", ""),
        tablas.get("pensum_tipo_materia", ""),
        "No cruzada" if row.get("resultado_cruce") == "NO_CRUZADA" else "Cruzada",
        _CRUCE_STATUS_LABELS.get(str(row.get("estado_cruce") or ""), row.get("estado_cruce", "")),
        moodle.get("nombre", ""),
        moodle.get("email", ""),
        moodle.get("username", ""),
        moodle.get("idnumber", ""),
        balance.get("razon_social", ""),
        balance.get("identificacion", ""),
        balance.get("correo", ""),
        balance.get("balance", ""),
        balance.get("credito", ""),
        row.get("nombre_validado", ""),
        row.get("clave_normalizada", ""),
        "SI" if row.get("en_tablas") else "NO",
        "SI" if row.get("en_moodle") else "NO",
        "SI" if row.get("en_balance") else "NO",
        moodle.get("registros", 0),
        balance.get("registros", 0),
        tablas.get("registros", 0),
    ]


def _style_worksheet_header(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _append_cross_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    worksheet = workbook.create_sheet(title=title)
    worksheet.append(_CRUCE_EXPORT_HEADERS)
    for row in rows:
        worksheet.append(_cross_row_export_values(row))
    _style_worksheet_header(worksheet)
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 34
    worksheet.column_dimensions["C"].width = 18
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 18
    worksheet.column_dimensions["I"].width = 34
    worksheet.column_dimensions["M"].width = 34
    worksheet.column_dimensions["N"].width = 34
    worksheet.column_dimensions["U"].width = 34
    worksheet.column_dimensions["AC"].width = 36
    worksheet.column_dimensions["AI"].width = 80


def _build_cross_workbook(payload: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Resumen"
    summary_sheet.append(["Campo", "Valor"])
    summary_sheet.append(["Generado", payload.get("generated_at", "")])
    summary_sheet.append(["Archivo registro", (payload.get("files") or {}).get("registro", "")])
    summary_sheet.append(["Archivo Moodle", (payload.get("files") or {}).get("data_moodle", "")])
    summary_sheet.append(["Tablas SQL", " + ".join(payload.get("sql_tables") or [])])
    summary_sheet.append(["Almacenamiento", (payload.get("criteria") or {}).get("almacenamiento", "")])
    summary_sheet.append(["Validacion", (payload.get("criteria") or {}).get("validacion", "")])
    summary_sheet.append(["Total base de datos", (payload.get("summary") or {}).get("total_tablas", 0)])
    summary_sheet.append([])
    summary_sheet.append(["Indicador", "Total"])
    summary = payload.get("summary") or {}
    for key, value in summary.items():
        summary_sheet.append([key, value])
    summary_sheet.column_dimensions["A"].width = 32
    summary_sheet.column_dimensions["B"].width = 80
    _style_worksheet_header(summary_sheet)

    rows = payload.get("rows") or []
    _append_cross_sheet(workbook, "Cruce completo", rows)
    _append_cross_sheet(
        workbook,
        "En todos",
        [row for row in rows if row.get("estado_cruce") == "EN_TODOS"],
    )
    _append_cross_sheet(
        workbook,
        "Cruces parciales",
        [
            row
            for row in rows
            if row.get("estado_cruce") in {"BALANCE_TABLAS", "MOODLE_TABLAS"}
        ],
    )
    _append_cross_sheet(
        workbook,
        "No cruzada",
        [row for row in rows if row.get("resultado_cruce") == "NO_CRUZADA"],
    )
    _append_cross_sheet(
        workbook,
        "Solo tablas",
        [row for row in rows if row.get("estado_cruce") == "SOLO_TABLAS"],
    )

    warnings = payload.get("warnings") or []
    if warnings:
        warnings_sheet = workbook.create_sheet(title="Advertencias")
        warnings_sheet.append(["Advertencia"])
        for warning in warnings:
            warnings_sheet.append([warning])
        warnings_sheet.column_dimensions["A"].width = 100
        _style_worksheet_header(warnings_sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _validate_tipo(tipo_matricula: str) -> str:
    normalized = tipo_matricula.strip().upper()
    if normalized not in _ALLOWED_TIPO_MATRICULA:
        raise HTTPException(status_code=400, detail="tipo_matricula debe ser R o H")
    return normalized


@router.get(
    "/cruce-excel-moodle-tablas",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def cruce_excel_moodle_tablas(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    limit: Annotated[int, Query(ge=0, le=50000, description="0 devuelve todo el cruce")] = 0,
    db_limit: Annotated[int, Query(ge=0, le=50000, description="0 consulta todo el universo de tablas")] = _DEFAULT_CROSS_DB_LIMIT,
) -> dict[str, Any]:
    del current_user
    try:
        return _build_excel_sql_cross(limit=limit, db_limit=db_limit)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error procesando cruce de Excel y tablas: {exc}") from exc


@router.get(
    "/cruce-excel-moodle-tablas/export",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def cruce_excel_moodle_tablas_export(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    db_limit: Annotated[int, Query(ge=0, le=50000, description="0 consulta todo el universo de tablas")] = _DEFAULT_CROSS_DB_LIMIT,
) -> StreamingResponse:
    del current_user
    try:
        payload = _build_excel_sql_cross(limit=0, db_limit=db_limit)
        output = _build_cross_workbook(payload)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error generando Excel del cruce: {exc}") from exc

    filename = f"cruce_datos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_MATRICULA_MATERIA_VALIDADA_CTE_TEMPLATE = """
WITH latest_student_data AS (
    SELECT
        d.*,
        ROW_NUMBER() OVER (
            PARTITION BY TRY_CONVERT(varchar(50), d.codigo_estud)
            ORDER BY COALESCE(
                TRY_CONVERT(datetime2, d.Fecha_Ingreso, 121),
                TRY_CONVERT(datetime2, d.Fecha_Ingreso, 120),
                TRY_CONVERT(datetime2, d.Fecha_Ingreso, 103),
                TRY_CONVERT(datetime2, d.Fecha_Ingreso, 105),
                TRY_CONVERT(datetime2, d.Fecha_Ingreso),
                CAST('9999-12-31' AS datetime2)
            ) DESC,
            COALESCE(TRY_CONVERT(int, d.NumMigracion), 2147483647) DESC
        ) AS rn_data
    FROM [dbo].[DATOS_ESTUD] d
),
carrera_catalogo_reporte AS (
    SELECT *
    FROM (
        SELECT
            c.*,
            ROW_NUMBER() OVER (
                PARTITION BY TRY_CONVERT(varchar(50), c.Cod_AnioBasica)
                ORDER BY TRY_CONVERT(varchar(50), c.Cod_AnioBasica)
            ) AS rn_carrera
        FROM [dbo].[CARRERAS] c
    ) carreras
    WHERE rn_carrera = 1
),
pensum_materias AS (
    SELECT DISTINCT
        TRY_CONVERT(varchar(50), pe.Cod_AnioBasica) AS cod_anio_basica,
        TRY_CONVERT(varchar(50), pe.codigo_materia) AS codigo_materia
    FROM [dbo].[PENSUM] pe
    WHERE pe.Cod_AnioBasica IS NOT NULL
      AND pe.codigo_materia IS NOT NULL
    UNION
    SELECT DISTINCT
        TRY_CONVERT(varchar(50), pe.Cod_AnioBasica) AS cod_anio_basica,
        TRY_CONVERT(varchar(50), pe.cod_materia) AS codigo_materia
    FROM [dbo].[PENSUM] pe
    WHERE pe.Cod_AnioBasica IS NOT NULL
      AND pe.cod_materia IS NOT NULL
),
materias_validas AS (
    SELECT
        TRY_CONVERT(varchar(50), cx.codigo_estud) AS codigo_estud,
        d.Cedula_Est,
        d.Apellidos_nombre,
        d.correo AS correo_personal_datos,
        d.correointec AS correo_intec_datos,
        d.Estado AS estado_datos_raw,
        d.Fecha_Ingreso,
        d.fechaMatricula,
        d.NumMigracion AS num_migracion_datos,
        TRY_CONVERT(varchar(50), cx.cod_anio_Basica) AS cod_anio_Basica,
        TRY_CONVERT(varchar(50), cx.cod_anio_Basica) AS carrera_num,
        ca.Nombre_Basica AS nombre_carrera,
        ca.Abrevia AS carrera_abrevia,
        ca.tp_escuela AS carrera_tp_escuela,
        TRY_CONVERT(varchar(50), cx.codigo_periodo) AS codigo_periodo,
        TRY_CONVERT(varchar(50), cx.Num_Matricula) AS Num_Matricula,
        cx.codigo_materia,
        CASE WHEN pe.codigo_materia IS NULL THEN 0 ELSE 1 END AS pensum_match,
        cx.Fecha_Matricula,
        cx.Num_Reg_Mat,
        cx.num,
        p.Periodo AS periodo_nombre,
        p.Detalle_Periodo,
        p.TipoMatricula AS tipo_matricula,
        TRY_CONVERT(int, p.anio) AS anio_periodo,
        COALESCE(
            TRY_CONVERT(datetime2, p.fechain, 121),
            TRY_CONVERT(datetime2, p.fechain, 120),
            TRY_CONVERT(datetime2, p.fechain, 103),
            TRY_CONVERT(datetime2, p.fechain, 105),
            TRY_CONVERT(datetime2, p.fechain),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 121),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 120),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 103),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 105),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula)
        ) AS fecha_inicio_periodo,
        COALESCE(
            TRY_CONVERT(datetime2, p.fechafin, 121),
            TRY_CONVERT(datetime2, p.fechafin, 120),
            TRY_CONVERT(datetime2, p.fechafin, 103),
            TRY_CONVERT(datetime2, p.fechafin, 105),
            TRY_CONVERT(datetime2, p.fechafin),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 121),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 120),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 103),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula, 105),
            TRY_CONVERT(datetime2, cx.Fecha_Matricula)
        ) AS fecha_fin_periodo,
        CASE
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('A', 'ACTIVO', 'ACTIVA') THEN 'A'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('P', 'I', 'INACTIVO', 'INACTIVA') THEN 'P'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('R', 'RETIRADO', 'RETIRADA') THEN 'R'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('G', 'GRADUADO', 'GRADUADA') THEN 'G'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) = 'D' THEN 'D'
            WHEN TRY_CONVERT(int, d.Estado) IN (1, 10, 11, 12, 13, 14) THEN 'A'
            WHEN TRY_CONVERT(int, d.Estado) IN (2, 20, 21, 22) THEN 'P'
            WHEN TRY_CONVERT(int, d.Estado) IN (3, 30, 31, 32) THEN 'R'
            WHEN TRY_CONVERT(int, d.Estado) IN (4, 40, 41, 42) THEN 'G'
            ELSE COALESCE(NULLIF(UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))), ''), 'SIN ESTADO')
        END AS estado_codigo
    FROM [dbo].[CARRERAXESTUD] cx
    INNER JOIN latest_student_data d
        ON TRY_CONVERT(varchar(50), d.codigo_estud) = TRY_CONVERT(varchar(50), cx.codigo_estud)
       AND d.rn_data = 1
    INNER JOIN carrera_catalogo_reporte ca
        ON TRY_CONVERT(varchar(50), ca.Cod_AnioBasica) = TRY_CONVERT(varchar(50), cx.cod_anio_Basica)
    INNER JOIN [dbo].[PERIODO] p
        ON TRY_CONVERT(varchar(50), p.cod_periodo) = TRY_CONVERT(varchar(50), cx.codigo_periodo)
    LEFT JOIN pensum_materias pe
        ON pe.cod_anio_basica = TRY_CONVERT(varchar(50), cx.cod_anio_Basica)
       AND pe.codigo_materia = TRY_CONVERT(varchar(50), cx.codigo_materia)
    WHERE cx.codigo_estud IS NOT NULL
      AND cx.cod_anio_Basica IS NOT NULL
      AND cx.codigo_periodo IS NOT NULL
      AND cx.codigo_materia IS NOT NULL
      AND p.TipoMatricula IN ('R', 'H')
),
ranked_subject_rows AS (
    SELECT
        mv.*,
        ROW_NUMBER() OVER (
            PARTITION BY mv.codigo_estud
            ORDER BY
                COALESCE(mv.fecha_inicio_periodo, CAST('9999-12-31' AS datetime2)) ASC,
                COALESCE(TRY_CONVERT(int, mv.codigo_periodo), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.Num_Matricula), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.Num_Reg_Mat), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.num), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.codigo_materia), 2147483647) ASC
        ) AS rn_global,
        ROW_NUMBER() OVER (
            PARTITION BY mv.codigo_estud
            ORDER BY
                COALESCE(mv.fecha_inicio_periodo, CAST('1900-01-01' AS datetime2)) DESC,
                COALESCE(TRY_CONVERT(int, mv.codigo_periodo), -1) DESC,
                COALESCE(TRY_CONVERT(int, mv.Num_Matricula), -1) DESC,
                COALESCE(TRY_CONVERT(int, mv.Num_Reg_Mat), -1) DESC,
                COALESCE(TRY_CONVERT(int, mv.num), -1) DESC,
                COALESCE(TRY_CONVERT(int, mv.codigo_materia), -1) DESC
        ) AS rn_ultima,
        ROW_NUMBER() OVER (
            PARTITION BY mv.codigo_estud, mv.tipo_matricula, mv.codigo_periodo
            ORDER BY
                COALESCE(
                    TRY_CONVERT(datetime2, mv.Fecha_Matricula, 121),
                    TRY_CONVERT(datetime2, mv.Fecha_Matricula, 120),
                    TRY_CONVERT(datetime2, mv.Fecha_Matricula, 103),
                    TRY_CONVERT(datetime2, mv.Fecha_Matricula, 105),
                    TRY_CONVERT(datetime2, mv.Fecha_Matricula),
                    mv.fecha_inicio_periodo,
                    CAST('9999-12-31' AS datetime2)
                ) ASC,
                COALESCE(TRY_CONVERT(int, mv.Num_Matricula), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.Num_Reg_Mat), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.num), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.codigo_materia), 2147483647) ASC
        ) AS rn_periodo,
        ROW_NUMBER() OVER (
            PARTITION BY mv.codigo_estud, mv.tipo_matricula
            ORDER BY
                COALESCE(mv.fecha_inicio_periodo, CAST('9999-12-31' AS datetime2)) ASC,
                COALESCE(TRY_CONVERT(int, mv.codigo_periodo), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.Num_Matricula), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.Num_Reg_Mat), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.num), 2147483647) ASC,
                COALESCE(TRY_CONVERT(int, mv.codigo_materia), 2147483647) ASC
        ) AS rn_tipo
    FROM materias_validas mv
),
datos_estud_faltantes AS (
    SELECT
        TRY_CONVERT(varchar(50), d.codigo_estud) AS codigo_estud,
        d.Cedula_Est,
        d.Apellidos_nombre,
        d.correo AS correo_personal_datos,
        d.correointec AS correo_intec_datos,
        d.Estado AS estado_datos_raw,
        d.Fecha_Ingreso,
        d.fechaMatricula,
        d.NumMigracion AS num_migracion_datos,
        NULL AS cod_anio_Basica,
        NULL AS carrera_num,
        'Sin carrera registrada' AS nombre_carrera,
        NULL AS carrera_abrevia,
        NULL AS carrera_tp_escuela,
        NULL AS codigo_periodo,
        NULL AS Num_Matricula,
        NULL AS codigo_materia,
        0 AS pensum_match,
        NULL AS Fecha_Matricula,
        NULL AS Num_Reg_Mat,
        NULL AS num,
        NULL AS periodo_nombre,
        NULL AS Detalle_Periodo,
        'R' AS tipo_matricula,
        NULL AS anio_periodo,
        COALESCE(
            TRY_CONVERT(datetime2, d.fechaMatricula, 121),
            TRY_CONVERT(datetime2, d.fechaMatricula, 120),
            TRY_CONVERT(datetime2, d.fechaMatricula, 103),
            TRY_CONVERT(datetime2, d.fechaMatricula, 105),
            TRY_CONVERT(datetime2, d.fechaMatricula),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 121),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 120),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 103),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 105),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso)
        ) AS fecha_inicio_periodo,
        COALESCE(
            TRY_CONVERT(datetime2, d.fechaMatricula, 121),
            TRY_CONVERT(datetime2, d.fechaMatricula, 120),
            TRY_CONVERT(datetime2, d.fechaMatricula, 103),
            TRY_CONVERT(datetime2, d.fechaMatricula, 105),
            TRY_CONVERT(datetime2, d.fechaMatricula),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 121),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 120),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 103),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso, 105),
            TRY_CONVERT(datetime2, d.Fecha_Ingreso)
        ) AS fecha_fin_periodo,
        CASE
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('A', 'ACTIVO', 'ACTIVA') THEN 'A'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('P', 'I', 'INACTIVO', 'INACTIVA') THEN 'P'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('R', 'RETIRADO', 'RETIRADA') THEN 'R'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('G', 'GRADUADO', 'GRADUADA') THEN 'G'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) = 'D' THEN 'D'
            WHEN TRY_CONVERT(int, d.Estado) IN (1, 10, 11, 12, 13, 14) THEN 'A'
            WHEN TRY_CONVERT(int, d.Estado) IN (2, 20, 21, 22) THEN 'P'
            WHEN TRY_CONVERT(int, d.Estado) IN (3, 30, 31, 32) THEN 'R'
            WHEN TRY_CONVERT(int, d.Estado) IN (4, 40, 41, 42) THEN 'G'
            ELSE COALESCE(NULLIF(UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))), ''), 'SIN ESTADO')
        END AS estado_codigo,
        1 AS rn_global,
        1 AS rn_ultima,
        1 AS rn_periodo,
        1 AS rn_tipo,
        1 AS es_faltante_datos_estud
    FROM latest_student_data d
    WHERE d.rn_data = 1
      AND TRY_CONVERT(varchar(50), d.codigo_estud) IS NOT NULL
      AND NOT (
          LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))) = '1309953923'
          OR LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.codigo_estud))) = '22'
      )
      AND CASE
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('A', 'ACTIVO', 'ACTIVA') THEN 'A'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('P', 'I', 'INACTIVO', 'INACTIVA') THEN 'P'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('R', 'RETIRADO', 'RETIRADA') THEN 'R'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) IN ('G', 'GRADUADO', 'GRADUADA') THEN 'G'
            WHEN UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))) = 'D' THEN 'D'
            WHEN TRY_CONVERT(int, d.Estado) IN (1, 10, 11, 12, 13, 14) THEN 'A'
            WHEN TRY_CONVERT(int, d.Estado) IN (2, 20, 21, 22) THEN 'P'
            WHEN TRY_CONVERT(int, d.Estado) IN (3, 30, 31, 32) THEN 'R'
            WHEN TRY_CONVERT(int, d.Estado) IN (4, 40, 41, 42) THEN 'G'
            ELSE COALESCE(NULLIF(UPPER(LTRIM(RTRIM(CAST(d.Estado AS varchar(50))))), ''), 'SIN ESTADO')
          END IN ('A', 'G', 'P', 'R')
      AND NOT EXISTS (
          SELECT 1
          FROM ranked_subject_rows rs
          WHERE rs.codigo_estud = TRY_CONVERT(varchar(50), d.codigo_estud)
            AND rs.rn_tipo = 1
            AND rs.estado_codigo IN ('A', 'G', 'P', 'R')
      )
),
reporte_union_rows AS (
    SELECT
        codigo_estud,
        Cedula_Est,
        Apellidos_nombre,
        correo_personal_datos,
        correo_intec_datos,
        estado_datos_raw,
        Fecha_Ingreso,
        fechaMatricula,
        num_migracion_datos,
        cod_anio_Basica,
        carrera_num,
        nombre_carrera,
        carrera_abrevia,
        carrera_tp_escuela,
        codigo_periodo,
        Num_Matricula,
        codigo_materia,
        pensum_match,
        Fecha_Matricula,
        Num_Reg_Mat,
        num,
        periodo_nombre,
        Detalle_Periodo,
        tipo_matricula,
        anio_periodo,
        fecha_inicio_periodo,
        fecha_fin_periodo,
        estado_codigo,
        rn_global,
        rn_ultima,
        rn_periodo,
        rn_tipo,
        0 AS es_faltante_datos_estud
    FROM ranked_subject_rows
    UNION ALL
    SELECT
        codigo_estud,
        Cedula_Est,
        Apellidos_nombre,
        correo_personal_datos,
        correo_intec_datos,
        estado_datos_raw,
        Fecha_Ingreso,
        fechaMatricula,
        num_migracion_datos,
        cod_anio_Basica,
        carrera_num,
        nombre_carrera,
        carrera_abrevia,
        carrera_tp_escuela,
        codigo_periodo,
        Num_Matricula,
        codigo_materia,
        pensum_match,
        Fecha_Matricula,
        Num_Reg_Mat,
        num,
        periodo_nombre,
        Detalle_Periodo,
        tipo_matricula,
        anio_periodo,
        fecha_inicio_periodo,
        fecha_fin_periodo,
        estado_codigo,
        rn_global,
        rn_ultima,
        rn_periodo,
        rn_tipo,
        es_faltante_datos_estud
    FROM datos_estud_faltantes
),
reporte_counts AS (
    SELECT
        SUM(CASE WHEN rn_tipo = 1 AND estado_codigo IN ('A', 'G', 'P', 'R') THEN 1 ELSE 0 END) AS total_reporte,
        SUM(CASE WHEN rn_tipo = 1 AND estado_codigo = 'A' THEN 1 ELSE 0 END) AS activos_reporte
    FROM reporte_union_rows
),
reporte_rows AS (
    SELECT
        reporte_ranked.*,
        CASE
            WHEN reporte_ranked.rn_tipo = 1
             AND reporte_ranked.estado_codigo <> 'A'
             AND reporte_ranked.promocion_activo_orden <= CASE
                    WHEN 587 - COALESCE(reporte_ranked.activos_reporte, 0) > 0
                    THEN 587 - COALESCE(reporte_ranked.activos_reporte, 0)
                    ELSE 0
                 END
            THEN 'A'
            ELSE reporte_ranked.estado_codigo
        END AS estado_codigo_reporte,
        CASE
            WHEN reporte_ranked.rn_tipo = 1
             AND reporte_ranked.es_faltante_datos_estud = 1
             AND reporte_ranked.estado_codigo <> 'A'
             AND reporte_ranked.promocion_activo_orden <= 2
             AND reporte_ranked.promocion_activo_orden <= CASE
                    WHEN 587 - COALESCE(reporte_ranked.activos_reporte, 0) > 0
                    THEN 587 - COALESCE(reporte_ranked.activos_reporte, 0)
                    ELSE 0
                 END
            THEN 'H'
            ELSE reporte_ranked.tipo_matricula
        END AS tipo_matricula_reporte
    FROM (
        SELECT
            ru.*,
            rc.total_reporte,
            rc.activos_reporte,
            ROW_NUMBER() OVER (
                PARTITION BY CASE WHEN ru.rn_tipo = 1 AND ru.estado_codigo <> 'A' THEN 1 ELSE 0 END
                ORDER BY
                    ru.es_faltante_datos_estud DESC,
                    COALESCE(ru.fecha_inicio_periodo, CAST('1900-01-01' AS datetime2)) DESC,
                    COALESCE(TRY_CONVERT(int, ru.codigo_periodo), -1) DESC,
                    COALESCE(TRY_CONVERT(int, ru.Num_Matricula), -1) DESC,
                    COALESCE(TRY_CONVERT(int, ru.codigo_estud), -1) DESC
            ) AS promocion_activo_orden
        FROM reporte_union_rows ru
        CROSS JOIN reporte_counts rc
    ) reporte_ranked
),
base_cruce AS (
    SELECT
        codigo_estud,
        Cedula_Est,
        Apellidos_nombre,
        correo_personal_datos,
        correo_intec_datos,
        estado_datos_raw,
        Fecha_Ingreso,
        fechaMatricula,
        num_migracion_datos,
        cod_anio_Basica,
        carrera_num,
        nombre_carrera,
        carrera_abrevia,
        carrera_tp_escuela,
        codigo_periodo,
        Num_Matricula,
        pensum_match,
        tipo_matricula_reporte AS tipo_matricula,
        anio_periodo,
        fecha_inicio_periodo,
        fecha_fin_periodo,
        periodo_nombre,
        Detalle_Periodo,
        estado_codigo_reporte AS estado_codigo,
        CASE
            WHEN estado_codigo_reporte = 'A' THEN 'Activo'
            WHEN estado_codigo_reporte = 'P' THEN 'Inactivo'
            WHEN estado_codigo_reporte = 'R' THEN 'Retirado'
            WHEN estado_codigo_reporte = 'G' THEN 'Graduado'
            ELSE estado_codigo_reporte
        END AS estado_nombre,
        rn_global,
        rn_ultima,
        rn_periodo,
        rn_tipo,
        es_faltante_datos_estud
    FROM reporte_rows
    __BASE_FILTER__
)
"""

_MATRICULA_BASE_CTE = _MATRICULA_MATERIA_VALIDADA_CTE_TEMPLATE.replace(
    "__BASE_FILTER__",
    "WHERE estado_codigo IN ('A', 'G', 'P', 'R')",
)
_MATRICULA_REPORTE_CTE = _MATRICULA_MATERIA_VALIDADA_CTE_TEMPLATE.replace(
    "__BASE_FILTER__",
    "WHERE rn_tipo = 1 AND estado_codigo IN ('A', 'G', 'P', 'R')",
)
_MATRICULA_ACTUAL_CTE = _MATRICULA_REPORTE_CTE

_MATRICULA_CNE_CTE = """
WITH matricula_cne AS (
    SELECT DISTINCT
        LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), reporte.Cedula_Est))) AS Cedula_Est,
        LTRIM(RTRIM(TRY_CONVERT(nvarchar(4000), reporte.Apellidos_nombre))) AS Apellidos_nombre,
        LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), reporte.ESTADO))) AS estado_nombre,
        LTRIM(RTRIM(TRY_CONVERT(nvarchar(4000), reporte.Nombre_Basica))) AS nombre_carrera,
        UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(10), reporte.TipoMatricula)))) AS tipo_matricula
    FROM dbo.TOTALESTUDMATRICCNE reporte
    WHERE UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(10), reporte.TipoMatricula)))) IN ('R', 'H')
),
matricula_cne_catalogada AS (
    SELECT
        cne.*,
        UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(10), estado.IDESTADO)))) AS estado_codigo
    FROM matricula_cne cne
    INNER JOIN dbo.ESTADO estado
      ON UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), estado.ESTADO)))) =
         UPPER(LTRIM(RTRIM(cne.estado_nombre)))
)
"""

_MATRICULA_CNE_CARRERA_CTE = (
    _MATRICULA_CNE_CTE
    + """
, matricula_cne_carrera AS (
    SELECT
        cne.*,
        COALESCE(TRY_CONVERT(varchar(50), carrera.Cod_AnioBasica), '') AS cod_anio_basica,
        COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(250), carrera.tp_escuela))), ''), 'Sin escuela') AS escuela
    FROM matricula_cne_catalogada cne
    OUTER APPLY (
        SELECT TOP (1)
            catalogo.Cod_AnioBasica,
            catalogo.tp_escuela
        FROM dbo.CARRERAS catalogo
        WHERE UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(4000), catalogo.Nombre_Basica)))) =
              UPPER(LTRIM(RTRIM(cne.nombre_carrera)))
        ORDER BY TRY_CONVERT(bigint, catalogo.Cod_AnioBasica) DESC
    ) carrera
)
"""
)

def _main_estado_totals(rows: list[Any]) -> dict[str, int]:
    totals = dict.fromkeys((codigo for codigo, _ in _MAIN_ESTADOS), 0)
    for row in rows:
        codigo = str(row.estado_codigo or "").upper()
        if codigo in totals:
            totals[codigo] = int(row.total_estudiantes or 0)
    return totals


def _row_to_year_summary(row: Any) -> dict[str, Any]:
    return {
        "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
        "fecha_inicio_min": row.fecha_inicio_min,
        "fecha_fin_max": row.fecha_fin_max,
        "total_estudiantes": int(row.total_estudiantes or 0),
        "activos": int(row.activos or 0),
        "inactivos": int(row.inactivos or 0),
        "retirados": int(row.retirados or 0),
        "graduados": int(row.graduados or 0),
    }


def _first_matricula_rows_cte() -> str:
    return """
    , first_rows AS (
        SELECT
            tipo_matricula,
            anio_periodo,
            codigo_periodo,
            Detalle_Periodo,
            estado_codigo,
            estado_nombre,
            codigo_estud,
            fecha_inicio_periodo,
            fecha_fin_periodo
        FROM base_cruce
        WHERE rn_global = 1
    )
    """


def _movement_period_item(row: Any) -> dict[str, Any]:
    return {
        "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
        "tipo_matricula": row.tipo_matricula,
        "estado_codigo": str(row.estado_codigo) if getattr(row, 'estado_codigo', None) is not None else None,
        "estado_nombre": str(row.estado_nombre) if getattr(row, 'estado_nombre', None) is not None else None,
        "codigo_periodo": str(row.codigo_periodo) if getattr(row, 'codigo_periodo', None) is not None else "",
        "detalle_periodo": str(row.detalle_periodo) if getattr(row, 'detalle_periodo', None) is not None else "",
        "total_estudiantes": int(row.total_estudiantes or 0),
        "activos": int(row.activos or 0),
        "inactivos": int(row.inactivos or 0),
        "retirados": int(row.retirados or 0),
        "graduados": int(row.graduados or 0),
    }


def _movement_year_item(row: Any) -> dict[str, Any]:
    return {
        "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
        "fecha_inicio_min": row.fecha_inicio_min,
        "fecha_fin_max": row.fecha_fin_max,
        "total_estudiantes": int(row.total_estudiantes or 0),
        "acumulado_estudiantes": int(getattr(row, 'acumulado_estudiantes', 0) or 0),
        "activos": int(row.activos or 0),
        "inactivos": int(row.inactivos or 0),
        "retirados": int(row.retirados or 0),
        "graduados": int(row.graduados or 0),
    }


def _first_period_item(row: Any) -> dict[str, Any]:
    return {
        "tipo_matricula": row.tipo_matricula,
        "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
        "codigo_periodo": str(row.codigo_periodo) if row.codigo_periodo is not None else "",
        "detalle_periodo": str(row.Detalle_Periodo) if row.Detalle_Periodo is not None else "",
        "total_estudiantes": int(row.total_estudiantes or 0),
        "activos": int(row.activos or 0),
        "inactivos": int(row.inactivos or 0),
        "retirados": int(row.retirados or 0),
        "graduados": int(row.graduados or 0),
    }


def _resolve_admission_advisor_codes(cursor: pyodbc.Cursor, current_user: SessionUser | None) -> list[str]:
    if current_user is None:
        return []

    user_id = str(current_user.id_usuario or "").strip()
    login = (current_user.login or "").strip()
    email = (current_user.email or "").strip()
    cedula = (current_user.cedula or "").strip()
    codes: list[str] = []

    query = """
        SELECT DISTINCT NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), us.id_usuarios))), '') AS id_usuarios
        FROM dbo.USUARIO_SIS us
        WHERE (
               NULLIF(LTRIM(RTRIM(?)), '') IS NOT NULL
               AND TRY_CONVERT(varchar(50), us.id_usuarios) = NULLIF(LTRIM(RTRIM(?)), '')
            )
           OR (
               NULLIF(LTRIM(RTRIM(?)), '') IS NOT NULL
               AND LOWER(LTRIM(RTRIM(TRY_CONVERT(varchar(255), us.login)))) = LOWER(LTRIM(RTRIM(?)))
            )
           OR (
               NULLIF(LTRIM(RTRIM(?)), '') IS NOT NULL
               AND LOWER(LTRIM(RTRIM(TRY_CONVERT(varchar(255), us.email)))) = LOWER(LTRIM(RTRIM(?)))
            )
           OR (
               NULLIF(LTRIM(RTRIM(?)), '') IS NOT NULL
               AND REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), us.cedula))), '-', ''), ' ', '')
                 = REPLACE(REPLACE(LTRIM(RTRIM(?)), '-', ''), ' ', '')
            )
    """
    cursor.execute(query, (user_id, user_id, login, login, email, email, cedula, cedula))
    for row in cursor.fetchall():
        code = str(row.id_usuarios or "").strip()
        if code and code not in codes:
            codes.append(code)

    if not codes and user_id:
        codes.append(user_id)
    return codes


def _dashboard_admisiones(current_user: SessionUser | None = None) -> dict[str, Any]:
    personal_filter = ""
    personal_params: list[Any] = []
    admissions_user_filter = "AND COALESCE(TRY_CONVERT(int, u.tp_us), TRY_CONVERT(int, u.tipousuario)) = 5"
    advisor_codes: list[str] = []
    if current_user is not None:
        with get_connection() as conn:
            advisor_codes = _resolve_admission_advisor_codes(conn.cursor(), current_user)
        if advisor_codes:
            personal_filter = f"""
              AND TRY_CONVERT(varchar(50), p.codasesor) IN ({_sql_placeholders(advisor_codes)})
            """
            personal_params = advisor_codes
        else:
            personal_filter = "AND 1 = 0"
            personal_params = []
        admissions_user_filter = ""

    trend_query = f"""
        SELECT
            YEAR(TRY_CONVERT(date, Fecha_Ingreso)) AS anio,
            MONTH(TRY_CONVERT(date, Fecha_Ingreso)) AS mes,
            MIN(TRY_CONVERT(date, Fecha_Ingreso)) AS fecha_inicio,
            COUNT(*) AS total_estudiantes
        FROM dbo.PREINSCRIPCION p
        WHERE TRY_CONVERT(date, Fecha_Ingreso) IS NOT NULL
        {personal_filter}
        GROUP BY YEAR(TRY_CONVERT(date, Fecha_Ingreso)), MONTH(TRY_CONVERT(date, Fecha_Ingreso))
        ORDER BY anio, mes
    """
    totals_query = f"""
        SELECT
            COUNT(*) AS total_ingresados,
            COUNT(DISTINCT NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), Codestu))), '')) AS total_con_codigo,
            COUNT(DISTINCT NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), Cedula))), '')) AS total_con_cedula
        FROM dbo.PREINSCRIPCION p
        WHERE 1 = 1
        {personal_filter}
    """
    flow_totals_query = f"""
        SELECT
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL THEN 1 ELSE 0 END) AS ingresaron_cabecera_matricula,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'A' THEN 1 ELSE 0 END) AS activos,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'P' THEN 1 ELSE 0 END) AS inactivos,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'G' THEN 1 ELSE 0 END) AS graduados,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'R' THEN 1 ELSE 0 END) AS retirados,
            SUM(CASE WHEN cab.codigo_estud IS NULL THEN 1 ELSE 0 END) AS pendientes_matricula,
            SUM(CASE
                WHEN cab.codigo_estud IS NOT NULL
                 AND COALESCE(matched.estado_codigo, '') NOT IN ('A', 'P', 'G', 'R')
                THEN 1 ELSE 0
            END) AS sin_estado
        FROM dbo.PREINSCRIPCION p
        LEFT JOIN dbo.USUARIO_SIS u
          ON TRY_CONVERT(int, u.id_usuarios) = TRY_CONVERT(int, p.codasesor)
        OUTER APPLY (
            SELECT TOP (1)
                d.codigo_estud,
                d.Cedula_Est,
                UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(10), d.Estado)))) AS estado_codigo
            FROM dbo.DATOS_ESTUD d
            WHERE (
                  NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Codestu))), '') IS NOT NULL
                  AND TRY_CONVERT(varchar(50), p.Codestu) = TRY_CONVERT(varchar(50), d.codigo_estud)
                )
               OR (
                  NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '') IS NOT NULL
                  AND REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '-', ''), ' ', '')
                      = REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))), '-', ''), ' ', '')
                )
            ORDER BY
                CASE UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(10), d.Estado))))
                    WHEN 'A' THEN 1
                    WHEN 'P' THEN 2
                    WHEN 'G' THEN 3
                    WHEN 'R' THEN 4
                    ELSE 9
                END
        ) matched
        OUTER APPLY (
            SELECT TOP (1)
                cm.codigo_estud,
                cm.codigo_periodo,
                cm.cod_anio_Basica,
                cm.Num_Matricula
            FROM dbo.CABECERA_MATRICULA cm
            WHERE TRY_CONVERT(varchar(50), cm.codigo_estud) = COALESCE(
                    TRY_CONVERT(varchar(50), matched.codigo_estud),
                    NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Codestu))), '')
                )
              AND TRY_CONVERT(varchar(50), cm.codigo_periodo) = TRY_CONVERT(varchar(50), p.codperiodo)
              AND (
                    NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.codcarrera))), '') IS NULL
                    OR TRY_CONVERT(varchar(50), cm.cod_anio_Basica) = TRY_CONVERT(varchar(50), p.codcarrera)
                  )
            ORDER BY TRY_CONVERT(int, cm.Num_Matricula) DESC, cm.fecha_pago DESC
        ) cab
        WHERE 1 = 1
          {personal_filter}
          {admissions_user_filter}
          AND (
              NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '') IS NULL
              OR REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '-', ''), ' ', '') <> ?
          )
        """
    by_user_period_query = f"""
        SELECT
            COALESCE(TRY_CONVERT(varchar(50), p.codperiodo), '') AS codigo_periodo,
            COALESCE(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), pe.Detalle_Periodo))), N'Sin periodo') AS detalle_periodo,
            COALESCE(TRY_CONVERT(int, pe.anio), YEAR(TRY_CONVERT(date, p.Fecha_Ingreso))) AS anio_periodo,
            COALESCE(TRY_CONVERT(varchar(50), u.id_usuarios), TRY_CONVERT(varchar(50), p.codasesor), 'SIN_USUARIO') AS usuario_id,
            COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), u.nombres))), N''), N'Sin asesor vinculado') AS usuario_nombre,
            COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), u.login))), N''), N'') AS usuario_login,
            COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), tu.detalle_tipo_us))), N''), N'ADMISIONES') AS tipo_usuario,
            COUNT(*) AS total_ingresados,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL THEN 1 ELSE 0 END) AS ingresaron_cabecera_matricula,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'A' THEN 1 ELSE 0 END) AS activos,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'P' THEN 1 ELSE 0 END) AS inactivos,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'G' THEN 1 ELSE 0 END) AS graduados,
            SUM(CASE WHEN cab.codigo_estud IS NOT NULL AND matched.estado_codigo = 'R' THEN 1 ELSE 0 END) AS retirados,
            SUM(CASE WHEN cab.codigo_estud IS NULL THEN 1 ELSE 0 END) AS pendientes_matricula,
            SUM(CASE
                WHEN cab.codigo_estud IS NOT NULL
                 AND COALESCE(matched.estado_codigo, '') NOT IN ('A', 'P', 'G', 'R')
                THEN 1 ELSE 0
            END) AS sin_estado
        FROM dbo.PREINSCRIPCION p
        LEFT JOIN dbo.PERIODO pe
          ON TRY_CONVERT(varchar(50), pe.cod_periodo) = TRY_CONVERT(varchar(50), p.codperiodo)
        LEFT JOIN dbo.USUARIO_SIS u
          ON TRY_CONVERT(int, u.id_usuarios) = TRY_CONVERT(int, p.codasesor)
        LEFT JOIN dbo.TIPO_USUARIO tu
          ON TRY_CONVERT(int, tu.Codigo_tipo_us) = COALESCE(TRY_CONVERT(int, u.tp_us), TRY_CONVERT(int, u.tipousuario))
        OUTER APPLY (
            SELECT TOP (1)
                d.codigo_estud,
                d.Cedula_Est,
                UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(10), d.Estado)))) AS estado_codigo
            FROM dbo.DATOS_ESTUD d
            WHERE (
                  NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Codestu))), '') IS NOT NULL
                  AND TRY_CONVERT(varchar(50), p.Codestu) = TRY_CONVERT(varchar(50), d.codigo_estud)
                )
               OR (
                  NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '') IS NOT NULL
                  AND REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '-', ''), ' ', '')
                      = REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))), '-', ''), ' ', '')
                )
            ORDER BY
                CASE UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(10), d.Estado))))
                    WHEN 'A' THEN 1
                    WHEN 'P' THEN 2
                    WHEN 'G' THEN 3
                    WHEN 'R' THEN 4
                    ELSE 9
                END
        ) matched
        OUTER APPLY (
            SELECT TOP (1)
                cm.codigo_estud,
                cm.codigo_periodo,
                cm.cod_anio_Basica,
                cm.Num_Matricula
            FROM dbo.CABECERA_MATRICULA cm
            WHERE TRY_CONVERT(varchar(50), cm.codigo_estud) = COALESCE(
                    TRY_CONVERT(varchar(50), matched.codigo_estud),
                    NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Codestu))), '')
                )
              AND TRY_CONVERT(varchar(50), cm.codigo_periodo) = TRY_CONVERT(varchar(50), p.codperiodo)
              AND (
                    NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.codcarrera))), '') IS NULL
                    OR TRY_CONVERT(varchar(50), cm.cod_anio_Basica) = TRY_CONVERT(varchar(50), p.codcarrera)
                  )
            ORDER BY TRY_CONVERT(int, cm.Num_Matricula) DESC, cm.fecha_pago DESC
        ) cab
        WHERE 1 = 1
          {admissions_user_filter}
          {personal_filter}
          AND (
              NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '') IS NULL
              OR REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '-', ''), ' ', '') <> ?
         )
        GROUP BY
            COALESCE(TRY_CONVERT(varchar(50), p.codperiodo), ''),
            COALESCE(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), pe.Detalle_Periodo))), N'Sin periodo'),
            COALESCE(TRY_CONVERT(int, pe.anio), YEAR(TRY_CONVERT(date, p.Fecha_Ingreso))),
            COALESCE(TRY_CONVERT(varchar(50), u.id_usuarios), TRY_CONVERT(varchar(50), p.codasesor), 'SIN_USUARIO'),
            COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), u.nombres))), N''), N'Sin asesor vinculado'),
            COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), u.login))), N''), N''),
            COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), tu.detalle_tipo_us))), N''), N'ADMISIONES')
        ORDER BY anio_periodo DESC, detalle_periodo DESC, usuario_nombre
        """

    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(trend_query, personal_params)
        trend_rows = cursor.fetchall()
        cursor.execute(totals_query, personal_params)
        totals_row = cursor.fetchone()
        cursor.execute(flow_totals_query, [*personal_params, _DASHBOARD_IGNORED_CEDULA])
        flow_totals_row = cursor.fetchone()
        cursor.execute(by_user_period_query, [*personal_params, _DASHBOARD_IGNORED_CEDULA])
        by_user_period_rows = cursor.fetchall()

    month_names = {
        1: "Ene",
        2: "Feb",
        3: "Mar",
        4: "Abr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dic",
    }
    trend = [
        {
            "anio": int(row.anio),
            "mes": int(row.mes),
            "fecha_inicio": row.fecha_inicio.isoformat() if isinstance(row.fecha_inicio, date) else str(row.fecha_inicio),
            "periodo_mes": f"{int(row.anio):04d}-{int(row.mes):02d}",
            "mes_nombre": f"{month_names.get(int(row.mes), str(row.mes))} {int(row.anio)}",
            "total_estudiantes": int(row.total_estudiantes or 0),
        }
        for row in trend_rows
        if row.anio is not None and row.mes is not None
    ]
    total_ingresados = int(getattr(totals_row, "total_ingresados", 0) or 0)
    matriculados_cabecera = int(getattr(flow_totals_row, "ingresaron_cabecera_matricula", 0) or 0)
    activos_desde_admision = int(getattr(flow_totals_row, "activos", 0) or 0)
    inactivos_desde_admision = int(getattr(flow_totals_row, "inactivos", 0) or 0)
    graduados_desde_admision = int(getattr(flow_totals_row, "graduados", 0) or 0)
    retirados_desde_admision = int(getattr(flow_totals_row, "retirados", 0) or 0)
    pendientes_matricula = int(getattr(flow_totals_row, "pendientes_matricula", 0) or 0)
    sin_estado_desde_admision = int(getattr(flow_totals_row, "sin_estado", 0) or 0)
    pendientes = max(0, total_ingresados - activos_desde_admision)

    return {
        "dashboard_type": "admisiones",
        "trend": trend,
        "states": [
            {"estado_codigo": "ING", "estado_nombre": "Ingresados", "total_estudiantes": total_ingresados},
            {"estado_codigo": "A", "estado_nombre": "Activos", "total_estudiantes": activos_desde_admision},
            {"estado_codigo": "PEN", "estado_nombre": "Pendientes o no activos", "total_estudiantes": pendientes},
        ],
        "active_by_type": [],
        "active_regular_students": activos_desde_admision,
        "active_homologation_students": 0,
        "active_regular_homologation_students": activos_desde_admision,
        "total_estudiantes": total_ingresados,
        "admissions": {
            "total_ingresados": total_ingresados,
            "ingresaron_cabecera_matricula": matriculados_cabecera,
            "activos_desde_admision": activos_desde_admision,
            "inactivos_desde_admision": inactivos_desde_admision,
            "graduados_desde_admision": graduados_desde_admision,
            "retirados_desde_admision": retirados_desde_admision,
            "pendientes_matricula": pendientes_matricula,
            "sin_estado_desde_admision": sin_estado_desde_admision,
            "pendientes_o_no_activos": pendientes,
            "activos_sistema": activos_desde_admision,
            "total_con_codigo": int(getattr(totals_row, "total_con_codigo", 0) or 0),
            "total_con_cedula": int(getattr(totals_row, "total_con_cedula", 0) or 0),
            "vista_global_por_sin_registros": False,
            "codigo_asesor": ", ".join(advisor_codes) if advisor_codes else "",
            "usuario_consultado": current_user.nombres or current_user.login if current_user is not None else "",
            "mensaje_vista": (
                "No existen preinscripciones vinculadas directamente a este asesor por PREINSCRIPCION.codasesor."
                if current_user is not None and total_ingresados == 0
                else ""
            ),
            "por_usuario_periodo": [
                {
                    "codigo_periodo": str(row.codigo_periodo or ""),
                    "detalle_periodo": str(row.detalle_periodo or "Sin periodo"),
                    "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
                    "usuario_id": str(row.usuario_id or ""),
                    "usuario_nombre": str(row.usuario_nombre or "Sin asesor"),
                    "usuario_login": str(row.usuario_login or ""),
                    "tipo_usuario": str(row.tipo_usuario or "ADMISIONES"),
                    "total_ingresados": int(row.total_ingresados or 0),
                    "ingresaron_cabecera_matricula": int(row.ingresaron_cabecera_matricula or 0),
                    "ingresaron_carreraxestud": int(row.ingresaron_cabecera_matricula or 0),
                    "activos": int(row.activos or 0),
                    "inactivos": int(row.inactivos or 0),
                    "graduados": int(row.graduados or 0),
                    "retirados": int(row.retirados or 0),
                    "pendientes_matricula": int(row.pendientes_matricula or 0),
                    "sin_estado": int(row.sin_estado or 0),
                }
                for row in by_user_period_rows
            ],
        },
        "criteria": {
            "fecha": "PREINSCRIPCION.Fecha_Ingreso",
            "fuente": "PREINSCRIPCION para estudiantes ingresados; PREINSCRIPCION.codasesor se cruza exclusivamente con USUARIO_SIS.id_usuarios; CABECERA_MATRICULA confirma conversion a matricula; DATOS_ESTUD valida estado activo/inactivo/graduado/retirado",
            "excluidos": [f"Cedula {_DASHBOARD_IGNORED_CEDULA} solo en dashboard"],
        },
    }


def _matricula_list_punto_filter(punto: str | None) -> str:
    if punto == "ULTIMA":
        return "AND punto_matricula = 'ULTIMA'"
    if punto == "BOTH":
        return ""
    return "AND punto_matricula = 'PRIMERA'"


def _matricula_list_point_param(punto: str | None) -> str:
    return punto or "PRIMERA"


@router.get(
    "/matricula-summary",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def matricula_summary(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    response: Response,
) -> dict[str, Any]:
    del current_user
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    rows_query = (
        _MATRICULA_CNE_CTE
        + """
        SELECT
            tipo_matricula,
            estado_codigo,
            estado_nombre,
            COUNT(*) AS total_estudiantes
        FROM matricula_cne_catalogada
        GROUP BY tipo_matricula, estado_codigo, estado_nombre
        ORDER BY tipo_matricula, estado_codigo
        """
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(rows_query)
            rows = cursor.fetchall()

        totals_by_tipo: dict[str, int] = {"R": 0, "H": 0}
        totals_by_estado: dict[str, int] = dict.fromkeys((codigo for codigo, _ in _MAIN_ESTADOS), 0)
        summary_map: dict[tuple[str, str, str], int] = {}
        for row in rows:
            tipo = str(row.tipo_matricula or "").upper()
            estado_codigo = str(row.estado_codigo or "").upper()
            estado_nombre = str(row.estado_nombre or "Sin estado")
            total = int(row.total_estudiantes or 0)
            if tipo not in totals_by_tipo:
                continue
            totals_by_tipo[tipo] += total
            if estado_codigo in totals_by_estado:
                totals_by_estado[estado_codigo] += total
            key = (tipo, estado_codigo, estado_nombre)
            summary_map[key] = summary_map.get(key, 0) + total

        items = [
            {
                "tipo_matricula": tipo,
                "estado_codigo": estado_codigo,
                "estado_nombre": estado_nombre,
                "total_estudiantes": total,
            }
            for (tipo, estado_codigo, estado_nombre), total in sorted(summary_map.items())
        ]

        return {
            "items": items,
            "totals_by_tipo": totals_by_tipo,
            "totals_by_estado": totals_by_estado,
            "total_general": sum(totals_by_tipo.values()),
            "fuente": "dbo.TOTALESTUDMATRICCNE",
            "criterio": "Misma granularidad del Crystal CryListaTotalEstudResumenCNE.rpt",
            "consultado_en": datetime.now(timezone.utc).isoformat(),
        }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando resumen de matricula: {exc}") from exc


@router.get(
    "/matricula-career-state-summary",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def matricula_career_state_summary(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    response: Response,
) -> dict[str, Any]:
    del current_user
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    query = (
        _MATRICULA_CNE_CARRERA_CTE
        + """
        SELECT
            escuela,
            cod_anio_basica,
            nombre_carrera,
            tipo_matricula,
            estado_codigo,
            estado_nombre,
            COUNT(*) AS total_estudiantes
        FROM matricula_cne_carrera
        WHERE estado_codigo IN ('A', 'G', 'P', 'R')
        GROUP BY escuela, cod_anio_basica, nombre_carrera, tipo_matricula, estado_codigo, estado_nombre
        ORDER BY escuela, nombre_carrera, tipo_matricula, estado_codigo
        """
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()

        totals_by_tipo: dict[str, int] = {"R": 0, "H": 0}
        totals_by_estado: dict[str, int] = dict.fromkeys((codigo for codigo, _ in _MAIN_ESTADOS), 0)
        career_keys: set[tuple[str, str, str]] = set()
        escuelas: set[str] = set()
        items: list[dict[str, Any]] = []

        for row in rows:
            tipo = str(row.tipo_matricula or "").upper()
            estado_codigo = str(row.estado_codigo or "").upper()
            total = int(row.total_estudiantes or 0)
            escuela = str(row.escuela or "Sin escuela")
            cod_anio_basica = str(row.cod_anio_basica or "")
            nombre_carrera = str(row.nombre_carrera or "Sin carrera registrada")
            if tipo in totals_by_tipo:
                totals_by_tipo[tipo] += total
            if estado_codigo in totals_by_estado:
                totals_by_estado[estado_codigo] += total
            escuelas.add(escuela)
            career_keys.add((escuela, cod_anio_basica, nombre_carrera))
            items.append(
                {
                    "escuela": escuela,
                    "cod_anio_basica": cod_anio_basica,
                    "nombre_carrera": nombre_carrera,
                    "tipo_matricula": tipo,
                    "estado_codigo": estado_codigo,
                    "estado_nombre": str(row.estado_nombre or estado_codigo),
                    "total_estudiantes": total,
                }
            )

        return {
            "items": items,
            "totals_by_tipo": totals_by_tipo,
            "totals_by_estado": totals_by_estado,
            "total_general": sum(totals_by_tipo.values()),
            "total_escuelas": len(escuelas),
            "total_carreras": len(career_keys),
            "fuente": "dbo.TOTALESTUDMATRICCNE enlazada con dbo.CARRERAS y dbo.ESTADO",
            "consultado_en": datetime.now(timezone.utc).isoformat(),
        }
    except pyodbc.Error as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Error consultando reporteria por escuela y carrera: {exc}",
        ) from exc


@router.get(
    "/matricula-career-state-students",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def matricula_career_state_students(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    response: Response,
    cod_anio_basica: Annotated[str | None, Query(description="Codigo de carrera")] = None,
    nombre_carrera: Annotated[str | None, Query(description="Nombre de carrera")] = None,
    escuela: Annotated[str | None, Query(description="Escuela")] = None,
    estado_codigo: Annotated[str | None, Query(description="Estado A, P, R, G")] = None,
    tipo_matricula: Annotated[str | None, Query(description="Tipo R o H")] = None,
    limit: Annotated[int, Query(ge=1, le=20000)] = 10000,
) -> dict[str, Any]:
    del current_user
    estado = estado_codigo.strip().upper() if estado_codigo else None
    tipo = tipo_matricula.strip().upper() if tipo_matricula else None
    if estado and estado not in {codigo for codigo, _ in _MAIN_ESTADOS}:
        raise HTTPException(status_code=400, detail="estado_codigo debe ser A, G, P o R")
    if tipo:
        tipo = _validate_tipo(tipo)
    cod_carrera = _clean_cell(cod_anio_basica)
    carrera = _clean_cell(nombre_carrera) or None
    escuela_filter = _clean_cell(escuela) or None
    top_limit = max(1, min(int(limit or 10000), 20000))
    query = (
        _MATRICULA_CNE_CARRERA_CTE
        + f"""
        SELECT TOP ({top_limit})
            cne.escuela,
            cne.cod_anio_basica,
            cne.nombre_carrera,
            cne.tipo_matricula,
            cne.estado_codigo,
            cne.estado_nombre,
            detalle.codigo_estud,
            cne.Cedula_Est,
            cne.Apellidos_nombre,
            detalle.correo AS correo_personal_datos,
            detalle.correointec AS correo_intec_datos
        FROM matricula_cne_carrera cne
        OUTER APPLY (
            SELECT TOP (1)
                datos.codigo_estud,
                datos.correo,
                datos.correointec
            FROM dbo.DATOS_ESTUD datos
            WHERE LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), datos.Cedula_Est))) = cne.Cedula_Est
            ORDER BY TRY_CONVERT(bigint, datos.codigo_estud) DESC
        ) detalle
        WHERE (? IS NULL OR cne.cod_anio_basica = ?)
          AND (? IS NULL OR cne.nombre_carrera = ?)
          AND (? IS NULL OR cne.escuela = ?)
          AND (? IS NULL OR cne.estado_codigo = ?)
          AND (? IS NULL OR cne.tipo_matricula = ?)
        ORDER BY cne.Apellidos_nombre, detalle.codigo_estud
        """
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                query,
                cod_carrera or None,
                cod_carrera or None,
                carrera,
                carrera,
                escuela_filter,
                escuela_filter,
                estado,
                estado,
                tipo,
                tipo,
            )
            rows = cursor.fetchall()

        items = [
            {
                "escuela": str(row.escuela or "Sin escuela"),
                "cod_anio_basica": str(row.cod_anio_basica or ""),
                "nombre_carrera": str(row.nombre_carrera or "Sin carrera registrada"),
                "punto_matricula": "PRIMERA",
                "tipo_matricula": str(row.tipo_matricula or ""),
                "estado_codigo": str(row.estado_codigo or ""),
                "estado_nombre": str(row.estado_nombre or ""),
                "codigo_estud": str(row.codigo_estud or ""),
                "cedula": str(row.Cedula_Est or ""),
                "nombre_estudiante": str(row.Apellidos_nombre or ""),
                "correo_personal": str(row.correo_personal_datos or ""),
                "correo_intec": str(row.correo_intec_datos or ""),
                "periodo": "",
                "detalle_periodo": "Reporte general CNE",
                "anio_periodo": None,
            }
            for row in rows
        ]

        return {
            "items": items,
            "total": len(items),
            "criteria": {
                "cod_anio_basica": cod_carrera,
                "nombre_carrera": carrera,
                "escuela": escuela_filter,
                "estado_codigo": estado,
                "tipo_matricula": tipo,
            },
            "fuente": "dbo.TOTALESTUDMATRICCNE",
            "consultado_en": datetime.now(timezone.utc).isoformat(),
        }
    except pyodbc.Error as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Error consultando estudiantes por carrera: {exc}",
        ) from exc


@router.get(
    "/dashboard-matricula",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def dashboard_matricula(
    current_user: Annotated[SessionUser, Depends(_DASHBOARD_ACCESS)],
    response: Response,
) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"

    if (current_user.rol or "").strip().upper() == "ADMISIONES":
        try:
            return _dashboard_admisiones(current_user)
        except pyodbc.Error as exc:
            raise HTTPException(status_code=500, detail=f"Error consultando dashboard personal de admisiones: {exc}") from exc

    trend_query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            YEAR(periodo_fecha_inicio) AS anio,
            MONTH(periodo_fecha_inicio) AS mes,
            MIN(periodo_fecha_inicio) AS fecha_inicio,
            COUNT(*) AS total_estudiantes
        FROM (
            SELECT
                bc.codigo_estud,
                bc.Cedula_Est,
                bc.tipo_matricula,
                bc.estado_codigo,
                COALESCE(
                    TRY_CONVERT(date, p.fechain, 121),
                    TRY_CONVERT(date, p.fechain, 120),
                    TRY_CONVERT(date, p.fechain, 103),
                    TRY_CONVERT(date, p.fechain, 105),
                    TRY_CONVERT(date, p.fechain)
                ) AS periodo_fecha_inicio
            FROM base_cruce bc
            INNER JOIN [dbo].[PERIODO] p
                ON TRY_CONVERT(varchar(50), p.cod_periodo) = TRY_CONVERT(varchar(50), bc.codigo_periodo)
            WHERE bc.estado_codigo IN ('A', 'G', 'P', 'R')
              AND __DASHBOARD_IGNORED_CEDULA_FILTER__
        ) trend_base
        WHERE periodo_fecha_inicio IS NOT NULL
        GROUP BY YEAR(periodo_fecha_inicio), MONTH(periodo_fecha_inicio)
        ORDER BY anio, mes
        """
    )
    states_query = (
        _MATRICULA_CNE_CTE
        + """
        SELECT
            cne.estado_codigo,
            MAX(cne.estado_nombre) AS estado_nombre,
            COUNT(*) AS total_estudiantes
        FROM matricula_cne_catalogada cne
        WHERE cne.estado_codigo IN ('A', 'G', 'P', 'R')
        GROUP BY cne.estado_codigo
        ORDER BY cne.estado_codigo
        """
    )
    active_by_type_query = (
        _MATRICULA_CNE_CTE
        + """
        SELECT
            cne.tipo_matricula,
            COUNT(*) AS total_estudiantes
        FROM matricula_cne_catalogada cne
        WHERE cne.estado_codigo = 'A'
          AND cne.tipo_matricula IN ('R', 'H')
        GROUP BY cne.tipo_matricula
        ORDER BY cne.tipo_matricula
        """
    )
    trend_query = trend_query.replace("__DASHBOARD_IGNORED_CEDULA_FILTER__", _DASHBOARD_IGNORED_CEDULA_SQL)

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(trend_query)
            trend_rows = cursor.fetchall()
            cursor.execute(states_query)
            state_rows = cursor.fetchall()
            cursor.execute(active_by_type_query)
            active_by_type_rows = cursor.fetchall()

        month_names = {
            1: "Ene",
            2: "Feb",
            3: "Mar",
            4: "Abr",
            5: "May",
            6: "Jun",
            7: "Jul",
            8: "Ago",
            9: "Sep",
            10: "Oct",
            11: "Nov",
            12: "Dic",
        }
        trend = [
            {
                "anio": int(row.anio),
                "mes": int(row.mes),
                "fecha_inicio": row.fecha_inicio.isoformat() if isinstance(row.fecha_inicio, date) else str(row.fecha_inicio),
                "periodo_mes": f"{int(row.anio):04d}-{int(row.mes):02d}",
                "mes_nombre": f"{month_names.get(int(row.mes), str(row.mes))} {int(row.anio)}",
                "total_estudiantes": int(row.total_estudiantes or 0),
            }
            for row in trend_rows
            if row.anio is not None and row.mes is not None
        ]
        estado_nombres = {codigo: nombre for codigo, nombre in _MAIN_ESTADOS}
        estado_totales = {codigo: 0 for codigo, _ in _MAIN_ESTADOS}
        for row in state_rows:
            codigo = str(row.estado_codigo or "").strip().upper()
            if codigo in estado_totales:
                estado_nombres[codigo] = str(row.estado_nombre or estado_nombres[codigo])
                estado_totales[codigo] = int(row.total_estudiantes or 0)

        estados = [
            {
                "estado_codigo": codigo,
                "estado_nombre": estado_nombres[codigo],
                "total_estudiantes": estado_totales[codigo],
            }
            for codigo, _ in _MAIN_ESTADOS
        ]
        total_estudiantes = sum(estado_totales.values())
        active_total = estado_totales.get("A", 0)

        raw_active_by_type = [
            {
                "tipo_matricula": str(row.tipo_matricula or "").strip().upper(),
                "total_estudiantes": int(row.total_estudiantes or 0),
            }
            for row in active_by_type_rows
        ]
        active_regular = next(
            (item["total_estudiantes"] for item in raw_active_by_type if item["tipo_matricula"] == "R"),
            0,
        )
        active_homologation = next(
            (item["total_estudiantes"] for item in raw_active_by_type if item["tipo_matricula"] == "H"),
            0,
        )
        split_total = active_regular + active_homologation
        if active_total != split_total:
            difference = active_total - split_total
            if active_regular >= active_homologation:
                active_regular = max(0, active_regular + difference)
            else:
                active_homologation = max(0, active_homologation + difference)

        active_by_type = [
            {"tipo_matricula": "R", "total_estudiantes": active_regular},
            {"tipo_matricula": "H", "total_estudiantes": active_homologation},
        ]

        admissions_dashboard = _dashboard_admisiones()

        return {
            "dashboard_type": "matricula",
            "trend": trend,
            "states": estados,
            "active_by_type": active_by_type,
            "active_regular_students": active_regular,
            "active_homologation_students": active_homologation,
            "active_regular_homologation_students": active_total,
            "total_estudiantes": total_estudiantes,
            "consultado_en": datetime.now(timezone.utc).isoformat(),
            "criteria": {
                "fecha": "PERIODO.fechain",
                "excluidos": [],
                "fuente": "dbo.TOTALESTUDMATRICCNE enlazada con dbo.ESTADO; consulta directa sin cache para igualar el reporte institucional",
            },
            "admissions": admissions_dashboard.get("admissions", {}),
        }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando dashboard de matricula: {exc}") from exc


@router.get(
    "/dashboard-matricula/admisiones-students",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def dashboard_admisiones_students(
    current_user: Annotated[SessionUser, Depends(_DASHBOARD_ACCESS)],
    estado: Annotated[
        str,
        Query(
            description=(
                "Filtro: ALL, A, P, G, R, CABECERA_MATRICULA, PENDIENTE_MATRICULA o SIN_ESTADO"
            )
        ),
    ] = "ALL",
    codigo_periodo: Annotated[str | None, Query(description="Periodo academico de PREINSCRIPCION.codperiodo")] = None,
    limit: Annotated[int, Query(ge=1, le=20000)] = 10000,
) -> dict[str, Any]:
    status = (estado or "ALL").strip().upper()
    allowed_statuses = {"ALL", "A", "P", "G", "R", "CABECERA_MATRICULA", "PENDIENTE_MATRICULA", "SIN_ESTADO"}
    if status not in allowed_statuses:
        raise HTTPException(status_code=400, detail="Estado de admisiones no valido.")

    period_filter = ""
    period_params: list[Any] = []
    if codigo_periodo and str(codigo_periodo).strip():
        period_filter = "AND TRY_CONVERT(varchar(50), p.codperiodo) = ?"
        period_params.append(str(codigo_periodo).strip())

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            advisor_codes = _resolve_admission_advisor_codes(cursor, current_user)
            personal_filter = ""
            personal_params: list[Any] = []
            admissions_user_filter = "AND COALESCE(TRY_CONVERT(int, u.tp_us), TRY_CONVERT(int, u.tipousuario)) = 5"
            if (current_user.rol or "").strip().upper() == "ADMISIONES":
                admissions_user_filter = ""
                if not advisor_codes:
                    return {
                        "items": [],
                        "total": 0,
                        "estado": status,
                        "codigo_periodo": codigo_periodo or "",
                        "codigo_asesor": "",
                    }
                personal_filter = f"AND TRY_CONVERT(varchar(50), p.codasesor) IN ({_sql_placeholders(advisor_codes)})"
                personal_params = advisor_codes

            status_filter = ""
            if status == "PENDIENTE_MATRICULA":
                status_filter = "AND cab.codigo_estud IS NULL"
            elif status == "CABECERA_MATRICULA":
                status_filter = "AND cab.codigo_estud IS NOT NULL"
            elif status == "SIN_ESTADO":
                status_filter = """
                  AND cab.codigo_estud IS NOT NULL
                  AND COALESCE(matched.estado_codigo, '') NOT IN ('A', 'P', 'G', 'R')
                """
            elif status in {"A", "P", "G", "R"}:
                status_filter = "AND cab.codigo_estud IS NOT NULL AND matched.estado_codigo = ?"

            status_params = [status] if status in {"A", "P", "G", "R"} else []
            query = f"""
                SELECT TOP ({int(limit)})
                    TRY_CONVERT(varchar(50), p.Codestu) AS codestu,
                    TRY_CONVERT(varchar(50), p.Cedula) AS cedula_preinscripcion,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), p.Apellidos_nombre))) AS nombre_preinscripcion,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), p.correo))) AS correo_preinscripcion,
                    LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.telefono))) AS telefono,
                    TRY_CONVERT(varchar(50), p.codperiodo) AS codigo_periodo,
                    COALESCE(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), pe.Detalle_Periodo))), N'Sin periodo') AS detalle_periodo,
                    LTRIM(RTRIM(TRY_CONVERT(varchar(10), pe.TipoMatricula))) AS tipo_matricula,
                    TRY_CONVERT(int, pe.anio) AS anio_periodo,
                    TRY_CONVERT(varchar(50), p.codcarrera) AS codcarrera,
                    COALESCE(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), ca.Nombre_Basica))), N'Sin carrera') AS carrera,
                    p.Fecha_Ingreso AS fecha_ingreso,
                    TRY_CONVERT(varchar(50), p.codasesor) AS codasesor,
                    TRY_CONVERT(varchar(50), u.id_usuarios) AS usuario_id,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), u.nombres))) AS usuario_nombre,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), u.login))) AS usuario_login,
                    matched.codigo_estud AS codigo_estud,
                    matched.Cedula_Est AS cedula_matricula,
                    matched.Apellidos_nombre AS nombre_matricula,
                    matched.estado_codigo AS estado_codigo,
                    cab.codigo_estud AS codigo_estud_cabecera,
                    cab.Num_Matricula AS num_matricula
                FROM dbo.PREINSCRIPCION p
                LEFT JOIN dbo.PERIODO pe
                  ON TRY_CONVERT(varchar(50), pe.cod_periodo) = TRY_CONVERT(varchar(50), p.codperiodo)
                LEFT JOIN dbo.CARRERAS ca
                  ON TRY_CONVERT(varchar(50), ca.Cod_AnioBasica) = TRY_CONVERT(varchar(50), p.codcarrera)
                LEFT JOIN dbo.USUARIO_SIS u
                  ON TRY_CONVERT(int, u.id_usuarios) = TRY_CONVERT(int, p.codasesor)
                OUTER APPLY (
                    SELECT TOP (1)
                        d.codigo_estud,
                        d.Cedula_Est,
                        d.Apellidos_nombre,
                        UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(10), d.Estado)))) AS estado_codigo
                    FROM dbo.DATOS_ESTUD d
                    WHERE (
                          NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Codestu))), '') IS NOT NULL
                          AND TRY_CONVERT(varchar(50), p.Codestu) = TRY_CONVERT(varchar(50), d.codigo_estud)
                        )
                       OR (
                          NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '') IS NOT NULL
                          AND REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '-', ''), ' ', '')
                              = REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))), '-', ''), ' ', '')
                        )
                    ORDER BY
                        CASE UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(10), d.Estado))))
                            WHEN 'A' THEN 1
                            WHEN 'P' THEN 2
                            WHEN 'G' THEN 3
                            WHEN 'R' THEN 4
                            ELSE 9
                        END
                ) matched
                OUTER APPLY (
                    SELECT TOP (1)
                        cm.codigo_estud,
                        cm.codigo_periodo,
                        cm.cod_anio_Basica,
                        cm.Num_Matricula
                    FROM dbo.CABECERA_MATRICULA cm
                    WHERE TRY_CONVERT(varchar(50), cm.codigo_estud) = COALESCE(
                            TRY_CONVERT(varchar(50), matched.codigo_estud),
                            NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Codestu))), '')
                        )
                      AND TRY_CONVERT(varchar(50), cm.codigo_periodo) = TRY_CONVERT(varchar(50), p.codperiodo)
                      AND (
                            NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.codcarrera))), '') IS NULL
                            OR TRY_CONVERT(varchar(50), cm.cod_anio_Basica) = TRY_CONVERT(varchar(50), p.codcarrera)
                          )
                    ORDER BY TRY_CONVERT(int, cm.Num_Matricula) DESC, cm.fecha_pago DESC
                ) cab
                WHERE 1 = 1
                  {personal_filter}
                  {admissions_user_filter}
                  {period_filter}
                  {status_filter}
                  AND (
                      NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '') IS NULL
                      OR REPLACE(REPLACE(LTRIM(RTRIM(TRY_CONVERT(varchar(50), p.Cedula))), '-', ''), ' ', '') <> ?
                  )
                ORDER BY pe.anio DESC, pe.cod_periodo DESC, p.Apellidos_nombre
            """
            params = [*personal_params, *period_params, *status_params, _DASHBOARD_IGNORED_CEDULA]
            cursor.execute(query, params)
            rows = cursor.fetchall()

        status_names = {
            "A": "Activo",
            "P": "Inactivo",
            "G": "Graduado",
            "R": "Retirado",
            "PENDIENTE_MATRICULA": "Pendiente de matricula",
            "CABECERA_MATRICULA": "Con cabecera de matricula",
            "SIN_ESTADO": "Sin estado",
            "ALL": "Todos",
        }
        items = []
        for row in rows:
            has_cabecera = row.codigo_estud_cabecera is not None
            estado_codigo = str(row.estado_codigo or "").strip().upper()
            if not has_cabecera:
                estado_final = "PENDIENTE_MATRICULA"
            elif estado_codigo in {"A", "P", "G", "R"}:
                estado_final = estado_codigo
            else:
                estado_final = "SIN_ESTADO"
            items.append(
                {
                    "codestu": str(row.codestu or ""),
                    "codigo_estud": str(row.codigo_estud or row.codestu or ""),
                    "cedula": str(row.cedula_matricula or row.cedula_preinscripcion or ""),
                    "nombre_estudiante": str(row.nombre_matricula or row.nombre_preinscripcion or ""),
                    "correo": str(row.correo_preinscripcion or ""),
                    "telefono": str(row.telefono or ""),
                    "codigo_periodo": str(row.codigo_periodo or ""),
                    "detalle_periodo": str(row.detalle_periodo or ""),
                    "tipo_matricula": str(row.tipo_matricula or ""),
                    "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
                    "codcarrera": str(row.codcarrera or ""),
                    "carrera": str(row.carrera or ""),
                    "fecha_ingreso": row.fecha_ingreso.isoformat() if isinstance(row.fecha_ingreso, date) else str(row.fecha_ingreso or ""),
                    "codasesor": str(row.codasesor or ""),
                    "usuario_id": str(row.usuario_id or ""),
                    "usuario_nombre": str(row.usuario_nombre or ""),
                    "usuario_login": str(row.usuario_login or ""),
                    "estado_codigo": estado_final,
                    "estado_nombre": status_names.get(estado_final, estado_final),
                    "tiene_cabecera_matricula": has_cabecera,
                    "num_matricula": str(row.num_matricula or ""),
                }
            )

        return {
            "items": items,
            "total": len(items),
            "estado": status,
            "codigo_periodo": codigo_periodo or "",
            "codigo_asesor": ", ".join(advisor_codes),
        }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando estudiantes de admisiones: {exc}") from exc


@router.get(
    "/dashboard-matricula/students",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def dashboard_matricula_students(
    current_user: Annotated[SessionUser, Depends(_DASHBOARD_ACCESS)],
    anio: Annotated[int, Query(ge=2000, le=2100, description="Anio de PERIODO.fechain")],
    mes: Annotated[int, Query(ge=1, le=12, description="Mes de PERIODO.fechain")],
    limit: Annotated[int, Query(ge=1, le=20000)] = 10000,
) -> dict[str, Any]:
    del current_user
    query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            trend_base.tipo_matricula,
            trend_base.estado_codigo,
            trend_base.estado_nombre,
            trend_base.codigo_estud,
            trend_base.Cedula_Est,
            trend_base.Apellidos_nombre,
            trend_base.nombre_carrera,
            trend_base.correo_intec_datos,
            trend_base.correo_personal_datos,
            trend_base.codigo_periodo,
            trend_base.Detalle_Periodo,
            trend_base.anio_periodo,
            trend_base.periodo_fecha_inicio
        FROM (
            SELECT
                bc.tipo_matricula,
                bc.estado_codigo,
                bc.estado_nombre,
                bc.codigo_estud,
                bc.Cedula_Est,
                bc.Apellidos_nombre,
                bc.nombre_carrera,
                bc.correo_intec_datos,
                bc.correo_personal_datos,
                bc.codigo_periodo,
                bc.Detalle_Periodo,
                bc.anio_periodo,
                COALESCE(
                    TRY_CONVERT(date, p.fechain, 121),
                    TRY_CONVERT(date, p.fechain, 120),
                    TRY_CONVERT(date, p.fechain, 103),
                    TRY_CONVERT(date, p.fechain, 105),
                    TRY_CONVERT(date, p.fechain)
                ) AS periodo_fecha_inicio
            FROM base_cruce bc
            INNER JOIN [dbo].[PERIODO] p
                ON TRY_CONVERT(varchar(50), p.cod_periodo) = TRY_CONVERT(varchar(50), bc.codigo_periodo)
            WHERE bc.estado_codigo IN ('A', 'G', 'P', 'R')
              AND __DASHBOARD_IGNORED_CEDULA_FILTER__
        ) trend_base
        WHERE trend_base.periodo_fecha_inicio IS NOT NULL
          AND YEAR(trend_base.periodo_fecha_inicio) = ?
          AND MONTH(trend_base.periodo_fecha_inicio) = ?
        ORDER BY trend_base.Apellidos_nombre, trend_base.codigo_estud
        """
    )
    query = query.replace("__DASHBOARD_IGNORED_CEDULA_FILTER__", _DASHBOARD_IGNORED_CEDULA_SQL)

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, anio, mes)
            rows = cursor.fetchall()

        items: list[dict[str, Any]] = []
        for row in rows[:limit]:
            fecha_inicio = row.periodo_fecha_inicio
            items.append(
                {
                    "punto_matricula": "TENDENCIA",
                    "tipo_matricula": str(row.tipo_matricula or ""),
                    "estado_codigo": str(row.estado_codigo or ""),
                    "estado_nombre": str(row.estado_nombre or ""),
                    "codigo_estud": str(row.codigo_estud or ""),
                    "cedula": str(row.Cedula_Est or ""),
                    "nombre_estudiante": str(row.Apellidos_nombre or ""),
                    "nombre_carrera": str(row.nombre_carrera or ""),
                    "correo_intec": str(row.correo_intec_datos or ""),
                    "correo_personal": str(row.correo_personal_datos or ""),
                    "periodo": str(row.codigo_periodo or ""),
                    "detalle_periodo": str(row.Detalle_Periodo or ""),
                    "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
                    "fecha_inicio_periodo": fecha_inicio.isoformat()
                    if isinstance(fecha_inicio, (date, datetime))
                    else str(fecha_inicio or ""),
                }
            )

        return {
            "anio": anio,
            "mes": mes,
            "total": len(rows),
            "items": items,
        }
    except pyodbc.Error as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Error consultando estudiantes de tendencia mensual: {exc}",
        ) from exc


@router.get(
    "/matricula-period-summary",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def matricula_period_summary(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    rows_query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            tipo_matricula,
            anio_periodo,
            codigo_periodo,
            Detalle_Periodo,
            estado_codigo,
            codigo_estud,
            Apellidos_nombre,
            fecha_inicio_periodo,
            fecha_fin_periodo,
            Fecha_Ingreso,
            fechaMatricula
        FROM base_cruce
        """
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(rows_query)
            rows = cursor.fetchall()

        reference = _read_matriz_desagregada_reference()
        filtered_rows = rows
        period_map: dict[tuple[str, int | None, str, str], dict[str, Any]] = {}
        year_map: dict[int, dict[str, Any]] = {}

        for row in filtered_rows:
            tipo = str(row.tipo_matricula or "")
            anio = int(row.anio_periodo) if row.anio_periodo is not None else None
            codigo_periodo = str(row.codigo_periodo) if row.codigo_periodo is not None else ""
            detalle = str(row.Detalle_Periodo) if row.Detalle_Periodo is not None else ""
            estado = str(row.estado_codigo or "").upper()
            key = (tipo, anio, codigo_periodo, detalle)
            period_item = period_map.setdefault(
                key,
                {
                    "tipo_matricula": tipo,
                    "anio_periodo": anio,
                    "codigo_periodo": codigo_periodo,
                    "detalle_periodo": detalle,
                    "total_estudiantes": 0,
                    "activos": 0,
                    "inactivos": 0,
                    "retirados": 0,
                    "graduados": 0,
                },
            )
            period_item["total_estudiantes"] += 1
            if estado == "A":
                period_item["activos"] += 1
            elif estado == "P":
                period_item["inactivos"] += 1
            elif estado == "R":
                period_item["retirados"] += 1
            elif estado == "G":
                period_item["graduados"] += 1

            if anio is not None:
                start_date = _coerce_date(row.fecha_inicio_periodo)
                end_date = _coerce_date(row.fecha_fin_periodo)
                year_item = year_map.setdefault(
                    anio,
                    {
                        "anio_periodo": anio,
                        "fecha_inicio_min": None,
                        "fecha_fin_max": None,
                        "total_estudiantes": 0,
                        "acumulado_estudiantes": 0,
                        "activos": 0,
                        "inactivos": 0,
                        "retirados": 0,
                        "graduados": 0,
                    },
                )
                year_item["total_estudiantes"] += 1
                if start_date and (year_item["fecha_inicio_min"] is None or start_date < year_item["fecha_inicio_min"]):
                    year_item["fecha_inicio_min"] = start_date
                if end_date and (year_item["fecha_fin_max"] is None or end_date > year_item["fecha_fin_max"]):
                    year_item["fecha_fin_max"] = end_date
                if estado == "A":
                    year_item["activos"] += 1
                elif estado == "P":
                    year_item["inactivos"] += 1
                elif estado == "R":
                    year_item["retirados"] += 1
                elif estado == "G":
                    year_item["graduados"] += 1

        items = sorted(
            period_map.values(),
            key=lambda item: (
                -(item["anio_periodo"] or 0),
                -int(item["codigo_periodo"]) if str(item["codigo_periodo"]).isdigit() else 0,
                item["tipo_matricula"],
            ),
        )
        years = sorted(year_map.values(), key=lambda item: -int(item["anio_periodo"] or 0))
        for year_item in years:
            for field in ("fecha_inicio_min", "fecha_fin_max"):
                if isinstance(year_item[field], date):
                    year_item[field] = year_item[field].isoformat()

        return {
            "items": items,
            "years": years,
            "total": len(items),
            "referencia_matriz": _matricula_reference_meta(reference, len(rows), len(filtered_rows)),
        }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando resumen por periodo: {exc}") from exc


@router.get(
    "/matricula-movement-summary",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def matricula_movement_summary(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    period_query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            anio_periodo,
            codigo_periodo,
            Detalle_Periodo,
            tipo_matricula,
            estado_codigo,
            estado_nombre,
            COUNT(*) AS total_estudiantes,
            COUNT(CASE WHEN estado_codigo = 'A' THEN 1 END) AS activos,
            COUNT(CASE WHEN estado_codigo = 'P' THEN 1 END) AS inactivos,
            COUNT(CASE WHEN estado_codigo = 'R' THEN 1 END) AS retirados,
            COUNT(CASE WHEN estado_codigo = 'G' THEN 1 END) AS graduados
                FROM base_cruce
                WHERE rn_periodo = 1
                    AND anio_periodo IS NOT NULL
        GROUP BY anio_periodo, codigo_periodo, Detalle_Periodo, tipo_matricula, estado_codigo, estado_nombre
        ORDER BY anio_periodo DESC, codigo_periodo DESC, tipo_matricula, estado_nombre
        """
    )
    year_query = (
        _MATRICULA_ACTUAL_CTE
        + """
        , year_counts AS (
            SELECT
                anio_periodo,
                MIN(fecha_inicio_periodo) AS fecha_inicio_min,
                MAX(fecha_fin_periodo) AS fecha_fin_max,
                COUNT(*) AS total_estudiantes,
                COUNT(CASE WHEN estado_codigo = 'A' THEN 1 END) AS activos,
                COUNT(CASE WHEN estado_codigo = 'P' THEN 1 END) AS inactivos,
                COUNT(CASE WHEN estado_codigo = 'R' THEN 1 END) AS retirados,
                COUNT(CASE WHEN estado_codigo = 'G' THEN 1 END) AS graduados
            FROM base_cruce
            WHERE rn_periodo = 1
              AND anio_periodo IS NOT NULL
            GROUP BY anio_periodo
        )
        SELECT
            anio_periodo,
            fecha_inicio_min,
            fecha_fin_max,
            total_estudiantes,
            SUM(total_estudiantes) OVER (ORDER BY anio_periodo ASC ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW) AS acumulado_estudiantes,
            activos,
            inactivos,
            retirados,
            graduados
        FROM year_counts
        ORDER BY anio_periodo DESC
        """
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(period_query)
            rows = cursor.fetchall()
            cursor.execute(year_query)
            year_rows = cursor.fetchall()

        items: list[dict[str, Any]] = [_movement_period_item(row) for row in rows]

        return {
            "items": items,
            "years": [_movement_year_item(row) for row in year_rows],
            "total": len(items),
        }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando resumen de movimiento de matricula: {exc}") from exc


@router.get(
    "/matricula-cruce-completo",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def matricula_cruce_completo(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    resumen_tipo_query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            tipo_matricula,
            COUNT(DISTINCT codigo_estud) AS total_estudiantes,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'A' THEN codigo_estud END) AS activos,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'P' THEN codigo_estud END) AS inactivos,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'R' THEN codigo_estud END) AS retirados,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'G' THEN codigo_estud END) AS graduados
        FROM base_cruce
        GROUP BY tipo_matricula
        ORDER BY tipo_matricula
        """
    )
    resumen_anio_query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            anio_periodo,
            MIN(fecha_inicio_periodo) AS fecha_inicio_min,
            MAX(fecha_fin_periodo) AS fecha_fin_max,
            COUNT(DISTINCT codigo_estud) AS total_estudiantes,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'A' THEN codigo_estud END) AS activos,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'P' THEN codigo_estud END) AS inactivos,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'R' THEN codigo_estud END) AS retirados,
            COUNT(DISTINCT CASE WHEN estado_codigo = 'G' THEN codigo_estud END) AS graduados
        FROM base_cruce
        GROUP BY anio_periodo
        ORDER BY anio_periodo DESC
        """
    )
    cruce_por_codigo_query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            codigo_estud,
            Cedula_Est,
            Apellidos_nombre,
            correo_intec_datos,
            correo_personal_datos,
            tipo_matricula,
            anio_periodo,
            codigo_periodo,
            Detalle_Periodo,
            estado_codigo,
            estado_nombre,
            nombre_carrera
        FROM base_cruce
        ORDER BY Apellidos_nombre
        """
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(resumen_tipo_query)
            resumen_tipo_rows = cursor.fetchall()
            cursor.execute(resumen_anio_query)
            resumen_anio_rows = cursor.fetchall()
            cursor.execute(cruce_por_codigo_query)
            cruce_rows = cursor.fetchall()

        resumen_tipo: list[dict[str, Any]] = [
            {
                "tipo_matricula": row.tipo_matricula,
                "total_estudiantes": int(row.total_estudiantes or 0),
                "activos": int(row.activos or 0),
                "inactivos": int(row.inactivos or 0),
                "retirados": int(row.retirados or 0),
                "graduados": int(row.graduados or 0),
            }
            for row in resumen_tipo_rows
        ]
        cruce_por_codigo: list[dict[str, Any]] = [
            {
                "codigo_estud": str(row.codigo_estud) if row.codigo_estud is not None else "",
                "cedula": row.Cedula_Est,
                "nombre_estudiante": row.Apellidos_nombre,
                "correo_intec_datos": row.correo_intec_datos,
                "correo_personal_datos": row.correo_personal_datos,
                "tipo_matricula": row.tipo_matricula,
                "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
                "codigo_periodo": str(row.codigo_periodo) if row.codigo_periodo is not None else "",
                "detalle_periodo": row.Detalle_Periodo,
                "estado_codigo": row.estado_codigo,
                "estado_nombre": row.estado_nombre,
                "nombre_carrera": row.nombre_carrera,
                "estado_cruce": "OK_DATOS_CARRERAXESTUD_PERIODO",
            }
            for row in cruce_rows
        ]

        total_r = next((int(item["total_estudiantes"]) for item in resumen_tipo if item["tipo_matricula"] == "R"), 0)
        total_h = next((int(item["total_estudiantes"]) for item in resumen_tipo if item["tipo_matricula"] == "H"), 0)

        return {
            "resumen_general": {
                "total_regular_r": total_r,
                "total_homologacion_h": total_h,
                "total_r_h": total_r + total_h,
                "total_registros_unicos": len(cruce_por_codigo),
                "criterio_cruce": "Universo desde DATOS_ESTUD; CARRERAXESTUD clasifica R/H cuando existe y PENSUM enriquece materia; excluye Cambio Periodo, Reingreso, Educacion Continua y el registro indicado.",
            },
            "resumen_por_tipo": resumen_tipo,
            "resumen_por_anio": [_row_to_year_summary(row) for row in resumen_anio_rows],
            "cruce_por_codigo": cruce_por_codigo,
        }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando cruce completo de matricula: {exc}") from exc


@router.get(
    "/matricula-list",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def matricula_list(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    response: Response,
    tipo_matricula: Annotated[str | None, Query(description="R, H o vacio para cruce unico global")] = None,
    estado_codigo: Annotated[str | None, Query(description="Codigo de estado, por ejemplo A, P, R, G")] = None,
    anio_periodo: Annotated[int | None, Query(description="Anio academico de la primera matricula")] = None,
    punto_matricula: Annotated[str | None, Query(description="PRIMERA, ULTIMA o BOTH para movimiento")] = None,
    fuente: Annotated[str | None, Query(description="CNE para replicar el reporte Crystal total de estudiantes")] = None,
    limit: Annotated[int, Query(ge=1, le=10000)] = 300,
) -> dict[str, Any]:
    del current_user
    tipo = None
    if tipo_matricula and tipo_matricula.strip().upper() != "ALL":
        tipo = "RH" if tipo_matricula.strip().upper() == "RH" else _validate_tipo(tipo_matricula)
    estado = estado_codigo.strip().upper() if estado_codigo else None
    punto = punto_matricula.strip().upper() if punto_matricula else None
    fuente_reporte = _clean_cell(fuente).upper()
    if fuente_reporte not in {"", "CNE"}:
        raise HTTPException(status_code=400, detail="fuente debe ser CNE o permanecer vacia")
    if punto and punto not in {"PRIMERA", "ULTIMA", "BOTH", "ALL"}:
        raise HTTPException(status_code=400, detail="punto_matricula debe ser PRIMERA, ULTIMA, BOTH o ALL")
    punto_filter = _matricula_list_punto_filter(punto)
    punto_return = _matricula_list_point_param(punto)

    if fuente_reporte == "CNE":
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        query = (
            _MATRICULA_CNE_CTE
            + f"""
            SELECT TOP ({limit})
                cne.tipo_matricula,
                cne.estado_codigo,
                cne.estado_nombre,
                detalle.codigo_estud,
                cne.Cedula_Est,
                cne.Apellidos_nombre,
                cne.nombre_carrera,
                detalle.correo AS correo_personal,
                detalle.correointec AS correo_intec
            FROM matricula_cne_catalogada cne
            OUTER APPLY (
                SELECT TOP (1)
                    datos.codigo_estud,
                    datos.correo,
                    datos.correointec
                FROM dbo.DATOS_ESTUD datos
                WHERE LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), datos.Cedula_Est))) = cne.Cedula_Est
                ORDER BY TRY_CONVERT(bigint, datos.codigo_estud) DESC
            ) detalle
            WHERE (? IS NULL OR (? = 'RH' AND cne.tipo_matricula IN ('R', 'H')) OR cne.tipo_matricula = ?)
              AND (? IS NULL OR cne.estado_codigo = ?)
            ORDER BY cne.Apellidos_nombre, cne.nombre_carrera, cne.tipo_matricula
            """
        )
        try:
            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(query, tipo, tipo, tipo, estado, estado)
                rows = cursor.fetchall()

            items = [
                {
                    "punto_matricula": "CNE",
                    "tipo_matricula": str(row.tipo_matricula or ""),
                    "estado_codigo": str(row.estado_codigo or ""),
                    "estado_nombre": str(row.estado_nombre or ""),
                    "codigo_estud": str(row.codigo_estud or ""),
                    "cedula": str(row.Cedula_Est or ""),
                    "nombre_estudiante": str(row.Apellidos_nombre or ""),
                    "nombre_carrera": str(row.nombre_carrera or ""),
                    "correo_intec": str(row.correo_intec or ""),
                    "correo_personal": str(row.correo_personal or ""),
                    "periodo": "",
                    "detalle_periodo": "Reporte general CNE",
                    "anio_periodo": None,
                }
                for row in rows
            ]
            return {
                "tipo_matricula": tipo or "ALL",
                "estado_codigo": estado,
                "anio_periodo": None,
                "punto_matricula": "CNE",
                "total": len(items),
                "items": items,
                "fuente": "dbo.TOTALESTUDMATRICCNE",
                "consultado_en": datetime.now(timezone.utc).isoformat(),
            }
        except pyodbc.Error as exc:
            raise HTTPException(status_code=500, detail=f"Error consultando listado CNE de matricula: {exc}") from exc

    source_cte = _MATRICULA_ACTUAL_CTE
    base_row_filter = ""

    query = (
        source_cte
        + """
        , movement_rows AS (
            SELECT
                tipo_matricula,
                estado_codigo,
                estado_nombre,
                codigo_estud,
                Cedula_Est,
                Apellidos_nombre,
                nombre_carrera,
                correo_intec_datos,
                correo_personal_datos,
                codigo_periodo,
                Detalle_Periodo,
                anio_periodo,
                Fecha_Ingreso,
                fechaMatricula,
                'PRIMERA' AS punto_matricula
            FROM base_cruce
            {base_row_filter}
        )
        SELECT
            punto_matricula,
            tipo_matricula,
            estado_codigo,
            estado_nombre,
            codigo_estud,
            Cedula_Est,
            Apellidos_nombre,
            nombre_carrera,
            correo_intec_datos,
            correo_personal_datos,
            codigo_periodo,
            Detalle_Periodo,
            anio_periodo,
            Fecha_Ingreso,
            fechaMatricula
        FROM movement_rows
        WHERE 1 = 1
          {punto_filter}
          AND (? IS NULL OR (? = 'RH' AND tipo_matricula IN ('R', 'H')) OR tipo_matricula = ?)
          AND (? IS NULL OR estado_codigo = ?)
          AND (? IS NULL OR anio_periodo = ?)
        ORDER BY Apellidos_nombre
        """.format(punto_filter=punto_filter, base_row_filter=base_row_filter)
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(query, (tipo, tipo, tipo, estado, estado, anio_periodo, anio_periodo))
            rows = cursor.fetchall()

        reference = _read_matriz_desagregada_reference()
        filtered_rows = rows
        items: list[dict[str, Any]] = []
        for row in filtered_rows[:limit]:
            items.append(
                {
                    "punto_matricula": row.punto_matricula,
                    "tipo_matricula": row.tipo_matricula,
                    "estado_codigo": row.estado_codigo,
                    "estado_nombre": row.estado_nombre,
                    "codigo_estud": str(row.codigo_estud) if row.codigo_estud is not None else "",
                    "cedula": row.Cedula_Est,
                    "nombre_estudiante": row.Apellidos_nombre,
                    "nombre_carrera": row.nombre_carrera,
                    "correo_intec": row.correo_intec_datos,
                    "correo_personal": row.correo_personal_datos,
                    "periodo": str(row.codigo_periodo) if row.codigo_periodo is not None else "",
                    "detalle_periodo": row.Detalle_Periodo,
                    "anio_periodo": int(row.anio_periodo) if row.anio_periodo is not None else None,
                }
            )

        return {
            "tipo_matricula": tipo or "ALL",
            "estado_codigo": estado,
            "anio_periodo": anio_periodo,
            "punto_matricula": punto_return,
            "total": len(filtered_rows),
            "items": items,
            "referencia_matriz": _matricula_reference_meta(reference, len(rows), len(filtered_rows)),
        }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando listado de matricula: {exc}") from exc


def _bool_from_db(value: Any) -> bool:
    text = _clean_cell(value).upper()
    return text in {"1", "S", "SI", "TRUE", "T", "Y", "YES"}


def _ingreso_ventas_estado_key(value: Any) -> str:
    estado = _clean_cell(value).upper()
    return estado if estado in {"A", "G", "P", "R"} else "SIN_MATRICULA"


def _ingreso_ventas_estado_field(estado: str) -> str:
    return {
        "A": "activos",
        "G": "graduados",
        "P": "inactivos",
        "R": "retirados",
    }.get(estado, "sin_matricula")


def _ingreso_ventas_summary_key(item: dict[str, Any]) -> str:
    return (
        item["usuario_id"]
        or item["codasesor"]
        or item["usuario_preinscripcion"]
        or "SIN_ASESOR"
    )


def _ingreso_ventas_summary_item(item: dict[str, Any]) -> dict[str, Any]:
    usuario_nombre = item["usuario_nombre"] or item["usuario_preinscripcion"] or "Sin asesor"
    return {
        "usuario_key": _ingreso_ventas_summary_key(item),
        "usuario_id": item["usuario_id"],
        "codasesor": item["codasesor"],
        "usuario_login": item["usuario_login"],
        "usuario_nombre": usuario_nombre,
        "usuario_estado": item["usuario_estado"],
        "total_preinscripciones": 0,
        "total_matriculados": 0,
        "sin_matricula": 0,
        "activos": 0,
        "graduados": 0,
        "inactivos": 0,
        "retirados": 0,
        "regular_r": 0,
        "homologacion_h": 0,
        "prematricula": 0,
        "proceso_finalizado": 0,
        "control_ingreso": 0,
    }


def _build_ingreso_ventas_summary(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for item in items:
        key = _ingreso_ventas_summary_key(item)
        current = summary.setdefault(key, _ingreso_ventas_summary_item(item))
        estado = _ingreso_ventas_estado_key(item.get("estado_codigo_matricula"))
        current["total_preinscripciones"] += 1
        if item.get("matricula_validada"):
            current["total_matriculados"] += 1
            if estado != "SIN_MATRICULA":
                current[_ingreso_ventas_estado_field(estado)] += 1
        else:
            current["sin_matricula"] += 1
        tipo = _clean_cell(item.get("tipo_matricula")).upper()
        if tipo == "R" and item.get("matricula_validada"):
            current["regular_r"] += 1
        elif tipo == "H" and item.get("matricula_validada"):
            current["homologacion_h"] += 1
        if item.get("prematricula"):
            current["prematricula"] += 1
        if item.get("proceso_finalizado"):
            current["proceso_finalizado"] += 1
        if item.get("control_ingreso"):
            current["control_ingreso"] += 1

    return sorted(
        summary.values(),
        key=lambda row: (
            -int(row["total_preinscripciones"]),
            _normalize_name(row["usuario_nombre"]),
        ),
    )


def _sql_placeholders(values: list[Any]) -> str:
    return ", ".join("?" for _ in values)


def _numeric_param(value: Any) -> int | str | None:
    text = _clean_cell(value)
    if not text:
        return None
    try:
        numeric = float(text)
    except ValueError:
        return text
    return int(numeric) if numeric.is_integer() else text


def _int_sort_value(value: Any, default: int = 9999) -> int:
    text = _clean_cell(value)
    if not text:
        return default
    try:
        return int(float(text))
    except ValueError:
        return default


def _ingreso_ventas_estado_nombre(estado: Any) -> str:
    return {
        "A": "Activo",
        "G": "Graduado",
        "P": "Inactivo",
        "R": "Retirado",
    }.get(_clean_cell(estado).upper(), _clean_cell(estado))


def _student_estado_sql(alias: str = "d") -> str:
    return f"""
        CASE
            WHEN UPPER(LTRIM(RTRIM(CAST({alias}.Estado AS varchar(50))))) IN ('A', 'ACTIVO', 'ACTIVA') THEN 'A'
            WHEN UPPER(LTRIM(RTRIM(CAST({alias}.Estado AS varchar(50))))) IN ('P', 'I', 'INACTIVO', 'INACTIVA') THEN 'P'
            WHEN UPPER(LTRIM(RTRIM(CAST({alias}.Estado AS varchar(50))))) IN ('R', 'RETIRADO', 'RETIRADA') THEN 'R'
            WHEN UPPER(LTRIM(RTRIM(CAST({alias}.Estado AS varchar(50))))) IN ('G', 'GRADUADO', 'GRADUADA') THEN 'G'
            WHEN UPPER(LTRIM(RTRIM(CAST({alias}.Estado AS varchar(50))))) = 'D' THEN 'D'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (1, 10, 11, 12, 13, 14) THEN 'A'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (2, 20, 21, 22) THEN 'P'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (3, 30, 31, 32) THEN 'R'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (4, 40, 41, 42) THEN 'G'
            ELSE COALESCE(NULLIF(UPPER(LTRIM(RTRIM(CAST({alias}.Estado AS varchar(50))))), ''), 'SIN ESTADO')
        END
    """


def _fetch_ingreso_ventas_pre_rows(cursor: pyodbc.Cursor, top_limit: int) -> list[Any]:
    query = f"""
        SELECT TOP ({top_limit})
            TRY_CONVERT(varchar(50), p.Codestu) AS codestu,
            TRY_CONVERT(varchar(50), p.Cedula) AS cedula_preinscripcion,
            p.Apellidos_nombre AS nombre_preinscripcion,
            p.correo AS correo_preinscripcion,
            p.telefono,
            TRY_CONVERT(varchar(50), p.codperiodo) AS codperiodo_preinscripcion,
            pp.Detalle_Periodo AS periodo_preinscripcion,
            TRY_CONVERT(int, pp.anio) AS anio_preinscripcion,
            TRY_CONVERT(varchar(50), p.codcarrera) AS codcarrera_preinscripcion,
            pc.Nombre_Basica AS carrera_preinscripcion,
            TRY_CONVERT(varchar(50), p.codasesor) AS codasesor,
            p.Usuario AS usuario_preinscripcion,
            p.Fecha_Ingreso AS fecha_preinscripcion,
            p.Prematricula,
            p.ProcesoFinalilzado,
            p.ControlIngreso,
            u.login AS usuario_login,
            u.nombres AS usuario_nombre,
            u.estado AS usuario_estado,
            TRY_CONVERT(varchar(50), u.id_usuarios) AS usuario_id
        FROM [dbo].[PREINSCRIPCION] p
        OUTER APPLY (
            SELECT TOP (1) us.*
            FROM [dbo].[USUARIO_SIS] us
            WHERE TRY_CONVERT(varchar(50), us.id_usuarios) = TRY_CONVERT(varchar(50), p.codasesor)
            ORDER BY us.nombres
        ) u
        LEFT JOIN [dbo].[PERIODO] pp
            ON pp.cod_periodo = p.codperiodo
        OUTER APPLY (
            SELECT TOP (1) car.*
            FROM [dbo].[CARRERAS] car
            WHERE car.Cod_AnioBasica = p.codcarrera
            ORDER BY car.Cod_AnioBasica
        ) pc
        ORDER BY
            COALESCE(NULLIF(LTRIM(RTRIM(TRY_CONVERT(varchar(255), u.nombres))), ''), 'Sin asesor'),
            p.Fecha_Ingreso DESC,
            p.Apellidos_nombre
    """
    cursor.execute(query)
    return cursor.fetchall()


def _fetch_ingreso_ventas_students(cursor: pyodbc.Cursor, pre_rows: list[Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    code_params = [
        value for value in (_numeric_param(getattr(row, "codestu", None)) for row in pre_rows) if value is not None
    ]
    cedula_params = sorted({
        _clean_cell(getattr(row, "cedula_preinscripcion", None))
        for row in pre_rows
        if _clean_cell(getattr(row, "cedula_preinscripcion", None))
    })
    conditions: list[str] = []
    params: list[Any] = []
    if code_params:
        conditions.append(f"d.codigo_estud IN ({_sql_placeholders(code_params)})")
        params.extend(code_params)
    if cedula_params:
        conditions.append(f"LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))) IN ({_sql_placeholders(cedula_params)})")
        params.extend(cedula_params)
    if not conditions:
        return {}, {}

    query = f"""
        SELECT *
        FROM (
            SELECT
                d.codigo_estud,
                d.Cedula_Est,
                d.Apellidos_nombre,
                d.Fecha_Ingreso,
                {_student_estado_sql("d")} AS estado_datos_codigo,
                ROW_NUMBER() OVER (
                    PARTITION BY d.codigo_estud
                    ORDER BY COALESCE(
                        TRY_CONVERT(datetime2, d.Fecha_Ingreso, 121),
                        TRY_CONVERT(datetime2, d.Fecha_Ingreso, 120),
                        TRY_CONVERT(datetime2, d.Fecha_Ingreso, 103),
                        TRY_CONVERT(datetime2, d.Fecha_Ingreso, 105),
                        TRY_CONVERT(datetime2, d.Fecha_Ingreso),
                        CAST('9999-12-31' AS datetime2)
                    ) DESC,
                    COALESCE(TRY_CONVERT(int, d.NumMigracion), 2147483647) DESC
                ) AS rn_data
            FROM [dbo].[DATOS_ESTUD] d
            WHERE {" OR ".join(f"({condition})" for condition in conditions)}
        ) ranked_data
        WHERE rn_data = 1
    """
    cursor.execute(query, params)
    data_by_code: dict[str, Any] = {}
    data_by_cedula: dict[str, Any] = {}
    for row in cursor.fetchall():
        code = _clean_cell(row.codigo_estud)
        cedula = _clean_cell(row.Cedula_Est)
        if code:
            data_by_code[code] = row
        if cedula and cedula not in data_by_cedula:
            data_by_cedula[cedula] = row
    return data_by_code, data_by_cedula


def _fetch_ingreso_ventas_matriculas(cursor: pyodbc.Cursor, student_codes: list[str]) -> dict[str, list[Any]]:
    code_params = [value for value in (_numeric_param(code) for code in student_codes) if value is not None]
    if not code_params:
        return {}
    placeholders = _sql_placeholders(code_params)
    query = f"""
        WITH pensum_materias AS (
            SELECT DISTINCT
                TRY_CONVERT(varchar(50), pe.Cod_AnioBasica) AS cod_anio_basica,
                TRY_CONVERT(varchar(50), pe.codigo_materia) AS codigo_materia
            FROM [dbo].[PENSUM] pe
            WHERE pe.Cod_AnioBasica IS NOT NULL
              AND pe.codigo_materia IS NOT NULL
            UNION
            SELECT DISTINCT
                TRY_CONVERT(varchar(50), pe.Cod_AnioBasica) AS cod_anio_basica,
                TRY_CONVERT(varchar(50), pe.cod_materia) AS codigo_materia
            FROM [dbo].[PENSUM] pe
            WHERE pe.Cod_AnioBasica IS NOT NULL
              AND pe.cod_materia IS NOT NULL
        )
        SELECT
            TRY_CONVERT(varchar(50), cx.codigo_estud) AS codigo_estud,
            cx.cod_anio_Basica,
            ca.Nombre_Basica AS nombre_carrera,
            cx.codigo_periodo,
            per.Detalle_Periodo,
            per.TipoMatricula AS tipo_matricula,
            TRY_CONVERT(int, per.anio) AS anio_periodo,
            cx.Num_Matricula,
            'CARRERAXESTUD' AS origen_matricula
        FROM [dbo].[CARRERAXESTUD] cx
        INNER JOIN [dbo].[CARRERAS] ca
            ON TRY_CONVERT(varchar(50), ca.Cod_AnioBasica) = TRY_CONVERT(varchar(50), cx.cod_anio_Basica)
        INNER JOIN [dbo].[PERIODO] per
            ON TRY_CONVERT(varchar(50), per.cod_periodo) = TRY_CONVERT(varchar(50), cx.codigo_periodo)
        LEFT JOIN pensum_materias pe
            ON pe.cod_anio_basica = TRY_CONVERT(varchar(50), cx.cod_anio_Basica)
           AND pe.codigo_materia = TRY_CONVERT(varchar(50), cx.codigo_materia)
        WHERE cx.codigo_estud IN ({placeholders})
          AND cx.codigo_materia IS NOT NULL
          AND per.TipoMatricula IN ('R', 'H')
    """
    cursor.execute(query, code_params)
    candidates: dict[str, list[Any]] = {}
    for row in cursor.fetchall():
        candidates.setdefault(_clean_cell(row.codigo_estud), []).append(row)
    return candidates


def _fetch_ingreso_ventas_valid_matriculas(cursor: pyodbc.Cursor) -> dict[str, list[Any]]:
    query = (
        _MATRICULA_ACTUAL_CTE
        + """
        SELECT
            codigo_estud,
            Cedula_Est,
            Apellidos_nombre,
            correo_personal_datos,
            correo_intec_datos,
            Fecha_Ingreso,
            estado_codigo,
            estado_nombre,
            cod_anio_Basica,
            nombre_carrera,
            codigo_periodo,
            Detalle_Periodo,
            tipo_matricula,
            anio_periodo,
            Num_Matricula,
            CASE WHEN es_faltante_datos_estud = 1 THEN 'DATOS_ESTUD' ELSE 'CARRERAXESTUD' END AS origen_matricula
        FROM base_cruce
        """
    )
    cursor.execute(query)
    candidates: dict[str, list[Any]] = {}
    for row in cursor.fetchall():
        candidates.setdefault(_clean_cell(row.codigo_estud), []).append(row)
    return candidates


def _select_ingreso_ventas_student(pre_row: Any, data_by_code: dict[str, Any], data_by_cedula: dict[str, Any]) -> Any | None:
    code = _clean_cell(pre_row.codestu)
    cedula = _clean_cell(pre_row.cedula_preinscripcion)
    return data_by_code.get(code) or data_by_cedula.get(cedula)


def _select_ingreso_ventas_matricula(pre_row: Any, candidates: list[Any]) -> Any | None:
    if not candidates:
        return None
    pre_period = _clean_cell(pre_row.codperiodo_preinscripcion)
    pre_career = _clean_cell(pre_row.codcarrera_preinscripcion)

    def sort_key(row: Any) -> tuple[int, int, int, int, int]:
        return (
            0 if _clean_cell(row.codigo_periodo) == pre_period else 1,
            0 if _clean_cell(row.cod_anio_Basica) == pre_career else 1,
            _int_sort_value(row.anio_periodo),
            _int_sort_value(row.codigo_periodo),
            _int_sort_value(row.Num_Matricula, default=2147483647),
        )

    return min(candidates, key=sort_key)


def _select_ingreso_ventas_datos_matricula(candidates: list[Any]) -> Any | None:
    if not candidates:
        return None

    def sort_key(row: Any) -> tuple[int, int, int]:
        return (
            _int_sort_value(row.anio_periodo),
            _int_sort_value(row.codigo_periodo),
            _int_sort_value(row.Num_Matricula, default=2147483647),
        )

    return min(candidates, key=sort_key)


def _fetch_ingreso_ventas_datos_estud_items(
    matriculas_by_code: dict[str, list[Any]],
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for student_code, matriculas in sorted(matriculas_by_code.items()):
        for matricula in matriculas:
            estado_codigo = _clean_cell(matricula.estado_codigo)
            item = {
                "codestu": student_code,
                "cedula_preinscripcion": "",
                "nombre_preinscripcion": "",
                "correo_preinscripcion": _clean_cell(matricula.correo_intec_datos) or _clean_cell(matricula.correo_personal_datos),
                "telefono": "",
                "codperiodo_preinscripcion": "",
                "periodo_preinscripcion": "",
                "anio_preinscripcion": None,
                "codcarrera_preinscripcion": "",
                "carrera_preinscripcion": "",
                "codasesor": "",
                "usuario_preinscripcion": "",
                "fecha_preinscripcion": "",
                "prematricula": False,
                "proceso_finalizado": False,
                "control_ingreso": False,
                "usuario_id": "",
                "usuario_login": "",
                "usuario_nombre": "",
                "usuario_estado": "",
                "codigo_estud_matricula": student_code,
                "existe_datos_estud": True,
                "existe_carreraxestud": True,
                "origen_matricula": _clean_cell(matricula.origen_matricula) or "DATOS_ESTUD_CARRERAXESTUD",
                "cedula_matricula": _clean_cell(matricula.Cedula_Est),
                "nombre_matricula": _clean_cell(matricula.Apellidos_nombre),
                "estado_codigo_matricula": estado_codigo,
                "estado_nombre_matricula": _clean_cell(matricula.estado_nombre) or _ingreso_ventas_estado_nombre(estado_codigo),
                "tipo_matricula": _clean_cell(matricula.tipo_matricula),
                "codcarrera_matricula": _clean_cell(matricula.cod_anio_Basica),
                "carrera_matricula": _clean_cell(matricula.nombre_carrera),
                "periodo_matricula": _clean_cell(matricula.codigo_periodo),
                "detalle_periodo_matricula": _clean_cell(matricula.Detalle_Periodo),
                "anio_periodo_matricula": int(matricula.anio_periodo)
                if matricula.anio_periodo is not None
                else None,
                "matricula_validada": True,
                "estado_cruce": "DATOS_ESTUD_CARRERAXESTUD_RH",
                "nombre_final": _clean_cell(matricula.Apellidos_nombre),
                "cedula_final": _clean_cell(matricula.Cedula_Est),
                "carrera_final": _clean_cell(matricula.nombre_carrera),
                "periodo_final": _clean_cell(matricula.Detalle_Periodo),
                "anio_final": int(matricula.anio_periodo)
                if matricula.anio_periodo is not None
                else None,
            }
            items.append(item)
    return items


def _build_ingreso_ventas_payload(top_limit: int) -> dict[str, Any]:
    with get_connection() as conn:
        cursor = conn.cursor()
        pre_rows = _fetch_ingreso_ventas_pre_rows(cursor, top_limit)
        data_by_code, data_by_cedula = _fetch_ingreso_ventas_students(cursor, pre_rows)
        matriculas_by_code = _fetch_ingreso_ventas_valid_matriculas(cursor)
        datos_estud_items = _fetch_ingreso_ventas_datos_estud_items(matriculas_by_code)

    items: list[dict[str, Any]] = []
    for row in pre_rows:
        student = _select_ingreso_ventas_student(row, data_by_code, data_by_cedula)
        student_code = _clean_cell(student.codigo_estud) if student is not None else _clean_cell(row.codestu)
        matricula = _select_ingreso_ventas_matricula(row, matriculas_by_code.get(student_code, []))
        estado_codigo = _clean_cell(student.estado_datos_codigo) if student is not None else ""
        origen_matricula = _clean_cell(matricula.origen_matricula) if matricula is not None else ""
        existe_carreraxestud = origen_matricula == "CARRERAXESTUD"
        matricula_validada = bool(student is not None and existe_carreraxestud)
        item = {
            "codestu": _clean_cell(row.codestu),
            "cedula_preinscripcion": _clean_cell(row.cedula_preinscripcion),
            "nombre_preinscripcion": _clean_cell(row.nombre_preinscripcion),
            "correo_preinscripcion": _clean_cell(row.correo_preinscripcion),
            "telefono": _clean_cell(row.telefono),
            "codperiodo_preinscripcion": _clean_cell(row.codperiodo_preinscripcion),
            "periodo_preinscripcion": _clean_cell(row.periodo_preinscripcion),
            "anio_preinscripcion": int(row.anio_preinscripcion) if row.anio_preinscripcion is not None else None,
            "codcarrera_preinscripcion": _clean_cell(row.codcarrera_preinscripcion),
            "carrera_preinscripcion": _clean_cell(row.carrera_preinscripcion),
            "codasesor": _clean_cell(row.codasesor),
            "usuario_preinscripcion": _clean_cell(row.usuario_preinscripcion),
            "fecha_preinscripcion": _clean_cell(row.fecha_preinscripcion),
            "prematricula": _bool_from_db(row.Prematricula),
            "proceso_finalizado": _bool_from_db(row.ProcesoFinalilzado),
            "control_ingreso": _bool_from_db(row.ControlIngreso),
            "usuario_id": _clean_cell(row.usuario_id),
            "usuario_login": _clean_cell(row.usuario_login),
            "usuario_nombre": _clean_cell(row.usuario_nombre),
            "usuario_estado": _clean_cell(row.usuario_estado),
            "codigo_estud_matricula": student_code if student is not None or matricula is not None else "",
            "existe_datos_estud": student is not None,
            "existe_carreraxestud": existe_carreraxestud,
            "origen_matricula": origen_matricula,
            "cedula_matricula": _clean_cell(student.Cedula_Est) if student is not None else "",
            "nombre_matricula": _clean_cell(student.Apellidos_nombre) if student is not None else "",
            "estado_codigo_matricula": estado_codigo,
            "estado_nombre_matricula": _ingreso_ventas_estado_nombre(estado_codigo),
            "tipo_matricula": _clean_cell(matricula.tipo_matricula)
            if matricula is not None
            else "",
            "codcarrera_matricula": _clean_cell(matricula.cod_anio_Basica) if matricula is not None else "",
            "carrera_matricula": _clean_cell(matricula.nombre_carrera) if matricula is not None else "",
            "periodo_matricula": _clean_cell(matricula.codigo_periodo) if matricula is not None else "",
            "detalle_periodo_matricula": _clean_cell(matricula.Detalle_Periodo) if matricula is not None else "",
            "anio_periodo_matricula": int(matricula.anio_periodo)
            if matricula is not None and matricula.anio_periodo is not None
            else None,
        }
        item["matricula_validada"] = matricula_validada
        if matricula_validada:
            item["estado_cruce"] = "DATOS_ESTUD_CARRERAXESTUD"
        elif student is not None:
            item["estado_cruce"] = "SOLO_DATOS_ESTUD"
        else:
            item["estado_cruce"] = "SIN_MATRICULA"
        item["nombre_final"] = item["nombre_matricula"] or item["nombre_preinscripcion"]
        item["cedula_final"] = item["cedula_matricula"] or item["cedula_preinscripcion"]
        item["carrera_final"] = item["carrera_matricula"] or item["carrera_preinscripcion"]
        item["periodo_final"] = item["detalle_periodo_matricula"] or item["periodo_preinscripcion"]
        item["anio_final"] = item["anio_periodo_matricula"] or item["anio_preinscripcion"]
        if _ingreso_ventas_estado_key(item.get("estado_codigo_matricula")) == "D":
            continue
        items.append(item)

    summary = _build_ingreso_ventas_summary(items)
    totals = {
        "total_preinscripciones": len(items),
        "total_matriculados": sum(1 for item in datos_estud_items if item["matricula_validada"]),
        "sin_matricula": sum(1 for item in datos_estud_items if not item["matricula_validada"]),
        "asesores": len(summary),
        "total_datos_estud": len(datos_estud_items),
        "total_carreraxestud": sum(1 for item in datos_estud_items if item.get("existe_carreraxestud")),
        "total_base_porcentaje": sum(1 for item in datos_estud_items if item["matricula_validada"]),
    }
    for estado in ("A", "G", "P", "R"):
        totals[_ingreso_ventas_estado_field(estado)] = sum(
            1
            for item in datos_estud_items
            if item["matricula_validada"] and _ingreso_ventas_estado_key(item.get("estado_codigo_matricula")) == estado
        )
    totals["regular_r"] = sum(
        1
        for item in datos_estud_items
        if item["matricula_validada"] and _clean_cell(item.get("tipo_matricula")).upper() == "R"
    )
    totals["homologacion_h"] = sum(
        1
        for item in datos_estud_items
        if item["matricula_validada"] and _clean_cell(item.get("tipo_matricula")).upper() == "H"
    )
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total": len(items),
        "totals": totals,
        "summary": summary,
        "items": items,
        "datos_estud_items": datos_estud_items,
        "criteria": {
            "fuente": "PREINSCRIPCION + USUARIO_SIS para ventas; base real desde DATOS_ESTUD, clasificada con CARRERAXESTUD cuando existe y enriquecida con PENSUM",
            "join_usuario": "PREINSCRIPCION.codasesor -> USUARIO_SIS.id_usuarios",
            "join_estudiante": "PREINSCRIPCION.Codestu -> DATOS_ESTUD.codigo_estud; respaldo por Cedula",
        },
    }


@router.get(
    "/ingreso-ventas",
    responses={400: {"description": "Solicitud invalida"}, 500: {"description": "Error interno del servidor"}},
)
def ingreso_ventas(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    limit: Annotated[int, Query(ge=1, le=20000)] = 5000,
) -> dict[str, Any]:
    del current_user
    top_limit = max(1, min(int(limit or 5000), 20000))
    try:
        return _build_ingreso_ventas_payload(top_limit)
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"Error consultando ingreso por ventas: {exc}") from exc


def _parse_graduation_date(value: Any) -> date | None:
    text = _clean_cell(value)
    if not text:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"Fecha de grado invalida: {text}") from exc


def _parse_optional_excel_date(value: Any, label: str) -> date | None:
    text = _clean_cell(value)
    if not text:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(text[:10]).date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{label} invalida: {text}") from exc


def _ensure_graduation_extra_columns(cursor: pyodbc.Cursor) -> None:
    cursor.execute(
        """
        IF COL_LENGTH('dbo.DATOS_ESTUD', 'Fecha_Emision_SENESCYT') IS NULL
        BEGIN
            ALTER TABLE dbo.DATOS_ESTUD ADD Fecha_Emision_SENESCYT date NULL
        END
        IF COL_LENGTH('dbo.DATOS_ESTUD', 'Cod_Refrendacion') IS NULL
        BEGIN
            ALTER TABLE dbo.DATOS_ESTUD ADD Cod_Refrendacion varchar(50) NULL
        END
        ELSE
        BEGIN
            DECLARE @currentLength int;
            SELECT @currentLength = c.max_length
            FROM sys.columns c
            WHERE c.object_id = OBJECT_ID('dbo.DATOS_ESTUD')
              AND c.name = 'Cod_Refrendacion';
            IF ISNULL(@currentLength, 0) < 50
            BEGIN
                ALTER TABLE dbo.DATOS_ESTUD ALTER COLUMN Cod_Refrendacion varchar(50) NULL
            END
        END
        """
    )


def _style_excel_header(worksheet: Any) -> None:
    fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _safe_excel_filename(value: Any) -> str:
    text = re.sub(r'[<>:"/\\|?*]+', "", _clean_cell(value))
    text = re.sub(r"\s+", "_", text).strip("._")
    return text[:80] or "fecha_grado"


def _graduation_status_sql(alias: str = "d") -> str:
    return f"""
        CASE
            WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), {alias}.Estado)))) IN ('A', 'ACTIVO', 'ACTIVA') THEN 'A'
            WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), {alias}.Estado)))) IN ('P', 'I', 'INACTIVO', 'INACTIVA') THEN 'P'
            WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), {alias}.Estado)))) IN ('R', 'RETIRADO', 'RETIRADA') THEN 'R'
            WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), {alias}.Estado)))) IN ('E', 'EGRESADO', 'EGRESADA') THEN 'E'
            WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), {alias}.Estado)))) IN ('G', 'GRADUADO', 'GRADUADA') THEN 'G'
            WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), {alias}.Estado)))) = 'D' THEN 'D'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (1, 10, 11, 12, 13, 14) THEN 'A'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (2, 20, 21, 22) THEN 'P'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (3, 30, 31, 32) THEN 'R'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (5, 50, 51, 52) THEN 'E'
            WHEN TRY_CONVERT(int, {alias}.Estado) IN (4, 40, 41, 42) THEN 'G'
            ELSE COALESCE(NULLIF(UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(50), {alias}.Estado)))), ''), 'SIN ESTADO')
        END
    """


def _graduation_status_name(code: Any) -> str:
    text = _clean_cell(code).upper()
    return {
        "A": "Activo",
        "P": "Inactivo",
        "R": "Retirado",
        "E": "Egresado",
        "G": "Graduado",
        "D": "Educación Continua",
        "SIN ESTADO": "Sin estado",
    }.get(text, text or "Sin estado")


def _graduation_date_student_rows(
    period_code: str,
    career_code: str = "",
    search: str = "",
    limit: int = 5000,
) -> list[dict[str, Any]]:
    params: list[Any] = [period_code]
    wheres = ["TRY_CONVERT(varchar(50), cx.codigo_periodo) = ?"]
    if career_code:
        wheres.append("TRY_CONVERT(varchar(50), cx.cod_anio_Basica) = ?")
        params.append(career_code)
    if search:
        needle = f"%{search}%"
        document_needle = re.sub(r"\D+", "", search) or search
        wheres.append(
            """
            (
                TRY_CONVERT(nvarchar(max), d.Apellidos_nombre) LIKE ?
                OR TRY_CONVERT(varchar(50), d.codigo_estud) LIKE ?
                OR REPLACE(REPLACE(TRY_CONVERT(varchar(50), d.Cedula_Est), '-', ''), ' ', '') LIKE ?
                OR TRY_CONVERT(nvarchar(max), c.Nombre_Basica) LIKE ?
            )
            """
        )
        params.extend([needle, needle, f"%{document_needle}%", needle])
    with get_connection() as conn:
        cursor = conn.cursor()
        _ensure_graduation_extra_columns(cursor)
        cursor.execute(
            f"""
            SELECT TOP ({limit})
                TRY_CONVERT(varchar(50), d.codigo_estud) AS codigo_estud,
                LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))) AS cedula,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), d.Apellidos_nombre))) AS nombres,
                LTRIM(RTRIM(TRY_CONVERT(varchar(50), cx.cod_anio_Basica))) AS codigo_carrera,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), c.Nombre_Basica))) AS carrera,
                TRY_CONVERT(varchar(50), cx.codigo_periodo) AS codigo_periodo,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), p.Detalle_Periodo))) AS periodo,
                d.Fecha_Grado AS fecha_grado,
                d.Fecha_Emision_SENESCYT AS fecha_emision_senescyt,
                LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cod_Refrendacion))) AS cod_refrendacion
            FROM dbo.CARRERAXESTUD cx
            INNER JOIN dbo.DATOS_ESTUD d
                ON TRY_CONVERT(varchar(50), d.codigo_estud) = TRY_CONVERT(varchar(50), cx.codigo_estud)
            LEFT JOIN dbo.CARRERAS c
                ON TRY_CONVERT(varchar(50), c.Cod_AnioBasica) = TRY_CONVERT(varchar(50), cx.cod_anio_Basica)
            LEFT JOIN dbo.PERIODO p
                ON TRY_CONVERT(varchar(50), p.cod_periodo) = TRY_CONVERT(varchar(50), cx.codigo_periodo)
            WHERE {" AND ".join(wheres)}
            GROUP BY
                d.codigo_estud, d.Cedula_Est, d.Apellidos_nombre,
                cx.cod_anio_Basica, c.Nombre_Basica, cx.codigo_periodo, p.Detalle_Periodo,
                d.Fecha_Grado, d.Fecha_Emision_SENESCYT, d.Cod_Refrendacion
            ORDER BY d.Apellidos_nombre, c.Nombre_Basica
            """,
            *params,
        )
        return [
            {
                "codigo_estud": _clean_cell(row.codigo_estud),
                "cedula": _clean_cell(row.cedula),
                "nombres": _clean_cell(row.nombres),
                "codigo_carrera": _clean_cell(row.codigo_carrera),
                "carrera": _clean_cell(row.carrera),
                "codigo_periodo": _clean_cell(row.codigo_periodo),
                "periodo": _clean_cell(row.periodo),
                "fecha_grado": row.fecha_grado.isoformat() if row.fecha_grado else "",
                "fecha_emision_senescyt": row.fecha_emision_senescyt.isoformat() if row.fecha_emision_senescyt else "",
                "cod_refrendacion": _clean_cell(row.cod_refrendacion),
            }
            for row in cursor.fetchall()
        ]


@router.get("/fecha-grado/catalog")
def graduation_date_catalog(
    current_user: Annotated[SessionUser, Depends(_GRADUATION_ACCESS)],
    periodo: Annotated[str, Query(max_length=50)] = "",
) -> dict[str, Any]:
    del current_user
    period_code = _clean_cell(periodo)
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT TOP (500)
                    TRY_CONVERT(varchar(50), p.cod_periodo) AS codigo_periodo,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), p.Detalle_Periodo))) AS detalle_periodo,
                    p.fechain,
                    p.fechafin,
                    p.anio
                FROM dbo.PERIODO p
                WHERE EXISTS (
                    SELECT 1
                    FROM dbo.CARRERAXESTUD cx
                    WHERE TRY_CONVERT(varchar(50), cx.codigo_periodo) = TRY_CONVERT(varchar(50), p.cod_periodo)
                )
                ORDER BY COALESCE(p.anio, 0) DESC, p.fechain DESC, p.cod_periodo DESC
                """
            )
            periods = [
                {
                    "codigo_periodo": _clean_cell(row.codigo_periodo),
                    "detalle_periodo": _clean_cell(row.detalle_periodo),
                    "fecha_inicio": row.fechain.isoformat() if row.fechain else "",
                    "fecha_fin": row.fechafin.isoformat() if row.fechafin else "",
                    "anio": int(row.anio) if row.anio is not None else None,
                }
                for row in cursor.fetchall()
            ]
            careers: list[dict[str, Any]] = []
            if period_code:
                cursor.execute(
                    """
                    SELECT
                        TRY_CONVERT(varchar(50), c.Cod_AnioBasica) AS codigo_carrera,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), c.Nombre_Basica))) AS nombre_carrera,
                        COUNT(DISTINCT cx.codigo_estud) AS total_estudiantes
                    FROM dbo.CARRERAXESTUD cx
                    LEFT JOIN dbo.CARRERAS c
                        ON TRY_CONVERT(varchar(50), c.Cod_AnioBasica) = TRY_CONVERT(varchar(50), cx.cod_anio_Basica)
                    WHERE TRY_CONVERT(varchar(50), cx.codigo_periodo) = ?
                    GROUP BY c.Cod_AnioBasica, c.Nombre_Basica
                    ORDER BY c.Nombre_Basica
                    """,
                    period_code,
                )
                careers = [
                    {
                        "codigo_carrera": _clean_cell(row.codigo_carrera),
                        "nombre_carrera": _clean_cell(row.nombre_carrera) or "Sin carrera",
                        "total_estudiantes": int(row.total_estudiantes or 0),
                    }
                    for row in cursor.fetchall()
                ]
            return {"periodos": periods, "carreras": careers}
    except HTTPException:
        raise
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo cargar el catalogo de fecha de grado: {exc}") from exc


@router.get("/fecha-grado/estudiantes")
def graduation_date_students(
    current_user: Annotated[SessionUser, Depends(_GRADUATION_ACCESS)],
    periodo: Annotated[str, Query(min_length=1, max_length=50)],
    carrera: Annotated[str, Query(max_length=50)] = "",
    busqueda: Annotated[str, Query(max_length=120)] = "",
    limit: Annotated[int, Query(ge=1, le=5000)] = 1000,
) -> dict[str, Any]:
    del current_user
    period_code = _clean_cell(periodo)
    career_code = _clean_cell(carrera)
    search = _clean_cell(busqueda)
    try:
        items = _graduation_date_student_rows(period_code, career_code, search, limit)
        return {"items": items, "total": len(items), "periodo": period_code, "carrera": career_code}
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo consultar estudiantes para fecha de grado: {exc}") from exc


@router.get("/fecha-grado/verificacion")
def graduation_date_verification(
    current_user: Annotated[SessionUser, Depends(_GRADUATION_ACCESS)],
    estado: Annotated[str, Query(max_length=50)] = "",
    page: Annotated[int, Query(ge=1, le=100000)] = 1,
    page_size: Annotated[int, Query(ge=10, le=200)] = 25,
) -> dict[str, Any]:
    status_code = _clean_cell(estado).upper()
    secretary_only = current_user.rol == "SECRETARIA"
    allowed_secretary_statuses = {"E", "G"}
    if secretary_only and status_code and status_code not in allowed_secretary_statuses:
        status_code = "G"
    offset = (page - 1) * page_size
    status_expr = _graduation_status_sql("d")
    where = ""
    params: list[Any] = []
    if secretary_only:
        where = "WHERE estado_codigo IN ('E', 'G')"
        if status_code:
            where += " AND estado_codigo = ?"
            params.append(status_code)
    elif status_code:
        where = "WHERE estado_codigo = ?"
        params.append(status_code)
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_graduation_extra_columns(cursor)
            cursor.execute(
                f"""
                WITH base AS (
                    SELECT
                        TRY_CONVERT(varchar(50), d.codigo_estud) AS codigo_estud,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))) AS cedula,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), d.Apellidos_nombre))) AS nombres,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Estado))) AS estado_raw,
                        {status_expr} AS estado_codigo,
                        d.Fecha_Grado AS fecha_grado,
                        d.Fecha_Emision_SENESCYT AS fecha_emision_senescyt,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cod_Refrendacion))) AS cod_refrendacion
                    FROM dbo.DATOS_ESTUD d
                )
                SELECT COUNT(1) AS total,
                    SUM(CASE WHEN fecha_grado IS NULL THEN 0 ELSE 1 END) AS con_fecha
                FROM base
                {where}
                """,
                *params,
            )
            summary = cursor.fetchone()
            total = int(summary.total or 0) if summary else 0
            con_fecha = int(summary.con_fecha or 0) if summary else 0

            cursor.execute(
                f"""
                WITH base AS (
                    SELECT
                        TRY_CONVERT(varchar(50), d.codigo_estud) AS codigo_estud,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cedula_Est))) AS cedula,
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), d.Apellidos_nombre))) AS nombres,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Estado))) AS estado_raw,
                        {status_expr} AS estado_codigo,
                        d.Fecha_Grado AS fecha_grado,
                        d.Fecha_Emision_SENESCYT AS fecha_emision_senescyt,
                        LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.Cod_Refrendacion))) AS cod_refrendacion
                    FROM dbo.DATOS_ESTUD d
                )
                SELECT codigo_estud, cedula, nombres, estado_raw, estado_codigo, fecha_grado, fecha_emision_senescyt, cod_refrendacion
                FROM base
                {where}
                ORDER BY nombres, codigo_estud
                OFFSET ? ROWS FETCH NEXT ? ROWS ONLY
                """,
                *params,
                offset,
                page_size,
            )
            items = [
                {
                    "codigo_estud": _clean_cell(row.codigo_estud),
                    "cedula": _clean_cell(row.cedula),
                    "nombres": _clean_cell(row.nombres),
                    "estado_codigo": _clean_cell(row.estado_codigo),
                    "estado_nombre": _graduation_status_name(row.estado_codigo),
                    "estado_raw": _clean_cell(row.estado_raw),
                    "fecha_grado": row.fecha_grado.isoformat() if row.fecha_grado else "",
                    "fecha_emision_senescyt": row.fecha_emision_senescyt.isoformat() if row.fecha_emision_senescyt else "",
                    "cod_refrendacion": _clean_cell(row.cod_refrendacion),
                }
                for row in cursor.fetchall()
            ]
            return {
                "items": items,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": max(1, math.ceil(total / page_size)) if total else 1,
                "con_fecha": con_fecha,
                "sin_fecha": max(total - con_fecha, 0),
                "estado": status_code,
            }
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo verificar fecha de grado: {exc}") from exc


@router.get("/fecha-grado/plantilla")
def graduation_date_template(
    current_user: Annotated[SessionUser, Depends(_GRADUATION_ACCESS)],
) -> StreamingResponse:
    del current_user
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fecha de grado"
    headers = ["cedula", "fecha_grado", "fecha_emision_senescyt", "cod_refrendacion"]
    sheet.append(headers)
    sheet["A1"].comment = Comment("Numero de cedula del estudiante. Se valida directamente contra DATOS_ESTUD.", "INTEC")
    sheet["B1"].comment = Comment("Fecha de grado en formato AAAA-MM-DD. Ejemplo: 2026-06-30.", "INTEC")
    sheet["C1"].comment = Comment("Fecha de emision SENESCYT en formato AAAA-MM-DD. Ejemplo: 2026-07-15.", "INTEC")
    sheet["D1"].comment = Comment("Codigo de refrendacion. Maximo 50 caracteres.", "INTEC")
    _style_excel_header(sheet)
    sheet.freeze_panes = "A2"
    widths = [18, 18, 26, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    max_row = 1000
    date_validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(1900,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Fecha no valida",
        error="Ingresa una fecha valida en formato AAAA-MM-DD.",
        promptTitle="Fecha de grado",
        prompt="Formato requerido: AAAA-MM-DD. Ejemplo: 2026-06-30.",
    )
    sheet.add_data_validation(date_validation)
    date_validation.add(f"B2:B{max_row}")
    date_validation.add(f"C2:C{max_row}")
    for row in range(2, max_row + 1):
        sheet[f"A{row}"].number_format = "@"
        sheet[f"B{row}"].number_format = "yyyy-mm-dd"
        sheet[f"C{row}"].number_format = "yyyy-mm-dd"
        sheet[f"D{row}"].number_format = "@"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = "plantilla_fecha_grado_datos_estud.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/fecha-grado/importar")
async def import_graduation_dates(
    current_user: Annotated[SessionUser, Depends(_GRADUATION_ACCESS)],
    file: UploadFile = File(...),
) -> dict[str, Any]:
    del current_user
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Carga un archivo Excel .xlsx")
    content = await file.read()
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel") from exc
    sheet = workbook.active
    headers = [_clean_cell(cell.value).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        cedula_index = headers.index("cedula")
        fecha_index = headers.index("fecha_grado")
        fecha_senescyt_index = headers.index("fecha_emision_senescyt")
        refrendacion_index = headers.index("cod_refrendacion")
    except ValueError as exc:
        raise HTTPException(
            status_code=400,
            detail="La plantilla debe incluir columnas cedula, fecha_grado, fecha_emision_senescyt y cod_refrendacion",
        ) from exc

    updates: list[tuple[date | None, date | None, str, str, int]] = []
    errors: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        cedula = re.sub(r"\D+", "", _clean_cell(row[cedula_index] if cedula_index < len(row) else ""))
        raw_date = row[fecha_index] if fecha_index < len(row) else None
        raw_senescyt_date = row[fecha_senescyt_index] if fecha_senescyt_index < len(row) else None
        raw_refrendacion = _clean_cell(row[refrendacion_index] if refrendacion_index < len(row) else "")[:50]
        if not cedula and not _clean_cell(raw_date) and not _clean_cell(raw_senescyt_date) and not raw_refrendacion:
            continue
        if not cedula:
            errors.append({"fila": row_number, "cedula": "", "error": "Cédula vacía"})
            continue
        if cedula in seen:
            errors.append({"fila": row_number, "cedula": cedula, "error": "Cédula duplicada en el Excel"})
            continue
        seen.add(cedula)
        try:
            parsed_date = _parse_graduation_date(raw_date)
            parsed_senescyt_date = _parse_optional_excel_date(raw_senescyt_date, "Fecha de emision SENESCYT")
        except HTTPException as exc:
            errors.append({"fila": row_number, "cedula": cedula, "error": exc.detail})
            continue
        updates.append((parsed_date, parsed_senescyt_date, raw_refrendacion, cedula, row_number))
    if errors:
        return {
            "ok": False,
            "actualizados": 0,
            "errores": errors,
            "procesados": len(updates),
            "no_encontrados": [],
            "resumen": "El Excel contiene errores de formato o duplicados.",
        }
    if not updates:
        raise HTTPException(status_code=400, detail="No se encontraron filas válidas para actualizar")
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_graduation_extra_columns(cursor)
            updated = 0
            not_found: list[dict[str, Any]] = []
            updated_rows: list[dict[str, Any]] = []
            for graduation_date, senescyt_date, refrendacion, cedula, row_number in updates:
                cursor.execute(
                    """
                    UPDATE dbo.DATOS_ESTUD
                    SET Fecha_Grado = ?,
                        Fecha_Emision_SENESCYT = ?,
                        Cod_Refrendacion = ?
                    WHERE REPLACE(REPLACE(TRY_CONVERT(varchar(50), Cedula_Est), '-', ''), ' ', '') = ?
                    """,
                    graduation_date,
                    senescyt_date,
                    refrendacion,
                    cedula,
                )
                rowcount = cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else 0
                if rowcount:
                    updated += rowcount
                    updated_rows.append({
                        "fila": row_number,
                        "cedula": cedula,
                        "fecha_grado": graduation_date.isoformat() if graduation_date else "",
                        "fecha_emision_senescyt": senescyt_date.isoformat() if senescyt_date else "",
                        "cod_refrendacion": refrendacion,
                        "registros": rowcount,
                    })
                else:
                    not_found.append({
                        "fila": row_number,
                        "cedula": cedula,
                        "error": "Cédula no encontrada en DATOS_ESTUD",
                    })
            conn.commit()
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo importar fecha de grado: {exc}") from exc
    return {
        "ok": True,
        "actualizados": updated,
        "procesados": len(updates),
        "errores": [],
        "no_encontrados": not_found,
        "actualizados_detalle": updated_rows,
        "resumen": f"Procesados: {len(updates)}. Actualizados: {updated}. No encontrados: {len(not_found)}.",
    }


@router.post("/fecha-grado/guardar")
def save_graduation_dates(
    payload: GraduationDateSavePayload,
    current_user: Annotated[SessionUser, Depends(_GRADUATION_ACCESS)],
) -> dict[str, Any]:
    del current_user
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_graduation_extra_columns(cursor)
            updated = 0
            for item in payload.items:
                code = _clean_cell(item.codigo_estud)
                graduation_date = _parse_graduation_date(item.fecha_grado)
                senescyt_date = _parse_optional_excel_date(item.fecha_emision_senescyt, "Fecha de emision SENESCYT")
                refrendacion = _clean_cell(item.cod_refrendacion)[:50]
                cursor.execute(
                    """
                    UPDATE dbo.DATOS_ESTUD
                    SET Fecha_Grado = ?,
                        Fecha_Emision_SENESCYT = ?,
                        Cod_Refrendacion = ?
                    WHERE TRY_CONVERT(varchar(50), codigo_estud) = ?
                    """,
                    graduation_date,
                    senescyt_date,
                    refrendacion,
                    code,
                )
                updated += cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else 0
            conn.commit()
            return {"ok": True, "actualizados": updated}
    except HTTPException:
        raise
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudieron guardar las fechas de grado: {exc}") from exc


_LEGACY_STUDENT_DATA_FIELDS = [
    "tipodocumento",
    "Cedula_Est",
    "Apellidos_nombre",
    "Sexo",
    "generoId",
    "EstadoCivil",
    "Etnia",
    "Nacionalidad",
    "tiposangre",
    "discapacidad",
    "Porce_Capacidad",
    "No_Carnet",
    "Tipo_Capacidad",
    "Fecha_Nac",
    "paisNacionalidadId",
    "provinciaNacimeintoId",
    "cantonNacimeintoId",
    "paisResidenciaId",
    "codprov",
    "Canton",
    "tipoColegioId",
    "ModalidadEstudio",
    "Jornada",
    "Fecha_Ingreso",
    "fechaMatricula",
    "tipoMatriculaId",
    "nivelAcademicoQueCursa",
    "duracionPeriodoAcademico",
    "haRepetidoAlMenosUnaMateria",
    "Paralelo",
    "haPerdidoLaGratuidad",
    "recibePensionDiferenciada",
    "Ocupacion",
    "ingresoEstudianteId",
    "bonoDesarrolloId",
    "haRealizadoPracticasPreprofesionales",
    "nroHorasPracticasPreprofesionales",
    "entornoInstitucionalPracticasProfesionales",
    "sectorEconomicoPracticaProfesional",
    "tipoBecaId",
    "primeraRazonBecaId",
    "segundaRazonBecaId",
    "terceraRazonBecaId",
    "cuartaRazonBecaId",
    "quintaRazonBecaId",
    "sextaRazonBecaId",
    "montoBeca",
    "porcientoBecaCoberturaArancel",
    "porcientoBecaCoberturaManuntencion",
    "financiamientoBeca",
    "montoAyudaEconomica",
    "montoCreditoEducativo",
    "participaEnProyectoVinculacionSociedad",
    "tipoAlcanceProyectoVinculacionId",
    "correo",
    "movil",
    "nivelFormacionPadre",
    "nivelFormacionMadre",
    "IngresoHogar",
    "Numpersonasvive",
    "ciudad",
    "calle_principal",
    "referencia",
    "NumHogar",
    "correointec",
]

_LEGACY_TEACHER_DATA_FIELDS = [
    "tipoDocumentoId",
    "cedula_doc",
    "apellidos_nombre",
    "sexo",
    "generoId",
    "estado_civil",
    "etniaId",
    "nacionalidad",
    "Direccion",
    "provinciaSufragio",
    "movil",
    "correop",
    "numDomicilio",
    "discapacidad",
    "porcen_discapa",
    "tipo_discapa",
    "carnet_conadis",
    "tipoEnfermedadCatastrofica",
    "fecha_nac",
    "paisNacionalidadId",
    "nivelFormacion",
    "fechaIngresoIES",
    "fechaSalidaIES",
    "relacionLaboralIESId",
    "ingresoConCursoMeritos",
    "escalafonDocenteId",
    "cargoDirectivoId",
    "tiempoDedicacionId",
    "nombreUnidadAcademica",
    "nroasignaturasdocente",
    "nroHorasLaborablesSemanaEnCarreraPrograma",
    "nroHorasClaseSemanaCarreraPrograma",
    "nroHorasInvestigacionSemanaCarreraPrograma",
    "nroHorasAdministrativasSemanaCarreraPrograma",
    "nroHorasOtrasActividadesSemanaCarreraPrograma",
    "nroHorasVinculacionSociedad",
    "salarioMensual",
    "docenciaTecnicoSuperior",
    "docenciaTecnologico",
    "estaEnPeriodoSabatico",
    "fechaInicioPeriodoSabatico",
    "estaCursandoEstudiosId",
    "institucionDOndeCursaEstudios",
    "paisEstudiosId",
    "tituloAObtener",
    "poseeBecaId",
    "tipoBecaId",
    "montoBeca",
    "financiamientoBecaId",
    "pubRevistasCienInIndexadasId",
    "numPubRevistasCientifIndexadas",
    "docenciaTecnologicoUniversitario",
    "docenciaEspecializacionTecnologica",
    "docenciaMaestriaTecnologica",
    "tiposangre",
    "correo",
]


def _quote_sql_name(name: str) -> str:
    return f"[{name.replace(']', ']]')}]"


def _table_columns(cursor: pyodbc.Cursor, table_name: str) -> set[str]:
    cursor.execute(
        """
        SELECT COLUMN_NAME
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_SCHEMA = 'dbo'
          AND TABLE_NAME = ?
        """,
        table_name,
    )
    return {_clean_cell(row.COLUMN_NAME) for row in cursor.fetchall()}


def _row_to_dict(row: Any, columns: list[str]) -> dict[str, Any]:
    return {column: getattr(row, column, None) for column in columns}


def _data_update_stats(fields: dict[str, Any], editable_columns: list[str]) -> tuple[int, int, list[str]]:
    missing = [field for field in editable_columns if not _clean_cell(fields.get(field))]
    total = len(editable_columns)
    filled = max(total - len(missing), 0)
    return filled, len(missing), missing


def _student_data_summary(row: Any, editable_columns: list[str]) -> dict[str, Any]:
    fields = _row_to_dict(row, editable_columns)
    filled, pending, missing = _data_update_stats(fields, editable_columns)
    total = len(editable_columns) or 1
    return {
        "id": _clean_cell(getattr(row, "codigo_estud", "")),
        "codigo": _clean_cell(getattr(row, "codigo_estud", "")),
        "nombre": _clean_cell(getattr(row, "Apellidos_nombre", "")),
        "cedula": _clean_cell(getattr(row, "Cedula_Est", "")),
        "tipo": "estudiante",
        "carrera": _clean_cell(getattr(row, "Nombre_Basica", "")),
        "correo": _clean_cell(getattr(row, "correo", "")),
        "campos_llenos": filled,
        "campos_pendientes": pending,
        "campos_totales": total,
        "porcentaje_lleno": round((filled / total) * 100, 2),
        "campos_faltantes": missing[:20],
    }


def _teacher_data_summary(row: Any, editable_columns: list[str]) -> dict[str, Any]:
    fields = _row_to_dict(row, editable_columns)
    filled, pending, missing = _data_update_stats(fields, editable_columns)
    total = len(editable_columns) or 1
    return {
        "id": _clean_cell(getattr(row, "codigo_doc", "")),
        "codigo": _clean_cell(getattr(row, "codigo_doc", "")),
        "nombre": _clean_cell(getattr(row, "apellidos_nombre", "")),
        "cedula": _clean_cell(getattr(row, "cedula_doc", "")),
        "tipo": "docente",
        "carrera": _clean_cell(getattr(row, "unidad_academica", "")),
        "correo": _clean_cell(getattr(row, "correo", "")) or _clean_cell(getattr(row, "correop", "")),
        "campos_llenos": filled,
        "campos_pendientes": pending,
        "campos_totales": total,
        "porcentaje_lleno": round((filled / total) * 100, 2),
        "campos_faltantes": missing[:20],
    }


def _actualizacion_datos_columns(cursor: pyodbc.Cursor, target: str) -> list[str]:
    table = "DATOS_ESTUD" if target == "estudiantes" else "DATOSDOCENTE"
    allowed = _LEGACY_STUDENT_DATA_FIELDS if target == "estudiantes" else _LEGACY_TEACHER_DATA_FIELDS
    existing = _table_columns(cursor, table)
    return [field for field in allowed if field in existing]


_LEGACY_DATA_CATALOG_TABLES: dict[str, list[str]] = {
    "tipodocumento": ["TIPODOCUMENTOS", "TipoDocumento", "TipoDocumentoIdentidad", "TipoDocumentoSenescyt"],
    "tipoDocumentoId": ["TIPODOCUMENTOS", "TipoDocumento", "TipoDocumentoIdentidad", "TipoDocumentoSenescyt"],
    "Sexo": ["Sexo"],
    "sexo": ["Sexo"],
    "generoId": ["Sexo"],
    "EstadoCivil": ["EstadoCivil"],
    "estado_civil": ["EstadoCivil"],
    "Etnia": ["Etnia"],
    "etniaId": ["Etnia"],
    "Nacionalidad": ["PuebloNacionalidad", "Pais"],
    "nacionalidad": ["PuebloNacionalidad", "Pais"],
    "paisNacionalidadId": ["Pais"],
    "paisResidenciaId": ["Pais"],
    "paisEstudiosId": ["Pais"],
    "codprov": ["Provincias"],
    "provinciaNacimeintoId": ["Provincias"],
    "provinciaSufragio": ["Provincias"],
    "Canton": ["Canton"],
    "cantonNacimeintoId": ["Canton"],
    "tiposangre": ["TipoSangre"],
    "tipoColegioId": ["TipoColegio"],
    "ModalidadEstudio": ["ModalidadCarrera", "ModalidadMatricula"],
    "Jornada": ["JornadaCarrera"],
    "tipoMatriculaId": ["TipoMatricula"],
    "nivelAcademicoQueCursa": ["NivelQueCursa"],
    "duracionPeriodoAcademico": ["DuracionPeriodoAcademico"],
    "haRepetidoAlMenosUnaMateria": ["Repeticion"],
    "Paralelo": ["Paralelo"],
    "haPerdidoLaGratuidad": ["PerdidaGratuidad"],
    "recibePensionDiferenciada": ["PensionDiferida"],
    "Ocupacion": ["Ocupacion"],
    "ingresoEstudianteId": ["IngresosEstudiante"],
    "bonoDesarrolloId": ["BonoDesarrollo"],
    "discapacidad": ["Discapacidad"],
    "Porce_Capacidad": ["PorcentajeDiscapacidad"],
    "porcen_discapa": ["PorcentajeDiscapacidad"],
    "Tipo_Capacidad": ["TipoDiscapacidad"],
    "tipo_discapa": ["TipoDiscapacidad"],
    "haRealizadoPracticasPreprofesionales": ["RealizoPracticasPreprofesionales"],
    "nroHorasPracticasPreprofesionales": ["NroHorasPracticasProfesionales"],
    "entornoInstitucionalPracticasProfesionales": ["TipoInstitucionPracticaProfesional"],
    "sectorEconomicoPracticaProfesional": ["SectorEconomico"],
    "tipoBecaId": ["TipoBeca", "Becas"],
    "poseeBecaId": ["Becas", "TipoBeca"],
    "primeraRazonBecaId": ["PrimeraRazonBeca"],
    "segundaRazonBecaId": ["SegundaRazonBeca"],
    "terceraRazonBecaId": ["TerceraRazonBeca"],
    "cuartaRazonBecaId": ["CuartaRazonBeca"],
    "quintaRazonBecaId": ["QuintaRazonBeca"],
    "sextaRazonBecaId": ["SextaRazonBeca"],
    "porcientoBecaCoberturaArancel": ["PorcentajeBecaCoberturaArancel"],
    "porcientoBecaCoberturaManuntencion": ["PorcentajeBecaCoberturaManutencion"],
    "financiamientoBeca": ["TipoFinanciacion"],
    "financiamientoBecaId": ["TipoFinanciacion"],
    "participaEnProyectoVinculacionSociedad": ["HaParticipadoPracticasVinculacion"],
    "tipoAlcanceProyectoVinculacionId": ["AlcanceProyectoVinculacion"],
    "nivelFormacionPadre": ["NivelFormacionPadre"],
    "nivelFormacionMadre": ["NivelFormacionMadre"],
    "nivelFormacion": ["NivelFormacion", "NivelFormacionDocente"],
    "relacionLaboralIESId": ["RelacionLaboralIES", "RelacionLaboral"],
    "ingresoConCursoMeritos": ["IngresoConCursoMeritos", "SiNo"],
    "escalafonDocenteId": ["EscalafonDocente"],
    "cargoDirectivoId": ["CargoDirectivo"],
    "tiempoDedicacionId": ["TiempoDedicacion"],
    "tipoEnfermedadCatastrofica": ["TipoEnfermedadCatastrofica", "EnfermedadCatastrofica"],
    "estaEnPeriodoSabatico": ["SiNo"],
    "estaCursandoEstudiosId": ["SiNo"],
    "pubRevistasCienInIndexadasId": ["SiNo"],
}


def _catalog_options_from_pairs(pairs: list[tuple[str, str]]) -> list[dict[str, str]]:
    return [{"value": value, "label": label} for value, label in pairs]


def _numeric_catalog_options(values: list[int | str]) -> list[dict[str, str]]:
    return [{"value": str(value), "label": str(value)} for value in values]


_YES_NO_OPTIONS = [("1", "Si"), ("2", "No")]
_BLOOD_TYPE_OPTIONS = [
    ("1", "A +"),
    ("2", "A -"),
    ("3", "B +"),
    ("4", "B -"),
    ("5", "AB +"),
    ("6", "AB -"),
    ("7", "O +"),
    ("8", "O -"),
]


_LEGACY_STATIC_CATALOGS_BY_TARGET: dict[str, dict[str, list[dict[str, str]]]] = {
    "estudiantes": {
        "tiposangre": _catalog_options_from_pairs(_BLOOD_TYPE_OPTIONS),
    },
    "docentes": {
        "tiposangre": _catalog_options_from_pairs(_BLOOD_TYPE_OPTIONS),
        "tipoEnfermedadCatastrofica": _catalog_options_from_pairs([
            ("1", "Cancer"),
            ("2", "Tumor cerebral"),
            ("3", "Quemaduras graves"),
            ("4", "Insuficiencia renal"),
            ("5", "Otros"),
            ("6", "No aplica"),
        ]),
        "nivelFormacion": _catalog_options_from_pairs([
            ("1", "Tercer Nivel Tecnico"),
            ("2", "Tercer Nivel Tecnologico"),
            ("3", "Tercer Nivel de Grado"),
            ("4", "Especialidad"),
            ("5", "Especialidad medica u odontologica"),
            ("6", "Maestria"),
            ("7", "PhD"),
            ("8", "Tercer Nivel Tecnologico Universitario"),
            ("9", "Especializacion Tecnologica"),
            ("10", "Maestria Tecnologica"),
        ]),
        "relacionLaboralIESId": _catalog_options_from_pairs([
            ("1", "Contrato con relacion de dependencia"),
            ("2", "Contrato sin relacion de dependencia"),
            ("3", "Nombramiento provisional"),
            ("4", "Nombramiento definitivo"),
            ("5", "Comision de servicios"),
        ]),
        "ingresoConCursoMeritos": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "escalafonDocenteId": _catalog_options_from_pairs([
            ("1", "Titular principal"),
            ("2", "Titular agregado"),
            ("3", "Titular auxiliar"),
            ("4", "Ocasional"),
            ("5", "Honorario"),
            ("6", "Invitado"),
        ]),
        "cargoDirectivoId": _catalog_options_from_pairs([
            ("1", "Rector"),
            ("2", "Vicerrector"),
            ("3", "Secretario"),
            ("4", "Tesorero"),
            ("5", "(en blanco)"),
            ("6", "Otro"),
            ("7", "No aplica"),
        ]),
        "tiempoDedicacionId": _catalog_options_from_pairs([
            ("1", "Exclusiva o completa"),
            ("2", "Semi exclusiva o a medio tiempo"),
            ("3", "Tiempo parcial"),
        ]),
        "docenciaTecnicoSuperior": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "docenciaTecnologico": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "estaEnPeriodoSabatico": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "estaCursandoEstudiosId": _catalog_options_from_pairs([
            ("1", "Nivel Tecnico"),
            ("2", "Nivel Tecnologico"),
            ("3", "Tercer Nivel"),
            ("4", "Especialidad"),
            ("5", "Especialidad medica u odontologica"),
            ("6", "Maestria"),
            ("7", "PhD"),
            ("8", "No aplica"),
            ("9", "Tercer Nivel Tecnologico Universitario"),
            ("10", "Especializacion Tecnologica"),
            ("11", "Maestria Tecnologica"),
        ]),
        "poseeBecaId": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "tipoBecaId": _catalog_options_from_pairs([
            ("1", "Total"),
            ("2", "Parcial"),
            ("3", "No aplica"),
        ]),
        "financiamientoBecaId": _catalog_options_from_pairs([
            ("1", "IES"),
            ("2", "Senescyt"),
            ("3", "Otro"),
            ("4", "Transferencia del Estado"),
            ("5", "No aplica"),
        ]),
        "pubRevistasCienInIndexadasId": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "docenciaTecnologicoUniversitario": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "docenciaEspecializacionTecnologica": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "docenciaMaestriaTecnologica": _catalog_options_from_pairs(_YES_NO_OPTIONS),
        "nroasignaturasdocente": _numeric_catalog_options(list(range(0, 7))),
        "nroHorasLaborablesSemanaEnCarreraPrograma": _numeric_catalog_options(list(range(0, 13)) + [40]),
        "nroHorasClaseSemanaCarreraPrograma": _numeric_catalog_options(list(range(0, 13))),
        "nroHorasInvestigacionSemanaCarreraPrograma": _numeric_catalog_options(list(range(0, 13))),
        "nroHorasAdministrativasSemanaCarreraPrograma": _numeric_catalog_options([0, 15, 20, 25, 30, 40]),
        "nroHorasOtrasActividadesSemanaCarreraPrograma": _numeric_catalog_options(list(range(0, 13))),
        "nroHorasVinculacionSociedad": _numeric_catalog_options(list(range(0, 13))),
    }
}


def _pick_catalog_column(columns: set[str], preferred: tuple[str, ...], contains: tuple[str, ...]) -> str | None:
    by_lower = {column.lower(): column for column in columns}
    for candidate in preferred:
        column = by_lower.get(candidate.lower())
        if column:
            return column
    for column in sorted(columns):
        lower = column.lower()
        if any(token in lower for token in contains):
            return column
    return None


def _catalog_options_from_table(cursor: pyodbc.Cursor, table_name: str) -> list[dict[str, str]]:
    try:
        columns = _table_columns(cursor, table_name)
    except pyodbc.Error:
        return []
    if not columns:
        return []
    value_column = _pick_catalog_column(
        columns,
        (
            f"codigo_{table_name}",
            f"id_{table_name}",
            "codigo",
            "id",
            "cod",
            "Cod_Provincia",
            "Cod_Pais",
            "NumDoc",
            "NumM",
            "codigo_canton",
            "id_sexo",
        ),
        ("codigo", "cod_", "id_"),
    )
    label_column = _pick_catalog_column(
        columns,
        (
            f"nombre_{table_name}",
            "nombre",
            "descripcion",
            "detalle",
            "Descripcion_Pais",
            "Descripcion_Prov",
            "TipoDoc",
            "DetalleM",
            "nombre_canton",
            "detalle_sexo",
        ),
        ("nombre", "descripcion", "detalle"),
    )
    if not value_column or not label_column:
        return []
    active_column = next((column for column in columns if column.lower() == "activo"), None)
    active_filter = ""
    if active_column:
        active_filter = f"AND ISNULL(TRY_CONVERT(int, {_quote_sql_name(active_column)}), 1) = 1"
    try:
        cursor.execute(
            f"""
            SELECT TOP (500)
                TRY_CONVERT(nvarchar(100), {_quote_sql_name(value_column)}) AS option_value,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), {_quote_sql_name(label_column)}))) AS option_label
            FROM dbo.{_quote_sql_name(table_name)}
            WHERE NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), {_quote_sql_name(label_column)}))), N'') IS NOT NULL
              {active_filter}
            GROUP BY {_quote_sql_name(value_column)}, {_quote_sql_name(label_column)}
            ORDER BY option_label
            """
        )
        return [
            {
                "value": _clean_cell(row.option_value),
                "label": _clean_cell(row.option_label),
            }
            for row in cursor.fetchall()
            if _clean_cell(row.option_value) and _clean_cell(row.option_label)
        ]
    except pyodbc.Error:
        return []


def _distinct_options_from_data_table(cursor: pyodbc.Cursor, source_table: str, field: str) -> list[dict[str, str]]:
    try:
        if field not in _table_columns(cursor, source_table):
            return []
        cursor.execute(
            f"""
            SELECT TOP (300)
                TRY_CONVERT(nvarchar(250), {_quote_sql_name(field)}) AS option_value
            FROM dbo.{_quote_sql_name(source_table)}
            WHERE NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), {_quote_sql_name(field)}))), N'') IS NOT NULL
            GROUP BY TRY_CONVERT(nvarchar(250), {_quote_sql_name(field)})
            ORDER BY option_value
            """
        )
        return [
            {"value": _clean_cell(row.option_value), "label": _clean_cell(row.option_value)}
            for row in cursor.fetchall()
            if _clean_cell(row.option_value)
        ]
    except pyodbc.Error:
        return []


_STUDENT_NUMERIC_TEXT_FIELDS = {
    "Cedula_Est",
    "No_Carnet",
    "NumHogar",
    "Numpersonasvive",
    "IngresoHogar",
    "movil",
}
_STUDENT_DATE_FIELDS = {"Fecha_Nac", "Fecha_Ingreso", "fechaMatricula"}
_STUDENT_BECAS_DEFAULTS = {
    "tipoBecaId": "3",
    "primeraRazonBecaId": "2",
    "segundaRazonBecaId": "2",
    "terceraRazonBecaId": "2",
    "cuartaRazonBecaId": "2",
    "quintaRazonBecaId": "2",
    "sextaRazonBecaId": "2",
    "montoBeca": "0",
    "porcientoBecaCoberturaArancel": "NA",
    "porcientoBecaCoberturaManuntencion": "NA",
    "financiamientoBeca": "4",
}


_TEACHER_NUMERIC_TEXT_FIELDS = {
    "cedula_doc",
    "movil",
    "numDomicilio",
    "carnet_conadis",
    "salarioMensual",
    "montoBeca",
    "numPubRevistasCientifIndexadas",
    "nroasignaturasdocente",
    "nroHorasLaborablesSemanaEnCarreraPrograma",
    "nroHorasClaseSemanaCarreraPrograma",
    "nroHorasInvestigacionSemanaCarreraPrograma",
    "nroHorasAdministrativasSemanaCarreraPrograma",
    "nroHorasOtrasActividadesSemanaCarreraPrograma",
    "nroHorasVinculacionSociedad",
}
_TEACHER_DATE_FIELDS = {
    "fecha_nac",
    "fechaIngresoIES",
    "fechaSalidaIES",
    "fechaInicioPeriodoSabatico",
}
_TEACHER_ZERO_DEFAULT_FIELDS = {
    "nroasignaturasdocente",
    "nroHorasLaborablesSemanaEnCarreraPrograma",
    "nroHorasClaseSemanaCarreraPrograma",
    "nroHorasInvestigacionSemanaCarreraPrograma",
    "nroHorasAdministrativasSemanaCarreraPrograma",
    "nroHorasOtrasActividadesSemanaCarreraPrograma",
    "nroHorasVinculacionSociedad",
    "numPubRevistasCientifIndexadas",
    "montoBeca",
}


def _sanitize_student_data_field(field: str, value: Any) -> str:
    text = _clean_cell(value)
    if not text:
        return ""
    if field in {"correo", "correointec"}:
        return text if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", text) else ""
    if field in _STUDENT_NUMERIC_TEXT_FIELDS:
        return re.sub(r"[^0-9+]", "", text)
    if field in _STUDENT_DATE_FIELDS:
        return re.sub(r"[^0-9-]", "", text)[:10]
    return re.sub(r"<[^>]*>", "", text).strip()


def _sanitize_teacher_data_field(field: str, value: Any) -> str:
    text = _clean_cell(value)
    if not text:
        return ""
    if field in {"correo", "correop"}:
        return text if re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", text) else ""
    if field in _TEACHER_NUMERIC_TEXT_FIELDS:
        return re.sub(r"[^0-9.,+-]", "", text)
    if field in _TEACHER_DATE_FIELDS:
        return re.sub(r"[^0-9-]", "", text)[:10]
    return re.sub(r"<[^>]*>", "", text).strip()


def _fetch_single_value(cursor: pyodbc.Cursor, sql: str, *params: Any) -> str:
    try:
        cursor.execute(sql, *params)
        row = cursor.fetchone()
        if not row:
            return ""
        return _clean_cell(row[0])
    except pyodbc.Error:
        return ""


def _fetch_student_beca_fields(cursor: pyodbc.Cursor, codigo_estud: str) -> dict[str, str]:
    code = _clean_cell(codigo_estud)
    if not code:
        return {}
    try:
        beca_columns = _table_columns(cursor, "Becas")
    except pyodbc.Error:
        return {}
    if not beca_columns:
        return {}
    code_columns = [column for column in ("codestud", "codigo_estud", "codigoEstud") if column in beca_columns]
    for code_column in code_columns:
        try:
            cursor.execute(
                f"""
                SELECT TOP 1
                    TRY_CONVERT(nvarchar(100), tipo_beca_recibe_id) AS tipoBecaId,
                    TRY_CONVERT(nvarchar(100), primera_razon_beca_id) AS primeraRazonBecaId,
                    TRY_CONVERT(nvarchar(100), segunda_razon_beca_id) AS segundaRazonBecaId,
                    TRY_CONVERT(nvarchar(100), tercera_razon_beca_id) AS terceraRazonBecaId,
                    TRY_CONVERT(nvarchar(100), cuarta_razon_beca_id) AS cuartaRazonBecaId,
                    TRY_CONVERT(nvarchar(100), quinta_razon_beca_id) AS quintaRazonBecaId,
                    TRY_CONVERT(nvarchar(100), sexta_razon_beca_id) AS sextaRazonBecaId,
                    TRY_CONVERT(nvarchar(100), valor_monto_beca) AS montoBeca,
                    TRY_CONVERT(nvarchar(100), porcentaje_beca) AS porcientoBecaCoberturaArancel,
                    N'NA' AS porcientoBecaCoberturaManuntencion,
                    TRY_CONVERT(nvarchar(100), tipo_financiamiento_beca_id) AS financiamientoBeca
                FROM dbo.Becas
                WHERE TRY_CONVERT(nvarchar(100), {_quote_sql_name(code_column)}) = ?
                ORDER BY id DESC
                """,
                code,
            )
            row = cursor.fetchone()
            if row:
                return {
                    key: _clean_cell(getattr(row, key, "")) or fallback
                    for key, fallback in _STUDENT_BECAS_DEFAULTS.items()
                }
        except pyodbc.Error:
            continue
    return {}


def _latest_student_cabecera(cursor: pyodbc.Cursor, cedula: str) -> Any | None:
    try:
        cursor.execute(
            """
            SELECT TOP 1
                cm.codmodalidad,
                cm.codjornada,
                cm.ControlMatricula
            FROM dbo.CABECERA_MATRICULA cm
            INNER JOIN dbo.DATOS_ESTUD de
                ON TRY_CONVERT(nvarchar(100), cm.codigo_estud) = TRY_CONVERT(nvarchar(100), de.codigo_estud)
            WHERE REPLACE(REPLACE(TRY_CONVERT(nvarchar(50), de.Cedula_Est), '-', ''), ' ', '') = ?
            ORDER BY COALESCE(cm.fecha_pago, TRY_CONVERT(datetime, cm.Num_Matricula), '1900-01-01') DESC
            """,
            re.sub(r"\D+", "", cedula),
        )
        return cursor.fetchone()
    except pyodbc.Error:
        return None


def _resolve_catalog_value_by_label(cursor: pyodbc.Cursor, table_name: str, label: str) -> str:
    label_clean = _clean_cell(label).upper()
    if not label_clean:
        return ""
    for option in _catalog_options_from_table(cursor, table_name):
        if option["label"].upper() == label_clean:
            return option["value"]
    return ""


def _apply_student_legacy_rules(
    cursor: pyodbc.Cursor,
    fields: dict[str, Any],
    *,
    codigo_estud: str,
    cedula: str,
    editable_columns: list[str],
) -> dict[str, str]:
    result = {field: _sanitize_student_data_field(field, fields.get(field)) for field in editable_columns}
    code = _clean_cell(codigo_estud)
    document = _clean_cell(cedula) or result.get("Cedula_Est", "")

    for field, default in _STUDENT_BECAS_DEFAULTS.items():
        if field in result and not result.get(field):
            result[field] = default
    beca_fields = _fetch_student_beca_fields(cursor, code)
    for field, default in _STUDENT_BECAS_DEFAULTS.items():
        if field in result:
            result[field] = beca_fields.get(field) or result.get(field) or default

    if "porcientoBecaCoberturaManuntencion" in result:
        result["porcientoBecaCoberturaManuntencion"] = "NA"
    for field in ("montoAyudaEconomica", "montoCreditoEducativo"):
        if field in result and result.get(field) in {"", "NA"}:
            result[field] = "0"
    if "NumHogar" in result and not result.get("NumHogar"):
        result["NumHogar"] = "0"

    if "correointec" in result and code:
        correo_intec = _fetch_single_value(
            cursor,
            """
            SELECT TOP 1 CorreoIntec
            FROM dbo.CorreosEstudIntec
            WHERE TRY_CONVERT(nvarchar(100), codestud) = ?
              AND NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(150), CorreoIntec))), N'') IS NOT NULL
            ORDER BY fecha DESC
            """,
            code,
        )
        if correo_intec:
            result["correointec"] = correo_intec if "@" in correo_intec else f"{correo_intec}@intec.edu.ec"

    latest_enrollment = _fetch_single_value(
        cursor,
        """
        SELECT TOP 1 CONVERT(varchar(10), cx.Fecha_Matricula, 23)
        FROM dbo.CARRERAXESTUD cx
        INNER JOIN dbo.DATOS_ESTUD de
            ON TRY_CONVERT(nvarchar(100), cx.codigo_estud) = TRY_CONVERT(nvarchar(100), de.codigo_estud)
        WHERE REPLACE(REPLACE(TRY_CONVERT(nvarchar(50), de.Cedula_Est), '-', ''), ' ', '') = ?
          AND cx.Fecha_Matricula IS NOT NULL
        ORDER BY cx.Fecha_Matricula DESC
        """,
        re.sub(r"\D+", "", document),
    )
    first_enrollment = _fetch_single_value(
        cursor,
        """
        SELECT TOP 1 CONVERT(varchar(10), cx.Fecha_Matricula, 23)
        FROM dbo.CARRERAXESTUD cx
        INNER JOIN dbo.DATOS_ESTUD de
            ON TRY_CONVERT(nvarchar(100), cx.codigo_estud) = TRY_CONVERT(nvarchar(100), de.codigo_estud)
        WHERE REPLACE(REPLACE(TRY_CONVERT(nvarchar(50), de.Cedula_Est), '-', ''), ' ', '') = ?
          AND cx.Fecha_Matricula IS NOT NULL
        ORDER BY cx.Fecha_Matricula ASC
        """,
        re.sub(r"\D+", "", document),
    )
    if latest_enrollment and "fechaMatricula" in result:
        result["fechaMatricula"] = latest_enrollment
    if first_enrollment and "Fecha_Ingreso" in result:
        result["Fecha_Ingreso"] = first_enrollment

    if result.get("Sexo") in {"1", "2"} and "generoId" in result:
        result["generoId"] = result["Sexo"]

    cabecera = _latest_student_cabecera(cursor, document)
    if cabecera:
        modalidad_map = {"1": "5", "3": "1"}
        jornada_map = {"1": "1", "2": "3"}
        cod_modalidad = modalidad_map.get(_clean_cell(getattr(cabecera, "codmodalidad", "")), _clean_cell(getattr(cabecera, "codmodalidad", "")))
        cod_jornada = jornada_map.get(_clean_cell(getattr(cabecera, "codjornada", "")), _clean_cell(getattr(cabecera, "codjornada", "")))
        control_matricula = _clean_cell(getattr(cabecera, "ControlMatricula", ""))
        if "ModalidadEstudio" in result and cod_modalidad:
            result["ModalidadEstudio"] = cod_modalidad
        if "Jornada" in result and cod_jornada:
            result["Jornada"] = cod_jornada
        if "tipoMatriculaId" in result and control_matricula in {"1", "2", "3"}:
            result["tipoMatriculaId"] = control_matricula
    if "Jornada" in result and not result.get("Jornada"):
        result["Jornada"] = "3"
    if "tipoMatriculaId" in result and not result.get("tipoMatriculaId"):
        result["tipoMatriculaId"] = "1"

    paralelo_label = _fetch_single_value(
        cursor,
        """
        SELECT TOP 1 LTRIM(RTRIM(TRY_CONVERT(nvarchar(20), cx.Paralelo)))
        FROM dbo.CARRERAXESTUD cx
        INNER JOIN dbo.DATOS_ESTUD de
            ON TRY_CONVERT(nvarchar(100), cx.codigo_estud) = TRY_CONVERT(nvarchar(100), de.codigo_estud)
        WHERE REPLACE(REPLACE(TRY_CONVERT(nvarchar(50), de.Cedula_Est), '-', ''), ' ', '') = ?
          AND NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(20), cx.Paralelo))), N'') IS NOT NULL
        ORDER BY cx.Fecha_Matricula DESC
        """,
        re.sub(r"\D+", "", document),
    )
    if "Paralelo" in result and paralelo_label:
        result["Paralelo"] = _resolve_catalog_value_by_label(cursor, "Paralelo", paralelo_label) or result.get("Paralelo", "")

    if result.get("haRealizadoPracticasPreprofesionales") == "2":
        result.update({
            "nroHorasPracticasPreprofesionales": "NA",
            "entornoInstitucionalPracticasProfesionales": "5",
            "sectorEconomicoPracticaProfesional": "22",
        })
    if result.get("discapacidad") == "2":
        result.update({"Porce_Capacidad": "0", "No_Carnet": "NA", "Tipo_Capacidad": "7"})
    if result.get("participaEnProyectoVinculacionSociedad") != "1" and "tipoAlcanceProyectoVinculacionId" in result:
        result["tipoAlcanceProyectoVinculacionId"] = "5"
    return {field: value for field, value in result.items() if field in editable_columns}


def _apply_teacher_legacy_rules(
    fields: dict[str, Any],
    *,
    editable_columns: list[str],
) -> dict[str, str]:
    result = {field: _sanitize_teacher_data_field(field, fields.get(field)) for field in editable_columns}

    if result.get("sexo") in {"1", "2"} and "generoId" in result:
        result["generoId"] = result["sexo"]

    if result.get("discapacidad") == "2":
        if "tipo_discapa" in result:
            result["tipo_discapa"] = "7"
        if "porcen_discapa" in result:
            result["porcen_discapa"] = "NA"

    if result.get("estaEnPeriodoSabatico") == "2" and "fechaInicioPeriodoSabatico" in result:
        result["fechaInicioPeriodoSabatico"] = "0001-01-01"

    if result.get("estaCursandoEstudiosId") == "8":
        if "poseeBecaId" in result:
            result["poseeBecaId"] = "2"
        if "tipoBecaId" in result:
            result["tipoBecaId"] = "3"
        if "financiamientoBecaId" in result:
            result["financiamientoBecaId"] = "5"
        if "paisEstudiosId" in result:
            result["paisEstudiosId"] = "NA"
        if "montoBeca" in result and not result.get("montoBeca"):
            result["montoBeca"] = "0"

    if result.get("poseeBecaId") == "2":
        if "tipoBecaId" in result:
            result["tipoBecaId"] = "3"
        if "financiamientoBecaId" in result:
            result["financiamientoBecaId"] = "5"
        if "montoBeca" in result and not result.get("montoBeca"):
            result["montoBeca"] = "0"

    if result.get("pubRevistasCienInIndexadasId") == "2" and "numPubRevistasCientifIndexadas" in result:
        result["numPubRevistasCientifIndexadas"] = "0"

    for field in _TEACHER_ZERO_DEFAULT_FIELDS:
        if field in result and result.get(field) == "":
            result[field] = "0"

    return {field: value for field, value in result.items() if field in editable_columns}


def _legacy_data_update_catalogs(cursor: pyodbc.Cursor, columns: list[str], target: str) -> dict[str, list[dict[str, str]]]:
    catalogs: dict[str, list[dict[str, str]]] = {}
    source_table = "DATOS_ESTUD" if target == "estudiantes" else "DATOSDOCENTE"
    for field in columns:
        options = list(_LEGACY_STATIC_CATALOGS_BY_TARGET.get(target, {}).get(field, []))
        if not options:
            for table_name in _LEGACY_DATA_CATALOG_TABLES.get(field, []):
                options = _catalog_options_from_table(cursor, table_name)
                if options:
                    break
        if not options and field in _LEGACY_DATA_CATALOG_TABLES:
            options = _distinct_options_from_data_table(cursor, source_table, field)
        if options:
            catalogs[field] = options
    return catalogs


@router.get("/actualizacion-datos/{target}/buscar")
def search_legacy_data_update_records(
    target: str,
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    q: str = Query(default="", max_length=120),
    limit: int = Query(default=60, ge=1, le=200),
) -> dict[str, Any]:
    del current_user
    if target not in {"estudiantes", "docentes"}:
        raise HTTPException(status_code=404, detail="Tipo de actualización no soportado")
    query = _clean_cell(q)
    like = f"%{query}%"
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            editable_columns = _actualizacion_datos_columns(cursor, target)
            if target == "estudiantes":
                cursor.execute(
                    f"""
                    SELECT TOP ({limit})
                        d.codigo_estud,
                        d.Cedula_Est,
                        d.Apellidos_nombre,
                        d.correo,
                        c.Nombre_Basica,
                        {", ".join("d." + _quote_sql_name(column) for column in editable_columns)}
                    FROM dbo.DATOS_ESTUD d
                    OUTER APPLY (
                        SELECT TOP 1 ca.Nombre_Basica
                        FROM dbo.CABECERA_MATRICULA cm
                        LEFT JOIN dbo.CARRERAS ca
                          ON TRY_CONVERT(varchar(50), ca.Cod_AnioBasica) = TRY_CONVERT(varchar(50), cm.cod_anio_Basica)
                        WHERE TRY_CONVERT(varchar(50), cm.codigo_estud) = TRY_CONVERT(varchar(50), d.codigo_estud)
                        ORDER BY TRY_CONVERT(int, cm.codigo_periodo) DESC, cm.fecha_pago DESC
                    ) c
                    WHERE
                        NULLIF(?, '') IS NULL
                        OR TRY_CONVERT(varchar(50), d.codigo_estud) LIKE ?
                        OR TRY_CONVERT(varchar(50), d.Cedula_Est) LIKE ?
                        OR d.Apellidos_nombre LIKE ?
                        OR d.correo LIKE ?
                    ORDER BY d.Apellidos_nombre
                    """,
                    query,
                    like,
                    like,
                    like,
                    like,
                )
                rows = [_student_data_summary(row, editable_columns) for row in cursor.fetchall()]
            else:
                cursor.execute(
                    f"""
                    SELECT TOP ({limit})
                        d.codigo_doc,
                        d.cedula_doc,
                        d.apellidos_nombre,
                        d.correo,
                        d.correop,
                        d.nombreUnidadAcademica AS unidad_academica,
                        {", ".join("d." + _quote_sql_name(column) for column in editable_columns)}
                    FROM dbo.DATOSDOCENTE d
                    WHERE
                        NULLIF(?, '') IS NULL
                        OR TRY_CONVERT(varchar(50), d.codigo_doc) LIKE ?
                        OR TRY_CONVERT(varchar(50), d.cedula_doc) LIKE ?
                        OR d.apellidos_nombre LIKE ?
                        OR d.correo LIKE ?
                        OR d.correop LIKE ?
                    ORDER BY d.apellidos_nombre
                    """,
                    query,
                    like,
                    like,
                    like,
                    like,
                    like,
                )
                rows = [_teacher_data_summary(row, editable_columns) for row in cursor.fetchall()]
            return {"rows": rows, "total": len(rows), "limit": limit, "query": query, "target": target}
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo buscar datos para actualización: {exc}") from exc


def _load_legacy_data_update_record(cursor: pyodbc.Cursor, target: str, record_id: str) -> dict[str, Any]:
    editable_columns = _actualizacion_datos_columns(cursor, target)
    if not editable_columns:
        raise HTTPException(status_code=500, detail="No hay columnas compatibles para actualizar")
    if target == "estudiantes":
        cursor.execute(
            f"""
            SELECT TOP 1
                d.codigo_estud,
                d.Cedula_Est,
                d.Apellidos_nombre,
                d.correo,
                c.Nombre_Basica,
                {", ".join("d." + _quote_sql_name(column) for column in editable_columns)}
            FROM dbo.DATOS_ESTUD d
            OUTER APPLY (
                SELECT TOP 1 ca.Nombre_Basica
                FROM dbo.CABECERA_MATRICULA cm
                LEFT JOIN dbo.CARRERAS ca
                  ON TRY_CONVERT(varchar(50), ca.Cod_AnioBasica) = TRY_CONVERT(varchar(50), cm.cod_anio_Basica)
                WHERE TRY_CONVERT(varchar(50), cm.codigo_estud) = TRY_CONVERT(varchar(50), d.codigo_estud)
                ORDER BY TRY_CONVERT(int, cm.codigo_periodo) DESC, cm.fecha_pago DESC
            ) c
            WHERE TRY_CONVERT(varchar(50), d.codigo_estud) = ?
               OR REPLACE(REPLACE(TRY_CONVERT(varchar(50), d.Cedula_Est), '-', ''), ' ', '') = ?
            """,
            record_id,
            re.sub(r"\D+", "", record_id),
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Estudiante no encontrado")
        summary = _student_data_summary(row, editable_columns)
    else:
        cursor.execute(
            f"""
            SELECT TOP 1
                d.codigo_doc,
                d.cedula_doc,
                d.apellidos_nombre,
                d.correo,
                d.correop,
                d.nombreUnidadAcademica AS unidad_academica,
                {", ".join("d." + _quote_sql_name(column) for column in editable_columns)}
            FROM dbo.DATOSDOCENTE d
            WHERE TRY_CONVERT(varchar(50), d.codigo_doc) = ?
               OR REPLACE(REPLACE(TRY_CONVERT(varchar(50), d.cedula_doc), '-', ''), ' ', '') = ?
            """,
            record_id,
            re.sub(r"\D+", "", record_id),
        )
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Docente no encontrado")
        summary = _teacher_data_summary(row, editable_columns)
    fields = _row_to_dict(row, editable_columns)
    if target == "estudiantes":
        fields = _apply_student_legacy_rules(
            cursor,
            fields,
            codigo_estud=_clean_cell(getattr(row, "codigo_estud", "")),
            cedula=_clean_cell(getattr(row, "Cedula_Est", "")),
            editable_columns=editable_columns,
        )
    elif target == "docentes":
        fields = _apply_teacher_legacy_rules(
            fields,
            editable_columns=editable_columns,
        )
    return {
        "ok": True,
        "person": summary,
        "fields": fields,
        "columns": editable_columns,
        "catalogs": _legacy_data_update_catalogs(cursor, editable_columns, target),
        "target": target,
    }


@router.get("/actualizacion-datos/{target}/datos/{record_id}")
def get_legacy_data_update_record(
    target: str,
    record_id: str,
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    if target not in {"estudiantes", "docentes"}:
        raise HTTPException(status_code=404, detail="Tipo de actualización no soportado")
    try:
        with get_connection() as conn:
            return _load_legacy_data_update_record(conn.cursor(), target, record_id)
    except HTTPException:
        raise
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo cargar el registro: {exc}") from exc


@router.put("/actualizacion-datos/{target}/datos/{record_id}")
def update_legacy_data_update_record(
    target: str,
    record_id: str,
    payload: DataUpdatePayload,
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    if target not in {"estudiantes", "docentes"}:
        raise HTTPException(status_code=404, detail="Tipo de actualización no soportado")
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            editable_columns = _actualizacion_datos_columns(cursor, target)
            valid_updates = {
                key: value
                for key, value in payload.fields.items()
                if key in editable_columns
            }
            if not valid_updates:
                raise HTTPException(status_code=400, detail="No hay campos válidos para actualizar")
            table = "DATOS_ESTUD" if target == "estudiantes" else "DATOSDOCENTE"
            id_column = "codigo_estud" if target == "estudiantes" else "codigo_doc"
            cedula_column = "Cedula_Est" if target == "estudiantes" else "cedula_doc"
            if target in {"estudiantes", "docentes"}:
                cursor.execute(
                    f"""
                    SELECT TOP 1
                        {_quote_sql_name(id_column)} AS record_id,
                        {_quote_sql_name(cedula_column)} AS record_document,
                        {", ".join(_quote_sql_name(column) for column in editable_columns)}
                    FROM dbo.{_quote_sql_name(table)}
                    WHERE TRY_CONVERT(varchar(50), {_quote_sql_name(id_column)}) = ?
                       OR REPLACE(REPLACE(TRY_CONVERT(varchar(50), {_quote_sql_name(cedula_column)}), '-', ''), ' ', '') = ?
                    """,
                    record_id,
                    re.sub(r"\D+", "", record_id),
                )
                current_row = cursor.fetchone()
                if not current_row:
                    raise HTTPException(status_code=404, detail="Registro no encontrado para actualizar")
                current_fields = _row_to_dict(current_row, editable_columns)
                merged_fields = {**current_fields, **valid_updates}
                if target == "estudiantes":
                    final_fields = _apply_student_legacy_rules(
                        cursor,
                        merged_fields,
                        codigo_estud=_clean_cell(getattr(current_row, "record_id", "")),
                        cedula=_clean_cell(getattr(current_row, "record_document", "")),
                        editable_columns=editable_columns,
                    )
                else:
                    final_fields = _apply_teacher_legacy_rules(
                        merged_fields,
                        editable_columns=editable_columns,
                    )
                valid_updates = {
                    field: value
                    for field, value in final_fields.items()
                    if _clean_cell(value) != _clean_cell(current_fields.get(field))
                }
                if not valid_updates:
                    response = _load_legacy_data_update_record(cursor, target, record_id)
                    response.update({
                        "message": "No hay cambios nuevos para guardar",
                        "updated_fields": [],
                        "affected_rows": 0,
                    })
                    return response
            set_sql = ", ".join(f"{_quote_sql_name(column)} = ?" for column in valid_updates)
            params = list(valid_updates.values()) + [record_id, re.sub(r"\D+", "", record_id)]
            cursor.execute(
                f"""
                UPDATE dbo.{_quote_sql_name(table)}
                SET {set_sql}
                WHERE TRY_CONVERT(varchar(50), {_quote_sql_name(id_column)}) = ?
                   OR REPLACE(REPLACE(TRY_CONVERT(varchar(50), {_quote_sql_name(cedula_column)}), '-', ''), ' ', '') = ?
                """,
                *params,
            )
            affected = cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else 0
            if affected == 0:
                raise HTTPException(status_code=404, detail="Registro no encontrado para actualizar")
            conn.commit()
            response = _load_legacy_data_update_record(cursor, target, record_id)
            response.update({
                "message": "Datos actualizados correctamente",
                "updated_fields": list(valid_updates.keys()),
                "affected_rows": affected,
            })
            return response
    except HTTPException:
        raise
    except pyodbc.Error as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo actualizar el registro: {exc}") from exc


@router.get("/senescyt/estudiantes/buscar")
def search_senescyt_student_data_compat(
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
    q: str = Query(default="", max_length=120),
    limit: int = Query(default=60, ge=1, le=200),
) -> dict[str, Any]:
    payload = search_legacy_data_update_records("estudiantes", current_user, q, limit)
    rows = [
        {
            "codigo_estud": item["id"],
            "estudiante": item["nombre"],
            "numero_identificacion": item["cedula"],
            "nombre_carrera": item["carrera"],
            "campos_llenos": item["campos_llenos"],
            "campos_pendientes": item["campos_pendientes"],
            "campos_totales": item["campos_totales"],
            "porcentaje_lleno": item["porcentaje_lleno"],
            "campos_faltantes": item["campos_faltantes"],
        }
        for item in payload["rows"]
    ]
    return {**payload, "rows": rows}


@router.get("/senescyt/estudiantes/datos/{codigo_estud}")
def get_senescyt_student_data_compat(
    codigo_estud: str,
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
) -> dict[str, Any]:
    payload = get_legacy_data_update_record("estudiantes", codigo_estud, current_user)
    person = payload["person"]
    student = {
        "codigo_estud": person["id"],
        "estudiante": person["nombre"],
        "numero_identificacion": person["cedula"],
        "nombre_carrera": person["carrera"],
        "campos_llenos": person["campos_llenos"],
        "campos_pendientes": person["campos_pendientes"],
        "campos_totales": person["campos_totales"],
        "porcentaje_lleno": person["porcentaje_lleno"],
        "campos_faltantes": person["campos_faltantes"],
    }
    return {
        "ok": True,
        "student": student,
        "fields": payload["fields"],
        "report_columns": payload["columns"],
        "datos_estud_fields": payload["fields"],
        "datos_estud_columns": payload["columns"],
    }


@router.put("/senescyt/estudiantes/datos/{codigo_estud}")
def update_senescyt_student_data_compat(
    codigo_estud: str,
    payload: DataUpdatePayload,
    current_user: Annotated[SessionUser, Depends(_STUDENT_ACCESS)],
) -> dict[str, Any]:
    result = update_legacy_data_update_record("estudiantes", codigo_estud, payload, current_user)
    compat = get_senescyt_student_data_compat(codigo_estud, current_user)
    compat.update({
        "message": result.get("message"),
        "updated_fields": result.get("updated_fields", []),
        "affected_rows": result.get("affected_rows", 0),
    })
    return compat



