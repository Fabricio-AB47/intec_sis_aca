from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
import json
from typing import Annotated, Any

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field, model_validator
import pyodbc
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.core.security import SessionUser, require_roles
from app.routers.practicas_institucionales import (
    _ADMIN_ROLES,
    _catalog_state_id,
    _clean,
    _document_compliance_summary,
    _fetch_all,
    _fetch_legacy_expediente,
    _required_document_codes,
    _required_documents_status,
    _responsible_assignment,
    _row_dict,
    _sync_titulacion_completion,
)
from app.services.db import get_practices_connection
from app.services.practices_operations import (
    assert_practices_database,
    calculate_actor_grade,
    effective_process_configuration,
    ensure_operations_schema,
    record_evaluation_history,
    save_titulation_reconciliation,
    update_compliance_enrollment_status,
    write_operations_audit,
)


router = APIRouter(prefix="/api/practicas/operaciones", tags=["practicas-operativas"])

_ALL_ACCESS = require_roles(
    "ADMINISTRADOR",
    "ACADEMICO",
    "RECTOR",
    "VICERRECTOR",
    "SOPORTE",
    "SECRETARIA",
    "DOCENTE",
    "ESTUDIANTE",
)
_ADMIN_ACCESS = require_roles("ADMINISTRADOR", "ACADEMICO", "RECTOR", "VICERRECTOR", "SOPORTE", "SECRETARIA")
_RESPONSIBLE_ACCESS = require_roles(
    "DOCENTE", "ADMINISTRADOR", "ACADEMICO", "RECTOR", "VICERRECTOR", "SOPORTE", "SECRETARIA"
)

_REQUIRED_HOURS = {"PPF": 240.0, "VIN": 60.0}
_REQUIRED_DOCUMENTS = {"PPF": 5, "VIN": 4}
_MINIMUM_PASSING_GRADE = 7.0


class EntityPayload(BaseModel):
    nombre: str = Field(min_length=3, max_length=250)
    ruc: str | None = Field(default=None, max_length=20)
    tipo_entidad: str | None = Field(default=None, max_length=80)
    sector_economico: str | None = Field(default=None, max_length=160)
    direccion: str | None = Field(default=None, max_length=500)
    contacto_nombre: str | None = Field(default=None, max_length=250)
    contacto_correo: str | None = Field(default=None, max_length=250)
    contacto_telefono: str | None = Field(default=None, max_length=30)
    activo: bool = True


class AgreementPayload(BaseModel):
    entidad_id: int = Field(gt=0)
    tipo_proceso_codigo: str = Field(pattern="^(PPF|VIN)$")
    codigo_convenio: str = Field(min_length=2, max_length=80)
    objeto: str | None = None
    fecha_inicio: date
    fecha_fin: date
    estado: str = Field(default="VIGENTE", max_length=30)
    archivo_url: str | None = Field(default=None, max_length=1000)
    activo: bool = True

    @model_validator(mode="after")
    def validate_dates(self) -> "AgreementPayload":
        if self.fecha_fin < self.fecha_inicio:
            raise ValueError("La fecha de fin del convenio no puede ser anterior a la fecha de inicio.")
        return self


class ProjectPayload(BaseModel):
    entidad_id: int | None = Field(default=None, gt=0)
    convenio_id: int | None = Field(default=None, gt=0)
    codigo_proyecto: str = Field(min_length=2, max_length=80)
    nombre: str = Field(min_length=3, max_length=300)
    linea_intervencion: str = Field(min_length=3, max_length=250)
    poblacion_objetivo: str | None = None
    beneficiarios_previstos: int | None = Field(default=None, ge=0)
    objetivo_general: str | None = None
    fecha_inicio: date
    fecha_fin: date
    estado: str = Field(default="PLANIFICADO", max_length=30)
    activo: bool = True

    @model_validator(mode="after")
    def validate_dates(self) -> "ProjectPayload":
        if self.fecha_fin < self.fecha_inicio:
            raise ValueError("La fecha de fin del proyecto no puede ser anterior a la fecha de inicio.")
        return self


class PlanPayload(BaseModel):
    entidad_id: int | None = Field(default=None, gt=0)
    convenio_id: int | None = Field(default=None, gt=0)
    proyecto_id: int | None = Field(default=None, gt=0)
    tutor_externo_nombre: str | None = Field(default=None, max_length=250)
    tutor_externo_correo: str | None = Field(default=None, max_length=250)
    tutor_externo_telefono: str | None = Field(default=None, max_length=30)
    objetivo_general: str | None = None
    resultados_aprendizaje: str | None = None
    actividades_planificadas: str | None = None
    fecha_inicio: date | None = None
    fecha_fin: date | None = None
    horas_planificadas: float = Field(default=0, ge=0, le=10000)
    estado: str = Field(default="BORRADOR", pattern="^(BORRADOR|APROBADO|EN_EJECUCION|FINALIZADO)$")

    @model_validator(mode="after")
    def validate_dates(self) -> "PlanPayload":
        if self.fecha_inicio and self.fecha_fin and self.fecha_fin < self.fecha_inicio:
            raise ValueError("La fecha de fin del plan no puede ser anterior a la fecha de inicio.")
        return self


class ActivityPayload(BaseModel):
    fecha_actividad: date
    descripcion: str = Field(min_length=3, max_length=1500)
    horas: float | None = Field(default=None, gt=0, le=24)
    hora_inicio: time | None = None
    hora_fin: time | None = None
    descanso_minutos: int = Field(default=0, ge=0, le=600)
    modalidad: str | None = Field(default=None, pattern="^(PRESENCIAL|VIRTUAL|HIBRIDA)$")
    lugar: str | None = Field(default=None, max_length=300)
    evidencia_url: str | None = Field(default=None, max_length=1000)
    evidencia_nombre: str | None = Field(default=None, max_length=300)

    @model_validator(mode="after")
    def validate_schedule(self) -> "ActivityPayload":
        has_start = self.hora_inicio is not None
        has_end = self.hora_fin is not None
        if has_start != has_end:
            raise ValueError("Registre tanto la hora de inicio como la hora de fin.")
        if has_start and has_end:
            start_minutes = self.hora_inicio.hour * 60 + self.hora_inicio.minute
            end_minutes = self.hora_fin.hour * 60 + self.hora_fin.minute
            if end_minutes <= start_minutes:
                raise ValueError("La hora de fin debe ser posterior a la hora de inicio.")
            worked_minutes = end_minutes - start_minutes - self.descanso_minutos
            if worked_minutes <= 0:
                raise ValueError("El descanso no puede consumir toda la jornada registrada.")
            self.horas = round(worked_minutes / 60, 2)
        elif self.horas is None:
            raise ValueError("Registre las horas o una jornada con hora de inicio y fin.")
        return self


class ActivityReviewPayload(BaseModel):
    estado_revision: str = Field(pattern="^(VALIDADO|OBSERVADO|RECHAZADO)$")
    observacion_revision: str | None = Field(default=None, max_length=1000)

    @model_validator(mode="after")
    def validate_observation(self) -> "ActivityReviewPayload":
        if self.estado_revision != "VALIDADO" and not _clean(self.observacion_revision):
            raise ValueError("Registra una observación para devolver o rechazar la actividad.")
        return self


class IndicatorPayload(BaseModel):
    indicador_id: int | None = Field(default=None, gt=0)
    nombre: str = Field(min_length=3, max_length=300)
    unidad_medida: str = Field(min_length=1, max_length=80)
    meta: float = Field(ge=0)
    resultado: float | None = Field(default=None, ge=0)
    evidencia_url: str | None = Field(default=None, max_length=1000)
    observacion: str | None = Field(default=None, max_length=1000)


class ClosurePayload(BaseModel):
    supervision_realizada: bool = False
    evaluacion_entidad: float | None = Field(default=None, ge=0, le=10)
    informe_final_validado: bool = False
    acta_aceptacion_validada: bool = False
    certificado_emitido: bool = False
    observacion: str | None = Field(default=None, max_length=1500)
    cerrar: bool = False


class FinalEvaluationPayload(BaseModel):
    accion: str = Field(pattern="^(ENVIAR_REVISION|DEVOLVER|HABILITAR_CALIFICACION|CALIFICAR)$")
    calificacion: float | None = Field(default=None, ge=0, le=10)
    observacion: str | None = Field(default=None, max_length=1500)

    @model_validator(mode="after")
    def validate_evaluation(self) -> "FinalEvaluationPayload":
        if self.accion != "CALIFICAR" and self.calificacion is not None:
            raise ValueError("La calificación solo se admite al ejecutar la acción CALIFICAR.")
        if self.accion == "DEVOLVER" and not _clean(self.observacion):
            raise ValueError("Registre la corrección solicitada antes de devolver el expediente.")
        return self


class ProcessConfigurationPayload(BaseModel):
    tipo_proceso_codigo: str = Field(pattern="^(PPF|VIN)$")
    codigo_carrera: str | None = Field(default=None, max_length=50)
    nivel: str | None = Field(default=None, max_length=50)
    codigo_periodo: str | None = Field(default=None, max_length=50)
    horas_requeridas: float = Field(gt=0, le=10000)
    documentos_requeridos: int = Field(ge=0, le=100)
    nota_minima_aprobacion: float = Field(ge=0, le=10)
    requiere_evaluacion_docente: bool = True
    requiere_evaluacion_tutor: bool = False
    requiere_autoevaluacion: bool = False
    requiere_resultado_vinculacion: bool = False
    peso_docente: float = Field(default=100, ge=0, le=100)
    peso_tutor: float = Field(default=0, ge=0, le=100)
    peso_autoevaluacion: float = Field(default=0, ge=0, le=100)

    @model_validator(mode="after")
    def validate_weights(self) -> "ProcessConfigurationPayload":
        expected_documents = _REQUIRED_DOCUMENTS[self.tipo_proceso_codigo]
        if self.documentos_requeridos != expected_documents:
            raise ValueError(
                f"El flujo {self.tipo_proceso_codigo} requiere exactamente "
                f"{expected_documents} documentos institucionales."
            )
        role_rules = (
            (self.requiere_evaluacion_docente, self.peso_docente, "docente académico"),
            (self.requiere_evaluacion_tutor, self.peso_tutor, "tutor empresarial"),
            (self.requiere_autoevaluacion, self.peso_autoevaluacion, "autoevaluación"),
        )
        for required, weight, label in role_rules:
            if required and weight <= 0:
                raise ValueError(f"El peso de {label} debe ser mayor a cero cuando su evaluación es obligatoria.")
            if not required and weight != 0:
                raise ValueError(f"El peso de {label} debe ser cero cuando su evaluación no es obligatoria.")
        if abs(sum(weight for required, weight, _ in role_rules if required) - 100) > 0.01:
            raise ValueError("Los pesos de las evaluaciones obligatorias deben sumar 100%.")
        if self.tipo_proceso_codigo == "PPF" and self.requiere_resultado_vinculacion:
            raise ValueError("Los resultados de vinculación solo se aplican al proceso VIN.")
        return self


class ActorEvaluationPayload(BaseModel):
    calificacion: float = Field(ge=0, le=10)
    evaluador_nombre: str | None = Field(default=None, max_length=250)
    evaluador_correo: str | None = Field(default=None, max_length=250)
    observacion: str | None = Field(default=None, max_length=1500)
    evidencia_url: str | None = Field(default=None, max_length=1000)


class VinculationResultPayload(BaseModel):
    beneficiarios_reales: int = Field(ge=0, le=10000000)
    resumen_impacto: str = Field(min_length=10, max_length=10000)
    observacion: str | None = Field(default=None, max_length=1500)
    evidencia_url: str | None = Field(default=None, max_length=1000)
    validar: bool = False


class VinculationProductPayload(BaseModel):
    producto_id: int | None = Field(default=None, gt=0)
    nombre: str = Field(min_length=3, max_length=300)
    descripcion: str | None = Field(default=None, max_length=1500)
    cantidad: float = Field(ge=0, le=10000000)
    unidad_medida: str = Field(min_length=1, max_length=80)
    evidencia_url: str | None = Field(default=None, max_length=1000)


class VinculationProductReviewPayload(BaseModel):
    estado_revision: str = Field(pattern="^(VALIDADO|OBSERVADO|RECHAZADO)$")
    observacion_revision: str | None = Field(default=None, max_length=1000)

    @model_validator(mode="after")
    def validate_observation(self) -> "VinculationProductReviewPayload":
        if self.estado_revision != "VALIDADO" and not _clean(self.observacion_revision):
            raise ValueError("Registre una observación para observar o rechazar el producto.")
        return self


class ReopenPayload(BaseModel):
    motivo: str = Field(min_length=10, max_length=1500)
    confirmar_reversion_titulacion: bool = False


def _grade_result(grade: float, minimum: float = _MINIMUM_PASSING_GRADE) -> str:
    return "APROBADO" if float(grade) >= float(minimum) else "REPROBADO"


def _configuration_for_expediente(
    cursor: pyodbc.Cursor,
    expediente: dict[str, Any],
    process: str,
) -> dict[str, Any]:
    return effective_process_configuration(
        cursor,
        process_code=process,
        career_code=expediente.get("cod_anio_basica"),
        level=expediente.get("semestre_numero") or expediente.get("semestre"),
        period_code=expediente.get("codigo_periodo"),
    )


def _actor_role(value: Any) -> str:
    role = _clean(value).upper()
    if role not in {"DOCENTE_ACADEMICO", "TUTOR_EMPRESARIAL", "AUTOEVALUACION"}:
        raise HTTPException(status_code=400, detail="El rol de evaluación no es válido.")
    return role


def _actor_evaluations(cursor: pyodbc.Cursor, expediente_id: int) -> list[dict[str, Any]]:
    cursor.execute(
        "SELECT * FROM ops.evaluacion_actor WHERE expediente_id = ? ORDER BY rol_evaluador",
        expediente_id,
    )
    return _fetch_all(cursor)


def _actor_grade_calculation(
    cursor: pyodbc.Cursor,
    expediente_id: int,
    configuration: dict[str, Any],
) -> dict[str, Any]:
    evaluations = _actor_evaluations(cursor, expediente_id)
    grade, missing, components = calculate_actor_grade(configuration, evaluations)
    return {
        "calificacion_calculada": grade,
        "roles_faltantes": missing,
        "componentes": components,
        "usa_evaluaciones_actores": any(
            bool(configuration.get(key))
            for key in (
                "requiere_evaluacion_docente",
                "requiere_evaluacion_tutor",
                "requiere_autoevaluacion",
            )
        ),
    }


def _process_code(value: Any) -> str:
    process = _clean(value).upper()
    if process not in {"PPF", "VIN"}:
        raise HTTPException(status_code=400, detail="El proceso debe ser PPF o VIN.")
    return process


def _current_login(user: SessionUser) -> str:
    return _clean(user.login or user.email) or "SISTEMA"


def _expediente_access(
    cursor: pyodbc.Cursor,
    expediente_id: int,
    current_user: SessionUser,
    *,
    responsible_only: bool = False,
) -> dict[str, Any]:
    row = _fetch_legacy_expediente(cursor, expediente_id)
    if not row:
        raise HTTPException(status_code=404, detail="No existe el expediente solicitado.")
    expediente = _row_dict(cursor, row)
    if current_user.rol in _ADMIN_ROLES:
        return expediente
    if current_user.rol == "ESTUDIANTE":
        if responsible_only:
            raise HTTPException(status_code=403, detail="Esta acción corresponde al responsable del proceso.")
        if current_user.codigo_estud is None or int(expediente.get("codigo_estud") or 0) != int(current_user.codigo_estud):
            raise HTTPException(status_code=403, detail="El expediente no pertenece al estudiante autenticado.")
        return expediente
    if current_user.rol == "DOCENTE":
        _responsible_assignment(cursor, expediente_id, current_user, require_approval=False)
        return expediente
    raise HTTPException(status_code=403, detail="No tiene acceso al expediente solicitado.")


def _teacher_filter(current_user: SessionUser, params: list[Any]) -> str:
    filters: list[str] = []
    if current_user.cedula:
        filters.append("LTRIM(RTRIM(rp.cedula_ruc)) = ?")
        params.append(_clean(current_user.cedula))
    for email in {_clean(current_user.email).lower(), _clean(current_user.login).lower()} - {""}:
        filters.append("LOWER(LTRIM(RTRIM(rp.correo))) = ?")
        params.append(email)
    if current_user.codigo_doc is not None:
        filters.append("TRY_CONVERT(bigint, rp.codigo_referencia) = ?")
        params.append(int(current_user.codigo_doc))
    if not filters:
        return "AND 1 = 0"
    return "AND EXISTS (SELECT 1 FROM pp.responsable_proceso rp WHERE rp.expediente_id = e.expediente_id AND rp.activo = 1 AND (" + " OR ".join(filters) + "))"


def _principal_responsible(cursor: pyodbc.Cursor, expediente_id: int) -> dict[str, Any] | None:
    cursor.execute(
        """
        SELECT TOP (1)
            responsable_proceso_id AS ResponsableProcesoId,
            nombres AS NombreResponsable,
            correo AS CorreoResponsable,
            cedula_ruc AS CedulaResponsable,
            CONVERT(varchar(50), codigo_referencia) AS CodigoDocente,
            puede_validar_documentos AS PuedeValidarDocumentos,
            puede_aprobar AS PuedeAprobar
        FROM pp.responsable_proceso
        WHERE expediente_id = ? AND activo = 1
        ORDER BY principal DESC, responsable_proceso_id DESC
        """,
        expediente_id,
    )
    row = cursor.fetchone()
    return _row_dict(cursor, row) if row else None


def _reconciliation_status(cursor: pyodbc.Cursor, expediente_id: int, process: str) -> dict[str, Any] | None:
    cursor.execute(
        """
        SELECT TOP (1)
            conciliacion_id AS ConciliacionId,
            estado AS Estado,
            intentos AS Intentos,
            proximo_intento AS ProximoIntento,
            ultimo_error AS UltimoError,
            fecha_ultimo_intento AS FechaUltimoIntento,
            fecha_completado AS FechaCompletado
        FROM ops.conciliacion_titulacion
        WHERE expediente_id = ? AND tipo_proceso_codigo = ?
        """,
        expediente_id,
        process,
    )
    row = cursor.fetchone()
    return _row_dict(cursor, row) if row else None


def _process_requirements(
    *,
    process: str,
    responsible: dict[str, Any] | None,
    plan: dict[str, Any] | None,
    documents: list[dict[str, Any]],
    summary: dict[str, Any],
    closure: dict[str, Any] | None,
    reconciliation: dict[str, Any] | None,
    evaluation: dict[str, Any] | None = None,
    indicators: list[dict[str, Any]] | None = None,
    configuration: dict[str, Any] | None = None,
    actor_evaluations: list[dict[str, Any]] | None = None,
    vinculation_result: dict[str, Any] | None = None,
    vinculation_products: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    rules = configuration or {}
    required_hours = float(rules.get("horas_requeridas") or _REQUIRED_HOURS[process])
    loaded_documents = sum(1 for item in documents if item.get("Cargado"))
    validated_documents = sum(1 for item in documents if item.get("Validado"))
    registered_hours = float(summary.get("horas_registradas") or 0)
    validated_hours = float(summary.get("horas_validadas") or 0)
    pending_activities = int(summary.get("pendientes") or 0)
    plan_state = _clean((plan or {}).get("estado")).upper()
    titulation_state = _clean((reconciliation or {}).get("Estado")).upper()
    evaluation_state = _clean((evaluation or {}).get("estado")).upper()
    evaluation_result = _clean((evaluation or {}).get("resultado")).upper()
    evaluation_grade = (evaluation or {}).get("calificacion")
    indicator_items = indicators or []
    completed_indicators = sum(1 for item in indicator_items if item.get("resultado") is not None)
    actor_items = actor_evaluations or []
    actor_roles = {_clean(item.get("rol_evaluador")).upper() for item in actor_items}
    required_actor_roles: list[str] = []
    if rules.get("requiere_evaluacion_docente"):
        required_actor_roles.append("DOCENTE_ACADEMICO")
    if rules.get("requiere_evaluacion_tutor"):
        required_actor_roles.append("TUTOR_EMPRESARIAL")
    if rules.get("requiere_autoevaluacion"):
        required_actor_roles.append("AUTOEVALUACION")
    completed_actor_roles = sum(1 for role in required_actor_roles if role in actor_roles)
    product_items = vinculation_products or []
    validated_products = sum(1 for item in product_items if _clean(item.get("estado_revision")).upper() == "VALIDADO")

    def requirement(
        code: str,
        title: str,
        detail: str,
        *,
        complete: bool,
        in_review: bool = False,
    ) -> dict[str, Any]:
        return {
            "codigo": code,
            "titulo": title,
            "detalle": detail,
            "estado": "COMPLETO" if complete else ("EN_REVISION" if in_review else "PENDIENTE"),
            "completo": complete,
        }

    requirements = [
        requirement(
            "INSCRIPCION",
            "Inscripción institucional",
            "La inscripción de cumplimiento está registrada y no modifica la matrícula académica.",
            complete=True,
        ),
        requirement(
            "RESPONSABLE",
            "Responsable académico asignado",
            _clean((responsible or {}).get("NombreResponsable")) or "Administración debe asignar al docente responsable.",
            complete=bool(responsible),
        ),
        requirement(
            "PLAN",
            "Plan de prácticas aprobado",
            f"Estado actual: {plan_state or 'SIN PLAN'}.",
            complete=plan_state in {"APROBADO", "EN_EJECUCION", "FINALIZADO"},
            in_review=bool(plan) and plan_state == "BORRADOR",
        ),
        requirement(
            "DOCUMENTOS",
            "Documentos obligatorios",
            f"{loaded_documents} de {len(documents)} cargados; {validated_documents} validados.",
            complete=bool(documents) and validated_documents == len(documents),
            in_review=bool(documents) and loaded_documents == len(documents),
        ),
        requirement(
            "BITACORA",
            "Bitácora y horas",
            f"{registered_hours:g} hora(s) registradas y {validated_hours:g} validadas de {required_hours:g} requeridas.",
            complete=validated_hours >= required_hours and pending_activities == 0,
            in_review=registered_hours >= required_hours or pending_activities > 0,
        ),
    ]
    if required_actor_roles:
        requirements.append(
            requirement(
                "EVALUACIONES_ACTORES",
                "Evaluaciones de responsables",
                f"{completed_actor_roles} de {len(required_actor_roles)} evaluación(es) obligatoria(s) registradas.",
                complete=completed_actor_roles == len(required_actor_roles),
                in_review=completed_actor_roles > 0,
            )
        )
    if process == "VIN":
        requirements.append(
            requirement(
                "INDICADORES",
                "Metas e indicadores",
                f"{completed_indicators} de {len(indicator_items)} indicador(es) cuentan con resultado.",
                complete=bool(indicator_items) and completed_indicators == len(indicator_items),
                in_review=bool(indicator_items),
            )
        )
        if rules.get("requiere_resultado_vinculacion"):
            result_state = _clean((vinculation_result or {}).get("estado")).upper()
            requirements.append(
                requirement(
                    "RESULTADOS_VINCULACION",
                    "Resultados, beneficiarios y productos",
                    f"Resultado: {result_state or 'PENDIENTE'}; {validated_products} de {len(product_items)} producto(s) validado(s).",
                    complete=result_state == "VALIDADO" and bool(product_items) and validated_products == len(product_items),
                    in_review=bool(vinculation_result) or bool(product_items),
                )
            )
    requirements.extend([
        requirement(
            "EVALUACION",
            "Revisión y calificación final",
            (
                f"Resultado: {evaluation_result}; calificación: {float(evaluation_grade):.2f}/10."
                if evaluation_state == "CALIFICADA" and evaluation_grade is not None
                else {
                    "EN_REVISION": "El expediente está siendo revisado por el responsable.",
                    "PENDIENTE_CALIFICACION": "La revisión terminó y el expediente está a la espera de calificación.",
                    "PENDIENTE_REVISION": "El expediente debe corregirse antes de volver a revisión.",
                }.get(evaluation_state, "El estudiante debe enviar el expediente a revisión.")
            ),
            complete=evaluation_state == "CALIFICADA",
            in_review=evaluation_state in {"EN_REVISION", "PENDIENTE_CALIFICACION"},
        ),
        requirement(
            "CIERRE",
            "Cierre y certificado",
            "Supervisión, evaluación, informe, acta y certificado deben estar confirmados.",
            complete=bool((closure or {}).get("fecha_cierre")),
            in_review=bool(closure),
        ),
        requirement(
            "TITULACION",
            "Cumplimiento reflejado en Titulación",
            (
                "No aplica porque el resultado final es reprobado."
                if evaluation_result == "REPROBADO"
                else _clean((reconciliation or {}).get("UltimoError")) or f"Estado de conciliación: {titulation_state or 'PENDIENTE'}."
            ),
            complete=titulation_state == "COMPLETADO" or evaluation_result == "REPROBADO",
            in_review=titulation_state in {"PENDIENTE", "PROCESANDO"},
        ),
    ])
    return requirements


def _evaluation_submission_errors(
    cursor: pyodbc.Cursor,
    expediente_id: int,
    process: str,
    configuration: dict[str, Any],
) -> list[str]:
    errors: list[str] = []
    cursor.execute("SELECT estado FROM ops.plan_proceso WHERE expediente_id = ?", expediente_id)
    plan_row = cursor.fetchone()
    plan_state = _clean(plan_row[0] if plan_row else "").upper()
    if plan_state not in {"APROBADO", "EN_EJECUCION", "FINALIZADO"}:
        errors.append("El plan debe estar aprobado antes de enviar el expediente a revisión.")

    cursor.execute(
        """
        SELECT ISNULL(SUM(CASE WHEN estado_revision <> N'RECHAZADO' THEN horas ELSE 0 END), 0)
        FROM ops.registro_actividad
        WHERE expediente_id = ?
        """,
        expediente_id,
    )
    registered_hours = float(cursor.fetchone()[0] or 0)
    required_hours = float(configuration["horas_requeridas"])
    if registered_hours < required_hours:
        errors.append(f"Registre al menos {required_hours:g} horas antes de solicitar la revisión.")

    documents = _required_documents_status(cursor, expediente_id, process)
    missing_documents = [
        _clean(item.get("Nombre") or item.get("Codigo"))
        for item in documents
        if not item.get("Cargado")
    ]
    if missing_documents:
        errors.append(f"Cargue los documentos obligatorios: {', '.join(missing_documents)}.")
    return errors


def _evaluation_grading_errors(
    cursor: pyodbc.Cursor,
    expediente_id: int,
    process: str,
    configuration: dict[str, Any],
) -> list[str]:
    errors: list[str] = []
    cursor.execute("SELECT estado FROM ops.plan_proceso WHERE expediente_id = ?", expediente_id)
    plan_row = cursor.fetchone()
    plan_state = _clean(plan_row[0] if plan_row else "").upper()
    if plan_state not in {"APROBADO", "EN_EJECUCION", "FINALIZADO"}:
        errors.append("El plan de prácticas aún no está aprobado.")

    cursor.execute(
        """
        SELECT
            ISNULL(SUM(CASE WHEN estado_revision = N'VALIDADO' THEN horas ELSE 0 END), 0),
            ISNULL(SUM(CASE WHEN estado_revision NOT IN (N'VALIDADO', N'RECHAZADO') THEN 1 ELSE 0 END), 0)
        FROM ops.registro_actividad
        WHERE expediente_id = ?
        """,
        expediente_id,
    )
    activity_row = cursor.fetchone()
    validated_hours = float(activity_row[0] or 0)
    pending_activities = int(activity_row[1] or 0)
    required_hours = float(configuration["horas_requeridas"])
    if validated_hours < required_hours:
        errors.append(f"Valide al menos {required_hours:g} horas antes de habilitar la calificación.")
    if pending_activities:
        errors.append("Revise todas las actividades observadas o pendientes.")

    documents = _required_documents_status(cursor, expediente_id, process)
    pending_documents = [
        _clean(item.get("Nombre") or item.get("Codigo"))
        for item in documents
        if not item.get("Validado")
    ]
    if pending_documents:
        errors.append(f"Valide los documentos obligatorios: {', '.join(pending_documents)}.")

    grade_calculation = _actor_grade_calculation(cursor, expediente_id, configuration)
    if grade_calculation["roles_faltantes"]:
        actor_labels = {
            "DOCENTE_ACADEMICO": "docente académico",
            "TUTOR_EMPRESARIAL": "tutor empresarial",
            "AUTOEVALUACION": "autoevaluación del estudiante",
        }
        missing_labels = [actor_labels.get(role, role) for role in grade_calculation["roles_faltantes"]]
        errors.append(f"Registre las evaluaciones obligatorias de: {', '.join(missing_labels)}.")

    if process == "VIN":
        cursor.execute(
            """
            SELECT COUNT_BIG(*),
                   ISNULL(SUM(CASE WHEN resultado IS NULL THEN 1 ELSE 0 END), 0)
            FROM ops.meta_indicador
            WHERE expediente_id = ?
            """,
            expediente_id,
        )
        indicator_row = cursor.fetchone()
        if int(indicator_row[0] or 0) == 0 or int(indicator_row[1] or 0) > 0:
            errors.append("Complete los resultados de todos los indicadores de vinculación.")
        if configuration.get("requiere_resultado_vinculacion"):
            cursor.execute(
                "SELECT estado FROM ops.resultado_vinculacion WHERE expediente_id = ?",
                expediente_id,
            )
            result_row = cursor.fetchone()
            if not result_row or _clean(result_row[0]).upper() != "VALIDADO":
                errors.append("Valide el resultado y el impacto alcanzado por el proyecto de vinculación.")
            cursor.execute(
                """
                SELECT COUNT_BIG(*),
                       ISNULL(SUM(CASE WHEN estado_revision <> N'VALIDADO' THEN 1 ELSE 0 END), 0)
                FROM ops.producto_vinculacion
                WHERE expediente_id = ?
                """,
                expediente_id,
            )
            product_row = cursor.fetchone()
            if int(product_row[0] or 0) == 0 or int(product_row[1] or 0) > 0:
                errors.append("Registre y valide todos los productos entregados en vinculación.")
    return errors


def _ensure_expediente_open(
    cursor: pyodbc.Cursor,
    expediente_id: int,
    *,
    allow_evaluation_transition: bool = False,
) -> None:
    cursor.execute("SELECT fecha_cierre FROM ops.cierre_proceso WHERE expediente_id = ?", expediente_id)
    closure_row = cursor.fetchone()
    if closure_row and closure_row[0] is not None:
        raise HTTPException(status_code=409, detail="El expediente ya está cerrado y no admite modificaciones.")
    if allow_evaluation_transition:
        return
    cursor.execute("SELECT estado FROM ops.evaluacion_practica WHERE expediente_id = ?", expediente_id)
    evaluation_row = cursor.fetchone()
    evaluation_state = _clean(evaluation_row[0] if evaluation_row else "").upper()
    if evaluation_state in {"PENDIENTE_CALIFICACION", "CALIFICADA"}:
        raise HTTPException(
            status_code=409,
            detail="La revisión concluyó; devuelva el expediente antes de modificar el plan, actividades o indicadores.",
        )


def _accessible_expedientes(
    cursor: pyodbc.Cursor,
    current_user: SessionUser,
    process: str,
    limit: int = 500,
) -> list[dict[str, Any]]:
    default_configuration = effective_process_configuration(cursor, process_code=process)
    params: list[Any] = []
    access_sql = ""
    if current_user.rol == "ESTUDIANTE":
        if current_user.codigo_estud is None:
            return []
        access_sql = "AND e.codigo_estud = ?"
        params.append(int(current_user.codigo_estud))
    elif current_user.rol == "DOCENTE":
        access_sql = _teacher_filter(current_user, params)
    elif current_user.rol not in _ADMIN_ROLES:
        return []
    required_document_codes = _required_document_codes(process)
    required_document_placeholders = ",".join("?" for _ in required_document_codes)
    cursor.execute(
        f"""
        SELECT TOP (?)
            e.expediente_id AS ExpedienteId,
            e.codigo_expediente AS CodigoExpediente,
            e.codigo_estud AS CodigoEstud,
            e.cedula_est AS Cedula,
            e.estudiante_snapshot AS Estudiante,
            e.cod_anio_basica AS CodigoCarrera,
            e.carrera_snapshot AS Carrera,
            e.codigo_periodo AS CodigoPeriodo,
            e.periodo_snapshot AS Periodo,
            e.fecha_inicio AS FechaInicio,
            e.fecha_fin AS FechaFin,
            e.horas_requeridas AS HorasRequeridas,
            e.horas_reconocidas AS HorasReconocidas,
            ee.codigo AS EstadoCodigo,
            ee.nombre AS Estado,
            ins.inscripcion_id AS InscripcionId,
            COALESCE(ins.estado, N'EN_PROCESO') AS EstadoInscripcion,
            ins.codigo_periodo_academico_origen AS CodigoPeriodoAcademicoOrigen,
            ins.codigo_periodo_institucional AS CodigoPeriodoInstitucional,
            CAST(COALESCE(ins.es_matricula_academica, 0) AS bit) AS EsMatriculaAcademica,
            evaluacion.evaluacion_id AS EvaluacionId,
            COALESCE(evaluacion.estado, N'PENDIENTE_REVISION') AS EstadoEvaluacion,
            evaluacion.calificacion AS CalificacionFinal,
            COALESCE(evaluacion.nota_minima_aprobacion, 7.00) AS NotaMinimaAprobacion,
            COALESCE(evaluacion.resultado, N'PENDIENTE') AS ResultadoEvaluacion,
            evaluacion.fecha_envio_revision AS FechaEnvioRevision,
            evaluacion.fecha_revision AS FechaRevisionEvaluacion,
            evaluacion.fecha_calificacion AS FechaCalificacion,
            op_plan.plan_id AS PlanId,
            op_plan.estado AS EstadoPlan,
            op_plan.entidad_id AS EntidadId,
            op_plan.convenio_id AS ConvenioId,
            op_plan.proyecto_id AS ProyectoId,
            ISNULL(activity_totals.HorasRegistradas, 0) AS HorasRegistradas,
            ISNULL(activity_totals.HorasValidadas, 0) AS HorasValidadas,
            ISNULL(activity_totals.Actividades, 0) AS Actividades,
            ISNULL(document_totals.DocumentosCargados, 0) AS DocumentosCargados,
            ISNULL(document_totals.DocumentosValidados, 0) AS DocumentosValidados,
            assigned_responsible.NombreResponsable,
            cierre.cierre_id AS CierreId,
            cierre.fecha_cierre AS FechaCierre
        FROM pp.expediente_practica e
        INNER JOIN cat.tipo_proceso tp ON tp.tipo_proceso_id = e.tipo_proceso_id
        INNER JOIN cat.estado_expediente ee ON ee.estado_expediente_id = e.estado_expediente_id
        LEFT JOIN ops.inscripcion_cumplimiento ins ON ins.expediente_id = e.expediente_id
        LEFT JOIN ops.evaluacion_practica evaluacion ON evaluacion.expediente_id = e.expediente_id
        LEFT JOIN ops.plan_proceso op_plan ON op_plan.expediente_id = e.expediente_id
        LEFT JOIN ops.cierre_proceso cierre ON cierre.expediente_id = e.expediente_id
        OUTER APPLY (
            SELECT
                SUM(ra.horas) AS HorasRegistradas,
                SUM(CASE WHEN ra.estado_revision = N'VALIDADO' THEN ra.horas ELSE 0 END) AS HorasValidadas,
                COUNT_BIG(*) AS Actividades
            FROM ops.registro_actividad ra
            WHERE ra.expediente_id = e.expediente_id
        ) AS activity_totals
        OUTER APPLY (
            SELECT
                SUM(CASE
                    WHEN latest.documento_id IS NOT NULL
                     AND latest.estado_codigo NOT IN (N'RECHAZADO', N'ANULADO') THEN 1
                    ELSE 0
                END) AS DocumentosCargados,
                SUM(CASE
                    WHEN latest.validado = 1 AND latest.estado_codigo = N'VALIDADO' THEN 1
                    ELSE 0
                END) AS DocumentosValidados
            FROM cat.tipo_documento_practica td
            OUTER APPLY (
                SELECT TOP (1)
                    dp.documento_id,
                    dp.validado,
                    ed.codigo AS estado_codigo
                FROM pp.documento_practica dp
                INNER JOIN cat.estado_documento ed
                    ON ed.estado_documento_id = dp.estado_documento_id
                WHERE dp.expediente_id = e.expediente_id
                  AND dp.tipo_documento_id = td.tipo_documento_id
                  AND ed.codigo <> N'ANULADO'
                ORDER BY dp.fecha_registro DESC, dp.documento_id DESC
            ) AS latest
            WHERE td.codigo IN ({required_document_placeholders})
              AND td.activo = 1
        ) AS document_totals
        OUTER APPLY (
            SELECT TOP (1) rp.nombres AS NombreResponsable
            FROM pp.responsable_proceso rp
            WHERE rp.expediente_id = e.expediente_id AND rp.activo = 1
            ORDER BY rp.principal DESC, rp.responsable_proceso_id DESC
        ) AS assigned_responsible
        WHERE tp.codigo = ?
          {access_sql}
        ORDER BY e.fecha_registro DESC, e.expediente_id DESC
        """,
        limit,
        *required_document_codes,
        process,
        *params,
    )
    items = _fetch_all(cursor)
    today = date.today()
    completed_states = {"APROBADO", "VALIDADO", "FINALIZADO", "CERRADO"}
    for item in items:
        state = _clean(item.get("EstadoCodigo")).upper()
        end = item.get("FechaFin")
        days = (end - today).days if isinstance(end, date) else None
        required_hours = float(item.get("HorasRequeridas") or default_configuration["horas_requeridas"])
        required_documents = len(required_document_codes)
        loaded_documents = min(int(item.get("DocumentosCargados") or 0), required_documents)
        validated_documents = min(int(item.get("DocumentosValidados") or 0), required_documents)
        evaluation_result = _clean(item.get("ResultadoEvaluacion")).upper()
        if evaluation_result == "REPROBADO":
            semaphore = "ROJO"
        elif item.get("FechaCierre") and evaluation_result == "APROBADO":
            semaphore = "VERDE"
        elif days is not None and days < 0:
            semaphore = "ROJO"
        elif days is not None and days <= 7:
            semaphore = "AMARILLO"
        elif (
            not item.get("PlanId")
            or float(item.get("HorasValidadas") or 0) < required_hours
            or int(item.get("DocumentosValidados") or 0) < required_documents
        ):
            semaphore = "AMARILLO"
        elif state in completed_states and evaluation_result == "APROBADO":
            semaphore = "VERDE"
        else:
            semaphore = "AMARILLO"
        item["TipoProcesoCodigo"] = process
        item["Semaforo"] = semaphore
        item["DiasRestantes"] = days
        item["HorasRequeridas"] = required_hours
        item["HorasRegistradas"] = float(item.get("HorasRegistradas") or 0)
        item["HorasValidadas"] = float(item.get("HorasValidadas") or 0)
        item["DocumentosCargados"] = loaded_documents
        item["DocumentosValidados"] = validated_documents
        item["DocumentosRequeridos"] = required_documents
        item["AvanceDocumental"] = round((loaded_documents / max(required_documents, 1)) * 100, 2)
        item["AvanceValidacionDocumental"] = round((validated_documents / max(required_documents, 1)) * 100, 2)
    return items


def _refresh_notifications(cursor: pyodbc.Cursor, current_user: SessionUser, process: str) -> int:
    items = _accessible_expedientes(cursor, current_user, process, 500)
    role = "DOCENTE" if current_user.rol == "DOCENTE" else ("ESTUDIANTE" if current_user.rol == "ESTUDIANTE" else "ADMINISTRATIVO")
    login = _current_login(current_user)
    cursor.execute(
        """
        SELECT clave_evento
        FROM ops.notificacion_proceso
        WHERE activa = 1
          AND tipo_proceso_codigo = ?
          AND destinatario_rol = ?
          AND ISNULL(destinatario_login, N'') = ?
        """,
        process,
        role,
        login,
    )
    existing_keys = {_clean(row[0]) for row in cursor.fetchall()}
    pending_rows: list[tuple[Any, ...]] = []
    for item in items:
        reasons: list[tuple[str, str, str]] = []
        if not item.get("PlanId"):
            reasons.append(("PLAN", "ADVERTENCIA", "El expediente aún no tiene plan operativo."))
        if item.get("Semaforo") == "ROJO":
            reasons.append(("VENCIDO", "CRITICA", "El plazo del expediente está vencido."))
        elif item.get("Semaforo") == "AMARILLO" and item.get("DiasRestantes") is not None:
            reasons.append(("POR_VENCER", "ADVERTENCIA", f"Restan {item.get('DiasRestantes')} día(s) para el cierre."))
        if int(item.get("DocumentosCargados") or 0) < int(item.get("DocumentosRequeridos") or 0):
            reasons.append(("DOCUMENTOS", "ADVERTENCIA", "Existen documentos obligatorios pendientes."))
        evaluation_state = _clean(item.get("EstadoEvaluacion")).upper()
        evaluation_result = _clean(item.get("ResultadoEvaluacion")).upper()
        if evaluation_state == "EN_REVISION":
            message = (
                "El expediente fue enviado y requiere revisión del responsable."
                if role != "ESTUDIANTE"
                else "Tu expediente está siendo revisado por el responsable."
            )
            reasons.append(("EVALUACION_REVISION", "INFORMATIVA", message))
        elif evaluation_state == "PENDIENTE_CALIFICACION":
            message = (
                "La revisión terminó y falta registrar la calificación final."
                if role != "ESTUDIANTE"
                else "La revisión terminó; tu expediente está a la espera de calificación."
            )
            reasons.append(("EVALUACION_CALIFICACION", "ADVERTENCIA", message))
        elif evaluation_state == "CALIFICADA" and evaluation_result in {"APROBADO", "REPROBADO"}:
            reasons.append((
                f"EVALUACION_{evaluation_result}",
                "INFORMATIVA" if evaluation_result == "APROBADO" else "CRITICA",
                f"Resultado final: {evaluation_result.lower()} con calificación {float(item.get('CalificacionFinal') or 0):.2f}/10.",
            ))
        for code, level, message in reasons:
            key = f"{process}:{item['ExpedienteId']}:{role}:{login}:{code}"
            if key in existing_keys:
                continue
            existing_keys.add(key)
            pending_rows.append((
                key,
                int(item["ExpedienteId"]),
                process,
                login,
                role,
                level,
                f"{process} · {item.get('Estudiante') or item.get('CodigoExpediente')}",
                message,
            ))
    if pending_rows:
        cursor.executemany(
            """
            INSERT INTO ops.notificacion_proceso (
                clave_evento, expediente_id, tipo_proceso_codigo, destinatario_login,
                destinatario_rol, nivel, titulo, mensaje
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            pending_rows,
        )
    return len(pending_rows)


@router.get("/catalogo")
def operations_catalog(
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    del current_user
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        database_name = assert_practices_database(cursor)
        cursor.execute("SELECT * FROM ops.entidad_receptora ORDER BY activo DESC, nombre")
        entities = _fetch_all(cursor)
        cursor.execute(
            """
            SELECT c.*, e.nombre AS entidad_nombre
            FROM ops.convenio_institucional c
            INNER JOIN ops.entidad_receptora e ON e.entidad_id = c.entidad_id
            ORDER BY c.activo DESC, c.fecha_fin DESC, c.codigo_convenio
            """
        )
        agreements = _fetch_all(cursor)
        cursor.execute(
            """
            SELECT p.*, e.nombre AS entidad_nombre, c.codigo_convenio
            FROM ops.proyecto_vinculacion p
            LEFT JOIN ops.entidad_receptora e ON e.entidad_id = p.entidad_id
            LEFT JOIN ops.convenio_institucional c ON c.convenio_id = p.convenio_id
            ORDER BY p.activo DESC, p.fecha_fin DESC, p.nombre
            """
        )
        projects = _fetch_all(cursor)
        cursor.execute(
            """
            SELECT *
            FROM ops.configuracion_proceso
            WHERE activo = 1
            ORDER BY tipo_proceso_codigo, codigo_periodo, codigo_carrera, nivel, configuracion_id
            """
        )
        configurations = _fetch_all(cursor)
        conn.commit()
    return {
        "entidades": entities,
        "convenios": agreements,
        "proyectos": projects,
        "configuraciones": configurations,
        "almacenamiento": {
            "base_datos": database_name,
            "esquema_operativo": "ops",
            "tabla_calificacion": "ops.evaluacion_practica",
            "fuente_academica": "INTECBDD (solo lectura)",
        },
    }


@router.put("/configuraciones")
def save_process_configuration(
    payload: ProcessConfigurationPayload,
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
) -> dict[str, Any]:
    process = _process_code(payload.tipo_proceso_codigo)
    career = _clean(payload.codigo_carrera) or None
    level = _clean(payload.nivel) or None
    period = _clean(payload.codigo_periodo) or None
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute(
            """
            SELECT configuracion_id
            FROM ops.configuracion_proceso WITH (UPDLOCK, HOLDLOCK)
            WHERE tipo_proceso_codigo = ?
              AND ((codigo_carrera = ?) OR (codigo_carrera IS NULL AND ? IS NULL))
              AND ((nivel = ?) OR (nivel IS NULL AND ? IS NULL))
              AND ((codigo_periodo = ?) OR (codigo_periodo IS NULL AND ? IS NULL))
            """,
            process,
            career,
            career,
            level,
            level,
            period,
            period,
        )
        existing = cursor.fetchone()
        values = (
            payload.horas_requeridas,
            payload.documentos_requeridos,
            payload.nota_minima_aprobacion,
            payload.requiere_evaluacion_docente,
            payload.requiere_evaluacion_tutor,
            payload.requiere_autoevaluacion,
            payload.requiere_resultado_vinculacion,
            payload.peso_docente,
            payload.peso_tutor,
            payload.peso_autoevaluacion,
        )
        if existing:
            configuration_id = int(existing[0])
            cursor.execute(
                """
                UPDATE ops.configuracion_proceso
                SET horas_requeridas = ?, documentos_requeridos = ?, nota_minima_aprobacion = ?,
                    requiere_evaluacion_docente = ?, requiere_evaluacion_tutor = ?,
                    requiere_autoevaluacion = ?, requiere_resultado_vinculacion = ?,
                    peso_docente = ?, peso_tutor = ?, peso_autoevaluacion = ?, activo = 1,
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE configuracion_id = ?
                """,
                *values,
                login,
                configuration_id,
            )
            action = "UPDATE"
        else:
            cursor.execute(
                """
                INSERT INTO ops.configuracion_proceso (
                    tipo_proceso_codigo, codigo_carrera, nivel, codigo_periodo,
                    horas_requeridas, documentos_requeridos, nota_minima_aprobacion,
                    requiere_evaluacion_docente, requiere_evaluacion_tutor,
                    requiere_autoevaluacion, requiere_resultado_vinculacion,
                    peso_docente, peso_tutor, peso_autoevaluacion, usuario_registro
                ) OUTPUT INSERTED.configuracion_id
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                process,
                career,
                level,
                period,
                *values,
                login,
            )
            configuration_id = int(cursor.fetchone()[0])
            action = "INSERT"
        cursor.execute(
            """
            UPDATE e
            SET e.horas_requeridas = ?, e.usuario_modifica = ?, e.fecha_modifica = SYSDATETIME()
            FROM pp.expediente_practica e
            INNER JOIN cat.tipo_proceso tp ON tp.tipo_proceso_id = e.tipo_proceso_id
            LEFT JOIN ops.cierre_proceso cierre ON cierre.expediente_id = e.expediente_id
            WHERE tp.codigo = ? AND cierre.fecha_cierre IS NULL
              AND (? IS NULL OR CONVERT(nvarchar(50), e.cod_anio_basica) = ?)
              AND (? IS NULL OR CONVERT(nvarchar(50), e.semestre_numero) = ? OR e.semestre = ?)
              AND (? IS NULL OR CONVERT(nvarchar(50), e.codigo_periodo) = ?)
            """,
            payload.horas_requeridas,
            login,
            process,
            career,
            career,
            level,
            level,
            level,
            period,
            period,
        )
        cursor.execute(
            """
            UPDATE evaluacion
            SET evaluacion.nota_minima_aprobacion = ?, evaluacion.usuario_modifica = ?,
                evaluacion.fecha_modifica = SYSDATETIME()
            FROM ops.evaluacion_practica evaluacion
            INNER JOIN pp.expediente_practica e ON e.expediente_id = evaluacion.expediente_id
            INNER JOIN cat.tipo_proceso tp ON tp.tipo_proceso_id = e.tipo_proceso_id
            LEFT JOIN ops.cierre_proceso cierre ON cierre.expediente_id = e.expediente_id
            WHERE tp.codigo = ? AND cierre.fecha_cierre IS NULL AND evaluacion.estado <> N'CALIFICADA'
              AND (? IS NULL OR CONVERT(nvarchar(50), e.cod_anio_basica) = ?)
              AND (? IS NULL OR CONVERT(nvarchar(50), e.semestre_numero) = ? OR e.semestre = ?)
              AND (? IS NULL OR CONVERT(nvarchar(50), e.codigo_periodo) = ?)
            """,
            payload.nota_minima_aprobacion,
            login,
            process,
            career,
            career,
            level,
            level,
            level,
            period,
            period,
        )
        write_operations_audit(
            cursor,
            entity="CONFIGURACION_PROCESO",
            entity_id=configuration_id,
            action=action,
            user=login,
            detail=f"{process}; horas={payload.horas_requeridas:g}; nota mínima={payload.nota_minima_aprobacion:g}",
        )
        conn.commit()
    return {
        "ok": True,
        "configuracion_id": configuration_id,
        "message": "Reglas del proceso guardadas en la base complementaria de prácticas.",
    }


@router.post("/entidades")
def save_entity(
    payload: EntityPayload,
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute(
            """
            INSERT INTO ops.entidad_receptora (
                nombre, ruc, tipo_entidad, sector_economico, direccion, contacto_nombre,
                contacto_correo, contacto_telefono, activo, usuario_registro
            ) OUTPUT INSERTED.entidad_id
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload.nombre,
            payload.ruc,
            payload.tipo_entidad,
            payload.sector_economico,
            payload.direccion,
            payload.contacto_nombre,
            payload.contacto_correo,
            payload.contacto_telefono,
            payload.activo,
            _current_login(current_user),
        )
        entity_id = int(cursor.fetchone()[0])
        write_operations_audit(cursor, entity="ENTIDAD_RECEPTORA", entity_id=entity_id, action="INSERT", user=_current_login(current_user), detail=payload.nombre)
        conn.commit()
    return {"ok": True, "entidad_id": entity_id, "message": "Entidad receptora registrada correctamente."}


@router.post("/convenios")
def save_agreement(
    payload: AgreementPayload,
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute("SELECT 1 FROM ops.entidad_receptora WHERE entidad_id = ? AND activo = 1", payload.entidad_id)
        if not cursor.fetchone():
            raise HTTPException(status_code=400, detail="La entidad receptora no existe o está inactiva.")
        cursor.execute(
            """
            INSERT INTO ops.convenio_institucional (
                entidad_id, tipo_proceso_codigo, codigo_convenio, objeto, fecha_inicio,
                fecha_fin, estado, archivo_url, activo, usuario_registro
            ) OUTPUT INSERTED.convenio_id
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload.entidad_id,
            payload.tipo_proceso_codigo,
            payload.codigo_convenio,
            payload.objeto,
            payload.fecha_inicio,
            payload.fecha_fin,
            payload.estado.upper(),
            payload.archivo_url,
            payload.activo,
            _current_login(current_user),
        )
        agreement_id = int(cursor.fetchone()[0])
        write_operations_audit(cursor, entity="CONVENIO", entity_id=agreement_id, action="INSERT", user=_current_login(current_user), detail=payload.codigo_convenio)
        conn.commit()
    return {"ok": True, "convenio_id": agreement_id, "message": "Convenio registrado correctamente."}


@router.post("/proyectos")
def save_project(
    payload: ProjectPayload,
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute(
            """
            INSERT INTO ops.proyecto_vinculacion (
                entidad_id, convenio_id, codigo_proyecto, nombre, linea_intervencion,
                poblacion_objetivo, beneficiarios_previstos, objetivo_general, fecha_inicio,
                fecha_fin, estado, activo, usuario_registro
            ) OUTPUT INSERTED.proyecto_id
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            payload.entidad_id,
            payload.convenio_id,
            payload.codigo_proyecto,
            payload.nombre,
            payload.linea_intervencion,
            payload.poblacion_objetivo,
            payload.beneficiarios_previstos,
            payload.objetivo_general,
            payload.fecha_inicio,
            payload.fecha_fin,
            payload.estado.upper(),
            payload.activo,
            _current_login(current_user),
        )
        project_id = int(cursor.fetchone()[0])
        write_operations_audit(cursor, entity="PROYECTO_VINCULACION", entity_id=project_id, action="INSERT", user=_current_login(current_user), detail=payload.nombre)
        conn.commit()
    return {"ok": True, "proyecto_id": project_id, "message": "Proyecto de vinculación registrado correctamente."}


@router.get("/expedientes/{expediente_id}")
def operational_detail(
    expediente_id: int,
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        database_name = assert_practices_database(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user)
        process = _process_code(expediente.get("tipo_proceso_codigo"))
        configuration = _configuration_for_expediente(cursor, expediente, process)
        cursor.execute("SELECT * FROM ops.plan_proceso WHERE expediente_id = ?", expediente_id)
        plan_row = cursor.fetchone()
        plan = _row_dict(cursor, plan_row) if plan_row else None
        cursor.execute("SELECT * FROM ops.registro_actividad WHERE expediente_id = ? ORDER BY fecha_actividad DESC, actividad_id DESC", expediente_id)
        activities = _fetch_all(cursor)
        cursor.execute("SELECT * FROM ops.meta_indicador WHERE expediente_id = ? ORDER BY indicador_id", expediente_id)
        indicators = _fetch_all(cursor)
        cursor.execute("SELECT * FROM ops.cierre_proceso WHERE expediente_id = ?", expediente_id)
        closure_row = cursor.fetchone()
        closure = _row_dict(cursor, closure_row) if closure_row else None
        cursor.execute("SELECT * FROM ops.evaluacion_practica WHERE expediente_id = ?", expediente_id)
        evaluation_row = cursor.fetchone()
        evaluation = _row_dict(cursor, evaluation_row) if evaluation_row else None
        actor_evaluations = _actor_evaluations(cursor, expediente_id)
        grade_calculation = _actor_grade_calculation(cursor, expediente_id, configuration)
        cursor.execute("SELECT * FROM ops.resultado_vinculacion WHERE expediente_id = ?", expediente_id)
        vinculation_result_row = cursor.fetchone()
        vinculation_result = _row_dict(cursor, vinculation_result_row) if vinculation_result_row else None
        cursor.execute(
            "SELECT * FROM ops.producto_vinculacion WHERE expediente_id = ? ORDER BY producto_id",
            expediente_id,
        )
        vinculation_products = _fetch_all(cursor)
        cursor.execute(
            "SELECT TOP (50) * FROM ops.historial_calificacion WHERE expediente_id = ? ORDER BY fecha DESC, historial_id DESC",
            expediente_id,
        )
        grade_history = _fetch_all(cursor)
        cursor.execute(
            "SELECT TOP (20) * FROM ops.reapertura_expediente WHERE expediente_id = ? ORDER BY fecha DESC, reapertura_id DESC",
            expediente_id,
        )
        reopenings = _fetch_all(cursor)
        responsible = _principal_responsible(cursor, expediente_id)
        documents = _required_documents_status(cursor, expediente_id, process)
        reconciliation = _reconciliation_status(cursor, expediente_id, process)
        cursor.execute(
            """
            SELECT
                ISNULL(SUM(horas), 0) AS horas_registradas,
                ISNULL(SUM(CASE WHEN estado_revision = N'VALIDADO' THEN horas ELSE 0 END), 0) AS horas_validadas,
                COUNT_BIG(*) AS actividades,
                SUM(CASE WHEN estado_revision = N'PENDIENTE' THEN 1 ELSE 0 END) AS pendientes
            FROM ops.registro_actividad WHERE expediente_id = ?
            """,
            expediente_id,
        )
        summary = _row_dict(cursor, cursor.fetchone())
        summary["horas_requeridas"] = float(configuration["horas_requeridas"])
        document_compliance = _document_compliance_summary(documents)
        summary["documentos_requeridos"] = document_compliance["required"]
        summary["documentos_cargados"] = document_compliance["loaded"]
        summary["documentos_validados"] = document_compliance["validated"]
        summary["documentos_pendientes"] = document_compliance["pending_upload"]
        summary["avance_documental_porcentaje"] = document_compliance["upload_percentage"]
        summary["avance_validacion_documental_porcentaje"] = document_compliance["validation_percentage"]
        requirements = _process_requirements(
            process=process,
            responsible=responsible,
            plan=plan,
            documents=documents,
            summary=summary,
            closure=closure,
            reconciliation=reconciliation,
            evaluation=evaluation,
            indicators=indicators,
            configuration=configuration,
            actor_evaluations=actor_evaluations,
            vinculation_result=vinculation_result,
            vinculation_products=vinculation_products,
        )
        completed_requirements = sum(1 for item in requirements if item["completo"])
        summary["avance_porcentaje"] = round((completed_requirements / len(requirements)) * 100, 2)
        summary["requisitos_completos"] = completed_requirements
        summary["requisitos_totales"] = len(requirements)
        conn.commit()
    return {
        "expediente": expediente,
        "tipo_proceso_codigo": process,
        "responsable": responsible,
        "plan": plan,
        "actividades": activities,
        "indicadores": indicators,
        "resultado_vinculacion": vinculation_result,
        "productos_vinculacion": vinculation_products,
        "evaluaciones_actores": actor_evaluations,
        "calculo_calificacion": grade_calculation,
        "evaluacion": evaluation,
        "historial_calificacion": grade_history,
        "reaperturas": reopenings,
        "cierre": closure,
        "documentos": documents,
        "conciliacion_titulacion": reconciliation,
        "requisitos": requirements,
        "resumen": summary,
        "configuracion": configuration,
        "almacenamiento": {
            "base_datos": database_name,
            "tabla_calificacion": "ops.evaluacion_practica",
            "tabla_historial": "ops.historial_calificacion",
        },
        "permisos": {
            "puede_editar_plan": current_user.rol in _ADMIN_ROLES or current_user.rol == "DOCENTE",
            "puede_registrar_actividad": current_user.rol in _ADMIN_ROLES or current_user.rol == "ESTUDIANTE",
            "puede_revisar_actividad": current_user.rol in _ADMIN_ROLES or current_user.rol == "DOCENTE",
            "puede_enviar_revision": current_user.rol in _ADMIN_ROLES or current_user.rol in {"DOCENTE", "ESTUDIANTE"},
            "puede_calificar": current_user.rol in _ADMIN_ROLES or current_user.rol == "DOCENTE",
            "puede_cerrar": current_user.rol in _ADMIN_ROLES or current_user.rol == "DOCENTE",
            "puede_registrar_evaluacion_actor": current_user.rol in _ADMIN_ROLES or current_user.rol in {"DOCENTE", "ESTUDIANTE"},
            "puede_registrar_resultado": current_user.rol in _ADMIN_ROLES or current_user.rol in {"DOCENTE", "ESTUDIANTE"},
            "puede_reabrir": current_user.rol in _ADMIN_ROLES,
        },
    }


@router.put("/expedientes/{expediente_id}/plan")
def save_plan(
    expediente_id: int,
    payload: PlanPayload,
    current_user: Annotated[SessionUser, Depends(_RESPONSIBLE_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user, responsible_only=True)
        _ensure_expediente_open(cursor, expediente_id)
        process = _process_code(expediente.get("tipo_proceso_codigo"))
        configuration = _configuration_for_expediente(cursor, expediente, process)
        required_hours = float(configuration["horas_requeridas"])
        if process == "PPF" and not payload.entidad_id:
            raise HTTPException(status_code=400, detail="Selecciona la entidad receptora de la práctica laboral.")
        if process == "PPF" and not payload.convenio_id:
            raise HTTPException(status_code=400, detail="Selecciona un convenio vigente para la práctica laboral.")
        if process == "VIN" and not payload.proyecto_id:
            raise HTTPException(status_code=400, detail="Selecciona el proyecto de vinculación.")
        if payload.estado != "BORRADOR":
            missing_fields: list[str] = []
            if not _clean(payload.objetivo_general):
                missing_fields.append("objetivo general")
            if not _clean(payload.actividades_planificadas):
                missing_fields.append("actividades planificadas")
            if not payload.fecha_inicio or not payload.fecha_fin:
                missing_fields.append("fechas de ejecución")
            if payload.horas_planificadas < required_hours:
                missing_fields.append(f"mínimo {required_hours:g} horas")
            if process == "PPF" and not _clean(payload.tutor_externo_nombre):
                missing_fields.append("tutor externo")
            if missing_fields:
                raise HTTPException(
                    status_code=400,
                    detail=f"Completa el plan antes de aprobarlo: {', '.join(missing_fields)}.",
                )
        if payload.entidad_id:
            cursor.execute("SELECT activo FROM ops.entidad_receptora WHERE entidad_id = ?", payload.entidad_id)
            entity_row = cursor.fetchone()
            if not entity_row or not bool(entity_row[0]):
                raise HTTPException(status_code=400, detail="La entidad receptora no existe o está inactiva.")
        if payload.convenio_id:
            cursor.execute(
                """
                SELECT entidad_id, tipo_proceso_codigo, fecha_inicio, fecha_fin, activo
                FROM ops.convenio_institucional WHERE convenio_id = ?
                """,
                payload.convenio_id,
            )
            agreement = cursor.fetchone()
            if not agreement or not bool(agreement[4]):
                raise HTTPException(status_code=400, detail="El convenio seleccionado no existe o está inactivo.")
            if _clean(agreement[1]).upper() != process:
                raise HTTPException(status_code=400, detail="El convenio no corresponde al tipo de proceso.")
            if payload.entidad_id and int(agreement[0]) != payload.entidad_id:
                raise HTTPException(status_code=400, detail="El convenio no pertenece a la entidad seleccionada.")
            if payload.fecha_inicio and payload.fecha_fin and (
                payload.fecha_inicio < agreement[2] or payload.fecha_fin > agreement[3]
            ):
                raise HTTPException(status_code=400, detail="Las fechas del plan deben estar dentro de la vigencia del convenio.")
        if payload.proyecto_id:
            cursor.execute(
                """
                SELECT fecha_inicio, fecha_fin, activo
                FROM ops.proyecto_vinculacion WHERE proyecto_id = ?
                """,
                payload.proyecto_id,
            )
            project = cursor.fetchone()
            if not project or not bool(project[2]):
                raise HTTPException(status_code=400, detail="El proyecto seleccionado no existe o está inactivo.")
            if payload.fecha_inicio and payload.fecha_fin and (
                payload.fecha_inicio < project[0] or payload.fecha_fin > project[1]
            ):
                raise HTTPException(status_code=400, detail="Las fechas del plan deben estar dentro de la vigencia del proyecto.")
        cursor.execute("SELECT plan_id FROM ops.plan_proceso WHERE expediente_id = ?", expediente_id)
        existing = cursor.fetchone()
        values = (
            payload.entidad_id,
            payload.convenio_id,
            payload.proyecto_id,
            payload.tutor_externo_nombre,
            payload.tutor_externo_correo,
            payload.tutor_externo_telefono,
            payload.objetivo_general,
            payload.resultados_aprendizaje,
            payload.actividades_planificadas,
            payload.fecha_inicio,
            payload.fecha_fin,
            payload.horas_planificadas,
            payload.estado,
        )
        if existing:
            plan_id = int(existing[0])
            cursor.execute(
                """
                UPDATE ops.plan_proceso
                SET entidad_id = ?, convenio_id = ?, proyecto_id = ?, tutor_externo_nombre = ?,
                    tutor_externo_correo = ?, tutor_externo_telefono = ?, objetivo_general = ?,
                    resultados_aprendizaje = ?, actividades_planificadas = ?, fecha_inicio = ?,
                    fecha_fin = ?, horas_planificadas = ?, estado = ?, usuario_modifica = ?,
                    fecha_modifica = SYSDATETIME()
                WHERE plan_id = ?
                """,
                *values,
                _current_login(current_user),
                plan_id,
            )
            action = "UPDATE"
        else:
            cursor.execute(
                """
                INSERT INTO ops.plan_proceso (
                    expediente_id, tipo_proceso_codigo, entidad_id, convenio_id, proyecto_id,
                    tutor_externo_nombre, tutor_externo_correo, tutor_externo_telefono,
                    objetivo_general, resultados_aprendizaje, actividades_planificadas,
                    fecha_inicio, fecha_fin, horas_planificadas, estado, usuario_registro
                ) OUTPUT INSERTED.plan_id
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                expediente_id,
                process,
                *values,
                _current_login(current_user),
            )
            plan_id = int(cursor.fetchone()[0])
            action = "INSERT"
        write_operations_audit(cursor, entity="PLAN_PROCESO", entity_id=plan_id, action=action, user=_current_login(current_user), detail=f"Expediente {expediente_id}")
        conn.commit()
    return {"ok": True, "plan_id": plan_id, "message": "Plan operativo guardado correctamente."}


@router.post("/expedientes/{expediente_id}/actividades")
def save_activity(
    expediente_id: int,
    payload: ActivityPayload,
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user)
        cursor.execute("SELECT estado FROM ops.plan_proceso WHERE expediente_id = ?", expediente_id)
        plan_row = cursor.fetchone()
        plan_state = _clean(plan_row[0] if plan_row else "").upper()
        if plan_state not in {"APROBADO", "EN_EJECUCION"}:
            raise HTTPException(
                status_code=409,
                detail="El responsable debe aprobar el plan antes de registrar actividades.",
            )
        _ensure_expediente_open(cursor, expediente_id)
        start = expediente.get("fecha_inicio")
        end = expediente.get("fecha_fin")
        if isinstance(start, date) and payload.fecha_actividad < start:
            raise HTTPException(status_code=400, detail="La actividad no puede ser anterior al inicio del expediente.")
        if isinstance(end, date) and payload.fecha_actividad > end:
            raise HTTPException(status_code=400, detail="La actividad no puede ser posterior al cierre del expediente.")
        activity_hours = float(payload.horas or 0)
        cursor.execute(
            "SELECT ISNULL(SUM(horas), 0) FROM ops.registro_actividad WHERE expediente_id = ? AND fecha_actividad = ? AND estado_revision <> N'RECHAZADO'",
            expediente_id,
            payload.fecha_actividad,
        )
        day_hours = float(cursor.fetchone()[0] or 0)
        if day_hours + activity_hours > 24:
            raise HTTPException(status_code=400, detail="La suma de horas del día no puede superar 24.")
        if payload.hora_inicio and payload.hora_fin:
            cursor.execute(
                """
                SELECT TOP (1) actividad_id
                FROM ops.registro_actividad WITH (UPDLOCK, HOLDLOCK)
                WHERE expediente_id = ? AND fecha_actividad = ?
                  AND estado_revision <> N'RECHAZADO'
                  AND hora_inicio IS NOT NULL AND hora_fin IS NOT NULL
                  AND hora_inicio < ? AND hora_fin > ?
                """,
                expediente_id,
                payload.fecha_actividad,
                payload.hora_fin,
                payload.hora_inicio,
            )
            if cursor.fetchone():
                raise HTTPException(status_code=409, detail="La jornada se cruza con otra actividad registrada ese día.")
        cursor.execute(
            """
            INSERT INTO ops.registro_actividad (
                expediente_id, fecha_actividad, descripcion, horas, hora_inicio, hora_fin,
                descanso_minutos, modalidad, lugar, origen_horas, evidencia_url,
                evidencia_nombre, usuario_registro
            ) OUTPUT INSERTED.actividad_id
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            expediente_id,
            payload.fecha_actividad,
            payload.descripcion,
            activity_hours,
            payload.hora_inicio,
            payload.hora_fin,
            payload.descanso_minutos,
            payload.modalidad,
            payload.lugar,
            "JORNADA_CALCULADA" if payload.hora_inicio else "MANUAL",
            payload.evidencia_url,
            payload.evidencia_nombre,
            _current_login(current_user),
        )
        activity_id = int(cursor.fetchone()[0])
        cursor.execute(
            """
            UPDATE ops.plan_proceso
            SET estado = N'EN_EJECUCION', usuario_modifica = ?, fecha_modifica = SYSDATETIME()
            WHERE expediente_id = ? AND estado = N'APROBADO'
            """,
            _current_login(current_user),
            expediente_id,
        )
        write_operations_audit(cursor, entity="REGISTRO_ACTIVIDAD", entity_id=activity_id, action="INSERT", user=_current_login(current_user), detail=f"Expediente {expediente_id}; {activity_hours} hora(s)")
        conn.commit()
    return {
        "ok": True,
        "actividad_id": activity_id,
        "horas_calculadas": activity_hours,
        "message": "Actividad registrada en la bitácora de la base complementaria.",
    }


@router.put("/actividades/{actividad_id}/revision")
def review_activity(
    actividad_id: int,
    payload: ActivityReviewPayload,
    current_user: Annotated[SessionUser, Depends(_RESPONSIBLE_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute("SELECT expediente_id FROM ops.registro_actividad WHERE actividad_id = ?", actividad_id)
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="No existe la actividad solicitada.")
        expediente_id = int(row[0])
        _expediente_access(cursor, expediente_id, current_user, responsible_only=True)
        _ensure_expediente_open(cursor, expediente_id)
        cursor.execute(
            """
            UPDATE ops.registro_actividad
            SET estado_revision = ?, observacion_revision = ?, revisado_por = ?,
                fecha_revision = SYSDATETIME(), usuario_modifica = ?, fecha_modifica = SYSDATETIME()
            WHERE actividad_id = ?
            """,
            payload.estado_revision,
            payload.observacion_revision,
            _current_login(current_user),
            _current_login(current_user),
            actividad_id,
        )
        cursor.execute(
            """
            SELECT ISNULL(SUM(CASE WHEN estado_revision = N'VALIDADO' THEN horas ELSE 0 END), 0)
            FROM ops.registro_actividad
            WHERE expediente_id = ?
            """,
            expediente_id,
        )
        validated_hours = float(cursor.fetchone()[0] or 0)
        cursor.execute(
            """
            UPDATE pp.expediente_practica
            SET horas_asistencia_validadas = ?, fecha_modifica = SYSDATETIME(), usuario_modifica = ?
            WHERE expediente_id = ?
            """,
            validated_hours,
            _current_login(current_user),
            expediente_id,
        )
        write_operations_audit(cursor, entity="REGISTRO_ACTIVIDAD", entity_id=actividad_id, action="REVISION", user=_current_login(current_user), detail=payload.estado_revision)
        conn.commit()
    return {"ok": True, "message": "Actividad revisada correctamente."}


@router.put("/expedientes/{expediente_id}/indicadores")
def save_indicator(
    expediente_id: int,
    payload: IndicatorPayload,
    current_user: Annotated[SessionUser, Depends(_RESPONSIBLE_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user, responsible_only=True)
        _ensure_expediente_open(cursor, expediente_id)
        if _process_code(expediente.get("tipo_proceso_codigo")) != "VIN":
            raise HTTPException(status_code=400, detail="Los indicadores corresponden únicamente a Vinculación.")
        if payload.indicador_id:
            cursor.execute(
                """
                UPDATE ops.meta_indicador
                SET nombre = ?, unidad_medida = ?, meta = ?, resultado = ?, evidencia_url = ?,
                    observacion = ?, usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE indicador_id = ? AND expediente_id = ?
                """,
                payload.nombre,
                payload.unidad_medida,
                payload.meta,
                payload.resultado,
                payload.evidencia_url,
                payload.observacion,
                _current_login(current_user),
                payload.indicador_id,
                expediente_id,
            )
            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="No existe el indicador solicitado.")
            indicator_id = payload.indicador_id
            action = "UPDATE"
        else:
            cursor.execute(
                """
                INSERT INTO ops.meta_indicador (
                    expediente_id, nombre, unidad_medida, meta, resultado,
                    evidencia_url, observacion, usuario_registro
                ) OUTPUT INSERTED.indicador_id
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                expediente_id,
                payload.nombre,
                payload.unidad_medida,
                payload.meta,
                payload.resultado,
                payload.evidencia_url,
                payload.observacion,
                _current_login(current_user),
            )
            indicator_id = int(cursor.fetchone()[0])
            action = "INSERT"
        write_operations_audit(cursor, entity="INDICADOR_VINCULACION", entity_id=indicator_id, action=action, user=_current_login(current_user), detail=payload.nombre)
        conn.commit()
    return {"ok": True, "indicador_id": indicator_id, "message": "Indicador guardado correctamente."}


@router.put("/expedientes/{expediente_id}/evaluaciones-actores/{rol_evaluador}")
def save_actor_evaluation(
    expediente_id: int,
    rol_evaluador: str,
    payload: ActorEvaluationPayload,
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    role = _actor_role(rol_evaluador)
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user)
        _ensure_expediente_open(cursor, expediente_id)
        process = _process_code(expediente.get("tipo_proceso_codigo"))
        configuration = _configuration_for_expediente(cursor, expediente, process)
        if current_user.rol == "ESTUDIANTE" and role != "AUTOEVALUACION":
            raise HTTPException(status_code=403, detail="El estudiante solo puede registrar su autoevaluación.")
        if current_user.rol == "DOCENTE" and role == "AUTOEVALUACION":
            raise HTTPException(status_code=403, detail="La autoevaluación corresponde al estudiante.")
        if role == "TUTOR_EMPRESARIAL" and process != "PPF":
            raise HTTPException(status_code=400, detail="La evaluación del tutor empresarial corresponde a prácticas laborales.")

        weight_key = {
            "DOCENTE_ACADEMICO": "peso_docente",
            "TUTOR_EMPRESARIAL": "peso_tutor",
            "AUTOEVALUACION": "peso_autoevaluacion",
        }[role]
        evaluator_name = _clean(payload.evaluador_nombre)
        evaluator_email = _clean(payload.evaluador_correo)
        if role == "TUTOR_EMPRESARIAL":
            cursor.execute(
                "SELECT tutor_externo_nombre, tutor_externo_correo FROM ops.plan_proceso WHERE expediente_id = ?",
                expediente_id,
            )
            tutor_row = cursor.fetchone()
            evaluator_name = evaluator_name or _clean(tutor_row[0] if tutor_row else "")
            evaluator_email = evaluator_email or _clean(tutor_row[1] if tutor_row else "")
            if not evaluator_name:
                raise HTTPException(status_code=400, detail="Identifique al tutor empresarial que emitió la evaluación.")
        else:
            evaluator_name = evaluator_name or _clean(current_user.nombres) or login
            evaluator_email = evaluator_email or _clean(current_user.email)

        state = "REGISTRADA" if current_user.rol == "ESTUDIANTE" else "VALIDADA"
        cursor.execute(
            """
            SELECT evaluacion_actor_id
            FROM ops.evaluacion_actor WITH (UPDLOCK, HOLDLOCK)
            WHERE expediente_id = ? AND rol_evaluador = ?
            """,
            expediente_id,
            role,
        )
        existing = cursor.fetchone()
        values = (
            payload.calificacion,
            float(configuration.get(weight_key) or 0),
            evaluator_name or None,
            evaluator_email or None,
            payload.observacion,
            payload.evidencia_url,
            state,
            login if state == "VALIDADA" else None,
        )
        if existing:
            actor_evaluation_id = int(existing[0])
            cursor.execute(
                """
                UPDATE ops.evaluacion_actor
                SET calificacion = ?, peso = ?, evaluador_nombre = ?, evaluador_correo = ?,
                    observacion = ?, evidencia_url = ?, estado = ?, validado_por = ?,
                    fecha_validacion = CASE WHEN ? = N'VALIDADA' THEN SYSDATETIME() ELSE NULL END,
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE evaluacion_actor_id = ?
                """,
                *values,
                state,
                login,
                actor_evaluation_id,
            )
            action = "UPDATE"
        else:
            cursor.execute(
                """
                INSERT INTO ops.evaluacion_actor (
                    expediente_id, rol_evaluador, calificacion, peso, evaluador_nombre,
                    evaluador_correo, observacion, evidencia_url, estado, validado_por,
                    fecha_validacion, usuario_registro
                ) OUTPUT INSERTED.evaluacion_actor_id
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                        CASE WHEN ? = N'VALIDADA' THEN SYSDATETIME() ELSE NULL END, ?)
                """,
                expediente_id,
                role,
                *values,
                state,
                login,
            )
            actor_evaluation_id = int(cursor.fetchone()[0])
            action = "INSERT"
        write_operations_audit(
            cursor,
            entity="EVALUACION_ACTOR",
            entity_id=actor_evaluation_id,
            action=action,
            user=login,
            detail=f"{role}; {payload.calificacion:.2f}/10; expediente {expediente_id}",
        )
        calculation = _actor_grade_calculation(cursor, expediente_id, configuration)
        conn.commit()
    return {
        "ok": True,
        "evaluacion_actor_id": actor_evaluation_id,
        "calculo_calificacion": calculation,
        "message": "Evaluación registrada en la base complementaria de prácticas.",
    }


@router.put("/expedientes/{expediente_id}/resultado-vinculacion")
def save_vinculation_result(
    expediente_id: int,
    payload: VinculationResultPayload,
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user)
        _ensure_expediente_open(cursor, expediente_id)
        if _process_code(expediente.get("tipo_proceso_codigo")) != "VIN":
            raise HTTPException(status_code=400, detail="Los resultados de impacto corresponden únicamente a Vinculación.")
        if payload.validar and current_user.rol not in _ADMIN_ROLES and current_user.rol != "DOCENTE":
            raise HTTPException(status_code=403, detail="Solo el responsable puede validar los resultados de vinculación.")
        cursor.execute(
            "SELECT resultado_vinculacion_id, estado FROM ops.resultado_vinculacion WITH (UPDLOCK, HOLDLOCK) WHERE expediente_id = ?",
            expediente_id,
        )
        existing = cursor.fetchone()
        if existing and _clean(existing[1]).upper() == "VALIDADO" and current_user.rol == "ESTUDIANTE":
            raise HTTPException(status_code=409, detail="El resultado ya fue validado y debe ser observado por el responsable antes de editarlo.")
        state = "VALIDADO" if payload.validar else "REGISTRADO"
        if existing:
            result_id = int(existing[0])
            cursor.execute(
                """
                UPDATE ops.resultado_vinculacion
                SET beneficiarios_reales = ?, resumen_impacto = ?, observacion = ?, evidencia_url = ?,
                    estado = ?, validado_por = CASE WHEN ? = N'VALIDADO' THEN ? ELSE NULL END,
                    fecha_validacion = CASE WHEN ? = N'VALIDADO' THEN SYSDATETIME() ELSE NULL END,
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE resultado_vinculacion_id = ?
                """,
                payload.beneficiarios_reales,
                payload.resumen_impacto,
                payload.observacion,
                payload.evidencia_url,
                state,
                state,
                login,
                state,
                login,
                result_id,
            )
            action = "UPDATE"
        else:
            cursor.execute(
                """
                INSERT INTO ops.resultado_vinculacion (
                    expediente_id, beneficiarios_reales, resumen_impacto, observacion,
                    evidencia_url, estado, validado_por, fecha_validacion, usuario_registro
                ) OUTPUT INSERTED.resultado_vinculacion_id
                VALUES (?, ?, ?, ?, ?, ?, CASE WHEN ? = N'VALIDADO' THEN ? ELSE NULL END,
                        CASE WHEN ? = N'VALIDADO' THEN SYSDATETIME() ELSE NULL END, ?)
                """,
                expediente_id,
                payload.beneficiarios_reales,
                payload.resumen_impacto,
                payload.observacion,
                payload.evidencia_url,
                state,
                state,
                login,
                state,
                login,
            )
            result_id = int(cursor.fetchone()[0])
            action = "INSERT"
        write_operations_audit(
            cursor,
            entity="RESULTADO_VINCULACION",
            entity_id=result_id,
            action=action,
            user=login,
            detail=f"Expediente {expediente_id}; beneficiarios={payload.beneficiarios_reales}; estado={state}",
        )
        conn.commit()
    return {"ok": True, "resultado_vinculacion_id": result_id, "message": "Resultados de vinculación guardados correctamente."}


@router.put("/expedientes/{expediente_id}/productos-vinculacion")
def save_vinculation_product(
    expediente_id: int,
    payload: VinculationProductPayload,
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user)
        _ensure_expediente_open(cursor, expediente_id)
        if _process_code(expediente.get("tipo_proceso_codigo")) != "VIN":
            raise HTTPException(status_code=400, detail="Los productos corresponden únicamente a Vinculación.")
        if payload.producto_id:
            cursor.execute(
                "SELECT estado_revision FROM ops.producto_vinculacion WHERE producto_id = ? AND expediente_id = ?",
                payload.producto_id,
                expediente_id,
            )
            existing = cursor.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="No existe el producto solicitado.")
            if _clean(existing[0]).upper() == "VALIDADO" and current_user.rol == "ESTUDIANTE":
                raise HTTPException(status_code=409, detail="El producto validado no puede ser editado por el estudiante.")
            cursor.execute(
                """
                UPDATE ops.producto_vinculacion
                SET nombre = ?, descripcion = ?, cantidad = ?, unidad_medida = ?, evidencia_url = ?,
                    estado_revision = N'PENDIENTE', observacion_revision = NULL,
                    revisado_por = NULL, fecha_revision = NULL,
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE producto_id = ? AND expediente_id = ?
                """,
                payload.nombre,
                payload.descripcion,
                payload.cantidad,
                payload.unidad_medida,
                payload.evidencia_url,
                login,
                payload.producto_id,
                expediente_id,
            )
            product_id = payload.producto_id
            action = "UPDATE"
        else:
            cursor.execute(
                """
                INSERT INTO ops.producto_vinculacion (
                    expediente_id, nombre, descripcion, cantidad, unidad_medida,
                    evidencia_url, usuario_registro
                ) OUTPUT INSERTED.producto_id
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                expediente_id,
                payload.nombre,
                payload.descripcion,
                payload.cantidad,
                payload.unidad_medida,
                payload.evidencia_url,
                login,
            )
            product_id = int(cursor.fetchone()[0])
            action = "INSERT"
        write_operations_audit(
            cursor,
            entity="PRODUCTO_VINCULACION",
            entity_id=product_id,
            action=action,
            user=login,
            detail=f"Expediente {expediente_id}; {payload.nombre}",
        )
        conn.commit()
    return {"ok": True, "producto_id": product_id, "message": "Producto de vinculación guardado correctamente."}


@router.put("/productos-vinculacion/{producto_id}/revision")
def review_vinculation_product(
    producto_id: int,
    payload: VinculationProductReviewPayload,
    current_user: Annotated[SessionUser, Depends(_RESPONSIBLE_ACCESS)],
) -> dict[str, Any]:
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute("SELECT expediente_id FROM ops.producto_vinculacion WHERE producto_id = ?", producto_id)
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="No existe el producto solicitado.")
        expediente_id = int(row[0])
        _expediente_access(cursor, expediente_id, current_user, responsible_only=True)
        _ensure_expediente_open(cursor, expediente_id)
        cursor.execute(
            """
            UPDATE ops.producto_vinculacion
            SET estado_revision = ?, observacion_revision = ?, revisado_por = ?,
                fecha_revision = SYSDATETIME(), usuario_modifica = ?, fecha_modifica = SYSDATETIME()
            WHERE producto_id = ?
            """,
            payload.estado_revision,
            payload.observacion_revision,
            login,
            login,
            producto_id,
        )
        write_operations_audit(
            cursor,
            entity="PRODUCTO_VINCULACION",
            entity_id=producto_id,
            action="REVISION",
            user=login,
            detail=payload.estado_revision,
        )
        conn.commit()
    return {"ok": True, "message": "Producto de vinculación revisado correctamente."}


@router.put("/expedientes/{expediente_id}/evaluacion")
def save_final_evaluation(
    expediente_id: int,
    payload: FinalEvaluationPayload,
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    action = payload.accion.upper()
    observation = _clean(payload.observacion) or None
    responsible_action = action != "ENVIAR_REVISION"
    login = _current_login(current_user)
    result = "PENDIENTE"
    database_name = ""
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        database_name = assert_practices_database(cursor)
        expediente = _expediente_access(
            cursor,
            expediente_id,
            current_user,
            responsible_only=responsible_action,
        )
        _ensure_expediente_open(cursor, expediente_id, allow_evaluation_transition=True)
        process = _process_code(expediente.get("tipo_proceso_codigo"))
        configuration = _configuration_for_expediente(cursor, expediente, process)
        minimum_grade = float(configuration["nota_minima_aprobacion"])
        cursor.execute(
            "SELECT evaluacion_id, estado, resultado FROM ops.evaluacion_practica WHERE expediente_id = ?",
            expediente_id,
        )
        evaluation_row = cursor.fetchone()
        if evaluation_row:
            evaluation_id = int(evaluation_row[0])
            current_state = _clean(evaluation_row[1]).upper()
        else:
            cursor.execute(
                """
                INSERT INTO ops.evaluacion_practica (expediente_id, nota_minima_aprobacion, usuario_registro)
                OUTPUT INSERTED.evaluacion_id
                VALUES (?, ?, ?)
                """,
                expediente_id,
                minimum_grade,
                login,
            )
            evaluation_id = int(cursor.fetchone()[0])
            current_state = "PENDIENTE_REVISION"

        if action == "ENVIAR_REVISION":
            if current_state in {"PENDIENTE_CALIFICACION", "CALIFICADA"}:
                raise HTTPException(
                    status_code=409,
                    detail="La revisión ya terminó y el expediente no puede volver a enviarse.",
                )
            errors = _evaluation_submission_errors(cursor, expediente_id, process, configuration)
            if errors:
                raise HTTPException(status_code=409, detail=" ".join(errors))
            cursor.execute(
                """
                UPDATE ops.evaluacion_practica
                SET estado = N'EN_REVISION', calificacion = NULL, resultado = N'PENDIENTE',
                    nota_minima_aprobacion = ?, origen_calificacion = NULL, detalle_calculo = NULL,
                    observacion_revision = ?, observacion_calificacion = NULL,
                    enviado_por = ?, fecha_envio_revision = SYSDATETIME(),
                    revisado_por = NULL, fecha_revision = NULL,
                    calificado_por = NULL, fecha_calificacion = NULL,
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE evaluacion_id = ?
                """,
                minimum_grade,
                observation,
                login,
                login,
                evaluation_id,
            )
            update_compliance_enrollment_status(
                cursor,
                expediente_id=expediente_id,
                state="EN_REVISION",
                user=login,
            )
            message = "Expediente enviado a revisión correctamente."
        elif action == "DEVOLVER":
            if current_state not in {"EN_REVISION", "PENDIENTE_CALIFICACION"}:
                raise HTTPException(status_code=409, detail="El expediente no se encuentra en una etapa revisable.")
            cursor.execute(
                """
                UPDATE ops.evaluacion_practica
                SET estado = N'PENDIENTE_REVISION', calificacion = NULL, resultado = N'PENDIENTE',
                    nota_minima_aprobacion = ?, origen_calificacion = NULL, detalle_calculo = NULL,
                    observacion_revision = ?, observacion_calificacion = NULL,
                    revisado_por = ?, fecha_revision = SYSDATETIME(),
                    calificado_por = NULL, fecha_calificacion = NULL,
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE evaluacion_id = ?
                """,
                minimum_grade,
                observation,
                login,
                login,
                evaluation_id,
            )
            update_compliance_enrollment_status(
                cursor,
                expediente_id=expediente_id,
                state="EN_PROCESO",
                user=login,
            )
            message = "Expediente devuelto para corrección."
        elif action == "HABILITAR_CALIFICACION":
            if current_state != "EN_REVISION":
                raise HTTPException(status_code=409, detail="El expediente debe estar en revisión antes de habilitar la calificación.")
            errors = _evaluation_grading_errors(cursor, expediente_id, process, configuration)
            if errors:
                raise HTTPException(status_code=409, detail=" ".join(errors))
            cursor.execute(
                """
                UPDATE ops.evaluacion_practica
                SET estado = N'PENDIENTE_CALIFICACION', calificacion = NULL, resultado = N'PENDIENTE',
                    nota_minima_aprobacion = ?, origen_calificacion = NULL, detalle_calculo = NULL,
                    observacion_revision = COALESCE(?, observacion_revision),
                    revisado_por = ?, fecha_revision = SYSDATETIME(),
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE evaluacion_id = ?
                """,
                minimum_grade,
                observation,
                login,
                login,
                evaluation_id,
            )
            update_compliance_enrollment_status(
                cursor,
                expediente_id=expediente_id,
                state="EN_REVISION",
                user=login,
            )
            message = "Revisión finalizada; el expediente está a la espera de calificación."
        else:
            if current_state != "PENDIENTE_CALIFICACION":
                raise HTTPException(status_code=409, detail="La calificación aún no está habilitada para este expediente.")
            errors = _evaluation_grading_errors(cursor, expediente_id, process, configuration)
            if errors:
                raise HTTPException(status_code=409, detail=" ".join(errors))
            grade_calculation = _actor_grade_calculation(cursor, expediente_id, configuration)
            calculated_grade = grade_calculation.get("calificacion_calculada")
            if grade_calculation.get("usa_evaluaciones_actores"):
                if calculated_grade is None:
                    raise HTTPException(status_code=409, detail="No se pudo calcular la calificación con las evaluaciones obligatorias.")
                grade = float(calculated_grade)
                grade_origin = "PROMEDIO_PONDERADO_ACTORES"
                calculation_detail = json.dumps(grade_calculation, ensure_ascii=True, default=str)
            else:
                if payload.calificacion is None:
                    raise HTTPException(status_code=400, detail="Ingrese la calificación final sobre 10.")
                grade = float(payload.calificacion)
                grade_origin = "MANUAL_RESPONSABLE"
                calculation_detail = json.dumps({"calificacion_manual": grade}, ensure_ascii=True)
            if grade < minimum_grade and not observation:
                raise HTTPException(status_code=400, detail="Registre una observación para justificar la reprobación.")
            result = _grade_result(grade, minimum_grade)
            cursor.execute(
                """
                UPDATE ops.evaluacion_practica
                SET estado = N'CALIFICADA', calificacion = ?, nota_minima_aprobacion = ?,
                    resultado = ?, origen_calificacion = ?, detalle_calculo = ?, observacion_calificacion = ?,
                    calificado_por = ?, fecha_calificacion = SYSDATETIME(),
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE evaluacion_id = ?
                """,
                grade,
                minimum_grade,
                result,
                grade_origin,
                calculation_detail,
                observation,
                login,
                login,
                evaluation_id,
            )
            new_state_code = "APROBADO" if result == "APROBADO" else "REPROBADO"
            new_state_id = _catalog_state_id(
                cursor,
                "cat.estado_expediente",
                "estado_expediente_id",
                new_state_code,
            )
            cursor.execute(
                """
                SELECT ISNULL(SUM(CASE WHEN estado_revision = N'VALIDADO' THEN horas ELSE 0 END), 0)
                FROM ops.registro_actividad
                WHERE expediente_id = ?
                """,
                expediente_id,
            )
            validated_hours = float(cursor.fetchone()[0] or 0)
            cursor.execute(
                """
                UPDATE pp.expediente_practica
                SET horas_requeridas = ?, horas_reconocidas = ?, horas_asistencia_validadas = ?,
                    estado_expediente_id = ?, fecha_fin = COALESCE(fecha_fin, CONVERT(date, SYSDATETIME())),
                    observacion = COALESCE(?, observacion),
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE expediente_id = ?
                """,
                float(configuration["horas_requeridas"]),
                validated_hours,
                validated_hours,
                new_state_id,
                observation,
                login,
                expediente_id,
            )
            if int(expediente.get("estado_expediente_id") or 0) != new_state_id:
                cursor.execute(
                    """
                    INSERT INTO pp.historial_estado_expediente (
                        expediente_id, estado_anterior_id, estado_nuevo_id, motivo, usuario_registro
                    ) VALUES (?, ?, ?, ?, ?)
                    """,
                    expediente_id,
                    expediente.get("estado_expediente_id"),
                    new_state_id,
                    observation or f"Resultado de calificación final: {result.lower()} ({grade:.2f}/10).",
                    login,
                )
            update_compliance_enrollment_status(
                cursor,
                expediente_id=expediente_id,
                state="EN_REVISION" if result == "APROBADO" else "NO_CUMPLIDO",
                user=login,
            )
            message = f"Calificación registrada: {grade:.2f}/10, resultado {result.lower()}."

        record_evaluation_history(
            cursor,
            evaluation_id=evaluation_id,
            action=action,
            user=login,
            observation=observation,
        )
        write_operations_audit(
            cursor,
            entity="EVALUACION_PRACTICA",
            entity_id=evaluation_id,
            action=action,
            user=login,
            detail=observation or result,
        )
        cursor.execute("SELECT * FROM ops.evaluacion_practica WHERE evaluacion_id = ?", evaluation_id)
        evaluation = _row_dict(cursor, cursor.fetchone())
        conn.commit()
    return {
        "ok": True,
        "evaluacion_id": evaluation_id,
        "evaluacion": evaluation,
        "message": message,
        "almacenamiento": f"{database_name}.ops.evaluacion_practica",
    }


@router.put("/expedientes/{expediente_id}/cierre")
def save_closure(
    expediente_id: int,
    payload: ClosurePayload,
    current_user: Annotated[SessionUser, Depends(_RESPONSIBLE_ACCESS)],
) -> dict[str, Any]:
    final_result = "PENDIENTE"
    final_grade: float | None = None
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user, responsible_only=True)
        process = _process_code(expediente.get("tipo_proceso_codigo"))
        configuration = _configuration_for_expediente(cursor, expediente, process)
        required_hours = float(configuration["horas_requeridas"])
        cursor.execute(
            "SELECT estado, calificacion, resultado FROM ops.evaluacion_practica WHERE expediente_id = ?",
            expediente_id,
        )
        evaluation_row = cursor.fetchone()
        evaluation_state = _clean(evaluation_row[0] if evaluation_row else "").upper()
        final_grade = float(evaluation_row[1]) if evaluation_row and evaluation_row[1] is not None else None
        final_result = _clean(evaluation_row[2] if evaluation_row else "PENDIENTE").upper()
        effective_grade = final_grade if final_grade is not None else payload.evaluacion_entidad
        cursor.execute("SELECT cierre_id, fecha_cierre FROM ops.cierre_proceso WHERE expediente_id = ?", expediente_id)
        existing = cursor.fetchone()
        if existing and existing[1] is not None:
            if payload.cerrar:
                conn.commit()
                return {
                    "ok": True,
                    "cierre_id": int(existing[0]),
                    "message": "El proceso ya se encuentra cerrado.",
                    "titulacion": None,
                }
            raise HTTPException(status_code=409, detail="El expediente ya está cerrado y no admite modificaciones.")
        if payload.cerrar:
            if evaluation_state != "CALIFICADA" or final_result not in {"APROBADO", "REPROBADO"}:
                raise HTTPException(
                    status_code=409,
                    detail="El expediente debe completar la revisión y contar con una calificación final antes del cierre.",
                )
            required_checks_complete = (
                payload.supervision_realizada
                and payload.informe_final_validado
                and payload.acta_aceptacion_validada
            )
            if not required_checks_complete or (final_result == "APROBADO" and not payload.certificado_emitido):
                requirement_label = (
                    "supervisión, informe, acta y certificado"
                    if final_result == "APROBADO"
                    else "supervisión, informe y acta"
                )
                raise HTTPException(status_code=400, detail=f"Completa {requirement_label} antes de cerrar.")
            state = _clean(expediente.get("estado_expediente_codigo")).upper()
            accepted_states = (
                {"APROBADO", "VALIDADO", "FINALIZADO", "CERRADO"}
                if final_result == "APROBADO"
                else {"REPROBADO"}
            )
            if state not in accepted_states:
                raise HTTPException(status_code=409, detail="El estado del expediente no coincide con el resultado de la calificación final.")
            cursor.execute("SELECT estado FROM ops.plan_proceso WHERE expediente_id = ?", expediente_id)
            plan_row = cursor.fetchone()
            plan_state = _clean(plan_row[0] if plan_row else "").upper()
            if plan_state not in {"APROBADO", "EN_EJECUCION", "FINALIZADO"}:
                raise HTTPException(status_code=409, detail="El plan de prácticas debe estar aprobado antes del cierre.")
            cursor.execute(
                """
                SELECT
                    ISNULL(SUM(CASE WHEN estado_revision = N'VALIDADO' THEN horas ELSE 0 END), 0),
                    SUM(CASE WHEN estado_revision = N'PENDIENTE' THEN 1 ELSE 0 END)
                FROM ops.registro_actividad
                WHERE expediente_id = ?
                """,
                expediente_id,
            )
            activity_summary = cursor.fetchone()
            validated_hours = float(activity_summary[0] or 0)
            pending_activities = int(activity_summary[1] or 0)
            if validated_hours < required_hours:
                raise HTTPException(
                    status_code=409,
                    detail=f"Se requieren {required_hours:g} horas validadas; actualmente existen {validated_hours:g}.",
                )
            if pending_activities:
                raise HTTPException(status_code=409, detail="Revisa todas las actividades pendientes antes de cerrar.")
            documents = _required_documents_status(cursor, expediente_id, process)
            missing_documents = [
                _clean(item.get("Nombre") or item.get("Codigo"))
                for item in documents
                if not item.get("Validado")
            ]
            if missing_documents:
                raise HTTPException(
                    status_code=409,
                    detail=f"Faltan documentos validados: {', '.join(missing_documents)}.",
                )
            if process == "VIN":
                cursor.execute(
                    """
                    SELECT COUNT_BIG(*),
                           SUM(CASE WHEN resultado IS NULL THEN 1 ELSE 0 END)
                    FROM ops.meta_indicador
                    WHERE expediente_id = ?
                    """,
                    expediente_id,
                )
                indicator_summary = cursor.fetchone()
                indicator_count = int(indicator_summary[0] or 0)
                indicators_without_result = int(indicator_summary[1] or 0)
                if indicator_count == 0 or indicators_without_result:
                    raise HTTPException(
                        status_code=409,
                        detail="Registra las metas e ingresa el resultado de todos los indicadores antes del cierre.",
                    )
        if existing:
            closure_id = int(existing[0])
            cursor.execute(
                """
                UPDATE ops.cierre_proceso
                SET supervision_realizada = ?, evaluacion_entidad = ?, informe_final_validado = ?,
                    acta_aceptacion_validada = ?, certificado_emitido = ?, observacion = ?,
                    fecha_cierre = CASE WHEN ? = 1 THEN SYSDATETIME() ELSE NULL END,
                    cerrado_por = CASE WHEN ? = 1 THEN ? ELSE NULL END,
                    usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE cierre_id = ?
                """,
                payload.supervision_realizada,
                effective_grade,
                payload.informe_final_validado,
                payload.acta_aceptacion_validada,
                payload.certificado_emitido,
                payload.observacion,
                payload.cerrar,
                payload.cerrar,
                _current_login(current_user),
                _current_login(current_user),
                closure_id,
            )
            action = "UPDATE"
        else:
            cursor.execute(
                """
                INSERT INTO ops.cierre_proceso (
                    expediente_id, supervision_realizada, evaluacion_entidad,
                    informe_final_validado, acta_aceptacion_validada, certificado_emitido,
                    observacion, fecha_cierre, cerrado_por, usuario_registro
                ) OUTPUT INSERTED.cierre_id
                VALUES (?, ?, ?, ?, ?, ?, ?, CASE WHEN ? = 1 THEN SYSDATETIME() ELSE NULL END,
                        CASE WHEN ? = 1 THEN ? ELSE NULL END, ?)
                """,
                expediente_id,
                payload.supervision_realizada,
                effective_grade,
                payload.informe_final_validado,
                payload.acta_aceptacion_validada,
                payload.certificado_emitido,
                payload.observacion,
                payload.cerrar,
                payload.cerrar,
                _current_login(current_user),
                _current_login(current_user),
            )
            closure_id = int(cursor.fetchone()[0])
            action = "INSERT"
        write_operations_audit(cursor, entity="CIERRE_PROCESO", entity_id=closure_id, action=action, user=_current_login(current_user), detail="CERRADO" if payload.cerrar else "SEGUIMIENTO")
        if payload.cerrar:
            cursor.execute(
                """
                UPDATE ops.plan_proceso
                SET estado = N'FINALIZADO', usuario_modifica = ?, fecha_modifica = SYSDATETIME()
                WHERE expediente_id = ?
                """,
                _current_login(current_user),
                expediente_id,
            )
            update_compliance_enrollment_status(
                cursor,
                expediente_id=expediente_id,
                state="CUMPLIDO" if final_result == "APROBADO" else "NO_CUMPLIDO",
                user=_current_login(current_user),
            )
        conn.commit()
    titulation: dict[str, Any] | None = None
    if payload.cerrar and final_result == "APROBADO":
        try:
            titulation = _sync_titulacion_completion(expediente_id, process, _current_login(current_user))
        except (pyodbc.Error, RuntimeError) as exc:
            titulation = {
                "sincronizado": False,
                "pendiente": True,
                "motivo": f"El cierre fue guardado, pero la conciliación con Titulación quedó pendiente: {exc}",
            }
        save_titulation_reconciliation(
            expediente_id,
            process,
            _current_login(current_user),
            titulation,
        )
    elif payload.cerrar and final_result == "REPROBADO":
        titulation = {
            "sincronizado": False,
            "pendiente": False,
            "motivo": "El resultado reprobado no habilita el requisito de Titulación.",
        }
    return {
        "ok": True,
        "cierre_id": closure_id,
        "message": (
            "Proceso aprobado, cerrado y conciliado correctamente."
            if payload.cerrar and final_result == "APROBADO"
            else "Proceso reprobado cerrado sin habilitar Titulación."
            if payload.cerrar
            else "Cierre operativo actualizado correctamente."
        ),
        "titulacion": titulation,
    }


@router.post("/expedientes/{expediente_id}/reapertura")
def reopen_operational_record(
    expediente_id: int,
    payload: ReopenPayload,
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
) -> dict[str, Any]:
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        expediente = _expediente_access(cursor, expediente_id, current_user)
        process = _process_code(expediente.get("tipo_proceso_codigo"))
        cursor.execute(
            """
            SELECT e.evaluacion_id, e.estado, e.resultado, e.calificacion, c.fecha_cierre
            FROM ops.evaluacion_practica e
            LEFT JOIN ops.cierre_proceso c ON c.expediente_id = e.expediente_id
            WHERE e.expediente_id = ?
            """,
            expediente_id,
        )
        evaluation_row = cursor.fetchone()
        if not evaluation_row or (_clean(evaluation_row[1]).upper() != "CALIFICADA" and evaluation_row[4] is None):
            raise HTTPException(status_code=409, detail="El expediente todavía está abierto y no requiere reapertura.")
        cursor.execute(
            "SELECT conciliacion_id, estado FROM ops.conciliacion_titulacion WHERE expediente_id = ? AND tipo_proceso_codigo = ?",
            expediente_id,
            process,
        )
        reconciliation = cursor.fetchone()
        titulation_completed = bool(reconciliation and _clean(reconciliation[1]).upper() == "COMPLETADO")
        if titulation_completed and not payload.confirmar_reversion_titulacion:
            raise HTTPException(
                status_code=409,
                detail=(
                    "El cumplimiento ya fue conciliado con Titulación. Confirme la reversión para reabrir; "
                    "la conciliación quedará marcada para revisión administrativa."
                ),
            )

        evaluation_id = int(evaluation_row[0])
        record_evaluation_history(
            cursor,
            evaluation_id=evaluation_id,
            action="ANTES_REAPERTURA",
            user=login,
            observation=payload.motivo,
        )
        cursor.execute(
            """
            INSERT INTO ops.reapertura_expediente (
                expediente_id, evaluacion_id, estado_anterior, resultado_anterior,
                calificacion_anterior, motivo, requiere_reversion_titulacion, usuario
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            expediente_id,
            evaluation_id,
            evaluation_row[1],
            evaluation_row[2],
            evaluation_row[3],
            payload.motivo,
            titulation_completed,
            login,
        )
        cursor.execute(
            """
            UPDATE ops.evaluacion_practica
            SET estado = N'PENDIENTE_REVISION', calificacion = NULL, resultado = N'PENDIENTE',
                origen_calificacion = NULL, detalle_calculo = NULL,
                observacion_revision = ?, observacion_calificacion = NULL,
                revisado_por = NULL, fecha_revision = NULL,
                calificado_por = NULL, fecha_calificacion = NULL,
                usuario_modifica = ?, fecha_modifica = SYSDATETIME()
            WHERE evaluacion_id = ?
            """,
            payload.motivo,
            login,
            evaluation_id,
        )
        record_evaluation_history(
            cursor,
            evaluation_id=evaluation_id,
            action="REAPERTURA",
            user=login,
            observation=payload.motivo,
        )
        cursor.execute(
            """
            UPDATE ops.cierre_proceso
            SET fecha_cierre = NULL, cerrado_por = NULL, certificado_emitido = 0,
                observacion = ?, usuario_modifica = ?, fecha_modifica = SYSDATETIME()
            WHERE expediente_id = ?
            """,
            payload.motivo,
            login,
            expediente_id,
        )
        cursor.execute(
            """
            UPDATE ops.plan_proceso
            SET estado = N'EN_EJECUCION', usuario_modifica = ?, fecha_modifica = SYSDATETIME()
            WHERE expediente_id = ?
            """,
            login,
            expediente_id,
        )
        new_state_id = _catalog_state_id(
            cursor,
            "cat.estado_expediente",
            "estado_expediente_id",
            "EN_PROCESO",
        )
        previous_state_id = expediente.get("estado_expediente_id")
        cursor.execute(
            """
            UPDATE pp.expediente_practica
            SET estado_expediente_id = ?, observacion = ?, usuario_modifica = ?, fecha_modifica = SYSDATETIME()
            WHERE expediente_id = ?
            """,
            new_state_id,
            payload.motivo,
            login,
            expediente_id,
        )
        if int(previous_state_id or 0) != new_state_id:
            cursor.execute(
                """
                INSERT INTO pp.historial_estado_expediente (
                    expediente_id, estado_anterior_id, estado_nuevo_id, motivo, usuario_registro
                ) VALUES (?, ?, ?, ?, ?)
                """,
                expediente_id,
                previous_state_id,
                new_state_id,
                f"Reapertura administrativa: {payload.motivo}",
                login,
            )
        update_compliance_enrollment_status(
            cursor,
            expediente_id=expediente_id,
            state="EN_PROCESO",
            user=login,
        )
        if reconciliation:
            cursor.execute(
                """
                UPDATE ops.conciliacion_titulacion
                SET estado = N'ERROR', proximo_intento = NULL,
                    ultimo_error = N'Expediente reabierto; requiere revisión o reversión administrativa en Titulación.',
                    usuario_solicita = ?, fecha_ultimo_intento = SYSDATETIME(), fecha_completado = NULL
                WHERE conciliacion_id = ?
                """,
                login,
                int(reconciliation[0]),
            )
        write_operations_audit(
            cursor,
            entity="EXPEDIENTE_PRACTICA",
            entity_id=expediente_id,
            action="REAPERTURA",
            user=login,
            detail=payload.motivo,
        )
        conn.commit()
    return {
        "ok": True,
        "message": "Expediente reabierto. La calificación anterior permanece en el historial de la base complementaria.",
        "requiere_revision_titulacion": titulation_completed,
    }


@router.get("/dashboard")
def operations_dashboard(
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
    tipo_proceso: str = Query(default="PPF", pattern="^(PPF|VIN)$"),
) -> dict[str, Any]:
    process = _process_code(tipo_proceso)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        items = _accessible_expedientes(cursor, current_user, process)
        conn.commit()
    total = len(items)
    summary = {
        "total": total,
        "verdes": sum(1 for item in items if item.get("Semaforo") == "VERDE"),
        "amarillos": sum(1 for item in items if item.get("Semaforo") == "AMARILLO"),
        "rojos": sum(1 for item in items if item.get("Semaforo") == "ROJO"),
        "con_plan": sum(1 for item in items if item.get("PlanId")),
        "cerrados": sum(1 for item in items if item.get("FechaCierre")),
        "en_revision": sum(1 for item in items if item.get("EstadoEvaluacion") == "EN_REVISION"),
        "pendientes_calificacion": sum(1 for item in items if item.get("EstadoEvaluacion") == "PENDIENTE_CALIFICACION"),
        "aprobados": sum(1 for item in items if item.get("ResultadoEvaluacion") == "APROBADO"),
        "reprobados": sum(1 for item in items if item.get("ResultadoEvaluacion") == "REPROBADO"),
        "horas_registradas": round(sum(float(item.get("HorasRegistradas") or 0) for item in items), 2),
        "horas_validadas": round(sum(float(item.get("HorasValidadas") or 0) for item in items), 2),
    }
    return {"tipo_proceso": process, "summary": summary, "items": items}


@router.get("/notificaciones")
def operations_notifications(
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
    tipo_proceso: str = Query(default="PPF", pattern="^(PPF|VIN)$"),
    refresh: bool = True,
) -> dict[str, Any]:
    process = _process_code(tipo_proceso)
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        generated = _refresh_notifications(cursor, current_user, process) if refresh else 0
        if current_user.rol in _ADMIN_ROLES:
            cursor.execute(
                """
                SELECT TOP (200) * FROM ops.notificacion_proceso
                WHERE tipo_proceso_codigo = ? AND activa = 1
                ORDER BY leida, fecha_registro DESC
                """,
                process,
            )
        else:
            cursor.execute(
                """
                SELECT TOP (200) * FROM ops.notificacion_proceso
                WHERE tipo_proceso_codigo = ? AND activa = 1 AND destinatario_login = ?
                ORDER BY leida, fecha_registro DESC
                """,
                process,
                login,
            )
        items = _fetch_all(cursor)
        conn.commit()
    return {"generated": generated, "unread": sum(1 for item in items if not item.get("leida")), "items": items}


@router.put("/notificaciones/{notification_id}/leer")
def read_notification(
    notification_id: int,
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
) -> dict[str, Any]:
    login = _current_login(current_user)
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        if current_user.rol in _ADMIN_ROLES:
            cursor.execute("UPDATE ops.notificacion_proceso SET leida = 1, fecha_lectura = SYSDATETIME() WHERE notificacion_id = ?", notification_id)
        else:
            cursor.execute(
                "UPDATE ops.notificacion_proceso SET leida = 1, fecha_lectura = SYSDATETIME() WHERE notificacion_id = ? AND destinatario_login = ?",
                notification_id,
                login,
            )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="No existe la notificación solicitada.")
        conn.commit()
    return {"ok": True}


def _report_rows(current_user: SessionUser, process: str) -> list[dict[str, Any]]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        items = _accessible_expedientes(cursor, current_user, process, 2000)
        conn.commit()
    return items


@router.get("/reportes/seguimiento.xlsx")
def operations_excel(
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
    tipo_proceso: str = Query(default="PPF", pattern="^(PPF|VIN)$"),
) -> StreamingResponse:
    process = _process_code(tipo_proceso)
    rows = _report_rows(current_user, process)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Seguimiento {process}"
    headers = ["Expediente", "Estudiante", "Cédula", "Carrera", "Período", "Estado", "Plan", "Etapa evaluación", "Calificación", "Resultado", "Horas registradas", "Horas validadas", "Documentos", "Semáforo"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="9E1B16")
        cell.alignment = Alignment(horizontal="center")
    for item in rows:
        sheet.append([
            item.get("CodigoExpediente"), item.get("Estudiante"), item.get("Cedula"),
            item.get("Carrera"), item.get("Periodo") or item.get("CodigoPeriodo"), item.get("Estado"),
            item.get("EstadoPlan") or "Pendiente", item.get("EstadoEvaluacion"), item.get("CalificacionFinal"),
            item.get("ResultadoEvaluacion"), item.get("HorasRegistradas"), item.get("HorasValidadas"),
            f"{item.get('DocumentosCargados', 0)}/{item.get('DocumentosRequeridos', 0)}", item.get("Semaforo"),
        ])
    widths = [20, 38, 16, 30, 30, 18, 18, 24, 14, 16, 18, 18, 16, 14]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"seguimiento_{process.lower()}_{date.today().isoformat()}.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/reportes/seguimiento.pdf")
def operations_pdf(
    current_user: Annotated[SessionUser, Depends(_ALL_ACCESS)],
    tipo_proceso: str = Query(default="PPF", pattern="^(PPF|VIN)$"),
) -> StreamingResponse:
    process = _process_code(tipo_proceso)
    rows = _report_rows(current_user, process)
    output = BytesIO()
    document = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    story: list[Any] = [Paragraph(f"Seguimiento institucional {process}", styles["Title"]), Spacer(1, 0.3 * cm)]
    data: list[list[Any]] = [["Expediente", "Estudiante", "Carrera", "Período", "Evaluación", "Nota", "Resultado", "Horas", "Docs.", "Semáforo"]]
    for item in rows:
        data.append([
            item.get("CodigoExpediente") or item.get("ExpedienteId"),
            Paragraph(_clean(item.get("Estudiante")), styles["BodyText"]),
            Paragraph(_clean(item.get("Carrera")), styles["BodyText"]),
            Paragraph(_clean(item.get("Periodo") or item.get("CodigoPeriodo")), styles["BodyText"]),
            item.get("EstadoEvaluacion"),
            "-" if item.get("CalificacionFinal") is None else f"{float(item.get('CalificacionFinal')):.2f}",
            item.get("ResultadoEvaluacion"),
            f"{item.get('HorasValidadas', 0):.2f}/{item.get('HorasRequeridas', 0):.2f}",
            f"{item.get('DocumentosCargados', 0)}/{item.get('DocumentosRequeridos', 0)}",
            item.get("Semaforo"),
        ])
    table = Table(data, repeatRows=1, colWidths=[2.5 * cm, 4.5 * cm, 3.5 * cm, 3.5 * cm, 2.8 * cm, 1.5 * cm, 2.2 * cm, 2.3 * cm, 1.8 * cm, 2 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9E1B16")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#9AA4B2")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F7F8")]),
    ]))
    story.append(table)
    document.build(story)
    output.seek(0)
    filename = f"seguimiento_{process.lower()}_{date.today().isoformat()}.pdf"
    return StreamingResponse(output, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/conciliaciones")
def reconciliation_list(
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
    estado: str | None = Query(default=None, max_length=30),
) -> dict[str, Any]:
    del current_user
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        if estado:
            cursor.execute("SELECT TOP (500) * FROM ops.conciliacion_titulacion WHERE estado = ? ORDER BY fecha_solicitud DESC", estado.upper())
        else:
            cursor.execute("SELECT TOP (500) * FROM ops.conciliacion_titulacion ORDER BY fecha_solicitud DESC")
        items = _fetch_all(cursor)
        conn.commit()
    return {"total": len(items), "items": items}


@router.post("/conciliaciones/{reconciliation_id}/reintentar")
def retry_reconciliation(
    reconciliation_id: int,
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
) -> dict[str, Any]:
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute("SELECT expediente_id, tipo_proceso_codigo FROM ops.conciliacion_titulacion WHERE conciliacion_id = ?", reconciliation_id)
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="No existe la conciliación solicitada.")
        expediente_id = int(row[0])
        process = _process_code(row[1])
        conn.commit()
    try:
        result = _sync_titulacion_completion(expediente_id, process, _current_login(current_user))
    except (pyodbc.Error, RuntimeError) as exc:
        result = {"sincronizado": False, "pendiente": True, "motivo": str(exc)[:1500]}
    save_titulation_reconciliation(expediente_id, process, _current_login(current_user), result)
    return {"ok": bool(result.get("sincronizado")), "result": result}


@router.get("/auditoria")
def operations_audit(
    current_user: Annotated[SessionUser, Depends(_ADMIN_ACCESS)],
    limit: int = Query(default=200, ge=1, le=1000),
) -> dict[str, Any]:
    del current_user
    with get_practices_connection() as conn:
        cursor = conn.cursor()
        ensure_operations_schema(cursor)
        cursor.execute("SELECT TOP (?) * FROM ops.auditoria_operativa ORDER BY fecha DESC, auditoria_id DESC", limit)
        items = _fetch_all(cursor)
        conn.commit()
    return {"total": len(items), "items": items, "generated_at": datetime.now().isoformat()}
