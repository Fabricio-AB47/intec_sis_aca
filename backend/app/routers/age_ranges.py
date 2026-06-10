from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.security import SessionUser, require_roles
from app.services.db import get_connection

router = APIRouter(prefix="/api/students/rango-edades", tags=["rango-edades"])

AllowedUser = Depends(require_roles("ADMINISTRADOR", "ACADEMICO", "RECTOR", "BIENESTAR", "VICERRECTOR"))

AGE_RANGE_ORDER = [
    ("Menor de 18", 0),
    ("18 a 29", 1),
    ("30 a 40", 2),
    ("41 a 50", 3),
    ("51 a 60", 4),
    ("61 o mas", 5),
    ("Sin fecha", 99),
]


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _clean_optional(value: Any) -> str | None:
    text = _clean(value)
    return text or None


def _limit(value: int, *, max_value: int = 10000) -> int:
    return max(1, min(value, max_value))


def _serializable(value: Any) -> Any:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bytes):
        return None
    return value


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _rows_from_cursor(cursor: Any) -> list[dict[str, Any]]:
    columns = [column[0] for column in cursor.description or []]
    return [{column: _serializable(value) for column, value in zip(columns, row)} for row in cursor.fetchall()]


def _age_ranges_sql(
    limit: int,
    *,
    periodo: str | None,
    carrera: str | None,
    estado: str | None,
    tipo_beca: str | None,
    buscar: str | None,
    rango_edad: str | None,
    calculation_date: date,
) -> tuple[str, list[Any]]:
    mat_filters: list[str] = []
    mat_params: list[Any] = []
    where_clauses = ["1 = 1"]
    where_params: list[Any] = []

    if periodo:
        mat_filters.append("AND CAST(ce.codigo_periodo AS varchar(30)) = ?")
        mat_params.append(periodo)
        where_clauses.append("mat.periodo_codigo IS NOT NULL")

    if carrera:
        mat_filters.append("AND CAST(ce.cod_anio_Basica AS varchar(30)) = ?")
        mat_params.append(carrera)
        where_clauses.append("mat.carrera_codigo IS NOT NULL")

    if estado:
        where_clauses.append("LTRIM(RTRIM(TRY_CONVERT(varchar(20), de.Estado))) = ?")
        where_params.append(estado)

    if tipo_beca:
        where_clauses.append("COALESCE(NULLIF(LTRIM(RTRIM(beca.tipo_beca)), ''), 'Sin beca') = ?")
        where_params.append(tipo_beca)

    if buscar:
        like = f"%{buscar}%"
        where_clauses.append(
            """
            (
                de.Apellidos_nombre LIKE ?
                OR de.Cedula_Est LIKE ?
                OR CAST(de.codigo_estud AS varchar(30)) LIKE ?
                OR COALESCE(NULLIF(LTRIM(RTRIM(beca.tipo_beca)), ''), 'Sin beca') LIKE ?
                OR LTRIM(RTRIM(mat.carrera)) LIKE ?
                OR TRY_CONVERT(nvarchar(255), de.correo) LIKE ?
                OR TRY_CONVERT(nvarchar(255), de.correointec) LIKE ?
                OR TRY_CONVERT(nvarchar(100), de.telefono) LIKE ?
                OR TRY_CONVERT(nvarchar(100), de.movil) LIKE ?
            )
            """
        )
        where_params.extend([like, like, like, like, like, like, like, like, like])

    mat_filter_sql = "\n".join(mat_filters)
    where_sql = " AND ".join(where_clauses)

    sql = f"""
        WITH base AS (
            SELECT
                CAST(de.codigo_estud AS varchar(30)) AS estudiante_codigo,
                LTRIM(RTRIM(TRY_CONVERT(varchar(50), de.Cedula_Est))) AS cedula,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), de.Apellidos_nombre))) AS estudiante,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), de.correo))) AS correo_personal,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), de.correointec))) AS correo_intec,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), de.telefono))) AS telefono,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), de.movil))) AS celular,
                LTRIM(RTRIM(TRY_CONVERT(varchar(20), de.Estado))) AS estado_codigo,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(120), es.ESTADO))) AS estado,
                fecha.fecha_nacimiento,
                hoy.fecha_actual AS fecha_calculo,
                edad.edad,
                CASE
                    WHEN edad.edad IS NULL THEN 'Sin fecha'
                    WHEN edad.edad < 18 THEN 'Menor de 18'
                    WHEN edad.edad BETWEEN 18 AND 29 THEN '18 a 29'
                    WHEN edad.edad BETWEEN 30 AND 40 THEN '30 a 40'
                    WHEN edad.edad BETWEEN 41 AND 50 THEN '41 a 50'
                    WHEN edad.edad BETWEEN 51 AND 60 THEN '51 a 60'
                    ELSE '61 o mas'
                END AS rango_edad,
                CASE
                    WHEN edad.edad IS NULL THEN 99
                    WHEN edad.edad < 18 THEN 0
                    WHEN edad.edad BETWEEN 18 AND 29 THEN 1
                    WHEN edad.edad BETWEEN 30 AND 40 THEN 2
                    WHEN edad.edad BETWEEN 41 AND 50 THEN 3
                    WHEN edad.edad BETWEEN 51 AND 60 THEN 4
                    ELSE 5
                END AS rango_orden,
                COALESCE(NULLIF(LTRIM(RTRIM(beca.tipo_beca)), ''), 'Sin beca') AS tipo_beca,
                CASE
                    WHEN COALESCE(NULLIF(LTRIM(RTRIM(beca.tipo_beca)), ''), '') = 'Beca Futuro Femenino' THEN CAST(50 AS decimal(10, 2))
                    ELSE COALESCE(TRY_CONVERT(decimal(10, 2), beca.porcentaje_beca), CAST(0 AS decimal(10, 2)))
                END AS porcentaje_beca,
                mat.periodo_codigo,
                mat.periodo,
                mat.carrera_codigo,
                mat.carrera
            FROM dbo.DATOS_ESTUD de
            CROSS APPLY (SELECT CAST(? AS date) AS fecha_actual) hoy
            LEFT JOIN dbo.ESTADO es ON de.Estado = es.IDESTADO
            OUTER APPLY (
                SELECT TRY_CONVERT(date, de.Fecha_Nac) AS fecha_nacimiento
            ) fecha
            OUTER APPLY (
                SELECT
                    CASE
                        WHEN fecha.fecha_nacimiento IS NULL
                          OR fecha.fecha_nacimiento < CONVERT(date, '19000101')
                          OR fecha.fecha_nacimiento > hoy.fecha_actual
                        THEN NULL
                        ELSE DATEDIFF(YEAR, fecha.fecha_nacimiento, hoy.fecha_actual)
                            - CASE
                                WHEN DATEADD(
                                    YEAR,
                                    DATEDIFF(YEAR, fecha.fecha_nacimiento, hoy.fecha_actual),
                                    fecha.fecha_nacimiento
                                  ) > hoy.fecha_actual
                                THEN 1
                                ELSE 0
                              END
                    END AS edad
            ) edad
            OUTER APPLY (
                SELECT TOP (1)
                    B.tipo_beca,
                    B.porcentaje_beca
                FROM dbo.Becas B
                WHERE TRY_CONVERT(varchar(50), B.codestud) = TRY_CONVERT(varchar(50), de.codigo_estud)
                ORDER BY
                    CASE WHEN NULLIF(LTRIM(RTRIM(B.tipo_beca)), '') IS NULL THEN 1 ELSE 0 END,
                    TRY_CONVERT(decimal(10, 2), B.porcentaje_beca) DESC
            ) beca
            OUTER APPLY (
                SELECT TOP (1)
                    CAST(ce.codigo_periodo AS varchar(30)) AS periodo_codigo,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), p.Detalle_Periodo))) AS periodo,
                    CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
                    LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), c.Nombre_Basica))) AS carrera,
                    p.anio,
                    p.Orden
                FROM dbo.CARRERAXESTUD ce
                LEFT JOIN dbo.PERIODO p ON ce.codigo_periodo = p.cod_periodo
                LEFT JOIN dbo.CARRERAS c ON ce.cod_anio_Basica = c.Cod_AnioBasica
                WHERE TRY_CONVERT(varchar(50), ce.codigo_estud) = TRY_CONVERT(varchar(50), de.codigo_estud)
                  {mat_filter_sql}
                ORDER BY
                    COALESCE(TRY_CONVERT(int, p.anio), 0) DESC,
                    COALESCE(TRY_CONVERT(int, p.Orden), 0) DESC,
                    TRY_CONVERT(int, ce.codigo_periodo) DESC
            ) mat
            WHERE {where_sql}
        )
        SELECT TOP ({limit})
            estudiante_codigo,
            cedula,
            estudiante,
            correo_personal,
            correo_intec,
            telefono,
            celular,
            estado_codigo,
            estado,
            fecha_nacimiento,
            fecha_calculo,
            edad,
            rango_edad,
            tipo_beca,
            porcentaje_beca,
            periodo_codigo,
            periodo,
            carrera_codigo,
            carrera
        FROM base
        WHERE (? IS NULL OR rango_edad = ?)
        ORDER BY rango_orden, estudiante
    """
    return sql, [calculation_date, *mat_params, *where_params, rango_edad, rango_edad]


def _summarize(rows: list[dict[str, Any]]) -> dict[str, Any]:
    range_summary = {
        label: {
            "rango_edad": label,
            "orden": order,
            "total": 0,
            "con_beca": 0,
            "sin_beca": 0,
            "porcentaje_beca_total": 0.0,
            "promedio_beca": 0.0,
        }
        for label, order in AGE_RANGE_ORDER
    }
    total_with_age = 0
    total_without_date = 0
    total_with_scholarship = 0

    for row in rows:
        range_name = _clean(row.get("rango_edad")) or "Sin fecha"
        bucket = range_summary.setdefault(
            range_name,
            {
                "rango_edad": range_name,
                "orden": 100,
                "total": 0,
                "con_beca": 0,
                "sin_beca": 0,
                "porcentaje_beca_total": 0.0,
                "promedio_beca": 0.0,
            },
        )
        scholarship_name = _clean(row.get("tipo_beca")) or "Sin beca"
        has_scholarship = scholarship_name.lower() != "sin beca"
        bucket["total"] += 1
        if has_scholarship:
            bucket["con_beca"] += 1
            bucket["porcentaje_beca_total"] += _number(row.get("porcentaje_beca"))
            total_with_scholarship += 1
        else:
            bucket["sin_beca"] += 1
        if row.get("edad") is None:
            total_without_date += 1
        else:
            total_with_age += 1

    ranges = []
    for item in range_summary.values():
        if item["con_beca"]:
            item["promedio_beca"] = round(item["porcentaje_beca_total"] / item["con_beca"], 2)
        item["porcentaje_beca_total"] = round(item["porcentaje_beca_total"], 2)
        ranges.append(item)
    ranges.sort(key=lambda item: item["orden"])

    return {
        "total": len(rows),
        "edad_calculada": total_with_age,
        "sin_fecha": total_without_date,
        "con_beca": total_with_scholarship,
        "sin_beca": len(rows) - total_with_scholarship,
        "rangos": ranges,
    }


def _execute_age_ranges(
    *,
    periodo: str | None,
    carrera: str | None,
    estado: str | None,
    tipo_beca: str | None,
    buscar: str | None,
    rango_edad: str | None,
    limit: int,
) -> dict[str, Any]:
    calculation_date = date.today()
    safe_limit = _limit(limit)
    filters = {
        "periodo": _clean_optional(periodo),
        "carrera": _clean_optional(carrera),
        "estado": _clean_optional(estado),
        "tipo_beca": _clean_optional(tipo_beca),
        "buscar": _clean_optional(buscar),
        "rango_edad": _clean_optional(rango_edad),
        "limit": safe_limit,
    }
    sql, sql_params = _age_ranges_sql(
        safe_limit,
        periodo=filters["periodo"],
        carrera=filters["carrera"],
        estado=filters["estado"],
        tipo_beca=filters["tipo_beca"],
        buscar=filters["buscar"],
        rango_edad=filters["rango_edad"],
        calculation_date=calculation_date,
    )

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(sql, sql_params)
            rows = _rows_from_cursor(cursor)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo generar el reporte de rango de edades",
        ) from exc

    summary = _summarize(rows)
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "fecha_calculo": calculation_date.isoformat(),
        "columns": [
            "estudiante_codigo",
            "cedula",
            "estudiante",
            "correo_personal",
            "correo_intec",
            "telefono",
            "celular",
            "estado",
            "fecha_nacimiento",
            "edad",
            "rango_edad",
            "tipo_beca",
            "porcentaje_beca",
            "periodo",
            "carrera",
        ],
        "rows": rows,
        "summary": summary,
        "ranges": summary["rangos"],
        "criteria": filters,
    }


@router.get("/catalog")
def catalog(_: SessionUser = AllowedUser) -> dict[str, Any]:
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT DISTINCT LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca))) AS tipo_beca
                FROM dbo.Becas
                WHERE tipo_beca IS NOT NULL
                  AND LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca))) <> ''
                ORDER BY LTRIM(RTRIM(TRY_CONVERT(nvarchar(250), tipo_beca)))
                """
            )
            becas = [_clean(row.tipo_beca) for row in cursor.fetchall() if _clean(row.tipo_beca)]
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="No se pudo cargar el catalogo de edades") from exc

    return {
        "becas": becas,
        "rangos": [{"value": label, "label": label} for label, _order in AGE_RANGE_ORDER],
        "estados": [
            {"value": "A", "label": "Activo"},
            {"value": "G", "label": "Graduado"},
            {"value": "P", "label": "Inactivo"},
            {"value": "R", "label": "Retirado"},
        ],
    }


@router.get("")
def list_age_ranges(
    periodo: str | None = Query(default=None),
    carrera: str | None = Query(default=None),
    estado: str | None = Query(default="A"),
    tipo_beca: str | None = Query(default=None),
    buscar: str | None = Query(default=None),
    rango_edad: str | None = Query(default=None),
    limit: int = Query(default=1000, ge=1, le=10000),
    _: SessionUser = AllowedUser,
) -> dict[str, Any]:
    return _execute_age_ranges(
        periodo=periodo,
        carrera=carrera,
        estado=estado,
        tipo_beca=tipo_beca,
        buscar=buscar,
        rango_edad=rango_edad,
        limit=limit,
    )


def _write_key_values(sheet: Any, rows: list[tuple[str, Any]]) -> None:
    for label, value in rows:
        sheet.append([label, value])
    for row in sheet.iter_rows():
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill("solid", fgColor="E8F1F8")


def _autosize(sheet: Any) -> None:
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 52)


def _workbook(payload: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    ws_rows = workbook.active
    ws_rows.title = "Estudiantes"
    columns = payload["columns"]
    ws_rows.append(columns)
    for cell in ws_rows[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5C83")
        cell.alignment = Alignment(horizontal="center")
    for row in payload["rows"]:
        ws_rows.append([row.get(column) for column in columns])
    _autosize(ws_rows)

    ws_summary = workbook.create_sheet("Resumen rangos")
    ws_summary.append(["Rango de edad", "Total", "Con beca", "Sin beca", "Promedio beca"])
    for cell in ws_summary[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5C83")
        cell.alignment = Alignment(horizontal="center")
    for item in payload["ranges"]:
        ws_summary.append([
            item["rango_edad"],
            item["total"],
            item["con_beca"],
            item["sin_beca"],
            item["promedio_beca"],
        ])
    _autosize(ws_summary)

    ws_meta = workbook.create_sheet("Filtros")
    _write_key_values(
        ws_meta,
        [
            ("Generado", payload["generated_at"]),
            ("Fecha calculo edad", payload["fecha_calculo"]),
            ("Total filas", payload["summary"]["total"]),
            ("Edad calculada", payload["summary"]["edad_calculada"]),
            ("Sin fecha valida", payload["summary"]["sin_fecha"]),
            *[(key, value or "") for key, value in payload["criteria"].items()],
        ],
    )
    _autosize(ws_meta)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/export")
def export_age_ranges(
    periodo: str | None = Query(default=None),
    carrera: str | None = Query(default=None),
    estado: str | None = Query(default="A"),
    tipo_beca: str | None = Query(default=None),
    buscar: str | None = Query(default=None),
    rango_edad: str | None = Query(default=None),
    limit: int = Query(default=10000, ge=1, le=10000),
    _: SessionUser = AllowedUser,
) -> StreamingResponse:
    payload = _execute_age_ranges(
        periodo=periodo,
        carrera=carrera,
        estado=estado,
        tipo_beca=tipo_beca,
        buscar=buscar,
        rango_edad=rango_edad,
        limit=limit,
    )
    filename = f"rango-edades-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        _workbook(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
