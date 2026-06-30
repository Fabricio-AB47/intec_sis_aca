from datetime import datetime
from io import BytesIO
import re
from typing import Annotated, Any
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill
import pandas as pd
from pydantic import BaseModel, Field

from app.core.security import SessionUser, require_roles
from app.routers.students import _MATRICULA_ACTUAL_CTE
from app.services.db import get_connection

router = APIRouter(prefix="/api/students/senescyt", tags=["senescyt"])

_SENESCYT_ACCESS = require_roles("ADMINISTRADOR", "ACADEMICO", "RECTOR")

_REPORT_COLUMNS = [
    "tipoDocumentoId",
    "numeroIdentificacion",
    "primerApellido",
    "segundoApellido",
    "primerNombre",
    "segundoNombre",
    "sexoId",
    "generoId",
    "estadocivilId",
    "etniaId",
    "pueblonacionalidadId",
    "tipoSangre",
    "discapacidad",
    "porcentajeDiscapacidad",
    "numCarnetConadis",
    "tipoDiscapacidad",
    "fechaNacimiento",
    "paisNacionalidadId",
    "provinciaNacimientoId",
    "cantonNacimientoId",
    "paisResidenciaId",
    "provinciaResidenciaId",
    "cantonResidenciaId",
    "tipoColegioId",
    "modalidadCarrera",
    "jornadaCarrera",
    "fechaInicioCarrera",
    "fechaMatricula",
    "tipoMatriculaId",
    "nivelAcademicoQueCursa",
    "duracionPeriodoAcademico",
    "haRepetidoAlMenosUnaMateria",
    "paraleloId",
    "haPerdidoLaGratuidad",
    "recibePensionDiferenciada",
    "estudianteocupacionId",
    "ingresosestudianteId",
    "bonodesarrolloId",
    "haRealizadoPracticasPreprofesionales",
    "nroHorasPracticasPreprofesionalesPorPeriodo",
    "entornoInstitucionalPracticasProfesionales",
    "sectorEconomicoPracticaProfesional",
    "tipoBecaId",
    "primeraRazonBecaId",
    "segundaRazonBecaId",
    "terceraRazonBecaId",
    "cuartaRazonBecaId",
    "quintaRazonBecaId",
    "sextaRazonBecaId",
    "montoBeca",
    "porcientoBecaCoberturaArancel",
    "porcientoBecaCoberturaManuntencion",
    "financiamientoBeca",
    "montoAyudaEconomica",
    "montoCreditoEducativo",
    "participaEnProyectoVinculacionSociedad",
    "tipoAlcanceProyectoVinculacionId",
    "correoElectronico",
    "numeroCelular",
    "nivelFormacionPadre",
    "nivelFormacionMadre",
    "ingresoTotalHogar",
    "cantidadMiembrosHogar",
]

_NUMERIC_COLUMNS = [
    "tipoDocumentoId",
    "sexoId",
    "generoId",
    "estadocivilId",
    "etniaId",
    "pueblonacionalidadId",
    "tipoSangre",
    "tipoDiscapacidad",
    "discapacidad",
    "porcentajeDiscapacidad",
    "tipoColegioId",
    "modalidadCarrera",
    "jornadaCarrera",
    "tipoMatriculaId",
    "nivelAcademicoQueCursa",
    "duracionPeriodoAcademico",
    "haRepetidoAlMenosUnaMateria",
    "paraleloId",
    "haPerdidoLaGratuidad",
    "recibePensionDiferenciada",
    "estudianteocupacionId",
    "ingresosestudianteId",
    "bonodesarrolloId",
    "haRealizadoPracticasPreprofesionales",
    "entornoInstitucionalPracticasProfesionales",
    "sectorEconomicoPracticaProfesional",
    "tipoBecaId",
    "primeraRazonBecaId",
    "segundaRazonBecaId",
    "terceraRazonBecaId",
    "cuartaRazonBecaId",
    "quintaRazonBecaId",
    "sextaRazonBecaId",
    "financiamientoBeca",
    "montoAyudaEconomica",
    "montoCreditoEducativo",
    "participaEnProyectoVinculacionSociedad",
    "tipoAlcanceProyectoVinculacionId",
    "nivelFormacionPadre",
    "nivelFormacionMadre",
    "ingresoTotalHogar",
    "cantidadMiembrosHogar",
]

_UPDATE_FIELD_MAP = {
    "tipoDocumentoId": "tipodocumento",
    "numeroIdentificacion": "Cedula_Est",
    "sexoId": "Sexo",
    "generoId": "generoId",
    "estadocivilId": "EstadoCivil",
    "etniaId": "Etnia",
    "pueblonacionalidadId": "Nacionalidad",
    "tipoSangre": "tiposangre",
    "discapacidad": "discapacidad",
    "porcentajeDiscapacidad": "Porce_Capacidad",
    "numCarnetConadis": "No_Carnet",
    "tipoDiscapacidad": "Tipo_Capacidad",
    "fechaNacimiento": "Fecha_Nac",
    "paisNacionalidadId": "paisNacionalidadId",
    "provinciaNacimientoId": "provinciaNacimeintoId",
    "cantonNacimientoId": "cantonNacimeintoId",
    "paisResidenciaId": "paisResidenciaId",
    "provinciaResidenciaId": "codprov",
    "cantonResidenciaId": "Canton",
    "tipoColegioId": "tipoColegioId",
    "modalidadCarrera": "ModalidadEstudio",
    "jornadaCarrera": "Jornada",
    "fechaInicioCarrera": "Fecha_Ingreso",
    "fechaMatricula": "fechaMatricula",
    "tipoMatriculaId": "tipoMatriculaId",
    "nivelAcademicoQueCursa": "nivelAcademicoQueCursa",
    "duracionPeriodoAcademico": "duracionPeriodoAcademico",
    "haRepetidoAlMenosUnaMateria": "haRepetidoAlMenosUnaMateria",
    "paraleloId": "Paralelo",
    "haPerdidoLaGratuidad": "haPerdidoLaGratuidad",
    "recibePensionDiferenciada": "recibePensionDiferenciada",
    "estudianteocupacionId": "Ocupacion",
    "ingresosestudianteId": "ingresoEstudianteId",
    "bonodesarrolloId": "bonoDesarrolloId",
    "haRealizadoPracticasPreprofesionales": "haRealizadoPracticasPreprofesionales",
    "nroHorasPracticasPreprofesionalesPorPeriodo": "nroHorasPracticasPreprofesionales",
    "entornoInstitucionalPracticasProfesionales": "entornoInstitucionalPracticasProfesionales",
    "sectorEconomicoPracticaProfesional": "sectorEconomicoPracticaProfesional",
    "tipoBecaId": "tipoBecaId",
    "primeraRazonBecaId": "primeraRazonBecaId",
    "segundaRazonBecaId": "segundaRazonBecaId",
    "terceraRazonBecaId": "terceraRazonBecaId",
    "cuartaRazonBecaId": "cuartaRazonBecaId",
    "quintaRazonBecaId": "quintaRazonBecaId",
    "sextaRazonBecaId": "sextaRazonBecaId",
    "montoBeca": "montoBeca",
    "porcientoBecaCoberturaArancel": "porcientoBecaCoberturaArancel",
    "porcientoBecaCoberturaManuntencion": "porcientoBecaCoberturaManuntencion",
    "financiamientoBeca": "financiamientoBeca",
    "montoAyudaEconomica": "montoAyudaEconomica",
    "montoCreditoEducativo": "montoCreditoEducativo",
    "participaEnProyectoVinculacionSociedad": "participaEnProyectoVinculacionSociedad",
    "tipoAlcanceProyectoVinculacionId": "tipoAlcanceProyectoVinculacionId",
    "correoElectronico": "correointec",
    "numeroCelular": "movil",
    "nivelFormacionPadre": "nivelFormacionPadre",
    "nivelFormacionMadre": "nivelFormacionMadre",
    "ingresoTotalHogar": "IngresoHogar",
    "cantidadMiembrosHogar": "Numpersonasvive",
}

_NAME_FIELDS = {"primerApellido", "segundoApellido", "primerNombre", "segundoNombre"}


class SenescytStudentUpdatePayload(BaseModel):
    fields: dict[str, Any] = Field(default_factory=dict)

_DASHBOARD_ACTIVE_COUNT_QUERY = (
    _MATRICULA_ACTUAL_CTE
    + """
SELECT COUNT(*)
FROM base_cruce bc
WHERE bc.estado_codigo = 'A'
  AND bc.cod_anio_Basica IS NOT NULL
  AND NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), bc.nombre_carrera))), '') IS NOT NULL
  AND UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), bc.nombre_carrera)))) NOT LIKE N'SIN CARRERA%';
"""
)

_QUERY = """
WITH UltimaMatricula AS (
    SELECT
        *,
        ROW_NUMBER() OVER (
            PARTITION BY codigo_estud
            ORDER BY fecha_pago DESC, codigo_periodo DESC, numcodigo DESC
        ) AS rn
    FROM dbo.CABECERA_MATRICULA
)
SELECT
    TRY_CONVERT(varchar(50), e.codigo_estud) AS codigoEstud,
    e.tipodocumento AS tipoDocumentoId,
    e.Cedula_Est AS numeroIdentificacion,
    e.Apellidos_nombre,
    e.Sexo AS sexoId,
    e.generoId,
    e.EstadoCivil AS estadocivilId,
    e.Etnia AS etniaId,
    e.Nacionalidad AS pueblonacionalidadId,
    e.tiposangre AS tipoSangre,
    e.discapacidad,
    e.Porce_Capacidad AS porcentajeDiscapacidad,
    e.No_Carnet AS numCarnetConadis,
    e.Tipo_Capacidad AS tipoDiscapacidad,
    e.Fecha_Nac AS fechaNacimiento,
    e.paisNacionalidadId,
    e.provinciaNacimeintoId AS provinciaNacimientoId,
    e.cantonNacimeintoId AS cantonNacimientoId,
    e.paisResidenciaId,
    COALESCE(e.codprov, e.provinciaNacimeintoId) AS provinciaResidenciaId,
    COALESCE(e.Canton, e.cantonNacimeintoId) AS cantonResidenciaId,
    e.tipoColegioId,
    e.ModalidadEstudio AS modalidadCarrera,
    e.Jornada AS jornadaCarrera,
    e.Fecha_Ingreso AS fechaInicioCarrera,
    e.fechaMatricula,
    e.tipoMatriculaId,
    e.nivelAcademicoQueCursa,
    e.duracionPeriodoAcademico,
    e.haRepetidoAlMenosUnaMateria,
    e.Paralelo AS paraleloId,
    e.haPerdidoLaGratuidad,
    e.recibePensionDiferenciada,
    e.Ocupacion AS estudianteocupacionId,
    e.ingresoEstudianteId AS ingresosestudianteId,
    e.bonoDesarrolloId AS bonodesarrolloId,
    e.haRealizadoPracticasPreprofesionales,
    e.nroHorasPracticasPreprofesionales AS nroHorasPracticasPreprofesionalesPorPeriodo,
    e.entornoInstitucionalPracticasProfesionales,
    e.sectorEconomicoPracticaProfesional,
    e.tipoBecaId,
    e.primeraRazonBecaId,
    e.segundaRazonBecaId,
    e.terceraRazonBecaId,
    e.cuartaRazonBecaId,
    e.quintaRazonBecaId,
    e.sextaRazonBecaId,
    e.montoBeca,
    e.porcientoBecaCoberturaArancel,
    e.porcientoBecaCoberturaManuntencion,
    e.financiamientoBeca,
    e.montoAyudaEconomica,
    e.montoCreditoEducativo,
    e.participaEnProyectoVinculacionSociedad,
    e.tipoAlcanceProyectoVinculacionId,
    c.CorreoIntec AS correoElectronico,
    e.movil AS numeroCelular,
    e.nivelFormacionPadre,
    e.nivelFormacionMadre,
    e.IngresoHogar AS ingresoTotalHogar,
    e.Numpersonasvive AS cantidadMiembrosHogar,
    ca.Nombre_Basica AS nombreCarrera
FROM dbo.DATOS_ESTUD e
INNER JOIN dbo.CorreosEstudIntec c
    ON TRY_CONVERT(varchar(50), e.codigo_estud) = TRY_CONVERT(varchar(50), c.codestud)
INNER JOIN UltimaMatricula m
    ON TRY_CONVERT(varchar(50), e.codigo_estud) = TRY_CONVERT(varchar(50), m.codigo_estud)
    AND m.rn = 1
INNER JOIN dbo.CARRERAS ca
    ON TRY_CONVERT(varchar(50), m.cod_anio_Basica) = TRY_CONVERT(varchar(50), ca.Cod_AnioBasica)
WHERE UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), e.Estado)))) IN (N'A', N'ACTIVO', N'ACTIVA');
"""


def _read_dataframe() -> pd.DataFrame:
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(_QUERY)
        columns = [column[0] for column in cursor.description]
        rows = [tuple(row) for row in cursor.fetchall()]
    return pd.DataFrame.from_records(rows, columns=columns)


def _count_scalar(sql: str) -> int:
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(sql)
        return int(cursor.fetchone()[0] or 0)


def _clean_income(value: Any) -> float:
    if pd.isna(value):
        return 0
    match = re.search(r"[-+]?\d+(?:[.,]\d+)?", str(value))
    if not match:
        return 0
    try:
        return float(match.group(0).replace(",", "."))
    except ValueError:
        return 0


def _clean_amount(value: Any) -> Any:
    if pd.isna(value):
        return "NA"
    text = str(value).strip()
    if not text or text.upper() == "NA":
        return "NA"
    try:
        number = float(text.replace(",", "."))
    except ValueError:
        return "NA"
    return number


def _normalize_conadis(value: Any) -> str:
    if pd.isna(value):
        return "NA"
    text = str(value).strip()
    if not text:
        return "NA"
    digits = re.sub(r"\D", "", text)
    if digits and int(digits) == 0:
        return "0"
    return digits if len(digits) == 10 else "NA"


def _format_province(value: Any) -> str | None:
    if pd.isna(value):
        return None
    digits = re.sub(r"\D", "", str(value).strip())
    if not digits:
        return None
    base = str(int(digits))
    if base == "50":
        return "05"
    if len(base) >= 2:
        return base[:2]
    return base.zfill(2)


def _split_names(full_name: Any) -> pd.Series:
    parts = str(full_name or "").split()
    return pd.Series(
        {
            "primerApellido": parts[0] if len(parts) > 0 else None,
            "segundoApellido": parts[1] if len(parts) > 1 else None,
            "primerNombre": parts[2] if len(parts) > 2 else None,
            "segundoNombre": " ".join(parts[3:]) if len(parts) > 3 else "NA",
        }
    )


def _filled(value: Any) -> bool:
    if value is None or pd.isna(value):
        return False
    text = str(value).strip()
    return bool(text) and text.upper() not in {"NA", "N/A", "NONE", "NULL"}


def _safe_filename(value: Any) -> str:
    name = re.sub(r'[<>:"/\\|?*]+', "", str(value or "Sin carrera")).strip()
    name = re.sub(r"\s+", " ", name)
    return name[:120] or "Sin carrera"


def _style_header(worksheet: Any) -> None:
    fill = PatternFill("solid", fgColor="DDEBF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _student_display_name(row: pd.Series) -> str:
    parts = [
        row.get("primerApellido"),
        row.get("segundoApellido"),
        row.get("primerNombre"),
        row.get("segundoNombre"),
    ]
    name = " ".join(str(part).strip() for part in parts if _filled(part))
    return name or "Sin nombre"


def _student_missing_detail(row: pd.Series) -> dict[str, Any] | None:
    missing = [column for column in _REPORT_COLUMNS if not _filled(row.get(column))]
    if not missing:
        return None
    filled = len(_REPORT_COLUMNS) - len(missing)
    return {
        "codigo_estud": str(row.get("codigoEstud") or "").strip(),
        "estudiante": _student_display_name(row),
        "numero_identificacion": str(row.get("numeroIdentificacion") or "").strip(),
        "campos_llenos": filled,
        "campos_pendientes": len(missing),
        "campos_totales": len(_REPORT_COLUMNS),
        "porcentaje_lleno": round((filled / max(len(_REPORT_COLUMNS), 1)) * 100, 2),
        "campos_faltantes": missing,
    }


def _normalize_payload_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text != "" else None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return value


def _fetch_datos_estud_raw(codigo_estud: str) -> dict[str, Any]:
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT *
            FROM dbo.DATOS_ESTUD
            WHERE TRY_CONVERT(varchar(50), codigo_estud) = TRY_CONVERT(varchar(50), ?)
            """,
            codigo_estud,
        )
        row = cursor.fetchone()
        if not row:
            return {"columns": [], "fields": {}}
        columns = [column[0] for column in cursor.description]
        values = {column: _normalize_payload_value(value) for column, value in zip(columns, row)}
        return {"columns": columns, "fields": values}


def _student_summary(row: pd.Series) -> dict[str, Any]:
    missing = [column for column in _REPORT_COLUMNS if not _filled(row.get(column))]
    filled = len(_REPORT_COLUMNS) - len(missing)
    return {
        "codigo_estud": str(row.get("codigoEstud") or "").strip(),
        "estudiante": _student_display_name(row),
        "numero_identificacion": str(row.get("numeroIdentificacion") or "").strip(),
        "nombre_carrera": str(row.get("nombreCarrera") or "Sin carrera").strip(),
        "campos_llenos": filled,
        "campos_pendientes": len(missing),
        "campos_totales": len(_REPORT_COLUMNS),
        "porcentaje_lleno": round((filled / max(len(_REPORT_COLUMNS), 1)) * 100, 2),
        "campos_faltantes": missing,
    }


def _load_normalized_students() -> pd.DataFrame:
    return _normalize_dataframe(_read_dataframe())


def _find_student_row(codigo_estud: str) -> pd.Series:
    dataframe = _load_normalized_students()
    matches = dataframe[
        dataframe["codigoEstud"].astype(str).str.strip() == str(codigo_estud).strip()
    ]
    if matches.empty:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado en datos SENECYT.")
    return matches.iloc[0]


def _student_detail_response(row: pd.Series) -> dict[str, Any]:
    raw = _fetch_datos_estud_raw(str(row.get("codigoEstud") or ""))
    return {
        "student": _student_summary(row),
        "fields": {column: _normalize_payload_value(row.get(column)) for column in _REPORT_COLUMNS},
        "report_columns": _REPORT_COLUMNS,
        "datos_estud_fields": raw["fields"],
        "datos_estud_columns": raw["columns"],
    }


def _normalize_dataframe(raw: pd.DataFrame) -> pd.DataFrame:
    df = raw.copy()
    df["nombreCarrera"] = df["nombreCarrera"].apply(
        lambda value: re.sub(r"\s+", " ", str(value or "Sin carrera")).strip() or "Sin carrera"
    )
    for column in _NUMERIC_COLUMNS:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    df["ingresoTotalHogar"] = df["ingresoTotalHogar"].apply(_clean_income)
    df["ingresoTotalHogar"] = pd.to_numeric(df["ingresoTotalHogar"], errors="coerce").fillna(0)

    df["porcentajeDiscapacidad"] = df["porcentajeDiscapacidad"].fillna(0)
    df["tipoSangre"] = df["tipoSangre"].fillna(0).replace(0, 7)
    df["tipoDiscapacidad"] = df["tipoDiscapacidad"].fillna(7).replace(0, 7)
    df["discapacidad"] = df["discapacidad"].fillna(0).replace(0, 2)
    df["pueblonacionalidadId"] = df["pueblonacionalidadId"].fillna(0).replace(0, 34)
    df["tipoColegioId"] = df["tipoColegioId"].fillna(0).replace(0, 1)
    df["financiamientoBeca"] = df["financiamientoBeca"].fillna(4).replace(0, 4)
    df["numeroCelular"] = df["numeroCelular"].where(df["numeroCelular"].isna(), df["numeroCelular"].astype(str))
    df["tipoDocumentoId"] = df["tipoDocumentoId"].fillna(0).replace(0, 1)
    df["haPerdidoLaGratuidad"] = df["haPerdidoLaGratuidad"].fillna(0).replace(0, 3)
    df["recibePensionDiferenciada"] = df["recibePensionDiferenciada"].fillna(0).replace(0, 2)
    df["haRepetidoAlMenosUnaMateria"] = df["haRepetidoAlMenosUnaMateria"].fillna(0).replace(0, 2)
    df["sectorEconomicoPracticaProfesional"] = df["sectorEconomicoPracticaProfesional"].fillna(0).replace(0, 22)
    df["tipoBecaId"] = df["tipoBecaId"].fillna(0).replace(0, 3)
    df["bonodesarrolloId"] = df["bonodesarrolloId"].fillna(0).replace(0, 2)
    df["paraleloId"] = df["paraleloId"].fillna(0).replace(0, 1)
    df["ingresosestudianteId"] = df["ingresosestudianteId"].fillna(0).replace(0, 4)
    for column in [
        "primeraRazonBecaId",
        "segundaRazonBecaId",
        "terceraRazonBecaId",
        "cuartaRazonBecaId",
        "quintaRazonBecaId",
        "sextaRazonBecaId",
    ]:
        df[column] = df[column].fillna(0).replace(0, 2)
    df["haRealizadoPracticasPreprofesionales"] = df["haRealizadoPracticasPreprofesionales"].fillna(0).replace(0, 2)
    df["entornoInstitucionalPracticasProfesionales"] = df["entornoInstitucionalPracticasProfesionales"].fillna(0).replace(0, 5)
    for column in ["montoAyudaEconomica", "montoCreditoEducativo"]:
        df[column] = df[column].apply(_clean_amount)
    df["montoBeca"] = pd.to_numeric(df["montoBeca"], errors="coerce").fillna(0)
    df["participaEnProyectoVinculacionSociedad"] = df["participaEnProyectoVinculacionSociedad"].fillna(0).replace(0, 3)
    df["tipoAlcanceProyectoVinculacionId"] = df["tipoAlcanceProyectoVinculacionId"].fillna(0).replace(0, 5)
    df["nroHorasPracticasPreprofesionalesPorPeriodo"] = df[
        "nroHorasPracticasPreprofesionalesPorPeriodo"
    ].apply(lambda value: "NA" if pd.isna(value) or str(value).strip() == "" else value)
    df["porcientoBecaCoberturaArancel"] = pd.to_numeric(
        df["porcientoBecaCoberturaArancel"], errors="coerce"
    ).apply(lambda value: "NA" if pd.isna(value) else int(value))
    df["numCarnetConadis"] = df["numCarnetConadis"].apply(_normalize_conadis)
    df["provinciaResidenciaId"] = df["provinciaResidenciaId"].apply(_format_province)

    names = df["Apellidos_nombre"].apply(_split_names)
    df = df.drop(columns=["Apellidos_nombre"])
    df = pd.concat(
        [
            df[["tipoDocumentoId", "numeroIdentificacion"]],
            names,
            df.drop(columns=["tipoDocumentoId", "numeroIdentificacion"]),
        ],
        axis=1,
    )
    return df[_REPORT_COLUMNS + ["nombreCarrera", "codigoEstud"]]


def _build_report() -> dict[str, Any]:
    raw = _read_dataframe()
    final = _normalize_dataframe(raw)
    total_report = len(final)
    total_active_dashboard = _count_scalar(_DASHBOARD_ACTIVE_COUNT_QUERY)
    total_active_datos_estud = total_report

    field_totals = {
        column: int(final[column].map(_filled).sum())
        for column in _REPORT_COLUMNS
    }
    total_cells = max(total_report * len(_REPORT_COLUMNS), 1)
    filled_cells = sum(field_totals.values())

    career_rows: list[dict[str, Any]] = []
    student_missing_rows: list[dict[str, Any]] = []
    if total_report:
        for career, group in final.groupby("nombreCarrera", dropna=False):
            filled = int(sum(group[column].map(_filled).sum() for column in _REPORT_COLUMNS))
            cells = max(len(group) * len(_REPORT_COLUMNS), 1)
            students_missing = [
                detail
                for _, row in group.iterrows()
                if (detail := _student_missing_detail(row)) is not None
            ]
            students_missing.sort(key=lambda item: (-item["campos_pendientes"], item["estudiante"]))
            career_name = str(career or "Sin carrera")
            for student in students_missing:
                student_missing_rows.append({"nombre_carrera": career_name, **student})
            career_rows.append(
                {
                    "nombre_carrera": career_name,
                    "total_estudiantes": int(len(group)),
                    "campos_llenos": filled,
                    "campos_totales": cells,
                    "campos_pendientes": int(cells - filled),
                    "estudiantes_con_pendientes": len(students_missing),
                    "porcentaje_lleno": round((filled / cells) * 100, 2),
                    "students_missing": students_missing,
                }
            )
    career_rows.sort(key=lambda item: item["nombre_carrera"])

    missing_fields = sorted(
        (
            {
                "campo": column,
                "llenos": filled,
                "pendientes": total_report - filled,
                "porcentaje_lleno": round((filled / max(total_report, 1)) * 100, 2),
            }
            for column, filled in field_totals.items()
        ),
        key=lambda item: (-item["pendientes"], item["campo"]),
    )

    warnings: list[str] = []
    if total_report != total_active_dashboard:
        warnings.append("El total exportable no coincide con los activos del tablero de matricula.")

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "dataframe": final,
        "summary": {
            "total_reporte": total_report,
            "total_activos_sistema": total_active_dashboard,
            "total_activos_datos_estud": total_active_datos_estud,
            "coincide_activos": total_report == total_active_dashboard,
            "total_carreras": len(career_rows),
            "total_columnas": len(_REPORT_COLUMNS),
            "campos_llenos": filled_cells,
            "campos_totales": total_cells,
            "porcentaje_lleno": round((filled_cells / total_cells) * 100, 2),
        },
        "careers": career_rows,
        "students_missing": student_missing_rows,
        "missing_fields": missing_fields[:15],
        "warnings": warnings,
        "criteria": {
            "fuente": "DATOS_ESTUD directo para los datos del estudiante; CARRERAXESTUD/PENSUM solo definen carrera y estado del reporte",
            "activos": "Mismo criterio del tablero de matricula; excluye estudiantes sin carrera registrada",
            "matricula": "Matricula actual validada contra carrera, pensum y estado del tablero",
            "export": "Un archivo Excel por carrera dentro de un ZIP",
        },
    }


def _dataframe_to_workbook_bytes(dataframe: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Datos")
        worksheet = writer.book["Datos"]
        _style_header(worksheet)
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 42)
            worksheet.column_dimensions[column_letter].width = max(width, 12)
    output.seek(0)
    return output.getvalue()


@router.get("/estudiantes/buscar")
def search_senescyt_students(
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
    q: Annotated[str, Query(max_length=120)] = "",
    limit: Annotated[int, Query(ge=1, le=100)] = 30,
) -> dict[str, Any]:
    del current_user
    try:
        dataframe = _load_normalized_students()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error buscando estudiantes SENECYT: {exc}") from exc

    text = q.strip().casefold()
    if text:
        searchable = (
            dataframe["codigoEstud"].fillna("").astype(str)
            + " "
            + dataframe["numeroIdentificacion"].fillna("").astype(str)
            + " "
            + dataframe["primerApellido"].fillna("").astype(str)
            + " "
            + dataframe["segundoApellido"].fillna("").astype(str)
            + " "
            + dataframe["primerNombre"].fillna("").astype(str)
            + " "
            + dataframe["segundoNombre"].fillna("").astype(str)
        ).str.casefold()
        dataframe = dataframe[searchable.str.contains(re.escape(text), regex=True, na=False)]

    rows = [_student_summary(row) for _, row in dataframe.head(limit).iterrows()]
    return {"rows": rows, "total": int(len(dataframe)), "limit": limit, "query": q}


@router.get("/estudiantes/datos/{codigo_estud}")
def get_senescyt_student_data(
    codigo_estud: str,
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    return _student_detail_response(_find_student_row(codigo_estud))


@router.put("/estudiantes/datos/{codigo_estud}")
def update_senescyt_student_data(
    codigo_estud: str,
    payload: SenescytStudentUpdatePayload,
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    current_row = _find_student_row(codigo_estud)
    updates: dict[str, Any] = {}

    name_changed = bool(_NAME_FIELDS.intersection(payload.fields))
    if name_changed:
        name_values = {
            field: _normalize_payload_value(payload.fields.get(field, current_row.get(field)))
            for field in _NAME_FIELDS
        }
        full_name = " ".join(str(name_values[field]).strip() for field in [
            "primerApellido",
            "segundoApellido",
            "primerNombre",
            "segundoNombre",
        ] if _filled(name_values[field]))
        updates["Apellidos_nombre"] = full_name or None

    for field, value in payload.fields.items():
        column = _UPDATE_FIELD_MAP.get(field)
        if not column:
            continue
        updates[column] = _normalize_payload_value(value)

    if not updates:
        raise HTTPException(status_code=400, detail="No hay campos validos para actualizar.")

    set_clause = ", ".join(f"{column} = ?" for column in updates)
    params = list(updates.values()) + [codigo_estud]
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                f"""
                UPDATE dbo.DATOS_ESTUD
                SET {set_clause}
                WHERE TRY_CONVERT(varchar(50), codigo_estud) = TRY_CONVERT(varchar(50), ?)
                """,
                params,
            )
            affected = cursor.rowcount
            conn.commit()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error actualizando DATOS_ESTUD: {exc}") from exc

    updated_row = _find_student_row(codigo_estud)
    return {
        "ok": True,
        "message": "Datos del estudiante actualizados.",
        "updated_fields": sorted(payload.fields.keys()),
        "affected_rows": int(affected or 0),
        **_student_detail_response(updated_row),
    }


@router.get("/estudiantes")
def senescyt_students_report(
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    try:
        report = _build_report()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error generando reporte SENECYT: {exc}") from exc
    return {
        "generated_at": report["generated_at"],
        "summary": report["summary"],
        "careers": report["careers"],
        "missing_fields": report["missing_fields"],
        "warnings": report["warnings"],
        "criteria": report["criteria"],
    }


@router.get("/estudiantes/export")
def senescyt_students_export(
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
    split_by_career: Annotated[bool, Query(description="Generar un Excel por carrera dentro del ZIP")] = True,
) -> StreamingResponse:
    del current_user
    try:
        report = _build_report()
        dataframe: pd.DataFrame = report["dataframe"]
        output = BytesIO()
        with ZipFile(output, "w", ZIP_DEFLATED) as archive:
            if split_by_career:
                for career, group in dataframe.groupby("nombreCarrera", dropna=False):
                    filename = f"EstudiantesPorCarrera/{_safe_filename(career)}.xlsx"
                    archive.writestr(
                        filename,
                        _dataframe_to_workbook_bytes(group.drop(columns=["nombreCarrera", "codigoEstud"])),
                    )
            else:
                archive.writestr(
                    "senescyt_estudiantes.xlsx",
                    _dataframe_to_workbook_bytes(dataframe.drop(columns=["nombreCarrera", "codigoEstud"])),
                )
        output.seek(0)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error exportando reporte SENECYT: {exc}") from exc

    filename = f"senescyt_estudiantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return StreamingResponse(
        output,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_TEACHER_REPORT_COLUMNS = [
    "tipoDocumentoId",
    "numeroIdentificacion",
    "primerApellido",
    "segundoApellido",
    "primerNombre",
    "segundoNombre",
    "sexoId",
    "generoId",
    "estadocivilId",
    "etniaId",
    "pueblonacionalidadId",
    "direccionDomiciliaria",
    "provinciaSufragio",
    "numeroCelular",
    "correoElectronico",
    "numDomicilio",
    "discapacidad",
    "porcentajeDiscapacidad",
    "numCarnetDiscapacidad",
    "tipoDiscapacidad",
    "tipoEnfermedadCatastrofica",
    "fechaNacimiento",
    "paisNacionalidadId",
    "nivelFormacion",
    "fechaIngresoIES",
    "fechaSalidaIES",
    "relacionLaboralIESId",
    "ingresoConCursoMeritos",
    "escalafonDocenteId",
    "cargoDirectivoId",
    "tiempoDedicacionId",
    "nombreUnidadAcademica",
    "nroasignaturasdocente",
    "nroHorasLaborablesSemanaEnCarreraPrograma",
    "nroHorasClaseSemanaCarreraPrograma",
    "nroHorasInvestigacionSemanaCarreraPrograma",
    "nroHorasAdministrativasSemanaCarreraPrograma",
    "nroHorasOtrasActividadesSemanaCarreraPrograma",
    "nroHorasVinculacionSociedad",
    "salarioMensual",
    "docenciaTecnicoSuperior",
    "docenciaTecnologico",
    "estaEnPeriodoSabatico",
    "fechaInicioPeriodoSabatico",
    "estaCursandoEstudiosId",
    "institucionDondeCursaEstudios",
    "paisEstudiosId",
    "tituloAObtener",
    "poseeBecaId",
    "tipoBecaId",
    "montoBeca",
    "financiamientoBecaId",
    "pubRevistasCienInIndexadasId",
    "numPubRevistasCientifIndexadas",
    "docenciaTecnologicoUniversitario",
    "docenciaEspecializacionTecnologica",
    "docenciaMaestriaTecnologica",
]

_STUDENT_AUDIT_COLUMNS = _REPORT_COLUMNS

_REPORT_TARGETS = {"estudiantes", "docentes"}
_EXPORT_MODES = {"completo", "faltantes"}


def _read_sql_dataframe(sql: str, params: list[Any] | None = None) -> pd.DataFrame:
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(sql, params or [])
        columns = [column[0] for column in cursor.description]
        rows = [tuple(row) for row in cursor.fetchall()]
    return pd.DataFrame.from_records(rows, columns=columns)


def _career_catalog() -> list[dict[str, Any]]:
    dataframe = _read_sql_dataframe(
        """
        SELECT
            TRY_CONVERT(varchar(50), Cod_AnioBasica) AS codigo_carrera,
            LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), Nombre_Basica))) AS nombre_carrera
        FROM dbo.CARRERAS
        WHERE NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), Nombre_Basica))), '') IS NOT NULL
        ORDER BY Nombre_Basica
        """
    )
    return [
        {
            "codigo_carrera": str(row.codigo_carrera or "").strip(),
            "nombre_carrera": str(row.nombre_carrera or "").strip(),
        }
        for row in dataframe.itertuples()
    ]


def _split_report_name(full_name: Any) -> dict[str, Any]:
    values = _split_names(full_name)
    return {key: _normalize_payload_value(values.get(key)) for key in _NAME_FIELDS}


def _prepare_student_audit_dataframe() -> pd.DataFrame:
    raw = _read_dataframe()
    if raw.empty:
        return pd.DataFrame(columns=_STUDENT_AUDIT_COLUMNS + ["codigo", "nombreCompleto", "nombreCarrera"])

    df = _normalize_dataframe(raw).rename(columns={"codigoEstud": "codigo"})
    for column in _STUDENT_AUDIT_COLUMNS:
        if column not in df.columns:
            df[column] = None
    df["nombreCompleto"] = df.apply(_student_display_name, axis=1)
    df["nombreCarrera"] = df["nombreCarrera"].apply(
        lambda value: re.sub(r"\s+", " ", str(value or "Sin carrera")).strip() or "Sin carrera"
    )
    return df[_STUDENT_AUDIT_COLUMNS + ["codigo", "nombreCompleto", "nombreCarrera"]]


def _read_teacher_audit_dataframe() -> pd.DataFrame:
    sql = """
    SELECT DISTINCT
        TRY_CONVERT(varchar(50), d.codigo_doc) AS codigo,
        d.apellidos_nombre AS nombreOriginal,
        d.tipoDocumentoId,
        LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.cedula_doc))) AS numeroIdentificacion,
        d.sexo AS sexoId,
        d.generoId,
        d.estado_civil AS estadocivilId,
        d.etniaId,
        CAST(NULL AS nvarchar(100)) AS pueblonacionalidadId,
        d.numDomicilio AS numDomicilio,
        d.Direccion AS direccionDomiciliaria,
        d.provinciaSufragio,
        COALESCE(
            NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), d.movil))), ''),
            NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), d.telefono))), '')
        ) AS numeroCelular,
        COALESCE(
            NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(320), d.correo))), ''),
            NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(320), d.correop))), '')
        ) AS correoElectronico,
        d.discapacidad,
        d.porcen_discapa AS porcentajeDiscapacidad,
        d.num_carnet_cona AS numCarnetDiscapacidad,
        d.tipo_discapa AS tipoDiscapacidad,
        d.tipoEnfermedadCatastrofica,
        d.fecha_nac AS fechaNacimiento,
        d.paisNacionalidadId,
        d.nivelFormacion,
        d.fechaIngresoIES,
        d.fechaSalidaIES,
        d.relacionLaboralIESId,
        d.ingresoConCursoMeritos,
        d.escalafonDocenteId,
        d.cargoDirectivoId,
        d.tiempoDedicacionId,
        d.nombreUnidadAcademica,
        d.nroasignaturasdocente AS nroasignaturasdocente,
        d.nroHorasLaborablesSemanaEnCarreraPrograma,
        d.nroHorasClaseSemanaCarreraPrograma,
        d.nroHorasInvestigacionSemanaCarreraPrograma,
        d.nroHorasAdministrativasSemanaCarreraPrograma,
        d.nroHorasOtrasActividadesSemanaCarreraPrograma,
        d.nroHorasVinculacionSociedad,
        d.salarioMensual,
        d.docenciaTecnicoSuperior,
        d.docenciaTecnologico,
        d.estaEnPeriodoSabatico,
        d.fechaInicioPeriodoSabatico,
        d.estaCursandoEstudiosId,
        d.institucionDOndeCursaEstudios AS institucionDondeCursaEstudios,
        d.paisEstudiosId,
        d.tituloAObtener,
        d.poseeBecaId,
        d.tipoBecaId,
        d.montoBeca,
        d.financiamientoBecaId,
        d.pubRevistasCienInIndexadasId,
        d.numPubRevistasCientifIndexadas,
        d.docenciaTecnologicoUniversitario,
        d.docenciaEspecializacionTecnologica,
        d.docenciaMaestriaTecnologica,
        COALESCE(
            NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), c.Nombre_Basica))), ''),
            N'Sin carrera'
        ) AS nombreCarrera
    FROM dbo.CARRERAXDOCENTE cd
    INNER JOIN dbo.CARRERAS c
        ON TRY_CONVERT(varchar(50), cd.cod_Anio_Basica) = TRY_CONVERT(varchar(50), c.Cod_AnioBasica)
    INNER JOIN dbo.DATOSDOCENTE d
        ON TRY_CONVERT(varchar(50), cd.codigo_doc) = TRY_CONVERT(varchar(50), d.codigo_doc)
    WHERE EXISTS (
        SELECT 1
        FROM dbo.USUARIOS u
        WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(50), u.cedula))) = LTRIM(RTRIM(TRY_CONVERT(varchar(50), d.cedula_doc)))
          AND UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(50), u.Estado)))) IN (N'A', N'ACTIVO', N'ACTIVA')
    )
    ORDER BY nombreCarrera, nombreOriginal
    """
    raw = _read_sql_dataframe(sql)
    if raw.empty:
        return pd.DataFrame(columns=_TEACHER_REPORT_COLUMNS + ["codigo", "nombreCompleto", "nombreCarrera"])

    names = raw["nombreOriginal"].apply(_split_report_name).apply(pd.Series)
    df = pd.concat([raw.drop(columns=["nombreOriginal"], errors="ignore"), names], axis=1)
    for column in _TEACHER_REPORT_COLUMNS:
        if column not in df.columns:
            df[column] = None
    df["nombreCompleto"] = df.apply(
        lambda row: " ".join(str(row.get(field) or "").strip() for field in [
            "primerApellido",
            "segundoApellido",
            "primerNombre",
            "segundoNombre",
        ] if _filled(row.get(field))) or "Sin nombre",
        axis=1,
    )
    df["nombreCarrera"] = df["nombreCarrera"].apply(
        lambda value: re.sub(r"\s+", " ", str(value or "Sin carrera")).strip() or "Sin carrera"
    )
    return df[_TEACHER_REPORT_COLUMNS + ["codigo", "nombreCompleto", "nombreCarrera"]]


def _load_senescyt_audit_dataframe(target: str) -> tuple[pd.DataFrame, list[str]]:
    if target == "estudiantes":
        return _prepare_student_audit_dataframe(), _STUDENT_AUDIT_COLUMNS
    if target == "docentes":
        return _read_teacher_audit_dataframe(), _TEACHER_REPORT_COLUMNS
    raise HTTPException(status_code=400, detail="Tipo de reporte SENESCYT no valido.")


def _normalize_career_filters(careers: list[str] | None) -> list[str]:
    selected: list[str] = []
    values = [careers] if isinstance(careers, str) else (careers or [])
    for value in values:
        for item in str(value or "").split("|"):
            name = re.sub(r"\s+", " ", item).strip()
            if name and name.casefold() not in {career.casefold() for career in selected}:
                selected.append(name)
    return selected


def _filter_by_career(dataframe: pd.DataFrame, careers: list[str] | None) -> pd.DataFrame:
    selected = _normalize_career_filters(careers)
    if not selected:
        return dataframe
    names = dataframe["nombreCarrera"].fillna("").astype(str).str.casefold()
    mask = pd.Series(False, index=dataframe.index)
    for career in selected:
        needle = career.casefold()
        mask = mask | names.eq(needle) | names.str.contains(re.escape(needle), regex=True, na=False)
    return dataframe[mask].copy()


_NO_APLICA_MARKERS = {"NA", "N/A", "N.A", "N.A.", "NO APLICA", "NO APLICA.", "NO_APLICA"}
_EMPTY_MARKERS = {"", "NONE", "NULL", "NULO", "SIN DATO", "SIN DATOS"}
_UNSELECTED_MARKERS = {
    "0",
    "0.0",
    "00",
    "000",
    "0000",
    "SELECCIONE",
    "-- SELECCIONE --",
    "- SELECCIONE -",
    "SELECCIONE ESTADO",
}

_ZERO_ALLOWED_FIELDS = {
    ("estudiantes", "ingresoTotalHogar"),
}

_ZERO_CONTEXT_FIELDS = {
    ("estudiantes", "porcentajeDiscapacidad"),
    ("estudiantes", "numCarnetConadis"),
    ("estudiantes", "montoBeca"),
    ("docentes", "porcentajeDiscapacidad"),
    ("docentes", "carnetConadis"),
    ("docentes", "numeroCarnetConadis"),
    ("docentes", "montoBeca"),
    ("docentes", "numPubRevistasCientifIndexadas"),
}

_STUDENT_CODE_FIELDS = {
    "tipoDocumentoId",
    "sexoId",
    "generoId",
    "estadocivilId",
    "etniaId",
    "pueblonacionalidadId",
    "tipoSangre",
    "discapacidad",
    "tipoDiscapacidad",
    "paisNacionalidadId",
    "provinciaNacimientoId",
    "cantonNacimientoId",
    "paisResidenciaId",
    "provinciaResidenciaId",
    "cantonResidenciaId",
    "tipoColegioId",
    "modalidadCarrera",
    "jornadaCarrera",
    "tipoMatriculaId",
    "nivelAcademicoQueCursa",
    "duracionPeriodoAcademico",
    "haRepetidoAlMenosUnaMateria",
    "paraleloId",
    "haPerdidoLaGratuidad",
    "recibePensionDiferenciada",
    "estudianteocupacionId",
    "ingresosestudianteId",
    "bonodesarrolloId",
    "haRealizadoPracticasPreprofesionales",
    "nroHorasPracticasPreprofesionalesPorPeriodo",
    "entornoInstitucionalPracticasProfesionales",
    "sectorEconomicoPracticaProfesional",
    "tipoBecaId",
    "primeraRazonBecaId",
    "segundaRazonBecaId",
    "terceraRazonBecaId",
    "cuartaRazonBecaId",
    "quintaRazonBecaId",
    "sextaRazonBecaId",
    "porcientoBecaCoberturaArancel",
    "porcientoBecaCoberturaManuntencion",
    "financiamientoBeca",
    "participaEnProyectoVinculacionSociedad",
    "tipoAlcanceProyectoVinculacionId",
    "nivelFormacionPadre",
    "nivelFormacionMadre",
}

_TEACHER_CODE_FIELDS = {
    "tipoDocumentoId",
    "sexoId",
    "generoId",
    "estadoCivilId",
    "etniaId",
    "nacionalidadId",
    "provinciaSufragio",
    "discapacidad",
    "tipoDiscapacidad",
    "carnetConadis",
    "paisNacionalidadId",
    "nivelFormacion",
    "relacionLaboralIESId",
    "ingresoConCursoMeritos",
    "escalafonDocenteId",
    "cargoDirectivoId",
    "tiempoDedicacionId",
    "docenciaTecnicoSuperior",
    "docenciaTecnologico",
    "docenciaTecnologicoUniversitario",
    "docenciaEspecializacionTecnologica",
    "docenciaMaestriaTecnologica",
    "estaEnPeriodoSabatico",
    "estaCursandoEstudiosId",
    "paisEstudiosId",
    "poseeBecaId",
    "tipoBecaId",
    "financiamientoBecaId",
    "pubRevistasCienInIndexadasId",
}

_DOCUMENT_TYPE_CEDULA = "1"
_DOCUMENT_TYPE_PASSPORT = "2"
_DOCUMENT_TYPE_LABELS = {
    _DOCUMENT_TYPE_CEDULA: "Cedula",
    _DOCUMENT_TYPE_PASSPORT: "Pasaporte",
}
_PASSPORT_PATTERNS = (
    ("Ecuador / Espana / Argentina", re.compile(r"^[A-Z]{3}\d{6}$")),
    ("Estados Unidos", re.compile(r"^[A-Z]\d{8}$")),
)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if re.fullmatch(r"-?\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _cell_code(value: Any) -> str:
    return re.sub(r"\s+", " ", _cell_text(value)).strip().upper()


def _is_zero_text(text: str) -> bool:
    if not text:
        return False
    try:
        return float(text.replace(",", ".")) == 0
    except ValueError:
        return False


def _is_no_code(value: Any) -> bool:
    return _cell_code(value) in {"2", "NO", "N", "FALSE", "FALSO"}


def _has_selected_code(value: Any) -> bool:
    code = _cell_code(value)
    return bool(code) and code not in _EMPTY_MARKERS and code not in _UNSELECTED_MARKERS and "SELECCIONE" not in code


def _normalize_document_number(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", _cell_text(value).upper())


def _infer_document_type(number: Any) -> str:
    normalized = _normalize_document_number(number)
    if re.fullmatch(r"\d{10}", normalized):
        return _DOCUMENT_TYPE_CEDULA
    if any(pattern.fullmatch(normalized) for _, pattern in _PASSPORT_PATTERNS):
        return _DOCUMENT_TYPE_PASSPORT
    return ""


def _document_country_format(number: Any) -> str:
    normalized = _normalize_document_number(number)
    for label, pattern in _PASSPORT_PATTERNS:
        if pattern.fullmatch(normalized):
            return label
    return ""


def _document_analysis(row: pd.Series) -> dict[str, Any]:
    type_code = _cell_code(row.get("tipoDocumentoId"))
    number = _normalize_document_number(row.get("numeroIdentificacion"))
    inferred_code = _infer_document_type(number)
    expected_code = inferred_code or (type_code if type_code in _DOCUMENT_TYPE_LABELS else "")
    is_type_selected = type_code in _DOCUMENT_TYPE_LABELS
    is_number_valid = bool(inferred_code)
    format_label = ""

    if inferred_code == _DOCUMENT_TYPE_CEDULA:
        format_label = "Cedula ecuatoriana: 10 digitos"
    elif inferred_code == _DOCUMENT_TYPE_PASSPORT:
        format_label = _document_country_format(number)

    is_consistent = bool(is_type_selected and is_number_valid and type_code == inferred_code)
    suggested_code = expected_code if expected_code in _DOCUMENT_TYPE_LABELS else ""
    issues: list[str] = []
    if not is_type_selected:
        issues.append("tipoDocumentoId debe ser 1 para cedula o 2 para pasaporte")
    if not number:
        issues.append("numeroIdentificacion esta vacio")
    elif not is_number_valid:
        issues.append(
            "numeroIdentificacion no cumple 10 digitos de cedula ni formato de pasaporte permitido"
        )
    if is_type_selected and inferred_code and type_code != inferred_code:
        issues.append(
            f"tipoDocumentoId registrado {type_code} no coincide con el documento; sugerido {inferred_code}"
        )

    return {
        "tipo_actual": type_code,
        "tipo_actual_label": _DOCUMENT_TYPE_LABELS.get(type_code, "Sin seleccionar"),
        "numero": number,
        "tipo_sugerido": suggested_code,
        "tipo_sugerido_label": _DOCUMENT_TYPE_LABELS.get(suggested_code, ""),
        "formato": format_label,
        "valido": is_consistent,
        "numero_valido": is_number_valid,
        "tipo_valido": is_type_selected,
        "observaciones": issues,
    }


def _is_no_beca_student(row: pd.Series) -> bool:
    return _cell_code(row.get("tipoBecaId")) == "3"


def _is_no_study_teacher(row: pd.Series) -> bool:
    return _cell_code(row.get("estaCursandoEstudiosId")) in {"8", "NA", "N/A", "NO APLICA"}


def _is_no_beca_teacher(row: pd.Series) -> bool:
    return _is_no_study_teacher(row) or _is_no_code(row.get("poseeBecaId"))


def _field_has_contextual_no_aplica_code(row: pd.Series, column: str, target: str, code: str) -> bool:
    if target == "estudiantes":
        if column == "pueblonacionalidadId" and code == "34":
            return True
        if column == "tipoDiscapacidad" and code == "7":
            return True
        if column == "entornoInstitucionalPracticasProfesionales" and code == "5":
            return True
        if column == "sectorEconomicoPracticaProfesional" and code == "22":
            return True
        if column == "financiamientoBeca" and code == "4":
            return True
        if column == "tipoAlcanceProyectoVinculacionId" and code == "5":
            return True

    if target == "docentes":
        if column == "tipoDiscapacidad" and code == "7":
            return True
        if column == "tipoBecaId" and code == "3":
            return True
        if column == "financiamientoBecaId" and code == "5":
            return True

    return False


def _field_allows_zero(row: pd.Series, column: str, target: str) -> bool:
    if (target, column) in _ZERO_ALLOWED_FIELDS:
        return True
    if target == "estudiantes":
        if column in {"porcentajeDiscapacidad", "numCarnetConadis"}:
            return _is_no_code(row.get("discapacidad"))
        if column == "montoBeca":
            return _is_no_beca_student(row)
    if target == "docentes":
        if column in {"porcentajeDiscapacidad", "carnetConadis", "numeroCarnetConadis"}:
            return _is_no_code(row.get("discapacidad"))
        if column == "montoBeca":
            return _is_no_beca_teacher(row)
        if column == "numPubRevistasCientifIndexadas":
            return _is_no_code(row.get("pubRevistasCienInIndexadasId"))
    return True


def _field_allows_no_aplica(row: pd.Series, column: str, target: str) -> bool:
    if target == "estudiantes":
        if column == "pueblonacionalidadId":
            etnia = row.get("etniaId")
            return _has_selected_code(etnia) and _cell_code(etnia) != "1"
        if column == "porcientoBecaCoberturaManuntencion":
            return True
        if column == "porcientoBecaCoberturaArancel":
            return _is_no_beca_student(row)
        if column in {"montoAyudaEconomica", "montoCreditoEducativo"}:
            return True
        if _is_no_code(row.get("discapacidad")) and column in {
            "porcentajeDiscapacidad",
            "numCarnetConadis",
            "tipoDiscapacidad",
        }:
            return True
        if _is_no_code(row.get("haRealizadoPracticasPreprofesionales")) and column in {
            "nroHorasPracticasPreprofesionalesPorPeriodo",
            "entornoInstitucionalPracticasProfesionales",
            "sectorEconomicoPracticaProfesional",
        }:
            return True
        if _is_no_beca_student(row) and column in {
            "primeraRazonBecaId",
            "segundaRazonBecaId",
            "terceraRazonBecaId",
            "cuartaRazonBecaId",
            "quintaRazonBecaId",
            "sextaRazonBecaId",
            "montoBeca",
            "porcientoBecaCoberturaArancel",
            "porcientoBecaCoberturaManuntencion",
            "financiamientoBeca",
        }:
            return True
        if _cell_code(row.get("participaEnProyectoVinculacionSociedad")) in {"2", "3", "NO", "N"} and column == "tipoAlcanceProyectoVinculacionId":
            return True

    if target == "docentes":
        if _is_no_code(row.get("discapacidad")) and column in {
            "tipoDiscapacidad",
            "porcentajeDiscapacidad",
            "carnetConadis",
            "numeroCarnetConadis",
        }:
            return True
        if _is_no_code(row.get("estaEnPeriodoSabatico")) and column == "fechaInicioPeriodoSabatico":
            return True
        if _is_no_study_teacher(row) and column in {
            "institucionDondeCursaEstudios",
            "paisEstudiosId",
            "tituloAObtener",
            "poseeBecaId",
            "tipoBecaId",
            "montoBeca",
            "financiamientoBecaId",
        }:
            return True
        if _is_no_beca_teacher(row) and column in {"tipoBecaId", "montoBeca", "financiamientoBecaId"}:
            return True
        if _is_no_code(row.get("pubRevistasCienInIndexadasId")) and column == "numPubRevistasCientifIndexadas":
            return True

    return False


def _audit_field_filled(row: pd.Series, column: str, target: str) -> bool:
    if column == "tipoDocumentoId":
        analysis = _document_analysis(row)
        return bool(analysis["tipo_valido"] and analysis["tipo_actual"] == analysis["tipo_sugerido"])
    if column == "numeroIdentificacion":
        return bool(_document_analysis(row)["numero_valido"])
    text = _cell_text(row.get(column))
    code = _cell_code(row.get(column))
    if code in _EMPTY_MARKERS or "SELECCIONE" in code:
        return False
    if _is_zero_text(text) and (target, column) in _ZERO_ALLOWED_FIELDS:
        return True
    if _is_zero_text(text) and (target, column) in _ZERO_CONTEXT_FIELDS:
        return _field_allows_zero(row, column, target)
    if code in _NO_APLICA_MARKERS:
        return _field_allows_no_aplica(row, column, target)
    if _field_has_contextual_no_aplica_code(row, column, target, code):
        return _field_allows_no_aplica(row, column, target)
    if code in {"0001-01-01", "0001-01-01 00:00:00"}:
        return _field_allows_no_aplica(row, column, target)
    code_fields = _STUDENT_CODE_FIELDS if target == "estudiantes" else _TEACHER_CODE_FIELDS
    if column in code_fields and code in _UNSELECTED_MARKERS:
        return False
    if _is_zero_text(text) and column in code_fields:
        return False
    return bool(text)


def _missing_columns(row: pd.Series, report_columns: list[str], target: str) -> list[str]:
    return [column for column in report_columns if not _audit_field_filled(row, column, target)]


def _count_filled(dataframe: pd.DataFrame, column: str, target: str) -> int:
    return int(sum(1 for _, row in dataframe.iterrows() if _audit_field_filled(row, column, target)))


def _excel_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _missing_export_record(
    row: pd.Series,
    target: str,
    report_columns: list[str],
    missing: list[str],
) -> dict[str, Any]:
    filled = len(report_columns) - len(missing)
    document = _document_analysis(row)
    record: dict[str, Any] = {
        "tipo": target,
        "codigo": str(row.get("codigo") or "").strip(),
        "identificacion": str(row.get("numeroIdentificacion") or "").strip(),
        "tipo_documento_actual": document.get("tipo_actual"),
        "tipo_documento_sugerido": document.get("tipo_sugerido"),
        "documento_formato": document.get("formato"),
        "documento_observaciones": "; ".join(document.get("observaciones") or []),
        "nombre": str(row.get("nombreCompleto") or "Sin nombre").strip(),
        "carrera": str(row.get("nombreCarrera") or "Sin carrera").strip(),
        "correo": str(row.get("correoElectronico") or row.get("correoPersonal") or "").strip(),
        "telefono": str(row.get("numeroCelular") or "").strip(),
        "campos_llenos": filled,
        "campos_pendientes": len(missing),
        "campos_totales": len(report_columns),
        "porcentaje_lleno": round((filled / max(len(report_columns), 1)) * 100, 2),
        "campos_faltantes": ", ".join(missing),
    }
    record.update({column: _excel_value(row.get(column)) for column in report_columns})
    return record


def _dedupe_missing_records(records: list[dict[str, Any]], include_career: bool) -> list[dict[str, Any]]:
    selected: dict[tuple[str, str], dict[str, Any]] = {}
    for record in sorted(
        records,
        key=lambda item: (
            -int(item.get("campos_pendientes") or 0),
            str(item.get("nombre") or ""),
            str(item.get("carrera") or ""),
        ),
    ):
        identity = str(record.get("identificacion") or record.get("codigo") or "").strip().casefold()
        if not identity:
            identity = f"sin-id:{record.get('nombre', '')}".casefold()
        career_key = str(record.get("carrera") or "").strip().casefold() if include_career else ""
        selected.setdefault((identity, career_key), record)
    return sorted(selected.values(), key=lambda item: (str(item.get("carrera") or ""), str(item.get("nombre") or "")))


def _unique_sheet_name(writer: pd.ExcelWriter, base: str) -> str:
    clean = _safe_filename(base).replace("'", "")[:31] or "Hoja"
    name = clean
    suffix = 2
    while name in writer.book.sheetnames:
        extra = f" {suffix}"
        name = f"{clean[:31 - len(extra)]}{extra}"
        suffix += 1
    return name


def _build_document_summary(dataframe: pd.DataFrame) -> dict[str, Any]:
    totals = {
        "total_registros": int(len(dataframe)),
        "documentos_validos": 0,
        "cedulas_validas": 0,
        "pasaportes_validos": 0,
        "tipo_incorrecto": 0,
        "numero_invalido": 0,
        "sin_tipo": 0,
        "sin_numero": 0,
        "pendientes": 0,
        "porcentaje_validos": 0,
        "reglas": [
            {
                "codigo": 1,
                "tipo": "Cedula",
                "formato": "10 digitos numericos",
            },
            {
                "codigo": 2,
                "tipo": "Pasaporte",
                "formato": "Ecuador/Espana/Argentina: 3 letras + 6 numeros; Estados Unidos: 1 letra + 8 numeros",
            },
        ],
    }
    if dataframe.empty:
        return totals

    for _, row in dataframe.iterrows():
        analysis = _document_analysis(row)
        suggested = analysis["tipo_sugerido"]
        if analysis["valido"]:
            totals["documentos_validos"] += 1
            if suggested == _DOCUMENT_TYPE_CEDULA:
                totals["cedulas_validas"] += 1
            elif suggested == _DOCUMENT_TYPE_PASSPORT:
                totals["pasaportes_validos"] += 1
        else:
            totals["pendientes"] += 1
        if not analysis["tipo_valido"]:
            totals["sin_tipo"] += 1
        elif analysis["tipo_sugerido"] and analysis["tipo_actual"] != analysis["tipo_sugerido"]:
            totals["tipo_incorrecto"] += 1
        if not analysis["numero"]:
            totals["sin_numero"] += 1
        elif not analysis["numero_valido"]:
            totals["numero_invalido"] += 1

    totals["porcentaje_validos"] = round(
        (totals["documentos_validos"] / max(totals["total_registros"], 1)) * 100,
        2,
    )
    return totals


def _build_senescyt_audit_from_dataframe(
    target: str,
    dataframe: pd.DataFrame,
    report_columns: list[str],
    selected_careers: list[str] | None = None,
) -> dict[str, Any]:
    selected_careers = selected_careers or []
    total_rows = int(len(dataframe))
    total_cells = max(total_rows * len(report_columns), 1)
    field_totals = {column: _count_filled(dataframe, column, target) for column in report_columns} if total_rows else {
        column: 0 for column in report_columns
    }
    filled_cells = int(sum(field_totals.values()))

    row_summaries: list[dict[str, Any]] = []
    missing_records: list[dict[str, Any]] = []
    missing_field_records: list[dict[str, Any]] = []
    for _, row in dataframe.iterrows():
        missing = _missing_columns(row, report_columns, target)
        filled = len(report_columns) - len(missing)
        document_analysis = _document_analysis(row)
        row_summary = {
            "codigo": str(row.get("codigo") or "").strip(),
            "identificacion": str(row.get("numeroIdentificacion") or "").strip(),
            "documento": document_analysis,
            "nombre": str(row.get("nombreCompleto") or "Sin nombre").strip(),
            "nombre_carrera": str(row.get("nombreCarrera") or "Sin carrera").strip(),
            "correo": str(row.get("correoElectronico") or row.get("correoPersonal") or "").strip(),
            "telefono": str(row.get("numeroCelular") or "").strip(),
            "campos_llenos": filled,
            "campos_pendientes": len(missing),
            "campos_totales": len(report_columns),
            "porcentaje_lleno": round((filled / max(len(report_columns), 1)) * 100, 2),
            "campos_faltantes": missing,
            "fields": {column: _excel_value(row.get(column)) for column in report_columns},
        }
        row_summaries.append(row_summary)
        if missing:
            missing_records.append(_missing_export_record(row, target, report_columns, missing))
            for column in missing:
                missing_field_records.append({
                    "tipo": target,
                    "codigo": row_summary["codigo"],
                    "identificacion": row_summary["identificacion"],
                    "nombre": row_summary["nombre"],
                    "carrera": row_summary["nombre_carrera"],
                    "correo": row_summary["correo"],
                    "telefono": row_summary["telefono"],
                    "campo": column,
                    "valor_actual": _excel_value(row.get(column)),
                })

    career_rows: list[dict[str, Any]] = []
    if total_rows:
        for career_name, group in dataframe.groupby("nombreCarrera", dropna=False):
            group_filled = int(sum(_count_filled(group, column, target) for column in report_columns))
            group_cells = max(len(group) * len(report_columns), 1)
            career_missing_students = [
                item for item in row_summaries if item["nombre_carrera"] == str(career_name or "Sin carrera")
            ]
            career_rows.append({
                "nombre_carrera": str(career_name or "Sin carrera"),
                "total_registros": int(len(group)),
                "campos_llenos": group_filled,
                "campos_totales": group_cells,
                "campos_pendientes": int(group_cells - group_filled),
                "registros_con_pendientes": sum(1 for item in career_missing_students if item["campos_pendientes"] > 0),
                "porcentaje_lleno": round((group_filled / group_cells) * 100, 2),
            })

    missing_fields = sorted(
        (
            {
                "campo": column,
                "llenos": filled,
                "pendientes": total_rows - filled,
                "porcentaje_lleno": round((filled / max(total_rows, 1)) * 100, 2),
            }
            for column, filled in field_totals.items()
        ),
        key=lambda item: (-item["pendientes"], item["campo"]),
    )
    row_summaries.sort(key=lambda item: (-item["campos_pendientes"], item["nombre"]))
    career_rows.sort(key=lambda item: item["nombre_carrera"])
    missing_records_by_career = _dedupe_missing_records(missing_records, include_career=True)
    missing_records_global = _dedupe_missing_records(missing_records, include_career=False)
    document_summary = _build_document_summary(dataframe)

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "target": target,
        "career_filter": selected_careers,
        "dataframe": dataframe,
        "report_columns": report_columns,
        "summary": {
            "total_registros": total_rows,
            "total_carreras": len(career_rows),
            "total_columnas": len(report_columns),
            "campos_llenos": filled_cells,
            "campos_totales": total_cells if total_rows else 0,
            "campos_pendientes": max((total_rows * len(report_columns)) - filled_cells, 0),
            "porcentaje_lleno": round((filled_cells / total_cells) * 100, 2) if total_rows else 0,
            "registros_con_pendientes": sum(1 for item in row_summaries if item["campos_pendientes"] > 0),
        },
        "documentos": document_summary,
        "careers": career_rows,
        "rows": [item for item in row_summaries if item["campos_pendientes"] > 0][:100],
        "missing_fields": missing_fields,
        "missing_records": missing_records_by_career,
        "missing_records_global": missing_records_global,
        "missing_field_records": missing_field_records,
    }


def _build_senescyt_audit(target: str, careers: list[str] | None = None) -> dict[str, Any]:
    selected_careers = _normalize_career_filters(careers)
    dataframe, report_columns = _load_senescyt_audit_dataframe(target)
    dataframe = _filter_by_career(dataframe, selected_careers)
    return _build_senescyt_audit_from_dataframe(target, dataframe, report_columns, selected_careers)


def _audit_export_workbook(report: dict[str, Any], mode: str) -> bytes:
    dataframe: pd.DataFrame = report["dataframe"]
    report_columns: list[str] = report["report_columns"]
    target = report["target"]
    missing_export_columns = [
        "tipo",
        "codigo",
        "identificacion",
        "tipo_documento_actual",
        "tipo_documento_sugerido",
        "documento_formato",
        "documento_observaciones",
        "nombre",
        "carrera",
        "correo",
        "telefono",
        "campos_llenos",
        "campos_pendientes",
        "campos_totales",
        "porcentaje_lleno",
        "campos_faltantes",
        *report_columns,
    ]

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if mode == "faltantes":
            global_rows = report.get("missing_records_global") or []
            career_rows = report.get("missing_records") or []
            detail_rows = report.get("missing_field_records") or []
            pd.DataFrame(global_rows, columns=missing_export_columns).to_excel(
                writer,
                index=False,
                sheet_name="Faltantes global",
            )
            pd.DataFrame(career_rows, columns=missing_export_columns).to_excel(
                writer,
                index=False,
                sheet_name="Faltantes carreras",
            )
            pd.DataFrame(detail_rows).to_excel(
                writer,
                index=False,
                sheet_name="Detalle campos",
            )
            if career_rows:
                career_dataframe = pd.DataFrame(career_rows, columns=missing_export_columns)
                for career_name, group in career_dataframe.groupby("carrera", dropna=False):
                    sheet_name = _unique_sheet_name(writer, f"F {career_name}")
                    group.to_excel(writer, index=False, sheet_name=sheet_name)
        else:
            export_columns = report_columns + ["codigo", "nombreCompleto", "nombreCarrera"]
            if dataframe.empty:
                pd.DataFrame(columns=export_columns).to_excel(writer, index=False, sheet_name="Datos")
            else:
                sheet_count = 0
                for career_name, group in dataframe.groupby("nombreCarrera", dropna=False):
                    sheet_count += 1
                    sheet_name = _unique_sheet_name(writer, str(career_name or f"Carrera {sheet_count}"))
                    group[export_columns].to_excel(writer, index=False, sheet_name=sheet_name)
                if sheet_count == 0:
                    dataframe[export_columns].to_excel(writer, index=False, sheet_name=target.title())

        for worksheet in writer.book.worksheets:
            _style_header(worksheet)
            worksheet.freeze_panes = "A2"
            for column in worksheet.columns:
                column_letter = column[0].column_letter
                width = min(max(len(str(cell.value or "")) for cell in column) + 2, 48)
                worksheet.column_dimensions[column_letter].width = max(width, 12)

    output.seek(0)
    return output.getvalue()


def _audit_export_zip(report: dict[str, Any], mode: str) -> bytes:
    dataframe: pd.DataFrame = report["dataframe"]
    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as archive:
        if dataframe.empty or "nombreCarrera" not in dataframe.columns:
            archive.writestr(
                f"senescyt_{report['target']}_{mode}.xlsx",
                _audit_export_workbook(report, mode),
            )
        else:
            career_series = (
                dataframe["nombreCarrera"]
                .fillna("Sin carrera")
                .astype(str)
                .str.strip()
                .replace({"": "Sin carrera"})
            )
            career_names = sorted(career_series.unique(), key=lambda value: str(value).casefold())
            for index, career_name in enumerate(career_names, start=1):
                group = dataframe[career_series == career_name].copy()
                career_report = _build_senescyt_audit_from_dataframe(
                    report["target"],
                    group,
                    report["report_columns"],
                    [str(career_name)],
                )
                filename = f"{index:02d}_{_safe_filename(str(career_name))}_{mode}.xlsx"
                archive.writestr(filename, _audit_export_workbook(career_report, mode))
    output.seek(0)
    return output.getvalue()


@router.get("/catalogo")
def senescyt_catalog(
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
) -> dict[str, Any]:
    del current_user
    try:
        careers = _career_catalog()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error cargando catálogo SENESCYT: {exc}") from exc
    return {
        "careers": careers,
        "targets": sorted(_REPORT_TARGETS),
        "export_modes": sorted(_EXPORT_MODES),
    }


@router.get("/datos")
def senescyt_audit_report(
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
    target: Annotated[str, Query(pattern="^(estudiantes|docentes)$")] = "estudiantes",
    carrera: Annotated[list[str] | None, Query()] = None,
) -> dict[str, Any]:
    del current_user
    try:
        report = _build_senescyt_audit(target, carrera)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error generando auditoria SENESCYT: {exc}") from exc

    return {
        "generated_at": report["generated_at"],
        "target": report["target"],
        "career_filter": report["career_filter"],
        "summary": report["summary"],
        "careers": report["careers"],
        "rows": report["rows"],
        "missing_fields": report["missing_fields"],
        "report_columns": report["report_columns"],
    }


@router.get("/datos/export")
def senescyt_audit_export(
    current_user: Annotated[SessionUser, Depends(_SENESCYT_ACCESS)],
    target: Annotated[str, Query(pattern="^(estudiantes|docentes)$")] = "estudiantes",
    mode: Annotated[str, Query(pattern="^(completo|faltantes)$")] = "completo",
    carrera: Annotated[list[str] | None, Query()] = None,
) -> StreamingResponse:
    del current_user
    try:
        report = _build_senescyt_audit(target, carrera)
        content = _audit_export_zip(report, mode)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Error exportando ZIP SENESCYT: {exc}") from exc

    selected_careers = _normalize_career_filters(carrera)
    suffix = _safe_filename("_".join(selected_careers[:3])) if selected_careers else "todas_las_carreras"
    if len(selected_careers) > 3:
        suffix = f"{suffix}_y_{len(selected_careers) - 3}_mas"
    filename = f"senescyt_{target}_{mode}_{suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
