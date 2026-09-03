from __future__ import annotations

import asyncio
import base64
import hashlib
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
from threading import Lock
from typing import Any, Literal
from urllib.parse import quote, urlencode
from uuid import uuid4

import httpx
from cryptography.fernet import Fernet, InvalidToken
from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from app.core.config import get_settings
from app.core.security import SessionUser, require_roles
from app.integrations.moodle.client import MoodleClient
from app.integrations.moodle.exceptions import MoodleError
from app.services.db import get_connection
from app.services.graph import graph_get, graph_patch, graph_post

router = APIRouter(prefix="/api/admin/credenciales", tags=["credenciales"])

AdminOnly = Depends(require_roles("ADMINISTRADOR"))

_MAX_FILE_BYTES = 8 * 1024 * 1024
_MAX_USERS = 300
_MAX_EMAIL_ATTEMPTS = 1000
_REPORT_TTL = timedelta(minutes=15)
_LICENSE_CACHE_TTL = timedelta(minutes=10)
_EXCEL_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PERMANENT_PASSWORD_NOTE = "Contraseña permanente; no requiere cambio en el primer ingreso."
_LICENSE_NAMES_BY_PERSON_TYPE = {
    "ESTUDIANTE": {
        "STANDARDWOFFPACK_STUDENT": "Office 365 A1 para estudiantes",
        "STANDARDPACK_STUDENT": "Office 365 Educación E1 para estudiantes",
        "ENTERPRISEPACK_STUDENT": "Office 365 Educación E3 para estudiantes",
        "ENTERPRISEPREMIUM_STUDENT": "Office 365 A5 para estudiantes",
        "ENTERPRISEPREMIUM_NOPSTNCONF_STUDENT": "Office 365 A5 sin conferencias para estudiantes",
        "M365EDU_A3_STUDENT": "Microsoft 365 A3 para estudiantes",
        "M365EDU_A3_STUUSEBNFT": "Microsoft 365 A3 para estudiantes",
        "M365EDU_A5_STUDENT": "Microsoft 365 A5 para estudiantes",
    },
    "PROFESOR": {
        "STANDARDWOFFPACK_FACULTY": "Office 365 A1 para profesores",
        "STANDARDPACK_FACULTY": "Office 365 Educación E1 para profesores",
        "ENTERPRISEPACK_FACULTY": "Office 365 Educación E3 para profesores",
        "ENTERPRISEPREMIUM_FACULTY": "Office 365 A5 para profesores",
        "ENTERPRISEPREMIUM_NOPSTNCONF_FACULTY": "Office 365 A5 sin conferencias para profesores",
        "M365EDU_A3_FACULTY": "Microsoft 365 A3 para profesores",
        "M365EDU_A5_FACULTY": "Microsoft 365 A5 para profesores",
    },
}
_DEFAULT_LICENSE_SKU_BY_PERSON_TYPE = {
    "ESTUDIANTE": "STANDARDWOFFPACK_STUDENT",
    "PROFESOR": "STANDARDWOFFPACK_FACULTY",
}
_TEMPLATE_HEADERS = [
    "primer_nombre",
    "segundo_nombre",
    "primer_apellido",
    "segundo_apellido",
    "cedula",
]
_HEADER_ALIASES = {
    "primer_nombre": {"primer_nombre", "primernombre", "nombre_1", "nombre1"},
    "segundo_nombre": {"segundo_nombre", "segundonombre", "nombre_2", "nombre2"},
    "primer_apellido": {"primer_apellido", "primerapellido", "apellido_1", "apellido1"},
    "segundo_apellido": {"segundo_apellido", "segundoapellido", "apellido_2", "apellido2"},
    "cedula": {"cedula", "identificacion", "documento", "numero_de_cedula"},
}


class CredentialPersonPayload(BaseModel):
    primer_nombre: str = Field(default="", max_length=120)
    segundo_nombre: str = Field(default="", max_length=120)
    primer_apellido: str = Field(default="", max_length=120)
    segundo_apellido: str = Field(default="", max_length=120)
    cedula: str = Field(default="", max_length=20)
    fila_origen: int | None = Field(default=None, ge=1)


class CredentialProvisionPayload(BaseModel):
    tipo_persona: Literal["ESTUDIANTE", "PROFESOR"] = "ESTUDIANTE"
    modo: Literal["INDIVIDUAL", "EXCEL"]
    usuarios: list[CredentialPersonPayload] = Field(min_length=1, max_length=_MAX_USERS)


@dataclass(slots=True)
class _CachedReport:
    owner: str
    filename: str
    content: bytes
    expires_at: datetime


@dataclass(frozen=True, slots=True)
class _EducationLicense:
    person_type: Literal["ESTUDIANTE", "PROFESOR"]
    sku_id: str
    sku_part_number: str
    name: str
    capability_status: str
    enabled_units: int
    consumed_units: int
    available_units: int


_REPORTS: dict[str, _CachedReport] = {}
_REPORTS_LOCK = Lock()
_IDENTITY_LOCK = asyncio.Lock()
_LICENSE_CACHE: dict[tuple[str, str, str], tuple[datetime, _EducationLicense]] = {}
_LICENSE_CACHE_LOCK = Lock()


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _archive_secret() -> str:
    settings = get_settings()
    configured = getattr(settings, "credential_archive_secret", None)
    if configured:
        if hasattr(configured, "get_secret_value"):
            secret = configured.get_secret_value().strip()
        else:
            secret = str(configured).strip()
        if secret:
            return secret
    return settings.signing_secret


def _credential_cipher() -> Fernet:
    digest = hashlib.sha256(
        f"intec-credential-archive-v1:{_archive_secret()}".encode("utf-8")
    ).digest()
    return Fernet(base64.urlsafe_b64encode(digest))


def _encrypt_credential_password(password: str) -> str:
    normalized = str(password or "")
    if not normalized:
        return ""
    token = _credential_cipher().encrypt(normalized.encode("utf-8")).decode("ascii")
    return f"v1:{token}"


def _decrypt_credential_password(value: Any) -> str:
    encrypted = str(value or "").strip()
    if not encrypted:
        return ""
    version, separator, token = encrypted.partition(":")
    if version != "v1" or not separator or not token:
        raise RuntimeError("El formato de la contraseña archivada no es válido")
    try:
        return _credential_cipher().decrypt(token.encode("ascii")).decode("utf-8")
    except (InvalidToken, UnicodeDecodeError, ValueError) as exc:
        raise RuntimeError(
            "No se pudo descifrar la contraseña archivada; verifique CREDENTIAL_ARCHIVE_SECRET"
        ) from exc


def _slug(value: Any) -> str:
    normalized = unicodedata.normalize("NFD", _clean(value).casefold())
    without_marks = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return re.sub(r"[^a-z0-9]+", "_", without_marks).strip("_")


def _credential_token(value: Any) -> str:
    return _slug(value).replace("_", "")


def _normalize_cedula(value: Any) -> str:
    raw = _clean(value)
    if raw.endswith(".0") and raw[:-2].isdigit():
        raw = raw[:-2]
    raw = re.sub(r"[\s.-]+", "", raw)
    if raw.isdigit() and len(raw) < 10:
        raw = raw.zfill(10)
    return raw


def _normalized_person(person: CredentialPersonPayload) -> dict[str, Any]:
    return {
        "primer_nombre": _clean(person.primer_nombre),
        "segundo_nombre": _clean(person.segundo_nombre),
        "primer_apellido": _clean(person.primer_apellido),
        "segundo_apellido": _clean(person.segundo_apellido),
        "cedula": _normalize_cedula(person.cedula),
        "fila_origen": person.fila_origen,
    }


def _person_errors(person: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not _credential_token(person["primer_nombre"]):
        errors.append("El primer nombre es obligatorio")
    if not _credential_token(person["primer_apellido"]):
        errors.append("El primer apellido es obligatorio")
    if not re.fullmatch(r"\d{10}", str(person["cedula"])):
        errors.append("La cédula debe contener exactamente 10 dígitos")
    for field, label in (
        ("primer_nombre", "primer nombre"),
        ("segundo_nombre", "segundo nombre"),
        ("primer_apellido", "primer apellido"),
        ("segundo_apellido", "segundo apellido"),
    ):
        value = str(person[field])
        if len(value) > 120:
            errors.append(f"El {label} supera 120 caracteres")
        if value and not any(char.isalpha() for char in value):
            errors.append(f"El {label} no contiene letras")
    return errors


def _email_base(person: dict[str, Any]) -> str:
    first_name = _credential_token(person["primer_nombre"])
    first_surname = _credential_token(person["primer_apellido"])
    return f"{first_name}.{first_surname}".strip(".")


def _fit_email_local_part(base: str, suffix: str) -> str:
    available = max(1, 64 - len(suffix))
    prefix = base[:available].rstrip("._-")
    return f"{prefix}{suffix}"[:64].strip("._-")


def _email_candidates(person: dict[str, Any]):
    base = _email_base(person)
    second_initial = _credential_token(person["segundo_apellido"])[:1]
    yield _fit_email_local_part(base, "")
    if second_initial:
        yield _fit_email_local_part(base, second_initial)
    for number in range(1, _MAX_EMAIL_ATTEMPTS + 1):
        yield _fit_email_local_part(base, f"{second_initial}{number}")


def _permanent_password(person: dict[str, Any], year: int | None = None) -> str:
    first_name = _credential_token(person["primer_nombre"])
    first_surname = _credential_token(person["primer_apellido"])
    cedula = str(person["cedula"])
    if not first_name or not first_surname or len(cedula) < 4:
        raise ValueError("No se puede generar la contraseña con la información indicada")
    surname = f"{first_surname[:1].upper()}{first_surname[1:].lower()}"
    return f"{first_name[:1].upper()}{surname}{cedula[-4:]}@{year or datetime.now().year}"


def _graph_domain() -> str:
    domain = _clean(get_settings().graph_user_domain).casefold().lstrip("@")
    if not domain:
        raise RuntimeError("El dominio de Microsoft 365 no está configurado")
    return domain


def _graph_is_configured() -> bool:
    settings = get_settings()
    return bool(_clean(settings.tenant_id) and _clean(settings.client_id) and _clean(settings.client_secret))


def _moodle_is_configured() -> bool:
    settings = get_settings()
    token = settings.moodle_token.get_secret_value().strip() if settings.moodle_token else ""
    return bool(
        settings.moodle_enabled
        and settings.moodle_reads_enabled
        and settings.moodle_writes_enabled
        and _clean(settings.moodle_base_url)
        and token
    )


def _ensure_tables(cursor: Any) -> None:
    cursor.execute(
        """
        IF OBJECT_ID(N'dbo.CREDENCIAL_IDENTIDAD', N'U') IS NULL
        BEGIN
            CREATE TABLE dbo.CREDENCIAL_IDENTIDAD (
                id BIGINT IDENTITY(1,1) NOT NULL CONSTRAINT PK_CREDENCIAL_IDENTIDAD PRIMARY KEY,
                cedula VARCHAR(20) NOT NULL,
                primer_nombre NVARCHAR(120) NOT NULL,
                segundo_nombre NVARCHAR(120) NULL,
                primer_apellido NVARCHAR(120) NOT NULL,
                segundo_apellido NVARCHAR(120) NULL,
                correo_institucional NVARCHAR(150) NOT NULL,
                graph_user_id NVARCHAR(100) NULL,
                moodle_user_id BIGINT NULL,
                estado_graph VARCHAR(50) NULL,
                estado_moodle VARCHAR(50) NULL,
                usuario_creacion NVARCHAR(100) NULL,
                fecha_creacion DATETIME2(0) NOT NULL CONSTRAINT DF_CREDENCIAL_IDENTIDAD_FECHA DEFAULT SYSDATETIME(),
                fecha_actualizacion DATETIME2(0) NULL
            );
            CREATE UNIQUE INDEX UX_CREDENCIAL_IDENTIDAD_CEDULA
                ON dbo.CREDENCIAL_IDENTIDAD (cedula);
            CREATE UNIQUE INDEX UX_CREDENCIAL_IDENTIDAD_CORREO
                ON dbo.CREDENCIAL_IDENTIDAD (correo_institucional);
        END;

        IF OBJECT_ID(N'dbo.CREDENCIAL_APROVISIONAMIENTO', N'U') IS NULL
        BEGIN
            CREATE TABLE dbo.CREDENCIAL_APROVISIONAMIENTO (
                id BIGINT IDENTITY(1,1) NOT NULL CONSTRAINT PK_CREDENCIAL_APROVISIONAMIENTO PRIMARY KEY,
                lote_id VARCHAR(36) NOT NULL,
                tipo_persona VARCHAR(20) NOT NULL,
                modo VARCHAR(20) NOT NULL,
                fila_origen INT NULL,
                cedula VARCHAR(20) NOT NULL,
                nombres NVARCHAR(250) NOT NULL,
                correo_institucional NVARCHAR(150) NULL,
                estado_graph VARCHAR(50) NOT NULL,
                error_graph NVARCHAR(1000) NULL,
                estado_licencia VARCHAR(50) NOT NULL,
                error_licencia NVARCHAR(1000) NULL,
                estado_moodle VARCHAR(50) NOT NULL,
                error_moodle NVARCHAR(1000) NULL,
                estado_general VARCHAR(50) NOT NULL,
                clave_emitida BIT NOT NULL CONSTRAINT DF_CREDENCIAL_APROV_CLAVE DEFAULT 0,
                clave_cifrada NVARCHAR(1000) NULL,
                observacion NVARCHAR(500) NULL,
                numero_descargas INT NOT NULL CONSTRAINT DF_CREDENCIAL_APROV_DESCARGAS DEFAULT 0,
                fecha_ultima_descarga DATETIME2(0) NULL,
                usuario_ultima_descarga NVARCHAR(100) NULL,
                usuario_creacion NVARCHAR(100) NULL,
                fecha_creacion DATETIME2(0) NOT NULL CONSTRAINT DF_CREDENCIAL_APROV_FECHA DEFAULT SYSDATETIME()
            );
            CREATE INDEX IX_CREDENCIAL_APROVISIONAMIENTO_LOTE
                ON dbo.CREDENCIAL_APROVISIONAMIENTO (lote_id, id);
            CREATE INDEX IX_CREDENCIAL_APROVISIONAMIENTO_CEDULA
                ON dbo.CREDENCIAL_APROVISIONAMIENTO (cedula, fecha_creacion DESC);
        END;
        """
    )
    # Dynamic DDL lets SQL Server add all migration columns in one round trip.
    cursor.execute(
        """
        IF COL_LENGTH(N'dbo.CREDENCIAL_APROVISIONAMIENTO', N'tipo_persona') IS NULL
            EXEC(N'ALTER TABLE dbo.CREDENCIAL_APROVISIONAMIENTO ADD tipo_persona VARCHAR(20) NULL');
        IF COL_LENGTH(N'dbo.CREDENCIAL_APROVISIONAMIENTO', N'clave_cifrada') IS NULL
            EXEC(N'ALTER TABLE dbo.CREDENCIAL_APROVISIONAMIENTO ADD clave_cifrada NVARCHAR(1000) NULL');
        IF COL_LENGTH(N'dbo.CREDENCIAL_APROVISIONAMIENTO', N'observacion') IS NULL
            EXEC(N'ALTER TABLE dbo.CREDENCIAL_APROVISIONAMIENTO ADD observacion NVARCHAR(500) NULL');
        IF COL_LENGTH(N'dbo.CREDENCIAL_APROVISIONAMIENTO', N'numero_descargas') IS NULL
            EXEC(N'ALTER TABLE dbo.CREDENCIAL_APROVISIONAMIENTO ADD numero_descargas INT NOT NULL CONSTRAINT DF_CREDENCIAL_APROV_DESCARGAS DEFAULT 0');
        IF COL_LENGTH(N'dbo.CREDENCIAL_APROVISIONAMIENTO', N'fecha_ultima_descarga') IS NULL
            EXEC(N'ALTER TABLE dbo.CREDENCIAL_APROVISIONAMIENTO ADD fecha_ultima_descarga DATETIME2(0) NULL');
        IF COL_LENGTH(N'dbo.CREDENCIAL_APROVISIONAMIENTO', N'usuario_ultima_descarga') IS NULL
            EXEC(N'ALTER TABLE dbo.CREDENCIAL_APROVISIONAMIENTO ADD usuario_ultima_descarga NVARCHAR(100) NULL');
        """
    )
    cursor.execute(
        """
        UPDATE dbo.CREDENCIAL_APROVISIONAMIENTO
           SET tipo_persona = 'ESTUDIANTE'
         WHERE tipo_persona IS NULL;

        IF EXISTS (
            SELECT 1
            FROM sys.columns
            WHERE object_id = OBJECT_ID(N'dbo.CREDENCIAL_APROVISIONAMIENTO')
              AND name = N'tipo_persona'
              AND is_nullable = 1
        )
            ALTER TABLE dbo.CREDENCIAL_APROVISIONAMIENTO
                ALTER COLUMN tipo_persona VARCHAR(20) NOT NULL;
        """
    )


def _table_exists(cursor: Any, table_name: str) -> bool:
    cursor.execute("SELECT CASE WHEN OBJECT_ID(?, N'U') IS NULL THEN 0 ELSE 1 END", table_name)
    row = cursor.fetchone()
    return bool(row and row[0])


def _valid_institutional_email(value: Any) -> str:
    email = _clean(value).casefold()
    try:
        domain = _graph_domain()
    except RuntimeError:
        return ""
    if re.fullmatch(rf"[a-z0-9.!#$%&'*+/=?^_`{{|}}~-]+@{re.escape(domain)}", email):
        return email
    return ""


def _identity_row(cursor: Any, cedula: str) -> dict[str, Any] | None:
    cursor.execute(
        """
        SELECT TOP (1) cedula, correo_institucional, graph_user_id, moodle_user_id,
               estado_graph, estado_moodle
        FROM dbo.CREDENCIAL_IDENTIDAD
        WHERE cedula = ?
        """,
        cedula,
    )
    row = cursor.fetchone()
    if not row:
        return None
    return {
        "cedula": _clean(row.cedula),
        "correo_institucional": _clean(row.correo_institucional).casefold(),
        "graph_user_id": _clean(row.graph_user_id),
        "moodle_user_id": int(row.moodle_user_id) if row.moodle_user_id is not None else None,
        "estado_graph": _clean(row.estado_graph),
        "estado_moodle": _clean(row.estado_moodle),
    }


def _legacy_email_for_cedula(cursor: Any, cedula: str) -> str:
    if _table_exists(cursor, "dbo.CREDENCIALES_CURSO"):
        cursor.execute(
            """
            SELECT TOP (1) usuario_generado
            FROM dbo.CREDENCIALES_CURSO
            WHERE cedula = ? AND NULLIF(LTRIM(RTRIM(usuario_generado)), '') IS NOT NULL
            ORDER BY id DESC
            """,
            cedula,
        )
        row = cursor.fetchone()
        if row:
            email = _valid_institutional_email(row[0])
            if email:
                return email

    if _table_exists(cursor, "dbo.DATOS_ESTUD"):
        if _table_exists(cursor, "dbo.CorreosEstudIntec"):
            cursor.execute(
                """
                SELECT TOP (1)
                    COALESCE(
                        NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(150), ce.CorreoIntec))), N''),
                        NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(150), de.correointec))), N'')
                    )
                FROM dbo.DATOS_ESTUD de
                LEFT JOIN dbo.CorreosEstudIntec ce
                  ON TRY_CONVERT(int, ce.codestud) = TRY_CONVERT(int, de.codigo_estud)
                WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(20), de.Cedula_Est))) = ?
                ORDER BY TRY_CONVERT(int, de.codigo_estud) DESC
                """,
                cedula,
            )
        else:
            cursor.execute(
                """
                SELECT TOP (1) LTRIM(RTRIM(TRY_CONVERT(nvarchar(150), correointec)))
                FROM dbo.DATOS_ESTUD
                WHERE LTRIM(RTRIM(TRY_CONVERT(varchar(20), Cedula_Est))) = ?
                ORDER BY TRY_CONVERT(int, codigo_estud) DESC
                """,
                cedula,
            )
        row = cursor.fetchone()
        if row:
            return _valid_institutional_email(row[0])
    return ""


def _existing_email_for_cedula(cedula: str) -> str:
    with get_connection() as conn:
        cursor = conn.cursor()
        _ensure_tables(cursor)
        identity = _identity_row(cursor, cedula)
        if identity:
            conn.commit()
            return str(identity["correo_institucional"])
        legacy = _legacy_email_for_cedula(cursor, cedula)
        conn.commit()
        return legacy


def _email_owner(cursor: Any, email: str) -> str:
    cursor.execute(
        "SELECT TOP (1) cedula FROM dbo.CREDENCIAL_IDENTIDAD WHERE correo_institucional = ?",
        email,
    )
    row = cursor.fetchone()
    if row:
        return _clean(row[0])

    if _table_exists(cursor, "dbo.CREDENCIALES_CURSO"):
        cursor.execute(
            """
            SELECT TOP (1) cedula
            FROM dbo.CREDENCIALES_CURSO
            WHERE LOWER(LTRIM(RTRIM(usuario_generado))) = ?
            ORDER BY id DESC
            """,
            email.casefold(),
        )
        row = cursor.fetchone()
        if row:
            return _clean(row[0])

    if _table_exists(cursor, "dbo.DATOS_ESTUD"):
        if _table_exists(cursor, "dbo.CorreosEstudIntec"):
            cursor.execute(
                """
                SELECT TOP (1) LTRIM(RTRIM(TRY_CONVERT(varchar(20), de.Cedula_Est)))
                FROM dbo.DATOS_ESTUD de
                LEFT JOIN dbo.CorreosEstudIntec ce
                  ON TRY_CONVERT(int, ce.codestud) = TRY_CONVERT(int, de.codigo_estud)
                WHERE LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(150), de.correointec)))) = ?
                   OR LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(150), ce.CorreoIntec)))) = ?
                ORDER BY TRY_CONVERT(int, de.codigo_estud) DESC
                """,
                email.casefold(),
                email.casefold(),
            )
        else:
            cursor.execute(
                """
                SELECT TOP (1) LTRIM(RTRIM(TRY_CONVERT(varchar(20), Cedula_Est)))
                FROM dbo.DATOS_ESTUD
                WHERE LOWER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(150), correointec)))) = ?
                ORDER BY TRY_CONVERT(int, codigo_estud) DESC
                """,
                email.casefold(),
            )
        row = cursor.fetchone()
        if row:
            return _clean(row[0])
    return ""


def _local_email_owner(email: str) -> str:
    with get_connection() as conn:
        cursor = conn.cursor()
        _ensure_tables(cursor)
        owner = _email_owner(cursor, email)
        conn.commit()
        return owner


def _reserve_identity(person: dict[str, Any], email: str, operator: str) -> str | None:
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_tables(cursor)
            current = _identity_row(cursor, str(person["cedula"]))
            if current:
                conn.commit()
                return str(current["correo_institucional"])
            cursor.execute(
                """
                INSERT INTO dbo.CREDENCIAL_IDENTIDAD (
                    cedula, primer_nombre, segundo_nombre, primer_apellido, segundo_apellido,
                    correo_institucional, usuario_creacion
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                str(person["cedula"]),
                str(person["primer_nombre"]),
                str(person["segundo_nombre"]) or None,
                str(person["primer_apellido"]),
                str(person["segundo_apellido"]) or None,
                email,
                operator,
            )
            conn.commit()
            return email
    except Exception:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_tables(cursor)
            current = _identity_row(cursor, str(person["cedula"]))
            conn.commit()
            return str(current["correo_institucional"]) if current else None


def _graph_error_detail(exc: Exception) -> str:
    if isinstance(exc, httpx.HTTPStatusError):
        try:
            payload = exc.response.json()
        except ValueError:
            return "Microsoft Graph respondió con un error"
        if isinstance(payload, dict) and isinstance(payload.get("error"), dict):
            code = _clean(payload["error"].get("code"))
            message = _clean(payload["error"].get("message"))
            return (f"{code}: {message}" if code else message)[:500]
    return _clean(exc)[:500] or "No fue posible completar la operación en Microsoft Graph"


def _license_target(
    person_type: Literal["ESTUDIANTE", "PROFESOR"],
) -> tuple[str, str]:
    settings = get_settings()
    if person_type == "PROFESOR":
        raw_part_number = settings.graph_faculty_license_sku_part_number
        raw_sku_id = settings.graph_faculty_license_sku_id
    else:
        raw_part_number = settings.graph_student_license_sku_part_number
        raw_sku_id = settings.graph_student_license_sku_id

    sku_part_number = (
        _clean(raw_part_number).upper()
        or _DEFAULT_LICENSE_SKU_BY_PERSON_TYPE[person_type]
    )
    if sku_part_number not in _LICENSE_NAMES_BY_PERSON_TYPE[person_type]:
        audience = "profesores" if person_type == "PROFESOR" else "estudiantes"
        raise RuntimeError(
            f"El SKU configurado no corresponde a una licencia educativa para {audience}"
        )
    return sku_part_number, _clean(raw_sku_id).casefold()


def _select_education_license(
    subscribed_skus: list[dict[str, Any]],
    person_type: Literal["ESTUDIANTE", "PROFESOR"],
    sku_part_number: str,
    configured_sku_id: str = "",
) -> _EducationLicense:
    normalized_part_number = _clean(sku_part_number).upper()
    license_names = _LICENSE_NAMES_BY_PERSON_TYPE[person_type]
    audience = "profesores" if person_type == "PROFESOR" else "estudiantes"
    if normalized_part_number not in license_names:
        raise RuntimeError(
            f"El SKU configurado no corresponde a una licencia educativa para {audience}"
        )

    matching = [
        item
        for item in subscribed_skus
        if _clean(item.get("skuPartNumber")).upper() == normalized_part_number
    ]
    if not matching:
        raise RuntimeError(
            f"El tenant no tiene contratada la licencia para {audience} {normalized_part_number}"
        )
    item = matching[0]
    sku_id = _clean(item.get("skuId")).casefold()
    if configured_sku_id and sku_id != configured_sku_id.casefold():
        raise RuntimeError(
            f"El GUID configurado no coincide con el SKU contratado para {audience}"
        )
    if _clean(item.get("appliesTo")).casefold() != "user":
        raise RuntimeError(f"La licencia para {audience} no se puede asignar a usuarios")

    capability_status = _clean(item.get("capabilityStatus"))
    if capability_status.casefold() != "enabled":
        raise RuntimeError(
            f"La licencia para {audience} está en estado {capability_status or 'desconocido'}"
        )

    prepaid = item.get("prepaidUnits") if isinstance(item.get("prepaidUnits"), dict) else {}
    try:
        enabled_units = max(0, int(prepaid.get("enabled") or 0))
    except (TypeError, ValueError):
        enabled_units = 0
    try:
        consumed_units = max(0, int(item.get("consumedUnits") or 0))
    except (TypeError, ValueError):
        consumed_units = 0

    return _EducationLicense(
        person_type=person_type,
        sku_id=sku_id,
        sku_part_number=normalized_part_number,
        name=license_names[normalized_part_number],
        capability_status=capability_status,
        enabled_units=enabled_units,
        consumed_units=consumed_units,
        available_units=max(0, enabled_units - consumed_units),
    )


def _education_license(
    person_type: Literal["ESTUDIANTE", "PROFESOR"],
    force_refresh: bool = False,
) -> _EducationLicense:
    target = _license_target(person_type)
    cache_key = (person_type, target[0], target[1])
    now = datetime.now(timezone.utc)
    with _LICENSE_CACHE_LOCK:
        cached = _LICENSE_CACHE.get(cache_key)
        if (
            not force_refresh
            and cached
            and cached[0] > now
        ):
            return cached[1]

    payload = graph_get(
        "https://graph.microsoft.com/v1.0/subscribedSkus"
        "?$select=skuId,skuPartNumber,capabilityStatus,consumedUnits,prepaidUnits,appliesTo"
    )
    raw_items = payload.get("value") if isinstance(payload, dict) else None
    if not isinstance(raw_items, list):
        raise RuntimeError("Microsoft Graph no devolvió las licencias contratadas")
    license_info = _select_education_license(
        [item for item in raw_items if isinstance(item, dict)],
        person_type,
        target[0],
        target[1],
    )
    with _LICENSE_CACHE_LOCK:
        _LICENSE_CACHE[cache_key] = (now + _LICENSE_CACHE_TTL, license_info)
    return license_info


def _license_configuration(
    person_type: Literal["ESTUDIANTE", "PROFESOR"],
) -> dict[str, Any]:
    default_part_number = _DEFAULT_LICENSE_SKU_BY_PERSON_TYPE[person_type]
    default_name = _LICENSE_NAMES_BY_PERSON_TYPE[person_type][default_part_number]
    license_info: _EducationLicense | None = None
    detail = ""
    if _graph_is_configured():
        try:
            license_info = _education_license(person_type)
        except Exception as exc:
            detail = _graph_error_detail(exc)

    return {
        "person_type": person_type,
        "configured": bool(license_info and license_info.available_units > 0),
        "name": license_info.name if license_info else default_name,
        "sku_part_number": license_info.sku_part_number if license_info else default_part_number,
        "status": (
            "DISPONIBLE"
            if license_info and license_info.available_units > 0
            else "SIN_CUPOS"
            if license_info
            else "NO_DISPONIBLE"
        ),
        "available_units": license_info.available_units if license_info else 0,
        "detail": detail,
    }


def _graph_user(email: str) -> dict[str, Any] | None:
    try:
        return graph_get(
            "https://graph.microsoft.com/v1.0/users/"
            f"{quote(email, safe='')}?$select=id,displayName,mail,userPrincipalName,employeeId,usageLocation,assignedLicenses"
        )
    except httpx.HTTPStatusError as exc:
        if exc.response.status_code == 404:
            return None
        raise


def _graph_users_by_employee_id(cedula: str) -> list[dict[str, Any]]:
    query = urlencode(
        {
            "$filter": f"employeeId eq '{cedula}'",
            "$select": "id,displayName,mail,userPrincipalName,employeeId,usageLocation,assignedLicenses",
            "$top": "2",
        }
    )
    payload = graph_get(f"https://graph.microsoft.com/v1.0/users?{query}")
    values = payload.get("value") if isinstance(payload, dict) else None
    if not isinstance(values, list):
        raise RuntimeError("Microsoft Graph devolvió una búsqueda de identidad no válida")
    return [item for item in values if isinstance(item, dict)]


def _create_graph_user(person: dict[str, Any], email: str, password: str) -> dict[str, Any]:
    settings = get_settings()
    given_name = " ".join(
        value for value in (str(person["primer_nombre"]), str(person["segundo_nombre"])) if value
    )
    surname = " ".join(
        value for value in (str(person["primer_apellido"]), str(person["segundo_apellido"])) if value
    )
    payload: dict[str, Any] = {
        "accountEnabled": True,
        "displayName": f"{given_name} {surname}".strip()[:256],
        "givenName": given_name[:64],
        "surname": surname[:64],
        "mailNickname": email.split("@", 1)[0],
        "userPrincipalName": email,
        "employeeId": str(person["cedula"]),
        "passwordProfile": {
            "forceChangePasswordNextSignIn": False,
            "password": password,
        },
        "passwordPolicies": "DisablePasswordExpiration",
    }
    usage_location = _clean(settings.graph_user_usage_location).upper()
    if usage_location:
        payload["usageLocation"] = usage_location[:2]
    return graph_post("https://graph.microsoft.com/v1.0/users", payload)


def _assign_graph_license(
    graph_user: dict[str, Any] | None,
    education_license: _EducationLicense,
) -> tuple[str, str]:
    if not graph_user:
        return "OMITIDA", ""

    assigned = graph_user.get("assignedLicenses")
    if isinstance(assigned, list) and any(
        _clean(item.get("skuId")).casefold() == education_license.sku_id
        for item in assigned
        if isinstance(item, dict)
    ):
        return f"YA_ASIGNADA_{education_license.person_type}", ""

    if education_license.available_units <= 0:
        audience = "profesores" if education_license.person_type == "PROFESOR" else "estudiantes"
        return f"SIN_CUPOS_{education_license.person_type}", f"No existen licencias para {audience} disponibles"

    user_id = _clean(graph_user.get("id"))
    if not user_id:
        return f"ERROR_LICENCIA_{education_license.person_type}", "Microsoft Graph no devolvió el identificador del usuario"
    settings = get_settings()
    if not _clean(graph_user.get("usageLocation")):
        usage_location = _clean(settings.graph_user_usage_location).upper()
        if usage_location:
            graph_patch(
                f"https://graph.microsoft.com/v1.0/users/{quote(user_id, safe='')}",
                {"usageLocation": usage_location[:2]},
            )
    graph_post(
        f"https://graph.microsoft.com/v1.0/users/{quote(user_id, safe='')}/assignLicense",
        {
            "addLicenses": [{"skuId": education_license.sku_id, "disabledPlans": []}],
            "removeLicenses": [],
        },
    )
    return f"ASIGNADA_{education_license.person_type}", ""


async def _moodle_users(client: MoodleClient, field: str, value: str) -> list[dict[str, Any]]:
    return await client.get_users_by_field(field, [value])


async def _resolve_email(person: dict[str, Any], operator: str, moodle: MoodleClient) -> str:
    cedula = str(person["cedula"])
    existing = await asyncio.to_thread(_existing_email_for_cedula, cedula)
    if existing:
        reserved = await asyncio.to_thread(_reserve_identity, person, existing, operator)
        if reserved:
            return reserved

    domain = _graph_domain()
    async with _IDENTITY_LOCK:
        existing = await asyncio.to_thread(_existing_email_for_cedula, cedula)
        if existing:
            reserved = await asyncio.to_thread(_reserve_identity, person, existing, operator)
            if reserved:
                return reserved

        remote_identity_emails: set[str] = set()
        if _graph_is_configured():
            try:
                graph_identities = await asyncio.to_thread(_graph_users_by_employee_id, cedula)
            except Exception as exc:
                raise RuntimeError(
                    f"No se pudo validar la cédula en Microsoft 365: {_graph_error_detail(exc)}"
                ) from exc
            if len(graph_identities) > 1:
                raise RuntimeError("La cédula está asociada a más de una cuenta Microsoft 365")
            if graph_identities:
                graph_email = _valid_institutional_email(
                    graph_identities[0].get("userPrincipalName") or graph_identities[0].get("mail")
                )
                if not graph_email:
                    raise RuntimeError(
                        "La cédula ya existe en Microsoft 365 con un correo fuera del dominio institucional"
                    )
                remote_identity_emails.add(graph_email)

        if _moodle_is_configured():
            try:
                moodle_identities = await _moodle_users(moodle, "idnumber", cedula)
            except MoodleError as exc:
                raise RuntimeError(f"No se pudo validar la cédula en Moodle: {_clean(exc)[:500]}") from exc
            for moodle_identity in moodle_identities:
                moodle_email = _valid_institutional_email(moodle_identity.get("email"))
                if moodle_email:
                    remote_identity_emails.add(moodle_email)

        if len(remote_identity_emails) > 1:
            raise RuntimeError(
                "La cédula tiene correos institucionales diferentes en Microsoft 365 y Moodle"
            )
        if remote_identity_emails:
            remote_email = next(iter(remote_identity_emails))
            local_owner = await asyncio.to_thread(_local_email_owner, remote_email)
            if local_owner and local_owner != cedula:
                raise RuntimeError("El correo de la identidad existente está reservado para otra cédula")
            reserved = await asyncio.to_thread(_reserve_identity, person, remote_email, operator)
            if reserved:
                return reserved

        for local_part in _email_candidates(person):
            email = f"{local_part}@{domain}"
            local_owner = await asyncio.to_thread(_local_email_owner, email)
            if local_owner and local_owner != cedula:
                continue

            graph_match: dict[str, Any] | None = None
            if _graph_is_configured():
                try:
                    graph_match = await asyncio.to_thread(_graph_user, email)
                except Exception as exc:
                    raise RuntimeError(
                        f"No se pudo validar el correo en Microsoft 365: {_graph_error_detail(exc)}"
                    ) from exc
            if graph_match and _clean(graph_match.get("employeeId")) != cedula:
                continue

            moodle_matches: list[dict[str, Any]] = []
            if _moodle_is_configured():
                try:
                    moodle_matches = await _moodle_users(moodle, "email", email)
                except MoodleError as exc:
                    raise RuntimeError(
                        f"No se pudo validar el correo en Moodle: {_clean(exc)[:500]}"
                    ) from exc
            if any(_clean(item.get("idnumber")) != cedula for item in moodle_matches):
                continue

            reserved = await asyncio.to_thread(_reserve_identity, person, email, operator)
            if reserved:
                return reserved

    raise RuntimeError("No se encontró un correo institucional disponible")


async def _provision_graph(
    person: dict[str, Any], email: str, password: str
) -> tuple[dict[str, Any] | None, str, str, bool]:
    if not _graph_is_configured():
        return None, "NO_CONFIGURADO", "Configure TENANT_ID, CLIENT_ID y CLIENT_SECRET", False
    try:
        existing = await asyncio.to_thread(_graph_user, email)
        if existing:
            owner = await asyncio.to_thread(_local_email_owner, email)
            employee_id = _clean(existing.get("employeeId"))
            if employee_id and employee_id != str(person["cedula"]):
                return existing, "CONFLICTO_GRAPH", "El correo pertenece a otra identidad Microsoft", False
            if owner and owner != str(person["cedula"]):
                return existing, "CONFLICTO_GRAPH", "El correo está reservado para otra cédula", False
            return existing, "EXISTENTE_GRAPH", "", False
        created = await asyncio.to_thread(_create_graph_user, person, email, password)
        return created, "CREADO_GRAPH", "", True
    except Exception as exc:
        return None, "ERROR_GRAPH", _graph_error_detail(exc), False


async def _provision_moodle(
    client: MoodleClient,
    person: dict[str, Any],
    email: str,
    password: str,
) -> tuple[dict[str, Any] | None, str, str, bool]:
    if not _moodle_is_configured():
        return None, "NO_CONFIGURADO", "Habilite lecturas y escrituras Moodle con un token válido", False
    cedula = str(person["cedula"])
    try:
        by_id = await _moodle_users(client, "idnumber", cedula)
        if by_id:
            existing = by_id[0]
            existing_email = _clean(existing.get("email")).casefold()
            if existing_email and existing_email != email.casefold():
                return (
                    existing,
                    "EXISTENTE_OTRO_CORREO",
                    f"La cédula ya existe en Moodle con {existing_email}",
                    False,
                )
            return existing, "EXISTENTE_MOODLE", "", False

        by_email = await _moodle_users(client, "email", email)
        if by_email:
            existing = by_email[0]
            if _clean(existing.get("idnumber")) != cedula:
                return existing, "CONFLICTO_MOODLE", "El correo pertenece a otra identidad Moodle", False
            return existing, "EXISTENTE_MOODLE", "", False

        given_names = " ".join(
            value for value in (str(person["primer_nombre"]), str(person["segundo_nombre"])) if value
        )
        surnames = " ".join(
            value for value in (str(person["primer_apellido"]), str(person["segundo_apellido"])) if value
        )
        created = await client.create_users(
            [
                {
                    "username": email.casefold(),
                    "password": password,
                    "firstname": given_names,
                    "lastname": surnames,
                    "email": email.casefold(),
                    "auth": "manual",
                    "idnumber": cedula,
                    "lang": "es",
                    "preferences": [
                        {"type": "auth_forcepasswordchange", "value": "0"},
                    ],
                }
            ]
        )
        if not created:
            return None, "ERROR_MOODLE", "Moodle no devolvió el usuario creado", False
        return created[0], "CREADO_MOODLE", "", True
    except MoodleError as exc:
        return None, "ERROR_MOODLE", _clean(exc)[:500], False


def _overall_status(graph_status: str, license_status: str, moodle_status: str) -> str:
    graph_ok = graph_status in {"CREADO_GRAPH", "EXISTENTE_GRAPH"}
    moodle_ok = moodle_status in {"CREADO_MOODLE", "EXISTENTE_MOODLE"}
    license_ok = license_status.startswith("ASIGNADA_") or license_status.startswith("YA_ASIGNADA_")
    if graph_ok and moodle_ok and license_ok:
        return "COMPLETO"
    if graph_ok or moodle_ok:
        return "PARCIAL"
    return "ERROR"


def _credential_observation(graph_created: bool, moodle_created: bool) -> str:
    systems = []
    if graph_created:
        systems.append("Microsoft 365")
    if moodle_created:
        systems.append("Moodle")
    if not systems:
        return "Cuenta existente; la contraseña no fue modificada ni archivada."
    return f"{_PERMANENT_PASSWORD_NOTE} Aplicada en {' y '.join(systems)}."


def _update_identity(
    person: dict[str, Any],
    email: str,
    graph_user: dict[str, Any] | None,
    graph_status: str,
    moodle_user: dict[str, Any] | None,
    moodle_status: str,
) -> None:
    graph_id = _clean(graph_user.get("id")) if graph_user else ""
    raw_moodle_id = moodle_user.get("id") if moodle_user else None
    try:
        moodle_id = int(raw_moodle_id) if raw_moodle_id is not None else None
    except (TypeError, ValueError):
        moodle_id = None
    with get_connection() as conn:
        cursor = conn.cursor()
        _ensure_tables(cursor)
        cursor.execute(
            """
            UPDATE dbo.CREDENCIAL_IDENTIDAD
               SET primer_nombre = ?, segundo_nombre = ?, primer_apellido = ?, segundo_apellido = ?,
                   correo_institucional = ?, graph_user_id = COALESCE(NULLIF(?, ''), graph_user_id),
                   moodle_user_id = COALESCE(?, moodle_user_id), estado_graph = ?, estado_moodle = ?,
                   fecha_actualizacion = SYSDATETIME()
             WHERE cedula = ?
            """,
            str(person["primer_nombre"]),
            str(person["segundo_nombre"]) or None,
            str(person["primer_apellido"]),
            str(person["segundo_apellido"]) or None,
            email,
            graph_id,
            moodle_id,
            graph_status,
            moodle_status,
            str(person["cedula"]),
        )
        conn.commit()


async def _provision_person(
    person: dict[str, Any],
    operator: str,
    moodle: MoodleClient,
    education_license: _EducationLicense,
) -> dict[str, Any]:
    email = await _resolve_email(person, operator, moodle)
    password = _permanent_password(person)
    graph_user, graph_status, graph_error, graph_created = await _provision_graph(
        person, email, password
    )

    license_status = "OMITIDA"
    license_error = ""
    if graph_user and graph_status in {"CREADO_GRAPH", "EXISTENTE_GRAPH"}:
        try:
            license_status, license_error = await asyncio.to_thread(
                _assign_graph_license,
                graph_user,
                education_license,
            )
        except Exception as exc:
            license_status = f"ERROR_LICENCIA_{education_license.person_type}"
            license_error = _graph_error_detail(exc)

    moodle_user, moodle_status, moodle_error, moodle_created = await _provision_moodle(
        moodle, person, email, password
    )
    await asyncio.to_thread(
        _update_identity,
        person,
        email,
        graph_user,
        graph_status,
        moodle_user,
        moodle_status,
    )
    password_was_issued = graph_created or moodle_created
    observation = _credential_observation(graph_created, moodle_created)
    return {
        **person,
        "correo_institucional": email,
        "clave_permanente": password if password_was_issued else "",
        "graph_user_id": _clean(graph_user.get("id")) if graph_user else "",
        "estado_graph": graph_status,
        "error_graph": graph_error,
        "estado_licencia": license_status,
        "error_licencia": license_error,
        "licencia_nombre": education_license.name,
        "licencia_sku_part_number": education_license.sku_part_number,
        "moodle_user_id": moodle_user.get("id") if moodle_user else None,
        "moodle_username": _clean(moodle_user.get("username")) if moodle_user else "",
        "estado_moodle": moodle_status,
        "error_moodle": moodle_error,
        "estado_general": _overall_status(graph_status, license_status, moodle_status),
        "clave_emitida": password_was_issued,
        "observacion": observation,
    }


def _full_name(person: dict[str, Any]) -> str:
    return " ".join(
        str(person.get(field) or "")
        for field in ("primer_nombre", "segundo_nombre", "primer_apellido", "segundo_apellido")
        if str(person.get(field) or "")
    )


def _record_audit(batch_id: str, mode: str, rows: list[dict[str, Any]], operator: str) -> None:
    with get_connection() as conn:
        cursor = conn.cursor()
        _ensure_tables(cursor)
        for row in rows:
            password = str(row.get("clave_permanente") or "")
            encrypted_password = (
                _encrypt_credential_password(password)
                if row.get("clave_emitida") and password
                else ""
            )
            cursor.execute(
                """
                INSERT INTO dbo.CREDENCIAL_APROVISIONAMIENTO (
                    lote_id, tipo_persona, modo, fila_origen, cedula, nombres, correo_institucional,
                    estado_graph, error_graph, estado_licencia, error_licencia,
                    estado_moodle, error_moodle, estado_general, clave_emitida,
                    clave_cifrada, observacion, usuario_creacion
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                batch_id,
                str(row.get("tipo_persona") or "ESTUDIANTE"),
                mode,
                row.get("fila_origen"),
                str(row.get("cedula") or ""),
                _full_name(row),
                str(row.get("correo_institucional") or "") or None,
                str(row.get("estado_graph") or "ERROR"),
                str(row.get("error_graph") or "") or None,
                str(row.get("estado_licencia") or "OMITIDA"),
                str(row.get("error_licencia") or "") or None,
                str(row.get("estado_moodle") or "ERROR"),
                str(row.get("error_moodle") or "") or None,
                str(row.get("estado_general") or "ERROR"),
                1 if row.get("clave_emitida") else 0,
                encrypted_password or None,
                str(row.get("observacion") or "")[:500] or None,
                operator,
            )
        conn.commit()


def _excel_safe(value: Any) -> Any:
    if value is None:
        return ""
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _report_bytes(batch_id: str, mode: str, rows: list[dict[str, Any]], operator: str) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Credenciales"
    headers = [
        "Lote", "Tipo de persona", "Tipo de creación", "Fila", "Primer nombre", "Segundo nombre",
        "Primer apellido", "Segundo apellido", "Cédula", "Correo Microsoft 365",
        "Contraseña permanente", "Observación", "Estado Microsoft 365",
        "Licencia Microsoft 365", "SKU licencia", "Estado licencia", "Usuario Moodle",
        "ID Moodle", "Estado Moodle", "Resultado general", "Detalle", "Fecha", "Administrador",
    ]
    worksheet.append(headers)
    generated_at = datetime.now().astimezone()
    for row in rows:
        detail = " | ".join(
            value
            for value in (
                str(row.get("error_graph") or ""),
                str(row.get("error_licencia") or ""),
                str(row.get("error_moodle") or ""),
                "; ".join(str(item) for item in row.get("errores", []) if item),
            )
            if value
        )
        worksheet.append(
            [
                batch_id, _excel_safe(row.get("tipo_persona")), mode,
                row.get("fila_origen") or "", _excel_safe(row.get("primer_nombre")),
                _excel_safe(row.get("segundo_nombre")), _excel_safe(row.get("primer_apellido")),
                _excel_safe(row.get("segundo_apellido")), _excel_safe(row.get("cedula")),
                _excel_safe(row.get("correo_institucional")), _excel_safe(row.get("clave_permanente")),
                _excel_safe(row.get("observacion")), _excel_safe(row.get("estado_graph")),
                _excel_safe(row.get("licencia_nombre")),
                _excel_safe(row.get("licencia_sku_part_number")),
                _excel_safe(row.get("estado_licencia")),
                _excel_safe(row.get("moodle_username") or row.get("correo_institucional")),
                row.get("moodle_user_id") or "", _excel_safe(row.get("estado_moodle")),
                _excel_safe(row.get("estado_general")), _excel_safe(detail),
                generated_at.strftime("%Y-%m-%d %H:%M:%S"), _excel_safe(operator),
            ]
        )

    header_fill = PatternFill("solid", fgColor="A91F1B")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = [38, 18, 18, 10, 22, 22, 24, 24, 16, 38, 28, 58, 24, 34, 32, 28, 38, 14, 24, 28, 60, 22, 24]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(1, index).column_letter].width = width
    for cell in worksheet["I"][1:]:
        cell.number_format = "@"
    worksheet.sheet_view.showGridLines = False

    instructions = workbook.create_sheet("Información")
    instructions.append(["Reporte de creación de credenciales"])
    if rows:
        instructions.append(
            [
                f"Perfil procesado: {_excel_safe(rows[0].get('tipo_persona'))}. "
                f"Licencia: {_excel_safe(rows[0].get('licencia_nombre'))}."
            ]
        )
    instructions.append([f"{_PERMANENT_PASSWORD_NOTE} Debe manejarse de forma confidencial."])
    instructions.append(["Las cuentas existentes no se restablecen y por eso no muestran contraseña."])
    instructions.append(["Puede volver a descargar una credencial archivada desde el historial."])
    instructions.column_dimensions["A"].width = 110
    instructions["A1"].font = Font(bold=True, color="A91F1B")

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _archived_credential_report_bytes(
    record: dict[str, Any],
    password: str,
    downloaded_by: str,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Credencial"
    worksheet.append(
        [
            "Tipo de persona",
            "Nombre",
            "Cédula",
            "Correo institucional",
            "Contraseña permanente",
            "Observación",
            "Fecha de creación",
            "Descargado por",
        ]
    )
    created_at = record.get("fecha_creacion")
    worksheet.append(
        [
            _excel_safe(record.get("tipo_persona")),
            _excel_safe(record.get("nombres")),
            _excel_safe(record.get("cedula")),
            _excel_safe(record.get("correo_institucional")),
            _excel_safe(password),
            _excel_safe(record.get("observacion")),
            created_at.strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(created_at, datetime)
            else _excel_safe(created_at),
            _excel_safe(downloaded_by),
        ]
    )
    header_fill = PatternFill("solid", fgColor="A91F1B")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column, width in zip(
        ("A", "B", "C", "D", "E", "F", "G", "H"),
        (18, 42, 16, 40, 30, 66, 24, 28),
        strict=True,
    ):
        worksheet.column_dimensions[column].width = width
    worksheet["C2"].number_format = "@"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:H2"
    worksheet.sheet_view.showGridLines = False

    information = workbook.create_sheet("Información")
    information.append(["Documento histórico de credencial institucional"])
    information.append([_PERMANENT_PASSWORD_NOTE])
    information.append(["Este documento contiene información confidencial y debe custodiarse."])
    information.append(["La descarga quedó registrada en el historial institucional."])
    information.column_dimensions["A"].width = 105
    information["A1"].font = Font(bold=True, color="A91F1B")

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _store_report(owner: str, filename: str, content: bytes) -> str:
    report_id = str(uuid4())
    now = datetime.now(timezone.utc)
    with _REPORTS_LOCK:
        expired = [key for key, item in _REPORTS.items() if item.expires_at <= now]
        for key in expired:
            _REPORTS.pop(key, None)
        _REPORTS[report_id] = _CachedReport(
            owner=owner,
            filename=filename,
            content=content,
            expires_at=now + _REPORT_TTL,
        )
    return report_id


def _take_report(report_id: str, owner: str) -> _CachedReport:
    now = datetime.now(timezone.utc)
    with _REPORTS_LOCK:
        item = _REPORTS.get(report_id)
        if not item or item.expires_at <= now:
            _REPORTS.pop(report_id, None)
            raise HTTPException(status_code=404, detail="El reporte expiró o ya fue descargado")
        if item.owner.casefold() != owner.casefold():
            raise HTTPException(status_code=403, detail="El reporte pertenece a otro usuario")
        return _REPORTS.pop(report_id)


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows[:10]):
        normalized = {_slug(value) for value in row if _clean(value)}
        recognized = sum(
            1 for aliases in _HEADER_ALIASES.values() if any(value in aliases for value in normalized)
        )
        if recognized >= 3:
            return index
    return 0


def _read_workbook(content: bytes, filename: str) -> list[CredentialPersonPayload]:
    if not filename.casefold().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Suba un archivo Excel con extensión .xlsx")
    if len(content) > _MAX_FILE_BYTES:
        raise HTTPException(status_code=400, detail="El archivo supera el máximo permitido de 8 MB")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel") from exc
    try:
        worksheet = workbook["Credenciales"] if "Credenciales" in workbook.sheetnames else workbook.active
        rows = list(worksheet.iter_rows(values_only=True, max_row=_MAX_USERS + 12))
    finally:
        workbook.close()
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo Excel está vacío")

    header_index = _find_header_row(rows)
    normalized_headers = [_slug(value) for value in rows[header_index]]
    columns: dict[str, int] = {}
    for index, header in enumerate(normalized_headers):
        for target, aliases in _HEADER_ALIASES.items():
            if target not in columns and header in aliases:
                columns[target] = index
    missing = [header for header in _TEMPLATE_HEADERS if header not in columns]
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan columnas obligatorias: {', '.join(missing)}")

    parsed: list[CredentialPersonPayload] = []
    for row_number, values in enumerate(rows[header_index + 1 :], start=header_index + 2):
        person_values = {
            field: values[index] if index < len(values) else "" for field, index in columns.items()
        }
        if not any(_clean(value) for value in person_values.values()):
            continue
        parsed.append(
            CredentialPersonPayload.model_construct(
                primer_nombre=_clean(person_values["primer_nombre"]),
                segundo_nombre=_clean(person_values["segundo_nombre"]),
                primer_apellido=_clean(person_values["primer_apellido"]),
                segundo_apellido=_clean(person_values["segundo_apellido"]),
                cedula=_normalize_cedula(person_values["cedula"]),
                fila_origen=row_number,
            )
        )
        if len(parsed) > _MAX_USERS:
            raise HTTPException(status_code=400, detail=f"El archivo supera {_MAX_USERS} usuarios")
    if not parsed:
        raise HTTPException(status_code=400, detail="No existen usuarios para analizar")
    return parsed


def _template_bytes(person_type: Literal["ESTUDIANTE", "PROFESOR"] = "ESTUDIANTE") -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Credenciales"
    worksheet.append(_TEMPLATE_HEADERS)
    worksheet.append(["María José", "Alejandra", "De la Cruz", "De la Torre", "0123456789"])
    header_fill = PatternFill("solid", fgColor="A91F1B")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = "A1:E2"
    for column, width in zip(("A", "B", "C", "D", "E"), (28, 28, 30, 30, 18), strict=True):
        worksheet.column_dimensions[column].width = width
    for cell in worksheet["E"][1:]:
        cell.number_format = "@"

    instructions = workbook.create_sheet("Instrucciones")
    audience = "profesores" if person_type == "PROFESOR" else "estudiantes"
    license_name = (
        "Office 365 A1 para profesores"
        if person_type == "PROFESOR"
        else "Office 365 A1 para estudiantes"
    )
    instructions.append([f"Creación de credenciales para {audience}: Microsoft 365 y Moodle"])
    instructions.append([f"Licencia asignada: {license_name}."])
    instructions.append(["No cambie los nombres de las cinco columnas de la plantilla."])
    instructions.append(["Los nombres y apellidos compuestos deben permanecer completos en su columna."])
    instructions.append(["La cédula debe contener 10 dígitos y conservarse como texto."])
    instructions.append([f"Máximo permitido por archivo: {_MAX_USERS} usuarios."])
    instructions.column_dimensions["A"].width = 105
    instructions["A1"].font = Font(bold=True, color="A91F1B")

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@router.get("/config")
def credential_config(response: Response, current_user: SessionUser = AdminOnly) -> dict[str, Any]:
    del current_user
    response.headers["Cache-Control"] = "no-store"
    settings = get_settings()
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_tables(cursor)
            cursor.execute("SELECT COUNT_BIG(1) FROM dbo.CREDENCIAL_IDENTIDAD")
            identity_count = int(cursor.fetchone()[0])
            conn.commit()
    except Exception as exc:
        raise HTTPException(status_code=500, detail="No se pudo inicializar el módulo de credenciales") from exc

    licenses = {
        person_type: _license_configuration(person_type)
        for person_type in ("ESTUDIANTE", "PROFESOR")
    }
    student_license = licenses["ESTUDIANTE"]

    return {
        "domain": _graph_domain(),
        "year": datetime.now().year,
        "graph_configured": _graph_is_configured(),
        "license_configured": bool(student_license["configured"]),
        "license_type": "ESTUDIANTE",
        "license_name": student_license["name"],
        "license_sku_part_number": student_license["sku_part_number"],
        "license_status": student_license["status"],
        "license_available_units": student_license["available_units"],
        "license_detail": student_license["detail"],
        "licenses": licenses,
        "moodle_configured": _moodle_is_configured(),
        "moodle_url": _clean(settings.moodle_base_url),
        "identity_count": identity_count,
        "max_users": _MAX_USERS,
        "report_ttl_minutes": int(_REPORT_TTL.total_seconds() // 60),
    }


@router.get("/template")
def download_template(
    tipo_persona: Literal["ESTUDIANTE", "PROFESOR"] = Query(default="ESTUDIANTE"),
    current_user: SessionUser = AdminOnly,
) -> StreamingResponse:
    del current_user
    suffix = "profesores" if tipo_persona == "PROFESOR" else "estudiantes"
    return StreamingResponse(
        BytesIO(_template_bytes(tipo_persona)),
        media_type=_EXCEL_MEDIA_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="plantilla_credenciales_{suffix}_microsoft_moodle.xlsx"'
            ),
            "Cache-Control": "no-store, private",
        },
    )


@router.post("/analyze")
async def analyze_credentials(
    response: Response,
    file: UploadFile = File(...),
    current_user: SessionUser = AdminOnly,
) -> dict[str, Any]:
    del current_user
    response.headers["Cache-Control"] = "no-store"
    content = await file.read(_MAX_FILE_BYTES + 1)
    people = _read_workbook(content, file.filename or "")
    normalized_people = [_normalized_person(person) for person in people]
    cedula_counts = Counter(str(person["cedula"]) for person in normalized_people if person["cedula"])
    used_preview: set[str] = set()
    rows: list[dict[str, Any]] = []
    valid_count = 0
    for person in normalized_people:
        errors = _person_errors(person)
        cedula = str(person["cedula"])
        if cedula and cedula_counts[cedula] > 1:
            errors.append("La cédula está repetida en el archivo")
        proposed_email = ""
        if not errors:
            existing = await asyncio.to_thread(_existing_email_for_cedula, cedula)
            if existing:
                proposed_email = existing
            else:
                domain = _graph_domain()
                for local_part in _email_candidates(person):
                    candidate = f"{local_part}@{domain}"
                    if candidate not in used_preview:
                        proposed_email = candidate
                        break
            used_preview.add(proposed_email)
            valid_count += 1
        rows.append(
            {
                **person,
                "correo_propuesto": proposed_email,
                "estado": "VALIDO" if not errors else "ERROR",
                "errores": errors,
            }
        )
    return {
        "rows": rows,
        "summary": {"total": len(rows), "validos": valid_count, "errores": len(rows) - valid_count},
        "filename": file.filename or "credenciales.xlsx",
    }


@router.post("/provision")
async def provision_credentials(
    payload: CredentialProvisionPayload,
    response: Response,
    current_user: SessionUser = AdminOnly,
) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store"
    operator = _clean(current_user.login)
    batch_id = str(uuid4())
    moodle = MoodleClient(get_settings())

    if not _graph_is_configured():
        audience = "profesores" if payload.tipo_persona == "PROFESOR" else "estudiantes"
        raise HTTPException(
            status_code=409,
            detail=f"Microsoft 365 no está configurado para asignar la licencia de {audience}",
        )
    try:
        education_license = await asyncio.to_thread(_education_license, payload.tipo_persona)
    except Exception as exc:
        audience = "profesores" if payload.tipo_persona == "PROFESOR" else "estudiantes"
        raise HTTPException(
            status_code=409,
            detail=f"No se pudo validar la licencia de {audience}: {_graph_error_detail(exc)}",
        ) from exc
    requested_licenses = len(payload.usuarios)
    if education_license.available_units < requested_licenses:
        raise HTTPException(
            status_code=409,
            detail=(
                f"No existen suficientes licencias {education_license.name}: "
                f"se requieren {requested_licenses} y hay {education_license.available_units} disponibles"
            ),
        )

    seen_cedulas: set[str] = set()
    results: list[dict[str, Any]] = []

    for index, model in enumerate(payload.usuarios, start=1):
        person = _normalized_person(model)
        person["tipo_persona"] = payload.tipo_persona
        person["fila_origen"] = person.get("fila_origen") or index
        errors = _person_errors(person)
        cedula = str(person["cedula"])
        if cedula in seen_cedulas:
            errors.append("La cédula está repetida en la solicitud")
        seen_cedulas.add(cedula)
        if errors:
            results.append(
                {
                    **person, "correo_institucional": "", "clave_permanente": "",
                    "estado_graph": "NO_PROCESADO", "error_graph": "",
                    "estado_licencia": "OMITIDA", "error_licencia": "",
                    "licencia_nombre": education_license.name,
                    "licencia_sku_part_number": education_license.sku_part_number,
                    "estado_moodle": "NO_PROCESADO", "error_moodle": "",
                    "estado_general": "ERROR_VALIDACION", "errores": errors, "clave_emitida": False,
                    "observacion": "No se generó una contraseña porque la fila no superó la validación.",
                }
            )
            continue
        try:
            results.append(await _provision_person(person, operator, moodle, education_license))
        except Exception as exc:
            results.append(
                {
                    **person, "correo_institucional": "", "clave_permanente": "",
                    "estado_graph": "NO_PROCESADO", "error_graph": "",
                    "estado_licencia": "OMITIDA", "error_licencia": "",
                    "licencia_nombre": education_license.name,
                    "licencia_sku_part_number": education_license.sku_part_number,
                    "estado_moodle": "NO_PROCESADO", "error_moodle": "",
                    "estado_general": "ERROR", "errores": [_clean(exc)[:500] or "No se pudo procesar la identidad"],
                    "clave_emitida": False,
                    "observacion": "No se generó una contraseña debido al error de aprovisionamiento.",
                }
            )

    await asyncio.to_thread(_record_audit, batch_id, payload.modo, results, operator)
    report_content = await asyncio.to_thread(_report_bytes, batch_id, payload.modo, results, operator)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_id = _store_report(
        operator,
        (
            f"reporte_credenciales_{payload.tipo_persona.casefold()}_"
            f"{payload.modo.casefold()}_{timestamp}.xlsx"
        ),
        report_content,
    )
    completed = sum(
        1 for row in results if str(row.get("estado_general")) in {"COMPLETO", "COMPLETO_CON_ADVERTENCIA"}
    )
    partial = sum(1 for row in results if row.get("estado_general") == "PARCIAL")
    failed = len(results) - completed - partial
    return {
        "ok": failed == 0 and partial == 0,
        "batch_id": batch_id,
        "tipo_persona": payload.tipo_persona,
        "rows": results,
        "summary": {"total": len(results), "completos": completed, "parciales": partial, "fallidos": failed},
        "report_id": report_id,
        "report_expires_minutes": int(_REPORT_TTL.total_seconds() // 60),
        "message": (
            f"Se procesaron {len(results)} identidad(es). "
            "Descargue el documento con los correos y contraseñas permanentes."
        ),
    }


@router.get("/reports/{report_id}")
def download_report(report_id: str, current_user: SessionUser = AdminOnly) -> StreamingResponse:
    report = _take_report(report_id, _clean(current_user.login))
    return StreamingResponse(
        BytesIO(report.content),
        media_type=_EXCEL_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{report.filename}"',
            "Cache-Control": "no-store, private",
        },
    )


@router.get("/history")
def credential_history(
    response: Response,
    limit: int = Query(default=100, ge=1, le=500),
    tipo_persona: Literal["ESTUDIANTE", "PROFESOR"] | None = Query(default=None),
    current_user: SessionUser = AdminOnly,
) -> dict[str, Any]:
    del current_user
    response.headers["Cache-Control"] = "no-store"
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_tables(cursor)
            type_filter = (
                "WHERE COALESCE(tipo_persona, 'ESTUDIANTE') = ?"
                if tipo_persona
                else ""
            )
            query = f"""
                SELECT TOP ({limit}) id, lote_id,
                       COALESCE(tipo_persona, 'ESTUDIANTE') AS tipo_persona,
                       modo, fila_origen, cedula, nombres,
                       correo_institucional, estado_graph, estado_licencia, estado_moodle,
                       estado_general, clave_emitida, clave_cifrada, observacion,
                       numero_descargas, fecha_ultima_descarga, usuario_ultima_descarga,
                       usuario_creacion, fecha_creacion
                FROM dbo.CREDENCIAL_APROVISIONAMIENTO
                {type_filter}
                ORDER BY fecha_creacion DESC, id DESC
            """
            if tipo_persona:
                cursor.execute(query, tipo_persona)
            else:
                cursor.execute(query)
            rows = []
            for row in cursor.fetchall():
                report_available = bool(_clean(row.clave_cifrada))
                observation = _clean(row.observacion)
                if not observation:
                    observation = (
                        "Registro anterior sin contraseña recuperable."
                        if row.clave_emitida
                        else "No se emitió una contraseña para esta cuenta."
                    )
                rows.append(
                    {
                        "id": int(row.id), "batch_id": _clean(row.lote_id),
                        "tipo_persona": _clean(row.tipo_persona), "modo": _clean(row.modo),
                        "fila_origen": int(row.fila_origen) if row.fila_origen is not None else None,
                        "cedula": _clean(row.cedula), "nombres": _clean(row.nombres),
                        "correo_institucional": _clean(row.correo_institucional),
                        "estado_graph": _clean(row.estado_graph),
                        "estado_licencia": _clean(row.estado_licencia),
                        "estado_moodle": _clean(row.estado_moodle),
                        "estado_general": _clean(row.estado_general),
                        "clave_emitida": bool(row.clave_emitida),
                        "reporte_disponible": report_available,
                        "observacion": observation,
                        "numero_descargas": int(row.numero_descargas or 0),
                        "fecha_ultima_descarga": (
                            row.fecha_ultima_descarga.isoformat()
                            if row.fecha_ultima_descarga
                            else None
                        ),
                        "usuario_ultima_descarga": _clean(row.usuario_ultima_descarga),
                        "usuario_creacion": _clean(row.usuario_creacion),
                        "fecha_creacion": row.fecha_creacion.isoformat() if row.fecha_creacion else None,
                    }
                )
            conn.commit()
    except Exception as exc:
        raise HTTPException(status_code=500, detail="No se pudo consultar el historial de credenciales") from exc
    return {"rows": rows, "count": len(rows)}


@router.get("/history/{record_id}/report")
def download_archived_credential_report(
    record_id: int,
    current_user: SessionUser = AdminOnly,
) -> StreamingResponse:
    operator = _clean(current_user.login)
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            _ensure_tables(cursor)
            cursor.execute(
                """
                SELECT TOP (1) id, lote_id,
                       COALESCE(tipo_persona, 'ESTUDIANTE') AS tipo_persona,
                       modo, cedula, nombres, correo_institucional, clave_cifrada,
                       observacion, fecha_creacion
                FROM dbo.CREDENCIAL_APROVISIONAMIENTO
                WHERE id = ?
                """,
                record_id,
            )
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="La credencial histórica no existe")
            encrypted_password = _clean(row.clave_cifrada)
            if not encrypted_password:
                raise HTTPException(
                    status_code=409,
                    detail="Este registro no conserva una contraseña descargable",
                )
            try:
                password = _decrypt_credential_password(encrypted_password)
            except RuntimeError as exc:
                raise HTTPException(status_code=409, detail=str(exc)) from exc

            observation = _clean(row.observacion) or _PERMANENT_PASSWORD_NOTE
            record = {
                "tipo_persona": _clean(row.tipo_persona),
                "nombres": _clean(row.nombres),
                "cedula": _clean(row.cedula),
                "correo_institucional": _clean(row.correo_institucional),
                "observacion": observation,
                "fecha_creacion": row.fecha_creacion,
            }
            report_content = _archived_credential_report_bytes(record, password, operator)
            cursor.execute(
                """
                UPDATE dbo.CREDENCIAL_APROVISIONAMIENTO
                   SET numero_descargas = COALESCE(numero_descargas, 0) + 1,
                       fecha_ultima_descarga = SYSDATETIME(),
                       usuario_ultima_descarga = ?
                 WHERE id = ?
                """,
                operator,
                record_id,
            )
            conn.commit()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail="No se pudo generar el documento histórico de la credencial",
        ) from exc

    person_type = str(record["tipo_persona"]).casefold()
    document = re.sub(r"[^0-9A-Za-z_-]+", "_", str(record["cedula"])) or str(record_id)
    filename = f"credencial_{person_type}_{document}.xlsx"
    return StreamingResponse(
        BytesIO(report_content),
        media_type=_EXCEL_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store, private",
        },
    )
