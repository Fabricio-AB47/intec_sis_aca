from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.core.security import SessionUser, require_roles
from app.services.db import get_connection

router = APIRouter(prefix="/api/students/reporteria-integral", tags=["reporteria-integral"])

AllowedUser = Depends(require_roles("ADMINISTRADOR", "ACADEMICO", "RECTOR"))


REPORTS: dict[str, dict[str, Any]] = {
    "matriculados": {
        "title": "Estudiantes matriculados",
        "description": "Equivalente agil de los reportes de matricula por periodo, carrera, estado y paralelo.",
        "category": "Academico",
        "source_tables": ["CARRERAXESTUD", "DATOS_ESTUD", "CARRERAS", "PERIODO", "PENSUM", "ESTADO"],
        "filters": ["periodo", "carrera", "estado", "buscar", "limite"],
        "estado_options": [
            {"value": "A", "label": "Activo"},
            {"value": "G", "label": "Graduado"},
            {"value": "P", "label": "Inactivo"},
            {"value": "R", "label": "Retirado"},
        ],
    },
    "becas_edades": {
        "title": "Becas y edades",
        "description": "Listado de estudiantes con beca, porcentaje, fecha de nacimiento, edad calculada y rango etario.",
        "category": "Estudiantes",
        "source_tables": ["DATOS_ESTUD", "Becas", "CARRERAXESTUD", "CARRERAS", "PERIODO", "ESTADO"],
        "filters": ["periodo", "carrera", "estado", "buscar", "limite"],
        "estado_options": [
            {"value": "A", "label": "Activo"},
            {"value": "G", "label": "Graduado"},
            {"value": "P", "label": "Inactivo"},
            {"value": "R", "label": "Retirado"},
        ],
    },
    "preinscritos": {
        "title": "Preinscritos y ventas",
        "description": "Cartera de aspirantes con asesor, medio de contacto, periodo, carrera y avance del ingreso.",
        "category": "Admision",
        "source_tables": ["PREINSCRIPCION", "PERIODO", "CARRERAS", "Provincias", "IN_LECONTACTO", "USUARIO_SIS"],
        "filters": ["periodo", "carrera", "estado", "buscar", "limite"],
        "estado_options": [
            {"value": "PREMATRICULA", "label": "Prematricula"},
            {"value": "FINALIZADO", "label": "Proceso finalizado"},
            {"value": "PENDIENTE", "label": "Pendiente"},
        ],
    },
    "docentes": {
        "title": "Docentes asignados",
        "description": "Docentes por periodo, carrera, materia, jornada y paralelo para reemplazar consultas Crystal.",
        "category": "Docencia",
        "source_tables": ["CARRERAXDOCENTE", "DATOSDOCENTE", "CARRERAS", "PENSUM", "PERIODO", "JORNADA"],
        "filters": ["periodo", "carrera", "estado", "buscar", "limite"],
        "estado_options": [
            {"value": "1", "label": "Sincronizado Moodle"},
            {"value": "0", "label": "Pendiente Moodle"},
        ],
    },
    "documentos": {
        "title": "Documentos de matricula",
        "description": "Control de cedula, titulo, deposito y convenio registrados en cabecera de matricula.",
        "category": "Documentacion",
        "source_tables": ["CABECERA_MATRICULA", "DATOS_ESTUD", "CARRERAS", "PERIODO"],
        "filters": ["periodo", "carrera", "estado", "buscar", "limite"],
        "estado_options": [
            {"value": "PENDIENTE", "label": "Pendiente"},
            {"value": "COMPLETO", "label": "Completo"},
        ],
    },
    "seguimiento": {
        "title": "Seguimiento estudiante",
        "description": "Bitacora activa de seguimiento por estudiante, periodo y materia.",
        "category": "Acompanamiento",
        "source_tables": ["SEGUIMIENTO_ESTUDIANTE", "DATOS_ESTUD", "PERIODO", "PENSUM", "CARRERAS"],
        "filters": ["periodo", "carrera", "buscar", "limite"],
        "estado_options": [],
    },
    "practicas": {
        "title": "Prácticas laborales",
        "description": "Practicas por estudiante, carrera, periodo, empresa, docente tutor y horas registradas.",
        "category": "Vinculacion",
        "source_tables": ["PRACTICASPROFESIONALES", "DATOS_ESTUD", "CARRERAS", "PERIODO", "EMPRESA", "DATOSDOCENTE"],
        "filters": ["periodo", "carrera", "buscar", "limite"],
        "estado_options": [],
    },
    "evaluacion_docente": {
        "title": "Evaluacion docente",
        "description": "Resultado consolidado de evaluaciones por docente, periodo, materia, jornada y paralelo.",
        "category": "Docencia",
        "source_tables": ["RESULTADO_EVALUACION", "DATOSDOCENTE", "CARRERAXDOCENTE", "PENSUM", "PERIODO", "CARRERAS"],
        "filters": ["periodo", "carrera", "buscar", "limite"],
        "estado_options": [],
    },
    "moodle_notas": {
        "title": "Notas Moodle sincronizadas",
        "description": "Notas importadas desde Moodle con componente, nota obtenida, porcentaje y estado.",
        "category": "Moodle",
        "source_tables": ["intec_estudiantenota", "DATOS_ESTUD", "PENSUM"],
        "filters": ["periodo", "buscar", "limite"],
        "estado_options": [
            {"value": "calificado", "label": "Calificado"},
            {"value": "pendiente", "label": "Pendiente"},
            {"value": "error", "label": "Error"},
        ],
    },
    "notas_carrera_materia": {
        "title": "Notas por carrera y periodo",
        "description": "Estudiantes activos por periodo, con materias y calificaciones regular u homologacion.",
        "category": "Academico",
        "source_tables": ["CARRERAXESTUD", "DATOS_ESTUD", "CARRERAS", "PENSUM", "PERIODO"],
        "filters": ["periodo", "carrera", "limite"],
        "estado_options": [],
    },
    "estud_per_c_m": {
        "title": "Estudiantes por periodo, carrera y materia",
        "description": "Estudiantes matriculados por periodo y carrera, materias tomadas y estado aprobado/reprobado segun promedio final.",
        "category": "Academico",
        "source_tables": ["CARRERAXESTUD", "DATOS_ESTUD", "CARRERAS", "PENSUM", "PERIODO"],
        "filters": ["periodo", "carrera", "estado", "buscar", "limite"],
        "estado_options": [
            {"value": "APROBADA", "label": "Aprobadas"},
            {"value": "REPROBADA", "label": "Reprobadas"},
            {"value": "PENDIENTE", "label": "Pendientes"},
        ],
    },
    "correos_intec": {
        "title": "Correos institucionales",
        "description": "Correos INTEC asignados a estudiantes, envio, estado y ultimo acceso Moodle.",
        "category": "Cuentas",
        "source_tables": ["CorreosEstudIntec", "DATOS_ESTUD", "PERIODO"],
        "filters": ["periodo", "estado", "buscar", "limite"],
        "estado_options": [
            {"value": "ACTIVO", "label": "Activo"},
            {"value": "ERROR", "label": "Error"},
            {"value": "PENDIENTE", "label": "Pendiente"},
        ],
    },
    "microsoft_audit": {
        "title": "Auditoria Microsoft 365",
        "description": "Auditoria de acciones sobre cuentas Microsoft 365 y licencias asignadas.",
        "category": "Cuentas",
        "source_tables": ["Microsoft365Audit"],
        "filters": ["estado", "buscar", "limite"],
        "estado_options": [
            {"value": "OK", "label": "Correcto"},
            {"value": "ERROR", "label": "Error"},
        ],
    },
    "pagos_matricula": {
        "title": "Pagos de matricula",
        "description": "Pagos registrados por estudiante, periodo, carrera, deposito y valor.",
        "category": "Financiero",
        "source_tables": ["REGISTROPAGOS", "DATOS_ESTUD", "PERIODO", "CARRERAS"],
        "filters": ["periodo", "carrera", "buscar", "limite"],
        "estado_options": [],
    },
}

FUNCTIONAL_INVENTORY = [
    {
        "module": "Academico y matricula",
        "legacy_sources": ["CARRERAXESTUD", "CABECERA_MATRICULA", "PENSUM", "CARRERAS", "PERIODO"],
        "capabilities": [
            "matriculas por carrera, periodo, paralelo y tipo",
            "promedios y notas finales",
            "mallas y materias por nivel",
        ],
    },
    {
        "module": "Admision y ventas",
        "legacy_sources": ["PREINSCRIPCION", "USUARIO_SIS", "IN_LECONTACTO", "Provincias"],
        "capabilities": [
            "aspirantes por asesor",
            "seguimiento de ingreso",
            "documentos de preinscripcion",
        ],
    },
    {
        "module": "Docencia",
        "legacy_sources": ["DATOSDOCENTE", "CARRERAXDOCENTE", "RESULTADO_EVALUACION", "CUESTIONARIOEVALUA"],
        "capabilities": [
            "asignacion docente",
            "evaluacion docente",
            "estado de sincronizacion Moodle",
        ],
    },
    {
        "module": "Documentacion y certificacion",
        "legacy_sources": ["CABECERA_MATRICULA", "PRACTICASPROFESIONALES", "EMPRESA", "Microsoft365Audit"],
        "capabilities": [
            "control documental",
            "prácticas laborales",
            "auditoria de correos institucionales",
        ],
    },
    {
        "module": "Integraciones modernas",
        "legacy_sources": ["intec_moodleconfig", "intec_logmatriculacion", "intec_estudiantenota", "auth_user"],
        "capabilities": [
            "sincronizacion Moodle",
            "logs operativos",
            "base Django/API para reemplazar pantallas WebForms",
        ],
    },
]

CRYSTAL_REPORTS: list[dict[str, Any]] = [
    {
        "key": "crystal_academico_estudiante_regular",
        "title": "Reporte académico por estudiante regular",
        "category": "Académico",
        "legacy_rpt": ["CryAcadxEstud.rpt", "ReporteAcad/CryAcadxEstud.rpt"],
        "legacy_pages": ["ReporteAcadxEstudDirecto.aspx", "ReporteAcad/ConsultaNotasEstud.aspx", "ReporteAcad/ReporteAcadxEstud.aspx"],
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PENSUM", "CARRERAS", "PERIODO"],
        "legacy_filters": ["codigo_estud", "PENSUM.verreporte=1", "PERIODO.estado='A'", "TipoMatricula='N'"],
        "modern_equivalent": "portal_academico.student_record_pdf_export / titulacion malla académica",
        "modern_format": ["PDF", "Excel"],
        "migration_status": "modernizado",
        "notes": "Usar reglas modernas de notas: R con P1/P2/P3/final; H con teórico/práctico/final.",
    },
    {
        "key": "crystal_academico_estudiante_homologacion",
        "title": "Reporte académico por estudiante homologación",
        "category": "Académico",
        "legacy_rpt": ["CryAcadxEstudHomo.rpt"],
        "legacy_pages": ["ReporteAcadxEstudDirecto.aspx", "ReporteAcad/ReporteAcadxEstud.aspx"],
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PENSUM", "CARRERAS", "PERIODO"],
        "legacy_filters": ["codigo_estud", "TipoMatricula<>'N'"],
        "modern_equivalent": "portal_academico.student_record_pdf_export",
        "modern_format": ["PDF", "Excel"],
        "migration_status": "modernizado",
        "notes": "Debe conservar detalle de teórico y práctico para homologación.",
    },
    {
        "key": "crystal_academico_estudiante_general",
        "title": "Reporte académico general / pase",
        "category": "Académico",
        "legacy_rpt": ["ReporteAcad/CryAcadxEstudGeneral.rpt", "ReporteAcad/CryAcadxEstudGeneralPase.rpt"],
        "legacy_pages": ["ReporteAcad/ConsultaNotasEstud.aspx", "ReporteAcad/ReporteAcadxEstud.aspx"],
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PENSUM", "CARRERAS", "PERIODO"],
        "legacy_filters": ["codigo_estud", "PENSUM.verreporte=1", "PERIODO.estado='A'"],
        "modern_equivalent": "portal_academico.student_record_pdf_export",
        "modern_format": ["PDF"],
        "migration_status": "base",
        "notes": "Pendiente separar visualmente formato general y formato pase si Secretaría los requiere distintos.",
    },
    {
        "key": "crystal_matriculados_periodo",
        "title": "Total de estudiantes por período/carrera",
        "category": "Reportería R/H",
        "legacy_rpt": [
            "ReporteAcad/CryListaTotalEstudPeriodo.rpt",
            "ReporteAcad/CryListaTotalEstudPeriodoCarrera.rpt",
            "ReporteAcad/CryListaTotalEstudPeriodoResumen.rpt",
        ],
        "legacy_pages": ["ReporteAcad/ListaTotalEstudPeriodo.aspx", "ReporteAcad/ListaTotalEstudPeriodoDeralle.aspx"],
        "source_tables": ["TOTALESTUDPERIODO", "TOTALESTUDMATRICULADOSPER", "TOTALESTUDCARRERAPERIODO"],
        "legacy_filters": ["anio", "cod_periodo"],
        "modern_equivalent": "reporteria-integral periodo / carrera",
        "modern_format": ["Excel"],
        "migration_status": "modernizado",
        "notes": "Ya se genera bajo reportería integral con separación Regular/Homologación.",
    },
    {
        "key": "crystal_matriculados_genero_estado_provincia",
        "title": "Resumen por género, estado, carrera y provincia",
        "category": "Reportería R/H",
        "legacy_rpt": [
            "ReporteAcad/CryListaTotalEstudAnioGenero.rpt",
            "ReporteAcad/CryListaTotalEstudAnioEstado.rpt",
            "ReporteAcad/CryListaTotalEstudAnioCarrera.rpt",
            "ReporteAcad/CryListaTotalEstudAnioProvincia.rpt",
            "ReporteAcad/CryResumenGeneroEstado.rpt",
            "ReporteAcad/CryResumenEstudCarrera.rpt",
            "ReporteAcad/CryResumenEstudGraduado.rpt",
        ],
        "legacy_pages": ["ReporteAcad/ListaEstudGeneroOtros.aspx", "ReporteAcad/ListaTotalEstudCNE.aspx"],
        "source_tables": ["ConsGenero", "ConsGeneroprovestado", "RESUMENCONSGENEROESTADO", "RESUMENNoESTUDCARRERA", "ResumenGradoAnio"],
        "legacy_filters": ["anio", "genero", "estado", "carrera"],
        "modern_equivalent": "reporteria-integral provincia, provincia_genero, provincia_carrera, genero, graduados_2025",
        "modern_format": ["Excel"],
        "migration_status": "modernizado",
        "notes": "Mantener los totales R/H y permitir exportación Excel para sustituir Crystal.",
    },
    {
        "key": "crystal_docentes_periodo",
        "title": "Profesores por período y materia",
        "category": "Docencia",
        "legacy_rpt": ["ReporteAcad/CrysListaProfesorMateria.rpt"],
        "legacy_pages": ["ReporteAcad/ListaProfesores.aspx"],
        "source_tables": ["DATOSDOCENTE", "CARRERAXDOCENTE", "PENSUM", "CARRERAS", "PERIODO"],
        "legacy_filters": ["codigo_periodo"],
        "modern_equivalent": "reporteria-integral docentes / portal docente",
        "modern_format": ["Excel"],
        "migration_status": "base",
        "notes": "Debe incluir paralelo, jornada, carrera, materia y docente asignado.",
    },
    {
        "key": "crystal_egresados",
        "title": "Lista de egresados",
        "category": "Titulación",
        "legacy_rpt": ["ReporteAcad/CryListaEgresados.rpt"],
        "legacy_pages": ["ReporteAcad/ListaEgresados.aspx"],
        "source_tables": ["ListaEgresadosCarrera"],
        "legacy_filters": ["codPeriodoMax", "cod_anio_Basica"],
        "modern_equivalent": "titulacion verificación y estudiantes aptos",
        "modern_format": ["Excel", "PDF"],
        "migration_status": "base",
        "notes": "Debe tomar malla completa, inglés A2+ - INTERMEDIATE, prácticas y vinculación con la sociedad.",
    },
    {
        "key": "crystal_practicas_profesionales",
        "title": "Prácticas preprofesionales",
        "category": "Prácticas",
        "legacy_rpt": ["ReporteAcad/CrysReporteParcticasProfesional.rpt", "ReporteAcad/CrysReporteParcticasProfesionalNivel.rpt"],
        "legacy_pages": [
            "ReporteAcad/RepPracticasProfesionales.aspx",
            "ReporteAcad/RepPracticasProfesionalesPorEstud.aspx",
            "ReporteAcad/RepPracticasVinculacion.aspx",
        ],
        "source_tables": ["PRACTICASPROFESIONALES", "DATOS_ESTUD", "CARRERAS", "PERIODO", "EMPRESA", "DATOSDOCENTE"],
        "legacy_filters": ["codigo_periodo", "codigo_estud"],
        "modern_equivalent": "practicas_institucionales / reporteria-integral practicas",
        "modern_format": ["Excel", "PDF"],
        "migration_status": "base",
        "notes": "Separar prácticas preprofesionales de vinculación con la sociedad; vinculación requiere 60 horas.",
    },
    {
        "key": "crystal_convenio_pagos",
        "title": "Convenio de pagos",
        "category": "Financiero",
        "legacy_rpt": ["RepFinanciero/CryConvenioPagos.rpt"],
        "legacy_pages": ["RepFinanciero/RepConvenioPagos.aspx", "RepFinanciero/RepConvenioPagosAdmin.aspx"],
        "source_tables": ["REGISTROPAGOS", "DATOS_ESTUD", "CABECERA_MATRICULA", "CARRERAS", "PERIODO"],
        "legacy_filters": ["Codestu"],
        "modern_equivalent": "preinscription carta/convenio de pago",
        "modern_format": ["PDF"],
        "migration_status": "modernizado",
        "notes": "Ya se maneja como PDF generado por backend, sin dependencia Crystal.",
    },
    {
        "key": "crystal_certificados_promocion",
        "title": "Certificados de promoción y matrícula",
        "category": "Certificados",
        "legacy_rpt": ["certificados/certificadosf.rpt", "certificados/certificado.rpt"],
        "legacy_pages": ["certificados/Repcertificados.aspx", "certificados/RepcertificadosEstud.aspx", "certificados/RepCD.aspx"],
        "source_tables": ["CARRERAXESTUD", "DATOS_ESTUD", "CARRERAS", "PENSUM", "PERIODO"],
        "legacy_filters": ["codigo_periodo", "codigo_estud", "codigo_materia", "PromedioFinal>=7"],
        "modern_equivalent": "certificados.py generar-pdf",
        "modern_format": ["PDF", "ZIP"],
        "migration_status": "modernizado",
        "notes": "Debe mantener generación individual y masiva.",
    },
    {
        "key": "crystal_certificados_educacion_continua",
        "title": "Certificados de educación continua",
        "category": "Educación continua",
        "legacy_rpt": ["certificados/certificadosEduCon.rpt", "certificados/certificadoEdContinua.rpt"],
        "legacy_pages": ["certificados/RepcertificadosEdContinua.aspx", "Reporteshtml/ListaEstudPeriodoOnline.aspx"],
        "source_tables": ["CABECERAEDUCONTINUA", "CursosEduContinua", "EstudiantesEdContinua", "CARRERAXESTUD"],
        "legacy_filters": ["cod_curso", "codigo_materia"],
        "modern_equivalent": "sisacademico_admin cursos_edu_continua / certificados pendiente PDF específico",
        "modern_format": ["PDF", "Excel"],
        "migration_status": "pendiente",
        "notes": "Crear generador PDF específico si se requiere conservar diseño de certificados de curso.",
    },
    {
        "key": "crystal_evaluacion_docente",
        "title": "Evaluación docente",
        "category": "Evaluación docente",
        "legacy_rpt": [
            "EncuestaDocEstud/Reportes/CryEvaluaciondocente.rpt",
            "EncuestaDocEstud/Reportes/CryEstudEvaluaProfe.rpt",
            "EncuestaDocEstud/Reportes/CryEstudEvaluaProfeAgrupado.rpt",
            "EncuestaDocEstud/Reportes/CryNumEstudEvaluaProfe.rpt",
            "EncuestaDocEstud/Reportes/EstudNOEvaluaProfe.rpt",
        ],
        "legacy_pages": [
            "EncuestaDocEstud/ReporteEncuesta.aspx",
            "EncuestaDocEstud/Reportes/EvaluacionEstud.aspx",
            "EncuestaDocEstud/Reportes/ListaEstudNOEvaluaProfe.aspx",
            "EncuestaDocEstud/Reportes/ListaEvaluacionEstud.aspx",
            "EncuestaDocEstud/Reportes/ListaEvaluacionEstudProfe.aspx",
        ],
        "source_tables": ["resultadoevaluaciondocente", "CUESTIO_ENCUESTA_CADETE", "CARRERAXDOCENTE", "DATOSDOCENTE", "PERIODO"],
        "legacy_filters": ["cod_periodo", "codigo_materia", "codigo_doc", "paralelo", "Encuesta=0"],
        "modern_equivalent": "teacher_evaluation.py reporte-docentes.pdf",
        "modern_format": ["PDF", "ZIP", "Excel"],
        "migration_status": "modernizado",
        "notes": "La generación moderna reemplaza Crystal con reportlab y descarga masiva.",
    },
]


def _modernized_legacy_report_payload(report: dict[str, Any]) -> dict[str, Any]:
    return {
        **report,
        "engine": "modern-reportlab-openpyxl",
        "source_engine": "Crystal Reports solo como referencia historica",
        "target_engine": "FastAPI + SQL Server + reportlab/openpyxl",
        "replacement_rule": "No ejecutar .rpt; reconstruir dataset SQL y generar PDF/Excel/HTML desde backend.",
    }


def _clean(value: str | None) -> str | None:
    text = str(value or "").strip()
    return text or None


def _limit(value: int) -> int:
    return max(1, min(value, 10000))


def _search_like(value: str | None) -> str | None:
    cleaned = _clean(value)
    return f"%{cleaned}%" if cleaned else None


def _serializable(value: Any) -> Any:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bytes):
        return None
    return value


def _rows_from_cursor(cursor: Any) -> tuple[list[str], list[dict[str, Any]]]:
    columns = [column[0] for column in cursor.description]
    rows = [
        {column: _serializable(value) for column, value in zip(columns, row)}
        for row in cursor.fetchall()
    ]
    return columns, rows


def _report_payload(report_key: str) -> dict[str, Any]:
    report = REPORTS[report_key]
    return {"key": report_key, **report}


def _base_params(
    periodo: str | None,
    carrera: str | None,
    estado: str | None,
    buscar: str | None,
    anio: str | None = None,
    genero: str | None = None,
) -> dict[str, str | None]:
    return {
        "periodo": _clean(periodo),
        "carrera": _clean(carrera),
        "estado": _clean(estado),
        "buscar": _search_like(buscar),
        "anio": _clean(anio),
        "genero": _search_like(genero),
    }


def _matriculados_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(ce.codigo_periodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            p.anio,
            CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            CAST(ce.codigo_estud AS varchar(30)) AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            de.Apellidos_nombre AS estudiante,
            de.Estado AS estado_codigo,
            LTRIM(RTRIM(es.ESTADO)) AS estado,
            LTRIM(RTRIM(ce.paralelo)) AS paralelo,
            ce.TipoMatricula AS tipo_matricula,
            MAX(pe.Semestre) AS nivel_maximo,
            COUNT(DISTINCT ce.codigo_materia) AS materias,
            AVG(CAST(ce.PromedioFinal AS decimal(18, 2))) AS promedio_final
        FROM dbo.CARRERAXESTUD ce
        INNER JOIN dbo.DATOS_ESTUD de ON ce.codigo_estud = de.codigo_estud
        INNER JOIN dbo.CARRERAS c ON ce.cod_anio_Basica = c.Cod_AnioBasica
        INNER JOIN dbo.PERIODO p ON ce.codigo_periodo = p.cod_periodo
        INNER JOIN dbo.PENSUM pe ON ce.codigo_materia = pe.codigo_materia
        LEFT JOIN dbo.ESTADO es ON de.Estado = es.IDESTADO
        WHERE c.Cod_AnioBasica NOT IN (12, 13)
          AND (? IS NULL OR CAST(ce.codigo_periodo AS varchar(30)) = ?)
          AND (? IS NULL OR CAST(c.Cod_AnioBasica AS varchar(30)) = ?)
          AND (? IS NULL OR de.Estado = ?)
          AND (
            ? IS NULL
            OR de.Apellidos_nombre LIKE ?
            OR de.Cedula_Est LIKE ?
            OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
            OR LTRIM(RTRIM(pe.Nomb_Materia)) LIKE ?
          )
        GROUP BY
            ce.codigo_periodo, p.Detalle_Periodo, p.anio, c.Cod_AnioBasica, c.Nombre_Basica,
            ce.codigo_estud, de.Cedula_Est, de.Apellidos_nombre, de.Estado, es.ESTADO,
            ce.paralelo, ce.TipoMatricula
        ORDER BY p.anio DESC, p.Detalle_Periodo DESC, c.Nombre_Basica, de.Apellidos_nombre
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        params["estado"], params["estado"],
        buscar, buscar, buscar, buscar, buscar,
    ]


def _becas_edades_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(de.codigo_estud AS varchar(30)) AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            de.Apellidos_nombre AS estudiante,
            de.Estado AS estado_codigo,
            LTRIM(RTRIM(es.ESTADO)) AS estado,
            fecha.fecha_nacimiento,
            edad.edad,
            CASE
                WHEN edad.edad IS NULL THEN 'Sin fecha'
                WHEN edad.edad < 18 THEN 'Menor de 18'
                WHEN edad.edad BETWEEN 18 AND 29 THEN '18 a 29'
                WHEN edad.edad BETWEEN 30 AND 40 THEN '30 a 40'
                WHEN edad.edad BETWEEN 41 AND 50 THEN '41 a 50'
                WHEN edad.edad BETWEEN 51 AND 60 THEN '51 a 60'
                ELSE '61 o mas'
            END AS rango_edad,
            COALESCE(NULLIF(LTRIM(RTRIM(beca.tipo_beca)), ''), 'Sin beca') AS tipo_beca,
            CASE
                WHEN COALESCE(NULLIF(LTRIM(RTRIM(beca.tipo_beca)), ''), '') = 'Beca Futuro Femenino' THEN CAST(50 AS decimal(10, 2))
                ELSE COALESCE(TRY_CONVERT(decimal(10, 2), beca.porcentaje_beca), CAST(0 AS decimal(10, 2)))
            END AS porcentaje_beca,
            mat.periodo_codigo,
            mat.periodo,
            mat.carrera_codigo,
            mat.carrera
        FROM dbo.DATOS_ESTUD de
        LEFT JOIN dbo.ESTADO es ON de.Estado = es.IDESTADO
        OUTER APPLY (
            SELECT TRY_CONVERT(date, de.Fecha_Nac) AS fecha_nacimiento
        ) fecha
        OUTER APPLY (
            SELECT
                CASE
                    WHEN fecha.fecha_nacimiento IS NULL THEN NULL
                    ELSE DATEDIFF(YEAR, fecha.fecha_nacimiento, CAST(GETDATE() AS date))
                        - CASE
                            WHEN DATEADD(
                                YEAR,
                                DATEDIFF(YEAR, fecha.fecha_nacimiento, CAST(GETDATE() AS date)),
                                fecha.fecha_nacimiento
                              ) > CAST(GETDATE() AS date)
                            THEN 1
                            ELSE 0
                          END
                END AS edad
        ) edad
        OUTER APPLY (
            SELECT TOP (1)
                B.tipo_beca,
                B.porcentaje_beca
            FROM dbo.Becas B
            WHERE TRY_CONVERT(varchar(50), B.codestud) = TRY_CONVERT(varchar(50), de.codigo_estud)
            ORDER BY
                CASE WHEN NULLIF(LTRIM(RTRIM(B.tipo_beca)), '') IS NULL THEN 1 ELSE 0 END,
                TRY_CONVERT(decimal(10, 2), B.porcentaje_beca) DESC
        ) beca
        OUTER APPLY (
            SELECT TOP (1)
                CAST(ce.codigo_periodo AS varchar(30)) AS periodo_codigo,
                LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
                CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
                LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
                p.anio,
                p.Orden
            FROM dbo.CARRERAXESTUD ce
            LEFT JOIN dbo.PERIODO p ON ce.codigo_periodo = p.cod_periodo
            LEFT JOIN dbo.CARRERAS c ON ce.cod_anio_Basica = c.Cod_AnioBasica
            WHERE TRY_CONVERT(varchar(50), ce.codigo_estud) = TRY_CONVERT(varchar(50), de.codigo_estud)
              AND (? IS NULL OR CAST(ce.codigo_periodo AS varchar(30)) = ?)
              AND (? IS NULL OR CAST(ce.cod_anio_Basica AS varchar(30)) = ?)
            ORDER BY
                COALESCE(TRY_CONVERT(int, p.anio), 0) DESC,
                COALESCE(TRY_CONVERT(int, p.Orden), 0) DESC,
                TRY_CONVERT(int, ce.codigo_periodo) DESC
        ) mat
        WHERE (? IS NULL OR mat.periodo_codigo IS NOT NULL)
          AND (? IS NULL OR mat.carrera_codigo IS NOT NULL)
          AND (? IS NULL OR de.Estado = ?)
          AND (
            ? IS NULL
            OR de.Apellidos_nombre LIKE ?
            OR de.Cedula_Est LIKE ?
            OR COALESCE(NULLIF(LTRIM(RTRIM(beca.tipo_beca)), ''), 'Sin beca') LIKE ?
            OR LTRIM(RTRIM(mat.carrera)) LIKE ?
          )
        ORDER BY
            CASE
                WHEN edad.edad IS NULL THEN 99
                WHEN edad.edad < 18 THEN 0
                WHEN edad.edad BETWEEN 18 AND 29 THEN 1
                WHEN edad.edad BETWEEN 30 AND 40 THEN 2
                WHEN edad.edad BETWEEN 41 AND 50 THEN 3
                WHEN edad.edad BETWEEN 51 AND 60 THEN 4
                ELSE 5
            END,
            de.Apellidos_nombre
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        params["periodo"],
        params["carrera"],
        params["estado"], params["estado"],
        buscar, buscar, buscar, buscar, buscar,
    ]


def _preinscritos_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            pr.Cedula AS cedula,
            pr.Apellidos_nombre AS estudiante,
            pr.correo,
            pr.telefono,
            pr.Fecha_Ingreso AS fecha_ingreso,
            CAST(pr.codperiodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            CAST(pr.codcarrera AS varchar(30)) AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            prov.Descripcion_Prov AS provincia,
            le.Lecontacto AS medio_contacto,
            pr.Prematricula AS prematricula,
            pr.ProcesoFinalilzado AS proceso_finalizado,
            pr.ControlIngreso AS control_ingreso,
            us.nombres AS asesor
        FROM dbo.PREINSCRIPCION pr
        LEFT JOIN dbo.PERIODO p ON pr.codperiodo = p.cod_periodo
        LEFT JOIN dbo.CARRERAS c ON pr.codcarrera = c.Cod_AnioBasica
        LEFT JOIN dbo.Provincias prov ON CAST(pr.codprov AS varchar(30)) = CAST(prov.Cod_Provincia AS varchar(30))
        LEFT JOIN dbo.IN_LECONTACTO le ON pr.codLecontacto = le.Num_Le
        LEFT JOIN dbo.USUARIO_SIS us ON pr.codasesor = us.id_usuarios
        WHERE (? IS NULL OR CAST(pr.codperiodo AS varchar(30)) = ?)
          AND (? IS NULL OR CAST(pr.codcarrera AS varchar(30)) = ?)
          AND (
            ? IS NULL
            OR (? = 'PREMATRICULA' AND ISNULL(pr.Prematricula, 0) = 1)
            OR (? = 'FINALIZADO' AND ISNULL(pr.ProcesoFinalilzado, 0) = 1)
            OR (? = 'PENDIENTE' AND ISNULL(pr.ProcesoFinalilzado, 0) = 0)
          )
          AND (
            ? IS NULL
            OR pr.Apellidos_nombre LIKE ?
            OR pr.Cedula LIKE ?
            OR pr.correo LIKE ?
            OR pr.telefono LIKE ?
            OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
            OR us.nombres LIKE ?
          )
        ORDER BY pr.Fecha_Ingreso DESC, pr.Apellidos_nombre
    """
    estado = params["estado"].upper() if params["estado"] else None
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        estado, estado, estado, estado,
        buscar, buscar, buscar, buscar, buscar, buscar, buscar,
    ]


def _docentes_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(cxd.codigo_periodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            CAST(d.codigo_doc AS varchar(30)) AS docente_codigo,
            LTRIM(RTRIM(d.cedula_doc)) AS cedula_docente,
            LTRIM(RTRIM(d.apellidos_nombre)) AS docente,
            LTRIM(RTRIM(d.correo)) AS correo,
            LTRIM(RTRIM(d.movil)) AS movil,
            CAST(cxd.codigo_materia AS varchar(30)) AS materia_codigo,
            pe.cod_materia AS materia_codigo_texto,
            LTRIM(RTRIM(pe.Nomb_Materia)) AS materia,
            pe.Semestre AS semestre,
            LTRIM(RTRIM(cxd.Paralelo)) AS paralelo,
            cxd.Cod_Jornada AS jornada_codigo,
            j.DetalleJ AS jornada,
            cxd.estadoMoodleDoc AS estado_moodle_doc
        FROM dbo.CARRERAXDOCENTE cxd
        INNER JOIN dbo.DATOSDOCENTE d ON cxd.codigo_doc = d.codigo_doc
        INNER JOIN dbo.CARRERAS c ON cxd.cod_Anio_Basica = c.Cod_AnioBasica
        INNER JOIN dbo.PENSUM pe ON cxd.codigo_materia = pe.codigo_materia
        INNER JOIN dbo.PERIODO p ON cxd.codigo_periodo = p.cod_periodo
        LEFT JOIN dbo.JORNADA j ON cxd.Cod_Jornada = j.NumJ
        WHERE (? IS NULL OR CAST(cxd.codigo_periodo AS varchar(30)) = ?)
          AND (? IS NULL OR CAST(c.Cod_AnioBasica AS varchar(30)) = ?)
          AND (? IS NULL OR ISNULL(CAST(cxd.estadoMoodleDoc AS varchar(1)), '0') = ?)
          AND (
            ? IS NULL
            OR LTRIM(RTRIM(d.apellidos_nombre)) LIKE ?
            OR LTRIM(RTRIM(d.cedula_doc)) LIKE ?
            OR LTRIM(RTRIM(pe.Nomb_Materia)) LIKE ?
            OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
          )
        ORDER BY p.anio DESC, p.Detalle_Periodo DESC, c.Nombre_Basica, d.apellidos_nombre
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        params["estado"], params["estado"],
        buscar, buscar, buscar, buscar, buscar,
    ]


def _documentos_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        WITH documentos AS (
            SELECT
                CAST(cm.codigo_periodo AS varchar(30)) AS periodo_codigo,
                LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
                CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
                LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
                CAST(cm.codigo_estud AS varchar(30)) AS estudiante_codigo,
                de.Cedula_Est AS cedula,
                de.Apellidos_nombre AS estudiante,
                cm.fecha_pago,
                cm.valor,
                cm.urlcedula,
                cm.urltitulo,
                cm.urldeposito,
                cm.urlconvenio,
                CASE
                    WHEN NULLIF(LTRIM(RTRIM(ISNULL(cm.urlcedula, ''))), '') IS NULL
                      OR NULLIF(LTRIM(RTRIM(ISNULL(cm.urltitulo, ''))), '') IS NULL
                      OR NULLIF(LTRIM(RTRIM(ISNULL(cm.urldeposito, ''))), '') IS NULL
                    THEN 'PENDIENTE'
                    ELSE 'COMPLETO'
                END AS estado_documentos
            FROM dbo.CABECERA_MATRICULA cm
            INNER JOIN dbo.DATOS_ESTUD de ON cm.codigo_estud = de.codigo_estud
            INNER JOIN dbo.CARRERAS c ON cm.cod_anio_Basica = c.Cod_AnioBasica
            INNER JOIN dbo.PERIODO p ON cm.codigo_periodo = p.cod_periodo
            WHERE (? IS NULL OR CAST(cm.codigo_periodo AS varchar(30)) = ?)
              AND (? IS NULL OR CAST(c.Cod_AnioBasica AS varchar(30)) = ?)
              AND (
                ? IS NULL
                OR de.Apellidos_nombre LIKE ?
                OR de.Cedula_Est LIKE ?
                OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
              )
        )
        SELECT TOP ({limit}) *
        FROM documentos
        WHERE (? IS NULL OR estado_documentos = ?)
        ORDER BY periodo DESC, carrera, estudiante
    """
    estado = params["estado"].upper() if params["estado"] else None
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        buscar, buscar, buscar, buscar,
        estado, estado,
    ]


def _seguimiento_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            se.fecha_registro,
            se.codigo_estud AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            COALESCE(de.Apellidos_nombre, se.nombres) AS estudiante,
            se.codigo_periodo AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            c.Cod_AnioBasica AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            se.codigo_materia AS materia_codigo,
            LTRIM(RTRIM(pe.Nomb_Materia)) AS materia,
            se.tipo_seguimiento,
            se.observacion,
            se.usuario_registro
        FROM dbo.SEGUIMIENTO_ESTUDIANTE se
        LEFT JOIN dbo.DATOS_ESTUD de ON CAST(se.codigo_estud AS varchar(50)) = CAST(de.codigo_estud AS varchar(50))
        LEFT JOIN dbo.PERIODO p ON CAST(se.codigo_periodo AS varchar(50)) = CAST(p.cod_periodo AS varchar(50))
        LEFT JOIN dbo.PENSUM pe ON CAST(se.codigo_materia AS varchar(50)) = CAST(pe.codigo_materia AS varchar(50))
        LEFT JOIN dbo.CARRERAS c ON pe.Cod_AnioBasica = c.Cod_AnioBasica
        WHERE se.activo = 1
          AND (? IS NULL OR se.codigo_periodo = ?)
          AND (? IS NULL OR CAST(c.Cod_AnioBasica AS varchar(30)) = ?)
          AND (
            ? IS NULL
            OR COALESCE(de.Apellidos_nombre, se.nombres) LIKE ?
            OR de.Cedula_Est LIKE ?
            OR LTRIM(RTRIM(pe.Nomb_Materia)) LIKE ?
            OR se.observacion LIKE ?
            OR se.tipo_seguimiento LIKE ?
          )
        ORDER BY se.fecha_registro DESC
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        buscar, buscar, buscar, buscar, buscar, buscar,
    ]


def _practicas_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(pp.codigo_periodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            CAST(pp.cod_anio_Basica AS varchar(30)) AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            CAST(pp.codigo_estud AS varchar(30)) AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            de.Apellidos_nombre AS estudiante,
            LTRIM(RTRIM(emp.Empresa)) AS empresa,
            pp.FechaInicio AS fecha_inicio,
            pp.FechaFinal AS fecha_final,
            pp.NoHoras AS horas,
            LTRIM(RTRIM(dd.apellidos_nombre)) AS docente_tutor,
            pp.Semestre AS semestre,
            pp.DetalleProyecto AS proyecto,
            LTRIM(RTRIM(pp.pathAr)) AS archivo
        FROM dbo.PRACTICASPROFESIONALES pp
        INNER JOIN dbo.DATOS_ESTUD de ON pp.codigo_estud = de.codigo_estud
        INNER JOIN dbo.CARRERAS c ON pp.cod_anio_Basica = c.Cod_AnioBasica
        INNER JOIN dbo.PERIODO p ON pp.codigo_periodo = p.cod_periodo
        LEFT JOIN dbo.EMPRESA emp ON pp.Cod_empresa = emp.Num_emp
        LEFT JOIN dbo.DATOSDOCENTE dd ON pp.CodDocente = dd.codigo_doc
        WHERE (? IS NULL OR CAST(pp.codigo_periodo AS varchar(30)) = ?)
          AND (? IS NULL OR CAST(pp.cod_anio_Basica AS varchar(30)) = ?)
          AND (
            ? IS NULL
            OR de.Apellidos_nombre LIKE ?
            OR de.Cedula_Est LIKE ?
            OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
            OR LTRIM(RTRIM(emp.Empresa)) LIKE ?
            OR LTRIM(RTRIM(dd.apellidos_nombre)) LIKE ?
            OR pp.DetalleProyecto LIKE ?
          )
        ORDER BY pp.FechaInicio DESC, de.Apellidos_nombre
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        buscar, buscar, buscar, buscar, buscar, buscar, buscar,
    ]


def _evaluacion_docente_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(re.Cod_periodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            CAST(re.Cod_Doc_Eval AS varchar(30)) AS docente_codigo,
            LTRIM(RTRIM(dd.apellidos_nombre)) AS docente,
            CAST(re.Cod_Materia AS varchar(30)) AS materia_codigo,
            LTRIM(RTRIM(pe.Nomb_Materia)) AS materia,
            LTRIM(RTRIM(re.Jornada)) AS jornada,
            LTRIM(RTRIM(re.Paralelo)) AS paralelo,
            COUNT(*) AS respuestas,
            AVG(CAST(re.Puntaje AS decimal(18, 2))) AS promedio_puntaje,
            AVG(CAST(re.Puntaje AS decimal(18, 2))) / 6 * 100 AS porcentaje,
            MIN(re.fecha) AS fecha_inicio,
            MAX(re.fecha) AS fecha_final
        FROM dbo.RESULTADO_EVALUACION re
        LEFT JOIN dbo.DATOSDOCENTE dd ON re.Cod_Doc_Eval = dd.codigo_doc
        LEFT JOIN dbo.CARRERAXDOCENTE cxd
            ON re.Cod_Doc_Eval = cxd.codigo_doc
            AND re.Cod_Materia = cxd.codigo_materia
            AND re.Cod_periodo = cxd.codigo_periodo
            AND ISNULL(LTRIM(RTRIM(re.Paralelo)), '') = ISNULL(LTRIM(RTRIM(cxd.Paralelo)), '')
        LEFT JOIN dbo.CARRERAS c ON cxd.cod_Anio_Basica = c.Cod_AnioBasica
        LEFT JOIN dbo.PENSUM pe ON re.Cod_Materia = pe.codigo_materia
        LEFT JOIN dbo.PERIODO p ON re.Cod_periodo = p.cod_periodo
        WHERE (? IS NULL OR CAST(re.Cod_periodo AS varchar(30)) = ?)
          AND (? IS NULL OR CAST(c.Cod_AnioBasica AS varchar(30)) = ?)
          AND (
            ? IS NULL
            OR LTRIM(RTRIM(dd.apellidos_nombre)) LIKE ?
            OR LTRIM(RTRIM(pe.Nomb_Materia)) LIKE ?
            OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
            OR re.Detalle_Preg LIKE ?
          )
        GROUP BY
            re.Cod_periodo, p.Detalle_Periodo, c.Cod_AnioBasica, c.Nombre_Basica,
            re.Cod_Doc_Eval, dd.apellidos_nombre, re.Cod_Materia, pe.Nomb_Materia,
            re.Jornada, re.Paralelo
        ORDER BY p.Detalle_Periodo DESC, dd.apellidos_nombre, pe.Nomb_Materia
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        buscar, buscar, buscar, buscar, buscar,
    ]


def _moodle_notas_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            en.periodo AS periodo_codigo,
            en.codigo_estudiante AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            de.Apellidos_nombre AS estudiante,
            en.codigo_materia AS materia_codigo,
            LTRIM(RTRIM(pe.Nomb_Materia)) AS materia,
            en.paralelo,
            en.tipo_matricula,
            en.componente_nota,
            en.nota_obtenida,
            en.nota_maxima,
            en.porcentaje,
            en.estado,
            en.moodle_course_id,
            en.moodle_grade_item_id,
            en.fecha_calificacion,
            en.fecha_sincronizacion,
            en.calificado_por
        FROM dbo.intec_estudiantenota en
        LEFT JOIN dbo.DATOS_ESTUD de ON CAST(en.codigo_estudiante AS varchar(50)) = CAST(de.codigo_estud AS varchar(50))
        LEFT JOIN dbo.PENSUM pe ON CAST(en.codigo_materia AS varchar(50)) = CAST(pe.codigo_materia AS varchar(50))
        WHERE (? IS NULL OR en.periodo = ?)
          AND (? IS NULL OR en.estado = ?)
          AND (
            ? IS NULL
            OR en.codigo_estudiante LIKE ?
            OR de.Cedula_Est LIKE ?
            OR de.Apellidos_nombre LIKE ?
            OR LTRIM(RTRIM(pe.Nomb_Materia)) LIKE ?
            OR en.componente_nota LIKE ?
            OR en.calificado_por LIKE ?
          )
        ORDER BY en.fecha_sincronizacion DESC, de.Apellidos_nombre
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["estado"], params["estado"],
        buscar, buscar, buscar, buscar, buscar, buscar, buscar,
    ]


def _notas_carrera_materia_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(ce.codigo_periodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            CAST(pe.codigo_materia AS varchar(30)) AS materia_codigo,
            LTRIM(RTRIM(pe.cod_materia)) AS materia_codigo_texto,
            LTRIM(RTRIM(pe.Nomb_Materia)) AS materia,
            pe.Semestre AS semestre,
            LTRIM(RTRIM(ce.paralelo)) AS paralelo,
            ce.TipoMatricula AS tipo_matricula,
            CASE
                WHEN UPPER(LTRIM(RTRIM(ISNULL(ce.TipoMatricula, '')))) = 'H'
                  OR UPPER(LTRIM(RTRIM(ISNULL(p.TipoMatricula, '')))) = 'H'
                  OR UPPER(LTRIM(RTRIM(ISNULL(p.Detalle_Periodo, '')))) LIKE '%HOMO%'
                THEN 'HOMOLOGACION'
                ELSE 'REGULAR'
            END AS esquema,
            CAST(ce.codigo_estud AS varchar(30)) AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            de.Apellidos_nombre AS estudiante,
            de.Estado AS estado_codigo,
            ce.teoriaHomo AS teoria_homo,
            ce.practicahomo AS practica_homo,
            ce.P1Tareas AS p1_tareas,
            ce.P1Proyectos AS p1_proyectos,
            ce.P1Examen AS p1_examen,
            ce.promP1 AS promedio_p1,
            ce.P2Tareas AS p2_tareas,
            ce.P2Proyectos AS p2_proyectos,
            ce.P2Examen AS p2_examen,
            ce.promP2 AS promedio_p2,
            ce.P3Tareas AS p3_tareas,
            ce.P3Proyectos AS p3_proyectos,
            ce.P3Examen AS p3_examen,
            ce.promP3 AS promedio_p3,
            ce.Asistencia AS asistencia,
            ce.Recuperacion AS recuperacion,
            ce.PromedioFinal AS promedio_final,
            LTRIM(RTRIM(ce.caprueba)) AS condicion,
            ce.estadoMoodle AS estado_moodle,
            ce.seguimiento,
            ce.observaciones
        FROM dbo.CARRERAXESTUD ce
        INNER JOIN dbo.DATOS_ESTUD de ON ce.codigo_estud = de.codigo_estud
        INNER JOIN dbo.CARRERAS c ON ce.cod_anio_Basica = c.Cod_AnioBasica
        INNER JOIN dbo.PENSUM pe ON ce.codigo_materia = pe.codigo_materia
        INNER JOIN dbo.PERIODO p ON ce.codigo_periodo = p.cod_periodo
        WHERE (? IS NULL OR CAST(ce.codigo_periodo AS varchar(30)) = ?)
          AND (? IS NULL OR CAST(c.Cod_AnioBasica AS varchar(30)) = ?)
          AND UPPER(LTRIM(RTRIM(ISNULL(de.Estado, '')))) = 'A'
          AND (
            ? IS NULL
            OR de.Apellidos_nombre LIKE ?
            OR de.Cedula_Est LIKE ?
            OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
            OR LTRIM(RTRIM(pe.Nomb_Materia)) LIKE ?
            OR LTRIM(RTRIM(ce.paralelo)) LIKE ?
          )
        ORDER BY p.anio DESC, p.Detalle_Periodo DESC, c.Nombre_Basica, pe.Semestre, pe.Nomb_Materia, de.Apellidos_nombre
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        buscar, buscar, buscar, buscar, buscar, buscar,
    ]


def _estud_per_c_m_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        WITH matriculas AS (
            SELECT
                CAST(ce.codigo_periodo AS varchar(30)) AS periodo_codigo,
                LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
                CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
                LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
                CAST(ce.codigo_estud AS varchar(30)) AS estudiante_codigo,
                de.Cedula_Est AS cedula,
                LTRIM(RTRIM(de.Apellidos_nombre)) AS estudiante,
                CAST(pe.codigo_materia AS varchar(30)) AS materia_codigo,
                LTRIM(RTRIM(pe.cod_materia)) AS codigo_interno,
                LTRIM(RTRIM(pe.Nomb_Materia)) AS materia,
                pe.Semestre AS nivel,
                pe.Creditos AS creditos,
                pe.Horas AS horas,
                LTRIM(RTRIM(ce.paralelo)) AS paralelo,
                LTRIM(RTRIM(ce.TipoMatricula)) AS tipo_matricula,
                CASE
                    WHEN UPPER(LTRIM(RTRIM(ISNULL(ce.TipoMatricula, '')))) = 'H' THEN 'HOMOLOGACION'
                    ELSE 'REGULAR'
                END AS esquema,
                ce.teoriaHomo AS teoria_homo,
                ce.practicahomo AS practica_homo,
                ce.promP1 AS promedio_p1,
                ce.promP2 AS promedio_p2,
                ce.promP3 AS promedio_p3,
                ce.PromedioFinal AS promedio_final,
                ce.Asistencia AS asistencia,
                ce.Num_Matricula AS num_matricula,
                ce.Fecha_Matricula AS fecha_matricula,
                CASE
                    WHEN ce.PromedioFinal IS NULL THEN 'PENDIENTE'
                    WHEN ce.PromedioFinal >= 7 THEN 'APROBADA'
                    WHEN ce.PromedioFinal < 7 THEN 'REPROBADA'
                    ELSE 'PENDIENTE'
                END AS estado_materia
            FROM dbo.CARRERAXESTUD ce
            INNER JOIN dbo.DATOS_ESTUD de ON ce.codigo_estud = de.codigo_estud
            INNER JOIN dbo.CARRERAS c ON ce.cod_anio_Basica = c.Cod_AnioBasica
            INNER JOIN dbo.PENSUM pe ON ce.codigo_materia = pe.codigo_materia
            INNER JOIN dbo.PERIODO p ON ce.codigo_periodo = p.cod_periodo
            WHERE (? IS NULL OR CAST(ce.codigo_periodo AS varchar(30)) = ?)
              AND (? IS NULL OR CAST(c.Cod_AnioBasica AS varchar(30)) = ?)
              AND (
                ? IS NULL
                OR de.Apellidos_nombre LIKE ?
                OR de.Cedula_Est LIKE ?
                OR LTRIM(RTRIM(c.Nombre_Basica)) LIKE ?
                OR LTRIM(RTRIM(pe.Nomb_Materia)) LIKE ?
                OR LTRIM(RTRIM(pe.cod_materia)) LIKE ?
                OR LTRIM(RTRIM(ce.paralelo)) LIKE ?
              )
        )
        SELECT TOP ({limit})
            periodo_codigo,
            periodo,
            carrera_codigo,
            carrera,
            estudiante_codigo,
            cedula,
            estudiante,
            materia_codigo,
            codigo_interno,
            materia,
            nivel,
            creditos,
            horas,
            paralelo,
            tipo_matricula,
            esquema,
            teoria_homo,
            practica_homo,
            promedio_p1,
            promedio_p2,
            promedio_p3,
            promedio_final,
            asistencia,
            num_matricula,
            fecha_matricula,
            estado_materia
        FROM matriculas
        WHERE (? IS NULL OR estado_materia = ?)
        ORDER BY periodo DESC, carrera, estudiante, nivel, materia
    """
    buscar = params["buscar"]
    estado = params["estado"].upper() if params["estado"] else None
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        buscar, buscar, buscar, buscar, buscar, buscar, buscar,
        estado, estado,
    ]


def _correos_intec_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(ci.Periodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            CAST(ci.codestud AS varchar(30)) AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            ci.Nombres AS estudiante,
            ci.CorreoPersonal AS correo_personal,
            ci.CorreoIntec AS correo_intec,
            ci.fecha AS fecha_creacion,
            ci.CorreoEnviado AS correo_enviado,
            ci.Estado AS estado,
            ci.Descripcion AS descripcion,
            ci.ultAccesoMoodle AS ultimo_acceso_moodle,
            ci.NumMigracion AS num_migracion,
            ci.TipoCursoMigra AS tipo_curso_migra
        FROM dbo.CorreosEstudIntec ci
        LEFT JOIN dbo.DATOS_ESTUD de ON ci.codestud = de.codigo_estud
        LEFT JOIN dbo.PERIODO p ON ci.Periodo = p.cod_periodo
        WHERE (? IS NULL OR CAST(ci.Periodo AS varchar(30)) = ?)
          AND (? IS NULL OR UPPER(ISNULL(ci.Estado, '')) = ?)
          AND (
            ? IS NULL
            OR ci.Nombres LIKE ?
            OR de.Cedula_Est LIKE ?
            OR ci.CorreoPersonal LIKE ?
            OR ci.CorreoIntec LIKE ?
            OR ci.Descripcion LIKE ?
          )
        ORDER BY ci.fecha DESC, ci.Nombres
    """
    buscar = params["buscar"]
    estado = params["estado"].upper() if params["estado"] else None
    return sql, [
        params["periodo"], params["periodo"],
        estado, estado,
        buscar, buscar, buscar, buscar, buscar, buscar,
    ]


def _microsoft_audit_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            fecha,
            correo,
            accion,
            estado,
            skuIdAsignado AS sku_id_asignado,
            mensaje_error
        FROM dbo.Microsoft365Audit
        WHERE (? IS NULL OR UPPER(ISNULL(estado, '')) = ?)
          AND (
            ? IS NULL
            OR correo LIKE ?
            OR accion LIKE ?
            OR estado LIKE ?
            OR skuIdAsignado LIKE ?
            OR mensaje_error LIKE ?
          )
        ORDER BY fecha DESC
    """
    buscar = params["buscar"]
    estado = params["estado"].upper() if params["estado"] else None
    return sql, [
        estado, estado,
        buscar, buscar, buscar, buscar, buscar, buscar,
    ]


def _pagos_matricula_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        SELECT TOP ({limit})
            CAST(rp.codperiodo AS varchar(30)) AS periodo_codigo,
            LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
            CAST(rp.cod_anio_Basica AS varchar(30)) AS carrera_codigo,
            LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
            CAST(rp.Codestu AS varchar(30)) AS estudiante_codigo,
            de.Cedula_Est AS cedula,
            de.Apellidos_nombre AS estudiante,
            rp.fechapago AS fecha_pago,
            rp.Detalle AS detalle,
            rp.Valor AS valor,
            rp.FechaRegistro AS fecha_registro,
            rp.usuarioreg AS usuario_registro,
            rp.NoDeposito AS no_deposito,
            rp.Banco AS banco,
            rp.FechaDeposito AS fecha_deposito,
            rp.ValorRegistrado AS valor_registrado,
            rp.urldeposito AS url_deposito
        FROM dbo.REGISTROPAGOS rp
        LEFT JOIN dbo.DATOS_ESTUD de ON rp.Codestu = de.codigo_estud
        LEFT JOIN dbo.PERIODO p ON rp.codperiodo = p.cod_periodo
        LEFT JOIN dbo.CARRERAS c ON rp.cod_anio_Basica = c.Cod_AnioBasica
        WHERE (? IS NULL OR CAST(rp.codperiodo AS varchar(30)) = ?)
          AND (? IS NULL OR CAST(rp.cod_anio_Basica AS varchar(30)) = ?)
          AND (
            ? IS NULL
            OR de.Apellidos_nombre LIKE ?
            OR de.Cedula_Est LIKE ?
            OR rp.Detalle LIKE ?
            OR rp.NoDeposito LIKE ?
            OR rp.Banco LIKE ?
          )
        ORDER BY rp.fechapago DESC, de.Apellidos_nombre
    """
    buscar = params["buscar"]
    return sql, [
        params["periodo"], params["periodo"],
        params["carrera"], params["carrera"],
        buscar, buscar, buscar, buscar, buscar, buscar,
    ]


STUDENT_ESTADO_OPTIONS = [
    {"value": "", "label": "Todos los estados"},
    {"value": "A", "label": "Activo"},
    {"value": "G", "label": "Graduado"},
    {"value": "P", "label": "Inactivo"},
    {"value": "R", "label": "Retirado"},
]


def _default_estado_for_report(report_key: str, estado: str | None) -> str | None:
    cleaned = _clean(estado)
    if cleaned:
        return cleaned
    if report_key == "graduados_2025":
        return "G"
    return None


REPORTS = {
    "provincia": {
        "title": "Provincia",
        "description": "Totales por provincia divididos en matrícula regular y homologación, filtrable por año.",
        "category": "Reportería R/H",
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PERIODO", "Provincias"],
        "filters": ["anio", "estado", "buscar", "limite"],
        "estado_options": STUDENT_ESTADO_OPTIONS,
    },
    "provincia_genero": {
        "title": "Provincia por género",
        "description": "Consolidado de estudiantes por provincia y género, dividido en regular y homologación.",
        "category": "Reportería R/H",
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PERIODO", "CARRERAS", "Provincias", "Sexo"],
        "filters": ["anio", "estado", "genero", "buscar", "limite"],
        "estado_options": STUDENT_ESTADO_OPTIONS,
    },
    "provincia_carrera": {
        "title": "Provincia por carreras",
        "description": "Consolidado de estudiantes por provincia y carrera, dividido en regular y homologación.",
        "category": "Reportería R/H",
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PERIODO", "CARRERAS", "Provincias"],
        "filters": ["anio", "estado", "carrera", "buscar", "limite"],
        "estado_options": STUDENT_ESTADO_OPTIONS,
    },
    "carrera": {
        "title": "Carrera",
        "description": "Totales por carrera divididos en matrícula regular y homologación.",
        "category": "Reportería R/H",
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PERIODO", "CARRERAS"],
        "filters": ["anio", "estado", "carrera", "buscar", "limite"],
        "estado_options": STUDENT_ESTADO_OPTIONS,
    },
    "graduados_2025": {
        "title": "Graduados",
        "description": "Listado de estudiantes graduados por año, provincia, carrera y género.",
        "category": "Reportería R/H",
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PERIODO", "CARRERAS", "ESTADO"],
        "filters": ["anio", "estado", "carrera", "genero", "buscar", "limite"],
        "estado_options": STUDENT_ESTADO_OPTIONS,
    },
    "genero": {
        "title": "Género",
        "description": "Distribución por género, dividida en matrícula regular y homologación.",
        "category": "Reportería R/H",
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PERIODO", "Sexo"],
        "filters": ["anio", "estado", "genero", "buscar", "limite"],
        "estado_options": STUDENT_ESTADO_OPTIONS,
    },
    "genero_docentes": {
        "title": "Género de docentes",
        "description": "Distribución de docentes por género y estado activo o inactivo.",
        "category": "Docencia",
        "source_tables": ["DATOSDOCENTE", "USUARIOS", "Sexo"],
        "filters": ["estado", "genero", "buscar", "limite"],
        "estado_options": [
            {"value": "A", "label": "Activo"},
            {"value": "P", "label": "Inactivo"},
        ],
    },
    "periodo": {
        "title": "Período",
        "description": "Totales de estudiantes por período académico, divididos en regular y homologación.",
        "category": "Reportería R/H",
        "source_tables": ["DATOS_ESTUD", "CARRERAXESTUD", "PERIODO"],
        "filters": ["anio", "estado", "carrera", "genero", "buscar", "limite"],
        "estado_options": STUDENT_ESTADO_OPTIONS,
    },
}

FUNCTIONAL_INVENTORY = [
    {
        "module": "Reportes por provincia",
        "legacy_sources": ["DATOS_ESTUD", "Provincias", "CARRERAXESTUD", "PERIODO"],
        "capabilities": ["Provincia por género", "Provincia por carrera", "Filtros por año y género"],
    },
    {
        "module": "Reportes academicos",
        "legacy_sources": ["CARRERAXESTUD", "CARRERAS", "PERIODO", "ESTADO"],
        "capabilities": ["Graduados", "Distribución por período"],
    },
]


def _student_report_base_where(params: dict[str, str | None]) -> tuple[str, list[Any]]:
    clauses = [
        "p.anio IS NOT NULL",
    ]
    values: list[Any] = []
    if params.get("anio"):
        clauses.append("CAST(p.anio AS varchar(10)) = ?")
        values.append(params["anio"])
    if params.get("estado"):
        clauses.append("LTRIM(RTRIM(de.Estado)) = ?")
        values.append(params["estado"])
    if params.get("periodo"):
        clauses.append("CAST(p.cod_periodo AS varchar(30)) = ?")
        values.append(params["periodo"])
    if params.get("carrera"):
        clauses.append("CAST(c.Cod_AnioBasica AS varchar(30)) = ?")
        values.append(params["carrera"])
    if params.get("genero"):
        clauses.append(
            """
            LOWER(COALESCE(NULLIF(LTRIM(RTRIM(s.detalle_sexo)), ''), NULLIF(LTRIM(RTRIM(de.generoId)), ''), 'Sin genero')) LIKE LOWER(?)
            """
        )
        values.append(params["genero"])
    if params.get("buscar"):
        clauses.append(
            """
            (
                de.Cedula_Est LIKE ?
                OR de.Apellidos_nombre LIKE ?
                OR c.Nombre_Basica LIKE ?
                OR p.Detalle_Periodo LIKE ?
                OR pr.Descripcion_Prov LIKE ?
            )
            """
        )
        values.extend([params["buscar"]] * 5)
    return " AND ".join(clauses), values


def _student_report_cte(params: dict[str, str | None]) -> tuple[str, list[Any]]:
    if not params.get("anio") and not params.get("periodo"):
        return _student_report_cne_cte(params)

    where_sql, values = _student_report_base_where(params)
    sql = f"""
        WITH base AS (
            SELECT DISTINCT
                CAST(de.codigo_estud AS varchar(30)) AS estudiante_codigo,
                LTRIM(RTRIM(de.Cedula_Est)) AS cedula,
                LTRIM(RTRIM(de.Apellidos_nombre)) AS estudiante,
                COALESCE(NULLIF(LTRIM(RTRIM(pr.Descripcion_Prov)), ''), NULLIF(LTRIM(RTRIM(de.ciudad)), ''), 'Sin provincia') AS provincia,
                COALESCE(NULLIF(LTRIM(RTRIM(s.detalle_sexo)), ''), NULLIF(LTRIM(RTRIM(de.generoId)), ''), 'Sin genero') AS genero,
                CAST(c.Cod_AnioBasica AS varchar(30)) AS carrera_codigo,
                LTRIM(RTRIM(c.Nombre_Basica)) AS carrera,
                CAST(p.cod_periodo AS varchar(30)) AS periodo_codigo,
                LTRIM(RTRIM(p.Detalle_Periodo)) AS periodo,
                p.anio,
                UPPER(COALESCE(NULLIF(LTRIM(RTRIM(p.TipoMatricula)), ''), NULLIF(LTRIM(RTRIM(ce.TipoMatricula)), ''), 'SIN TIPO')) AS tipo_matricula,
                LTRIM(RTRIM(de.Estado)) AS estado_codigo,
                COALESCE(NULLIF(LTRIM(RTRIM(es.ESTADO)), ''), LTRIM(RTRIM(de.Estado)), 'Sin estado') AS estado
            FROM dbo.CARRERAXESTUD ce
            INNER JOIN dbo.DATOS_ESTUD de ON ce.codigo_estud = de.codigo_estud
            INNER JOIN dbo.PERIODO p ON ce.codigo_periodo = p.cod_periodo
            INNER JOIN dbo.CARRERAS c ON ce.cod_anio_Basica = c.Cod_AnioBasica
            LEFT JOIN dbo.Provincias pr ON TRY_CONVERT(int, pr.Cod_Provincia) = TRY_CONVERT(int, de.codprov)
            LEFT JOIN dbo.Sexo s ON de.Sexo = s.id_sexo
            LEFT JOIN dbo.ESTADO es ON de.Estado = es.IDESTADO
            WHERE {where_sql}
        )
    """
    return sql, values


def _student_report_cne_cte(params: dict[str, str | None]) -> tuple[str, list[Any]]:
    clauses = ["cne.estado_codigo IN ('A', 'G', 'P', 'R')"]
    values: list[Any] = []
    if params.get("estado"):
        clauses.append("cne.estado_codigo = ?")
        values.append(params["estado"])
    if params.get("carrera"):
        clauses.append("TRY_CONVERT(varchar(30), carrera.Cod_AnioBasica) = ?")
        values.append(params["carrera"])
    if params.get("genero"):
        clauses.append(
            "LOWER(COALESCE(NULLIF(LTRIM(RTRIM(sexo.detalle_sexo)), ''), NULLIF(LTRIM(RTRIM(datos.generoId)), ''), 'Sin genero')) LIKE LOWER(?)"
        )
        values.append(params["genero"])
    if params.get("buscar"):
        clauses.append(
            """
            (
                cne.Cedula_Est LIKE ?
                OR cne.Apellidos_nombre LIKE ?
                OR cne.nombre_carrera LIKE ?
                OR provincia.Descripcion_Prov LIKE ?
            )
            """
        )
        values.extend([params["buscar"]] * 4)

    where_sql = " AND ".join(clauses)
    sql = f"""
        WITH cne_raw AS (
            SELECT DISTINCT
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), reporte.Cedula_Est))) AS Cedula_Est,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(4000), reporte.Apellidos_nombre))) AS Apellidos_nombre,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), reporte.ESTADO))) AS estado_nombre,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(4000), reporte.Nombre_Basica))) AS nombre_carrera,
                UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(10), reporte.TipoMatricula)))) AS tipo_matricula
            FROM dbo.TOTALESTUDMATRICCNE reporte
            WHERE UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(10), reporte.TipoMatricula)))) IN ('R', 'H')
        ),
        cne AS (
            SELECT
                cne_raw.*,
                UPPER(LTRIM(RTRIM(TRY_CONVERT(varchar(10), estado.IDESTADO)))) AS estado_codigo
            FROM cne_raw
            INNER JOIN dbo.ESTADO estado
              ON UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), estado.ESTADO)))) = UPPER(cne_raw.estado_nombre)
        ),
        base AS (
            SELECT
                CONCAT(
                    cne.Cedula_Est, '|',
                    cne.Apellidos_nombre, '|',
                    cne.nombre_carrera, '|',
                    cne.tipo_matricula, '|',
                    cne.estado_codigo
                ) AS estudiante_codigo,
                cne.Cedula_Est AS cedula,
                cne.Apellidos_nombre AS estudiante,
                COALESCE(NULLIF(LTRIM(RTRIM(provincia.Descripcion_Prov)), ''), NULLIF(LTRIM(RTRIM(datos.ciudad)), ''), 'Sin provincia') AS provincia,
                COALESCE(NULLIF(LTRIM(RTRIM(sexo.detalle_sexo)), ''), NULLIF(LTRIM(RTRIM(datos.generoId)), ''), 'Sin genero') AS genero,
                COALESCE(TRY_CONVERT(varchar(30), carrera.Cod_AnioBasica), '') AS carrera_codigo,
                cne.nombre_carrera AS carrera,
                CAST('' AS varchar(30)) AS periodo_codigo,
                CAST('Estado actual CNE' AS varchar(100)) AS periodo,
                CAST(NULL AS int) AS anio,
                cne.tipo_matricula,
                cne.estado_codigo,
                cne.estado_nombre AS estado
            FROM cne
            OUTER APPLY (
                SELECT TOP (1) detalle.*
                FROM dbo.DATOS_ESTUD detalle
                WHERE LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), detalle.Cedula_Est))) = cne.Cedula_Est
                ORDER BY TRY_CONVERT(bigint, detalle.codigo_estud) DESC
            ) datos
            OUTER APPLY (
                SELECT TOP (1) catalogo.Cod_AnioBasica
                FROM dbo.CARRERAS catalogo
                WHERE UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(4000), catalogo.Nombre_Basica)))) = UPPER(cne.nombre_carrera)
                ORDER BY TRY_CONVERT(bigint, catalogo.Cod_AnioBasica) DESC
            ) carrera
            LEFT JOIN dbo.Provincias provincia
              ON TRY_CONVERT(int, provincia.Cod_Provincia) = TRY_CONVERT(int, datos.codprov)
            LEFT JOIN dbo.Sexo sexo ON datos.Sexo = sexo.id_sexo
            WHERE {where_sql}
        )
    """
    return sql, values


def _split_count_columns() -> str:
    return """
            COUNT(DISTINCT CASE WHEN tipo_matricula = 'R' THEN estudiante_codigo END) AS regular,
            COUNT(DISTINCT CASE WHEN tipo_matricula = 'H' THEN estudiante_codigo END) AS homologacion,
            COUNT(DISTINCT estudiante_codigo) AS total_estudiantes
    """


def _provincia_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    cte, values = _student_report_cte(params)
    sql = f"""
        {cte}
        SELECT TOP ({limit})
            anio,
            provincia,
            {_split_count_columns()}
        FROM base
        GROUP BY anio, provincia
        ORDER BY anio DESC, provincia
    """
    return sql, values


def _provincia_genero_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    cte, values = _student_report_cte(params)
    sql = f"""
        {cte}
        SELECT TOP ({limit})
            anio,
            provincia,
            genero,
            {_split_count_columns()}
        FROM base
        GROUP BY anio, provincia, genero
        ORDER BY anio DESC, provincia, genero
    """
    return sql, values


def _provincia_carrera_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    cte, values = _student_report_cte(params)
    sql = f"""
        {cte}
        SELECT TOP ({limit})
            anio,
            provincia,
            carrera_codigo,
            carrera,
            {_split_count_columns()}
        FROM base
        GROUP BY anio, provincia, carrera_codigo, carrera
        ORDER BY anio DESC, provincia, carrera
    """
    return sql, values


def _carrera_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    cte, values = _student_report_cte(params)
    sql = f"""
        {cte}
        SELECT TOP ({limit})
            anio,
            carrera_codigo,
            carrera,
            {_split_count_columns()}
        FROM base
        GROUP BY anio, carrera_codigo, carrera
        ORDER BY anio DESC, carrera
    """
    return sql, values


def _graduados_2025_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    cte, values = _student_report_cte(params)
    sql = f"""
        {cte}
        SELECT TOP ({limit})
            cedula,
            estudiante,
            provincia,
            genero,
            carrera,
            periodo,
            anio,
            tipo_matricula,
            estado
        FROM base
        WHERE estado_codigo = 'G' OR LOWER(estado) LIKE '%graduado%'
        ORDER BY anio DESC, carrera, estudiante
    """
    return sql, values


def _genero_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    cte, values = _student_report_cte(params)
    sql = f"""
        {cte}
        SELECT TOP ({limit})
            anio,
            genero,
            {_split_count_columns()}
        FROM base
        GROUP BY anio, genero
        ORDER BY anio DESC, genero
    """
    return sql, values


def _genero_docentes_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    sql = f"""
        WITH docentes AS (
            SELECT
                TRY_CONVERT(varchar(50), d.codigo_doc) AS docente_codigo,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), d.cedula_doc))) AS cedula,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(4000), d.apellidos_nombre))) AS docente,
                LTRIM(RTRIM(TRY_CONVERT(nvarchar(255), d.correo))) AS correo,
                CASE
                    WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), d.sexo))))
                         IN (N'1', N'M', N'MASCULINO', N'HOMBRE') THEN N'Masculino'
                    WHEN UPPER(LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), d.sexo))))
                         IN (N'2', N'F', N'FEMENINO', N'MUJER') THEN N'Femenino'
                    ELSE N'Sin registrar'
                END AS genero,
                UPPER(COALESCE(
                    NULLIF(LTRIM(RTRIM(TRY_CONVERT(nvarchar(20), usuario.Estado))), N''),
                    N'SIN_ESTADO'
                )) AS estado_codigo
            FROM dbo.DATOSDOCENTE d
            OUTER APPLY (
                SELECT TOP (1) u.Estado
                FROM dbo.USUARIOS u
                WHERE COALESCE(TRY_CONVERT(int, u.tipo_usuario), 2) <> 1
                  AND (
                        TRY_CONVERT(int, u.Codigo_Usuario) = TRY_CONVERT(int, d.codigo_doc)
                     OR LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), u.cedula))) =
                        LTRIM(RTRIM(TRY_CONVERT(nvarchar(100), d.cedula_doc)))
                  )
                ORDER BY
                    CASE WHEN TRY_CONVERT(int, u.Codigo_Usuario) = TRY_CONVERT(int, d.codigo_doc) THEN 0 ELSE 1 END,
                    TRY_CONVERT(int, u.Codigo_Usuario)
            ) usuario
        ),
        filtrados AS (
            SELECT *
            FROM docentes
            WHERE (? IS NULL OR estado_codigo = UPPER(?))
              AND (? IS NULL OR LOWER(genero) LIKE LOWER(?))
              AND (
                    ? IS NULL
                 OR docente LIKE ?
                 OR cedula LIKE ?
                 OR correo LIKE ?
              )
        )
        SELECT TOP ({limit})
            genero,
            SUM(CASE WHEN estado_codigo = 'A' THEN 1 ELSE 0 END) AS activos,
            SUM(CASE WHEN estado_codigo = 'P' THEN 1 ELSE 0 END) AS inactivos,
            SUM(CASE WHEN estado_codigo NOT IN ('A', 'P') THEN 1 ELSE 0 END) AS sin_estado,
            COUNT(*) AS total
        FROM filtrados
        GROUP BY genero
        ORDER BY
            CASE genero WHEN N'Femenino' THEN 1 WHEN N'Masculino' THEN 2 ELSE 3 END
    """
    buscar = params["buscar"]
    return sql, [
        params["estado"], params["estado"],
        params["genero"], params["genero"],
        buscar, buscar, buscar, buscar,
    ]


def _periodo_query(limit: int, params: dict[str, str | None]) -> tuple[str, list[Any]]:
    cte, values = _student_report_cte(params)
    sql = f"""
        {cte}
        SELECT TOP ({limit})
            anio,
            periodo_codigo,
            periodo,
            {_split_count_columns()}
        FROM base
        GROUP BY anio, periodo_codigo, periodo
        ORDER BY anio DESC, TRY_CONVERT(int, periodo_codigo) DESC
    """
    return sql, values


QUERY_BUILDERS = {
    "provincia": _provincia_query,
    "provincia_genero": _provincia_genero_query,
    "provincia_carrera": _provincia_carrera_query,
    "carrera": _carrera_query,
    "graduados_2025": _graduados_2025_query,
    "genero": _genero_query,
    "genero_docentes": _genero_docentes_query,
    "periodo": _periodo_query,
}


def _execute_report(
    report_key: str,
    *,
    periodo: str | None,
    carrera: str | None,
    estado: str | None,
    buscar: str | None,
    anio: str | None,
    genero: str | None,
    limit: int,
) -> dict[str, Any]:
    if report_key not in REPORTS:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reporte no disponible")

    safe_limit = _limit(limit)
    effective_estado = _default_estado_for_report(report_key, estado)
    filters = _base_params(periodo, carrera, effective_estado, buscar, anio, genero)
    sql, sql_params = QUERY_BUILDERS[report_key](safe_limit, filters)

    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(sql, sql_params)
            columns, rows = _rows_from_cursor(cursor)
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"No se pudo generar el reporte {report_key}",
        ) from exc

    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": (
            "dbo.DATOSDOCENTE + dbo.USUARIOS"
            if report_key == "genero_docentes"
            else "dbo.TOTALESTUDMATRICCNE"
            if report_key in {"provincia", "provincia_genero", "provincia_carrera", "carrera", "graduados_2025", "genero", "periodo"}
            and not filters["anio"]
            and not filters["periodo"]
            else "historico por CARRERAXESTUD y PERIODO"
        ),
        "report": _report_payload(report_key),
        "columns": columns,
        "rows": rows,
        "total": len(rows),
        "criteria": {
            "periodo": filters["periodo"],
            "carrera": filters["carrera"],
            "estado": filters["estado"],
            "anio": filters["anio"],
            "genero": _clean(genero),
            "buscar": _clean(buscar),
            "limit": safe_limit,
        },
    }


def _catalog_options() -> dict[str, list[dict[str, str]]]:
    options: dict[str, list[dict[str, str]]] = {"periodos": [], "carreras": [], "anios": []}
    try:
        with get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    CAST(anio_valor AS varchar(10)) AS value,
                    CAST(anio_valor AS varchar(10)) AS label
                FROM (
                    SELECT DISTINCT TRY_CONVERT(int, anio) AS anio_valor
                    FROM dbo.PERIODO p
                    INNER JOIN dbo.CARRERAXESTUD ce ON ce.codigo_periodo = p.cod_periodo
                    WHERE TRY_CONVERT(int, p.anio) IS NOT NULL
                ) years
                ORDER BY anio_valor DESC
                """
            )
            options["anios"] = [{"value": str(row[0]), "label": str(row[1])} for row in cursor.fetchall()]

            cursor.execute(
                """
                SELECT TOP (500)
                    CAST(cod_periodo AS varchar(30)) AS value,
                    CONCAT(CAST(cod_periodo AS varchar(30)), ' - ', LTRIM(RTRIM(Detalle_Periodo))) AS label
                FROM dbo.PERIODO
                WHERE TRY_CONVERT(int, anio) IS NOT NULL
                ORDER BY cod_periodo DESC
                """
            )
            options["periodos"] = [{"value": str(row[0]), "label": str(row[1])} for row in cursor.fetchall()]

            cursor.execute(
                """
                SELECT
                    CAST(Cod_AnioBasica AS varchar(30)) AS value,
                    CONCAT(CAST(Cod_AnioBasica AS varchar(30)), ' - ', LTRIM(RTRIM(Nombre_Basica))) AS label
                FROM dbo.CARRERAS
                ORDER BY LTRIM(RTRIM(Nombre_Basica))
                """
            )
            options["carreras"] = [{"value": str(row[0]), "label": str(row[1])} for row in cursor.fetchall()]
    except Exception:
        return options
    return options


@router.get("/catalog")
def catalog(_: SessionUser = AllowedUser) -> dict[str, Any]:
    options = _catalog_options()
    return {
        "reports": [_report_payload(key) for key in REPORTS],
        "functional_inventory": FUNCTIONAL_INVENTORY,
        "periodos": options["periodos"],
        "carreras": options["carreras"],
        "anios": options["anios"],
    }


def _modernized_legacy_reports_catalog() -> dict[str, Any]:
    totals = {
        "total": len(CRYSTAL_REPORTS),
        "modernizado": sum(1 for report in CRYSTAL_REPORTS if report["migration_status"] == "modernizado"),
        "base": sum(1 for report in CRYSTAL_REPORTS if report["migration_status"] == "base"),
        "pendiente": sum(1 for report in CRYSTAL_REPORTS if report["migration_status"] == "pendiente"),
    }
    return {
        "project": "SisAcademicoV1",
        "source_engine": "Crystal Reports solo como referencia historica",
        "target_engine": "FastAPI + SQL Server + reportlab/openpyxl",
        "strategy": "No se ejecutan .rpt. Cada reporte heredado se convierte en dataset SQL parametrizado y salida PDF/Excel/HTML desde backend.",
        "totals": totals,
        "reports": [_modernized_legacy_report_payload(report) for report in CRYSTAL_REPORTS],
    }


@router.get("/modern-catalog")
def modernized_legacy_reports_catalog(_: SessionUser = AllowedUser) -> dict[str, Any]:
    return _modernized_legacy_reports_catalog()


@router.get("/modern-catalog/{report_key}")
def modernized_legacy_report_detail(report_key: str, _: SessionUser = AllowedUser) -> dict[str, Any]:
    for report in CRYSTAL_REPORTS:
        if report["key"] == report_key:
            return _modernized_legacy_report_payload(report)
    raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Reporte heredado no inventariado")


@router.get("/crystal-catalog")
def crystal_catalog(_: SessionUser = AllowedUser) -> dict[str, Any]:
    payload = _modernized_legacy_reports_catalog()
    payload["deprecated"] = True
    payload["replacement_endpoint"] = "/api/students/reporteria-integral/modern-catalog"
    return payload


@router.get("/crystal-catalog/{report_key}")
def crystal_report_detail(report_key: str, _: SessionUser = AllowedUser) -> dict[str, Any]:
    payload = modernized_legacy_report_detail(report_key, _)
    payload["deprecated"] = True
    payload["replacement_endpoint"] = f"/api/students/reporteria-integral/modern-catalog/{report_key}"
    return payload


@router.get("")
def run_report(
    response: Response,
    report_key: str = Query(default="provincia_genero"),
    periodo: str | None = Query(default=None),
    carrera: str | None = Query(default=None),
    estado: str | None = Query(default=None),
    buscar: str | None = Query(default=None),
    anio: str | None = Query(default=None),
    genero: str | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=10000),
    _: SessionUser = AllowedUser,
) -> dict[str, Any]:
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return _execute_report(
        report_key,
        periodo=periodo,
        carrera=carrera,
        estado=estado,
        buscar=buscar,
        anio=anio,
        genero=genero,
        limit=limit,
    )


def _write_key_values(sheet: Any, rows: list[tuple[str, Any]]) -> None:
    for label, value in rows:
        sheet.append([label, value])
    for row in sheet.iter_rows():
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill("solid", fgColor="E8F1F8")


def _autosize(sheet: Any) -> None:
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 12), 48)


def _is_total_column(column: str) -> bool:
    normalized = column.lower()
    if "codigo" in normalized or "cedula" in normalized or normalized == "anio":
        return False
    return (
        normalized in {"regular", "homologacion", "total", "cantidad", "graduados"}
        or normalized.startswith("total_")
        or normalized.endswith("_total")
        or "total_estudiantes" in normalized
    )


def _numeric_value(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value or "0").replace(",", "."))
    except ValueError:
        return 0.0


def _build_workbook(payload: dict[str, Any]) -> BytesIO:
    workbook = Workbook()
    ws_data = workbook.active
    ws_data.title = "Datos"

    report = payload["report"]
    columns = payload["columns"]
    rows = payload["rows"]

    ws_data.append(columns)
    for cell in ws_data[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F5C83")
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws_data.append([row.get(column) for column in columns])
    total_columns = [column for column in columns if _is_total_column(column)]
    if rows and total_columns:
        total_row: list[Any] = []
        for index, column in enumerate(columns):
            if index == 0:
                total_row.append("Total")
            elif column in total_columns:
                total_row.append(sum(_numeric_value(row.get(column)) for row in rows))
            else:
                total_row.append("")
        ws_data.append(total_row)
        for cell in ws_data[ws_data.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF4F6")
    _autosize(ws_data)

    ws_meta = workbook.create_sheet("Reporte")
    _write_key_values(
        ws_meta,
        [
            ("Reporte", report["title"]),
            ("Clave", report["key"]),
            ("Categoria", report["category"]),
            ("Descripcion", report["description"]),
            ("Generado", payload["generated_at"]),
            ("Total filas", payload["total"]),
        ],
    )
    _autosize(ws_meta)

    ws_filters = workbook.create_sheet("Filtros")
    _write_key_values(ws_filters, [(key, value or "") for key, value in payload["criteria"].items()])
    _autosize(ws_filters)

    ws_sources = workbook.create_sheet("Fuentes")
    ws_sources.append(["Tabla o vista"])
    ws_sources["A1"].font = Font(bold=True, color="FFFFFF")
    ws_sources["A1"].fill = PatternFill("solid", fgColor="1F5C83")
    for source in report["source_tables"]:
        ws_sources.append([source])
    _autosize(ws_sources)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/export")
def export_report(
    report_key: str = Query(default="provincia_genero"),
    periodo: str | None = Query(default=None),
    carrera: str | None = Query(default=None),
    estado: str | None = Query(default=None),
    buscar: str | None = Query(default=None),
    anio: str | None = Query(default=None),
    genero: str | None = Query(default=None),
    limit: int = Query(default=5000, ge=1, le=10000),
    _: SessionUser = AllowedUser,
) -> StreamingResponse:
    payload = _execute_report(
        report_key,
        periodo=periodo,
        carrera=carrera,
        estado=estado,
        buscar=buscar,
        anio=anio,
        genero=genero,
        limit=limit,
    )
    buffer = _build_workbook(payload)
    filename = f"reporteria-integral-{report_key}-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{report_key}")
def run_individual_report(
    report_key: str,
    periodo: str | None = Query(default=None),
    carrera: str | None = Query(default=None),
    estado: str | None = Query(default=None),
    buscar: str | None = Query(default=None),
    anio: str | None = Query(default=None),
    genero: str | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=10000),
    _: SessionUser = AllowedUser,
) -> dict[str, Any]:
    return _execute_report(
        report_key,
        periodo=periodo,
        carrera=carrera,
        estado=estado,
        buscar=buscar,
        anio=anio,
        genero=genero,
        limit=limit,
    )


@router.get("/{report_key}/export")
def export_individual_report(
    report_key: str,
    periodo: str | None = Query(default=None),
    carrera: str | None = Query(default=None),
    estado: str | None = Query(default=None),
    buscar: str | None = Query(default=None),
    anio: str | None = Query(default=None),
    genero: str | None = Query(default=None),
    limit: int = Query(default=5000, ge=1, le=10000),
    _: SessionUser = AllowedUser,
) -> StreamingResponse:
    payload = _execute_report(
        report_key,
        periodo=periodo,
        carrera=carrera,
        estado=estado,
        buscar=buscar,
        anio=anio,
        genero=genero,
        limit=limit,
    )
    buffer = _build_workbook(payload)
    filename = f"reporte-{report_key}-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
