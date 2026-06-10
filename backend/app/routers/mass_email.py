from __future__ import annotations

import base64
from datetime import date, datetime
from html import escape
from io import BytesIO
import json
from pathlib import Path
import re
import unicodedata
from typing import Annotated, Any
from urllib.parse import quote

import httpx
from openpyxl import load_workbook
import pyodbc
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel, Field

from app.core.config import get_settings
from app.core.security import SessionUser, require_roles
from app.routers.certificate_renamer import _cedula_candidates, _extract_pdf_text
from app.services.db import get_connection
from app.services.graph import graph_post

router = APIRouter(prefix="/api/admin/correos-masivos", tags=["correos-masivos"])

MassEmailAccess = Depends(
    require_roles(
        "ADMINISTRADOR",
        "ACADEMICO",
        "ADMISIONES",
        "BIENESTAR",
        "FINANCIERO",
        "RECTOR",
        "VICERRECTOR",
        "SOPORTE",
    )
)

MAX_CEDULAS = 1000
MAX_RECIPIENTS_PER_SEND = 300
MAX_ATTACHMENT_BYTES = 3 * 1024 * 1024
MAX_EXCEL_BYTES = 12 * 1024 * 1024
MAX_EXCEL_ROWS = 5000
_PROJECT_ROOT = Path(__file__).resolve().parents[3]
_INTEC_LOGO_PATH = _PROJECT_ROOT / "frontend" / "public" / "Intec-Logowithslogangray.svg"
_INTEC_LOGO_CONTENT_ID = "intec-logo-reporteria@intec.edu.ec"
_INTEC_LOGO_FILENAME = "intec-logo-reporteria.png"
_INTEC_LOGO_CACHE: dict[str, Any] = {}
_INTEC_RED = "#931913"
_INTEC_BLUE = "#8DBBC7"
_INTEC_WHITE = "#ffffff"
_INTEC_GRAY = "#C7C6C6"
_INTEC_DARK_GRAY = "#777777"
_INTEC_NAME = "Instituto Superior Tecnológico de Técnicas Empresariales y del Conocimiento INTEC"
_EXCEL_COLUMN_ALIASES = {
    "cedula": {
        "cedula",
        "cedula_est",
        "cedula_estudiante",
        "numero_cedula",
        "num_cedula",
        "identificacion",
        "identificacion_estudiante",
        "numero_identificacion",
        "documento",
        "dni",
        "ci",
        "cc",
    },
    "nombre": {
        "nombre",
        "nombres",
        "estudiante",
        "apellidos_nombre",
        "apellidos_nombres",
        "alumno",
        "participante",
        "nombre_completo",
        "beneficiario",
    },
    "correo": {
        "correo",
        "email",
        "mail",
        "correo_personal",
        "correo_electronico",
        "email_personal",
    },
    "documento": {
        "documento",
        "archivo",
        "nombre_archivo",
        "certificado",
        "pdf",
        "adjunto",
        "ruta",
        "ruta_documento",
        "documento_personal",
    },
    "carrera": {
        "carrera",
        "programa",
        "nombre_carrera",
        "curso",
        "materia",
    },
    "periodo": {
        "periodo",
        "periodo_academico",
        "detalle_periodo",
        "periodo_matricula",
    },
    "referencia": {
        "referencia",
        "observacion",
        "observaciones",
        "detalle",
        "comentario",
        "nota",
    },
}


class MassEmailResolvePayload(BaseModel):
    cedulas: str | list[str] = Field(default_factory=list)
    include_personal: bool = True
    include_intec: bool = True
    include_docentes: bool = True
    include_administrativos: bool = True


class MassEmailRecipient(BaseModel):
    id: str
    cedula: str
    email: str
    nombres: str | None = None
    codigo: str | None = None
    login: str | None = None
    tipo_usuario: str | None = None
    email_tipo: str | None = None
    source_table: str | None = None


class MassEmailSendResult(BaseModel):
    ok: bool
    sent: int
    failed: int
    skipped: int = 0
    skipped_attachments: int = 0
    attachment_count: int = 0
    send_mode: str = "individual"
    recipients: list[dict[str, Any]] = Field(default_factory=list)
    message: str | None = None


class MassEmailExcelRow(BaseModel):
    excel_row: int
    cedula: str = ""
    nombre_excel: str = ""
    correo_excel: str = ""
    documento: str = ""
    carrera: str = ""
    periodo: str = ""
    referencia: str = ""
    estado: str
    motivo: str = ""
    destinatarios: int = 0
    raw: dict[str, Any] = Field(default_factory=dict)


class MassEmailExcelResponse(BaseModel):
    filename: str
    sheet: str
    columns: list[str] = Field(default_factory=list)
    detected_columns: dict[str, str | None] = Field(default_factory=dict)
    rows: list[MassEmailExcelRow] = Field(default_factory=list)
    items: list[dict[str, Any]] = Field(default_factory=list)
    not_found: list[str] = Field(default_factory=list)
    sources: dict[str, int] = Field(default_factory=dict)
    summary: dict[str, Any] = Field(default_factory=dict)
    warnings: list[str] = Field(default_factory=list)
    graph_mail_sender: str | None = None


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _excel_clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _slug(value: Any) -> str:
    text = unicodedata.normalize("NFD", _excel_clean(value).lower())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _excel_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _normalize_excel_cedula(value: Any) -> str:
    digits = re.sub(r"\D+", "", _excel_clean(value))
    if not digits:
        return ""
    if len(digits) == 13 and digits.endswith("001"):
        digits = digits[:10]
    elif len(digits) > 10:
        match = re.search(r"\d{10}", digits)
        digits = match.group(0) if match else digits[-10:]
    return digits.zfill(10) if len(digits) < 10 else digits


def _find_excel_header_row(rows: list[tuple[Any, ...]]) -> int:
    for index, row in enumerate(rows[:12]):
        non_empty = [_excel_clean(cell) for cell in row if _excel_clean(cell)]
        if len(non_empty) < 2:
            continue
        slugs = {_slug(cell) for cell in non_empty}
        alias_hits = sum(
            1
            for aliases in _EXCEL_COLUMN_ALIASES.values()
            if any(slug in aliases for slug in slugs)
        )
        if alias_hits >= 1:
            return index
    for index, row in enumerate(rows[:12]):
        if len([cell for cell in row if _excel_clean(cell)]) >= 2:
            return index
    return 0


def _unique_excel_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        raw = _excel_clean(value) or f"columna_{index + 1}"
        base = _slug(raw) or f"columna_{index + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _detect_excel_columns(headers: list[str]) -> dict[str, str | None]:
    detected: dict[str, str | None] = {key: None for key in _EXCEL_COLUMN_ALIASES}
    for header in headers:
        normalized = _slug(header)
        for field_name, aliases in _EXCEL_COLUMN_ALIASES.items():
            if detected[field_name]:
                continue
            if normalized in aliases:
                detected[field_name] = header
    return detected


def _read_mass_email_excel(content: bytes, filename: str) -> tuple[str, list[str], list[dict[str, Any]], list[str]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix not in {"xlsx", "xlsm"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Sube un archivo Excel .xlsx o .xlsm.")
    if len(content) > MAX_EXCEL_BYTES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel supera el máximo permitido de 12 MB.")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No se pudo leer el Excel.") from exc

    sheet = workbook.active
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel no contiene filas.")

    header_index = _find_excel_header_row(raw_rows)
    headers = _unique_excel_headers(list(raw_rows[header_index]))
    data_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    for row_number, row in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
        values = list(row)
        if not any(_excel_clean(value) for value in values):
            continue
        item = {
            headers[index]: _excel_value(values[index]) if index < len(values) else None
            for index in range(len(headers))
        }
        item["_excel_row"] = row_number
        data_rows.append(item)
        if len(data_rows) >= MAX_EXCEL_ROWS:
            warnings.append(f"Se procesaron las primeras {MAX_EXCEL_ROWS} filas con información.")
            break

    if not data_rows:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El Excel no contiene datos después de la cabecera.")
    return sheet.title, headers, data_rows, warnings


def _valid_email(value: str | None) -> str:
    email = _clean(value)
    if not email or not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email):
        return ""
    return email


def _parse_emails(raw: str | list[str]) -> list[str]:
    values = raw if isinstance(raw, list) else re.split(r"[\s,;]+", raw)
    emails: list[str] = []
    seen: set[str] = set()
    for value in values:
        email = _valid_email(value)
        key = email.lower()
        if not email or key in seen:
            continue
        seen.add(key)
        emails.append(email)
    return emails


def _parse_attachment_assignments(raw: str) -> dict[str, str]:
    if not _clean(raw):
        return {}

    try:
        payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Asignación de adjuntos inválida.",
        ) from exc

    if not isinstance(payload, dict):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Asignación de adjuntos inválida.",
        )

    assignments: dict[str, str] = {}
    for filename, cedula in payload.items():
        normalized_filename = _clean(filename).lower()
        normalized_cedula = _clean(cedula)
        if normalized_filename and normalized_cedula:
            assignments[normalized_filename] = normalized_cedula
    return assignments


def _parse_cedulas(raw: str | list[str]) -> list[str]:
    if isinstance(raw, list):
        values = raw
    else:
        values = re.split(r"[\s,;]+", raw)

    cedulas: list[str] = []
    seen: set[str] = set()
    for value in values:
        cedula = _clean(value)
        if not cedula or cedula in seen:
            continue
        seen.add(cedula)
        cedulas.append(cedula)

    if len(cedulas) > MAX_CEDULAS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se permite consultar hasta {MAX_CEDULAS} cédulas por solicitud.",
        )
    return cedulas


def _chunks(values: list[str], size: int) -> list[list[str]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def _cedula_cte(cedulas: list[str]) -> str:
    if not cedulas:
        return ""
    parts = ["SELECT CAST(? AS VARCHAR(50)) AS cedula"]
    parts.extend("UNION ALL SELECT CAST(? AS VARCHAR(50))" for _ in cedulas[1:])
    return "WITH input_cedulas(cedula) AS (\n" + "\n".join(parts) + "\n)"


def _source_join(current: str | None, source: str) -> str:
    sources = [_clean(item) for item in _clean(current).split("/") if _clean(item)]
    if source not in sources:
        sources.append(source)
    return "/".join(sources)


def _add_recipient(
    recipients: list[dict[str, Any]],
    index_by_key: dict[tuple[str, str], int],
    *,
    cedula: str,
    email: str | None,
    nombres: str | None,
    codigo: str | None,
    login: str | None,
    tipo_usuario: str,
    email_tipo: str,
    source_table: str,
) -> None:
    normalized_email = _valid_email(email)
    normalized_cedula = _clean(cedula)
    if not normalized_email or not normalized_cedula:
        return

    key = (normalized_cedula, normalized_email.lower())
    if key in index_by_key:
        item = recipients[index_by_key[key]]
        item["source_table"] = _source_join(item.get("source_table"), source_table)
        return

    item = {
        "id": f"{normalized_cedula}:{normalized_email.lower()}",
        "cedula": normalized_cedula,
        "email": normalized_email,
        "nombres": _clean(nombres),
        "codigo": _clean(codigo),
        "login": _clean(login),
        "tipo_usuario": tipo_usuario,
        "email_tipo": email_tipo,
        "source_table": source_table,
    }
    index_by_key[key] = len(recipients)
    recipients.append(item)


def _rows(cursor: pyodbc.Cursor, sql: str, params: list[str]) -> list[Any]:
    cursor.execute(sql, params)
    return list(cursor.fetchall())


def _search_recipients(query: str, limit: int) -> dict[str, Any]:
    normalized_query = _clean(query)
    if len(normalized_query) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ingresa al menos 2 caracteres para buscar.")

    safe_limit = max(1, min(limit, 200))
    pattern = f"%{normalized_query}%"
    recipients: list[dict[str, Any]] = []
    index_by_key: dict[tuple[str, str], int] = {}

    try:
        with get_connection() as conn:
            cursor = conn.cursor()

            student_sql = f"""
            SELECT TOP ({safe_limit})
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), d.Cedula_Est))) AS cedula,
                TRY_CONVERT(VARCHAR(50), d.codigo_estud) AS codigo,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), d.Apellidos_nombre))) AS nombres,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correo))) AS datos_correo_personal,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correointec))) AS datos_correo_intec,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), ce.CorreoPersonal))) AS correo_personal,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), ce.CorreoIntec))) AS correo_intec
            FROM dbo.DATOS_ESTUD d
            LEFT JOIN dbo.CorreosEstudIntec ce
                ON ce.codestud = d.codigo_estud
            WHERE
                TRY_CONVERT(VARCHAR(50), d.Cedula_Est) LIKE ?
                OR TRY_CONVERT(NVARCHAR(250), d.Apellidos_nombre) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), d.correo) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), d.correointec) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), ce.CorreoPersonal) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), ce.CorreoIntec) LIKE ?
            ORDER BY d.Apellidos_nombre
            """
            for row in _rows(cursor, student_sql, [pattern] * 6):
                _add_recipient(
                    recipients,
                    index_by_key,
                    cedula=row.cedula,
                    email=_clean(row.correo_intec) or _clean(row.datos_correo_intec),
                    nombres=row.nombres,
                    codigo=row.codigo,
                    login=None,
                    tipo_usuario="ESTUDIANTE",
                    email_tipo="INTEC",
                    source_table="DATOS_ESTUD/CorreosEstudIntec",
                )
                _add_recipient(
                    recipients,
                    index_by_key,
                    cedula=row.cedula,
                    email=_clean(row.correo_personal) or _clean(row.datos_correo_personal),
                    nombres=row.nombres,
                    codigo=row.codigo,
                    login=None,
                    tipo_usuario="ESTUDIANTE",
                    email_tipo="PERSONAL",
                    source_table="DATOS_ESTUD/CorreosEstudIntec",
                )

            teacher_sql = f"""
            SELECT TOP ({safe_limit})
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), dd.cedula_doc))) AS cedula,
                TRY_CONVERT(VARCHAR(50), dd.codigo_doc) AS codigo,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), dd.apellidos_nombre))) AS nombres,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correo))) AS correo_docente,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correop))) AS correo_personal,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), u.login))) AS login
            FROM dbo.DATOSDOCENTE dd
            LEFT JOIN dbo.USUARIOS u
                ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), u.cedula))) = LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), dd.cedula_doc)))
            WHERE
                TRY_CONVERT(VARCHAR(50), dd.cedula_doc) LIKE ?
                OR TRY_CONVERT(NVARCHAR(250), dd.apellidos_nombre) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), dd.correo) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), dd.correop) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), u.login) LIKE ?
            ORDER BY dd.apellidos_nombre
            """
            for row in _rows(cursor, teacher_sql, [pattern] * 5):
                email = _clean(row.correo_personal) or _clean(row.correo_docente)
                if not _valid_email(email) and "@" in _clean(row.login):
                    email = _clean(row.login)
                _add_recipient(
                    recipients,
                    index_by_key,
                    cedula=row.cedula,
                    email=email,
                    nombres=row.nombres,
                    codigo=row.codigo,
                    login=row.login,
                    tipo_usuario="DOCENTE",
                    email_tipo="DOCENTE",
                    source_table="DATOSDOCENTE/USUARIOS",
                )

            users_sql = f"""
            SELECT TOP ({safe_limit})
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), u.cedula))) AS cedula,
                TRY_CONVERT(VARCHAR(50), u.Codigo_Usuario) AS codigo,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), u.login))) AS login,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), tu.detalle_tipo_us))) AS tipo_usuario,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), us.nombres))) AS nombres_sis,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.email))) AS email_sis,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), d.Apellidos_nombre))) AS nombres_estud,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correo))) AS correo_estud,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correointec))) AS correo_intec,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), dd.apellidos_nombre))) AS nombres_doc,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correo))) AS correo_doc,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correop))) AS correop_doc
            FROM dbo.USUARIOS u
            LEFT JOIN dbo.TIPO_USUARIO tu
                ON u.tipo_usuario = tu.Codigo_tipo_us
            LEFT JOIN dbo.USUARIO_SIS us
                ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.login))) = LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), u.login)))
            LEFT JOIN dbo.DATOS_ESTUD d
                ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), d.Cedula_Est))) = LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), u.cedula)))
            LEFT JOIN dbo.DATOSDOCENTE dd
                ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), dd.cedula_doc))) = LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), u.cedula)))
            WHERE
                TRY_CONVERT(VARCHAR(50), u.cedula) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), u.login) LIKE ?
                OR TRY_CONVERT(NVARCHAR(250), us.nombres) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), us.email) LIKE ?
                OR TRY_CONVERT(NVARCHAR(250), d.Apellidos_nombre) LIKE ?
                OR TRY_CONVERT(NVARCHAR(250), dd.apellidos_nombre) LIKE ?
            ORDER BY COALESCE(us.nombres, d.Apellidos_nombre, dd.apellidos_nombre, u.login)
            """
            for row in _rows(cursor, users_sql, [pattern] * 6):
                email = (
                    _clean(row.email_sis)
                    or _clean(row.correo_intec)
                    or _clean(row.correop_doc)
                    or _clean(row.correo_doc)
                    or _clean(row.correo_estud)
                )
                if not _valid_email(email) and "@" in _clean(row.login):
                    email = _clean(row.login)
                _add_recipient(
                    recipients,
                    index_by_key,
                    cedula=row.cedula,
                    email=email,
                    nombres=_clean(row.nombres_sis) or _clean(row.nombres_doc) or _clean(row.nombres_estud),
                    codigo=row.codigo,
                    login=row.login,
                    tipo_usuario=_clean(row.tipo_usuario) or "USUARIO",
                    email_tipo="USUARIO",
                    source_table="USUARIOS/USUARIO_SIS",
                )

            sis_sql = f"""
            SELECT TOP ({safe_limit})
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.login))) AS cedula,
                TRY_CONVERT(VARCHAR(50), us.id_usuarios) AS codigo,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.login))) AS login,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), us.nombres))) AS nombres,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.email))) AS email
            FROM dbo.USUARIO_SIS us
            WHERE
                TRY_CONVERT(VARCHAR(150), us.login) LIKE ?
                OR TRY_CONVERT(NVARCHAR(250), us.nombres) LIKE ?
                OR TRY_CONVERT(VARCHAR(150), us.email) LIKE ?
            ORDER BY us.nombres
            """
            for row in _rows(cursor, sis_sql, [pattern] * 3):
                _add_recipient(
                    recipients,
                    index_by_key,
                    cedula=row.cedula,
                    email=row.email,
                    nombres=row.nombres,
                    codigo=row.codigo,
                    login=row.login,
                    tipo_usuario="ADMINISTRATIVO",
                    email_tipo="ADMINISTRATIVO",
                    source_table="USUARIO_SIS",
                )
    except pyodbc.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo buscar usuarios en SQL Server",
        ) from exc

    return {
        "query": normalized_query,
        "items": recipients[:safe_limit],
        "total": len(recipients[:safe_limit]),
        "graph_mail_sender": get_settings().graph_mail_sender,
    }


def _resolve_recipients(payload: MassEmailResolvePayload) -> dict[str, Any]:
    cedulas = _parse_cedulas(payload.cedulas)
    recipients: list[dict[str, Any]] = []
    index_by_key: dict[tuple[str, str], int] = {}

    if not cedulas:
        return {
            "cedulas": [],
            "items": [],
            "total": 0,
            "not_found": [],
            "sources": {},
            "graph_mail_sender": get_settings().graph_mail_sender,
        }

    cte = _cedula_cte(cedulas)

    try:
        with get_connection() as conn:
            cursor = conn.cursor()

            student_sql = f"""
            {cte}
            SELECT
                c.cedula,
                TRY_CONVERT(VARCHAR(50), d.codigo_estud) AS codigo,
                LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), d.Apellidos_nombre))) AS nombres,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correo))) AS datos_correo_personal,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correointec))) AS datos_correo_intec,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), ce.CorreoPersonal))) AS correo_personal,
                LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), ce.CorreoIntec))) AS correo_intec
            FROM input_cedulas c
            INNER JOIN dbo.DATOS_ESTUD d
                ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), d.Cedula_Est))) = c.cedula
            LEFT JOIN dbo.CorreosEstudIntec ce
                ON ce.codestud = d.codigo_estud
            """
            for row in _rows(cursor, student_sql, cedulas):
                if payload.include_intec:
                    _add_recipient(
                        recipients,
                        index_by_key,
                        cedula=row.cedula,
                        email=_clean(row.correo_intec) or _clean(row.datos_correo_intec),
                        nombres=row.nombres,
                        codigo=row.codigo,
                        login=None,
                        tipo_usuario="ESTUDIANTE",
                        email_tipo="INTEC",
                        source_table="DATOS_ESTUD/CorreosEstudIntec",
                    )
                if payload.include_personal:
                    _add_recipient(
                        recipients,
                        index_by_key,
                        cedula=row.cedula,
                        email=_clean(row.correo_personal) or _clean(row.datos_correo_personal),
                        nombres=row.nombres,
                        codigo=row.codigo,
                        login=None,
                        tipo_usuario="ESTUDIANTE",
                        email_tipo="PERSONAL",
                        source_table="DATOS_ESTUD/CorreosEstudIntec",
                    )

            if payload.include_docentes:
                teacher_sql = f"""
                {cte}
                SELECT
                    c.cedula,
                    TRY_CONVERT(VARCHAR(50), dd.codigo_doc) AS codigo,
                    LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), dd.apellidos_nombre))) AS nombres,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correo))) AS correo_docente,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correop))) AS correo_personal,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), u.login))) AS login
                FROM input_cedulas c
                INNER JOIN dbo.DATOSDOCENTE dd
                    ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), dd.cedula_doc))) = c.cedula
                LEFT JOIN dbo.USUARIOS u
                    ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), u.cedula))) = c.cedula
                """
                for row in _rows(cursor, teacher_sql, cedulas):
                    email = _clean(row.correo_personal) or _clean(row.correo_docente)
                    if not _valid_email(email) and "@" in _clean(row.login):
                        email = _clean(row.login)
                    _add_recipient(
                        recipients,
                        index_by_key,
                        cedula=row.cedula,
                        email=email,
                        nombres=row.nombres,
                        codigo=row.codigo,
                        login=row.login,
                        tipo_usuario="DOCENTE",
                        email_tipo="DOCENTE",
                        source_table="DATOSDOCENTE/USUARIOS",
                    )

            if payload.include_administrativos:
                users_sql = f"""
                {cte}
                SELECT
                    c.cedula,
                    TRY_CONVERT(VARCHAR(50), u.Codigo_Usuario) AS codigo,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), u.login))) AS login,
                    LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), tu.detalle_tipo_us))) AS tipo_usuario,
                    LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), us.nombres))) AS nombres_sis,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.email))) AS email_sis,
                    LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), d.Apellidos_nombre))) AS nombres_estud,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correo))) AS correo_estud,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), d.correointec))) AS correo_intec,
                    LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), dd.apellidos_nombre))) AS nombres_doc,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correo))) AS correo_doc,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), dd.correop))) AS correop_doc
                FROM input_cedulas c
                INNER JOIN dbo.USUARIOS u
                    ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), u.cedula))) = c.cedula
                LEFT JOIN dbo.TIPO_USUARIO tu
                    ON u.tipo_usuario = tu.Codigo_tipo_us
                LEFT JOIN dbo.USUARIO_SIS us
                    ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.login))) = LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), u.login)))
                LEFT JOIN dbo.DATOS_ESTUD d
                    ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), d.Cedula_Est))) = c.cedula
                LEFT JOIN dbo.DATOSDOCENTE dd
                    ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(50), dd.cedula_doc))) = c.cedula
                """
                for row in _rows(cursor, users_sql, cedulas):
                    email = (
                        _clean(row.email_sis)
                        or _clean(row.correo_intec)
                        or _clean(row.correop_doc)
                        or _clean(row.correo_doc)
                        or _clean(row.correo_estud)
                    )
                    if not _valid_email(email) and "@" in _clean(row.login):
                        email = _clean(row.login)
                    _add_recipient(
                        recipients,
                        index_by_key,
                        cedula=row.cedula,
                        email=email,
                        nombres=_clean(row.nombres_sis) or _clean(row.nombres_doc) or _clean(row.nombres_estud),
                        codigo=row.codigo,
                        login=row.login,
                        tipo_usuario=_clean(row.tipo_usuario) or "USUARIO",
                        email_tipo="USUARIO",
                        source_table="USUARIOS/USUARIO_SIS",
                    )

                sis_sql = f"""
                {cte}
                SELECT
                    c.cedula,
                    TRY_CONVERT(VARCHAR(50), us.id_usuarios) AS codigo,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.login))) AS login,
                    LTRIM(RTRIM(TRY_CONVERT(NVARCHAR(250), us.nombres))) AS nombres,
                    LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.email))) AS email
                FROM input_cedulas c
                INNER JOIN dbo.USUARIO_SIS us
                    ON LTRIM(RTRIM(TRY_CONVERT(VARCHAR(150), us.login))) = c.cedula
                """
                for row in _rows(cursor, sis_sql, cedulas):
                    _add_recipient(
                        recipients,
                        index_by_key,
                        cedula=row.cedula,
                        email=row.email,
                        nombres=row.nombres,
                        codigo=row.codigo,
                        login=row.login,
                        tipo_usuario="ADMINISTRATIVO",
                        email_tipo="ADMINISTRATIVO",
                        source_table="USUARIO_SIS",
                    )
    except pyodbc.Error as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="No se pudo consultar destinatarios en SQL Server",
        ) from exc

    found_cedulas = {item["cedula"] for item in recipients}
    sources: dict[str, int] = {}
    for item in recipients:
        source = _clean(item.get("source_table")) or "SIN_FUENTE"
        sources[source] = sources.get(source, 0) + 1

    return {
        "cedulas": cedulas,
        "items": recipients,
        "total": len(recipients),
        "not_found": [cedula for cedula in cedulas if cedula not in found_cedulas],
        "sources": sources,
        "graph_mail_sender": get_settings().graph_mail_sender,
    }


def _graph_mail_sender() -> str:
    settings = get_settings()
    sender = _clean(settings.graph_mail_sender or settings.smtp_from)
    if not sender:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Configura GRAPH_MAIL_SENDER para enviar correos por Microsoft Graph.",
        )
    return sender


def _graph_attachment(filename: str, content_type: str, content: bytes) -> dict[str, Any]:
    return {
        "@odata.type": "#microsoft.graph.fileAttachment",
        "name": filename,
        "contentType": content_type or "application/octet-stream",
        "contentBytes": base64.b64encode(content).decode("ascii"),
    }


def _fallback_logo_png() -> bytes:
    try:
        from PIL import Image, ImageDraw, ImageFont

        image = Image.new("RGBA", (720, 230), (255, 255, 255, 255))
        draw = ImageDraw.Draw(image)
        try:
            title_font = ImageFont.truetype("arialbd.ttf", 116)
            subtitle_font = ImageFont.truetype("arialbd.ttf", 23)
            small_font = ImageFont.truetype("arial.ttf", 18)
        except OSError:
            title_font = ImageFont.load_default()
            subtitle_font = ImageFont.load_default()
            small_font = ImageFont.load_default()

        draw.rounded_rectangle((0, 0, 719, 229), radius=20, outline=_INTEC_GRAY, width=2, fill=_INTEC_WHITE)
        draw.rectangle((0, 0, 720, 12), fill=_INTEC_RED)
        draw.text((32, 30), "int", fill=_INTEC_DARK_GRAY, font=title_font)
        draw.text((242, 30), "e", fill=_INTEC_RED, font=title_font)
        draw.text((310, 30), "c", fill=_INTEC_DARK_GRAY, font=title_font)
        draw.rectangle((34, 158, 220, 166), fill=_INTEC_BLUE)
        draw.text((34, 176), "INSTITUTO SUPERIOR TECNOLÓGICO DE", fill=_INTEC_DARK_GRAY, font=subtitle_font)
        draw.text((34, 202), "TÉCNICAS EMPRESARIALES Y DEL CONOCIMIENTO", fill=_INTEC_DARK_GRAY, font=small_font)

        buffer = BytesIO()
        image.save(buffer, format="PNG")
        return buffer.getvalue()
    except Exception:
        return b""


def _logo_png_content() -> bytes | None:
    if "png_content" in _INTEC_LOGO_CACHE:
        cached = _INTEC_LOGO_CACHE["png_content"]
        return bytes(cached) if cached else None

    content: bytes | None = None
    if _INTEC_LOGO_PATH.exists():
        try:
            from reportlab.graphics import renderPM
            from svglib.svglib import svg2rlg

            drawing = svg2rlg(str(_INTEC_LOGO_PATH))
            content = renderPM.drawToString(drawing, fmt="PNG")
        except Exception:
            content = None

    if not content:
        content = _fallback_logo_png()

    _INTEC_LOGO_CACHE["png_content"] = content or None
    return content or None


def _graph_inline_logo_attachment() -> dict[str, Any] | None:
    if "attachment" in _INTEC_LOGO_CACHE:
        cached = _INTEC_LOGO_CACHE["attachment"]
        return dict(cached) if cached else None

    content = _logo_png_content()
    if not content:
        _INTEC_LOGO_CACHE["attachment"] = None
        return None

    attachment = {
        "@odata.type": "#microsoft.graph.fileAttachment",
        "name": _INTEC_LOGO_FILENAME,
        "contentType": "image/png",
        "contentBytes": base64.b64encode(content).decode("ascii"),
        "isInline": True,
        "contentId": _INTEC_LOGO_CONTENT_ID,
    }
    _INTEC_LOGO_CACHE["attachment"] = attachment
    return dict(attachment)


def _message_text_to_html(body: str) -> str:
    blocks = [block.strip() for block in re.split(r"\n\s*\n", body.strip()) if block.strip()]
    if not blocks:
        return "<p>Mensaje institucional INTEC.</p>"

    paragraphs = []
    for block in blocks:
        lines = [escape(line.strip()) for line in block.splitlines()]
        paragraphs.append(
            '<p style="margin:0 0 14px;color:#19233f !important;font-size:15px;line-height:1.58;">'
            + "<br>".join(lines)
            + "</p>"
        )
    return "\n".join(paragraphs)


def _branded_email_html(body: str, *, include_logo: bool) -> str:
    message_html = _message_text_to_html(body)
    logo_markup = (
        f'<table role="presentation" align="center" cellspacing="0" cellpadding="0" style="margin:0 auto 12px;">'
        f'<tr><td align="center" style="text-align:center;">'
        f'<img src="cid:{_INTEC_LOGO_CONTENT_ID}" alt="INTEC" align="center" '
        'width="210" style="display:block;width:210px;max-width:210px;height:auto;margin:0 auto;border:0;outline:none;text-decoration:none;background:#ffffff;" />'
        f'</td></tr></table>'
        if include_logo
        else (
            '<table role="presentation" align="center" cellspacing="0" cellpadding="0" style="margin:0 auto 12px;">'
            '<tr><td align="center" style="text-align:center;font-size:40px;font-weight:900;line-height:1;color:#777777;">'
            'int<span style="color:#931913;">e</span>c'
            '</td></tr></table>'
        )
    )
    return f"""<!doctype html>
<html lang="es">
  <head>
    <meta charset="utf-8">
    <meta name="color-scheme" content="light">
    <meta name="supported-color-schemes" content="light">
    <style>
      :root {{ color-scheme: light only; supported-color-schemes: light; }}
      body, table, td, p, div, span {{ color-scheme: light only; }}
      @media (prefers-color-scheme: dark) {{
        .intec-body {{ background:#C7C6C6 !important; }}
        .intec-card, .intec-header, .intec-content {{ background:#ffffff !important; color:#19233f !important; }}
        .intec-footer {{ background:#f7fbfc !important; color:#777777 !important; }}
      }}
    </style>
  </head>
  <body class="intec-body" bgcolor="{_INTEC_GRAY}" style="margin:0;padding:0;background-color:{_INTEC_GRAY} !important;font-family:Segoe UI,Arial,sans-serif;color:#19233f !important;color-scheme:light only;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" bgcolor="{_INTEC_GRAY}" style="background-color:{_INTEC_GRAY} !important;padding:24px 12px;">
      <tr>
        <td align="center">
          <table class="intec-card" role="presentation" width="100%" cellspacing="0" cellpadding="0" bgcolor="{_INTEC_WHITE}" style="max-width:720px;background-color:{_INTEC_WHITE} !important;border:1px solid {_INTEC_GRAY};border-radius:12px;overflow:hidden;">
            <tr>
              <td style="height:10px;background:{_INTEC_RED};font-size:0;line-height:0;">&nbsp;</td>
            </tr>
            <tr>
              <td class="intec-header" align="center" bgcolor="{_INTEC_WHITE}" style="padding:22px 26px 16px;border-bottom:1px solid {_INTEC_GRAY};background-color:{_INTEC_WHITE} !important;text-align:center;">
                {logo_markup}
                <table role="presentation" align="center" cellspacing="0" cellpadding="0" style="margin:0 auto;">
                  <tr>
                    <td align="center" style="text-align:center;font-size:12px;letter-spacing:3px;text-transform:uppercase;color:{_INTEC_RED};font-weight:800;">
                      Comunicación institucional
                    </td>
                  </tr>
                  <tr>
                    <td align="center" style="padding-top:7px;text-align:center;">
                      <table role="presentation" align="center" cellspacing="0" cellpadding="0" width="120" style="width:120px;margin:0 auto;">
                        <tr>
                          <td bgcolor="{_INTEC_BLUE}" style="height:5px;background:{_INTEC_BLUE};font-size:0;line-height:0;">&nbsp;</td>
                        </tr>
                      </table>
                    </td>
                  </tr>
                </table>
              </td>
            </tr>
            <tr>
              <td class="intec-content" bgcolor="{_INTEC_WHITE}" style="padding:24px 26px;font-size:15px;line-height:1.58;color:#19233f !important;background-color:{_INTEC_WHITE} !important;">
                {message_html}
              </td>
            </tr>
            <tr>
              <td class="intec-footer" bgcolor="#f7fbfc" style="padding:16px 26px;background-color:#f7fbfc !important;border-top:1px solid {_INTEC_GRAY};color:{_INTEC_DARK_GRAY} !important;font-size:12px;line-height:1.45;">
                {_INTEC_NAME}
              </td>
            </tr>
          </table>
        </td>
      </tr>
    </table>
  </body>
</html>"""


def _attachment_cedulas(filename: str, content: bytes) -> list[str]:
    candidates = _cedula_candidates(filename)
    if candidates:
        return candidates
    if filename.lower().endswith(".pdf"):
        try:
            return _cedula_candidates(_extract_pdf_text(content))
        except Exception:
            return []
    return []


async def _read_attachment_groups(
    files: list[UploadFile] | None,
    *,
    common_files: list[UploadFile] | None = None,
    student_files: list[UploadFile] | None = None,
    match_by_cedula: bool,
    assignments: dict[str, str] | None = None,
) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]], int, int]:
    common_attachments: list[dict[str, Any]] = []
    attachments_by_cedula: dict[str, list[dict[str, Any]]] = {}
    attachment_count = 0
    skipped_attachments = 0
    assignment_by_filename = assignments or {}

    async def register_upload(upload: UploadFile, attachment_scope: str) -> None:
        nonlocal attachment_count, skipped_attachments
        filename = _clean(upload.filename)
        if not filename:
            return
        content = await upload.read()
        size = len(content)
        if size > MAX_ATTACHMENT_BYTES:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"El archivo {filename} supera el límite individual de 3 MB.",
            )
        attachment = _graph_attachment(filename, upload.content_type or "application/octet-stream", content)
        attachment_count += 1

        if attachment_scope == "common":
            common_attachments.append(attachment)
            return

        assigned_cedula = assignment_by_filename.get(filename.lower())
        if assigned_cedula:
            attachments_by_cedula.setdefault(assigned_cedula, []).append(attachment)
            return

        if match_by_cedula:
            cedulas = _attachment_cedulas(filename, content)
            if cedulas:
                for cedula in cedulas:
                    attachments_by_cedula.setdefault(cedula, []).append(attachment)
                return

        if attachment_scope == "student":
            skipped_attachments += 1
            return

        common_attachments.append(attachment)

    for upload in common_files or []:
        await register_upload(upload, "common")

    for upload in student_files or []:
        await register_upload(upload, "student")

    for upload in files or []:
        await register_upload(upload, "legacy")

    return common_attachments, attachments_by_cedula, attachment_count, skipped_attachments


def _graph_error_message(error: Exception) -> str:
    if isinstance(error, httpx.HTTPStatusError):
        response_text = error.response.text[:800] if error.response is not None else ""
        return response_text or str(error)
    return str(error)


def _graph_email_recipient(email: str, name: str | None = None) -> dict[str, Any]:
    return {
        "emailAddress": {
            "address": email,
            "name": name or email,
        }
    }


def _send_graph_mail(
    *,
    sender: str,
    recipients: list[MassEmailRecipient],
    cc_emails: list[str],
    subject: str,
    body: str,
    attachments: list[dict[str, Any]],
    use_bcc: bool = False,
) -> None:
    recipient_key = "bccRecipients" if use_bcc else "toRecipients"
    message_attachments = list(attachments)
    logo_attachment = _graph_inline_logo_attachment()
    if logo_attachment:
        message_attachments.insert(0, logo_attachment)

    payload: dict[str, Any] = {
        "message": {
            "subject": subject,
            "body": {
                "contentType": "HTML",
                "content": _branded_email_html(body, include_logo=bool(logo_attachment)),
            },
            recipient_key: [
                _graph_email_recipient(recipient.email, recipient.nombres or recipient.email)
                for recipient in recipients
            ],
        },
        "saveToSentItems": True,
    }
    if cc_emails:
        payload["message"]["ccRecipients"] = [_graph_email_recipient(email) for email in cc_emails]
    if message_attachments:
        payload["message"]["attachments"] = message_attachments

    graph_post(
        f"https://graph.microsoft.com/v1.0/users/{quote(sender, safe='')}/sendMail",
        payload,
    )


def _dedupe_recipients(recipients: list[MassEmailRecipient]) -> list[MassEmailRecipient]:
    result: list[MassEmailRecipient] = []
    seen: set[str] = set()
    for recipient in recipients:
        key = _clean(recipient.email).lower()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(recipient)
    return result


@router.post("/resolver")
def resolve_mass_email_recipients(
    payload: MassEmailResolvePayload,
    _: SessionUser = MassEmailAccess,
) -> dict[str, Any]:
    return _resolve_recipients(payload)


@router.get("/buscar-usuarios")
def search_mass_email_users(
    query: Annotated[str, Query(min_length=2, max_length=120)],
    limit: Annotated[int, Query(ge=1, le=200)] = 50,
    _: SessionUser = MassEmailAccess,
) -> dict[str, Any]:
    return _search_recipients(query, limit)


@router.post("/excel/analizar")
async def analyze_mass_email_excel(
    file: Annotated[UploadFile, File()],
    include_personal: Annotated[bool, Form()] = True,
    include_intec: Annotated[bool, Form()] = True,
    include_docentes: Annotated[bool, Form()] = True,
    include_administrativos: Annotated[bool, Form()] = True,
    _: SessionUser = MassEmailAccess,
) -> MassEmailExcelResponse:
    filename = file.filename or "destinatarios.xlsx"
    content = await file.read()
    sheet_name, columns, excel_rows, warnings = _read_mass_email_excel(content, filename)
    detected = _detect_excel_columns(columns)

    parsed_rows: list[dict[str, Any]] = []
    cedulas: list[str] = []
    duplicate_cedulas: set[str] = set()
    seen_cedulas: set[str] = set()

    for raw in excel_rows:
        cedula_column = detected.get("cedula")
        cedula = _normalize_excel_cedula(raw.get(cedula_column)) if cedula_column else ""
        if not cedula:
            for key, value in raw.items():
                if key == "_excel_row":
                    continue
                candidate_digits = re.sub(r"\D+", "", _excel_clean(value))
                if len(candidate_digits) == 10 or (len(candidate_digits) == 13 and candidate_digits.endswith("001")):
                    cedula = _normalize_excel_cedula(value)
                    break

        if cedula:
            if cedula in seen_cedulas:
                duplicate_cedulas.add(cedula)
            seen_cedulas.add(cedula)
            cedulas.append(cedula)

        parsed_rows.append(
            {
                "excel_row": int(raw.get("_excel_row") or 0),
                "cedula": cedula,
                "nombre_excel": _excel_clean(raw.get(detected.get("nombre"))) if detected.get("nombre") else "",
                "correo_excel": _valid_email(_excel_clean(raw.get(detected.get("correo")))) if detected.get("correo") else "",
                "documento": _excel_clean(raw.get(detected.get("documento"))) if detected.get("documento") else "",
                "carrera": _excel_clean(raw.get(detected.get("carrera"))) if detected.get("carrera") else "",
                "periodo": _excel_clean(raw.get(detected.get("periodo"))) if detected.get("periodo") else "",
                "referencia": _excel_clean(raw.get(detected.get("referencia"))) if detected.get("referencia") else "",
                "raw": {key: value for key, value in raw.items() if key != "_excel_row"},
            }
        )

    resolved_items: list[dict[str, Any]] = []
    resolved_not_found: list[str] = []
    resolved_sources: dict[str, int] = {}
    graph_mail_sender = get_settings().graph_mail_sender
    for cedula_chunk in _chunks(sorted(seen_cedulas), MAX_CEDULAS):
        resolved_chunk = _resolve_recipients(
            MassEmailResolvePayload(
                cedulas=cedula_chunk,
                include_personal=include_personal,
                include_intec=include_intec,
                include_docentes=include_docentes,
                include_administrativos=include_administrativos,
            )
        )
        resolved_items.extend(resolved_chunk.get("items", []))
        resolved_not_found.extend(resolved_chunk.get("not_found", []))
        graph_mail_sender = resolved_chunk.get("graph_mail_sender") or graph_mail_sender
        for source, count in (resolved_chunk.get("sources") or {}).items():
            resolved_sources[source] = resolved_sources.get(source, 0) + count

    resolved = {
        "items": resolved_items,
        "not_found": resolved_not_found,
        "sources": resolved_sources,
        "graph_mail_sender": graph_mail_sender,
    }
    recipients_by_cedula: dict[str, list[dict[str, Any]]] = {}
    for item in resolved.get("items", []):
        recipients_by_cedula.setdefault(_clean(item.get("cedula")), []).append(item)

    response_rows: list[MassEmailExcelRow] = []
    missing_cedula = 0
    ready = 0
    without_recipient = 0
    with_document_reference = 0
    with_excel_email = 0

    for row in parsed_rows:
        cedula = row["cedula"]
        destination_count = len(recipients_by_cedula.get(cedula, [])) if cedula else 0
        if row["documento"] or row["referencia"]:
            with_document_reference += 1
        if row["correo_excel"]:
            with_excel_email += 1

        if not cedula:
            missing_cedula += 1
            estado = "SIN_CEDULA"
            motivo = "No se encontró una cédula válida en la fila."
        elif destination_count:
            ready += 1
            estado = "LISTO"
            motivo = f"{destination_count} destinatario(s) encontrado(s) por cédula."
        else:
            without_recipient += 1
            estado = "SIN_CORREO"
            motivo = "La cédula existe en el Excel, pero no se encontró correo en las fuentes seleccionadas."

        response_rows.append(
            MassEmailExcelRow(
                **row,
                estado=estado,
                motivo=motivo,
                destinatarios=destination_count,
            )
        )

    summary = {
        "total": len(response_rows),
        "con_cedula": len([row for row in response_rows if row.cedula]),
        "listos": ready,
        "sin_correo": without_recipient,
        "sin_cedula": missing_cedula,
        "cedulas_unicas": len(seen_cedulas),
        "cedulas_duplicadas": len(duplicate_cedulas),
        "filas_con_documento": with_document_reference,
        "filas_con_correo_excel": with_excel_email,
        "destinatarios": len(resolved.get("items", [])),
    }

    return MassEmailExcelResponse(
        filename=filename,
        sheet=sheet_name,
        columns=columns,
        detected_columns=detected,
        rows=response_rows,
        items=resolved.get("items", []),
        not_found=resolved.get("not_found", []),
        sources=resolved.get("sources", {}),
        summary=summary,
        warnings=warnings,
        graph_mail_sender=resolved.get("graph_mail_sender"),
    )


@router.post("/enviar")
async def send_mass_email(
    subject: Annotated[str, Form()],
    body: Annotated[str, Form()],
    recipients_json: Annotated[str, Form()] = "[]",
    cedulas: Annotated[str, Form()] = "",
    manual_emails: Annotated[str, Form()] = "",
    cc_emails: Annotated[str, Form()] = "",
    match_attachments_by_cedula: Annotated[bool, Form()] = False,
    send_mode: Annotated[str, Form()] = "individual",
    attachment_assignments_json: Annotated[str, Form()] = "{}",
    files: Annotated[list[UploadFile] | None, File()] = None,
    common_files: Annotated[list[UploadFile] | None, File()] = None,
    student_files: Annotated[list[UploadFile] | None, File()] = None,
    current_user: SessionUser = MassEmailAccess,
) -> MassEmailSendResult:
    normalized_subject = _clean(subject)
    normalized_body = _clean(body)
    normalized_send_mode = _clean(send_mode).lower() or "individual"
    if normalized_send_mode not in {"individual", "single"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Modo de envío inválido.")
    if not normalized_subject:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ingresa el asunto del correo.")
    if not normalized_body:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ingresa el cuerpo del correo.")

    try:
        raw_recipients = json.loads(recipients_json or "[]")
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Lista de destinatarios inválida.") from exc

    recipients: list[MassEmailRecipient] = []
    if isinstance(raw_recipients, list):
        recipients = [MassEmailRecipient.model_validate(item) for item in raw_recipients if isinstance(item, dict)]

    if not recipients and cedulas.strip():
        resolved = _resolve_recipients(MassEmailResolvePayload(cedulas=cedulas))
        recipients = [MassEmailRecipient.model_validate(item) for item in resolved["items"]]

    for index, email in enumerate(_parse_emails(manual_emails), start=1):
        recipients.append(
            MassEmailRecipient(
                id=f"manual:{email.lower()}",
                cedula="",
                email=email,
                nombres=email,
                codigo=str(index),
                tipo_usuario="MANUAL",
                email_tipo="MANUAL",
                source_table="MANUAL",
            )
        )

    recipients = _dedupe_recipients(recipients)
    parsed_cc_emails = _parse_emails(cc_emails)

    if not recipients:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="No hay destinatarios seleccionados.")
    if len(recipients) > MAX_RECIPIENTS_PER_SEND:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Solo se permite enviar hasta {MAX_RECIPIENTS_PER_SEND} destinatarios por lote.",
        )

    sender = _graph_mail_sender()
    effective_match_by_cedula = match_attachments_by_cedula or normalized_send_mode == "single"
    attachment_assignments = _parse_attachment_assignments(attachment_assignments_json)
    common_attachments, attachments_by_cedula, attachment_count, skipped_attachments = await _read_attachment_groups(
        files,
        common_files=common_files,
        student_files=student_files,
        match_by_cedula=effective_match_by_cedula,
        assignments=attachment_assignments,
    )
    personal_attachment_count = sum(len(items) for items in attachments_by_cedula.values())

    results: list[dict[str, Any]] = []
    sent = 0
    failed = 0

    if normalized_send_mode == "single":
        try:
            _send_graph_mail(
                sender=sender,
                recipients=recipients,
                cc_emails=parsed_cc_emails,
                subject=normalized_subject,
                body=normalized_body,
                attachments=common_attachments,
                use_bcc=True,
            )
            sent = 1
            results = [
                {
                    **recipient.model_dump(),
                    "status": "INCLUIDO",
                    "attachment_count": len(common_attachments),
                }
                for recipient in recipients
            ]
        except (RuntimeError, httpx.HTTPError) as exc:
            failed = 1
            error_message = _graph_error_message(exc)
            results = [
                {
                    **recipient.model_dump(),
                    "status": "ERROR",
                    "error": error_message,
                    "attachment_count": 0,
                }
                for recipient in recipients
            ]

        return MassEmailSendResult(
            ok=failed == 0,
            sent=sent,
            failed=failed,
            skipped=personal_attachment_count + skipped_attachments,
            skipped_attachments=skipped_attachments,
            attachment_count=attachment_count,
            send_mode=normalized_send_mode,
            recipients=results,
            message=(
                f"Envío único solicitado por {current_user.login}. "
                f"Adjuntos personalizados omitidos: {personal_attachment_count}. "
                f"Adjuntos sin asignación omitidos: {skipped_attachments}."
            ),
        )

    for recipient in recipients:
        recipient_attachments = (
            common_attachments + attachments_by_cedula.get(_clean(recipient.cedula), [])
            if effective_match_by_cedula
            else common_attachments
        )
        try:
            _send_graph_mail(
                sender=sender,
                recipients=[recipient],
                cc_emails=parsed_cc_emails,
                subject=normalized_subject,
                body=normalized_body,
                attachments=recipient_attachments,
            )
            sent += 1
            results.append({**recipient.model_dump(), "status": "ENVIADO", "attachment_count": len(recipient_attachments)})
        except (RuntimeError, httpx.HTTPError) as exc:
            failed += 1
            results.append({**recipient.model_dump(), "status": "ERROR", "error": _graph_error_message(exc)})

    return MassEmailSendResult(
        ok=failed == 0,
        sent=sent,
        failed=failed,
        skipped=skipped_attachments,
        skipped_attachments=skipped_attachments,
        send_mode=normalized_send_mode,
        attachment_count=attachment_count,
        recipients=results,
        message=(
            f"Envío solicitado por {current_user.login}. "
            f"Adjuntos de estudiante sin asignación omitidos: {skipped_attachments}."
        ),
    )
