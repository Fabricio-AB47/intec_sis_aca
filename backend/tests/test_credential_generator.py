import asyncio
import unittest
from io import BytesIO
from itertools import islice
from types import SimpleNamespace
from urllib.parse import parse_qs, urlsplit
from unittest.mock import MagicMock, patch

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.routers import credential_generator as credentials


def person(**overrides: str) -> dict[str, str | int | None]:
    values: dict[str, str | int | None] = {
        "primer_nombre": "María José",
        "segundo_nombre": "Alejandra",
        "primer_apellido": "De la Cruz",
        "segundo_apellido": "De la Torre",
        "cedula": "0123456789",
        "fila_origen": 2,
    }
    values.update(overrides)
    return values


class CredentialRuleTests(unittest.TestCase):
    def test_compound_names_are_complete_units(self) -> None:
        identity = person()

        self.assertEqual(credentials._email_base(identity), "mariajose.delacruz")
        self.assertEqual(
            credentials._permanent_password(identity, 2026),
            "MDelacruz6789@2026",
        )

    def test_collision_sequence_uses_second_surname_initial_then_numbers(self) -> None:
        candidates = list(islice(credentials._email_candidates(person()), 5))

        self.assertEqual(
            candidates,
            [
                "mariajose.delacruz",
                "mariajose.delacruzd",
                "mariajose.delacruzd1",
                "mariajose.delacruzd2",
                "mariajose.delacruzd3",
            ],
        )

    def test_collision_sequence_without_second_surname_starts_at_one(self) -> None:
        candidates = list(
            islice(credentials._email_candidates(person(segundo_apellido="")), 4)
        )

        self.assertEqual(
            candidates,
            [
                "mariajose.delacruz",
                "mariajose.delacruz1",
                "mariajose.delacruz2",
                "mariajose.delacruz3",
            ],
        )

    def test_email_local_part_never_exceeds_sixty_four_characters(self) -> None:
        identity = person(
            primer_nombre="Nombre Compuesto Muy Extenso " * 3,
            primer_apellido="Apellido Compuesto Muy Extenso " * 3,
        )

        candidates = list(islice(credentials._email_candidates(identity), 3))

        self.assertTrue(all(len(candidate) <= 64 for candidate in candidates))
        self.assertTrue(candidates[2].endswith("d1"))

    def test_excel_keeps_long_value_for_row_level_validation(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Credenciales"
        worksheet.append(credentials._TEMPLATE_HEADERS)
        worksheet.append(["A" * 121, "", "Pérez", "", "0123456789"])
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()

        parsed = credentials._read_workbook(stream.getvalue(), "usuarios.xlsx")
        normalized = credentials._normalized_person(parsed[0])

        self.assertIn("El primer nombre supera 120 caracteres", credentials._person_errors(normalized))

    def test_graph_employee_lookup_uses_exact_document_filter(self) -> None:
        captured_url = ""

        def fake_graph_get(url: str) -> dict[str, object]:
            nonlocal captured_url
            captured_url = url
            return {"value": [{"id": "graph-1", "employeeId": "0123456789"}]}

        with patch.object(credentials, "graph_get", side_effect=fake_graph_get):
            users = credentials._graph_users_by_employee_id("0123456789")

        query = parse_qs(urlsplit(captured_url).query)
        self.assertEqual(query["$filter"], ["employeeId eq '0123456789'"])
        self.assertEqual(users[0]["id"], "graph-1")

    def test_graph_creation_keeps_permanent_password_and_identity(self) -> None:
        captured_payload: dict[str, object] = {}

        def fake_graph_post(_url: str, payload: dict[str, object]) -> dict[str, object]:
            captured_payload.update(payload)
            return {"id": "graph-1", **payload}

        settings = SimpleNamespace(graph_user_usage_location="EC")
        with (
            patch.object(credentials, "get_settings", return_value=settings),
            patch.object(credentials, "graph_post", side_effect=fake_graph_post),
        ):
            created = credentials._create_graph_user(
                person(),
                "mariajose.delacruz@intec.edu.ec",
                "MDelacruz6789@2026",
            )

        self.assertEqual(created["id"], "graph-1")
        self.assertEqual(captured_payload["employeeId"], "0123456789")
        self.assertEqual(captured_payload["usageLocation"], "EC")
        password_profile = captured_payload["passwordProfile"]
        self.assertIsInstance(password_profile, dict)
        self.assertFalse(password_profile["forceChangePasswordNextSignIn"])
        self.assertEqual(captured_payload["passwordPolicies"], "DisablePasswordExpiration")

    def test_archived_password_is_encrypted_and_can_be_recovered(self) -> None:
        first_settings = SimpleNamespace(
            credential_archive_secret="archive-secret-one",
            signing_secret="session-secret",
        )
        second_settings = SimpleNamespace(
            credential_archive_secret="archive-secret-two",
            signing_secret="session-secret",
        )
        password = "MDelacruz6789@2026"

        with patch.object(credentials, "get_settings", return_value=first_settings):
            encrypted = credentials._encrypt_credential_password(password)
            recovered = credentials._decrypt_credential_password(encrypted)

        self.assertTrue(encrypted.startswith("v1:"))
        self.assertNotIn(password, encrypted)
        self.assertEqual(recovered, password)
        with (
            patch.object(credentials, "get_settings", return_value=second_settings),
            self.assertRaisesRegex(RuntimeError, "CREDENTIAL_ARCHIVE_SECRET"),
        ):
            credentials._decrypt_credential_password(encrypted)

    def test_audit_persists_only_encrypted_password(self) -> None:
        settings = SimpleNamespace(
            credential_archive_secret="archive-secret",
            signing_secret="session-secret",
        )
        connection = MagicMock()
        connection.__enter__.return_value = connection
        cursor = MagicMock()
        connection.cursor.return_value = cursor
        password = "MDelacruz6789@2026"
        row = {
            **person(),
            "tipo_persona": "ESTUDIANTE",
            "correo_institucional": "mariajose.delacruz@intec.edu.ec",
            "estado_graph": "CREADO_GRAPH",
            "estado_licencia": "ASIGNADA_ESTUDIANTE",
            "estado_moodle": "CREADO_MOODLE",
            "estado_general": "COMPLETO",
            "clave_emitida": True,
            "clave_permanente": password,
            "observacion": credentials._PERMANENT_PASSWORD_NOTE,
        }

        with (
            patch.object(credentials, "get_settings", return_value=settings),
            patch.object(credentials, "get_connection", return_value=connection),
            patch.object(credentials, "_ensure_tables"),
        ):
            credentials._record_audit("batch", "INDIVIDUAL", [row], "admin")

        parameters = cursor.execute.call_args.args[1:]
        encrypted_password = parameters[-3]
        self.assertEqual(len(parameters), 18)
        self.assertIsInstance(encrypted_password, str)
        self.assertNotIn(password, encrypted_password)
        with patch.object(credentials, "get_settings", return_value=settings):
            self.assertEqual(
                credentials._decrypt_credential_password(encrypted_password),
                password,
            )

    def test_email_resolution_advances_through_real_collision_sequence(self) -> None:
        identity = person()
        owners = {
            "mariajose.delacruz@intec.edu.ec": "1111111111",
            "mariajose.delacruzd@intec.edu.ec": "2222222222",
        }

        with (
            patch.object(credentials, "_existing_email_for_cedula", return_value=""),
            patch.object(credentials, "_graph_domain", return_value="intec.edu.ec"),
            patch.object(credentials, "_graph_is_configured", return_value=False),
            patch.object(credentials, "_moodle_is_configured", return_value=False),
            patch.object(
                credentials,
                "_local_email_owner",
                side_effect=lambda email: owners.get(email, ""),
            ),
            patch.object(
                credentials,
                "_reserve_identity",
                side_effect=lambda _person, email, _operator: email,
            ),
        ):
            email = asyncio.run(
                credentials._resolve_email(identity, "admin", MagicMock())
            )

        self.assertEqual(email, "mariajose.delacruzd1@intec.edu.ec")

    def test_student_license_selection_accepts_a1_and_rejects_faculty(self) -> None:
        subscribed = [
            {
                "skuPartNumber": "STANDARDWOFFPACK_STUDENT",
                "skuId": "314c4481-f395-4525-be8b-2ec4bb1e9d91",
                "capabilityStatus": "Enabled",
                "appliesTo": "User",
                "consumedUnits": 25,
                "prepaidUnits": {"enabled": 100},
            },
            {
                "skuPartNumber": "STANDARDWOFFPACK_FACULTY",
                "skuId": "94763226-9b3c-4e75-a931-5c89701abe66",
                "capabilityStatus": "Enabled",
                "appliesTo": "User",
                "consumedUnits": 1,
                "prepaidUnits": {"enabled": 100},
            },
        ]

        selected = credentials._select_education_license(
            subscribed,
            "ESTUDIANTE",
            "STANDARDWOFFPACK_STUDENT",
        )

        self.assertEqual(selected.name, "Office 365 A1 para estudiantes")
        self.assertEqual(selected.available_units, 75)
        with self.assertRaisesRegex(RuntimeError, "no corresponde"):
            credentials._select_education_license(
                subscribed,
                "ESTUDIANTE",
                "STANDARDWOFFPACK_FACULTY",
            )

        faculty = credentials._select_education_license(
            subscribed,
            "PROFESOR",
            "STANDARDWOFFPACK_FACULTY",
        )
        self.assertEqual(faculty.name, "Office 365 A1 para profesores")
        self.assertEqual(faculty.sku_id, "94763226-9b3c-4e75-a931-5c89701abe66")

    def test_license_assignment_uses_student_sku(self) -> None:
        student_license = credentials._EducationLicense(
            person_type="ESTUDIANTE",
            sku_id="314c4481-f395-4525-be8b-2ec4bb1e9d91",
            sku_part_number="STANDARDWOFFPACK_STUDENT",
            name="Office 365 A1 para estudiantes",
            capability_status="Enabled",
            enabled_units=100,
            consumed_units=25,
            available_units=75,
        )
        captured: dict[str, object] = {}

        def fake_graph_post(url: str, payload: dict[str, object]) -> dict[str, object]:
            captured["url"] = url
            captured["payload"] = payload
            return {"ok": True}

        settings = SimpleNamespace(graph_user_usage_location="EC")
        with (
            patch.object(credentials, "get_settings", return_value=settings),
            patch.object(credentials, "graph_post", side_effect=fake_graph_post),
        ):
            status, error = credentials._assign_graph_license(
                {"id": "graph-1", "usageLocation": "EC", "assignedLicenses": []},
                student_license,
            )

        self.assertEqual(status, "ASIGNADA_ESTUDIANTE")
        self.assertEqual(error, "")
        payload = captured["payload"]
        self.assertIsInstance(payload, dict)
        self.assertEqual(
            payload["addLicenses"][0]["skuId"],
            "314c4481-f395-4525-be8b-2ec4bb1e9d91",
        )

    def test_faculty_license_assignment_uses_independent_sku(self) -> None:
        faculty_license = credentials._EducationLicense(
            person_type="PROFESOR",
            sku_id="94763226-9b3c-4e75-a931-5c89701abe66",
            sku_part_number="STANDARDWOFFPACK_FACULTY",
            name="Office 365 A1 para profesores",
            capability_status="Enabled",
            enabled_units=100,
            consumed_units=10,
            available_units=90,
        )
        captured: dict[str, object] = {}

        def fake_graph_post(_url: str, payload: dict[str, object]) -> dict[str, object]:
            captured.update(payload)
            return {"ok": True}

        with (
            patch.object(
                credentials,
                "get_settings",
                return_value=SimpleNamespace(graph_user_usage_location="EC"),
            ),
            patch.object(credentials, "graph_post", side_effect=fake_graph_post),
        ):
            status, error = credentials._assign_graph_license(
                {"id": "graph-2", "usageLocation": "EC", "assignedLicenses": []},
                faculty_license,
            )

        self.assertEqual(status, "ASIGNADA_PROFESOR")
        self.assertEqual(error, "")
        self.assertEqual(
            captured["addLicenses"][0]["skuId"],
            "94763226-9b3c-4e75-a931-5c89701abe66",
        )

    def test_each_profile_downloads_a_template_with_its_exact_license(self) -> None:
        cases = (
            ("ESTUDIANTE", "Office 365 A1 para estudiantes"),
            ("PROFESOR", "Office 365 A1 para profesores"),
        )

        for person_type, expected_license in cases:
            with self.subTest(person_type=person_type):
                workbook = load_workbook(BytesIO(credentials._template_bytes(person_type)))
                try:
                    instructions = " ".join(
                        str(cell.value or "") for cell in workbook["Instrucciones"]["A"]
                    )
                finally:
                    workbook.close()
                self.assertIn(expected_license, instructions)

    def test_faculty_report_identifies_profile_and_license(self) -> None:
        row = {
            **person(),
            "tipo_persona": "PROFESOR",
            "correo_institucional": "mariajose.delacruz@intec.edu.ec",
            "clave_permanente": "MDelacruz6789@2026",
            "estado_graph": "CREADO_GRAPH",
            "licencia_nombre": "Office 365 A1 para profesores",
            "licencia_sku_part_number": "STANDARDWOFFPACK_FACULTY",
            "estado_licencia": "ASIGNADA_PROFESOR",
            "moodle_username": "mariajose.delacruz@intec.edu.ec",
            "moodle_user_id": 25,
            "estado_moodle": "CREADO_MOODLE",
            "estado_general": "COMPLETO",
            "observacion": credentials._PERMANENT_PASSWORD_NOTE,
        }

        workbook = load_workbook(
            BytesIO(credentials._report_bytes("batch", "INDIVIDUAL", [row], "admin"))
        )
        try:
            headers = [cell.value for cell in workbook["Credenciales"][1]]
            values = [cell.value for cell in workbook["Credenciales"][2]]
            information = " ".join(
                str(cell.value or "") for cell in workbook["Información"]["A"]
            )
        finally:
            workbook.close()

        self.assertIn("Tipo de persona", headers)
        self.assertIn("Licencia Microsoft 365", headers)
        self.assertIn("Contraseña permanente", headers)
        self.assertIn("Observación", headers)
        self.assertIn("PROFESOR", values)
        self.assertIn("Office 365 A1 para profesores", information)
        self.assertIn(credentials._PERMANENT_PASSWORD_NOTE, information)

    def test_archived_report_includes_permanent_password_and_observation(self) -> None:
        report = credentials._archived_credential_report_bytes(
            {
                "tipo_persona": "ESTUDIANTE",
                "nombres": "María José De la Cruz",
                "cedula": "0123456789",
                "correo_institucional": "mariajose.delacruz@intec.edu.ec",
                "observacion": credentials._PERMANENT_PASSWORD_NOTE,
                "fecha_creacion": None,
            },
            "MDelacruz6789@2026",
            "admin",
        )
        workbook = load_workbook(BytesIO(report))
        try:
            headers = [cell.value for cell in workbook["Credencial"][1]]
            values = [cell.value for cell in workbook["Credencial"][2]]
        finally:
            workbook.close()

        self.assertIn("Contraseña permanente", headers)
        self.assertIn("MDelacruz6789@2026", values)
        self.assertIn(credentials._PERMANENT_PASSWORD_NOTE, values)

    def test_history_report_can_be_downloaded_repeatedly_and_is_audited(self) -> None:
        settings = SimpleNamespace(
            credential_archive_secret="archive-secret",
            signing_secret="session-secret",
        )
        with patch.object(credentials, "get_settings", return_value=settings):
            encrypted = credentials._encrypt_credential_password("MDelacruz6789@2026")
        database_row = SimpleNamespace(
            id=7,
            lote_id="batch",
            tipo_persona="ESTUDIANTE",
            modo="INDIVIDUAL",
            cedula="0123456789",
            nombres="María José De la Cruz",
            correo_institucional="mariajose.delacruz@intec.edu.ec",
            clave_cifrada=encrypted,
            observacion=credentials._PERMANENT_PASSWORD_NOTE,
            fecha_creacion=None,
        )
        connection = MagicMock()
        connection.__enter__.return_value = connection
        cursor = MagicMock()
        cursor.fetchone.return_value = database_row
        connection.cursor.return_value = cursor
        current_user = credentials.SessionUser(login="admin", rol="ADMINISTRADOR")

        reports = []
        with (
            patch.object(credentials, "get_settings", return_value=settings),
            patch.object(credentials, "get_connection", return_value=connection),
            patch.object(credentials, "_ensure_tables"),
        ):
            for _ in range(2):
                response = credentials.download_archived_credential_report(7, current_user)

                async def read_response() -> bytes:
                    chunks = [chunk async for chunk in response.body_iterator]
                    return b"".join(
                        chunk if isinstance(chunk, bytes) else chunk.encode("utf-8")
                        for chunk in chunks
                    )

                reports.append(asyncio.run(read_response()))

        self.assertEqual(len(reports), 2)
        for report in reports:
            workbook = load_workbook(BytesIO(report), read_only=True)
            try:
                values = [cell.value for cell in workbook["Credencial"][2]]
            finally:
                workbook.close()
            self.assertIn("MDelacruz6789@2026", values)
        update_calls = [
            call for call in cursor.execute.call_args_list
            if "numero_descargas" in str(call.args[0])
        ]
        self.assertEqual(len(update_calls), 2)
        self.assertEqual(connection.commit.call_count, 2)

    def test_report_is_owned_and_downloaded_only_once(self) -> None:
        report_id = credentials._store_report("admin", "reporte.xlsx", b"report")

        report = credentials._take_report(report_id, "ADMIN")

        self.assertEqual(report.content, b"report")
        with self.assertRaises(HTTPException) as captured:
            credentials._take_report(report_id, "admin")
        self.assertEqual(captured.exception.status_code, 404)

    def test_result_is_not_complete_without_student_license(self) -> None:
        self.assertEqual(
            credentials._overall_status(
                "CREADO_GRAPH",
                "NO_CONFIGURADA",
                "CREADO_MOODLE",
            ),
            "PARCIAL",
        )
        self.assertEqual(
            credentials._overall_status(
                "CREADO_GRAPH",
                "ASIGNADA_ESTUDIANTE",
                "CREADO_MOODLE",
            ),
            "COMPLETO",
        )


if __name__ == "__main__":
    unittest.main()
