import secrets
import unittest
from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook

from app.routers.institutional_email import (
    _TEMPLATE_HEADERS,
    _build_template_bytes,
    _normalize_excel_cedula,
    _read_workbook,
    _unique_student_code,
    _upsert_student_email,
    _validate_institutional_email,
    _validate_password,
)
from app.services.screen_access import SCREEN_CATALOG


TEST_PASSWORD = secrets.token_urlsafe(12)


def workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Actualización"
    worksheet.append(["cedula", "correo_intec_nuevo", "password_nueva"])
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class InstitutionalEmailValidationTests(unittest.TestCase):
    def test_bulk_template_uses_cedula_as_the_only_identity_column(self) -> None:
        self.assertEqual(
            _TEMPLATE_HEADERS,
            ["cedula", "correo_intec_nuevo", "password_nueva"],
        )
        self.assertNotIn("codigo_estud", _TEMPLATE_HEADERS)

    def test_downloaded_template_is_a_plain_excel_compatible_range(self) -> None:
        content = _build_template_bytes(["0102246014", "1712345678"])

        with ZipFile(BytesIO(content)) as archive:
            self.assertIsNone(archive.testzip())
            self.assertFalse(any(name.startswith("xl/tables/") for name in archive.namelist()))

        workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
        worksheet = workbook["Actualización"]
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(worksheet.auto_filter.ref, "A1:C3")
        self.assertEqual(
            [worksheet.cell(1, column).value for column in range(1, 4)],
            _TEMPLATE_HEADERS,
        )
        self.assertEqual(worksheet["A2"].value, "0102246014")
        self.assertIsNone(worksheet["B2"].value)
        self.assertIsNone(worksheet["C2"].value)
        workbook.close()

    def test_student_code_must_be_unique_and_positive(self) -> None:
        self.assertEqual(_unique_student_code({"_codigos_estud": [42]}), 42)
        self.assertIsNone(_unique_student_code({"_codigos_estud": []}))
        self.assertIsNone(_unique_student_code({"_codigos_estud": [42, 84]}))
        self.assertIsNone(_unique_student_code({"codigo_estud": 0}))

    def test_normalizes_numeric_cedula_preserving_leading_zero(self) -> None:
        self.assertEqual(_normalize_excel_cedula(123456789), "0123456789")
        self.assertEqual(_normalize_excel_cedula(Decimal("123456789")), "0123456789")
        self.assertEqual(_normalize_excel_cedula("0123456789"), "0123456789")

    def test_rejects_invalid_institutional_domain(self) -> None:
        self.assertEqual(
            _validate_institutional_email(" Persona@INTEC.EDU.EC "),
            "persona@intec.edu.ec",
        )
        with self.assertRaises(ValueError):
            _validate_institutional_email("persona@example.com")

    def test_password_respects_database_column_limits(self) -> None:
        self.assertEqual(_validate_password(TEST_PASSWORD), TEST_PASSWORD)
        for value in ("12345", "x" * 21, f" {TEST_PASSWORD}", f"{TEST_PASSWORD}\n"):
            with self.subTest(value=value), self.assertRaises(ValueError):
                _validate_password(value)

    def test_template_reference_rows_are_ignored_until_changed(self) -> None:
        parsed = _read_workbook(
            workbook_bytes(
                [
                    [123456789, None, None],
                    ["1712345678", "estudiante@intec.edu.ec", TEST_PASSWORD],
                ]
            ),
            "actualizacion_correo_intec.xlsx",
        )
        self.assertEqual(
            parsed,
            [
                {
                    "row": 3,
                    "cedula": "1712345678",
                    "correo_intec": "estudiante@intec.edu.ec",
                    "password": TEST_PASSWORD,
                }
            ],
        )

    def test_empty_update_workbook_is_rejected(self) -> None:
        with self.assertRaises(HTTPException) as context:
            _read_workbook(
                workbook_bytes([[123456789, None, None]]),
                "actualizacion_correo_intec.xlsx",
            )
        self.assertEqual(context.exception.status_code, 400)


class RecordingCursor:
    def __init__(self) -> None:
        self.query = ""
        self.params: tuple[object, ...] = ()

    def execute(self, query: str, *params: object) -> None:
        self.query = query
        self.params = params


class InstitutionalEmailPersistenceTests(unittest.TestCase):
    def test_upsert_updates_registry_and_student_table_atomically(self) -> None:
        cursor = RecordingCursor()
        _upsert_student_email(
            cursor,
            {
                "codigo_estud": 42,
                "estudiante": "ESTUDIANTE DE PRUEBA",
                "correo_personal": "personal@example.com",
                "periodo_codigo": 1060,
                "correo_nuevo": "estudiante@intec.edu.ec",
                "password_nueva": TEST_PASSWORD,
            },
        )

        normalized_query = " ".join(cursor.query.split()).upper()
        self.assertIn("UPDATE DBO.CORREOSESTUDINTEC", normalized_query)
        self.assertIn("INSERT INTO DBO.CORREOSESTUDINTEC", normalized_query)
        self.assertIn("UPDATE DBO.DATOS_ESTUD", normalized_query)
        self.assertIn("CORREOINTEC = ?", normalized_query)
        self.assertIn("CLAVE = ?", normalized_query)
        self.assertEqual(cursor.query.count("?"), len(cursor.params))
        self.assertEqual(len(cursor.params), 17)

    def test_screen_is_available_for_assignment(self) -> None:
        screen = next(
            item for item in SCREEN_CATALOG
            if item["page"] == "actualizar-correo-intec"
        )
        self.assertEqual(screen["group"], "Actualizaciones")
        self.assertIn("correo institucional", screen["description"].lower())


if __name__ == "__main__":
    unittest.main()
