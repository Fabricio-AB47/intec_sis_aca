import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas

from app.services.curriculum_updater import (
    analyze_curriculum,
    generate_curriculum_workbook,
    parse_pea_pdf,
)


def base_workbook_bytes(*, title_row: bool = False, blank_target: bool = False) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Malla"
    header_row = 2 if title_row else 1
    if title_row:
        sheet.cell(1, 1, "Carrera en Desarrollo de Software - 01-2-A")
        sheet.merge_cells("A1:L1")

    headers = (
        "Asignaturas",
        "Período Académico",
        "Unidades de organización curricular",
        "Horas / Créditos",
        "Componetes del aprendizaje",
        "",
        "",
        "Práctica Pre Profesional",
        "Vinculación",
        "Trabajo de titulación",
        "Total horas carrera",
        "No. Créditos",
    )
    for column, value in enumerate(headers, start=1):
        sheet.cell(header_row, column, value)
    sheet.merge_cells(start_row=header_row, start_column=1, end_row=header_row + 1, end_column=1)
    sheet.merge_cells(start_row=header_row, start_column=2, end_row=header_row + 1, end_column=2)
    sheet.merge_cells(start_row=header_row, start_column=3, end_row=header_row + 1, end_column=3)
    sheet.merge_cells(start_row=header_row, start_column=4, end_row=header_row + 1, end_column=4)
    sheet.merge_cells(start_row=header_row, start_column=5, end_row=header_row, end_column=7)
    for column in range(8, 13):
        sheet.merge_cells(start_row=header_row, start_column=column, end_row=header_row + 1, end_column=column)
    sheet.cell(header_row + 1, 5, "Aprendizaje en contacto con el docente")
    sheet.cell(header_row + 1, 6, "Aprendizaje práctico/experimental")
    sheet.cell(header_row + 1, 7, "Aprendizaje autónomo")

    first_row = header_row + 2
    values = (
        "Sistemas Operativos",
        1,
        "Profesional",
        3,
        f"=D{first_row}*16",
        30,
        30,
        None,
        None,
        None,
        f"=SUM(E{first_row}:J{first_row})",
        f"=K{first_row}/48",
    )
    for column, value in enumerate(values, start=1):
        sheet.cell(first_row, column, value)

    second_row = first_row + 1
    values = (
        "Inteligencia Artificial 1",
        1,
        "Profesional",
        3,
        f"=D{second_row}*16",
        30,
        30,
        None,
        None,
        None,
        f"=SUM(E{second_row}:J{second_row})",
        f"=K{second_row}/48",
    )
    for column, value in enumerate(values, start=1):
        sheet.cell(second_row, column, value)

    subtotal_row = second_row + 1
    for column in range(4, 13):
        letter = chr(64 + column)
        sheet.cell(subtotal_row, column, f"=SUM({letter}{first_row}:{letter}{second_row})")

    if blank_target:
        target = workbook.create_sheet("Malla Desarrollo")
        workbook.move_sheet(target, offset=-1)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def workbook_with_unrelated_enriched_sheet() -> bytes:
    workbook = load_workbook(BytesIO(base_workbook_bytes()))
    unrelated = workbook.create_sheet("Malla Otra Carrera")
    headers = (
        "Asignaturas",
        "Período Académico",
        "Unidades de organización curricular",
        "Campos de formación",
        "Resultados de aprendizaje",
        "Contenidos Mínimos",
        "Horas / Créditos",
    )
    for column, value in enumerate(headers, start=1):
        unrelated.cell(1, column, value)
    unrelated.append([None] * len(headers))
    unrelated.append(["Materia sin relación", 1, "Profesional", "Tecnologías", "Resultado", "Unidad", 3])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def pea_pdf_bytes(
    subject: str = "Sistemas Operativos",
    career: str = "Desarrollo de Software",
) -> bytes:
    output = BytesIO()
    document = canvas.Canvas(output)
    lines = [
        "PROGRAMA DE ESTUDIOS DE ASIGNATURA - PEA",
        f"Carrera: {career}",
        f"Código de la asignatura DS-01 Nombre de la asignatura {subject}",
        "Nivel de la asignatura 1ero Unidad de organización curricular Profesional Campo de Formación Tecnologías",
        "Resultados de Aprendizaje de la asignatura y como aporta al perfil profesional:",
        "Analiza la arquitectura del sistema operativo y administra sus recursos de forma segura.",
        "ALINEAMIENTO CURRICULAR:",
        "CONTENIDOS DE LA ASIGNATURA",
        "UNIDAD 1: Procesos de un Sistema Operativo",
        "Resultado de Aprendizaje: Gestiona procesos y memoria.",
        "Contenidos Horas de la Unidad Observaciones",
        "UNIDAD 2: Arquitectura y tipos de sistemas operativos",
        "Resultado de Aprendizaje: Compara arquitecturas y selecciona la apropiada.",
        "Contenidos Horas de la Unidad Observaciones",
        "Descripción complementaria " + ("para validar la extracción de texto seleccionable. " * 10),
    ]
    y = 800
    for line in lines:
        document.drawString(35, y, line[:115])
        y -= 26
    document.save()
    return output.getvalue()


def syllabus_pdf_bytes(subject: str = "Sistemas Operativos") -> bytes:
    output = BytesIO()
    document = canvas.Canvas(output)
    document.drawString(35, 805, "SÍLABO DE LA ASIGNATURA")
    document.drawString(35, 780, "Carrera: Desarrollo de Software")
    document.drawString(35, 755, f"Código de la asignatura DS-01 Nombre de la asignatura {subject}")
    document.drawString(35, 720, "PLAN DE ESTUDIOS")
    document.drawString(35, 695, "Resultados de Aprendizaje")
    document.drawString(35, 665, "Perfil de egreso de la carrera")
    document.drawString(250, 665, "Resultado de aprendizaje")
    document.drawString(470, 665, "Evidencia")
    document.drawString(250, 645, "de la asignatura")
    document.drawString(35, 620, "Perfil profesional de referencia.")
    document.drawString(250, 620, "Resultado exclusivo del sílabo para la asignatura.")
    document.drawString(470, 620, "Evaluación práctica.")
    document.drawString(250, 600, "Integra conocimientos y resuelve problemas del contexto.")
    document.drawString(35, 560, "Información complementaria para validar la extracción estructurada. " * 4)
    document.showPage()
    document.drawString(35, 805, "CONTENIDOS DE LA ASIGNATURA")
    document.drawString(60, 770, "UNIDAD")
    document.drawString(180, 770, "TEMA")
    document.drawString(350, 770, "SEMANA")
    document.drawString(60, 730, "Fundamentos de datos")
    document.drawString(180, 730, "Conceptos, fuentes y ciclo de vida")
    document.drawString(60, 680, "Evaluación")
    document.drawString(180, 680, "Evaluación de fundamentos")
    document.drawString(60, 630, "Modelos analíticos")
    document.drawString(180, 630, "Diseño y aplicación de modelos")
    document.drawString(60, 580, "Gobierno de datos")
    document.drawString(180, 580, "Calidad, seguridad y gobierno")
    document.drawString(60, 520, "Estrategias Metodológicas")
    document.save()
    return output.getvalue()


def catalog_search_pdf_bytes(subject_text: str, *, label: str = "") -> bytes:
    output = BytesIO()
    document = canvas.Canvas(output)
    lines = [
        "DOCUMENTO ACADÉMICO INSTITUCIONAL",
        f"{label}: {subject_text}" if label else subject_text,
        "Carrera: Tecnología Superior",
        "Resultados de Aprendizaje de la asignatura:",
        "Resultado general que no se utilizará para completar la malla.",
        "CONTENIDOS DE LA ASIGNATURA",
        "UNIDAD 1: Fundamentos",
        "Resultado de Aprendizaje: Identifica y aplica los fundamentos de la unidad.",
        "Contenidos Horas de la Unidad Observaciones",
        *(
            "Información académica complementaria para validar la búsqueda dentro del documento PDF."
            for _ in range(8)
        ),
    ]
    y = 805
    for line in lines:
        document.drawString(35, y, line)
        y -= 28
    document.save()
    return output.getvalue()


def cover_title_pdf_bytes(
    title_lines: tuple[str, ...],
    *,
    repeated_title_lines: tuple[str, ...] = (),
) -> bytes:
    output = BytesIO()
    document = canvas.Canvas(output)
    y = 805
    for line in (*title_lines, *repeated_title_lines, "CONTROL DE CAMBIOS"):
        document.drawString(35, y, line)
        y -= 28
    document.showPage()
    lines = [
        "DATOS GENERALES DE LA ASIGNATURA",
        "Carrera: Tecnología Superior",
        "CONTENIDOS DE LA ASIGNATURA",
        "UNIDAD 1: Fundamentos",
        "Resultado de Aprendizaje: Aplica los fundamentos de la unidad.",
        "Contenidos Horas de la Unidad Observaciones",
        *(
            "Información académica complementaria para completar la extracción del documento."
            for _ in range(10)
        ),
    ]
    y = 805
    for line in lines:
        document.drawString(35, y, line)
        y -= 28
    document.save()
    return output.getvalue()


def style_signature(cell: object) -> tuple[str, ...]:
    return tuple(
        str(getattr(cell, attribute))
        for attribute in ("font", "fill", "border", "alignment", "protection", "number_format")
    )


class CurriculumUpdaterTests(unittest.TestCase):
    def test_parses_pea_fields_and_units(self) -> None:
        result = parse_pea_pdf(pea_pdf_bytes(), "PEA Sistemas Operativos.pdf")

        self.assertEqual(result["subject_name"], "Sistemas Operativos")
        self.assertEqual(result["career_name"], "Desarrollo de Software")
        self.assertEqual(result["field"], "Tecnologías")
        self.assertEqual(
            result["learning_outcomes"],
            "1. Gestiona procesos y memoria.\n"
            "2. Compara arquitecturas y selecciona la apropiada.",
        )
        self.assertNotIn("Analiza la arquitectura", result["learning_outcomes"])
        self.assertEqual(len(result["units"]), 2)
        self.assertIn("UNIDAD 2: Arquitectura", result["minimum_contents"])

    def test_does_not_use_general_syllabus_result_as_a_unit_outcome(self) -> None:
        result = parse_pea_pdf(syllabus_pdf_bytes(), "Sílabo Sistemas Operativos.pdf")

        self.assertEqual(result["document_type"], "SILABO")
        self.assertEqual(result["subject_name"], "Sistemas Operativos")
        self.assertEqual(result["learning_outcomes"], "")
        self.assertTrue(any("resultado general se omitió" in warning for warning in result["warnings"]))
        self.assertEqual(len(result["units"]), 3)
        self.assertIn("UNIDAD 1: Fundamentos de datos", result["minimum_contents"])
        self.assertNotIn("Evaluación", result["minimum_contents"])

    def test_detects_base_and_matches_pea(self) -> None:
        result = analyze_curriculum(
            base_workbook_bytes(),
            "Malla Desarrollo Software.xlsx",
            [("PEA Sistemas Operativos.pdf", pea_pdf_bytes())],
            "Desarrollo de Software",
        )

        self.assertTrue(result["workbook"]["target_will_be_created"])
        self.assertEqual(result["workbook"]["subject_count"], 2)
        self.assertEqual(result["summary"]["ready"], 1)
        matched = next(row for row in result["rows"] if row["document_index"] is not None)
        self.assertEqual(matched["subject_name"], "Sistemas Operativos")
        self.assertTrue(matched["apply_recommended"])

    def test_matches_common_subject_even_when_pdf_career_is_different(self) -> None:
        result = analyze_curriculum(
            base_workbook_bytes(),
            "Malla Desarrollo Software.xlsx",
            [("PEA Sistemas Operativos.pdf", pea_pdf_bytes(career="Ciberseguridad"))],
            "Desarrollo de Software",
        )

        matched = next(row for row in result["rows"] if row["subject_name"] == "Sistemas Operativos")
        self.assertEqual(matched["status"], "LISTO")
        self.assertEqual(matched["match_type"], "NOMBRE_EXACTO")
        self.assertTrue(matched["apply_recommended"])
        self.assertEqual(result["summary"]["unmatched_documents"], 0)
        self.assertFalse(any("carrera no coincide" in warning.lower() for warning in matched["warnings"]))

    def test_ignores_pdf_when_subject_does_not_exist_in_workbook(self) -> None:
        result = analyze_curriculum(
            base_workbook_bytes(),
            "Malla Desarrollo Software.xlsx",
            [("PEA Materia desconocida.pdf", pea_pdf_bytes("Materia completamente desconocida"))],
            "Desarrollo de Software",
        )

        self.assertEqual(result["summary"]["ready"], 0)
        self.assertEqual(result["summary"]["unmatched_documents"], 1)
        self.assertEqual(result["unmatched_documents"][0]["filename"], "PEA Materia desconocida.pdf")
        self.assertTrue(all(row["document_index"] is None for row in result["rows"]))
        self.assertTrue(all(not row["apply_recommended"] for row in result["rows"]))

    def test_matches_four_documents_using_subject_names_inside_the_pdf(self) -> None:
        result = analyze_curriculum(
            base_workbook_bytes(),
            "Malla Desarrollo Software.xlsx",
            [
                ("documento-001.pdf", catalog_search_pdf_bytes("Sis te mas Ope ra ti vos")),
                ("Silabo documento-002.pdf", catalog_search_pdf_bytes("Sis te mas Ope ra ti vos")),
                ("documento-003.pdf", catalog_search_pdf_bytes("Inteligencia Artificial I", label="Materia")),
                ("Silabo documento-004.pdf", catalog_search_pdf_bytes("Inteligencia Artificial I", label="Materia")),
            ],
            "Desarrollo de Software",
        )

        systems = next(row for row in result["rows"] if row["subject_name"] == "Sistemas Operativos")
        artificial_intelligence = next(
            row for row in result["rows"] if row["subject_name"] == "Inteligencia Artificial 1"
        )
        self.assertEqual(result["summary"]["documents"], 4)
        self.assertEqual(result["summary"]["unmatched_documents"], 0)
        self.assertEqual(len(systems["document_indices"]), 2)
        self.assertEqual(len(artificial_intelligence["document_indices"]), 2)
        self.assertEqual(systems["match_type"], "NOMBRE_EXACTO")
        self.assertEqual(artificial_intelligence["match_type"], "NOMBRE_EXACTO")

    def test_uses_multiline_first_page_title_for_similar_subjects(self) -> None:
        catalog = (
            "Administración de Base de Datos",
            "Seguridad de base de datos",
            "Fundamentos de Seguridad Informática y ciberseguridad",
            "Gestión de incidentes de seguridad - Ciberinteligencia",
            "Gestión de riesgos de la seguridad informática",
            "Investigación forense digital",
            "Legislación Empresarial y Laboral",
            "Titulación",
            "Hacking Etico 1",
            "Hacking Etico 2",
        )
        cases = (
            (
                "documento-legislacion.pdf",
                (
                    "PROGRAMA DE ESTUDIO",
                    "DE LA ASIGNATURA",
                    "LEGISLACIÓN",
                    "EMPRESARIAL Y",
                    "LEGISLACIÓN LABORAL",
                ),
                (),
                "Legislación Empresarial y Laboral",
            ),
            (
                "documento-incidentes.pdf",
                (
                    "PROGRAMA DE ESTUDIO DE LA",
                    "ASIGNATURA GESTIÓN DE INCIDENTES",
                    "DE SEGURIDAD y CIBERINTELIGENCIA Página 1 de 8",
                ),
                (),
                "Gestión de incidentes de seguridad - Ciberinteligencia",
            ),
            (
                "documento-base-datos.pdf",
                (
                    "PROGRAMA DE ESTUDIO DE LA",
                    "ASIGNATURA SEGURIDAD EN BASE",
                    "DE DATOS Página 1 de 6",
                ),
                (),
                "Seguridad de base de datos",
            ),
            (
                "PEA INVESTIGACIÓN DIGITAL.pdf",
                (
                    "PROGRAMA DE ESTUDIO DE LA",
                    "ASIGNATURA GESTION DE",
                    "RIESGOS DE LA SEGURIDAD INF. Página 1 de 6",
                ),
                (),
                "Gestión de riesgos de la seguridad informática",
            ),
            (
                "documento-hacking.pdf",
                (
                    "PROGRAMA DE ESTUDIO DE",
                    "LA ASIGNATURA HACKING",
                    "ÉTICO 2 Página 1 de 9",
                ),
                (
                    "PROGRAMA DE ESTUDIO DE",
                    "LA ASIGNATURA HACKING",
                    "ÉTICO 1",
                ),
                "Hacking Etico 2",
            ),
        )

        for filename, title_lines, repeated_title_lines, expected in cases:
            with self.subTest(filename=filename):
                result = parse_pea_pdf(
                    cover_title_pdf_bytes(
                        title_lines,
                        repeated_title_lines=repeated_title_lines,
                    ),
                    filename,
                    subject_catalog=catalog,
                )
                self.assertEqual(result["subject_name"], expected)

    def test_combines_pea_and_syllabus_for_the_same_subject(self) -> None:
        result = analyze_curriculum(
            base_workbook_bytes(),
            "Malla Desarrollo Software.xlsx",
            [
                ("PEA Sistemas Operativos.pdf", pea_pdf_bytes()),
                ("Sílabo Sistemas Operativos.pdf", syllabus_pdf_bytes()),
            ],
            "Desarrollo de Software",
        )

        matched = next(row for row in result["rows"] if row["subject_name"] == "Sistemas Operativos")
        self.assertEqual(len(matched["document_indices"]), 2)
        self.assertEqual(
            matched["source_files"],
            ["PEA Sistemas Operativos.pdf", "Sílabo Sistemas Operativos.pdf"],
        )
        self.assertIn("1. Gestiona procesos y memoria.", matched["proposal"]["learning_outcomes"])
        self.assertIn("2. Compara arquitecturas", matched["proposal"]["learning_outcomes"])
        self.assertIn("Fundamentos de datos", matched["proposal"]["minimum_contents"])
        self.assertTrue(any("Se combinaron PEA y sílabo" in warning for warning in matched["warnings"]))

    def test_generates_enriched_sheet_without_changing_base(self) -> None:
        source = base_workbook_bytes()
        analysis = analyze_curriculum(
            source,
            "Malla Desarrollo Software.xlsx",
            [("PEA Sistemas Operativos.pdf", pea_pdf_bytes())],
            "Desarrollo de Software",
        )
        updates = [
            {
                "row_number": row["row_number"],
                "subject_name": row["subject_name"],
                "period": row["period"],
                "apply": row["apply_recommended"],
                "status": row["status"],
                "source_file": row["source_file"],
                "proposal": row["proposal"],
            }
            for row in analysis["rows"]
        ]

        content, _filename, metadata = generate_curriculum_workbook(
            source,
            "Malla Desarrollo Software.xlsx",
            "Desarrollo de Software",
            updates,
            "pruebas",
        )
        workbook = load_workbook(BytesIO(content), data_only=False)
        target = workbook[metadata["target_sheet"]]

        self.assertEqual(workbook["Malla"]["D3"].value, 3)
        self.assertEqual(target["D1"].value, "Campos de formación")
        self.assertEqual(target["E1"].value, "Resultados de aprendizaje")
        self.assertEqual(target["F1"].value, "Contenidos Mínimos")
        self.assertEqual(target["H3"].value, "=G3*16")
        self.assertEqual(target["Q1"].value, "Horas / Créditos")
        self.assertEqual(target["D3"].value, "Tecnologías")
        self.assertEqual(
            target["E3"].value,
            "1. Gestiona procesos y memoria.\n"
            "2. Compara arquitecturas y selecciona la apropiada.",
        )
        self.assertIn("UNIDAD 1", target["F3"].value)
        self.assertEqual(target.max_column, 25)
        self.assertEqual(target.column_dimensions["A"].width, 28.42578125)
        self.assertEqual(target.column_dimensions["D"].width, 15.140625)
        self.assertEqual(target.column_dimensions["E"].width, 54.7109375)
        self.assertEqual(target.column_dimensions["F"].width, 40.0)
        self.assertEqual(target["A3"].fill.fgColor.rgb, "FFFFFF00")
        self.assertEqual(target["Q1"].fill.fgColor.theme, 9)
        self.assertEqual(target["R2"].alignment.text_rotation, 90)
        self.assertEqual(target["R3"].number_format, "#,##0.00")
        self.assertEqual(target.page_setup.orientation, "portrait")
        self.assertNotIn("Resultado OCR", workbook.sheetnames)
        self.assertEqual(metadata["format"], "Malla ADM")
        self.assertEqual(metadata["applied"], 1)

    def test_removes_an_existing_ocr_result_sheet_from_download(self) -> None:
        workbook = load_workbook(BytesIO(base_workbook_bytes()))
        workbook.create_sheet("Resultado OCR")
        source = BytesIO()
        workbook.save(source)

        content, _filename, _metadata = generate_curriculum_workbook(
            source.getvalue(),
            "Malla Desarrollo Software.xlsx",
            "Desarrollo de Software",
            [],
            "pruebas",
        )
        generated = load_workbook(BytesIO(content), data_only=False)

        self.assertNotIn("Resultado OCR", generated.sheetnames)

    def test_uses_existing_empty_career_sheet_without_moving_the_adm_template(self) -> None:
        source = base_workbook_bytes(title_row=True, blank_target=True)
        analysis = analyze_curriculum(source, "Malla Desarrollo.xlsx", [], "Desarrollo")
        content, _filename, metadata = generate_curriculum_workbook(
            source,
            "Malla Desarrollo.xlsx",
            "Desarrollo",
            [],
            "pruebas",
        )
        workbook = load_workbook(BytesIO(content), data_only=False)
        target = workbook[metadata["target_sheet"]]

        self.assertEqual(analysis["workbook"]["header_row"], 2)
        self.assertEqual(metadata["target_sheet"], "Malla Desarrollo")
        self.assertEqual(target["D1"].value, "Campos de formación")
        self.assertEqual(target["A3"].value, "Sistemas Operativos")
        self.assertNotIn("A1:Y1", {str(value) for value in target.merged_cells.ranges})
        self.assertEqual(target.freeze_panes, "G3")

    def test_replicates_the_immutable_malla_adm_format(self) -> None:
        content, _filename, metadata = generate_curriculum_workbook(
            base_workbook_bytes(),
            "Malla Desarrollo Software.xlsx",
            "Desarrollo de Software",
            [],
            "pruebas",
        )
        generated = load_workbook(BytesIO(content), data_only=False)[metadata["target_sheet"]]
        template_path = Path(__file__).resolve().parents[1] / "app" / "templates" / "malla_adm_template.xlsx"
        template = load_workbook(template_path, data_only=False)["Malla ADM"]

        self.assertEqual(generated.max_row, template.max_row)
        self.assertEqual(generated.max_column, template.max_column)
        self.assertEqual(
            {str(value) for value in generated.merged_cells.ranges},
            {str(value) for value in template.merged_cells.ranges},
        )
        for row in range(1, template.max_row + 1):
            self.assertEqual(generated.row_dimensions[row].height, template.row_dimensions[row].height)
            for column in range(1, template.max_column + 1):
                self.assertEqual(
                    style_signature(generated.cell(row, column)),
                    style_signature(template.cell(row, column)),
                )
        for column in template.column_dimensions:
            self.assertEqual(
                generated.column_dimensions[column].width,
                template.column_dimensions[column].width,
            )

    def test_recognizes_generated_target_on_next_analysis(self) -> None:
        source = base_workbook_bytes()
        generated, _filename, _metadata = generate_curriculum_workbook(
            source,
            "Malla Desarrollo Software.xlsx",
            "Desarrollo de Software",
            [],
            "pruebas",
        )

        result = analyze_curriculum(generated, "malla-actualizada.xlsx", [], "Desarrollo de Software")

        self.assertTrue(result["workbook"]["target_exists"])
        self.assertFalse(result["workbook"]["target_will_be_created"])

    def test_does_not_reuse_enriched_sheet_from_another_career(self) -> None:
        result = analyze_curriculum(
            workbook_with_unrelated_enriched_sheet(),
            "Malla Desarrollo Software.xlsx",
            [],
            "Desarrollo de Software",
        )

        self.assertFalse(result["workbook"]["target_exists"])
        self.assertEqual(result["workbook"]["target_sheet"], "Malla Desarrollo de Software")

    def test_updates_exact_row_when_subject_names_are_repeated(self) -> None:
        workbook = load_workbook(BytesIO(base_workbook_bytes()))
        workbook["Malla"]["A4"] = "Sistemas Operativos"
        source = BytesIO()
        workbook.save(source)
        updates = [
            {
                "row_number": 4,
                "subject_name": "Sistemas Operativos",
                "period": "1",
                "apply": True,
                "status": "REVISION_MANUAL",
                "source_file": "PEA repetido.pdf",
                "proposal": {
                    "field": "Tecnologías",
                    "learning_outcomes": "Resultado dirigido a la segunda fila.",
                    "minimum_contents": "UNIDAD 1: Contenido",
                },
            }
        ]

        generated, _filename, metadata = generate_curriculum_workbook(
            source.getvalue(),
            "Malla Desarrollo Software.xlsx",
            "Desarrollo de Software",
            updates,
            "pruebas",
        )
        result = load_workbook(BytesIO(generated), data_only=False)[metadata["target_sheet"]]

        self.assertIsNone(result["D3"].value)
        self.assertEqual(result["D4"].value, "Tecnologías")


if __name__ == "__main__":
    unittest.main()
